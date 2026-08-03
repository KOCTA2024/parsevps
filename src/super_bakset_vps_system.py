#!/usr/bin/env python3
"""SUPER_BASKET LIVE ADVISOR v12.4.1.

PROGRAMMER INTEGRATION:
    python super_basket_vps_system_FINAL_v12_4_FIVE_CHECKPOINTS_STAT_PROJECTION_FIX.py run --match /path/match.json --checkpoint 1
    python super_basket_vps_system_FINAL_v12_4_FIVE_CHECKPOINTS_STAT_PROJECTION_FIX.py watch --inbox /srv/basket/inbox --outbox /srv/basket/outbox

The parser remains the source of live match, history, lines and statistics.
This single file:
1) calculates exact P_history -> line-specific P_scenario -> live projections -> P_live -> P_raw -> P_final;
2) mines and explains repeated game-state scenarios and their impact on each total/team-IT line;
3) evaluates quarter, half and match totals plus both teams' individual totals;
4) detects history/live conflict and FAKE OVER / FAKE UNDER when statistics are available;
5) supports five explicit production checkpoints without confusing provider clock semantics;
6) ranks 1-2 strongest real/reference lines and keeps PASS silent in Telegram;
7) stores deterministic delivery/audit data in JSON and SQLite.

The v12.4.1 advisor is deterministic and does not require GPT. It supports PREMATCH, EARLY_LIVE_Q1, EARLY_LIVE_Q2, HT and Q4_CONFIRMATION; preserves fresh parser live projections; distinguishes missing statistics from present-but-inconsistent statistics; and builds history-anchored reference lines when a relevant bookmaker total/IT is absent. Optional legacy GPT and Excel-audit integrations require external packages only when explicitly enabled.
"""
from __future__ import annotations

import argparse
import base64
import gzip
import html
import hashlib
import json
import math
import os
import re
import sqlite3
import statistics
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from copy import deepcopy
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Callable, Iterable, Literal, Optional


# ===== schema_adapter.py =====
def to_number(value: Any) -> Optional[float]:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(',', '.')
    if not text:
        return None
    if text.endswith('%'):
        text = text[:-1]
    try:
        return float(text)
    except ValueError:
        return None

def to_int(value: Any) -> Optional[int]:
    number = to_number(value)
    return None if number is None else int(round(number))

def first(mapping: dict[str, Any], names: Iterable[str]) -> Any:
    for name in names:
        if name in mapping and mapping[name] not in (None, ''):
            return mapping[name]
    return None

def alias_value(mapping: dict[str, Any], canonical: str, aliases: dict[str, list[str]]) -> Any:
    return first(mapping, aliases.get(canonical, [canonical]))

def percentile(values: list[float], probability: float) -> Optional[float]:
    if not values:
        return None
    ordered = sorted(values)
    if len(ordered) == 1:
        return ordered[0]
    position = (len(ordered) - 1) * probability
    lower = int(position)
    upper = min(lower + 1, len(ordered) - 1)
    fraction = position - lower
    return ordered[lower] * (1 - fraction) + ordered[upper] * fraction

def clock_to_seconds(clock: Any) -> Optional[int]:
    if clock is None:
        return None
    if isinstance(clock, (int, float)):
        return max(0, int(round(float(clock))))
    text = str(clock).strip()
    match = re.fullmatch('(\\d{1,2}):(\\d{2})', text)
    if match:
        return int(match.group(1)) * 60 + int(match.group(2))
    number = to_number(text)
    if number is not None:
        return max(0, int(round(number)))
    return None

def _format_info(mapping: dict[str, Any], tournament: str, config: dict[str, Any]) -> dict[str, Any]:
    """Resolve game duration with explicit JSON taking priority over name heuristics."""
    settings = config['match_format']
    rules = mapping.get('rules', {}) if isinstance(mapping.get('rules'), dict) else {}
    explicit_quarter = to_int(first(rules, ['quarter_minutes', 'period_minutes']))
    if explicit_quarter is None:
        explicit_quarter = to_int(first(mapping, ['quarter_minutes', 'period_minutes', 'q_duration_min']))
    quarters = to_int(first(rules, ['quarters', 'regulation_quarters']))
    if quarters is None:
        quarters = to_int(first(mapping, ['quarters_count', 'regulation_quarters']))
    quarters = quarters or int(settings.get('default_quarters', 4))
    source = 'explicit_json' if explicit_quarter else 'tournament_fallback'
    quarter_minutes = explicit_quarter
    if quarter_minutes is None:
        for pattern in settings.get('ten_minute_league_patterns', []):
            if re.search(pattern, tournament or '', flags=re.IGNORECASE):
                quarter_minutes = 10
                break
    if quarter_minutes is None:
        for pattern in settings.get('twelve_minute_league_patterns', []):
            if re.search(pattern, tournament or '', flags=re.IGNORECASE):
                quarter_minutes = 12
                break
    if quarter_minutes is None:
        quarter_minutes = int(settings.get('default_quarter_minutes', 10))
        source = 'default_fallback'
    regulation_minutes = to_int(first(rules, ['regulation_minutes']))
    if regulation_minutes is None:
        regulation_minutes = to_int(first(mapping, ['regulation_minutes']))
    regulation_minutes = regulation_minutes or quarters * quarter_minutes
    competition_type = str(
        first(mapping, ['competition_type', 'season_type', 'league_type'])
        or first(rules, ['competition_type'])
        or ''
    ).strip()
    format_key = str(first(rules, ['format_key']) or first(mapping, ['format_key']) or '').strip()
    if not format_key:
        format_key = f'{competition_type or "GENERIC"}_{quarters}x{quarter_minutes}'
    warnings: list[str] = []
    if source != 'explicit_json':
        warnings.append('QUARTER_MINUTES_NOT_EXPLICIT')
    if regulation_minutes != quarters * quarter_minutes:
        warnings.append('REGULATION_DURATION_INCONSISTENT')
    return {
        'quarters': quarters,
        'quarter_minutes': quarter_minutes,
        'regulation_minutes': regulation_minutes,
        'overtime_minutes': to_int(first(rules, ['overtime_minutes'])) or 5,
        'competition_type': competition_type,
        'format_key': format_key,
        'source': source,
        'warnings': warnings,
    }

def _stage(elapsed_seconds: int, full_seconds: int, quarter_seconds: int, explicit: str) -> str:
    if elapsed_seconds <= 0:
        return 'PRE_MATCH'
    half = full_seconds // 2
    after_three = quarter_seconds * 3
    if elapsed_seconds == half:
        return 'HT'
    if elapsed_seconds == after_three:
        return 'AFTER_3Q'
    if elapsed_seconds >= after_three:
        return 'Q4_CONFIRMATION'
    if elapsed_seconds < half:
        return 'EARLY_LIVE'
    if 'HT' in (explicit or '').upper():
        return 'HT'
    return 'CURRENT_Q1_Q3'

_STATUS_QUARTER_RE = re.compile("\\((\\d+)[^\\d)]*?чверть(?:\\s*(\\d+)')?\\s*\\)", re.IGNORECASE)
_STATUS_BREAK_RE = re.compile('після\\s*Q(\\d+)', re.IGNORECASE)
_STATUS_FINISHED_RE = re.compile('\\bFT\\b|FINAL|FINISHED|ENDED|ЗАВЕРШЕНО|КІНЕЦЬ', re.IGNORECASE)

def _parse_status_clock(status: str, quarter_seconds: int, full_seconds: int) -> Optional[tuple[int, Optional[int], Optional[int]]]:
    """Best-effort parser for the raw provider 'st' status string, used only as a fallback
    when the payload has no numeric match_minute_played/period fields (this feed's 'match'
    block is empty and only raw_data.main_match.st carries live-time info), e.g.:
      "Live (2-а чверть 1')"  -> mid-quarter: quarter=2, minute=1
      "Live (4-а чверть)"     -> quarter just started: quarter=4, minute=0
      "Перерва (після Q2)"    -> half-time break: exactly 2 quarters elapsed
      "Finished"              -> full game elapsed
    Returns (elapsed_seconds, period, period_played_seconds) or None if unrecognised.
    """
    text = (status or '').strip()
    if not text:
        return None
    upper = text.upper()
    if upper.startswith('LIVE'):
        match = _STATUS_QUARTER_RE.search(text)
        if match:
            period = int(match.group(1))
            played_minutes = int(match.group(2)) if match.group(2) else 0
            played_seconds = min(quarter_seconds, played_minutes * 60)
            elapsed_seconds = min(full_seconds, (period - 1) * quarter_seconds + played_seconds)
            return (elapsed_seconds, period, played_seconds)
        # "Live" but sub-stage text not recognised (e.g. overtime) - don't fall back to
        # PRE_MATCH; conservatively treat as deep in the 4th quarter.
        return (max(0, full_seconds - 1), None, None)
    match = _STATUS_BREAK_RE.search(text)
    if match:
        completed = int(match.group(1))
        return (min(full_seconds, completed * quarter_seconds), None, None)
    if _STATUS_FINISHED_RE.search(upper):
        return (full_seconds, None, None)
    return None

def _game_key(row: dict[str, Any]) -> str:
    match_id = str(first(row, ['mid', 'match_id', 'id']) or '').strip()
    if match_id:
        return 'id:' + match_id
    parts = [row.get('dt'), row.get('ht'), row.get('at'), row.get('hs'), row.get('as_')]
    return 'fallback:' + '|'.join((str(part or '') for part in parts))

def _technical(row: dict[str, Any]) -> bool:
    home = to_int(first(row, ['hs', 'home_score', 'homeScore']))
    away = to_int(first(row, ['as_', 'away_score', 'awayScore']))
    return (home, away) in {(20, 0), (0, 20)}

def _team_side(row: dict[str, Any], team: str) -> Optional[str]:
    if str(first(row, ['ht', 'home_team', 'homeTeam']) or '').strip() == team:
        return 'home'
    if str(first(row, ['at', 'away_team', 'awayTeam']) or '').strip() == team:
        return 'away'
    return None

def _raw_stat(row: dict[str, Any], side: str, metric: str, quarter: Optional[int]=None) -> Optional[float]:
    prefix = 'h' if side == 'home' else 'a'
    codes = {'FGA': 'fga', 'FGM': 'fgm', '2PA': '2pa', '2PM': '2pm', '3PA': '3pa', '3PM': '3pm', 'FTA': 'fta', 'FTM': 'ftm', 'ORB': 'orb', 'DRB': 'drb', 'TO': 'tov', 'FOULS': 'fls'}
    suffix = str(quarter) if quarter else 'm'
    value = row.get(f'{prefix}{codes[metric]}{suffix}')
    return to_number(value)

def canonical_game(row: dict[str, Any], perspective_team: Optional[str]=None, config: Optional[dict[str, Any]]=None) -> dict[str, Any]:
    home_team = str(first(row, ['ht', 'home_team', 'homeTeam']) or '')
    away_team = str(first(row, ['at', 'away_team', 'awayTeam']) or '')
    home_score = to_number(first(row, ['hs', 'home_score', 'homeScore']))
    away_score = to_number(first(row, ['as_', 'away_score', 'awayScore']))
    quarters: list[dict[str, Optional[float]]] = []
    for number in range(1, 5):
        home = to_number(first(row, [f'q{number}h', f'home_q{number}']))
        away = to_number(first(row, [f'q{number}a', f'away_q{number}']))
        total = to_number(first(row, [f'q{number}t', f'q{number}_total']))
        if total is None and home is not None and (away is not None):
            total = home + away
        quarters.append({'home': home, 'away': away, 'total': total})
    total = to_number(first(row, ['tot', 'total', 'match_total']))
    if total is None and home_score is not None and (away_score is not None):
        total = home_score + away_score
    side = _team_side(row, perspective_team) if perspective_team else None
    team_score = home_score if side == 'home' else away_score if side == 'away' else None
    opponent_score = away_score if side == 'home' else home_score if side == 'away' else None
    team_quarters = [quarter.get(side) if side else None for quarter in quarters]
    opponent_side = 'away' if side == 'home' else 'home' if side == 'away' else None
    opponent_quarters = [quarter.get(opponent_side) if opponent_side else None for quarter in quarters]
    stats: dict[str, dict[str, Optional[float]]] = {'home': {}, 'away': {}}
    quarter_stats: dict[str, list[dict[str, Optional[float]]]] = {'home': [], 'away': []}
    for game_side in ('home', 'away'):
        for metric in ('FGA', 'FGM', '2PA', '2PM', '3PA', '3PM', 'FTA', 'FTM', 'ORB', 'DRB', 'TO', 'FOULS'):
            stats[game_side][metric] = _raw_stat(row, game_side, metric)
        for number in range(1, 5):
            quarter_stats[game_side].append({metric: _raw_stat(row, game_side, metric, number) for metric in ('FGA', 'FGM', '2PA', '2PM', '3PA', '3PM', 'FTA', 'FTM', 'ORB', 'DRB', 'TO', 'FOULS')})
    game_format = _format_info(row, str(first(row, ['tour', 'tournament']) or ''), config) if config else {
        'quarters': 4,
        'quarter_minutes': None,
        'regulation_minutes': None,
        'competition_type': '',
        'format_key': 'UNKNOWN',
        'source': 'unavailable',
        'warnings': ['FORMAT_NOT_RESOLVED'],
    }
    return {'id': str(first(row, ['mid', 'match_id', 'id']) or ''), 'date': first(row, ['dt', 'date', 'start_time']), 'status': first(row, ['st', 'status']), 'tournament': first(row, ['tour', 'tournament']), 'home_team': home_team, 'away_team': away_team, 'home_score': home_score, 'away_score': away_score, 'total': total, 'quarters': quarters, 'h1_total': sum((q['total'] for q in quarters[:2] if q['total'] is not None)) if all((q['total'] is not None for q in quarters[:2])) else None, 'h2_total': sum((q['total'] for q in quarters[2:] if q['total'] is not None)) if all((q['total'] is not None for q in quarters[2:])) else None, 'perspective_team': perspective_team, 'perspective_side': side, 'team_score': team_score, 'opponent_score': opponent_score, 'team_quarters': team_quarters, 'opponent_quarters': opponent_quarters, 'stats': stats, 'quarter_stats': quarter_stats, 'format': game_format, 'raw': row}

def _parse_history_date(value: Any) -> Optional[datetime]:
    """Parse provider history dates deterministically; unknown dates sort after dated rows."""
    if value in (None, ''):
        return None
    if isinstance(value, datetime):
        return value if value.tzinfo else value.replace(tzinfo=timezone.utc)
    if isinstance(value, (int, float)):
        try:
            timestamp = float(value)
            if timestamp > 10_000_000_000:
                timestamp /= 1000.0
            return datetime.fromtimestamp(timestamp, tz=timezone.utc)
        except (ValueError, OSError, OverflowError):
            return None
    text = str(value).strip()
    if not text:
        return None
    normalized = text.replace('Z', '+00:00')
    for candidate in (normalized, normalized.replace('/', '-')):
        try:
            parsed = datetime.fromisoformat(candidate)
            return parsed if parsed.tzinfo else parsed.replace(tzinfo=timezone.utc)
        except ValueError:
            pass
    for fmt in ('%d.%m.%Y %H:%M', '%d.%m.%Y', '%d-%m-%Y %H:%M', '%d-%m-%Y', '%Y-%m-%d %H:%M:%S', '%Y-%m-%d'):
        try:
            return datetime.strptime(text, fmt).replace(tzinfo=timezone.utc)
        except ValueError:
            continue
    return None


def _sort_history_newest_first(games: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], int]:
    indexed = list(enumerate(games))
    dated = sum(_parse_history_date(game.get('date')) is not None for game in games)
    floor = datetime.min.replace(tzinfo=timezone.utc)
    indexed.sort(key=lambda pair: (_parse_history_date(pair[1].get('date')) or floor, -pair[0]), reverse=True)
    return [game for _, game in indexed], dated


def _limit_history(games: list[dict[str, Any]], limit: int=35) -> tuple[list[dict[str, Any]], int]:
    ordered, _ = _sort_history_newest_first(games)
    return ordered[:max(0, int(limit))], max(0, len(ordered) - max(0, int(limit)))


def _filter_history(rows: list[dict[str, Any]], current_id: str, team: Optional[str], config: dict[str, Any]) -> tuple[list[dict[str, Any]], dict[str, int]]:
    seen: set[str] = set()
    output: list[dict[str, Any]] = []
    counters = {'current': 0, 'technical': 0, 'duplicate': 0, 'dated': 0}
    for row in rows or []:
        row_id = str(first(row, ['mid', 'match_id', 'id']) or '')
        if current_id and row_id == current_id:
            counters['current'] += 1
            continue
        if _technical(row):
            counters['technical'] += 1
            continue
        key = _game_key(row)
        if key in seen:
            counters['duplicate'] += 1
            continue
        seen.add(key)
        output.append(canonical_game(row, team, config))
    output, dated = _sort_history_newest_first(output)
    counters['dated'] = dated
    return (output, counters)

def _apply_history_format_override(games: list[dict[str, Any]], override: dict[str, Any]) -> int:
    """Temporary fallback for legacy parsers; explicit per-game metadata remains preferred."""
    last_n = to_int(override.get('last_n_games')) or 0
    last_format = str(override.get('last_n_format') or '').strip()
    remaining_format = str(override.get('remaining_format') or '').strip()
    if last_n <= 0 or not last_format:
        return 0
    changed = 0
    newest_first = str(override.get('history_order') or 'newest_first').lower() != 'oldest_first'
    selected = set(range(min(last_n, len(games)))) if newest_first else set(range(max(0, len(games) - last_n), len(games)))
    for index, game in enumerate(games):
        key = last_format if index in selected else remaining_format
        if not key:
            continue
        match = re.search(r'(\d+)x(10|12)', key, flags=re.IGNORECASE)
        if not match:
            continue
        quarters, quarter_minutes = int(match.group(1)), int(match.group(2))
        game['format'] = {
            'quarters': quarters,
            'quarter_minutes': quarter_minutes,
            'regulation_minutes': quarters * quarter_minutes,
            'competition_type': key.rsplit('_', 1)[0],
            'format_key': key,
            'source': 'history_format_override',
            'warnings': ['LEGACY_FORMAT_OVERRIDE_USED'],
        }
        changed += 1
    return changed

def _live_stats(source: dict[str, Any], raw_main: dict[str, Any], side: str, aliases: dict[str, list[str]], actual_points: float) -> tuple[dict[str, Optional[float]], list[str]]:
    live = source.get('live_team_stats', {}).get(side, {}) or {}
    box_key = 'team_a_1h' if side == 'home' else 'team_b_1h'
    box = source.get('live_boxscore', {}).get(box_key, {}) or {}
    prefix = 'h' if side == 'home' else 'a'
    raw_codes = {'FGA': 'fgam', 'FGM': 'fgmm', '2PA': '2pam', '2PM': '2pmm', '3PA': '3pam', '3PM': '3pmm', 'FTA': 'ftam', 'FTM': 'ftmm', 'ORB': 'orbm', 'DRB': 'drbm', 'TO': 'tovm', 'FOULS': 'flsm'}
    result: dict[str, Optional[float]] = {'POINTS': actual_points}
    missing: list[str] = []
    for metric in ('FGA', 'FGM', '2PA', '2PM', '3PA', '3PM', 'FTA', 'FTM', 'ORB', 'DRB', 'TO', 'FOULS'):
        value = alias_value(live, metric, aliases)
        if value in (None, ''):
            value = alias_value(box, metric, aliases)
        if value in (None, ''):
            value = raw_main.get(prefix + raw_codes[metric])
        result[metric] = to_number(value)
        if result[metric] is None:
            missing.append(f'live_stats.{side}.{metric}')
    fga, fgm, three_made, fta, orb, turnovers = (result.get(key) for key in ('FGA', 'FGM', '3PM', 'FTA', 'ORB', 'TO'))
    result['Poss'] = fga - orb + turnovers + 0.44 * fta if None not in (fga, orb, turnovers, fta) else None
    result['eFG'] = (fgm + 0.5 * three_made) / fga if fga and fgm is not None and (three_made is not None) else None
    result['FTr'] = fta / fga if fga and fta is not None else None
    result['OffRtg'] = actual_points / result['Poss'] * 100 if result.get('Poss') else None
    result['TO_rate'] = turnovers / result['Poss'] if result.get('Poss') and turnovers is not None else None
    result['ORB_per_possession'] = orb / result['Poss'] if result.get('Poss') and orb is not None else None
    result['FTA_per_possession'] = fta / result['Poss'] if result.get('Poss') and fta is not None else None
    return (result, missing)

def _extract_bonus_context(source: dict[str, Any], raw_main: dict[str, Any]) -> dict[str, Any]:
    """Read explicit Q4 team-foul/bonus flags without inventing them from cumulative fouls."""
    containers = [source.get('live_context'), source.get('bonus_context'), source.get('match'), raw_main]
    containers = [item for item in containers if isinstance(item, dict)]
    def lookup(names: list[str]) -> Any:
        for container in containers:
            value = first(container, names)
            if value not in (None, ''):
                return value
        return None
    def boolish(value: Any) -> Optional[bool]:
        if isinstance(value, bool): return value
        if value in (None, ''): return None
        text = str(value).strip().lower()
        if text in {'1','true','yes','on','bonus','in_bonus'}: return True
        if text in {'0','false','no','off','not_in_bonus'}: return False
        return None
    home_fouls = to_int(lookup(['home_q4_team_fouls','q4_home_team_fouls','home_team_fouls_q4','home_period_fouls','hfls4']))
    away_fouls = to_int(lookup(['away_q4_team_fouls','q4_away_team_fouls','away_team_fouls_q4','away_period_fouls','afls4']))
    home_bonus = boolish(lookup(['home_in_bonus','home_bonus','q4_home_bonus','home_team_in_bonus']))
    away_bonus = boolish(lookup(['away_in_bonus','away_bonus','q4_away_bonus','away_team_in_bonus']))
    exact = home_bonus is not None or away_bonus is not None or home_fouls is not None or away_fouls is not None
    return {'home_q4_team_fouls': home_fouls, 'away_q4_team_fouls': away_fouls, 'home_in_bonus': home_bonus, 'away_in_bonus': away_bonus, 'exact_available': exact}


def adapt_match(source: dict[str, Any], config: dict[str, Any], strict: bool=False) -> dict[str, Any]:
    match = source.get('match', {}) or {}
    raw_data = source.get('raw_data', {}) or {}
    raw_main = raw_data.get('main_match', {}) or {}
    current_id = str(first(match, ['id', 'match_id']) or first(raw_main, ['mid', 'id']) or '')
    home_team = str(first(raw_main, ['ht', 'home_team']) or first(match, ['home_team', 'home']) or '')
    away_team = str(first(raw_main, ['at', 'away_team']) or first(match, ['away_team', 'away']) or '')
    if not home_team or not away_team:
        name = str(match.get('name') or '')
        split = re.split('\\s+vs\\s+', name, maxsplit=1, flags=re.IGNORECASE)
        if len(split) == 2:
            home_team, away_team = (home_team or split[0], away_team or split[1])
    tournament = str(first(match, ['tournament', 'league']) or first(raw_main, ['tour']) or '')
    format_mapping = deepcopy(match)
    if isinstance(source.get('rules'), dict) and not isinstance(format_mapping.get('rules'), dict):
        format_mapping['rules'] = deepcopy(source['rules'])
    match_format = _format_info(format_mapping, tournament, config)
    quarter_minutes = int(match_format['quarter_minutes'])
    quarter_seconds = quarter_minutes * 60
    full_seconds = int(match_format['regulation_minutes']) * 60
    home_score = to_number(match.get('score', {}).get('home'))
    away_score = to_number(match.get('score', {}).get('away'))
    if home_score is None:
        home_score = to_number(first(raw_main, ['hs', 'home_score'])) or 0.0
    if away_score is None:
        away_score = to_number(first(raw_main, ['as_', 'away_score'])) or 0.0
    explicit_stage = str(first(match, ['stage', 'status']) or first(raw_main, ['st']) or '')
    analysis_context = source.get('analysis_context', {}) if isinstance(source.get('analysis_context'), dict) else {}
    trigger_checkpoint = to_int(first(analysis_context, ['trigger_checkpoint', 'checkpoint']))
    elapsed_raw = first(match, ['match_minute_played', 'elapsed_minutes'])
    period_raw = first(match, ['period', 'quarter', 'current_quarter'])
    period_played_raw = first(match, ['period_minute_played', 'quarter_minute_played'])
    period_left_raw = first(match, ['period_minute_left', 'quarter_minute_left'])
    elapsed_minutes = to_number(elapsed_raw)
    period = to_int(period_raw)
    period_played = to_number(period_played_raw)
    period_left = to_number(period_left_raw)
    time_reliable = elapsed_raw not in (None, '') or (period is not None and (period_played_raw not in (None, '') or period_left_raw not in (None, '')))
    if elapsed_minutes is None and period and (period_played is not None):
        elapsed_minutes = (period - 1) * quarter_minutes + period_played
    # Some parser snapshots provide only the clock remaining in the current quarter.
    # Convert it to played time before calculating stage, pace and live projections.
    if elapsed_minutes is None and period and period_played is None and period_left is not None:
        safe_left = max(0.0, min(float(quarter_minutes), float(period_left)))
        period_played = float(quarter_minutes) - safe_left
        elapsed_minutes = (period - 1) * quarter_minutes + period_played
    if elapsed_minutes is None and period is None:
        # No numeric time fields in the payload at all (e.g. 'match' block empty) - fall
        # back to parsing the provider's textual status ("Live (N-а чверть M')",
        # "Перерва (після QN)", "Finished") so the match isn't misclassified as PRE_MATCH.
        status_clock = _parse_status_clock(explicit_stage, quarter_seconds, full_seconds)
        if status_clock is not None:
            status_elapsed_seconds, status_period, status_period_played_seconds = status_clock
            elapsed_minutes = status_elapsed_seconds / 60.0
            period = status_period
            if status_period_played_seconds is not None:
                period_played = status_period_played_seconds / 60.0
            time_reliable = True
    if elapsed_minutes is None:
        elapsed_minutes = 0.0
    elapsed_seconds = max(0, min(full_seconds, int(round(elapsed_minutes * 60))))
    if period is None and elapsed_seconds < full_seconds:
        period = min(4, elapsed_seconds // quarter_seconds + 1)
    if period_left is None and period:
        period_elapsed_seconds = max(0, elapsed_seconds - (period - 1) * quarter_seconds)
        period_left_seconds = max(0, quarter_seconds - period_elapsed_seconds)
    else:
        period_left_seconds = int(round((period_left or 0) * 60))
    stage = _stage(elapsed_seconds, full_seconds, quarter_seconds, explicit_stage)
    raw_game = canonical_game(raw_main, config=config)
    quarters: list[dict[str, Optional[float]]] = []
    match_quarters = match.get('quarters', {}) or {}
    for number in range(1, 5):
        q_source = match_quarters.get(f'q{number}') or match_quarters.get(f'q{number}_live') or {}
        raw_q = raw_game['quarters'][number - 1]
        q_home = to_number(q_source.get('home'))
        q_away = to_number(q_source.get('away'))
        if q_home is None:
            q_home = raw_q.get('home')
        if q_away is None:
            q_away = raw_q.get('away')
        q_total = q_home + q_away if q_home is not None and q_away is not None else raw_q.get('total')
        quarters.append({'home': q_home, 'away': q_away, 'total': q_total})
    team_a_history, count_a = _filter_history(raw_data.get('team_a_hist', []), current_id, home_team, config)
    team_b_history, count_b = _filter_history(raw_data.get('team_b_hist', []), current_id, away_team, config)
    h2h_history, count_h2h = _filter_history(raw_data.get('h2h_hist', []), current_id, None, config)
    format_override = source.get('history_format_override', {}) if isinstance(source.get('history_format_override'), dict) else {}
    override_count = sum(
        _apply_history_format_override(pool, format_override)
        for pool in (team_a_history, team_b_history, h2h_history)
    )
    current_regulation = int(match_format['regulation_minutes'])
    def split_format(pool: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
        same: list[dict[str, Any]] = []
        cross: list[dict[str, Any]] = []
        for game in pool:
            game_regulation = to_int((game.get('format') or {}).get('regulation_minutes'))
            if game_regulation == current_regulation:
                same.append(game)
            else:
                cross.append(game)
        return same, cross
    team_a_same, team_a_cross = split_format(team_a_history)
    team_b_same, team_b_cross = split_format(team_b_history)
    h2h_same, h2h_cross = split_format(h2h_history)
    history_limit = int(config.get('history', {}).get('last_n', 35))
    team_a_same, team_a_trimmed = _limit_history(team_a_same, history_limit)
    team_b_same, team_b_trimmed = _limit_history(team_b_same, history_limit)
    h2h_limit = int(config.get('history', {}).get('h2h_last_n', 5))
    h2h_same, h2h_trimmed = _limit_history(h2h_same, h2h_limit)
    aliases = config.get('aliases', {})
    home_stats, missing_home = _live_stats(source, raw_main, 'home', aliases, home_score)
    away_stats, missing_away = _live_stats(source, raw_main, 'away', aliases, away_score)
    core = ('FGA', 'FTA', 'ORB', 'TO', 'Poss', 'eFG')
    found_home = sum((home_stats.get(metric) is not None for metric in core))
    found_away = sum((away_stats.get(metric) is not None for metric in core))
    if found_home == len(core) and found_away == len(core):
        stat_support = 'ON'
    elif found_home >= 3 and found_away >= 3:
        stat_support = 'LIMITED'
    else:
        stat_support = 'OFF'
    schema_errors: list[str] = []
    if not current_id:
        schema_errors.append('match.id')
    if not home_team:
        schema_errors.append('match.home_team/raw_data.main_match.ht')
    if not away_team:
        schema_errors.append('match.away_team/raw_data.main_match.at')
    explicit_upper = explicit_stage.upper()
    explicit_live_hint = any(token in explicit_upper for token in ('LIVE', 'Q1', 'Q2', 'Q3', 'Q4', 'ЧВЕРТ', 'QUARTER'))
    if explicit_live_hint and not time_reliable:
        schema_errors.append('match.live_time')
    if strict and schema_errors:
        raise ValueError('Schema errors: ' + ', '.join(schema_errors))
    exclusions = {'current': count_a['current'] + count_b['current'] + count_h2h['current'], 'technical': count_a['technical'] + count_b['technical'] + count_h2h['technical'], 'duplicate': count_a['duplicate'] + count_b['duplicate'] + count_h2h['duplicate']}
    parser_blocks = {
        key: deepcopy(source.get(key))
        for key in (
            'history_by_exact_line',
            'scenario_patterns_by_line',
            'checkpoint_matrices',
            'stat_conditioned_line_profiles',
            'quarter_result_profile',
            'stat_alignment',
            'history_zones',
            'stat_zones',
            'projections',
        )
        if source.get(key) is not None
    }
    return {
        'match_id': current_id,
        'name': match.get('name') or f'{home_team} vs {away_team}',
        'home_team': home_team,
        'away_team': away_team,
        'tournament': tournament,
        'explicit_stage': explicit_stage,
        'stage': stage,
        'trigger_checkpoint': trigger_checkpoint,
        'current_quarter': period,
        'quarter_minutes': quarter_minutes,
        'quarter_seconds': quarter_seconds,
        'full_game_seconds': full_seconds,
        'elapsed_game_seconds': elapsed_seconds,
        'remaining_game_seconds': full_seconds - elapsed_seconds,
        'quarter_seconds_remaining': period_left_seconds,
        'clock': f'{period_left_seconds // 60:02d}:{period_left_seconds % 60:02d}' if period is not None else None,
        'score': {
            'home': home_score,
            'away': away_score,
            'total': home_score + away_score,
            'margin_home': home_score - away_score,
        },
        'quarters': quarters,
        'series_context': deepcopy(match.get('series_context', {})),
        'bonus_context': _extract_bonus_context(source, raw_main),
        'format': match_format,
        'live_stats': {'home': home_stats, 'away': away_stats},
        'stat_support': stat_support,
        'history': {'team_a': team_a_same, 'team_b': team_b_same, 'h2h': h2h_same},
        'history_cross_format': {'team_a': team_a_cross, 'team_b': team_b_cross, 'h2h': h2h_cross},
        'raw_main': raw_main,
        'parser_blocks': parser_blocks,
        'data_gate': {
            'history_team_a_n': len(team_a_same),
            'history_team_b_n': len(team_b_same),
            'pooled_n': len(team_a_same) + len(team_b_same),
            'h2h_n': len(h2h_same),
            'cross_format_team_a_n': len(team_a_cross),
            'cross_format_team_b_n': len(team_b_cross),
            'cross_format_h2h_n': len(h2h_cross),
            'cross_format_exact_hits_used': False,
            'cross_format_normalized_baseline_allowed': bool(team_a_cross or team_b_cross),
            'history_format_override_games': override_count,
            'current_match_excluded': True,
            'current_games_excluded': exclusions['current'],
            'technical_games_excluded': exclusions['technical'],
            'duplicate_games_excluded': exclusions['duplicate'],
            'strict_last_n': history_limit,
            'team_a_history_trimmed': team_a_trimmed,
            'team_b_history_trimmed': team_b_trimmed,
            'h2h_history_trimmed': h2h_trimmed,
            'dated_history_rows': count_a.get('dated', 0) + count_b.get('dated', 0) + count_h2h.get('dated', 0),
            'stats_found': stat_support != 'OFF',
            'stat_support': stat_support,
            'missing_fields': sorted(set(missing_home + missing_away)),
            'schema_errors': schema_errors,
            'time_reliable': time_reliable,
        },
    }


# ===== coursework_remaining_forecast.py =====
def _coursework_distribution(values: Iterable[float]) -> dict[str, Any]:
    rows = [float(value) for value in values if value is not None and math.isfinite(float(value))]
    if not rows:
        return {
            'n': 0, 'mean': None, 'median': None, 'standard_deviation': None,
            'p10': None, 'p25': None, 'p75': None, 'p90': None,
            'minimum': None, 'maximum': None,
        }
    return {
        'n': len(rows),
        'mean': statistics.fmean(rows),
        'median': statistics.median(rows),
        'standard_deviation': statistics.stdev(rows) if len(rows) > 1 else 0.0,
        'p10': percentile(rows, 0.10),
        'p25': percentile(rows, 0.25),
        'p75': percentile(rows, 0.75),
        'p90': percentile(rows, 0.90),
        'minimum': min(rows),
        'maximum': max(rows),
    }


def build_coursework_remaining_forecast(canonical: dict[str, Any]) -> dict[str, Any]:
    """Independent checkpoint forecast used as one conservative projection component.

    It deduplicates the already same-format canonical pools, rebuilds the score strictly
    from completed quarters when a queue checkpoint is present, and never uses points
    from the first minute of the next quarter in the prior-quarter state.
    """
    trigger = to_int(canonical.get('trigger_checkpoint'))
    if trigger in (1, 2, 3):
        completed = int(trigger)
    elif canonical.get('stage') == 'HT':
        completed = 2
    elif canonical.get('stage') in {'AFTER_3Q', 'Q4_CONFIRMATION'}:
        completed = 3
    else:
        completed = max(0, min(4, int(canonical.get('elapsed_game_seconds', 0) // canonical['quarter_seconds'])))

    quarters = canonical.get('quarters') or []
    completed_rows = quarters[:completed]
    complete_box = completed > 0 and all(
        isinstance(row, dict) and row.get('home') is not None and row.get('away') is not None
        for row in completed_rows
    )
    if complete_box:
        checkpoint_total = float(sum(float(row['home']) + float(row['away']) for row in completed_rows))
    else:
        checkpoint_total = float((canonical.get('score') or {}).get('total') or 0.0)

    unique: dict[str, dict[str, Any]] = {}
    for pool_name in ('team_a', 'team_b', 'h2h'):
        for game in (canonical.get('history', {}).get(pool_name) or []):
            game_id = str(game.get('id') or _game_key(game.get('raw') or game))
            if game_id not in unique:
                unique[game_id] = game

    def historical_segment_value(game: dict[str, Any], quarter_numbers: list[int]) -> Optional[float]:
        game_quarters = game.get('quarters') or []
        values: list[float] = []
        for number in quarter_numbers:
            if number < 1 or number > len(game_quarters):
                return None
            row = game_quarters[number - 1]
            value = to_number((row or {}).get('total'))
            if value is None:
                return None
            values.append(float(value))
        return sum(values)

    remaining_quarters = list(range(completed + 1, 5))
    remaining_values: list[float] = []
    sample_ids: list[str] = []
    for game_id, game in unique.items():
        value = historical_segment_value(game, remaining_quarters)
        if value is not None:
            remaining_values.append(value)
            sample_ids.append(game_id)
    remaining_distribution = _coursework_distribution(remaining_values)
    n = int(remaining_distribution['n'])
    readiness = 'READY' if n >= 20 else 'REVIEW_REQUIRED' if n >= 8 else 'INSUFFICIENT_DATA'
    remaining_median = to_number(remaining_distribution.get('median'))
    final_total = checkpoint_total + remaining_median if remaining_median is not None else None

    segment_projections: dict[str, Optional[float]] = {'MATCH': final_total}
    # First half total at/after Q1.
    if completed == 0:
        h1_values = [historical_segment_value(game, [1, 2]) for game in unique.values()]
        h1_dist = _coursework_distribution(value for value in h1_values if value is not None)
        segment_projections['H1'] = h1_dist.get('median')
    elif completed == 1:
        q1_total = checkpoint_total
        q2_values = [historical_segment_value(game, [2]) for game in unique.values()]
        q2_dist = _coursework_distribution(value for value in q2_values if value is not None)
        segment_projections['H1'] = q1_total + float(q2_dist['median']) if q2_dist.get('median') is not None else None

    # Second half total at HT or after Q3.
    if completed <= 2:
        h2_values = [historical_segment_value(game, [3, 4]) for game in unique.values()]
        h2_dist = _coursework_distribution(value for value in h2_values if value is not None)
        segment_projections['H2'] = h2_dist.get('median')
    elif completed == 3:
        q3 = quarters[2] if len(quarters) >= 3 else {}
        q3_total = to_number((q3 or {}).get('total'))
        q4_values = [historical_segment_value(game, [4]) for game in unique.values()]
        q4_dist = _coursework_distribution(value for value in q4_values if value is not None)
        segment_projections['H2'] = (
            float(q3_total) + float(q4_dist['median'])
            if q3_total is not None and q4_dist.get('median') is not None else None
        )

    for quarter in range(max(1, completed + 1), 5):
        values = [historical_segment_value(game, [quarter]) for game in unique.values()]
        distribution = _coursework_distribution(value for value in values if value is not None)
        segment_projections[f'Q{quarter}'] = distribution.get('median')

    return {
        'model': 'COURSEWORK_REMAINING_QUARTER_DISTRIBUTION',
        'checkpoint': completed,
        'checkpoint_total_points': checkpoint_total,
        'checkpoint_score_source': 'COMPLETED_QUARTERS' if complete_box else 'LIVE_SCORE_FALLBACK',
        'future_quarters': remaining_quarters,
        'historical_remaining_distribution': remaining_distribution,
        'forecast_final_total_points': final_total,
        'segment_projections': segment_projections,
        'sample_game_ids': sample_ids,
        'data_readiness': readiness,
        'eligible_as_projection_component': readiness in {'READY', 'REVIEW_REQUIRED'},
    }

# ===== market_parser.py =====
SUPPORTED_BUCKETS = {'match_total', 'half_total', 'quarter_total', 'team_it', 'home_ind_total', 'away_ind_total'}

def _scope_text(row: dict[str, Any]) -> str:
    return str(row.get('scope') or row.get('segment') or row.get('period') or '').upper().replace(' ', '')

def _team_from_row(bucket: str, row: dict[str, Any], canonical: dict[str, Any]) -> Optional[str]:
    if bucket == 'home_ind_total':
        return canonical['home_team']
    if bucket == 'away_ind_total':
        return canonical['away_team']
    raw = row.get('team') or row.get('team_name') or row.get('participant')
    if raw in ('home', 'HOME', 'team_a', 'A'):
        return canonical['home_team']
    if raw in ('away', 'AWAY', 'team_b', 'B'):
        return canonical['away_team']
    return str(raw) if raw else None

def _market_type(bucket: str, scope: str, team: Optional[str]) -> tuple[Optional[str], str]:
    is_team = bucket in {'team_it', 'home_ind_total', 'away_ind_total'} or team is not None
    if bucket == 'match_total' and (not is_team):
        return ('MATCH_TOTAL', 'MATCH')
    if bucket == 'half_total' and (not is_team):
        if scope.startswith('H1') or scope in {'1H', 'FIRSTHALF'}:
            return ('H1_TOTAL', 'H1')
        if scope.startswith('H2') or scope in {'2H', 'SECONDHALF'}:
            return ('H2_TOTAL', 'H2')
    if bucket == 'quarter_total' and (not is_team):
        quarter = next((q for q in ('Q1', 'Q2', 'Q3', 'Q4') if q in scope), scope)
        return ('CURRENT_QUARTER_TOTAL', quarter)
    if is_team:
        if scope.startswith('H1') or scope in {'1H', 'FIRSTHALF'}:
            return ('TEAM_IT_H1', 'H1')
        if scope.startswith('H2') or scope in {'2H', 'SECONDHALF'}:
            return ('TEAM_IT_H2', 'H2')
        quarter = next((q for q in ('Q1', 'Q2', 'Q3', 'Q4') if q in scope), None)
        if quarter:
            return ('CURRENT_QUARTER_TEAM_IT', quarter)
        return ('TEAM_IT_MATCH', 'MATCH')
    return (None, scope or 'UNKNOWN')

def _current_quarter_issue(market_type: str, segment: str, canonical: dict[str, Any]) -> Optional[str]:
    """Validate quarter markets without blocking the next quarter at a boundary.

    After Q1/HT/Q3 the bookmaker can already publish Q2/Q3/Q4 lines before the
    next quarter has a score or a running clock. Those are valid forward-looking
    markets and must be evaluated. A genuinely past or unrelated future quarter
    remains blocked.
    """
    if market_type not in {'CURRENT_QUARTER_TOTAL', 'CURRENT_QUARTER_TEAM_IT'}:
        return None
    current = to_int(canonical.get('current_quarter'))
    target = int(segment[1:]) if segment.startswith('Q') and segment[1:].isdigit() else None
    trigger = to_int(canonical.get('trigger_checkpoint'))
    expected = trigger + 1 if trigger in {1, 2, 3} else current
    if target is None:
        return 'UNKNOWN_QUARTER'
    if target < 1 or target > 4:
        return 'INVALID_QUARTER'
    # The next quarter at an explicit checkpoint is valid even before its first
    # possession, so exact current-quarter score/clock are not required yet.
    if expected is not None and target == expected:
        if current is not None and current > target:
            return 'PAST_QUARTER'
        return None
    if current is None:
        return 'NO_CURRENT_QUARTER'
    if target > current:
        return 'FUTURE_QUARTER'
    if target < current:
        return 'PAST_QUARTER'
    if canonical.get('clock') is None:
        return 'NO_EXACT_CURRENT_QUARTER_TIME'
    quarters = canonical.get('quarters') or []
    quarter_score = quarters[target - 1] if len(quarters) >= target else {}
    if quarter_score.get('total') is None:
        return 'NO_CURRENT_QUARTER_SCORE'
    return None

def parse_markets(source: dict[str, Any], canonical: dict[str, Any], config: dict[str, Any]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    # math_script.py (the actual upstream producer) writes the enriched bookmaker
    # lines dict under the top-level key 'lines' (see its "lines": bookmaker_lines
    # assembly) - not 'bookmaker_lines'/'bookmaker_markets'/'markets'. Keep those
    # older names as fallbacks too in case other producers use them.
    containers = source.get('lines') or source.get('bookmaker_lines') or source.get('bookmaker_markets') or source.get('markets') or {}
    aliases = config.get('aliases', {})
    odds_min = float(config.get('odds_min', 1.44))
    odds_max = float(config.get('odds_max', 10.0))
    evaluations: list[dict[str, Any]] = []
    audit: list[dict[str, Any]] = []
    sequence = 0
    source_meta = source.get('meta') if isinstance(source.get('meta'), dict) else {}
    root_bookmaker = str(
        source_meta.get('source')
        or source_meta.get('bookmaker')
        or source.get('bookmaker')
        or 'unknown'
    )
    for bucket, rows in containers.items():
        if isinstance(rows, dict) and bucket in SUPPORTED_BUCKETS and any(key in rows for key in ('line', 'overOdd', 'underOdd', 'over_odd', 'under_odd')):
            rows = [rows]
        if not isinstance(rows, list):
            continue
        if bucket not in SUPPORTED_BUCKETS:
            for row in rows:
                audit.append({'bucket': bucket, 'supported': False, 'reason': 'UNSUPPORTED_MARKET', 'raw': row})
            continue
        for row in rows:
            if not isinstance(row, dict):
                audit.append({'bucket': bucket, 'supported': False, 'reason': 'INVALID_MARKET_ROW', 'raw': row})
                continue
            scope = _scope_text(row)
            team = _team_from_row(bucket, row, canonical)
            market_type, segment = _market_type(bucket, scope, team)
            line = to_number(alias_value(row, 'LINE', aliases))
            real_line = bool(row.get('is_real_bookmaker_line', True))
            bookmaker = str(row.get('bookmaker') or row.get('source') or root_bookmaker or 'unknown')
            current_issue = _current_quarter_issue(market_type or '', segment, canonical)
            base_reasons: list[str] = []
            if market_type is None:
                base_reasons.append('UNSUPPORTED_MARKET')
            if line is None:
                base_reasons.append('NO_LINE')
            if not real_line:
                base_reasons.append('SYNTHETIC_LINE')
            if current_issue:
                base_reasons.append(current_issue)
            over_odds = to_number(alias_value(row, 'OVER_ODDS', aliases))
            under_odds = to_number(alias_value(row, 'UNDER_ODDS', aliases))
            audit_row = {'source_id': row.get('id'), 'bucket': bucket, 'market_type': market_type, 'team': team, 'segment': segment, 'line': line, 'over_odds': over_odds, 'under_odds': under_odds, 'bookmaker': bookmaker, 'real_line': real_line, 'issues': list(base_reasons)}
            if over_odds is None:
                audit_row['issues'].append('NO_OVER_ODDS')
            if under_odds is None:
                audit_row['issues'].append('NO_UNDER_ODDS')
            audit.append(audit_row)
            for side, odds in (('OVER', over_odds), ('UNDER', under_odds)):
                reasons = list(base_reasons)
                if odds is None:
                    reasons.append('NO_ODDS')
                if odds is not None and odds < odds_min:
                    reasons.append('ODDS_BELOW_MINIMUM')
                if odds is not None and odds > odds_max:
                    reasons.append('ODDS_ABOVE_MAXIMUM')
                sequence += 1
                safe_line = 'na' if line is None else str(line).replace('.', '_')
                market_id = str(row.get('id') or f'{bucket}_{segment}_{safe_line}_{sequence}')
                evaluations.append({'market_id': f'{market_id}_{side.lower()}_{sequence}', 'source_market_id': row.get('id'), 'market_type': market_type or 'UNSUPPORTED', 'team': team, 'segment': segment, 'side': side, 'line': line, 'odds': odds, 'bookmaker': bookmaker, 'source_bucket': bucket, 'source_scope': scope or None, 'raw_line_row': deepcopy(row), 'parser_issues': reasons, 'eligible_market': not reasons})
    return (evaluations, audit)

# ===== history_engine.py =====
def normal_cdf(value: float) -> float:
    return 0.5 * (1.0 + math.erf(value / math.sqrt(2.0)))

def settle(result: float, line: float, side: str) -> str:
    if result == line:
        return 'push'
    won = result > line if side.upper() == 'OVER' else result < line
    return 'win' if won else 'loss'

def smoothed_probability(wins: int, valid_n: int, alpha: float=1.0, beta: float=1.0) -> float:
    return (wins + alpha) / (valid_n + alpha + beta)

def _segment_value(game: dict[str, Any], market: dict[str, Any], team_name: Optional[str]=None, opponent_allowed: bool=False) -> Optional[float]:
    market_type = market['market_type']
    segment = market.get('segment', 'MATCH')
    if market_type == 'MATCH_TOTAL':
        return game.get('total')
    if market_type == 'H1_TOTAL':
        return game.get('h1_total')
    if market_type == 'H2_TOTAL':
        return game.get('h2_total')
    if market_type == 'CURRENT_QUARTER_TOTAL':
        if segment.startswith('Q') and segment[1:].isdigit():
            return game['quarters'][int(segment[1:]) - 1].get('total')
        return None
    if market_type.startswith('TEAM_IT') or market_type == 'CURRENT_QUARTER_TEAM_IT':
        if team_name:
            if game.get('home_team') == team_name:
                side = 'home'
            elif game.get('away_team') == team_name:
                side = 'away'
            else:
                return None
            if opponent_allowed:
                side = 'away' if side == 'home' else 'home'
            if segment == 'MATCH':
                return game.get('home_score') if side == 'home' else game.get('away_score')
            if segment == 'H1':
                values = [game['quarters'][i].get(side) for i in (0, 1)]
                return sum(values) if all((value is not None for value in values)) else None
            if segment == 'H2':
                values = [game['quarters'][i].get(side) for i in (2, 3)]
                return sum(values) if all((value is not None for value in values)) else None
            if segment.startswith('Q') and segment[1:].isdigit():
                return game['quarters'][int(segment[1:]) - 1].get(side)
        return None
    return None

def exact_breakdown(values: list[Optional[float]], line: float, side: str, alpha: float, beta: float) -> dict[str, Any]:
    valid = [float(value) for value in values if value is not None]
    results = [settle(value, line, side) for value in valid]
    wins = results.count('win')
    losses = results.count('loss')
    pushes = results.count('push')
    n = len(valid)
    return {'wins': wins, 'losses': losses, 'pushes': pushes, 'n': n, 'valid_n': n, 'raw_pct': wins / n if n else None, 'raw_hit_pct': wins / n if n else None, 'p_smoothed': smoothed_probability(wins, n, alpha, beta) if n else None, 'values': valid}

def _distribution(values: list[float], line: float, side: str, min_normal_n: int, alpha: float, beta: float) -> dict[str, Any]:
    if not values:
        return {'n': 0, 'available': False, 'p_distribution': None}
    exact = exact_breakdown(values, line, side, alpha, beta)
    mean = statistics.fmean(values)
    median = statistics.median(values)
    standard_deviation = statistics.stdev(values) if len(values) > 1 else 0.0
    empirical = exact['p_smoothed']
    normal_probability = None
    if len(values) >= min_normal_n and standard_deviation > 0:
        z = (line - mean) / standard_deviation
        normal_probability = 1.0 - normal_cdf(z) if side == 'OVER' else normal_cdf(z)
    available = [probability for probability in (empirical, normal_probability) if probability is not None]
    probability = sum(available) / len(available)
    return {'n': len(values), 'available': True, 'mean': mean, 'median': median, 'standard_deviation': standard_deviation, 'empirical_percentile_line': sum((value <= line for value in values)) / len(values), 'normal_cdf_probability': normal_probability, 'p_distribution': probability}

def _weighted_available(components: dict[str, Optional[float]], weights: dict[str, float]) -> tuple[float, dict[str, float]]:
    active = {key: float(weights[key]) for key, value in components.items() if value is not None and key in weights and (weights[key] > 0)}
    total = sum(active.values())
    if not active or total <= 0:
        return (0.5, {})
    normalized = {key: weight / total for key, weight in active.items()}
    return (sum((float(components[key]) * normalized[key] for key in normalized)), normalized)

def _team_scored_allowed_values(game: dict[str, Any], market: dict[str, Any], team: str) -> tuple[Optional[float], Optional[float]]:
    side = _side_for_team(game, team)
    if side is None:
        return None, None
    opponent_side = 'away' if side == 'home' else 'home'
    segment = market.get('segment')
    if segment == 'MATCH':
        return to_number(game.get(f'{side}_score')), to_number(game.get(f'{opponent_side}_score'))
    quarters = game.get('quarters') or []
    indices = {'H1':[0,1], 'H2':[2,3], 'Q1':[0], 'Q2':[1], 'Q3':[2], 'Q4':[3]}.get(segment, [])
    if not indices or any(i >= len(quarters) for i in indices):
        return None, None
    own = [to_number(quarters[i].get(side)) for i in indices]
    allowed = [to_number(quarters[i].get(opponent_side)) for i in indices]
    if any(v is None for v in own + allowed):
        return None, None
    return sum(own), sum(allowed)


def calculate_total_scored_allowed(market: dict[str, Any], canonical: dict[str, Any], config: dict[str, Any]) -> dict[str, Any]:
    """Cross-match attack/defence interaction for non-Team-IT totals."""
    alpha = float(config['smoothing']['alpha']); beta = float(config['smoothing']['beta'])
    line = float(market['line']); side = market['side']
    home, away = canonical['home_team'], canonical['away_team']
    a_own=[]; a_allowed=[]; b_own=[]; b_allowed=[]
    for game in canonical['history']['team_a']:
        own, allowed = _team_scored_allowed_values(game, market, home)
        if own is not None: a_own.append(own)
        if allowed is not None: a_allowed.append(allowed)
    for game in canonical['history']['team_b']:
        own, allowed = _team_scored_allowed_values(game, market, away)
        if own is not None: b_own.append(own)
        if allowed is not None: b_allowed.append(allowed)
    expected_home = None
    expected_away = None
    if a_own and b_allowed: expected_home = 0.55 * statistics.mean(a_own) + 0.45 * statistics.mean(b_allowed)
    if b_own and a_allowed: expected_away = 0.55 * statistics.mean(b_own) + 0.45 * statistics.mean(a_allowed)
    expected_total = expected_home + expected_away if expected_home is not None and expected_away is not None else None
    sigma = statistics.pstdev([v for v in a_own+a_allowed+b_own+b_allowed]) if len(a_own+a_allowed+b_own+b_allowed) >= 2 else 0.0
    probability = None
    if expected_total is not None:
        effective_sigma = max(float(config.get('history', {}).get('scored_allowed_sigma_floor', 8.0)), sigma)
        z = (expected_total - line) / effective_sigma
        probability = normal_cdf(z) if side == 'OVER' else normal_cdf(-z)
    return {
        'available': probability is not None,
        'expected_home': expected_home, 'expected_away': expected_away, 'expected_total': expected_total,
        'p_scored_allowed': probability,
        'samples': {'home_scored':len(a_own),'home_allowed':len(a_allowed),'away_scored':len(b_own),'away_allowed':len(b_allowed)},
    }


def calculate_total_history(market: dict[str, Any], canonical: dict[str, Any], config: dict[str, Any]) -> dict[str, Any]:
    alpha = float(config['smoothing']['alpha'])
    beta = float(config['smoothing']['beta'])
    line = float(market['line'])
    side = market['side']
    team_a_values = [_segment_value(game, market) for game in canonical['history']['team_a']]
    team_b_values = [_segment_value(game, market) for game in canonical['history']['team_b']]
    h2h_values = [_segment_value(game, market) for game in canonical['history']['h2h']]
    team_a = exact_breakdown(team_a_values, line, side, alpha, beta)
    team_b = exact_breakdown(team_b_values, line, side, alpha, beta)
    pooled_values = [value for value in team_a_values + team_b_values if value is not None]
    pooled = exact_breakdown(pooled_values, line, side, alpha, beta)
    h2h = exact_breakdown(h2h_values, line, side, alpha, beta)
    pooled_probability = pooled['p_smoothed'] if pooled['p_smoothed'] is not None else 0.5
    h2h_k = float(config['credibility']['h2h_k'])
    h2h_credibility = h2h['n'] / (h2h['n'] + h2h_k) if h2h['n'] else 0.0
    h2h_probability = h2h['p_smoothed'] if h2h['p_smoothed'] is not None else pooled_probability
    h2h['credibility'] = h2h_credibility
    h2h['p_shrunk'] = h2h_credibility * h2h_probability + (1 - h2h_credibility) * pooled_probability
    last5_values = [value for value in team_a_values[:5] + team_b_values[:5] if value is not None]
    last5 = exact_breakdown(last5_values, line, side, alpha, beta)
    form_k = float(config['credibility']['form_k'])
    form_credibility = last5['n'] / (last5['n'] + form_k) if last5['n'] else 0.0
    form_probability = last5['p_smoothed'] if last5['p_smoothed'] is not None else pooled_probability
    last5['credibility'] = form_credibility
    last5['p_shrunk'] = form_credibility * form_probability + (1 - form_credibility) * pooled_probability
    distribution = _distribution(pooled_values, line, side, int(config['credibility']['normal_min_sample']), alpha, beta)
    scored_allowed = calculate_total_scored_allowed(market, canonical, config)
    components = {'exact': pooled['p_smoothed'], 'form': last5.get('p_shrunk') if last5['n'] else None, 'h2h': h2h.get('p_shrunk') if h2h['n'] else None, 'distribution': distribution.get('p_distribution'), 'scored_allowed': scored_allowed.get('p_scored_allowed')}
    p_hist, normalized = _weighted_available(components, config['history_weights'])
    for block in (team_a, team_b, pooled, h2h, last5):
        block.pop('values', None)
    history_zone_rate = pooled.get('raw_pct')
    return {
        'team_a': team_a, 'team_b': team_b, 'pooled': pooled, 'h2h': h2h,
        'last5': last5, 'distribution': distribution, 'scored_allowed': scored_allowed,
        'components': components, 'component_weights': normalized, 'p_hist': p_hist,
        # The user-facing history gate is the exact-line pooled raw hit zone.
        # P_hist may include form/distribution/scored-allowed, but it may not
        # replace the mandatory 75% exact historical zone.
        'history_zone_rate': history_zone_rate,
        'history_zone_hits': pooled.get('wins'),
        'history_zone_n': pooled.get('n'),
        'history_zone_source': 'POOLED_EXACT_LINE_RAW',
    }

def _current_team_score(canonical: dict[str, Any], team: str, segment: str) -> float:
    side = 'home' if team == canonical['home_team'] else 'away'
    if segment == 'MATCH':
        return float(canonical['score'][side])
    if segment == 'H1':
        return sum((float(q.get(side) or 0) for q in canonical['quarters'][:2]))
    if segment == 'H2':
        return sum((float(q.get(side) or 0) for q in canonical['quarters'][2:]))
    if segment.startswith('Q') and segment[1:].isdigit():
        return float(canonical['quarters'][int(segment[1:]) - 1].get(side) or 0)
    return 0.0

def calculate_team_it_history(market: dict[str, Any], canonical: dict[str, Any], config: dict[str, Any]) -> dict[str, Any]:
    alpha = float(config['smoothing']['alpha'])
    beta = float(config['smoothing']['beta'])
    team = market['team']
    opponent = canonical['away_team'] if team == canonical['home_team'] else canonical['home_team']
    own_pool = canonical['history']['team_a'] if team == canonical['home_team'] else canonical['history']['team_b']
    opponent_pool = canonical['history']['team_b'] if team == canonical['home_team'] else canonical['history']['team_a']
    own_values = [_segment_value(game, market, team) for game in own_pool]
    allowed_values = [_segment_value(game, market, opponent, opponent_allowed=True) for game in opponent_pool]
    h2h_values = [_segment_value(game, market, team) for game in canonical['history']['h2h']]
    own = exact_breakdown(own_values, float(market['line']), market['side'], alpha, beta)
    allowed = exact_breakdown(allowed_values, float(market['line']), market['side'], alpha, beta)
    h2h = exact_breakdown(h2h_values, float(market['line']), market['side'], alpha, beta)
    weights = config['team_it']
    components = {'own_scored': own['p_smoothed'], 'opponent_allowed': allowed['p_smoothed'], 'h2h_it': h2h['p_smoothed']}
    configured = {'own_scored': float(weights['own_weight']), 'opponent_allowed': float(weights['opponent_allowed_weight']), 'h2h_it': float(weights['h2h_weight'])}
    p_hist, normalized = _weighted_available(components, configured)
    raw_zone_components = {
        'own_scored': own.get('raw_pct'),
        'opponent_allowed': allowed.get('raw_pct'),
        'h2h_it': h2h.get('raw_pct') if h2h.get('n') else None,
    }
    history_zone_rate, history_zone_weights = _weighted_available(raw_zone_components, configured)
    current = _current_team_score(canonical, team, market.get('segment', 'MATCH'))
    required = max(0.0, float(market['line']) - current)
    if market.get('segment') == 'MATCH':
        remaining_minutes = canonical['remaining_game_seconds'] / 60
    elif market.get('segment') == 'H1':
        remaining_minutes = max(0.0, canonical['full_game_seconds'] / 2 - canonical['elapsed_game_seconds']) / 60
    elif market.get('segment') == 'H2':
        half_start = canonical['full_game_seconds'] / 2
        elapsed_half = max(0.0, canonical['elapsed_game_seconds'] - half_start)
        remaining_minutes = max(0.0, half_start - elapsed_half) / 60
    else:
        remaining_minutes = canonical['quarter_seconds_remaining'] / 60
    required_ppm = required / remaining_minutes if remaining_minutes > 0 else None
    live_side = 'home' if team == canonical['home_team'] else 'away'
    poss = canonical['live_stats'][live_side].get('Poss')
    elapsed_minutes = canonical['elapsed_game_seconds'] / 60
    possessions_per_minute = poss / elapsed_minutes if poss and elapsed_minutes > 0 else None
    remaining_possessions = possessions_per_minute * remaining_minutes if possessions_per_minute else None
    required_ppp = required / remaining_possessions if remaining_possessions and remaining_possessions > 0 else None
    weakest_values = [value for key, value in components.items() if key != 'h2h_it' and value is not None]
    weakest = min(weakest_values) if weakest_values else None
    for block in (own, allowed, h2h):
        block.pop('values', None)
    return {
        'team_a': own if team == canonical['home_team'] else {},
        'team_b': own if team == canonical['away_team'] else {},
        'pooled': {}, 'h2h': h2h, 'last5': {}, 'distribution': {},
        'own_scored': own, 'opponent_allowed': allowed, 'h2h_it': h2h,
        'opponent': opponent, 'components': components, 'component_weights': normalized,
        'weakest_gate': weakest, 'required_live': required,
        'remaining_minutes': remaining_minutes,
        'required_points_per_minute': required_ppm,
        'required_points_per_possession': required_ppp,
        'p_hist_IT': p_hist, 'p_hist': p_hist,
        'history_zone_rate': history_zone_rate,
        'history_zone_weights': history_zone_weights,
        'history_zone_source': 'TEAM_IT_OWN_ALLOWED_H2H_RAW',
    }

def calculate_history(market: dict[str, Any], canonical: dict[str, Any], config: dict[str, Any]) -> dict[str, Any]:
    if market['market_type'].startswith('TEAM_IT') or market['market_type'] == 'CURRENT_QUARTER_TEAM_IT':
        return calculate_team_it_history(market, canonical, config)
    return calculate_total_history(market, canonical, config)

def segment_value(game: dict[str, Any], market: dict[str, Any], team_name: Optional[str]=None) -> Optional[float]:
    return _segment_value(game, market, team_name)

# ===== scenario_engine.py =====
PatternMatcher = Callable[[dict[str, Any]], bool]

def _team_state(canonical: dict[str, Any], team: str) -> dict[str, Any]:
    side = 'home' if team == canonical['home_team'] else 'away'
    opponent = 'away' if side == 'home' else 'home'
    team_q = [quarter.get(side) for quarter in canonical['quarters']]
    opp_q = [quarter.get(opponent) for quarter in canonical['quarters']]
    elapsed = canonical['elapsed_game_seconds']
    q_seconds = canonical['quarter_seconds']

    # A checkpoint job is usually queued when the provider already shows the first
    # minute of the next quarter (Q2/Q3/Q4). In that snapshot elapsed % q_seconds
    # is no longer zero, so relying only on exact clock boundaries silently disables
    # PATTERN_13..15 after Q1/HT/Q3. The queue source is authoritative here.
    trigger_checkpoint = to_int(canonical.get('trigger_checkpoint'))
    checkpoint_boundary = trigger_checkpoint in (1, 2, 3)
    if checkpoint_boundary:
        completed = int(trigger_checkpoint)
        boundary = True
    else:
        completed = min(4, elapsed // q_seconds)
        boundary = elapsed % q_seconds == 0

    # Do not contaminate an after-quarter scenario with points already scored in
    # the next quarter. When the completed-quarter boxscore is available, rebuild
    # the checkpoint score strictly from Q1..Qn; otherwise preserve the live-score
    # fallback used by the previous implementation.
    checkpoint_team_values = team_q[:int(completed)]
    checkpoint_opp_values = opp_q[:int(completed)]
    checkpoint_boxscore_complete = (
        int(completed) > 0
        and all(value is not None for value in checkpoint_team_values + checkpoint_opp_values)
    )
    if boundary and checkpoint_boxscore_complete:
        score = float(sum(checkpoint_team_values))
        opponent_score = float(sum(checkpoint_opp_values))
        total = score + opponent_score
        checkpoint_score_source = 'COMPLETED_QUARTERS'
    else:
        score = canonical['score'][side]
        opponent_score = canonical['score'][opponent]
        total = canonical['score']['total']
        checkpoint_score_source = 'LIVE_SCORE_FALLBACK'

    return {
        'side': side,
        'team_q': team_q,
        'opp_q': opp_q,
        'completed_quarters': int(completed),
        'at_boundary': boundary,
        'current_quarter': canonical.get('current_quarter'),
        'score': score,
        'opponent_score': opponent_score,
        'total': total,
        'checkpoint_score_source': checkpoint_score_source,
    }

def _game_margin(game: dict[str, Any], after_quarters: int) -> Optional[float]:
    team_values = game['team_quarters'][:after_quarters]
    opponent_values = game['opponent_quarters'][:after_quarters]
    if not all((value is not None for value in team_values + opponent_values)):
        return None
    return sum(team_values) - sum(opponent_values)

def _game_total(game: dict[str, Any], after_quarters: int) -> Optional[float]:
    values = [quarter.get('total') for quarter in game['quarters'][:after_quarters]]
    return sum(values) if all((value is not None for value in values)) else None

def _bucket(value: float, size: float) -> tuple[float, float]:
    low = value // size * size
    return (low, low + size)

def _within_bucket(value: Optional[float], bounds: tuple[float, float]) -> bool:
    return value is not None and bounds[0] <= value < bounds[1]

def _active_patterns(
    canonical: dict[str, Any],
    team: str,
    config: dict[str, Any],
    market: Optional[dict[str, Any]]=None,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    state = _team_state(canonical, team)
    team_q, opp_q = (state['team_q'], state['opp_q'])
    active: list[dict[str, Any]] = []
    rejected: list[dict[str, Any]] = []

    def add(pattern_id: str, name: str, group: str, condition: bool, matcher: PatternMatcher, reason: str='CURRENT_CONDITION_FALSE') -> None:
        record = {'pattern_id': pattern_id, 'name': name, 'pattern_group': group, 'team': team, 'matcher': matcher}
        if condition:
            active.append(record)
        else:
            rejected.append({key: value for key, value in record.items() if key != 'matcher'} | {'rejection_reason': reason})
    completed_indices = range(state['completed_quarters'])
    won_any = any((team_q[i] is not None and opp_q[i] is not None and (team_q[i] > opp_q[i]) for i in completed_indices))
    scored_21 = any((team_q[i] is not None and team_q[i] >= 21 for i in completed_indices))
    add('PATTERN_01', 'WON_AT_LEAST_ONE_QUARTER', 'quarter_strength', won_any, lambda game: any((a is not None and b is not None and (a > b) for a, b in zip(game['team_quarters'], game['opponent_quarters']))))
    add('PATTERN_02', 'SCORED_21_PLUS_IN_ANY_QUARTER', 'quarter_strength', scored_21, lambda game: any((value is not None and value >= 21 for value in game['team_quarters'])))
    q1_known = team_q[0] is not None and opp_q[0] is not None and (canonical['elapsed_game_seconds'] >= canonical['quarter_seconds'])
    add('PATTERN_03', 'SCORED_18_PLUS_IN_Q1', 'quarter_strength', q1_known and team_q[0] >= 18, lambda game: game['team_quarters'][0] is not None and game['team_quarters'][0] >= 18)
    add('PATTERN_04', 'WON_Q1', 'quarter_result', q1_known and team_q[0] > opp_q[0], lambda game: None not in (game['team_quarters'][0], game['opponent_quarters'][0]) and game['team_quarters'][0] > game['opponent_quarters'][0])
    add('PATTERN_05', 'WON_Q1_AND_SCORED_18_PLUS', 'quarter_result', q1_known and team_q[0] > opp_q[0] and (team_q[0] >= 18), lambda game: None not in (game['team_quarters'][0], game['opponent_quarters'][0]) and game['team_quarters'][0] > game['opponent_quarters'][0] and (game['team_quarters'][0] >= 18))
    add('PATTERN_06', 'LOST_Q1', 'quarter_result', q1_known and team_q[0] < opp_q[0], lambda game: None not in (game['team_quarters'][0], game['opponent_quarters'][0]) and game['team_quarters'][0] < game['opponent_quarters'][0])
    add('PATTERN_07', 'LED_AFTER_Q1', 'score_state', q1_known and team_q[0] > opp_q[0], lambda game: (_game_margin(game, 1) or 0) > 0)
    add('PATTERN_08', 'TRAILED_AFTER_Q1', 'score_state', q1_known and team_q[0] < opp_q[0], lambda game: (_game_margin(game, 1) or 0) < 0)
    ht_known = canonical['elapsed_game_seconds'] >= canonical['quarter_seconds'] * 2 and all((value is not None for value in team_q[:2] + opp_q[:2]))
    current_ht_margin = sum(team_q[:2]) - sum(opp_q[:2]) if ht_known else 0
    add('PATTERN_09', 'LED_AT_HT', 'score_state', ht_known and current_ht_margin > 0, lambda game: (_game_margin(game, 2) or 0) > 0)
    add('PATTERN_10', 'TRAILED_AT_HT', 'score_state', ht_known and current_ht_margin < 0, lambda game: (_game_margin(game, 2) or 0) < 0)

    # Line-relative quarter scanners required by v5.5. They are active only
    # when the current market itself is a quarter total/quarter Team IT, so a
    # 21+ observation can never leak into an unrelated full-match line.
    line = to_number((market or {}).get('line'))
    side = str((market or {}).get('side') or '').upper()
    market_type = str((market or {}).get('market_type') or '')
    if line is not None and market_type in {'CURRENT_QUARTER_TOTAL', 'CURRENT_QUARTER_TEAM_IT'}:
        threshold = math.floor(line) + 1 if side == 'OVER' else math.floor(line)
        comparison = (lambda value, target=threshold: value is not None and value >= target) if side == 'OVER' else (lambda value, target=threshold: value is not None and value <= target)
        if market_type == 'CURRENT_QUARTER_TEAM_IT':
            current_values = [team_q[i] for i in completed_indices]
            opponent_values = [opp_q[i] for i in completed_indices]
            add(
                'PATTERN_21', 'TEAM_HIT_EXACT_QUARTER_LINE_BEFORE', 'line_threshold',
                any(comparison(value) for value in current_values),
                lambda game, predicate=comparison: any(predicate(value) for value in game['team_quarters']),
            )
            add(
                'PATTERN_22', 'OPPONENT_ALLOWED_EXACT_QUARTER_LINE_BEFORE', 'allowed_threshold',
                any(comparison(value) for value in opponent_values),
                lambda game, predicate=comparison: any(predicate(value) for value in game['opponent_quarters']),
            )
        else:
            current_totals = [
                team_q[i] + opp_q[i]
                for i in completed_indices
                if team_q[i] is not None and opp_q[i] is not None
            ]
            add(
                'PATTERN_21', 'MATCH_HIT_EXACT_QUARTER_LINE_BEFORE', 'line_threshold',
                any(comparison(value) for value in current_totals),
                lambda game, predicate=comparison: any(
                    predicate(quarter.get('total')) for quarter in game['quarters']
                ),
            )

    # Preserve the named quarter/result and sweep-state family. The current
    # condition selects the historical subset; the exact bookmaker line is
    # still evaluated afterwards by segment_value/exact_breakdown.
    team_won_match = lambda game: (
        game.get('team_score') is not None
        and game.get('opponent_score') is not None
        and game['team_score'] > game['opponent_score']
    )
    add(
        'PATTERN_23', 'WON_Q1_AND_WON_MATCH', 'match_result',
        q1_known and team_q[0] > opp_q[0],
        lambda game: (
            None not in (game['team_quarters'][0], game['opponent_quarters'][0])
            and game['team_quarters'][0] > game['opponent_quarters'][0]
            and team_won_match(game)
        ),
    )
    add(
        'PATTERN_24', 'WON_Q1_WITH_23_PLUS_AND_WON_MATCH', 'match_result',
        q1_known and team_q[0] > opp_q[0] and team_q[0] >= 23,
        lambda game: (
            None not in (game['team_quarters'][0], game['opponent_quarters'][0])
            and game['team_quarters'][0] > game['opponent_quarters'][0]
            and game['team_quarters'][0] >= 23
            and team_won_match(game)
        ),
    )
    completed = state['completed_quarters']
    if completed == 2:
        won_first = all(team_q[i] is not None and opp_q[i] is not None and team_q[i] > opp_q[i] for i in range(2))
        lost_first = all(team_q[i] is not None and opp_q[i] is not None and team_q[i] < opp_q[i] for i in range(2))
        add('PATTERN_25', 'AFTER_WON_Q1_Q2', 'sequence_state', won_first, lambda game: all(a is not None and b is not None and a > b for a, b in zip(game['team_quarters'][:2], game['opponent_quarters'][:2])))
        add('PATTERN_26', 'AFTER_LOST_Q1_Q2_ANTI_SWEEP', 'sequence_state', lost_first, lambda game: all(a is not None and b is not None and a < b for a, b in zip(game['team_quarters'][:2], game['opponent_quarters'][:2])))
    elif completed >= 3:
        won_first = all(team_q[i] is not None and opp_q[i] is not None and team_q[i] > opp_q[i] for i in range(3))
        lost_first = all(team_q[i] is not None and opp_q[i] is not None and team_q[i] < opp_q[i] for i in range(3))
        add('PATTERN_27', 'AFTER_WON_Q1_Q2_Q3_SWEEP_PATH', 'sequence_state', won_first, lambda game: all(a is not None and b is not None and a > b for a, b in zip(game['team_quarters'][:3], game['opponent_quarters'][:3])))
        add('PATTERN_28', 'AFTER_LOST_Q1_Q2_Q3_ANTI_SWEEP', 'sequence_state', lost_first, lambda game: all(a is not None and b is not None and a < b for a, b in zip(game['team_quarters'][:3], game['opponent_quarters'][:3])))

    opponent_allowed_21 = any(opp_q[i] is not None and opp_q[i] >= 21 for i in completed_indices)
    add('PATTERN_29', 'OPPONENT_ALLOWED_21_PLUS_IN_A_QUARTER', 'allowed_threshold', opponent_allowed_21, lambda game: any(value is not None and value >= 21 for value in game['opponent_quarters']))
    scored_23 = any(team_q[i] is not None and team_q[i] >= 23 for i in completed_indices)
    add('PATTERN_30', 'SCORED_23_PLUS_IN_A_QUARTER', 'quarter_strength', scored_23, lambda game: any(value is not None and value >= 23 for value in game['team_quarters']))
    if q1_known:
        bounds = _bucket(team_q[0] - opp_q[0], float(config['patterns']['margin_bucket_size']))
        add('PATTERN_11', 'Q1_MARGIN_BUCKET', 'margin_state', True, lambda game, b=bounds: _within_bucket(_game_margin(game, 1), b))
    else:
        rejected.append({'pattern_id': 'PATTERN_11', 'name': 'Q1_MARGIN_BUCKET', 'pattern_group': 'margin_state', 'team': team, 'rejection_reason': 'Q1_NOT_COMPLETE'})
    if ht_known:
        bounds = _bucket(current_ht_margin, float(config['patterns']['margin_bucket_size']))
        add('PATTERN_12', 'HT_MARGIN_BUCKET', 'margin_state', True, lambda game, b=bounds: _within_bucket(_game_margin(game, 2), b))
    else:
        rejected.append({'pattern_id': 'PATTERN_12', 'name': 'HT_MARGIN_BUCKET', 'pattern_group': 'margin_state', 'team': team, 'rejection_reason': 'HT_NOT_AVAILABLE'})
    if state['at_boundary'] and state['completed_quarters'] in (1, 2, 3):
        checkpoint = state['completed_quarters']
        margin_bounds = _bucket(state['score'] - state['opponent_score'], float(config['patterns']['margin_bucket_size']))
        total_bounds = _bucket(state['total'], float(config['patterns']['total_bucket_size']))
        score_bounds = _bucket(state['score'], float(config['patterns']['team_score_bucket_size']))
        add('PATTERN_13', 'CURRENT_MARGIN_BUCKET', 'margin_state', True, lambda game, n=checkpoint, b=margin_bounds: _within_bucket(_game_margin(game, n), b))
        add('PATTERN_14', 'CURRENT_TOTAL_BUCKET', 'total_state', True, lambda game, n=checkpoint, b=total_bounds: _within_bucket(_game_total(game, n), b))
        add('PATTERN_15', 'CURRENT_TEAM_SCORE_BUCKET', 'total_state', True, lambda game, n=checkpoint, b=score_bounds: _within_bucket(sum(game['team_quarters'][:n]) if all((value is not None for value in game['team_quarters'][:n])) else None, b))
        rejected.append({'pattern_id': 'PATTERN_16', 'name': 'SAME_STAGE', 'pattern_group': 'time_state', 'team': team, 'rejection_reason': 'HISTORICAL_STAGE_SNAPSHOT_NOT_AVAILABLE'})
        rejected.append({'pattern_id': 'PATTERN_17', 'name': 'SAME_QUARTER_NUMBER', 'pattern_group': 'time_state', 'team': team, 'rejection_reason': 'HISTORICAL_STAGE_SNAPSHOT_NOT_AVAILABLE'})
    else:
        for pattern_id, name, group in (('PATTERN_13', 'CURRENT_MARGIN_BUCKET', 'margin_state'), ('PATTERN_14', 'CURRENT_TOTAL_BUCKET', 'total_state'), ('PATTERN_15', 'CURRENT_TEAM_SCORE_BUCKET', 'total_state'), ('PATTERN_16', 'SAME_STAGE', 'time_state'), ('PATTERN_17', 'SAME_QUARTER_NUMBER', 'time_state'), ('PATTERN_18', 'SAME_MINUTE_BUCKET', 'time_state')):
            rejected.append({'pattern_id': pattern_id, 'name': name, 'pattern_group': group, 'team': team, 'rejection_reason': 'HISTORICAL_CHECKPOINT_NOT_AVAILABLE'})
    for pattern_id, name in (('PATTERN_19', 'FAVORITE_LEADS'), ('PATTERN_20', 'FAVORITE_TRAILS')):
        rejected.append({'pattern_id': pattern_id, 'name': name, 'pattern_group': 'favorite_state', 'team': team, 'rejection_reason': 'HISTORICAL_CLOSING_HANDICAP_NOT_AVAILABLE'})
    return (active, rejected)

def calculate_scenario(market: dict[str, Any], canonical: dict[str, Any], history: dict[str, Any], config: dict[str, Any]) -> dict[str, Any]:
    teams = [market['team']] if market.get('team') else [canonical['home_team'], canonical['away_team']]
    patterns_found: list[dict[str, Any]] = []
    patterns_rejected: list[dict[str, Any]] = []
    alpha = float(config['smoothing']['alpha'])
    beta = float(config['smoothing']['beta'])
    pattern_k = float(config['credibility']['pattern_k'])
    minimum = int(config['credibility']['pattern_min_sample'])
    p_hist = float(history['p_hist'])
    for team in teams:
        active, inactive = _active_patterns(canonical, team, config, market)
        patterns_rejected.extend(inactive)
        pool = canonical['history']['team_a'] if team == canonical['home_team'] else canonical['history']['team_b']
        for pattern in active:
            matched = [game for game in pool if pattern['matcher'](game)]
            values = [segment_value(game, market, team if market.get('team') else None) for game in matched]
            valid_outcomes = [float(value) for value in values if value is not None]
            breakdown = exact_breakdown(values, float(market['line']), market['side'], alpha, beta)
            credibility = breakdown['n'] / (breakdown['n'] + pattern_k) if breakdown['n'] else 0.0
            neutral = 0.50
            smoothed = breakdown['p_smoothed'] if breakdown['p_smoothed'] is not None else neutral
            # Scenario must be independent from P_hist. Small samples shrink to
            # neutral 50%, never back to the historical line probability.
            shrunk = credibility * smoothed + (1 - credibility) * neutral
            coverage = breakdown['n'] / len(pool) if pool else 0.0
            coverage_cap_applied = coverage < 0.10 and shrunk > neutral
            if coverage_cap_applied:
                shrunk = neutral
            sample_quality = min(1.0, breakdown['n'] / 10.0)
            specificity = float(config['patterns']['specificity'].get(pattern['pattern_id'], 0.7))
            distance_quality = 1.0
            rank = credibility * sample_quality * specificity * distance_quality
            result = {
                'pattern_id': pattern['pattern_id'],
                'name': pattern['name'],
                'pattern_group': pattern['pattern_group'],
                'team': team,
                'matched_games': breakdown['n'],
                'market_hits': breakdown['wins'],
                'market_losses': breakdown['losses'],
                'pushes': breakdown['pushes'],
                'raw_hit_pct': breakdown['raw_pct'],
                'smoothed_probability': breakdown['p_smoothed'],
                'credibility': credibility,
                'shrunk_probability': shrunk,
                'sample_quality': sample_quality,
                'sample_tier': 'INFORMATIONAL' if breakdown['n'] < 3 else 'CAUTION' if breakdown['n'] < 6 else 'USABLE' if breakdown['n'] < 10 else 'STRONG',
                'coverage': coverage,
                'coverage_cap_applied': coverage_cap_applied,
                'specificity': specificity,
                'distance_match_quality': distance_quality,
                'pattern_rank': rank,
                'outcome_mean': statistics.fmean(valid_outcomes) if valid_outcomes else None,
                'outcome_median': statistics.median(valid_outcomes) if valid_outcomes else None,
                'outcome_standard_deviation': statistics.stdev(valid_outcomes) if len(valid_outcomes) > 1 else None,
                'used_in_scenario': False,
                'rejection_reason': None,
            }
            if breakdown['n'] < minimum:
                result['rejection_reason'] = 'SAMPLE_BELOW_PATTERN_MINIMUM'
                patterns_rejected.append(result)
            else:
                patterns_found.append(result)
    # A group is de-duplicated within one team only. For a match-total market,
    # home and away histories are independent pools and both must contribute;
    # the previous global group key silently discarded one team's scenario.
    best_by_group: dict[tuple[str, str], dict[str, Any]] = {}
    for pattern in patterns_found:
        group_key = (str(pattern.get('team') or ''), str(pattern['pattern_group']))
        if group_key not in best_by_group or pattern['pattern_rank'] > best_by_group[group_key]['pattern_rank']:
            best_by_group[group_key] = pattern
    patterns_used = list(best_by_group.values())
    for pattern in patterns_used:
        pattern['used_in_scenario'] = True
    for pattern in patterns_found:
        if pattern not in patterns_used:
            rejected = dict(pattern)
            rejected['rejection_reason'] = 'DOUBLE_COUNT_TEAM_GROUP_LOWER_RANK'
            patterns_rejected.append(rejected)
    independence = float(config['patterns'].get('independence_factor', 0.9))
    for pattern in patterns_used:
        pattern['pattern_weight'] = pattern['credibility'] * pattern['sample_quality'] * pattern['specificity'] * independence * pattern['distance_match_quality']
    weight_sum = sum((pattern['pattern_weight'] for pattern in patterns_used))
    if weight_sum > 0:
        raw = sum((pattern['pattern_weight'] * pattern['shrunk_probability'] for pattern in patterns_used)) / weight_sum
        effective_sample = sum((pattern['pattern_weight'] * pattern['matched_games'] for pattern in patterns_used))
        scenario_k = float(config['credibility']['scenario_k'])
        credibility = effective_sample / (effective_sample + scenario_k)
        probability = credibility * raw + (1 - credibility) * 0.50
        support = 'ON'
    else:
        raw = 0.50
        effective_sample = 0.0
        credibility = 0.0
        probability = 0.50
        support = 'OFF'
    outcome_items = [
        (float(pattern['outcome_median']), float(pattern.get('pattern_weight') or 0.0))
        for pattern in patterns_used
        if pattern.get('outcome_median') is not None and float(pattern.get('pattern_weight') or 0.0) > 0
    ]
    outcome_weight = sum(weight for _, weight in outcome_items)
    outcome_center = (
        sum(value * weight for value, weight in outcome_items) / outcome_weight
        if outcome_weight > 0
        else None
    )
    return {
        'patterns_found': patterns_found,
        'patterns_used': patterns_used,
        'patterns_rejected': patterns_rejected,
        'p_scenario_raw': raw,
        'effective_sample': effective_sample,
        'scenario_credibility': credibility,
        'p_scenario': probability,
        'scenario_support': support,
        'outcome_center': outcome_center,
        'outcome_center_source': 'matched_pattern_outcome_medians' if outcome_center is not None else None,
    }

def choose_best_per_group(patterns: list[dict[str, Any]]) -> list[dict[str, Any]]:
    best: dict[tuple[str, str], dict[str, Any]] = {}
    for pattern in patterns:
        group_key = (str(pattern.get('team') or ''), str(pattern['pattern_group']))
        if group_key not in best or pattern['pattern_rank'] > best[group_key]['pattern_rank']:
            best[group_key] = pattern
    return list(best.values())

# ===== live_projection_engine.py =====
def safe_div(numerator: Optional[float], denominator: Optional[float]) -> Optional[float]:
    if numerator is None or denominator in (None, 0):
        return None
    return numerator / denominator

def calculate_team_metrics(stats: dict[str, Any], points: float, elapsed_minutes: float) -> dict[str, Optional[float]]:
    fga = stats.get('FGA')
    fgm = stats.get('FGM')
    two_pa = stats.get('2PA')
    two_pm = stats.get('2PM')
    three_pa = stats.get('3PA')
    three_pm = stats.get('3PM')
    fta = stats.get('FTA')
    ftm = stats.get('FTM')
    orb = stats.get('ORB')
    turnovers = stats.get('TO')
    poss = fga - orb + turnovers + 0.44 * fta if None not in (fga, orb, turnovers, fta) else stats.get('Poss')
    return {'Poss': poss, 'eFG': safe_div(fgm + 0.5 * three_pm if fgm is not None and three_pm is not None else None, fga), 'FTr': safe_div(fta, fga), 'OffRtg': points / poss * 100 if poss else None, '2P%': safe_div(two_pm, two_pa), '3P%': safe_div(three_pm, three_pa), 'FT%': safe_div(ftm, fta), 'TO_rate': safe_div(turnovers, poss), 'ORB_per_possession': safe_div(orb, poss), 'FTA_per_possession': safe_div(fta, poss), 'FGA_per_minute': safe_div(fga, elapsed_minutes), 'Poss_per_minute': safe_div(poss, elapsed_minutes)}

def _segment_clock(market: dict[str, Any], canonical: dict[str, Any]) -> dict[str, float]:
    market_type = market['market_type']
    segment = market.get('segment', 'MATCH')
    elapsed_game = float(canonical['elapsed_game_seconds'])
    full_game = float(canonical['full_game_seconds'])
    quarter = float(canonical['quarter_seconds'])
    side = 'home' if market.get('team') == canonical['home_team'] else 'away' if market.get('team') else None
    if market_type in {'MATCH_TOTAL', 'TEAM_IT_MATCH'}:
        full = full_game
        elapsed = elapsed_game
        current = float(canonical['score'][side]) if side else float(canonical['score']['total'])
    elif segment == 'H1':
        full = full_game / 2
        elapsed = min(elapsed_game, full)
        quarters = canonical['quarters'][:2]
        current = sum((float(q.get(side) or 0) for q in quarters)) if side else sum((float(q.get('total') or 0) for q in quarters))
    elif segment == 'H2':
        full = full_game / 2
        elapsed = max(0.0, min(full, elapsed_game - full_game / 2))
        quarters = canonical['quarters'][2:]
        current = sum((float(q.get(side) or 0) for q in quarters)) if side else sum((float(q.get('total') or 0) for q in quarters))
    elif segment.startswith('Q') and segment[1:].isdigit():
        target = int(segment[1:])
        full = quarter
        if target == canonical.get('current_quarter'):
            elapsed = max(0.0, quarter - float(canonical['quarter_seconds_remaining']))
        elif target < (canonical.get('current_quarter') or 0):
            elapsed = quarter
        else:
            elapsed = 0.0
        q = canonical['quarters'][target - 1]
        current = float(q.get(side) or 0) if side else float(q.get('total') or 0)
    else:
        full, elapsed, current = (full_game, elapsed_game, float(canonical['score']['total']))
    return {'full_seconds': full, 'elapsed_seconds': elapsed, 'remaining_seconds': max(0.0, full - elapsed), 'current_points': current}

def _history_values(market: dict[str, Any], canonical: dict[str, Any]) -> list[float]:
    values: list[float] = []
    team = market.get('team')
    pools = [canonical['history']['team_a'], canonical['history']['team_b']] if not team else [canonical['history']['team_a'] if team == canonical['home_team'] else canonical['history']['team_b']]
    for pool in pools:
        for game in pool:
            value = segment_value(game, market, team)
            if value is not None:
                values.append(float(value))
    return values

def _game_possessions(game: dict[str, Any], side: str) -> Optional[float]:
    stats = game.get('stats', {}).get(side, {})
    fields = [stats.get('FGA'), stats.get('ORB'), stats.get('TO'), stats.get('FTA')]
    if any((value is None for value in fields)):
        return None
    return fields[0] - fields[1] + fields[2] + 0.44 * fields[3]

def _historical_pace_and_ppp(canonical: dict[str, Any], team: str) -> tuple[Optional[float], Optional[float], Optional[float]]:
    pool = canonical['history']['team_a'] if team == canonical['home_team'] else canonical['history']['team_b']
    opponent_pool = canonical['history']['team_b'] if team == canonical['home_team'] else canonical['history']['team_a']
    cross_pool = canonical.get('history_cross_format', {}).get('team_a' if team == canonical['home_team'] else 'team_b', [])
    cross_opponent_pool = canonical.get('history_cross_format', {}).get('team_b' if team == canonical['home_team'] else 'team_a', [])

    def collect(games: list[dict[str, Any]], allowed_mode: bool=False) -> tuple[list[float], list[float]]:
        paces: list[float] = []
        rates: list[float] = []
        for game in games:
            side = game.get('perspective_side')
            if not side:
                continue
            other = 'away' if side == 'home' else 'home'
            team_poss = _game_possessions(game, side)
            opp_poss = _game_possessions(game, other)
            game_minutes = to_number((game.get('format') or {}).get('regulation_minutes')) or canonical['full_game_seconds'] / 60
            if team_poss and opp_poss and game_minutes:
                paces.append((team_poss + opp_poss) / 2 / game_minutes)
            target_poss = _game_possessions(game, other if allowed_mode else side)
            target_score = game.get('opponent_score') if allowed_mode else game.get('team_score')
            if target_poss and target_score is not None:
                rates.append(float(target_score) / target_poss)
        return paces, rates

    same_paces, same_offense = collect(pool)
    same_allowed_paces, same_allowed = collect(opponent_pool, allowed_mode=True)
    cross_paces, cross_offense = collect(cross_pool)
    cross_allowed_paces, cross_allowed = collect(cross_opponent_pool, allowed_mode=True)

    def conservative_blend(same: list[float], cross: list[float]) -> Optional[float]:
        same_value = statistics.median(same) if same else None
        cross_value = statistics.median(cross) if cross else None
        if same_value is not None and cross_value is not None:
            return 0.75 * same_value + 0.25 * cross_value
        if same_value is not None:
            return same_value
        if cross_value is not None:
            return cross_value
        return None

    return (
        conservative_blend(same_paces, cross_paces),
        conservative_blend(same_offense, cross_offense),
        conservative_blend(same_allowed, cross_allowed),
    )

def _previous_quarter_pace(market: dict[str, Any], canonical: dict[str, Any], clock: dict[str, float]) -> Optional[float]:
    current = canonical.get('current_quarter') or 1
    side = 'home' if market.get('team') == canonical['home_team'] else 'away' if market.get('team') else None
    index = current - 2
    if index < 0 or index >= len(canonical['quarters']):
        return None
    q = canonical['quarters'][index]
    points = q.get(side) if side else q.get('total')
    if points is None:
        return None
    return float(points) / canonical['quarter_seconds']


def _detect_run_context(canonical: dict[str, Any], team_metrics: dict[str, dict[str, Optional[float]]]) -> dict[str, Any]:
    """Classify a live run without play-by-play.

    BOTH_SIDES_RUNOUT means both offenses are participating in a high-rate segment.
    SOLO_RUN means one team owns at least 72% of a sufficiently large segment. A
    solo run is not automatically Match Over: the opponent must still show attack
    volume/bounce potential; otherwise only the running team's IT may benefit.
    """
    quarter_seconds = float(canonical.get('quarter_seconds') or 600)
    current_q = canonical.get('current_quarter')
    stage = str(canonical.get('stage') or '')
    if stage == 'HT':
        index = 1
        elapsed = quarter_seconds
    elif stage == 'AFTER_3Q':
        index = 2
        elapsed = quarter_seconds
    elif current_q and 1 <= int(current_q) <= 4:
        index = int(current_q) - 1
        elapsed = max(0.0, quarter_seconds - float(canonical.get('quarter_seconds_remaining') or quarter_seconds))
    else:
        return {'label': 'OFF', 'reason': 'NO_SEGMENT'}
    quarters = canonical.get('quarters') or []
    if index >= len(quarters):
        return {'label': 'OFF', 'reason': 'NO_QUARTER'}
    q = quarters[index] or {}
    home = to_number(q.get('home'))
    away = to_number(q.get('away'))
    if home is None or away is None or elapsed < 120:
        return {'label': 'OFF', 'reason': 'SEGMENT_TOO_EARLY_OR_MISSING'}
    total = home + away
    minutes = elapsed / 60.0
    ppm = total / minutes if minutes > 0 else 0.0
    share = max(home, away) / total if total > 0 else 0.0
    leader = 'home' if home >= away else 'away'
    opponent = 'away' if leader == 'home' else 'home'
    opponent_metrics = team_metrics.get(opponent) or {}
    opponent_bounce = bool(
        (opponent_metrics.get('eFG') is not None and float(opponent_metrics['eFG']) < 0.46)
        and (opponent_metrics.get('FGA_per_minute') is not None and float(opponent_metrics['FGA_per_minute']) >= 1.20)
        and ((opponent_metrics.get('FTr') or 0.0) >= 0.20 or (opponent_metrics.get('ORB_per_possession') or 0.0) >= 0.07)
    )
    if ppm >= 4.4 and min(home, away) >= max(8.0, 0.22 * total):
        label = 'BOTH_SIDES_RUNOUT'
    elif total >= 14 and share >= 0.72:
        label = 'SOLO_RUN_WITH_OPPONENT_BOUNCE' if opponent_bounce else 'SOLO_RUN_ONE_SIDED'
    else:
        label = 'NO_STRONG_RUN'
    return {
        'label': label,
        'quarter': index + 1,
        'home_points': home,
        'away_points': away,
        'segment_points_per_minute': ppm,
        'leader': leader,
        'leader_share': share,
        'opponent_bounce': opponent_bounce,
    }

def _trimmed_weighted_mean(items: list[tuple[str, float, float]]) -> tuple[float, set[str]]:
    included = list(items)
    excluded: set[str] = set()
    if len(included) >= 5:
        ordered = sorted(included, key=lambda item: item[1])
        excluded.update({ordered[0][0], ordered[-1][0]})
        included = ordered[1:-1]
    total_weight = sum((weight for _, _, weight in included))
    if total_weight <= 0:
        return (statistics.median((value for _, value, _ in included)), excluded)
    return (sum((value * weight for _, value, weight in included)) / total_weight, excluded)

def _stage_sigma(market_type: str, stage: str, config: dict[str, Any]) -> float:
    sigma_key = market_type
    if market_type in {'TEAM_IT_H1', 'TEAM_IT_H2'}:
        sigma_key = 'TEAM_IT_HALF'
    elif market_type == 'CURRENT_QUARTER_TOTAL':
        sigma_key = 'QUARTER_TOTAL'
    elif market_type == 'CURRENT_QUARTER_TEAM_IT':
        sigma_key = 'QUARTER_TEAM_IT'
    settings = config['sigma'].get(sigma_key, {'default': 10.0})
    return float(settings.get(stage, settings.get('default', 10.0)))

def _parser_projection_components(market: dict[str, Any], canonical: dict[str, Any], clock: dict[str, float]) -> dict[str, dict[str, Any]]:
    blocks = canonical.get('parser_blocks', {})
    projections = blocks.get('projections') if isinstance(blocks.get('projections'), dict) else {}
    conditioned = blocks.get('stat_conditioned_line_profiles') if isinstance(blocks.get('stat_conditioned_line_profiles'), dict) else {}
    live_meta = conditioned.get('live_calibrated') if isinstance(conditioned.get('live_calibrated'), dict) else {}
    elapsed_minutes = canonical['elapsed_game_seconds'] / 60
    parser_elapsed = to_number(live_meta.get('min_played'))
    snapshot_ok = parser_elapsed is None or abs(parser_elapsed - elapsed_minutes) <= 1.5
    team_side = 'home' if market.get('team') == canonical['home_team'] else 'away' if market.get('team') else None
    result: dict[str, dict[str, Any]] = {}
    coursework = canonical.get('coursework_forecast') if isinstance(canonical.get('coursework_forecast'), dict) else {}

    def add(name: str, value: Any, reason: Optional[str]=None, suggested_weight: Optional[float]=None) -> None:
        numeric = to_number(value)
        if numeric is not None and numeric < clock['current_points']:
            reason = 'PARSER_PROJECTION_BELOW_CURRENT_SCORE'
        result[name] = {
            'value': numeric,
            'available': numeric is not None and reason is None,
            'exclusion_reason': reason if reason else None,
            'suggested_weight': suggested_weight,
        }

    market_type = market['market_type']
    if market_type in {'MATCH_TOTAL', 'TEAM_IT_MATCH'}:
        live = projections.get('live_calibrated') if isinstance(projections.get('live_calibrated'), dict) else {}
        segment = projections.get('segment_projection') if isinstance(projections.get('segment_projection'), dict) else {}
        pre = projections.get('pre_match_stat') if isinstance(projections.get('pre_match_stat'), dict) else {}
        suffix = f'{team_side}_final' if team_side else 'total'
        live_value = live.get(suffix)
        segment_value_ = segment.get(suffix)
        pre_value = pre.get(suffix)
        live_reason = None if snapshot_ok and live.get('valid', True) else 'PARSER_PROJECTION_STALE_OR_INVALID'
        add('projection_parser_live_calibrated', live_value, live_reason)
        divergence_reason = None
        if to_number(live_value) is not None and to_number(segment_value_) is not None:
            divergence = abs(float(segment_value_) - float(live_value))
            if divergence > max(25.0, abs(float(live_value)) * 0.18):
                divergence_reason = 'PARSER_SEGMENT_DIVERGES_FROM_LIVE_CALIBRATED'
        if not snapshot_ok:
            divergence_reason = 'PARSER_PROJECTION_STALE_OR_INVALID'
        add('projection_parser_segment', segment_value_, divergence_reason)
        add('projection_parser_pre_match', pre_value)
    elif market_type in {'CURRENT_QUARTER_TOTAL', 'CURRENT_QUARTER_TEAM_IT'}:
        segment_key = str(market.get('segment') or '').lower()
        quarter_projection = projections.get(segment_key) if isinstance(projections.get(segment_key), dict) else {}
        if team_side:
            key = 'team_a_center' if team_side == 'home' else 'team_b_center'
        else:
            key = 'total_center'
        add('projection_parser_current_quarter', quarter_projection.get(key))

    # The newer coursework model supplies an independent remaining-quarter median.
    # It is available only for combined totals, never for team IT, so segment semantics
    # cannot be mixed (e.g. TEAM_IT_MATCH vs TEAM_IT_H2).
    if market_type in {'MATCH_TOTAL', 'H1_TOTAL', 'H2_TOTAL', 'CURRENT_QUARTER_TOTAL'}:
        segment_key = str(market.get('segment') or 'MATCH')
        value = (coursework.get('segment_projections') or {}).get(segment_key)
        readiness = coursework.get('data_readiness')
        reason = None if coursework.get('eligible_as_projection_component') else 'COURSEWORK_SAMPLE_INSUFFICIENT'
        weight = 0.12 if readiness == 'READY' else 0.06 if readiness == 'REVIEW_REQUIRED' else 0.0
        add('projection_coursework_remaining', value, reason, suggested_weight=weight)
    return result

def calculate_live_projection(market: dict[str, Any], canonical: dict[str, Any], history: dict[str, Any], scenario: dict[str, Any], config: dict[str, Any], stat: Optional[dict[str, Any]]=None) -> dict[str, Any]:
    clock = _segment_clock(market, canonical)
    elapsed_minutes = canonical['elapsed_game_seconds'] / 60
    team_metrics = {side: calculate_team_metrics(canonical['live_stats'][side], canonical['score'][side], elapsed_minutes) for side in ('home', 'away')}
    poss_values = [team_metrics[side]['Poss'] for side in ('home', 'away')]
    game_possessions = sum(poss_values) / 2 if all((value is not None for value in poss_values)) else None
    combined_ppp = None
    if all((value not in (None, 0) for value in poss_values)):
        combined_ppp = canonical['score']['home'] / poss_values[0] + canonical['score']['away'] / poss_values[1]
    simple = clock['current_points'] / clock['elapsed_seconds'] * clock['full_seconds'] if clock['elapsed_seconds'] > 0 else None
    values = _history_values(market, canonical)
    baseline = statistics.median(values) if values else None
    historical_rate = baseline / clock['full_seconds'] if baseline is not None and clock['full_seconds'] else None
    history_projection = clock['current_points'] + historical_rate * clock['remaining_seconds'] if historical_rate is not None else None
    scenario_projection = None
    scenario_projection_method = None
    scenario_center = to_number(scenario.get('outcome_center'))
    if history_projection is not None and scenario_center is not None and clock['full_seconds']:
        scenario_rate = scenario_center / clock['full_seconds']
        scenario_rate_projection = clock['current_points'] + scenario_rate * clock['remaining_seconds']
        scenario_credibility = max(0.0, min(0.65, float(scenario.get('scenario_credibility') or 0.0)))
        scenario_projection = scenario_credibility * scenario_rate_projection + (1 - scenario_credibility) * history_projection
        scenario_projection_method = 'MATCHED_PATTERN_OUTCOME_DISTRIBUTION'
    elif history_projection is not None:
        delta = (scenario['p_scenario'] - history['p_hist']) * float(config['projection']['scenario_projection_span'])
        direction = 1 if market['side'] == 'OVER' else -1
        scenario_projection = history_projection + direction * delta
        scenario_projection_method = 'PROBABILITY_DELTA_FALLBACK'
    current_rate = clock['current_points'] / clock['elapsed_seconds'] if clock['elapsed_seconds'] > 0 else None
    previous_rate = _previous_quarter_pace(market, canonical, clock)
    segment_rates = [(current_rate, 0.45), (previous_rate, 0.25), (historical_rate, 0.3)]
    available_rates = [(value, weight) for value, weight in segment_rates if value is not None]
    rate_weight = sum((weight for _, weight in available_rates))
    blended_rate = sum((value * weight for value, weight in available_rates)) / rate_weight if rate_weight else None
    segment_projection = clock['current_points'] + blended_rate * clock['remaining_seconds'] if blended_rate is not None else None
    stat_adjusted = None
    stat_details: dict[str, Any] = {'team_metrics': team_metrics, 'game_possessions': game_possessions, 'combined_ppp': combined_ppp}
    if game_possessions is not None and elapsed_minutes > 0:
        remaining_minutes = clock['remaining_seconds'] / 60
        current_pace = game_possessions / elapsed_minutes
        home_pace, home_offense, home_allowed = _historical_pace_and_ppp(canonical, canonical['home_team'])
        away_pace, away_offense, away_allowed = _historical_pace_and_ppp(canonical, canonical['away_team'])
        historical_paces = [value for value in (home_pace, away_pace) if value is not None]
        historical_pace = statistics.median(historical_paces) if historical_paces else current_pace
        scenario_pace = historical_pace
        pace_weights = config['projection']['regression']
        margin = abs(float(canonical['score']['home']) - float(canonical['score']['away']))
        # Close games preserve a real high pace longer; blowouts regress harder.
        if margin <= 10:
            current_pace_w, history_pace_w, scenario_pace_w = 0.55, 0.30, 0.15
            pace_context = 'CLOSE_PACE_PRESERVED'
        elif margin >= 18:
            current_pace_w, history_pace_w, scenario_pace_w = 0.30, 0.55, 0.15
            pace_context = 'BLOWOUT_REGRESSION'
        else:
            current_pace_w = float(pace_weights['current_pace'])
            history_pace_w = float(pace_weights['history_pace'])
            scenario_pace_w = float(pace_weights['scenario_pace'])
            pace_context = 'STANDARD_REGRESSION'
        blended_future_pace = current_pace * current_pace_w + historical_pace * history_pace_w + scenario_pace * scenario_pace_w
        future_possessions = blended_future_pace * remaining_minutes
        run_context = _detect_run_context(canonical, team_metrics)
        run_label = run_context.get('label')
        if run_label == 'BOTH_SIDES_RUNOUT' and margin <= 10:
            future_possessions *= 1.03
        elif run_label == 'SOLO_RUN_WITH_OPPONENT_BOUNCE':
            future_possessions *= 1.015
        elif run_label == 'SOLO_RUN_ONE_SIDED' and not market.get('team'):
            # A one-team burst with a dead opponent is not a match-total Over signal.
            future_possessions *= 0.98
        regressed: dict[str, float] = {}
        efficiency_context: dict[str, Any] = {}
        for side, offense, opponent_allowed in (('home', home_offense, away_allowed), ('away', away_offense, home_allowed)):
            current_ppp = safe_div(canonical['score'][side], team_metrics[side]['Poss'])
            efg = team_metrics[side].get('eFG')
            fga_pm = team_metrics[side].get('FGA_per_minute')
            ftr = team_metrics[side].get('FTr')
            orb_rate = team_metrics[side].get('ORB_per_possession')
            to_rate = team_metrics[side].get('TO_rate')
            high_volume = fga_pm is not None and fga_pm >= 1.20
            bounce = bool(efg is not None and efg < 0.46 and high_volume and ((ftr or 0) >= 0.22 or (orb_rate or 0) >= 0.075))
            overheat = bool(efg is not None and efg > 0.64 and not high_volume and (ftr or 0) < 0.20)
            turnover_drag = bool(to_rate is not None and to_rate >= 0.22)
            if bounce:
                local_weights = {'current': 0.15, 'offense': 0.35, 'allowed': 0.35, 'scenario': 0.15}
                label = 'LOW_EFG_HIGH_VOLUME_BOUNCE'
            elif overheat:
                local_weights = {'current': 0.15, 'offense': 0.38, 'allowed': 0.37, 'scenario': 0.10}
                label = 'HIGH_EFG_LOW_SUPPORT_REGRESSION'
            else:
                local_weights = {
                    'current': float(pace_weights['current_ppp']),
                    'offense': float(pace_weights['historical_offense_ppp']),
                    'allowed': float(pace_weights['opponent_allowed_ppp']),
                    'scenario': float(pace_weights['scenario_ppp']),
                }
                label = 'STANDARD_PPP_REGRESSION'
            base_values = {'current': current_ppp, 'offense': offense, 'allowed': opponent_allowed, 'scenario': offense}
            available = [(base_values[key], local_weights[key]) for key in base_values if base_values[key] is not None]
            total = sum((weight for _, weight in available))
            value = sum((v * w for v, w in available)) / total if total else 1.0
            if turnover_drag:
                value *= 0.97
                label += '+TO_DRAG'
            regressed[side] = value
            efficiency_context[side] = {
                'label': label, 'eFG': efg, 'FGA_per_minute': fga_pm,
                'FTr': ftr, 'ORB_per_possession': orb_rate, 'TO_rate': to_rate,
                'weights': local_weights,
            }
        if market.get('team'):
            side = 'home' if market['team'] == canonical['home_team'] else 'away'
            stat_adjusted = clock['current_points'] + future_possessions * regressed[side]
        else:
            stat_adjusted = clock['current_points'] + future_possessions * (regressed['home'] + regressed['away'])
        adjustment_rate = 0.0
        adjustment_events: list[dict[str, Any]] = []
        indicators = (stat or {}).get('indicators', {})
        adjustments = config['projection'].get('adjustments', {})
        adjustment_rules = [
            ('EFG_VERY_HIGH_NO_VOLUME', bool(indicators.get('score_or_efg_high') and indicators.get('volume_low')), 'efg_very_high_no_volume'),
            ('LOW_EFG_HIGH_VOLUME_BOUNCE', bool(indicators.get('score_or_efg_low') and indicators.get('volume_high')), 'low_efg_high_volume_bounce'),
            ('FTR_HIGH', bool(indicators.get('fta_high')), 'ftr_high'),
            ('ORB_HIGH', bool(indicators.get('orb_high')), 'orb_high'),
            ('TO_HIGH', bool(indicators.get('to_high')), 'to_high'),
            ('OPPONENT_ALLOWS', bool(indicators.get('opponent_allows')), 'opponent_allows'),
            ('OPPONENT_SUPPRESSES', bool(indicators.get('opponent_suppresses')), 'opponent_suppresses'),
        ]
        for rule_id, active, config_key in adjustment_rules:
            if not active:
                continue
            delta = float(adjustments.get(config_key, 0.0))
            adjustment_rate += delta
            adjustment_events.append({'rule_id': rule_id, 'delta': delta})
        adjustment_rate = max(-0.08, min(0.08, adjustment_rate))
        if stat_adjusted is not None:
            future_points = max(0.0, stat_adjusted - clock['current_points'])
            stat_adjusted = clock['current_points'] + future_points * (1 + adjustment_rate)
        stat_details.update({
            'current_pace': current_pace, 'historical_pace': historical_pace,
            'scenario_pace': scenario_pace, 'blended_future_pace': blended_future_pace,
            'future_possessions': future_possessions, 'regressed_ppp': regressed,
            'pace_context': pace_context, 'run_context': run_context,
            'efficiency_context': efficiency_context,
            'adjustment_rate': adjustment_rate, 'adjustment_events': adjustment_events,
        })
    parser_components = _parser_projection_components(market, canonical, clock)
    parser_available_values = [
        item['value'] for item in parser_components.values()
        if item.get('available') and item.get('value') is not None
    ]
    control_values = [value for value in (history_projection, scenario_projection, stat_adjusted, segment_projection) if value is not None]
    control_values.extend(parser_available_values[:1])
    control = statistics.median(control_values) if control_values else simple
    configured_weights = config['projection']['weights']
    component_values = {'projection_simple': simple, 'projection_segment': segment_projection, 'projection_history': history_projection, 'projection_scenario': scenario_projection, 'projection_stat_adjusted': stat_adjusted, 'projection_control': control}
    component_weights = {'projection_simple': float(config['projection']['simple_information_weight']), 'projection_segment': float(configured_weights['segment']), 'projection_history': float(configured_weights['history']), 'projection_scenario': float(configured_weights['scenario']), 'projection_stat_adjusted': float(configured_weights['stat_adjusted']), 'projection_control': float(configured_weights['control'])}
    parser_weights = {
        'projection_parser_live_calibrated': 0.20,
        'projection_parser_segment': 0.08,
        'projection_parser_pre_match': 0.08,
        'projection_parser_current_quarter': 0.18,
    }
    for key, item in parser_components.items():
        component_values[key] = item.get('value') if item.get('available') else None
        component_weights[key] = float(item.get('suggested_weight') if item.get('suggested_weight') is not None else parser_weights.get(key, 0.08))
    items = [(key, float(value), component_weights[key]) for key, value in component_values.items() if value is not None]
    line = float(market['line'])
    if items:
        projection_used, trimmed = _trimmed_weighted_mean(items)
    else:
        projection_used, trimmed = (line, set())
    components: dict[str, dict[str, Any]] = {}
    for key, value in component_values.items():
        parser_exclusion = parser_components.get(key, {}).get('exclusion_reason')
        components[key] = {'value': parser_components.get(key, {}).get('value', value), 'weight': component_weights[key], 'available': value is not None, 'included': value is not None and key not in trimmed, 'exclusion_reason': 'TRIMMED_EXTREME' if key in trimmed else parser_exclusion or ('UNAVAILABLE' if value is None else None)}
    line_edge = projection_used - line if market['side'] == 'OVER' else line - projection_used
    sigma = _stage_sigma(market['market_type'], canonical['stage'], config)
    if (stat or {}).get('stat_support') == 'OFF':
        sigma *= 1.20
    z_score = line_edge / sigma
    p_live = normal_cdf(z_score)
    return {'clock': canonical.get('clock'), 'elapsed_seconds': clock['elapsed_seconds'], 'remaining_seconds': clock['remaining_seconds'], 'elapsed_game_seconds': canonical['elapsed_game_seconds'], 'remaining_game_seconds': canonical['remaining_game_seconds'], 'current_points': clock['current_points'], 'components': components, 'projection_simple': simple, 'projection_segment': segment_projection, 'projection_model_live': segment_projection, 'projection_history': history_projection, 'projection_scenario': scenario_projection, 'scenario_projection_method': scenario_projection_method, 'projection_stat_adjusted': stat_adjusted, 'projection_control': control, 'projection_used': projection_used, 'Projection_used': projection_used, 'line': line, 'line_edge': line_edge, 'line_edge_over': projection_used - line, 'line_edge_under': line - projection_used, 'sigma': sigma, 'z_score': z_score, 'p_live': p_live, 'stat_projection_details': stat_details}

# ===== stat_gate_engine.py =====
METRICS = ('scored', 'allowed', 'period_total', 'FGA', 'Poss', '2PA', '3PA', 'FTA', 'FTr', 'ORB', 'TO', 'fouls', 'eFG', 'OffRtg', 'allowed_FGA', 'allowed_eFG', 'allowed_FTA', 'allowed_ORB', 'allowed_Poss', 'allowed_OffRtg', 'forced_TO')

def _zone(value: Optional[float], thresholds: Optional[dict[str, Any]]) -> Optional[str]:
    if value is None or not thresholds:
        return None
    if value <= thresholds['p25']:
        return 'LOW'
    if value >= thresholds['p90']:
        return 'VERY_HIGH'
    if value >= thresholds['p75']:
        return 'HIGH'
    return 'MID'

def _side_for_team(game: dict[str, Any], team: str) -> Optional[str]:
    if game.get('home_team') == team:
        return 'home'
    if game.get('away_team') == team:
        return 'away'
    return None

def _aggregate_stats(game: dict[str, Any], side: str, scope: str) -> dict[str, Optional[float]]:
    if scope == 'MATCH':
        return dict(game.get('stats', {}).get(side, {}))
    indices = [0, 1] if scope == 'H1' else [2, 3] if scope == 'H2' else [int(scope[1:]) - 1] if scope.startswith('Q') and scope[1:].isdigit() else []
    rows = game.get('quarter_stats', {}).get(side, [])
    output: dict[str, Optional[float]] = {}
    for metric in ('FGA', 'FGM', '2PA', '2PM', '3PA', '3PM', 'FTA', 'FTM', 'ORB', 'DRB', 'TO', 'FOULS'):
        values = [rows[index].get(metric) for index in indices if index < len(rows)]
        output[metric] = sum(values) if values and all((value is not None for value in values)) else None
    return output

def _score_for_scope(game: dict[str, Any], side: str, scope: str) -> Optional[float]:
    if scope == 'MATCH':
        return game.get('home_score') if side == 'home' else game.get('away_score')
    indices = [0, 1] if scope == 'H1' else [2, 3] if scope == 'H2' else [int(scope[1:]) - 1] if scope.startswith('Q') and scope[1:].isdigit() else []
    values = [game['quarters'][index].get(side) for index in indices]
    return sum(values) if values and all((value is not None for value in values)) else None

def _metric_value(game: dict[str, Any], team: str, scope: str, metric: str) -> Optional[float]:
    side = _side_for_team(game, team)
    if not side:
        return None
    opponent = 'away' if side == 'home' else 'home'
    scored = _score_for_scope(game, side, scope)
    allowed = _score_for_scope(game, opponent, scope)
    stats = _aggregate_stats(game, side, scope)
    opp_stats = _aggregate_stats(game, opponent, scope)
    full_minutes = 40.0 if scope == 'MATCH' else 20.0 if scope in {'H1', 'H2'} else 10.0
    team_metrics = calculate_team_metrics(stats, scored or 0.0, full_minutes)
    opp_metrics = calculate_team_metrics(opp_stats, allowed or 0.0, full_minutes)
    mapping = {'scored': scored, 'allowed': allowed, 'period_total': scored + allowed if scored is not None and allowed is not None else None, 'FGA': stats.get('FGA'), 'Poss': team_metrics.get('Poss'), '2PA': stats.get('2PA'), '3PA': stats.get('3PA'), 'FTA': stats.get('FTA'), 'FTr': team_metrics.get('FTr'), 'ORB': stats.get('ORB'), 'TO': stats.get('TO'), 'fouls': stats.get('FOULS'), 'eFG': team_metrics.get('eFG'), 'OffRtg': team_metrics.get('OffRtg'), 'allowed_FGA': opp_stats.get('FGA'), 'allowed_eFG': opp_metrics.get('eFG'), 'allowed_FTA': opp_stats.get('FTA'), 'allowed_ORB': opp_stats.get('ORB'), 'allowed_Poss': opp_metrics.get('Poss'), 'allowed_OffRtg': opp_metrics.get('OffRtg'), 'forced_TO': opp_stats.get('TO')}
    return mapping.get(metric)

class ZoneIndex:

    def __init__(self, zones_data: Optional[dict[str, Any]]) -> None:
        self._index: dict[tuple[str, str, str], dict[str, Any]] = {}
        for row in (zones_data or {}).get('team_relative_zone_thresholds', []):
            if isinstance(row, dict) and all((key in row for key in ('team', 'scope', 'metric'))):
                self._index[str(row['team']), str(row['scope']).upper(), str(row['metric'])] = row

    def get(self, team: str, scope: str, metric: str) -> Optional[dict[str, Any]]:
        return self._index.get((team, scope.upper(), metric))

def _fallback_thresholds(canonical: dict[str, Any], team: str, scope: str, metric: str) -> Optional[dict[str, Any]]:
    pool = canonical['history']['team_a'] if team == canonical['home_team'] else canonical['history']['team_b']
    values = [_metric_value(game, team, scope, metric) for game in pool]
    valid = [float(value) for value in values if value is not None]
    if not valid:
        return None
    return {'team': team, 'scope': scope, 'metric': metric, 'n': len(valid), 'mean': sum(valid) / len(valid), 'p25': percentile(valid, 0.25), 'p50': percentile(valid, 0.5), 'p75': percentile(valid, 0.75), 'p90': percentile(valid, 0.9), 'source': 'match_file_last35_fallback'}

def _current_raw_stats(canonical: dict[str, Any], side: str, scope: str) -> dict[str, Optional[float]]:
    if scope in {'MATCH', 'H1'}:
        return dict(canonical['live_stats'][side])
    raw = canonical.get('raw_main', {})
    prefix = 'h' if side == 'home' else 'a'
    codes = {'FGA': 'fga', 'FGM': 'fgm', '2PA': '2pa', '2PM': '2pm', '3PA': '3pa', '3PM': '3pm', 'FTA': 'fta', 'FTM': 'ftm', 'ORB': 'orb', 'DRB': 'drb', 'TO': 'tov', 'FOULS': 'fls'}
    indices = [2, 3] if scope == 'H2' else [int(scope[1:])] if scope.startswith('Q') and scope[1:].isdigit() else []
    result: dict[str, Optional[float]] = {}
    for metric, code in codes.items():
        values = [to_number(raw.get(f'{prefix}{code}{index}')) for index in indices]
        result[metric] = sum(values) if values and all((value is not None for value in values)) else None
    return result

def _current_score(canonical: dict[str, Any], side: str, scope: str) -> float:
    if scope == 'MATCH':
        return float(canonical['score'][side])
    indices = [0, 1] if scope == 'H1' else [2, 3] if scope == 'H2' else [int(scope[1:]) - 1] if scope.startswith('Q') and scope[1:].isdigit() else []
    return sum((float(canonical['quarters'][index].get(side) or 0) for index in indices))

def _scope_timing(canonical: dict[str, Any], scope: str) -> dict[str, Any]:
    """Return elapsed/target seconds inside the evaluated market scope.

    Stat-zone thresholds describe a completed MATCH/H1/H2/Qn segment.  Live boxscore
    counters are cumulative only up to the snapshot.  Comparing a HT FGA count directly
    with a full-match FGA distribution creates a false LOW-volume profile.  This helper
    makes the time base explicit before cumulative metrics are classified.
    """
    elapsed_game = max(0, int(canonical.get('elapsed_game_seconds') or 0))
    full_game = max(1, int(canonical.get('full_game_seconds') or 1))
    quarter = max(1, int(canonical.get('quarter_seconds') or full_game // 4 or 1))
    half = full_game // 2
    normalized_scope = str(scope or 'MATCH').upper()
    if normalized_scope == 'MATCH':
        start, target = 0, full_game
    elif normalized_scope == 'H1':
        start, target = 0, half
    elif normalized_scope == 'H2':
        start, target = half, half
    elif normalized_scope.startswith('Q') and normalized_scope[1:].isdigit():
        number = max(1, min(4, int(normalized_scope[1:])))
        start, target = (number - 1) * quarter, quarter
    else:
        start, target = 0, full_game
    elapsed_scope = max(0, min(target, elapsed_game - start))
    ratio = elapsed_scope / target if target > 0 else 0.0
    # Checkpoint calculations begin no earlier than one completed quarter for full-match
    # markets.  Below 25% of a scope the projection is retained for diagnostics only and
    # the stat gate may not become a hard AGAINST blocker.
    reliable = ratio >= 0.25
    factor = 1.0
    if 0 < elapsed_scope < target:
        factor = min(4.0, target / elapsed_scope)
    return {
        'scope': normalized_scope,
        'scope_start_seconds': start,
        'elapsed_scope_seconds': elapsed_scope,
        'target_scope_seconds': target,
        'elapsed_ratio': ratio,
        'projection_factor': factor,
        'reliable_for_hard_gate': reliable,
        'scope_complete': elapsed_scope >= target,
    }


def _scale_count(value: Optional[float], factor: float) -> Optional[float]:
    return None if value is None else float(value) * factor


def _project_counting_stats(stats: dict[str, Optional[float]], factor: float) -> dict[str, Optional[float]]:
    counting = {'FGA', 'FGM', '2PA', '2PM', '3PA', '3PM', 'FTA', 'FTM', 'ORB', 'DRB', 'TO', 'FOULS'}
    return {
        key: _scale_count(value, factor) if key in counting else value
        for key, value in stats.items()
    }


def _current_metric_map_legacy(canonical: dict[str, Any], side: str, scope: str) -> dict[str, Optional[float]]:
    """Pre-v5.3 metric map retained only for live-projection adjustments.

    Keeping the projection channel stable prevents a stat-gate time-base correction from
    silently changing already validated PLAY/RISK probabilities.  It is never used for
    hard stat-gate blocking.
    """
    opponent = 'away' if side == 'home' else 'home'
    stats = _current_raw_stats(canonical, side, scope)
    opp_stats = _current_raw_stats(canonical, opponent, scope)
    scored = _current_score(canonical, side, scope)
    allowed = _current_score(canonical, opponent, scope)
    elapsed_minutes = max(1 / 60, canonical['elapsed_game_seconds'] / 60)
    metrics = calculate_team_metrics(stats, scored, elapsed_minutes)
    opp_metrics = calculate_team_metrics(opp_stats, allowed, elapsed_minutes)
    return {'scored': scored, 'allowed': allowed, 'period_total': scored + allowed, 'FGA': stats.get('FGA'), 'Poss': metrics.get('Poss'), '2PA': stats.get('2PA'), '3PA': stats.get('3PA'), 'FTA': stats.get('FTA'), 'FTr': metrics.get('FTr'), 'ORB': stats.get('ORB'), 'TO': stats.get('TO'), 'fouls': stats.get('FOULS'), 'eFG': metrics.get('eFG'), 'OffRtg': metrics.get('OffRtg'), 'allowed_FGA': opp_stats.get('FGA'), 'allowed_eFG': opp_metrics.get('eFG'), 'allowed_FTA': opp_stats.get('FTA'), 'allowed_ORB': opp_stats.get('ORB'), 'allowed_Poss': opp_metrics.get('Poss'), 'allowed_OffRtg': opp_metrics.get('OffRtg'), 'forced_TO': opp_stats.get('TO')}


def _current_metric_map(canonical: dict[str, Any], side: str, scope: str) -> dict[str, Any]:
    opponent = 'away' if side == 'home' else 'home'
    raw_stats = _current_raw_stats(canonical, side, scope)
    raw_opp_stats = _current_raw_stats(canonical, opponent, scope)
    raw_scored = _current_score(canonical, side, scope)
    raw_allowed = _current_score(canonical, opponent, scope)
    timing = _scope_timing(canonical, scope)
    factor = float(timing['projection_factor'])

    # Only cumulative/counting quantities are projected to the completed scope.  Ratios
    # such as eFG/FTr/OffRtg remain mathematically unchanged because numerator and
    # denominator are scaled together.
    stats = _project_counting_stats(raw_stats, factor)
    opp_stats = _project_counting_stats(raw_opp_stats, factor)
    scored = raw_scored * factor
    allowed = raw_allowed * factor
    target_minutes = max(1 / 60, float(timing['target_scope_seconds']) / 60.0)
    metrics = calculate_team_metrics(stats, scored, target_minutes)
    opp_metrics = calculate_team_metrics(opp_stats, allowed, target_minutes)
    return {
        'scored': scored,
        'allowed': allowed,
        'period_total': scored + allowed,
        'FGA': stats.get('FGA'),
        'Poss': metrics.get('Poss'),
        '2PA': stats.get('2PA'),
        '3PA': stats.get('3PA'),
        'FTA': stats.get('FTA'),
        'FTr': metrics.get('FTr'),
        'ORB': stats.get('ORB'),
        'TO': stats.get('TO'),
        'fouls': stats.get('FOULS'),
        'eFG': metrics.get('eFG'),
        'OffRtg': metrics.get('OffRtg'),
        'allowed_FGA': opp_stats.get('FGA'),
        'allowed_eFG': opp_metrics.get('eFG'),
        'allowed_FTA': opp_stats.get('FTA'),
        'allowed_ORB': opp_stats.get('ORB'),
        'allowed_Poss': opp_metrics.get('Poss'),
        'allowed_OffRtg': opp_metrics.get('OffRtg'),
        'forced_TO': opp_stats.get('TO'),
        '_normalization': {
            **timing,
            'raw_score': raw_scored,
            'projected_score': scored,
            'raw_allowed': raw_allowed,
            'projected_allowed': allowed,
            'raw_stats': raw_stats,
            'projected_stats': stats,
        },
    }

def _is_high(zone: Optional[str]) -> bool:
    return zone in {'HIGH', 'VERY_HIGH'}

def _is_low(zone: Optional[str]) -> bool:
    return zone == 'LOW'

def classify_fake_profiles(flags: dict[str, bool]) -> tuple[bool, bool]:
    fake_over = flags.get('score_or_efg_high', False) and flags.get('volume_low', False) and flags.get('fta_low', False) and (not flags.get('orb_high', False))
    fake_under = flags.get('score_or_efg_low', False) and flags.get('volume_high', False) and (flags.get('orb_high', False) or flags.get('fta_high', False))
    return (fake_over, fake_under)

def calculate_stat_gate(market: dict[str, Any], canonical: dict[str, Any], zones_data: Optional[dict[str, Any]], *, project_counts_to_scope_end: bool=True) -> dict[str, Any]:
    scope = market.get('segment') or 'MATCH'
    if scope not in {'MATCH', 'H1', 'H2', 'Q1', 'Q2', 'Q3', 'Q4'}:
        scope = 'MATCH'
    index = ZoneIndex(zones_data)
    comparisons: dict[str, list[dict[str, Any]]] = {'team_a': [], 'team_b': []}
    zone_maps: dict[str, dict[str, Optional[str]]] = {'team_a': {}, 'team_b': {}}
    normalization: dict[str, dict[str, Any]] = {}
    current_maps: dict[str, dict[str, Any]] = {}
    for label, team, side in (('team_a', canonical['home_team'], 'home'), ('team_b', canonical['away_team'], 'away')):
        current = _current_metric_map(canonical, side, scope) if project_counts_to_scope_end else _current_metric_map_legacy(canonical, side, scope)
        current_maps[label] = current
        normalization[label] = deepcopy(current.get('_normalization') or {
            'scope': scope,
            'projection_factor': 1.0,
            'reliable_for_hard_gate': True,
            'legacy_projection_profile': True,
        })
        for metric in METRICS:
            thresholds = index.get(team, scope, metric) or _fallback_thresholds(canonical, team, scope, metric)
            value = current.get(metric)
            zone = _zone(value, thresholds)
            zone_maps[label][metric] = zone
            comparisons[label].append({'metric': metric, 'current_value': value, 'p25': thresholds.get('p25') if thresholds else None, 'p50': thresholds.get('p50') if thresholds else None, 'p75': thresholds.get('p75') if thresholds else None, 'p90': thresholds.get('p90') if thresholds else None, 'n': thresholds.get('n') if thresholds else 0, 'zone': zone, 'source': thresholds.get('source', 'compact_json') if thresholds else 'missing'})
    maps = [zone_maps['team_a'], zone_maps['team_b']]
    volume_high = any((_is_high(mapping.get('FGA')) or _is_high(mapping.get('Poss')) for mapping in maps))
    volume_low = all((_is_low(mapping.get('FGA')) or _is_low(mapping.get('Poss')) for mapping in maps))
    fta_high = any((_is_high(mapping.get('FTA')) or _is_high(mapping.get('FTr')) for mapping in maps))
    fta_low = all((_is_low(mapping.get('FTA')) or _is_low(mapping.get('FTr')) for mapping in maps))
    orb_high = any((_is_high(mapping.get('ORB')) for mapping in maps))
    orb_low = all((_is_low(mapping.get('ORB')) for mapping in maps))
    to_high = any((_is_high(mapping.get('TO')) for mapping in maps))
    to_not_high = all((not _is_high(mapping.get('TO')) for mapping in maps if mapping.get('TO') is not None))
    efg_low = any((_is_low(mapping.get('eFG')) for mapping in maps))
    efg_not_low = all((not _is_low(mapping.get('eFG')) for mapping in maps if mapping.get('eFG') is not None))
    opponent_allows = any((_is_high(mapping.get('allowed_FGA')) or _is_high(mapping.get('allowed_eFG')) or _is_high(mapping.get('allowed_Poss')) for mapping in maps))
    opponent_suppresses = any((_is_low(mapping.get('allowed_FGA')) or _is_low(mapping.get('allowed_eFG')) or _is_low(mapping.get('allowed_Poss')) for mapping in maps))
    score_or_efg_high = any((_is_high(mapping.get('scored')) or _is_high(mapping.get('eFG')) for mapping in maps))
    score_or_efg_low = any((_is_low(mapping.get('scored')) or _is_low(mapping.get('eFG')) for mapping in maps))
    flags = {'score_or_efg_high': score_or_efg_high, 'score_or_efg_low': score_or_efg_low, 'volume_high': volume_high, 'volume_low': volume_low, 'fta_high': fta_high, 'fta_low': fta_low, 'orb_high': orb_high, 'orb_low': orb_low, 'to_high': to_high, 'to_not_high': to_not_high, 'efg_low': efg_low, 'efg_not_low': efg_not_low, 'opponent_allows': opponent_allows, 'opponent_suppresses': opponent_suppresses}
    fake_over, fake_under = classify_fake_profiles(flags)
    over_channels = [name for name, active in {'FGA_OR_POSS_HIGH': volume_high, 'FTA_OR_FTR_HIGH': fta_high, 'ORB_ACTIVE': orb_high, 'TO_NOT_HIGH': to_not_high, 'EFG_NOT_LOW': efg_not_low, 'OPPONENT_ALLOWS': opponent_allows}.items() if active]
    under_channels = [name for name, active in {'FGA_OR_POSS_LOW': volume_low, 'FTA_OR_FTR_LOW': fta_low, 'ORB_LOW': orb_low, 'TO_HIGH_OR_EMPTY': to_high, 'EFG_LOW_WITHOUT_VOLUME': efg_low and (not volume_high), 'OPPONENT_SUPPRESSES': opponent_suppresses}.items() if active]
    canonical_support = canonical['stat_support']
    core_metrics = ('FGA', 'FTA', 'ORB', 'TO', 'Poss', 'eFG')
    available_core = sum(
        current_maps[label].get(metric) is not None
        for label in ('team_a', 'team_b')
        for metric in core_metrics
    )
    timing_reliable = all(
        bool(normalization[label].get('reliable_for_hard_gate'))
        for label in ('team_a', 'team_b')
    )
    if not project_counts_to_scope_end:
        # Exact pre-v5.3 behavior for the projection adjustment channel only.
        stat_support = canonical_support
        timing_reliable = True
    elif canonical_support == 'OFF' or available_core == 0:
        stat_support = 'OFF'
    elif canonical_support == 'LIMITED' or available_core < 6 or not timing_reliable:
        stat_support = 'LIMITED'
    else:
        stat_support = 'ON'
    if market['side'] == 'OVER':
        confirmed = len(over_channels) >= 3 and (not fake_over)
        against = len(under_channels) >= 3 and (not fake_under)
    else:
        confirmed = len(under_channels) >= 3 and (not fake_under)
        against = len(over_channels) >= 3 and (not fake_over)
    # An AGAINST result is a hard blocker only when the current counting stats were
    # compared on the same completed-segment time base and the snapshot is sufficiently
    # mature.  Otherwise it is diagnostic LIMITED support, never a false hard PASS.
    status = (
        'OFF' if stat_support == 'OFF'
        else 'CONFIRMED' if confirmed
        else 'AGAINST' if against and timing_reliable
        else 'LIMITED'
    )
    return {'scope': scope, 'team_a': comparisons['team_a'], 'team_b': comparisons['team_b'], 'zones': zone_maps, 'normalization': normalization, 'values_projected_to_scope_end': project_counts_to_scope_end and any(float(item.get('projection_factor') or 1.0) > 1.0 for item in normalization.values()), 'projection_profile': 'SCOPE_NORMALIZED_GATE' if project_counts_to_scope_end else 'LEGACY_LIVE_ADJUSTMENT', 'over_positive_channels': over_channels, 'under_positive_channels': under_channels, 'over_gate_score': len(over_channels), 'under_gate_score': len(under_channels), 'over_gate_status': 'CONFIRMED' if len(over_channels) >= 3 and (not fake_over) else 'NOT_CONFIRMED', 'under_gate_status': 'CONFIRMED' if len(under_channels) >= 3 and (not fake_under) else 'NOT_CONFIRMED', 'fake_over': fake_over, 'fake_under': fake_under, 'overheat_status': 'ON' if fake_over else 'OFF', 'bounce_risk': 'ON' if fake_under else 'OFF', 'real_over': len(over_channels) >= 3 and (not fake_over), 'real_under': len(under_channels) >= 3 and (not fake_under), 'indicators': flags, 'stat_support': stat_support, 'stat_gate_status': status}

# ===== q4_context_engine.py =====
def calculate_zone_rule_credibility(market: dict[str, Any], stat: dict[str, Any], zones_data: Optional[dict[str, Any]], config: dict[str, Any]) -> dict[str, Any]:
    matrix = (zones_data or {}).get('zone_hit_matrix') or []
    scope = 'MATCH' if market.get('segment') in {'MATCH','H1','H2'} else str(market.get('segment') or '').upper()
    if market.get('side') == 'OVER' and stat.get('fake_over'):
        rule_id = 'FAKE_OVER_BLOCK'
    elif market.get('side') == 'UNDER' and stat.get('fake_under'):
        rule_id = 'FAKE_UNDER_BLOCK'
    elif stat.get('stat_gate_status') == 'CONFIRMED':
        rule_id = 'OVER_Q_STRONG_01' if market.get('side') == 'OVER' else 'UNDER_Q_STRONG_01'
    else:
        return {'available': False, 'reason': 'NO_ACTIVE_ZONE_RULE'}
    row = next((item for item in matrix if item.get('rule_id') == rule_id and str(item.get('scope')).upper() == scope), None)
    if not row:
        return {'available': False, 'rule_id': rule_id, 'scope': scope, 'reason': 'MATRIX_ROW_MISSING'}
    cases = to_int(row.get('cases')) or 0
    hit_pct = to_number(row.get('hit_pct'))
    if hit_pct is None or cases <= 0:
        return {'available': False, 'rule_id': rule_id, 'scope': scope, 'cases': cases, 'reason': 'EMPTY_SAMPLE'}
    raw = hit_pct / 100.0
    k = float(config.get('credibility', {}).get('zone_matrix_k', 15.0))
    credibility = cases / (cases + k)
    shrunk = credibility * raw + (1.0 - credibility) * 0.5
    return {'available': True, 'rule_id': rule_id, 'scope': scope, 'cases': cases, 'raw_hit_rate': raw, 'credibility': credibility, 'shrunk_hit_rate': shrunk, 'quality': row.get('zone')}


def calculate_foul_conversion(pre_fta: Optional[float], pre_fouls: Optional[float], q3_fta: Optional[float], q3_fouls: Optional[float]) -> dict[str, Any]:
    pre = safe_div(pre_fta, pre_fouls)
    q3 = safe_div(q3_fta, q3_fouls)
    available = pre is not None or q3 is not None
    weighted = None
    if pre is not None and q3 is not None: weighted = 0.65 * pre + 0.35 * q3
    elif pre is not None: weighted = pre
    elif q3 is not None: weighted = q3
    level = 'UNKNOWN'
    if weighted is not None:
        level = 'LOW' if weighted < 0.70 else 'HIGH' if weighted >= 1.05 else 'MID'
    return {'available': available, 'pre_q4_fta_per_foul': pre, 'q3_fta_per_foul': q3, 'score': weighted, 'level': level}


def classify_blowout_context(abs_margin: Optional[float], indicators: dict[str, Any], q3_total: Optional[float], threshold: float) -> dict[str, Any]:
    if abs_margin is None or abs_margin < threshold:
        return {'active': False, 'classification': 'NONE', 'dry_evidence': 0, 'kill_evidence': 0}
    dry = sum(bool(indicators.get(key)) for key in ('volume_low','orb_low','to_high'))
    kill = sum(bool(indicators.get(key)) for key in ('volume_high','orb_high','fta_high','to_not_high'))
    if q3_total is not None:
        if q3_total <= 36: dry += 1
        elif q3_total >= 48: kill += 1
    classification = 'DRY' if dry >= kill + 1 and dry >= 2 else 'KILL' if kill >= dry + 1 and kill >= 2 else 'MIXED'
    return {'active': True, 'classification': classification, 'dry_evidence': dry, 'kill_evidence': kill, 'q3_total': q3_total}


def calculate_ot_tail(market: dict[str, Any], live_margin: float, minutes_left: float, bonus_path: bool, exact_bonus: bool) -> dict[str, Any]:
    applicable = market.get('market_type') == 'MATCH_TOTAL' and market.get('segment') == 'MATCH' and minutes_left <= 5.0
    score = 0.0
    if applicable:
        score += 0.45 if live_margin <= 2 else 0.30 if live_margin <= 5 else 0.0
        score += 0.25 if minutes_left <= 2.0 else 0.15 if minutes_left <= 4.0 else 0.0
        score += 0.20 if bonus_path else 0.0
        score += 0.10 if exact_bonus and bonus_path else 0.0
    return {'applicable': applicable, 'score': min(1.0, score), 'level': 'HIGH' if score >= 0.70 else 'MEDIUM' if score >= 0.45 else 'LOW'}


def weighted_harmonic_mean(values: dict[str, Optional[float]], weights: dict[str, float], epsilon: float=1e-06) -> Optional[float]:
    active = [(float(weights[key]), float(value)) for key, value in values.items() if key in weights and value is not None and (weights[key] > 0)]
    if not active:
        return None
    numerator = sum((weight for weight, _ in active))
    denominator = sum((weight / max(value, epsilon) for weight, value in active))
    return numerator / denominator if denominator else None

def _quarter_sum(raw: dict[str, Any], code: str, quarters: list[int]) -> Optional[float]:
    values: list[float] = []
    for quarter in quarters:
        home = to_number(raw.get(f'h{code}{quarter}'))
        away = to_number(raw.get(f'a{code}{quarter}'))
        if home is None or away is None:
            return None
        values.append(home + away)
    return sum(values)

def _quarter_total(canonical: dict[str, Any], number: int) -> Optional[float]:
    if number < 1 or number > len(canonical['quarters']):
        return None
    return canonical['quarters'][number - 1].get('total')

def is_q4_context_market(market: dict[str, Any], canonical: dict[str, Any]) -> bool:
    if market['market_type'] in {'CURRENT_QUARTER_TOTAL', 'CURRENT_QUARTER_TEAM_IT'} and market.get('segment') == 'Q4':
        return True
    return market['market_type'] in {'MATCH_TOTAL', 'TEAM_IT_MATCH'} and canonical['stage'] in {'AFTER_3Q', 'Q4_CONFIRMATION'}

def calculate_q4_context(market: dict[str, Any], canonical: dict[str, Any], history: dict[str, Any], scenario: dict[str, Any], live: dict[str, Any], stat: dict[str, Any], config: dict[str, Any]) -> dict[str, Any]:
    applicable = is_q4_context_market(market, canonical)
    if not applicable:
        return {'applicable': False, 'status': 'OFF'}
    exact_time = canonical.get('clock') is not None and canonical.get('current_quarter') == 4
    raw = canonical.get('raw_main', {})
    pre_fouls = _quarter_sum(raw, 'fls', [1, 2, 3])
    pre_fta = _quarter_sum(raw, 'fta', [1, 2, 3])
    pre_fga = _quarter_sum(raw, 'fga', [1, 2, 3])
    q3_fouls = _quarter_sum(raw, 'fls', [3])
    q3_fta = _quarter_sum(raw, 'fta', [3])
    q3_total = _quarter_total(canonical, 3)
    total_after_3q_values = [_quarter_total(canonical, number) for number in (1, 2, 3)]
    total_after_3q = sum(total_after_3q_values) if all((value is not None for value in total_after_3q_values)) else None
    q3_home = sum((float(canonical['quarters'][i].get('home') or 0) for i in range(3)))
    q3_away = sum((float(canonical['quarters'][i].get('away') or 0) for i in range(3)))
    abs_margin_3q = abs(q3_home - q3_away) if total_after_3q is not None else None
    live_margin = abs(float(canonical['score']['margin_home']))
    q4_total = _quarter_total(canonical, 4) or 0.0
    q4_minutes_left = canonical.get('quarter_seconds_remaining', 0) / 60
    pre_ftr = pre_fta / pre_fga if pre_fta is not None and pre_fga else None
    thresholds = deepcopy(config['q4']['thresholds'])
    duration_factor = float(canonical.get('quarter_minutes') or 10) / 10.0
    for key in ('pre_fouls_low', 'pre_fouls_high', 'pre_fouls_very_high', 'q3_fouls_high', 'pre_fta_high', 'pre_fta_low', 'q3_fta_high', 'total_after_3q_high'):
        thresholds[key] = float(thresholds[key]) * duration_factor
    playoff = bool((canonical.get('series_context') or {}).get('is_playoff'))
    must_win = bool((canonical.get('series_context') or {}).get('must_win'))
    chase_margin = abs_margin_3q is not None and thresholds['chase_margin_low'] <= abs_margin_3q <= thresholds['chase_margin_high'] or thresholds['chase_margin_low'] - 2 <= live_margin <= thresholds['chase_margin_high']
    foul_tail = 0.0
    foul_tail += 0.22 if chase_margin else 0.0
    foul_tail += 0.18 if pre_fouls is not None and pre_fouls >= thresholds['pre_fouls_high'] else 0.0
    foul_tail += 0.2 if pre_fta is not None and pre_fta >= thresholds['pre_fta_high'] else 0.0
    foul_tail += 0.15 if q3_fouls is not None and q3_fouls >= thresholds['q3_fouls_high'] else 0.0
    foul_tail += 0.1 if q3_fta is not None and q3_fta >= thresholds['q3_fta_high'] else 0.0
    foul_tail += 0.1 if total_after_3q is not None and total_after_3q >= thresholds['total_after_3q_high'] else 0.0
    foul_tail += 0.05 if playoff else 0.0
    indicators = stat.get('indicators', {})
    volume_low = bool(indicators.get('volume_low'))
    orb_low = bool(indicators.get('orb_low'))
    to_high = bool(indicators.get('to_high'))
    orb_high = bool(indicators.get('orb_high'))
    to_not_high = bool(indicators.get('to_not_high'))
    fta_high = bool(indicators.get('fta_high'))
    efg_not_low = bool(indicators.get('efg_not_low'))
    no_chase = not chase_margin and (not playoff) and (not must_win)
    dry = 0.0
    dry += 0.22 if pre_fouls is not None and pre_fouls <= thresholds['pre_fouls_low'] else 0.0
    dry += 0.18 if pre_fta is not None and pre_fta < thresholds['pre_fta_low'] else 0.0
    blowout = classify_blowout_context(abs_margin_3q, indicators, q3_total, thresholds['blowout_margin'])
    dry += 0.18 if blowout.get('classification') == 'DRY' else 0.0
    dry += 0.15 if volume_low else 0.0
    dry += 0.12 if orb_low else 0.0
    dry += 0.1 if to_high else 0.0
    dry += 0.05 if no_chase else 0.0
    q4_three_pa = _quarter_sum(raw, '3pa', [4])
    three_pa_chase = bool(q4_three_pa is not None and q4_three_pa >= 8 * duration_factor and chase_margin)
    bonus_context = canonical.get('bonus_context') or {}
    exact_bonus_threshold = int(config['q4'].get('bonus_team_fouls', 4))
    exact_bonus_flag = bool(bonus_context.get('home_in_bonus') or bonus_context.get('away_in_bonus'))
    exact_bonus_fouls = any(
        value is not None and int(value) >= exact_bonus_threshold
        for value in (
            bonus_context.get('home_q4_team_fouls'),
            bonus_context.get('away_q4_team_fouls'),
        )
    )
    exact_bonus_path = bool(exact_bonus_flag or exact_bonus_fouls) if bonus_context.get('exact_available') else False
    inferred_bonus_path = bool((pre_fouls is not None and pre_fouls >= thresholds['pre_fouls_high']) or (pre_fta is not None and pre_fta >= thresholds['pre_fta_high']))
    bonus_path = exact_bonus_path if bonus_context.get('exact_available') else inferred_bonus_path
    bonus_source = 'EXACT_FLAG' if exact_bonus_flag else 'EXACT_Q4_FOULS' if bonus_context.get('exact_available') else 'INFERRED_PRE_Q4'
    foul_conversion = calculate_foul_conversion(pre_fta, pre_fouls, q3_fta, q3_fouls)
    leader_ft_path = bool(chase_margin and fta_high)
    kill_chase = 0.0
    kill_chase += 0.2 if 4 <= live_margin <= 10 else 0.0
    kill_chase += 0.18 if three_pa_chase else 0.0
    kill_chase += 0.17 if bonus_path else 0.0
    kill_chase += 0.15 if orb_high else 0.0
    kill_chase += 0.15 if to_not_high else 0.0
    kill_chase += 0.1 if leader_ft_path else 0.0
    kill_chase += 0.05 if playoff else 0.0
    volume = 0.0
    volume += 0.25 if indicators.get('volume_high') else 0.0
    volume += 0.2 if fta_high else 0.0
    volume += 0.2 if orb_high else 0.0
    volume += 0.2 if to_not_high else 0.0
    volume += 0.15 if efg_not_low else 0.0
    epsilon = float(config['q4'].get('epsilon', 1e-06))
    under_values = {'hist': history['p_hist'], 'scenario': scenario['p_scenario'], 'live': live['p_live'], 'dry': dry, 'no_foul_tail': 1 - foul_tail, 'no_kill_chase': 1 - kill_chase}
    over_values = {'hist': history['p_hist'], 'scenario': scenario['p_scenario'], 'live': live['p_live'], 'foul_tail': foul_tail, 'kill_chase': kill_chase, 'volume': volume}
    under_gate = weighted_harmonic_mean(under_values, config['q4']['under_weights'], epsilon)
    over_gate = weighted_harmonic_mean(over_values, config['q4']['over_weights'], epsilon)
    line_edge_over = live['line_edge_over']
    over_boost = 0.05 if foul_tail >= 0.7 and line_edge_over >= 4 else 0.03 if foul_tail >= 0.55 and line_edge_over >= 0 else 0.0
    if foul_conversion.get('level') == 'LOW':
        over_boost = 0.0
    line_edge_bonus = 0.05 if line_edge_over >= 7 else 0.03 if line_edge_over >= 4 else 0.0
    context_gate = under_gate if market['side'] == 'UNDER' else over_gate
    ot_tail = calculate_ot_tail(market, live_margin, q4_minutes_left, bonus_path, bool(bonus_context.get('exact_available')))
    mandatory_missing = []
    if not exact_time:
        mandatory_missing.append('exact_q4_time_or_start')
    if total_after_3q is None:
        mandatory_missing.append('total_after_3q')
    if abs_margin_3q is None:
        mandatory_missing.append('abs_margin_3q')
    return {'applicable': True, 'status': 'ON' if not mandatory_missing else 'MISSING_CONTEXT', 'exact_time': exact_time, 'mandatory_missing': mandatory_missing, 'duration_factor_vs_4x10': duration_factor, 'duration_adjusted_thresholds': thresholds, 'pre_q4_fouls_total': pre_fouls, 'pre_q4_fta_total': pre_fta, 'pre_q4_ftr': pre_ftr, 'q3_fouls_total': q3_fouls, 'q3_fta_total': q3_fta, 'q3_total': q3_total, 'total_after_3q': total_after_3q, 'abs_margin_3q': abs_margin_3q, 'q4_current_total': q4_total, 'q4_minutes_left': q4_minutes_left, 'live_margin_q4': live_margin, 'foul_tail_score': foul_tail, 'dry_score': dry, 'kill_chase_score': kill_chase, 'volume_score': volume, 'under_gate_h': under_gate, 'over_gate_h': over_gate, 'context_gate': context_gate, 'over_boost': over_boost, 'line_edge_bonus': line_edge_bonus, 'bonus_path': bonus_path, 'bonus_source': bonus_source, 'bonus_context': bonus_context, 'foul_conversion': foul_conversion, 'blowout_context': blowout, 'ot_tail': ot_tail, 'three_pa_chase': three_pa_chase, 'leader_ft_path': leader_ft_path, 'chase_margin': chase_margin}

# ===== super_basket_calculator.py =====
def _normalize_weights(weights: dict[str, float]) -> tuple[dict[str, float], bool]:
    total = sum((float(value) for value in weights.values()))
    if total <= 0:
        return ({'hist': 1.0, 'scenario': 0.0, 'live': 0.0}, True)
    normalized = {key: float(value) / total for key, value in weights.items()}
    return (normalized, abs(total - 1.0) > 1e-12)

def _cap(rule_id: str, cap: float, reason: str, inputs: Optional[dict[str, Any]]=None) -> dict[str, Any]:
    return {'rule_id': rule_id, 'cap': float(cap), 'reason': reason, 'inputs': inputs or {}}

def _blocker(rule_id: str, reason: str, inputs: Optional[dict[str, Any]]=None) -> dict[str, Any]:
    return {'rule_id': rule_id, 'reason': reason, 'inputs': inputs or {}}

def _trace_step(step: str, applied: bool, formula: str, inputs: dict[str, Any], before: Optional[float], after: Optional[float], reason_codes: Optional[list[str]]=None) -> dict[str, Any]:
    return {
        'step': step,
        'applied': bool(applied),
        'formula': formula,
        'inputs': inputs,
        'probability_before': before,
        'probability_after': after,
        'reason_codes': reason_codes or [],
    }

def _strong_history_risk_band(
    probability: float,
    p_hist: Optional[float],
) -> bool:
    return (
        0.60 <= float(probability) < 0.65
        and p_hist is not None
        and float(p_hist) > 0.80
    )

def _verdict(
    probability: float,
    blockers: list[dict[str, Any]],
    strong_clean: bool,
    p_hist: Optional[float]=None,
) -> str:
    if blockers or probability < 0.60:
        return 'PASS'
    if probability < 0.75:
        return 'RISK PLAY'
    if probability < 0.8:
        return 'LIVE PLAY'
    return 'PLAY' if strong_clean else 'LIVE PLAY'

def _router(market: dict[str, Any], canonical: dict[str, Any]) -> dict[str, Any]:
    market_type = market['market_type']
    stage = canonical['stage']
    current = canonical.get('current_quarter')
    status, reason, cap = ('ALLOW', 'SUPPORTED_BY_STAGE_ROUTER', None)
    hard_block = False
    trigger_checkpoint = canonical.get('trigger_checkpoint')
    in_q2_window = canonical['quarter_seconds'] <= canonical['elapsed_game_seconds'] < canonical['full_game_seconds'] / 2
    if trigger_checkpoint == 1 and market_type not in {'H1_TOTAL', 'TEAM_IT_H1'}:
        # Absolute checkpoint-level protection. The Q1 job may start parsing a
        # little later (already in Q2 or even at HT), so stage inference alone
        # is not sufficient. A job triggered after Q1 may only emit first-half
        # total / first-half team-IT; full-match total and TEAM_IT_MATCH are
        # always blocked for this job.
        status, reason, hard_block = ('BLOCK', 'PRODUCTION_ROUTER_BLOCK_AFTER_Q1_ONLY_H1_TOTAL_AND_TEAM_IT_H1', True)
    elif market_type in {'MATCH_TOTAL', 'TEAM_IT_MATCH'} and (current == 2 or in_q2_window):
        # Checkpoint #1 (stage_monitor.js) opens the analysis window right after
        # Q1 ends, i.e. during Q2 — the whole Q2 window, until half-time (which
        # is Checkpoint #2). In that window a full-match total/team-IT signal is
        # too early and noisy; only the half-scoped markets (H1_TOTAL/TEAM_IT_H1)
        # are allowed to fire here.
        # NOTE: gated on elapsed game time (in_q2_window), not only on the
        # provider's raw `current_quarter` field. Some feeds keep `period` at 1
        # during the break right after Q1 ends (score/time already reflect a
        # finished Q1) until Q2 officially tips off, which let full-match total
        # signals slip through the old `current == 2`-only check for that
        # window.
        status, reason, hard_block = ('BLOCK', 'MATCH_TOTAL_BLOCKED_DURING_Q2_AFTER_CHECKPOINT1_ONLY_HALF_MARKETS', True)
    elif market_type == 'H1_TOTAL' or market_type == 'TEAM_IT_H1':
        if canonical['elapsed_game_seconds'] >= canonical['full_game_seconds'] / 2:
            status, reason, hard_block = ('BLOCK', 'H1_ALREADY_COMPLETE', True)
        elif current == 1:
            status, reason = ('DOWNGRADE', 'H1_BEFORE_Q1_COMPLETION')
    elif market_type == 'H2_TOTAL' or market_type == 'TEAM_IT_H2':
        if canonical['elapsed_game_seconds'] < canonical['full_game_seconds'] / 2:
            status, reason, hard_block = ('BLOCK', 'FUTURE_H2_BEFORE_HT', True)
    elif market_type in {'CURRENT_QUARTER_TOTAL', 'CURRENT_QUARTER_TEAM_IT'}:
        if market.get('segment') in {'Q2', 'Q3'}:
            status, reason, cap = ('DOWNGRADE', 'Q2_Q3_STANDALONE_NO_CLEAN_PLAY', 0.74)
        elif market.get('segment') == 'Q4':
            status, reason = ('CONTEXT_GATE', 'Q4_REQUIRES_CONTEXT_GATE')
    elif market_type in {'MATCH_TOTAL', 'TEAM_IT_MATCH'} and stage == 'AFTER_3Q':
        status, reason = ('PRIORITY', 'AFTER_3Q_PRIORITY')
    elif market_type in {'MATCH_TOTAL', 'TEAM_IT_MATCH', 'H2_TOTAL', 'TEAM_IT_H2'} and stage == 'HT':
        status, reason = ('PRIORITY', 'HT_PRIORITY')
    return {'status': status, 'reason': reason, 'cap': cap, 'hard_block': hard_block}

def _strong_edge_threshold(market_type: str, config: dict[str, Any]) -> float:
    return float(config['strong_live_edge'].get(market_type, 7.0))


def calculate_live_dominance(
    market: dict[str, Any],
    canonical: dict[str, Any],
    live: dict[str, Any],
    stat: dict[str, Any],
    config: dict[str, Any],
) -> dict[str, Any]:
    """Constrained v5.5 live-dominance exception; never bypasses blockers."""
    market_type = str(market.get('market_type') or '')
    thresholds = config.get('live_dominance', {}).get('edge_thresholds', {})
    required_edge = to_number(thresholds.get(market_type))
    side = str(market.get('side') or '').upper()
    line = to_number(market.get('line'))
    edge = to_number(live.get('line_edge_over' if side == 'OVER' else 'line_edge_under'))
    model_live = to_number(live.get('projection_model_live'))
    stat_adjusted = to_number(live.get('projection_stat_adjusted'))

    def supports(value: Optional[float]) -> bool:
        if value is None or line is None:
            return False
        return value > line if side == 'OVER' else value < line

    fake_profile = bool(stat.get('fake_over')) if side == 'OVER' else bool(stat.get('fake_under'))
    checks = {
        'live_stage': canonical.get('stage') != 'PRE_MATCH',
        'supported_market': required_edge is not None,
        'stat_support_on': stat.get('stat_support') == 'ON',
        'stat_gate_confirmed': stat.get('stat_gate_status') == 'CONFIRMED',
        'clean_pace_profile': not fake_profile,
        'p_live_at_least_65': float(live.get('p_live') or 0.0) >= 0.65,
        'edge_threshold_met': required_edge is not None and edge is not None and edge >= required_edge,
        'projection_model_live_aligned': supports(model_live),
        'projection_stat_adjusted_aligned': supports(stat_adjusted),
    }
    active = all(checks.values())
    return {
        'active': active,
        'rule_id': 'LIVE_DOMINANCE',
        'side': side,
        'line': line,
        'live_edge': edge,
        'required_edge': required_edge,
        'p_live': live.get('p_live'),
        'projection_model_live': model_live,
        'projection_stat_adjusted': stat_adjusted,
        'checks': checks,
        'failed_checks': [name for name, passed in checks.items() if not passed],
        'max_live_weight': float(config.get('live_dominance', {}).get('max_live_weight', 0.80)),
    }

def _empty_evaluation(market: dict[str, Any], blockers: list[dict[str, Any]]) -> dict[str, Any]:
    return {**market, 'history': {'p_hist': 0.5}, 'scenario': {'p_scenario': 0.5, 'scenario_support': 'OFF', 'patterns_found': [], 'patterns_used': [], 'patterns_rejected': []}, 'live': {'projection_used': None, 'p_live': 0.5}, 'stat_comparison': {'stat_gate_status': 'OFF', 'fake_over': False, 'fake_under': False}, 'q4_context': {'applicable': False, 'status': 'OFF'}, 'weights': {'original': {}, 'normalized': {}, 'normalization_applied': False}, 'p_raw': 0.5, 'router': {'status': 'BLOCK', 'reason': 'INVALID_MARKET'}, 'caps': [], 'blockers': blockers, 'hard_conflict': True, 'p_final': 0.5, 'verdict': 'PASS', 'p_trace': [_trace_step('INVALID_MARKET', True, 'hard_block', {'blockers': blockers}, None, 0.5, [item['rule_id'] for item in blockers])]}

def _dedupe_markets(markets: list[dict[str, Any]], odds_min: float) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    grouped: dict[tuple[Any, ...], list[dict[str, Any]]] = {}
    for market in markets:
        key = (
            market.get('market_type'),
            market.get('team'),
            market.get('segment'),
            market.get('side'),
            to_number(market.get('line')),
        )
        grouped.setdefault(key, []).append(market)
    unique: list[dict[str, Any]] = []
    for key, offers in grouped.items():
        ordered = sorted(offers, key=lambda item: float(item.get('odds') or 0.0), reverse=True)
        selected = deepcopy(ordered[0])
        selected['offers'] = [
            {
                'market_id': item.get('market_id'),
                'bookmaker': item.get('bookmaker'),
                'odds': item.get('odds'),
                'source_market_id': item.get('source_market_id'),
            }
            for item in ordered
        ]
        selected['duplicate_offer_count'] = max(0, len(ordered) - 1)
        issues = [issue for issue in selected.get('parser_issues', []) if issue != 'ODDS_BELOW_MINIMUM']
        if selected.get('odds') is None or float(selected['odds']) < odds_min:
            issues.append('ODDS_BELOW_MINIMUM')
        selected['parser_issues'] = sorted(set(issues))
        selected['eligible_market'] = not selected['parser_issues']
        stable_key = '|'.join(str(part) for part in key)
        selected['math_market_key'] = hashlib.sha256(stable_key.encode('utf-8')).hexdigest()[:16]
        unique.append(selected)
    unique.sort(key=lambda item: (str(item.get('market_type')), str(item.get('team')), str(item.get('segment')), float(item.get('line') or 0), str(item.get('side'))))
    return unique, {
        'offer_sides_before_deduplication': len(markets),
        'unique_market_sides': len(unique),
        'duplicate_offers_removed': len(markets) - len(unique),
    }


def _market_semantic_issues(market: dict[str, Any], canonical: dict[str, Any]) -> list[dict[str, Any]]:
    """Hard guards against comparing a line from one segment with another segment's score."""
    issues: list[dict[str, Any]] = []
    line = to_number(market.get('line'))
    if line is None:
        return issues
    try:
        clock = _segment_clock(market, canonical)
    except Exception:
        return issues
    current_points = to_number(clock.get('current_points'))
    remaining_seconds = to_number(clock.get('remaining_seconds'))
    if current_points is not None and remaining_seconds is not None and remaining_seconds > 0 and line <= current_points:
        issues.append(_blocker(
            'LINE_BELOW_CURRENT_SCOPE_SCORE',
            'Bookmaker line is not above the points already scored in the same market scope; probable segment mismatch',
            {
                'market_type': market.get('market_type'),
                'segment': market.get('segment'),
                'source_bucket': market.get('source_bucket'),
                'source_scope': market.get('source_scope'),
                'line': line,
                'current_scope_points': current_points,
            },
        ))
    # A match-scoped line must come from a Match-scoped row. The parser derives the
    # market from scope, but retaining this guard protects against hand-edited/enriched JSON.
    source_scope = str(market.get('source_scope') or '').upper().replace(' ', '')
    segment = str(market.get('segment') or '').upper()
    if source_scope and segment == 'MATCH' and source_scope not in {'MATCH', 'FULLMATCH', 'FT'}:
        issues.append(_blocker(
            'SOURCE_SCOPE_SEGMENT_MISMATCH',
            'Source scope does not match normalized market segment',
            {'source_scope': source_scope, 'normalized_segment': segment},
        ))
    return issues

class SuperBasketCalculator:

    def __init__(
        self,
        config: dict[str, Any],
        zones_data: Optional[dict[str, Any]]=None,
        zones_metadata: Optional[dict[str, Any]]=None,
    ) -> None:
        self.config = deepcopy(config)
        self.zones_data = zones_data or {}
        self.zones_metadata = zones_metadata or _zone_table_metadata(
            self.zones_data,
            source='PROVIDED_OBJECT' if self.zones_data else 'UNRESOLVED',
        )

    @classmethod
    def from_files(cls, config_path: str | Path, zones_path: Optional[str | Path]=None) -> 'SuperBasketCalculator':
        with open(config_path, 'r', encoding='utf-8') as handle:
            config = json.load(handle)
        zones, metadata = resolve_team_relative_zones({}, zones_path=zones_path)
        return cls(config, zones, metadata)

    def evaluate_market(self, market: dict[str, Any], canonical: dict[str, Any]) -> dict[str, Any]:
        initial_blockers = [_blocker(issue, issue.replace('_', ' ').title()) for issue in market.get('parser_issues', [])]
        initial_blockers.extend(_market_semantic_issues(market, canonical))
        if market.get('line') is None or market.get('odds') is None or market.get('market_type') == 'UNSUPPORTED' or initial_blockers:
            return _empty_evaluation(market, initial_blockers)
        router_preview = _router(market, canonical)
        if router_preview.get('hard_block'):
            blocked = _empty_evaluation(market, [_blocker('PRODUCTION_ROUTER_BLOCK', router_preview['reason'])])
            blocked['router'] = router_preview
            return blocked
        history = calculate_history(market, canonical, self.config)
        scenario = calculate_scenario(market, canonical, history, self.config)
        stat = calculate_stat_gate(market, canonical, self.zones_data, project_counts_to_scope_end=True)
        projection_stat = calculate_stat_gate(market, canonical, self.zones_data, project_counts_to_scope_end=False)
        live = calculate_live_projection(market, canonical, history, scenario, self.config, projection_stat)
        q4 = calculate_q4_context(market, canonical, history, scenario, live, projection_stat, self.config)
        zone_credibility = calculate_zone_rule_credibility(market, stat, self.zones_data, self.config)
        stat['legacy_projection_adjustment_profile'] = {
            'stat_gate_status': projection_stat.get('stat_gate_status'),
            'indicators': deepcopy(projection_stat.get('indicators') or {}),
        }
        stage_key = canonical['stage']
        original_weights = deepcopy(self.config['stage_weights'].get(stage_key, self.config['stage_weights']['EARLY_LIVE']))
        base_weights, normalization_applied = _normalize_weights(original_weights)
        live_dominance = calculate_live_dominance(market, canonical, live, stat, self.config)
        normalized_weights = deepcopy(base_weights)
        if live_dominance['active']:
            live_weight = min(0.80, float(live_dominance['max_live_weight']))
            non_live = max(0.0, 1.0 - live_weight)
            base_non_live = base_weights['hist'] + base_weights['scenario']
            if base_non_live > 0:
                normalized_weights['hist'] = non_live * base_weights['hist'] / base_non_live
                normalized_weights['scenario'] = non_live * base_weights['scenario'] / base_non_live
            else:
                normalized_weights['hist'] = non_live
                normalized_weights['scenario'] = 0.0
            normalized_weights['live'] = live_weight
        gates_cfg = self.config.get('signal_gates', {})
        history_zone_rate = to_number(history.get('history_zone_rate'))
        history_zone_min = float(gates_cfg.get('history_zone_min', 0.75))
        strong_edge_for_reversal = _strong_edge_threshold(market['market_type'], self.config)
        fake_profile = bool(
            (market['side'] == 'OVER' and stat.get('fake_over'))
            or (market['side'] == 'UNDER' and stat.get('fake_under'))
        )
        live_reversal_active = bool(
            gates_cfg.get('allow_live_reversal', True)
            and canonical['stage'] != 'PRE_MATCH'
            and live.get('p_live', 0.0) >= float(gates_cfg.get('live_reversal_p_live_min', 0.80))
            and scenario.get('p_scenario', 0.0) >= float(gates_cfg.get('live_reversal_p_scenario_min', 0.68))
            and live.get('line_edge', -999.0) >= strong_edge_for_reversal
            and stat.get('stat_gate_status') == 'CONFIRMED'
            and not fake_profile
        )
        if live_reversal_active:
            # Explicit project reversal weights: history 11.5%, scenario 8.5%, live 80%.
            normalized_weights = {'hist': 0.115, 'scenario': 0.085, 'live': 0.80}
        p_raw = normalized_weights['hist'] * history['p_hist'] + normalized_weights['scenario'] * scenario['p_scenario'] + normalized_weights['live'] * live['p_live']
        router = _router(market, canonical)
        caps: list[dict[str, Any]] = []
        blockers = list(initial_blockers)
        gates_cfg = self.config.get('signal_gates', {})
        history_zone_rate = to_number(history.get('history_zone_rate'))
        history_zone_min = float(gates_cfg.get('history_zone_min', 0.75))
        live_edge_min = float(gates_cfg.get('live_edge_min_points', 3.0))
        scenario_direction_min = float(gates_cfg.get('scenario_direction_min', 0.50))
        if (history_zone_rate is None or history_zone_rate < history_zone_min) and not live_reversal_active:
            blockers.append(_blocker(
                'HISTORY_ZONE_BELOW_75',
                'Signal requires at least a 75% exact-line historical zone in the same direction',
                {'history_zone_rate': history_zone_rate, 'required': history_zone_min, 'source': history.get('history_zone_source')},
            ))
        if canonical['stage'] != 'PRE_MATCH' and float(live.get('line_edge') or -999.0) < live_edge_min:
            blockers.append(_blocker(
                'LIVE_EDGE_BELOW_3',
                'Live projection must be at least 3 points beyond the bookmaker line in the signal direction',
                {'line_edge': live.get('line_edge'), 'required': live_edge_min, 'projection_used': live.get('projection_used'), 'line': market.get('line')},
            ))
        if scenario.get('scenario_support') == 'ON' and float(scenario.get('p_scenario') or 0.0) < scenario_direction_min:
            blockers.append(_blocker(
                'SCENARIO_DIRECTION_CONFLICT',
                'Matched historical states point against the evaluated side',
                {'p_scenario': scenario.get('p_scenario'), 'required': scenario_direction_min},
            ))
        if canonical['data_gate']['schema_errors']:
            blockers.append(_blocker('SCHEMA_ERROR', 'Required canonical fields are missing', {'paths': canonical['data_gate']['schema_errors']}))
        if router.get('cap') is not None:
            caps.append(_cap('PRODUCTION_ROUTER_DOWNGRADE', router['cap'], router['reason']))
        if router.get('hard_block'):
            blockers.append(_blocker('PRODUCTION_ROUTER_BLOCK', router['reason']))
        if market.get('segment') == 'Q3' and market['market_type'] in {'CURRENT_QUARTER_TOTAL', 'CURRENT_QUARTER_TEAM_IT'}:
            if p_raw < 0.80:
                blockers.append(_blocker('Q3_EXCEPTIONAL_PROBABILITY_BELOW_80', 'Standalone Q3 requires model probability of at least 80%', {'p_raw': p_raw}))
            if stat.get('stat_support') != 'ON':
                blockers.append(_blocker('Q3_EXCEPTIONAL_STATS_REQUIRED', 'Standalone Q3 requires complete live statistics'))
        same_format_n = int(canonical['data_gate'].get('pooled_n') or 0)
        if same_format_n == 0:
            caps.append(_cap('NO_SAME_FORMAT_HISTORY', 0.67, 'No same-duration history is available'))
            blockers.append(_blocker('NO_SAME_FORMAT_HISTORY', 'Exact-line history from a different game duration cannot be used'))
        elif same_format_n < 20:
            caps.append(_cap('SMALL_SAME_FORMAT_SAMPLE', self.config['caps']['small_sample'], 'Same-format history sample is below 20 games', {'same_format_pooled_n': same_format_n}))
        live_mode = canonical['stage'] != 'PRE_MATCH'
        if live_mode and stat['stat_support'] == 'OFF':
            # A physically consistent score/clock fallback remains eligible, but
            # it must never be promoted to a clean PLAY without live statistics.
            caps.append(_cap('NO_STATS_FALLBACK', self.config['caps']['stat_off'], 'Live statistics unavailable; RISK PLAY only'))
        elif live_mode and stat['stat_support'] == 'LIMITED':
            caps.append(_cap('STAT_SUPPORT_LIMITED', self.config['caps']['stat_limited'], 'Incomplete live statistics'))
        credibility_cfg = self.config.get('credibility', {})
        zone_credibility_mature = (
            zone_credibility.get('available')
            and zone_credibility.get('cases', 0) >= int(credibility_cfg.get('zone_matrix_min_cases', 10))
        )
        zone_credibility_low = (
            zone_credibility_mature
            and zone_credibility.get('shrunk_hit_rate', 0.5) < float(credibility_cfg.get('zone_matrix_low_threshold', 0.55))
        )
        if stat['stat_gate_status'] == 'AGAINST':
            if zone_credibility_low:
                # An AGAINST gate learned from a demonstrably weak zone rule
                # must not be treated as a hard contradiction. Keep the
                # candidate eligible only as capped RISK and expose the weak
                # evidence in the audit trail.
                caps.append(_cap(
                    'STAT_GATE_AGAINST_LOW_CREDIBILITY',
                    self.config['caps'].get('zone_matrix_low', 0.74),
                    'Opposing stat-zone rule has weak out-of-sample credibility; downgrade to RISK instead of hard block',
                    zone_credibility,
                ))
            else:
                blockers.append(_blocker('STAT_GATE_DIRECTLY_AGAINST', 'Team-relative stat channels oppose this side', {'side': market['side']}))
        if market['side'] == 'OVER' and stat['fake_over']:
            caps.append(_cap('FAKE_OVER', self.config['caps']['fake_over'], 'High score/efficiency is not supported by volume'))
        if market['side'] == 'UNDER' and stat['fake_under']:
            caps.append(_cap('FAKE_UNDER', self.config['caps']['fake_under'], 'Low score has high-volume bounce risk'))
        if zone_credibility_low and stat['stat_gate_status'] != 'AGAINST':
            caps.append(_cap('ZONE_MATRIX_LOW_CREDIBILITY', self.config['caps'].get('zone_matrix_low', 0.74), 'Active stat-zone rule has weak out-of-sample credibility', zone_credibility))
        strong_edge = _strong_edge_threshold(market['market_type'], self.config)
        if history['p_hist'] >= 0.9 and live['line_edge'] <= -strong_edge:
            blockers.append(_blocker('STRONG_HISTORY_LIVE_CONFLICT', 'History side is blocked by a strong opposite live projection', {'p_hist': history['p_hist'], 'line_edge': live['line_edge'], 'threshold': strong_edge}))
        if market['market_type'].startswith('TEAM_IT') or market['market_type'] == 'CURRENT_QUARTER_TEAM_IT':
            opponent_allowed = history.get('opponent_allowed', {})
            weakest = history.get('weakest_gate')
            if not opponent_allowed or opponent_allowed.get('n', 0) == 0:
                blockers.append(_blocker('TEAM_IT_NO_OPPONENT_ALLOWED', 'Opponent allowed history is mandatory'))
            if weakest is None:
                blockers.append(_blocker('TEAM_IT_WEAKEST_MISSING', 'Own scored/opponent allowed gate is unavailable'))
            elif weakest < 0.7:
                caps.append(_cap('TEAM_IT_WEAKEST_BELOW_70', self.config['caps']['team_it_weak'], 'Weakest Team IT gate below 70%', {'weakest': weakest}))
                blockers.append(_blocker('TEAM_IT_WEAKEST_BLOCK', 'Weakest Team IT gate below 70%', {'weakest': weakest}))
            elif weakest < 0.75:
                caps.append(_cap('TEAM_IT_WEAKEST_70_74', self.config['caps']['team_it_70_74'], 'Weakest Team IT gate is 70-74%', {'weakest': weakest}))
            elif weakest < 0.8:
                caps.append(_cap('TEAM_IT_WEAKEST_75_79', self.config['caps']['team_it_75_79'], 'Weakest Team IT gate is 75-79%', {'weakest': weakest}))
            required_ppm = history.get('required_points_per_minute')
            if market['side'] == 'OVER' and required_ppm is not None and (required_ppm > float(self.config['team_it']['unrealistic_points_per_minute'])):
                blockers.append(_blocker('TEAM_IT_REQUIRED_LIVE_UNREALISTIC', 'Required scoring rate is mathematically unrealistic', {'required_points_per_minute': required_ppm}))
        context_probability = p_raw
        if q4.get('applicable'):
            if q4.get('mandatory_missing'):
                blockers.append(_blocker('Q4_MISSING_MANDATORY_CONTEXT', 'Q4 exact time/score context is incomplete', {'missing': q4['mandatory_missing']}))
            if market['side'] == 'UNDER':
                if q4.get('ot_tail', {}).get('score', 0.0) >= 0.70:
                    caps.append(_cap('Q4_MATCH_UNDER_OT_TAIL', self.config['caps'].get('ot_tail_under', 0.68), 'Close late game creates overtime/foul-tail risk', q4.get('ot_tail')))
                gate = q4.get('under_gate_h')
                if gate is not None:
                    gate_adjusted = gate + 0.03 if q4['dry_score'] >= 0.7 and live['line_edge_under'] > 0 else gate
                    context_probability = min(p_raw, gate_adjusted)
                if q4['dry_score'] < 0.55:
                    blockers.append(_blocker('Q4_UNDER_NO_DRY', 'Q4 Under requires DryScore at least 0.55', {'dry_score': q4['dry_score']}))
                elif q4['dry_score'] < 0.7 and live['line_edge_under'] < strong_edge:
                    blockers.append(_blocker('Q4_UNDER_MEDIUM_DRY_NO_STRONG_EDGE', 'DryScore 0.55-0.69 requires a strong projection edge below the line', {'dry_score': q4['dry_score'], 'line_edge_under': live['line_edge_under'], 'required_edge': strong_edge}))
                if q4['foul_tail_score'] >= 0.7 or q4['kill_chase_score'] >= 0.65:
                    caps.append(_cap('Q4_UNDER_DANGER', self.config['caps']['q4_danger'], 'Foul-tail or kill/chase risk blocks clean Under', {'foul_tail': q4['foul_tail_score'], 'kill_chase': q4['kill_chase_score']}))
            else:
                if q4.get('foul_conversion', {}).get('level') == 'LOW':
                    caps.append(_cap('Q4_LOW_FOUL_CONVERSION', self.config['caps'].get('q4_low_foul_conversion', 0.74), 'Many fouls are not converting into free throws; no automatic Over boost', q4.get('foul_conversion')))
                gate = q4.get('over_gate_h')
                if gate is not None:
                    context_probability = min(p_raw + q4['over_boost'], gate + q4['line_edge_bonus'])
                if live['p_live'] < 0.6 or live['projection_used'] < market['line'] or projection_stat['indicators'].get('to_high') or (not projection_stat['indicators'].get('efg_not_low')):
                    blockers.append(_blocker('Q4_OVER_CONFIRMATION_FAILED', 'Q4 Over needs P_live >=60%, projection above line, TO not high and eFG not low'))
        active_cap = min((item['cap'] for item in caps), default=1.0)
        p_final = max(0.0, min(1.0, context_probability, active_cap))
        alignment = (
            (live_reversal_active or (history_zone_rate is not None and history_zone_rate >= history_zone_min))
            and scenario['p_scenario'] >= float(self.config.get('signal_gates', {}).get('scenario_direction_min', 0.50))
            and live['line_edge'] >= float(self.config.get('signal_gates', {}).get('live_edge_min_points', 3.0))
        )
        sample_sufficient = canonical['data_gate']['pooled_n'] >= 20
        strong_clean = not blockers and (not caps) and alignment and (stat['stat_gate_status'] == 'CONFIRMED') and sample_sufficient
        verdict = _verdict(
            p_final,
            blockers,
            strong_clean,
            history.get('p_hist'),
        )
        p_trace = [
            _trace_step('P_HIST', True, 'weighted exact + form + H2H + distribution (or Team IT formula)', history.get('components', {'team_it': history.get('component_weights')}), None, history['p_hist']),
            _trace_step('P_SCENARIO', scenario.get('scenario_support') == 'ON', 'independent matched-pattern groups with sample shrinkage', {'effective_sample': scenario.get('effective_sample'), 'patterns_used': [item.get('pattern_id') for item in scenario.get('patterns_used', [])], 'outcome_center': scenario.get('outcome_center')}, None, scenario['p_scenario']),
            _trace_step('P_LIVE', canonical['stage'] != 'PRE_MATCH', 'Phi(line edge / sigma) from conservative multi-component projection', {'projection_used': live.get('projection_used'), 'line': market['line'], 'line_edge': live.get('line_edge'), 'sigma': live.get('sigma'), 'scenario_projection_method': live.get('scenario_projection_method')}, None, live['p_live']),
            _trace_step('ALIGNED_SIGNAL_GATES', True, '75% exact history zone + 3 point live edge + non-opposing scenario; exceptional reversal requires 80% live weight conditions', {'history_zone_rate': history_zone_rate, 'history_zone_min': history_zone_min, 'line_edge': live.get('line_edge'), 'live_edge_min': float(self.config.get('signal_gates', {}).get('live_edge_min_points', 3.0)), 'p_scenario': scenario.get('p_scenario'), 'live_reversal_active': live_reversal_active}, None, p_raw, [item['rule_id'] for item in blockers if item['rule_id'] in {'HISTORY_ZONE_BELOW_75','LIVE_EDGE_BELOW_3','SCENARIO_DIRECTION_CONFLICT'}]),
            _trace_step('STAGE_WEIGHTS', True, 'w_hist*P_hist + w_scenario*P_scenario + w_live*P_live; LIVE_DOMINANCE may raise live weight to at most 0.80', {'stage': stage_key, 'base_weights': base_weights, 'effective_weights': normalized_weights, 'normalization_applied': normalization_applied, 'live_dominance': live_dominance}, None, p_raw, ['LIVE_DOMINANCE'] if live_dominance['active'] else []),
            _trace_step('PRODUCTION_ROUTER', router['status'] != 'ALLOW', 'router may allow, cap, prioritize or block the market', router, p_raw, p_raw, [router['reason']]),
            _trace_step('STAT_GATE', stat.get('stat_gate_status') != 'OFF', 'team-relative 3-of-5 confirmation gate', {'support': stat.get('stat_support'), 'status': stat.get('stat_gate_status'), 'over_score': stat.get('over_gate_score'), 'under_score': stat.get('under_gate_score')}, p_raw, p_raw, [f"STAT_{stat.get('stat_gate_status')}"]),
            _trace_step('FAKE_PROFILE', bool((market['side'] == 'OVER' and stat.get('fake_over')) or (market['side'] == 'UNDER' and stat.get('fake_under'))), 'fake over/under applies cap only to the evaluated side', {'fake_over': stat.get('fake_over'), 'fake_under': stat.get('fake_under'), 'evaluated_side': market['side']}, p_raw, p_raw, [item['rule_id'] for item in caps if item['rule_id'].startswith('FAKE_')]),
            _trace_step('LIVE_HISTORY_CONFLICT', any(item['rule_id'] == 'STRONG_HISTORY_LIVE_CONFLICT' for item in blockers), 'strong opposite live edge blocks the history side', {'p_hist': history['p_hist'], 'line_edge': live.get('line_edge'), 'required_edge': strong_edge}, p_raw, p_raw, [item['rule_id'] for item in blockers if 'CONFLICT' in item['rule_id']]),
            _trace_step('FORMAT_AND_SAMPLE_GATE', same_format_n < 20, 'exact-line hits use same regulation duration only; cross-format games are normalized baseline only', {'format': canonical.get('format'), 'same_format_pooled_n': same_format_n, 'cross_format_team_a_n': canonical['data_gate'].get('cross_format_team_a_n'), 'cross_format_team_b_n': canonical['data_gate'].get('cross_format_team_b_n')}, p_raw, p_raw, [item['rule_id'] for item in caps if 'SAMPLE' in item['rule_id'] or 'FORMAT' in item['rule_id']]),
            _trace_step('TEAM_IT_GATE', market['market_type'].startswith('TEAM_IT') or market['market_type'] == 'CURRENT_QUARTER_TEAM_IT', '0.50 own + 0.35 opponent allowed + 0.15 H2H; weakest gate controls cap/block', {'weakest_gate': history.get('weakest_gate'), 'required_points_per_minute': history.get('required_points_per_minute')}, p_raw, p_raw, [item['rule_id'] for item in caps + blockers if item['rule_id'].startswith('TEAM_IT_')]),
            _trace_step('Q4_CONTEXT', bool(q4.get('applicable')), 'harmonic context gate after P_raw', {'status': q4.get('status'), 'foul_tail': q4.get('foul_tail_score'), 'dry': q4.get('dry_score'), 'kill_chase': q4.get('kill_chase_score'), 'volume': q4.get('volume_score'), 'foul_conversion': q4.get('foul_conversion'), 'blowout': q4.get('blowout_context'), 'ot_tail': q4.get('ot_tail'), 'context_gate': q4.get('context_gate')}, p_raw, context_probability, [item['rule_id'] for item in caps + blockers if item['rule_id'].startswith('Q4_')]),
            _trace_step('ACTIVE_CAPS', bool(caps), 'P_capped = min(P_context, all active caps)', {'active_cap': active_cap, 'caps': caps}, context_probability, p_final, [item['rule_id'] for item in caps]),
            _trace_step('HARD_BLOCKERS', bool(blockers), 'any hard blocker forces PASS without inventing a replacement market', {'blockers': blockers}, p_final, p_final, [item['rule_id'] for item in blockers]),
            _trace_step('P_FINAL_RULE', True, 'clamp(P_context, active caps); blockers control verdict', {'strong_clean': strong_clean}, p_final, p_final, [verdict]),
        ]
        return {
            **market, 'history': history, 'scenario': scenario, 'live': live,
            'stat_comparison': stat, 'q4_context': q4,
            'zone_credibility': zone_credibility, 'live_dominance': live_dominance,
            'live_reversal': {
                'active': live_reversal_active,
                'history_zone_rate': history_zone_rate,
                'required_history_zone': history_zone_min,
                'required_live_edge': strong_edge_for_reversal,
            },
            'weights': {
                'original': original_weights, 'base_normalized': base_weights,
                'normalized': normalized_weights,
                'normalization_applied': normalization_applied,
                'live_dominance_applied': live_dominance['active'],
            },
            'p_raw': p_raw, 'router': router, 'caps': caps, 'blockers': blockers,
            'hard_conflict': bool(blockers), 'p_final': p_final, 'verdict': verdict,
            'p_trace': p_trace,
            'strong_requirements': {
                'aligned': alignment,
                'history_zone_pass': live_reversal_active or (history_zone_rate is not None and history_zone_rate >= history_zone_min),
                'live_edge_pass': live['line_edge'] >= float(self.config.get('signal_gates', {}).get('live_edge_min_points', 3.0)),
                'scenario_direction_pass': scenario['p_scenario'] >= float(self.config.get('signal_gates', {}).get('scenario_direction_min', 0.50)),
                'stat_confirmation': stat['stat_gate_status'] == 'CONFIRMED',
                'sample_sufficient': sample_sufficient, 'clean': strong_clean,
            },
        }

    def calculate(self, source: dict[str, Any], dispatch_threshold: Optional[float]=None, strict_schema: bool=False) -> dict[str, Any]:
        if not _valid_zone_table(self.zones_data):
            self.zones_data, self.zones_metadata = resolve_team_relative_zones(source)
        canonical = adapt_match(source, self.config, strict_schema)
        canonical['data_gate']['team_relative_zones'] = deepcopy(self.zones_metadata)
        canonical['coursework_forecast'] = build_coursework_remaining_forecast(canonical)
        parsed_market_sides, audit = parse_markets(source, canonical, self.config)
        canonical['data_gate']['lines_found'] = sum((row.get('line') is not None for row in audit))
        markets, dedupe_summary = _dedupe_markets(parsed_market_sides, float(self.config['odds_min']))
        evaluations = [self.evaluate_market(market, canonical) for market in markets]
        distinct_line_keys = {
            (item.get('market_type'), item.get('team'), item.get('segment'), to_number(item.get('line')))
            for item in markets
        }
        evaluated_keys = {
            (item.get('market_type'), item.get('team'), item.get('segment'), item.get('side'), to_number(item.get('line')))
            for item in evaluations
        }
        expected_keys = {
            (item.get('market_type'), item.get('team'), item.get('segment'), item.get('side'), to_number(item.get('line')))
            for item in markets
        }
        line_coverage = {
            'offer_sides_parsed': len(parsed_market_sides),
            'unique_line_sides_expected': len(expected_keys),
            'unique_line_sides_evaluated': len(evaluated_keys),
            'distinct_line_variants_evaluated': len(distinct_line_keys),
            'all_unique_line_sides_evaluated': evaluated_keys == expected_keys,
            'missing_line_side_keys': [list(key) for key in sorted(expected_keys - evaluated_keys, key=str)],
        }
        threshold = float(dispatch_threshold if dispatch_threshold is not None else self.config.get('dispatch_threshold', 0.7))
        candidates = [
            evaluation for evaluation in evaluations
            if evaluation['p_final'] >= threshold
            and (not evaluation['blockers'])
            and (evaluation.get('odds') is not None)
            and (evaluation['odds'] >= float(self.config['odds_min']))
        ]
        candidates.sort(key=lambda item: (item['p_final'], item.get('odds') or 0), reverse=True)
        best = candidates[0] if candidates else None
        payload_candidates = []
        for item in candidates:
            payload_candidates.append({'market_id': item['market_id'], 'market_type': item['market_type'], 'team': item.get('team'), 'segment': item['segment'], 'side': item['side'], 'line': item['line'], 'odds': item['odds'], 'team_a_hits_n': [item['history'].get('team_a', {}).get('wins'), item['history'].get('team_a', {}).get('n')], 'team_b_hits_n': [item['history'].get('team_b', {}).get('wins'), item['history'].get('team_b', {}).get('n')], 'h2h_hits_n': [item['history'].get('h2h', {}).get('wins'), item['history'].get('h2h', {}).get('n')], 'p_hist': item['history']['p_hist'], 'patterns': item['scenario']['patterns_used'], 'p_scenario': item['scenario']['p_scenario'], 'projection_used': item['live']['projection_used'], 'p_live': item['live']['p_live'], 'stat_zones': item['stat_comparison'].get('zones'), 'fake_over': item['stat_comparison'].get('fake_over'), 'fake_under': item['stat_comparison'].get('fake_under'), 'p_raw': item['p_raw'], 'caps': item['caps'], 'blockers': item['blockers'], 'p_final': item['p_final'], 'verdict': item['verdict']})
        unhashed = deepcopy(source)
        unhashed.pop('super_basket_calculation', None)
        unhashed.pop('super_basket_system', None)
        snapshot_hash = hashlib.sha256(json.dumps(unhashed, ensure_ascii=False, sort_keys=True, separators=(',', ':')).encode('utf-8')).hexdigest()
        calculation = {'engine_version': str(self.config.get('engine_version', '5.0')), 'calculated_at': source.get('meta', {}).get('generated_at') or source.get('generated_at'), 'input_snapshot_hash': snapshot_hash, 'canonical_snapshot': {'match_id': canonical['match_id'], 'name': canonical['name'], 'home_team': canonical['home_team'], 'away_team': canonical['away_team'], 'tournament': canonical['tournament'], 'stage': canonical['stage'], 'explicit_stage': canonical['explicit_stage'], 'trigger_checkpoint': canonical.get('trigger_checkpoint'), 'current_quarter': canonical['current_quarter'], 'quarter_minutes': canonical['quarter_minutes'], 'clock': canonical['clock'], 'score': canonical['score'], 'quarters': canonical['quarters'], 'elapsed_game_seconds': canonical['elapsed_game_seconds'], 'remaining_game_seconds': canonical['remaining_game_seconds'], 'stat_support': canonical['stat_support'], 'format': canonical.get('format')}, 'coursework_forecast': canonical.get('coursework_forecast'), 'data_gate': canonical['data_gate'], 'market_audit': dedupe_summary, 'line_coverage': line_coverage, 'markets_detected': audit, 'market_evaluations': evaluations, 'candidates': candidates, 'best_candidate': best, 'gpt_dispatch': {'threshold': threshold, 'eligible': bool(payload_candidates), 'candidate_count': len(payload_candidates), 'payload': {'match_id': canonical['match_id'], 'stage': canonical['stage'], 'candidates': payload_candidates}}}
        output = deepcopy(source)
        output['super_basket_calculation'] = calculation
        return output

def load_json(path: str | Path) -> dict[str, Any]:
    with open(path, 'r', encoding='utf-8') as handle:
        return json.load(handle)

def save_json(path: str | Path, data: dict[str, Any]) -> None:
    target = Path(path)
    target.parent.mkdir(parents=True, exist_ok=True)
    temporary = target.with_name(f'.{target.name}.{os.getpid()}.tmp')
    with open(temporary, 'w', encoding='utf-8') as handle:
        json.dump(data, handle, ensure_ascii=False, indent=2, sort_keys=False)
        handle.write('\n')
        handle.flush()
        os.fsync(handle.fileno())
    os.replace(temporary, target)


DEFAULT_ZONES_FILENAME = '02_team_relative_stat_zones_COMPACT.json'
EMBEDDED_ZONES_SHA256 = '57b66232f3bb2f2eeb8612ef7717a62e9328fa6553e28492fdce5adc6345bdf2'
EMBEDDED_ZONES_GZIP_B64 = '''H4sIABo2YWoC/+y9e7Nlx1Un+FVu6B9molXbmbny6R4mQqZtQxhsY4qenunouHElXUlll6o0t0oCw3SEZBvwHwSYh0WEgQGMTfDX
jPwQtiVL+grnfqPJPPucXGufvfYjM/e5MtMOQvics2+tzJ2PtX7r/YfPPHnhletXry7fuL558uDxo2c++czT66tXL2+uH149ffDG
9eWTp1dPL//g8aPrJ5cvPH71tasXnl6+ITvxzLPPvHz96Prm6un1i5dXT+M/U0LZe8Lek+K+NJ/U5pOQ/ujV66evPH7x8snrr756
dfPVZz75h888fvji5SsPnjx9fPPVA90HLz96fHP94jOffHrz+vWzz7zw+s3N9aOnl69ePX3hlcvr33/h4esvxkFeunn86uVr1zcv
xGcPHl4/Of75w6vnrx/Gb8/85hf+t//lV19T5tmL3/qN/5Q+3HvNxS+//huf/fVfjZ/uvRbEsxf/+dNf+t8v00//66/G73GCrzx4
evni9UsPHj142r/+F+JKfOJ3H714fXMRn128/uT6ycVV/O3q5euLtDT3jktzESfzIL7b08dPrx5e4MwuXnp8c/H846ev7P/8ybMX
jx4/vbi5+r2Llx8+fj7+5f7vn3TP/Pdnn3ny+PX4r+Lk/+sfHj5fvhQpxFmI3/iDV+9/+nMP4gzTL5ePrl5NP3/q6iuvxP168sr1
V155cHP5u48evHH5xpPLT8VfvvqVp1dPXrk8/sP/Sf7P3ZefxDeKe5AWMv7j3V/s3rn9o9s/2b17+8343x/dvn375sXun3cf3L69
+2F38caTi/gX6dlPd+/dfi3+7TfjP34ap5UGf5T2ePdPt1/ffRT/5s3d+/Ef/Wj3zsW9i90/7N7fvbv72b3dh7ff2H9P9G6/Ef//
B7t34rOfXvxfF/re7mcXu59Hol+7/ZPu4lNXr8T3vo7vcPG5xw+vv/zgYk/wx7uf3r55+/Xbv764/frF7VtpoDjJv7lQ92QX5xLP
4stpFX79fvryQjw08QvIe1qliaZjG8/x1cMHL16+HGcclxXM4ffnx7+/9jgO/OLwdyfiaj148uTBo5fjRlw/fPHJZTquLz28ejmO
FAdZOLz//dnTffwN+XtffvnLX/nSyT5+5vXnHz+6/NTN9Zff+IO0gfevHn/19atHl1+8ev3BdfqfLz98/PTJ5fFfj3bye3EffrD7
cPdB3LK4GXH1dz/c/WS/hd+Na/7h7Z/F///B7Z+mzXknLWj+Ev/y/fj8a7dvjTb3H7qL3d/22/ofuH393nFH437CcD93/0y3Sw63
6/Px/P/O06ubyCrIvol74k527TOffe5Z9cXnnoX432fuP/fsF770qWfvf+HZ1x4/efLs9Usv12zqK1/+9PWjL/wfNyeb+ltXX/nK
1fMPLu9fP7z33Bv9zfzcK1evPb5+GP/38cO448d/yd3Ov4tL+178v3d2P7h9O+3ju/FO/Om93bfiZYoXar+5349PP4r7vn+Uvn6Y
dnP3wXYXVQ42luzib8vLh5HrkR2091zpDmp+B23Y+t698txn//Mrv/9f5MkW3f/8/cv7N49fu4oDffbBy1d75nl1E79efTn+8Oj6
+Zury9+5enTvtx68HHftSGW0V9/d/f3uu3GP4vp+GNc53rXdjy52f3VY6p6TvhMf7h+kS9c//CAu/w/Szxe7f0rP7u3+bv9P9tt5
uodf/NRzF7/2+NX9usRXvvmVJxe/9vprSxfzhNFWzKOIFf//7m7/zkP72/K1z79wenCuXrx6NV7gz71yc/XiH7wSz0Y8Ob/zSsQe
z8cDFEk9ujz+S+awvBOX8524KXuG/f39fvx492+H5U6n5V/TrU+b0m/Y19LSM7f6r6NoTszgJ/G/REv2pyF+jA8SvQ8in//ztGOR
j7wfCb0VOUn8k49mDow6OTD8VOYPhbwnznEo1MShsHd9KJYg8clhWfrz0xNyus9kKcn6TSzeo9cfPuSXr3/CLWD/pInj/rdnn3k9
os+oNVw9vLx5/eEBxaZPlw9eTDg6Qu3L3778nftf+sLnP3sp5P59b75ynV7wt19Ph+PmE7+VFuDifgLE6UUfvHh9AODx24sPbq5f
IJg8/vTC40cvHmF6nNJF+gcXV0+fXr3wlYuXoyZy8YXPf/Iibv8nvhj3eg/8L/7DRTwCn/jM/Zvj1/tf2OPxw7frz3x2/zXqDvHb
49dei2QfPb2ICsH1oyeR9sOHj3/vQOmNxw9ff/X6E9cvvfTghYjRXvjqf7x4cp1m1E8jUXnx+mp/4J9G/eXlp69cxt/SC71wFaf9
YpqfE5/w4hNBXDx+9PCrFw9eunjhJs70Xtq8vbLx5PU4g5unT55JRxCX8nc//58q13Kvx5ws5vE3upp7dSX9mycXr+9Vn/Fy9kt0
XM3+W1q+9CkqPHFdD2sa7xtd4tGippe8uX6S9KrRmkbaz51jBT/z3Oc+fbk/kZ/6zS/82ufo+v3a1WsXn7j41MPHL3xlvyzkFF68
cPXayeLt/+7i8fg87m/pJ9KC7N/7+defDpcuvgFdPOYlb65fTssSyV3cPHjylf+YRr/YT+P6jetH6V2TErkf5+LBk/i2L7/CvWV/
WNa+Zq/oTr3n68xRwRdNrzV4z/2bpxdNRyCeCXr1mBd+/vHrj164Ji/bT4Z923gPn0k8J7Giy6SzR9558+D3l3lOpPBaD19LGdC/
E27D8cmrJ3vs1IuFJ5evXcXbFtdHyr29Iz7SveHjtSi6PmlcJ3W/spGEEee4f3Mbo/6H2xg/2hg4bgyQjQHXmbwvz1/HOVze/e7A
/3C7o093xx03R5LNUaYTH/vm6F9uTt4cRVmaIJtz59vyW8/d/7Vf/yVTCx8rT+MwcwUW+HcAllls3++JON0TcdwTQfYkqXuHHbnz
nVC/3IlfkJ2AX+7EL8hO6F/uxC/GTpQJ8l9uxhnMNLzIXrbZtJlmpi0+/WIlRfpEgzPH5fJkuZzqHOQVc+J8Vh9+0dQv+KJltTeQ
RfOy8yovmr/zRYNfpEUL02tmyZpZ21k3r/Gcdc30L9RBGymJMnAHzepOhY9v0UbS5WNfNxitm86mD2r78Lqzd3NDh8bsImFALdvb
2K9n7OP9AipzuoAqiwXp6MnzS/aJYtv46rVTv6BrN7IfKMevneic/LgWD35BF8/OLB5leQ464ecByfkWT/+7WbzAnzzTgf+YTt5a
YfExrJ900+KWLp+WnfRrbH81fsF9/MN+sS5feNxH6+59g+n3FJHym5/+Lxe7f9x9ePv13Y8vdn8ZP7yZYlfS/+5+wqzzzePf663N
/+frVw8fPP1qWsJXrx4+vHxy9eprD6/3G1dKey+2zkJYnYswnIuwLiO8++N9pN0HKTDyOylu+Vdu//z2a/d230pxym3bt5Z06e6t
pqvORBfORLdh6/5+9+7tW/d2/9KHr83snNqQMtm4Tcmq85CF85DVxWTfihufAog/uNj97Z7ej9KHFECehnuree/W06/YwQLi6pzE
4ZzES/f0W7v3dz/cxwO/N7l7soES2acmMmobMrANGV1OJoXt/11PKe7cTyI3/fPJBTfbECVrvxVFtTlF2JyirqP4j4nO7se3bzdi
hxlyxXhhjpbakBZsSEvX0fr+7t24nT9PSWj7nLPdz/Zh5V/b5pKsoF57W9aQVucjDecjXX6RPthnMXwYh/j6xe7bt99I4KN5AxeI
VuzbEkW1OUXYnGLF5kQUsUeF+yySxj1hadVsBU9IbUUItiJUtd59mlUk9d62PG0V6brdWEFXnYkunIlu8dbtU5jeasXDIyrlWHhM
QrWTgHYSpfj3L+Ie7WH0LOxto1W8upOE1FaEYCtCFetdnmnKbwpYOuLN9cMHV89vORrZtrMPpe5uKLi7oXTpUG9F5LeXcFvcxQlq
NbdxipTajhRsR6rmTk5W2pi+fmbVlq4hTG/ahlTVWajCWajqMqoZTVzsvpNyp3c/aERuCySLEdsSPbUxPdiYni6nd1KLpuHWTJEq
vSeTdNRGdGAjOoWn//29ieCHqTJMsj4neP3h7VupfkASVNNrv3InV5On+7E5bXVG2nBG2rqYdsITCVrsKzT8cardMCP+7TZEydZt
RVFtThE2p6hLKSbfeGSSfxIpp/3+Wissm6FXDszmiKkticGWxHQ5sds3W5naCY1yznVKQLUSgFYCuoZAxMwJN3+Ugjy2WNMJenXr
O0VMbUkMtiRWvgd/dvtnEQH8ydHI35cWanTPL1MtdsqvIKm2Jwnbkyx1u//lvrjau8l7v4EaPk2tmNvPkFLbkYLtSJXy+b9M2uHt
1+J1e2sPr/piebd/2r4JawiX78cqquosVOEsVEs37K/2e//OXp1562Ao+9nkXvlFxWaJHtmiDYipLYnBlsR0KbGohLx/sfv2/sOH
Mz5EWL+t0ySLRfsSPbUxPdiYnq6g9/PdDxJsmFHGZROtYm41SUhtRQi2IlTKhv46ip8fJmP8/7N3zf8saei7Hzcu/CLR4h1Ypqg2
pwibUyzenH3t4zm5UE+nfAs4ImoLIrAFkdK13XOr2z9ON+jitCBxii7+Wl9WegY/qbONU6xwVA2i7mIQuItBStWUb6eCrvGSvt9b
I7+9jyt/b1Gz35h4+S6vpqzORhnORrl0E9/eI48/3f1oD8/fTuD8O6kCfGK+zVEuq4kXs9H1lNXZKMPZKJdy4e/sw0K+ufvp7Z+n
e94IvqepFePuGVJqO1KwHSldTOr9uIH/di9y1b7Q/tzyyy1IFl+VJXpqY3qwMb3yy5CSYH64z4n58e7fLnbf2yfD/GD/w7f3XDN5
Zt5qjOSvGKY4wr9mDHUHY8AdjFGaKfCdZMjYx3n2+W0pLqc56WyRaDH6WKaoNqcIm1MsRRl/u3t3n5UzYxhqIFS8CTwVtQkV2IRK
8QL3NqN7ycwaUeIeJ+61uw8apdEqwsUyaR1VdRaqcBaqpVIq7fiPkmu6DxDYt+a5/VoEJcdWSy0IbjXxYkC3nrI6G2U4G2VdQTnp
1R9eHMNz+o5obzYLnnWUixnfSrLqPGThPGR1FdlsDyH58P+4twe/k6oY/NUs74TzDVR1JctHUXcyCtzJKLptlET6nb2x+YNm3Xkl
6dZNnqSrzkQXzkS3desiQ/9mL53fTzjqXzbxPhYP0rydyyOos48AZx+hcbNPGxxO73JY9FmvJ033dlO66kx04Ux0dSndD1Kj0aRK
NqogU7SKtY5JQmorQrAVIV1F6O3Z1VYNlKqQJkNGbUMGtiFTjhzf3YfKvJ8sIocopll1TcKKm7JIkx70bQiqrQnC1gR1OcH3IvP6
Sfz0jb2rbK5WS8luz1GtuBWLJNX2JGF7klU3p09ViRvebbEzLLW6HeFJqe1IwXakyld+X9otXbVmOcxRqpDCLBm1DRnYhkyx/N3b
BtvNrRyd8hXmiKgtiMAWRErX9u93/7T71uS6uioaZE0rCahWAtBKQBcTeHf3UYrA3iS2Y5pa8XmdIaW2IwXbkSo/wR/ufrj7aUoL
/O7R85m+fKdP3Wy2FpfQL5aQRcTVOYnDOYmXytX/e/feoSn6D5s13ClaxTdpkpDaihBsRUgXE0px8T+PV/LtXlw3c7B5ihVrP0tO
bUsOtiVXvhs/36sFKZR0/2nLEmariReXxFhPWZ2NMpyNcmnBjH/o7fFRxu1v6O3fbHKtlqkWX60VJNX2JGF7krqGZG9I3xfF32Rv
JujV7coUMbUlMdiSWMUe3L4ZUcS/ta/9iE7Nmo+JqC2IwBZEqtY28rWf7n6ygfI8Ta1unXlSajtSsB2p8pVP8iRhs5+nEnJbFGNY
IFnsJF2ipzamBxvTK3V3/mOS6sl/+uG93b/u9ZUP9umgSftMZWcv0ojp177+z4wwWPbPtYxFNvLMA6m7GgjuaiBdMtDt232CdqN1
gKNTbAVgiagtiMAWREq193/afWf3rRmroq6kUhziz5BQ7SSgnYQuJnFahjTFj78dMcJeaZkW7XrxIqwnTeX8pnTVmejCmejqMrrv
3n491R74+z7hujn+eJZgMQ6Yp6Y2pQabUitFAN+NZL57sfvuPuzxo77mb1/v90dzkYnLhX3XEi6rt7yaqjoLVTgLVV1GNfVhSJGp
HyYU8P19WMO+/tN82WyzHeWygqTryarzkIXzkNWlZH/WR6MmFvrdeIv/KhVs/DAVJ4rY4vt9cvW0yPJnGIKKrrPQV2emD2emr4vp
70uHfJCKPcfT0jeDuNj9dTw+79z+dYLnfVzyz/ZRLm82C73GIYvFYut46o7HgzseT1eM9+Htnx2PZLK1fH0/SP8lDvL+UnKWW3Uu
SwehJ+NMI6izjwBnH0GXjZDOSzxNt984AOh0ruIR+5vdR818YBXpivu+jq46E104E93ye/rh7TdmC7zrajrF6jtLRG1BBLYgUqrE
f3ePsH4Q2em+d+iCWXgbmuVrvkRQbU0QtiZYty8f9W6tRi/JFK1iH0ki9G7fEa3RQslTKrZRTpBR25CBbcjoYjJ/loKxD7k5+yLD
H+5+3iqjloiWS6dFimpzirA5xVJZ9M+3f7yPH45k+tTyv9hv+nvNNvtVhIsvyDqq6ixU4SxUS2/T9w6hNPsKDn1boMZgpXmKxRFK
C+TUtuRgW3K6hlzqzphqoX3vVyLiT/fy5wdP2hZVmspGKIYiheTVecnDecmXgpbv7QuM7i13/d392e6Hu5+0GEUXKZYZQ5fJqW3J
wbbkioye309687296HuvEUFOkCoGkFN01EZ0YCM6uoJOKmR3cNF9Yx+p+dY+4fGnM9BNb0u8mJmtp6zORhnORlm3UU4ht5EjppCP
d2cj/s3G1IsRRAFpdT7ScD7Suo30d48flspI6I2pt17IOdLqfKThfKQb72T8em/37X3ww3vb7uIE5dYdnCKrzkMWzkO2edcOvV+m
kaBfhWYWaFIsuA3BIi/JGoKwNUFdRjClx+1NVH0P1r84uNvf73Oim1HiavIVyHE9bXVG2nBG2sWoM+oLH80ozLaSSnGbToaEaicB
7SRKO2/+y76Z27tJbqWI0wuaRNfcEWw18eLLsZ6yOhtlOBvl0mvxL1F33l+5RPVv9h9/kq5c6/Ytky3fuBU01Rlowhlolm7Tv+4d
Zu/ufrDniSnSIyXeT1vRlw0eixSp+XwLcmpbcrAtOV1KLt69PWKfC2R0q/eWp1ZchmOGlNqOFGxHSheTSi1Cerfxexe7v0zxqqmL
dEqgb2RZK0kXs621dNWZ6MKZ6JaysP83VcpPVRS/lbyM05u1fP0mKdG9aSGjtiED25DRi2T+27N7Mpfxl6unD964vvyDx4+uL5++
cnP95JXHD1+M//K/0hbsB9fU7Tfjf3+Ugo9SAGHydPxwVCvs1eunNw9eSNv6wuOb6xfjL496c/qr11fxkxKdevaZ15SJc3OdiB+N
2P+aPrr4q4L9x5B+9Z0atoKvmUc8XI9/j50IaBWYuUh2Lrqz7XN57frmweMXL58+fnr18HRCWnQGJwQ2T0jj4mjdmcOETJxm+4Q+
89nnTuchTQcuJbn0CwP7EdM84oPjPA7LleYh/Rbz+OLjJ09GE3GdhnxYbKfUcSKu8/44E9/t/2Y/lRD/RrdPRn1xtCih08rbw1x8
3pqAKyLyzsQl22BFYDwJ0wU8IDpPArfF510JW5zWz9wfTUHHFc7roPCI5inYPIVNzsVn7t+cTiEyEO+Px1N0sl+SNIv4BMTxYMTr
7d1xS0RnpdjgZHzhS58a8ZHuuBwSWQhe2LwcZovl+NKnR+P7TkoNhykgF/N5CgEPpthiDve/MF4CPzoRwCzBJiz0pcevP3wyvhnK
ODlzM/BYui3W4Poznx0fS2MAj6U2AY+lMR6PpXXkWEZGCxscy5de+tLTl0f8MzJs7/1xb0LoQBh35KGRWQrt7JF9gekcyONBMbYz
TmwnfC85OROPA26ahLxrVM7gvslNNu44H34DPQTcQA+AG9h/OW6gsriBXtsNl4njuEQaqyyMOY5ruw2nwrI65fcHaIrbQWY1eovN
eunxzQtxJizDwZOzxHM2PDUTKEUaglKSa6Q/xpagFNcFj4DJ+w33aeLux+vt814JkaAUkLsPIF2++xFnK5lRlI7bCLJOWqpVGoAg
997lIz2hBLhKJUCtUwIMrwSQueiN57KgBBzhBODKUA2A3LFQebbVkgbgKMo7sOMD8BaIu4/zUGKLebB3K76hQVXEdXt5ZfpLZPVx
KqG/aKFX3mCDLWI0gAiesiTFBUFdUeKRjStm2+fAKACWaiG4LW5DBUAtKgCE9QKjADgyhS2OJ68ASBTU0gBRAKx1WVBrLVFQm/a5
MCIRIk7IIpGTRKZZOqsFBSB0AQEC8gwp2KMJW8xiLJGhkxSlnEMLUCu0ADmvH7drAWpRC9DB2Hw2owpqEUQqS0Ck9BJBpLBug+PJ
IwERr2zI98VHnO9DRgIifg0iawGRfeqQ0a2EyEWc0JuJX1YNiJtChI1m1QDka0pusXPzaoAGiWqAdgp3ELcPhCZKXPAbrhHHdInh
B3UAw1pd7HZTYXUATcAbqwO0WjzUKh0gjjPPefW2E5nVASKkB2LW3kMTBqY4i6fYqQ3PzNTVD5124siWfZylM5pc/QB7m2p/9UNk
4cqQq2+kq5sirFICXIn1vVZ4wgolIClnaEKTGkUoAiuF5l1Ve5xgrRIQ9TFiGQG0wAPam7VCgbrJ4vCeACJQKWu2rIXGb7EwEzq2
FQpQydbEFQAKpYTVRCcJG6wKowjEUZDxWMbgSl0BqlJhhGVNAE8IagJ2Q00AFjWBBejrmoUSLGoCIBV1BXhNXAEG0ZYG6grwAO3T
YWWjJ0jmLCIJFpUBPXIFUE1A4m1VW2wKJ5cBmem8RrTJIkxoAmK9P2CDZeBxpJMER1rnEEdahVDSCUc0gXilNjibEzZB34HO7DzK
fwlA/QEOZDZYRqljhcx8TIfOR0V7Mxk84RBwRA5POQQMOgS2wwS8JqCpQ6BnIf0GRqVA5Q002hN/AGy5SKw7YP564QJtIo5nVQEy
lfOoArBSFZDza7Ix311wB2ipLEIV1AVc5zVCFe0xaiGoDTdqQhWACHJtVgVUJ60NWUDEc2uMI1YApYIl/gAQSlZNUa9QBfodU72K
lE+RogZ40xoRpNfoAeEE7mY9ILDOgFpUo1c7AySdEGjWH4D33dTGneglPSB5stDxx/sDAi6M2GIeU/4AZSTnD0juLqKrKVS0a0OC
9GJIEIkfY0OCBmBrgyVh9ABH/VeW8QjgtshNtoURSaZznIfasi4Bu8UUOEWgh1RHl4BBlwBqAUQFUJUwSy/6AwqsUhusBasCkO3w
E1pAc6yaXgoJsmtDBcwWM2C1gCIleYNJ8FoABEcwpOeMyYYYk72CDUTtlAaQluQITkIEIcHncIWEVWDPS3oFIAqWgCDFyE4rtZ3w
nVAAzAoFoDkiSK9XAAJR4cBYNC9o6mg0br9uBw3AOLvhMt1nbD/kggPDbwO1P224RKwOIDiTB6cD6G3u+aQOoLnAVN4Ru+GaTLkD
AjiEKZL4AzAmKDJqErZQ6Q7QJTqATIFsx3m5ZPC3x+sfohS3VmYJ4ToAlW+/jmJdQhWOOuZqLGgBHqjy7xDrenQIBOLHFpVoZjSd
CWXAeXrLHB5thwDPu3zR4tT8JtOZ0wekVdTmKg21yJNIFKeQAcTlU5tMjOHX1lDF20pkRciLLMpYJ+oQz2gq3J1zkLzogPt11Azi
A3u8cs50Wa11iY1uMx9GOQBPRRngRgH6CTTqB9psdJoZFSH5zo58EeehcIsUyguQG20RI7Yiv6PxvpaN+1PU5Sc2mgqnLgiiLii7
l/IHv0EwElMITNYko8gXfhsOyMjRJKhxcfgAM0ORzzbbxKgP4ChqBgRh4NiTqzeaylikS2IYEuyKoDIltzotrCIRjyUxxCixZKXa
ajITCoWmboVAwlMsItIQ8OBaC9vs0ZRWEcEXCdMUKqVCyKx3hi70bofeuBhPlBYh65/Jz+DDlqKcUy2soBkiBsPkiNy0eJys3RZd
TAQaKY2qoXaKOIiyHpZ2FTxu5UYgbEa9iMebhkN4NjRDUz6tNp0Swx0Dn+PEm9pgM04woWokzylRehAaSqCsyWzNmuZUjgj5gkNo
31u99vAnPji6qiKq1kdDg4tXwGy8cRNah+u8OTKHEIcNmIgg4uHpWUXPGxKvQq0jnjKnT2B1mt1Pd+/tGxZ+szT3eKitBkQghI0b
Gu1vawafzjUgp5jin0HIk5kMeVo3+qxDYSDHaIIByVg01KFQNYUJ0w7xrRDTjmZ9uyPf5bqRWW08klUenRcYAWd7P4EjvoSDQ24U
O7RueMUZ6NVMqNCcdX7dkKxPAOZ9An4mI3TlDnOOacHlZGligDFTwddrB+VSAYKdsPurXqAegHxv2zikAmhRtb2s/cuQXPkFO3fV
oAxEdzQceCICLZ8rWbfYnEnfrQ/xrxpzIpTHczkvlg0zqxqWxWbWaZXPlZGAKDvCMQRnzmWcXbnQE1ITbOdUyAwz7ajKpvpkXYCM
qJMZxnrIrDMVtpANgos1zcMA8yiWfxuKeVpEJw+XrZBAFB9jiOJDjPGDeCqvVGhaCs78Tlx8wBzI0Hz55nNw18bf67o9mMu6nfdt
zsT7lL33hBgPJHrNdJ5NtPXxZKAcd02ncOJqpsoxFoGckJ3uLY/7yST1qY/16/WA5PTTmUeYOD9l12LaNdm0cmASJtoaNGLaNfmz
KeSaCQvxEzH8ombohXRZehn5CBk3kzG7bgrLMTF2OSamat0n3EtgPMmdyMLpWAznEAUjiHvJVQ3Phr2QIgBuKbxA1b02i2rnEl+r
eO5yrqtaC2l97aAcpDUkYkAKA3xMO2himzY1o7PiJawFeabulfmIFS6Sam3c+rpxOXlmuHdd5bteN+ZEVAphWXoW0Ia6V52IRrd2
IqsRpCKJjYChDB5OTUQrD9ZkTYqooBk0BQWnspk4ucalDKQmRUSyOstNHaGcaBFdl3yCk5qPNiF1ziplx7wFWATUMsBaEl6iaa4w
KBpgbpqW4f58XTFg0kvbuesMogX08qBJbE283No7OBkxAmujxm3bS08IcamCHCdAxgdUhNNA1tB0/qfMsybleB/5UVQepRPHcxii
CqGkIeZZ8ALrxOh4RU/9EZNzgnVYlig3vOfBzASDrxt8phaMo8VGFmvB+JrRF+yzxEANnrXPYgBchLNVC8DbZ8Ni6N2AGdqakacK
KEHQJLnTI5Sm2dMoDSJgkFVvzqZzjg20YSKXs+qd2exNEus0D2fXS53lfE2gRQSKSjeuHZSLtFBExEWxjpEWAKSomraShGcHG2om
sFCe0WyC8GCxICNhYpyzskqawhKWDdUpUOvGnDLOchHWpsk4u5xZaaUgWJbYAXv+fbADKuI399pUMYwpE1DkTCpgeTUbsavKfqVU
JixkTC11XAS7R/wIZbVrkFwTsdO0ZteUg02UG2jXp0saR666NoIUwNQSDbSW1mUdh7sXLgUHZz2X/GvXJKaXDc4Fcg3qWswqlVB7
EaftszQW22zPAxYSH30YwCfPlWsOHanWHLxuWv7JJAeJKZghOSbRPJuqjjqPaqbqNGhUMyMYFmvVK72qzIkfWGf5WocGMaWoGXw6
oZEW7OQRLVpW1Hp3vy4w0JLCYBNFzTHo04i6BeB1+2E2uWGKl/gaRKtXGGi1dQTRGo/KnUblThPlzvua0Xk8S8uTmFnnsKpbbtY8
S/Iv3GzQQeVKs644RYsRzmbCiNpRGUyrjfEYPYyA1gJGYNogiJSL06wZnTeXrK19a+p2l8G0UlBjSWBNtGomqXDdwGyRr6WXNW0v
OwFrHS0kaDaJOViRKqiI4V8LT+pPK+epOZAUEAdRxTmmZGcEtr2JZ7/Rh3yJXnKmYBudY/VARsmJ9UH8PpmoQXhN1AZUTDq+5Txs
tQx8oQAIueiHFOODgdZoQ0J0FVFetWxbh4WQAz2byCrlapuQLjbQFiY0r72Bk3iWLTUITdp7SRIfkHgD1xEnqyZCXBEnq2ja+CkD
bZyINJn3qg6EzJdSRKlnBGQ3me+CylX8QXdRCK3UNNfm6VlqRnIoCryheXoodlefx9WZeYMyHU5NZOZlA9pqG0thLp4Z5uJRgEkY
lCM5iz6sNRmvyr4blOO3mJRiSTEg0hxgNX9cl2+XNJvsv7Dx3GmMOBcYcO4U5ttZXfv6XIbdoNQCCNR3CYPA14dQfQ65nLqQBZNC
nQYQDQFR8231AeSz6Ia12bkiNYPq+fWDs9ZcDMFTOqA8VD4E0nlHoOlNO2Mqp8DZVyR9/6lK7AYrsde+Ppcah5mUNC/O0ry4HGoE
1deNyTiRtAnIivLztS/NZ78BTdxUeNNoiBkyXBWqh5/oymIx382AJKi4b8RyQMXCOuJACFB75iYsvhA5XMj3LgplYwUG/cnUUAOL
56VCFkYTR2kU2bpNDHL42AQqgwyq3VawWW31zGDe7gs0tMSCYWqcGO8DKb+jVetqMIlrQxM4cR8jNFF+Jte5eBIcfxIZLbsJNZnc
U1V9T6dy0wbhoJjhNMEedfMSsNhApkoymIsmFUIDhan4SiM0MM1bMaHPxuWOeBkr3ijipxER1Au8tSk8SUCOHo6CPehh8fvv3r65
bwv8k0Oz0492H96+uSoJTZMYB8FwUpmvCk2BPzHPlQ0/gM+a5MA5w2TAAU7Ado6EWMj6GYzxc55GFI95GlFqZsVaa66y3WmgQ9k0
MuPUxDZsEEEpw5a1ViSZwdWPjhdEE9USK0S6zlHrMAlkDySQXbYcxIyeNWETBlPrHeESGOxDFkC1LACMRtf5CgDTz4ya9BpeOgsI
ZlhqxzCjWF7bNuzNybApE400BoxIAi3GSpLGgGADAc2iYcmzXNKkTlM2pLhxHATZ7Ka9zqBZE3eEyc6IPLLLyz7IgPT1Ix9Foeas
Vma025O5amWjErys0XicGRu+sMkjE8bSxFcyEtPUAS/QgGyUJ8XmLG1gKEjkeIRrDctOxa6mjScBshV5b8ESkMVuZCfBCEXEbnCO
SF0ZVclmsXfJcn0UfWRvJG7OsLy0bp8Fu0uBtK8dtJm0fcjKIX4FvCd1wWGDJRmxRJsvCVkP6uZQmLPr2ifA8SUzz5ccdpSULfd0
gJfJ8PlE0AQAkw3M+f1hi/fnsUCuyny0LI86x/i+cvQhM/4kaqJyK9iLq1I4hBtUcgGb70nUo5QyGaFoE4WZxbimZCYwZZuklvGy
QgGSrO35jCBgUHYyx6xsAjxiTtGmBnfImfkgCtkygznEjCIc0AKk8bCSwg26FMKoRbzsDJPnN2Cc9XhZLePlAIiXsXtZxMXoiJG0
lqKvn8AYLZMDEIi2gOYNmp7UtP9jrIwc2jB+aEtqKrTs+EgwAN46yTBGjXzZdrJl4DFalsGjz1X0md29eJTGI4hRgRTNBQuqfhZj
sSQ7Uh4OgQLqyHjfW7Z7DJddN4uWPd4y33LJ7gwtqyW0rPMLA76wJlqZQm2w4YU5HGaVIu2++7aHB7QcKA5TpNOfH+bgFJ4zXuaK
LmirSLii9Jp0ZZESq9Sk5A+dKz0k868ZZoNVCTweK6sJrMyXe9Dts2CxsqXGfy1JGQ4pSEgMGGL832AuHEc0TGhtu/VANQHlDZCq
WgOUqcnGcTYbhTjdtb//hNGMRGPkPJKUSEeAsiXhxeDsBlvBXtqo0gqPbd+FTrW9ci+l+DR5oDI6UJ21IQPlFKoffNnUYBEoS9wi
qfgL24LRYIVpmRjY1ACpE9OyIvXNGmYwA5QPFXb3ZxWxKgSUZop6aXX9JFi2KVbn0LlCWQorzMqeBHsDKcdB4u8BSBqdr58BY1UW
RE0QrO9JkiMoWw4AZ1VWhjErb4AkYB4rE+XIUNA2DhZt2nEOK+s+FOwQnwhekQ5lpIyxAY/xielL/Sw4yeTMbNJJtWiAJdOyY8Ay
XxRUtqw8g5bNqP7TFmgZlm3L6DVDDmuJfjARl1w2NIfEADSG/oB0pMFO8Fg/QkuNzfC0DA07zkpe56KkzYfeQReCPZ761EMg5DAk
b7sgs6E7iFRtQjXLvMt5pq+XepbYln2ZQ8qWpEQaCJIYlYXhc++cALvBgowYomOuh2WNJarz7RPgWBKnvqvtb2iRVXkDozasB8se
TZaSIAFvEQkYggSC2GIneAU38iKRKwAnG3M8qxm+QwokRQQPESJ6YWnjEWPKmIhexsqWaLfAg2WsGpbMzvUTmMDKxAEfyBElJiai
0qlSnU6vxcoOjyrtM463VQtyWiO3rZ8GwzhtZ5j0DmJV3moTJtAyMSp74njRpKO4lyQIQ9r6KXBmZZMjxRzm6RnM06MRGKZ+6DFW
dlliWR6xVvv99FIMxrnMfXoxCsOjtU8GUj1UGatIVbVAQpel9fWTGMulgVWdE0skNKBhvxeiMNS2UFmvNCxP2K0cbrisH3cpEGMu
hc+1vDALlb0mbaOBxMMeWsgdoLJw5Kh52cBbWcHrQ2fBOWyJ5HwOwohQOVidsYCQkdHlcIPU0EMZ0yzzlmIwqOR1Ezl9on0WvO3f
GxKDYT0p70wiZRxFy2GDBZkJSiP3g9qVHXJi2T6BBfXdmfVVvUsv6F1HYOi1WDmCDNKVChNeXWctAgFMq4ugoDAaUxeA5agSGIwI
SZZkUDn1MHXEABKBIfpA0h4qR/CqZNGtnc3506S1WhYeGDHsUGo7DKp3ocygOp/0h6wMbbqeeKM8oqUg8aYEKLsqKxL/CGYlyN0Y
vgGfdSSMrBC4Tqf+aZL5h5YvouBavDcWo2itadwS5s6kbntogtF4JHJCTjoeudVe3DLRNocxeta4D4MAckXriqDpsTCsdToJkKpw
NDKJ6+ZCzqRqPZNj0SHRziItL0odTiDpL40T4CKbgyZt9ABd6AqbiSbojfGmIEIbjxqLsIAbgRV2aL01R3wObYswhtZALiOQ4Al0
jgLB9aAbz+FIgEriGUalQio2M1CX2dvm8gPZe0AzSgRNaUFThu7atp9Dc9oR27TWAzcI6aSmQyBNYJS2beeQk98hdN5BjhaL8EWh
H07sfTQZc0es0VeJC8c2eVEJ2EJ4MrjbSuKkCrT7XZYYRBWzpttmImxIh3eahHQoEqROfVZWBhLSYcrQ93LaIJeQJUl035QlzjQe
32kgTjIH/SAlR2yQkrOUQUjZiWJTgxBb0LabhTrBijRCijGUxioPOQw5AQ6CMnKVpYQypN1mc6YCtaSTueLDMSDlmACs+hSH3OTS
Z99TSmpQelix9nu7d3cf7N6NM/zBse/eqkRCWghAzzckBVrZTdQOPgDktJWexVgPwQRlk54jJ4VhSkYfI3GzEBBtaCc70l+5egqZ
mdIGkJ5pTKvZVr0n7TFLRsZLQassAgZ5WJIvYAe1kkkUlBmmC5RMIKNuUl1OzPZpEDSdq3rJYTSuziuumcrFk7a8on2+fzooVp8y
szWXTsq6lA16czJokomBFNnoS24dbFNomPIk4DGV/q2dQJZDpIIpiue8xcvlS0sGzSiaFk0enavAlsOuPlRHUUfqWqmVda2qbzDB
y6S4nJopmEwK9tS/aoZd9FQFArs0jaQ1JOUsB4eJzgeoP1VEgJJk/HiMemNWDprE8P5UX857J0kDEKUQHqf6PSBkoxS7ZJi5o65y
poUVLRLqm+XoJbs3fQmvY96m87QRtCZRzrSGhjG2eTXuj/msneGzhhxO1Tr4mPPghYSF9pK2/kIOYK/h8qMLajbXvDcn1VPcMimd
7OxE6CYpGCu0b94B9pamZvfKUwt0/A/DNXRS0yy9pQ4zEhIQs7bklqpFoBtBZQa6sAQ1fdHRUMtAN676PHugMNt3unb0OaCr873Q
imb+5ZsBtICyqp0CwxtJfyTFtjn3k4mPJSNzVyJ0fY8qJj4jA91kYCVXwoXqd+eQ7lzj5hmHecmwY6DrSL6fmAklPu3cXLTR98cv
60cvOwzbJiF79QeMAbvWkAhmS7J7JJqCgiSlvUKEDLUT4ETOXAM42gGq+maPwa5EabNYy80V6RZqAfD6GflaK9zVdoC3+kRzoMpK
TCM1JBqWBMMCSSH1SlW/9YQgTQ3uSI5QlCRBZEEa5QYEYYkrN8pWTM+JB1Nr2SjLOLiLfG3YdIxn6a3SlIW7WiuiikiS4TuAu4du
LgdzvfPNc7kruKvWwt3xddSs/unrryMPdyXpisQFKGN2c/Oi83BXqIlOIST0EhRNVNLNG8Be0gRvAwApEGcASIE4CBgxHbGvtdiF
TKdSyKFEGMEy2BWdZ/p1ANuBThcxTFgBdrHaAC3zYNlMhjLJBCvBLli8FSR7D/kT+MmaviVTYDgj6hk0jWMLqy4sgd3UTJpYdVH/
O3w+3Aj0KKQ639XLrxjUOQd2ay2dsAR15zrvTbbAK9rlEcdHMKIZMEI5vq8fdAxzwZCaT4ey3VwXPIp0tQ61UxgLG8VwfL6SfvXB
5qy6alQL1W/gLIBZmIvemaWmztWvOg9z9YZ2XVi261LbIWfUlRqNuk6baoE1gXLjggqbXcbJESqxq3MKZ+xBb+8XJSw0svcgfWiU
YBzEha66D17NDFiIG4C0wpMSLboaHNFAhMOIFCtd+2owEFfPoExLPGeqdXCO65zDrAprIC5Ce1hg9M3vzUNcCaTGq1FcT9tDvbcD
xlXQvgPcFfW+cypH/3gTD5rNeQAJYLhjiFRIOaEqZKasOlnkctHL+JYwZ8H2d4bJDsslg09ELZBYH892w1O0lFz16HPGXKL1BbYb
Hh5NU+bo0Av4FrsXSMuC++nEi5KReWOuEUThyxVwD5EKx354JFzF6trxx+hWim7cMIJPftL17z0GuLS65pgFUVStqreZCVqwM4PW
Yy69CHEDSbpWllQ5BuEcsepQJ6aw1XNYJ22gWdroJYwrZno9b7Teq/yVfJBG9ZgcxtWMh1Y3Y1y9iHGdI2hKO83BXBsU6bcoq/nH
BMyVqdipn4pdUAlZY+xCCl30ueCTTXUXQqssu6xh6dWxC3p97AJtWqg1tzeHenpHFQTal2LE9mBWv60NENMfJ8zVa2DufHM+M9UL
tOa9ecEuPVu+OAxgBemRp5s3fyK4KHLavuX4oXMlKGFJQTZtcmUY0sqy10TTU1cQ7DSbN4eSwWFQG+mTR0xcCDxCEfCYz5qjnfJy
ALXmGvV55NhBlMDdFRlzhEmhSVeSHkWpgx7yKdIWxxd5Habz5WibpOxU12yPJJTSp9lyhYMzd8RBJ4PCLGzPNMM5BO8emuG4klDx
6VQ55BGkMB8eRDATffJaDuIYBCusjqQ025LJVOpc0ylyODgePTo4XkHAnQfRduwYSKytJI5OK0n7WE+svoZafU3L8RuLpjAbWEu7
UUHXcurGwJh2y5PcqSNhS9o2rf3YvynZCH0SYqEn23AVjs3B5ImgscAWjIcmhsNCMu9I0VyFldoMsTySjsWu5cTxwjh01tlcrjUu
cdAme3lTFX1wJFcmGGyWl8oQBavahSGDly1G9FsiCIiV3tF+qXqDSbD707f2PrCFXh8+bBA4whY0aQhknNliSe4z4c9q3uXraD+y
LZaESTkgdjnB5pVIyqd003WdiIiQuA6LU9gErF3yWCFkJdPJThGwkFuDuZ6x7sFC6qyzyZ7wwRHxcLicbhRs6hyEOm96iNd4H7Dh
cyRw4nBiWJv8H+K0vrZ77/bt3c8vdv9U0DwPaJx0r3UrbNZzakAe9M+z1RMYdp/m2/dxWXdYXFidVDcsGp7pPZ0Fa6CrAA6L3Upa
+2CqxnHRNLDpKPGuCSbfHKhoJe1364cmnSVpfAoG68SBgucCJeKKkHY5QtTvAvadXmfuC/nVQ8Pew2hUQyzJhnFmGT5UrGyr75+O
urZPn24a9eZk1BQSIR3pMu0sb0nWgliVErqungT2cSUBcoTRMCbWyRi5ooGxyTQx3+PAS2mHqnrg3Dl2vgEImrCwClP9XtPu0jju
IQBmOQ7yROoWDY19i0mAMG3LaKzAPEtPj5k35Jh533LMaI9amlZuhUfwITuwuR6ZjOxdgcTUe99Ji1lxJuW8h1YZd8lyebIvUk0w
ejHRSLhqFuwe9RUyDukBNBW2vyXHFi3E9u9OEgTqVmTEDO/igoz7SpOVQLaAwIPAU9yLhq047SuNBm56Gjoza+NuPwm86AeiqBiN
8RSgUPBbjE51doONYG9s6hwudTZ7xyvhdTY1q4gLFCbt6RTebEnHHqmGPfKW5qXWoGK6O3xUf6CVdHX1BCZQsRjMILChFZIC4/oZ
zAFjPwGMARGaRoOfFmVHVS0A4+E2sHafmRCLotG5CxI6rRCyeBJEHHq/zGmfvBSlVr8AY2gsh6dALGWSyYYFgFXA3C9XECjb9BmI
rBiIPNkVqnDUMURWJNYi8iNSd02Sav/QZ9EdCq+5Muii5hHymZQDNQ+PSerchPYJ1FJUf8AYgEwOtzrbK3MY2Qx1Aq4N5EQOXdHQ
HP7SRkvq3cdexYcwjIMqZoUhqpivP+ysxA2mUy67UnxcX5BHo2mIexEUVjLW+4OeeYzsNPjQKu84hOwGKqI7J7ufQ8jaEFwI1mP8
hVYkh9YohxDZBLHBktx/bu6ccoaKyXalVRP4eHjSNEQmL6/Y+LNKFUGthci+8wMEoD0CQUf6fkgSdAy2fS48Sk5d37GVawgpDB57
5yZbcsgoOVWyChKbwEeAqMqOKCzD5LI2PbZ69AmMrAcCBFhX+KCTdf0M5o3H1ODi2XoSaNzSqoxtwQJGHrCIYdMPsg1YPLmMTcAS
Rnbx7FP/AVbWHiRAG1Lqu3p8BiDLE4C8LX6BeYCsh5a2Dc3IMI+RB2oBZ7sQHYKHloHHMFkopWYHFrTilTDDildFw4+l0YRE2MBo
Aks2ZLLgdraMiGs4YiMRqGdNZJMpN0WD3i0+hkV87DSWL9G9X98Iksp5jPCRxJVv6yUbK2q9S/VSjvPwNnlGDIYbavTS+lQLWxzl
bPwWZMNkZsCxGXA65DbDLDwzEZtcNQvWfOwtMh4LlhSYMS4Q+7Ek8clWqPYlYezH1DbFOjoq4SmsA8c0SNkwNR+yda5h9ClwfD4v
D6wHx7RuapiAxrlQvdprVu3bwBuQoQvgcsC68l3QCNYjKJN4OmXy6utBRzznyk6oXsbGA5FBA4E2wUV6BTy2U/CYBGui8RBUV78E
M/BYD4S2RohG6wrbybjFomkswmOzsZail03IwlALNjUhC6YMYfIqyPp9YE3IWBnLsPWFDaZT1Z9BLsLiXLYLvWRBZuUC36is4bBx
RuRg0AEunaV9ymjbBA1oRdbBQfUkFjGyWZ+1VzTwGCN70gFxvOaBZu3VH7MZhMxXOEUFrP6I8QiZ1f00c7p9wzJzEAy8pPZJUuyL
dsIzCI+1CPUvz+Njk/pC57BKd4wc2puPZWd1OOLj4DqpMj5ODTvin+pWUccBZHsSyVhWeLhqFixANt7RUHHSB08TFuAc6YrpbVnI
iV4HkO2SuhwmcxWqZrBaYd/AhKuLgyxUSQZz1fvzAEBZakH2BCZ7Uog493hPkFn49r3gYXLkShKw4YbrAIWU1Cn2CQsomqja6pAt
2zaiF6dK5OVsUh+e0oGX35LOogiRHCpTriwwaj6xD/1hqMt5ktaDYCVgA7pQ5nhakdlHmNjQ4z9oh0dAm0XLYtoX3zKdMSu1g2tr
MVjT4s2ZaYdXOgHm2rhDGeK9dIlIypJONdioJkftpdZ4RrXMYQycJ2PCSeU8LalZv2kXxvhZkfg4Yl9QbOFC0zW9/lh20KB0YlaX
bFC6bbuSLJKWgiBpCxinKJ3Grh0KiBhVLrRMg4lZlkuLgCFYKUGuZfQxoAbcApBsGzw0AINp24KR6JR6RbiqoX6lJp7IQWwluNpC
w5w/M9GPqXR83hLtSEdj8INcUxIHIAOQXFMpm87gRLCGti5XVE342uY+WUJiEmKqLwUqt+RLX12wZgNxyQBui8fT4vG0iooIkgOu
tpgFC7jjayLg9oY2ng6Y+xd1EkVbfchNVmXMNzEznuogbk2iau0sxoxr4JL2S8l3so13TKFvOQix9qz1CXGd1G03eAaEp153mOTk
BMY6R+CASYBRpyUlA8oiOVelAZKYO5vCMbL1PEVn9J00cyFzh52yVYpCwSLJoFL887BXYZzNN3bv3H5zVfIfbdRI2t0qLvXPkkaN
/qRR4+KgfOfpY/PHA7BRBNkYRDYGS9apslFnWk1rTTrLkiw/NFMZ0pnSnPZHXxx63CVUkpeVcrH540mYwPKIXPf1iMYC2VZgSoEf
GrEfJLcdHvXlUce9owMuLGn5OWFnliftsxfHG7eKxgbqChcV8pqS/sC6eBNHDU0d7iHnC6ctoUUXSke7Yfq5kmr+oEmlU+EUEWjS
kh4epvA1x+1Sse0x6VKqNlnTcZNni73G85K65VLVy0ON+q+SbtLUBmayCQhbSZe+Ftu8GYeT3DICcUoXXoOJ3r8YTgy0DLZ2HvUk
6wLag4MqZKkT3VtT04GspQXbSW9pQQqVU/6kSTEa2cdlI8tRoKuEySXLY8Vif11HlFVZNzK3/N7nAlnpRnoMhwgiVyWWnUQ1NX7R
JlS++7jXckfSFjCnFg+bQwdI5WszfZU7x3AKLpMXSuXZVBNl5PfUzWOWYh3WvyUnSA+AcBxY6YkgFUSQemkr15i9YDpEdJpFeYpq
VbnsdbpSuq8k0U9EdwHyHZcRwEoVlhdfLYLEYwtvhQELp3loAk9ZanVXNugESERhpCgbHZSFwFF16ahzIJEcbwg4tsZrpQkTtyvw
qVoEiVjtnVwooAxMEUXJlY3Ine1UoI8U0wb0Siga1+5Jw5jC9+RAYj5JpB97wLcUk83YlocbY0SNwwHZxjzcZCnHFXs4YsS0noTC
WAeTvcEKE9dc6WBjiAie9CgYdr8INIuI2DyMtoX8YJbzD8ANV8OhcEnHENHlW0GUU5qAqLA0SOGSjuSLJvJFcaEq9dvHgUSih9I2
WXgyUQktHY4FiRawc1sUKYhSDh3BjhqFscREJkzhcWGFmIiXwAWDKNG7kFPNoyoY/yAznL3v02DSShRiBmyVPOFwIlFSiSSTyIA2
kWQsTtSBdKvRDrBA2eHJMadMkahmpUPlu98f80LB6ctirNvZztQNOuYWhAEryoAXk3nXXCoeJ/LaFyxHxa5/T0aaJjAkiY9fIUiS
aPbNW5tQjPKVq8xesmRm8hoLNKbqgBr9VXEjJCkPmDwG2ZyoUu6nXj7q8HGYE6EIKcowsCaScqiKmBMLR503JwpsI5yH1g6BIhFS
eo2UggWkOMi4IxqJYWNZRdl4HE5MHTcRJ+bKlwkzeoyqUiR6RaxQgmAeKHoi+VHwm0pTIszDRLREGcYShRFjvnQ5x1zYUrGPhkST
x6KGRFk62hglalrhRgUbJnJpvCWlS7WCspE5S6KYR4l4H0vvBGdJdIbphmuydZbAxMK7z4iXs70YDxPFTIZHmf4CiyBRk+JdGkjs
slGWgEQA7CFr7QpTFiyDxFTpRavjWQ2pZid2tNnnVwOqwlEtVrkqZuSqVpWe2BmMiNzgaDQ85bDkQMlQKcNYjDiIHNX9l6MjG8MN
PAkdD1F6V745Y0mcMyRO9s5dPySHDwWH1VALcpggIEtv05Qd0c0bEot8HLDWkug6MExdi/hR47ECtN5F+Fa7yrypPt6SgNlaES7G
K6dIS0FQxpHG2UGIfMNMFEsSltdfLwNEoFAts0/UexUmRKmwQrLrFfBQEvPlBD4cxNH5slFXeps1SRHQtDMZnrc1yFQvwEO7VJDB
0QD4wledUHwcrm4uZpNy/SxqPhKblShROGqBr9ksl8RaHo8DiMT9q8awzZFMj8IzO2bDgQ7G2UiBloj0pcONMaJVtK8mihyFRa+i
Zpl9d1FzLTymHDxUCy7Siaj85cE4K6LjlIlKK6JegIfE4KEKUq+Wx+LgIdVcuCy36mPJwkNFO4YZTeChzIkMqRCSRF3CnjSmWnFY
eHhoovjKjCaECA8xayGVMBfYARRS4Co65VJHc7/ChqhX40NnGA180CBMkcKQpm5kdgf6QNPjDhjS5EASK65x2hGA7nXluzPaqxi7
U2gZZYUHztUNOuYVNNDHNKpAeg1CHHiNONN8ESjV6y2IuRpbsoITQZrdOcmQZogklapykXk7fTzZYCQGGccDZokF0WsM50g9fhRg
t68IGI1dvGOz6UCaphZmLyEG7Hg0qgVJeLZcPmrzKUA4cMhHzWPKS0A/d/A0BUgXDzsDFaUzRCbaQFQEmr1IvcA+qszFUxgzNEPk
o7F4rw3a2wxWurWi4r2Z4546eJGmHAF7cuSYpZTfozG/x4fy8zWGjkBC+QDNfYD2PiBGOL3CCDedyEPvNnE7Ob7DRlEsw3T+Dg7q
CRYhaeCCZo/kQcFVDcpYHIMBEo5GYhcNOBKOpr2kwfimePCxpPAdA5w9DUBF66qqeN0xtFQ4pMJYf4U3h5bXOElbXDXkSDhJdLGQ
k0QTkkjElzQVb8nhTOlpToMhLoBxVU4llpWS6eybgTmSFBqK0EeSk2RISyfadssEXz74BOR0yWeGPQQi5tQABHNKIXMUmDKpRKok
fmstjK6VSwzstIi9LAlQsKQgF+FbroZvzVomhUbjoNGkEbUhbR0seEyEckqZ+gUYsTNJbCk0Fc/QM+gIOqpfACbqhdityOpTuwox
5EhTAUmm0KjUy/kDhtp0bPV782I6R6I7iead+HsOgYzIxQKmzaj6CUyYLuObGsjedeGTVknjjMFmdpByY5TTOWgtCnA4OYffibN5
8/abu5/e/vnuvd07Zd2y1LCegWC7ZU1bE9eOvdgoS7qJkc1ErdO1I8/VcRqWW+PLnCJCNicdCdbOgG2cQgt7oFJGK99pmkdWtegT
HTIELSkiNer7RpGqe55kQsuqlR8noIfl7LnpHL61w3JFTasqPq/eX6ZKiZi1Yk3Ualo/4A1T4Vt4kmBD+nkSMwqtY2rlsI7p2sHn
agZyBVkm4yfXDjjOIj9JBTXjjrWDqNuqNWYKr1RVRlw7Ht/2iq36xYd+Vb0lX+3HBVoOM/DtlKynhylUHia+pEsqKiew8mG8MH2P
imPvHCmzTp0qvFiP+eHGpFZYukFG8YWYFp1QZqKGf+no7JaAD9SNTgrIDjpc0fvtK7dkOg/8zDdgTVFSApMZD4Qe5i2sv3oTyd7D
9n5me9Y238rKaFrlMZelTWXwqYvdUxd708GfSOyO7FYG4qmIJ4vGYVqFRZ9MRNQC87pTbRsLa+e03MRKyYkWUrSCBjmI67miqu7q
uqqp7NqR54uOknfXggWrmibxiJoZlIDVVV3+1g7MV+IPg55lpBC/Qt8USeiRVUNvWmJ07aBjmHoi7s36+serN3e+nh6NNRy7XEPt
mAxYDYaUeyReP/CkEJJxpLW5DQpqhue6ts5JE1L9qG5fx3DVDSvtG6YhsGhb47VVAtuk9YoWVILJAFpMjVg7JIuMSOcp01vuDlhV
GlR8nCQe/HiqVNVJYqVjyroh3DFxA8C64Yln5VyYZFO1JM4T9t1loUFC8R2nJvo8Ohaturrtn0OrURMlQZ9EfzDG0NtNux146ZrW
Yb5iMlcz2BItvGkJxhwGlm4fDT1VdbdvVX+pgnCd0rfmq+fneJrkTdAorbH7asiRcyle0ratPH8jU8GwHPedEoecxBY+6WvAgqB7
YI2hYCnf5yROdWZSUN511bC10onuWDV2EWBd1W117cjrm0ghWCeNVgHfHSrffQVgNax1VUxU6Vo7MA9YhcTtPlhRj5WKPGG/nrYg
rhl9wbga2AJpVagV7tq4utApyixxVzeZcr9+0BumRDlp4iM1aaEo+97vB7EmDIEa0pqa8TnQOrBFbAPoYMnMKsbl9lZ16V075rz8
4luIT5QaWjvkYuF7vq0Dtoaquq0sTFKawCQPJPq0L8x5xEnKk2gAqHrrieaLiRFZTLGNiysQu4rk+czYNW2xUNmys29nbnSDsOKw
K8megY73SZO6cqJldLY6qrekj60kXQgMH5xhm97//kr3WFhTca50dA61ilngOFHVfv3Fm7KyUoRE8rRaNHIosLJ6amV1rDvUdZoY
WVXTLCaMrJHFOpVdHin4AlyOzUlIzhqMB0/NKTTmC0ZsIdxa/WlFn6cw0aOe2joJdFy/LStaPB3U1NPWRoGWZiKBWbZm5DnQOugw
BQTSSFoiPQuhyndneN+gLjcnbtG6LFXde/M6G3UzxNNtWRtrINnwWlaNPkasA+XIMS0oBx12VM2gTA36JSRVo5wvtG+SgxYIgbWg
T7d0XT/uzdjUKQBRq/OYgdxJL7NC0oHE0PcOwrB1y9rxOcfdGcyQCz2bhmtN71KrIjTfteksIHKxZ9PAfm+ZYg+TXVzXjspCVyM9
LSNPyrn2WVSHEAFpFdGFhgm3q08VD1116qCgsART8MGREkzGykBsPP0tCAeTrFTCNEgs3upKpJZh2jQRVhbqtn8WuUpaFIOU7ARH
oKvUmsYVy6Y1mA164rrrovW/BS4wwLW+Z/X6m8djV73UMY0mczVtOt+HyWoEaZpUaFIkT1kjTHNCN6082zYtYgORy1y4OLbJjU18
3P6+kXDPdFMBzxxYuw84d2vN/uv6LvmBmcShCuVJTh9yRL86/m5lyyWPbmOHLjCH58GjlPd+rQJX1m1paCkyRI8mfSykIxFybrVp
bkWjJYPBOgYNOAZoOXqB1ehF5cDMfbChOwol6zt7vA62DxBw/enMgdQp9tvXvjbTWskNnPeouZA8LI1rvl5vWNFV6aSvjWT72gz6
aNeeeiYmzE915MUbB4jywFYfes4mazRNHwZNKhAKkpnVJ9fkJA5ZOYOx5HGkPygHvUyph3FNCyWSswxsCyVHfSq1Wz2206iT7kkC
8za43ieq+p0nGicN0xQEWz6ZxF/VchYW72KmfrIAkgbxgpj+rSIe7QhOa99/wq0ZMS4YTJxRKSteEwtRMCR7CzovNC2IGCJcbpN1
XL+kAeexaDe0eAssHkdbfw/m4w0kLfJFlBFHuEDkFSTeIDjbuhpjXjhUd0kqDbAJLaaaF05jYamGXkAun5IY0V31FGY6JIlZH2RN
gkdRZ6Qo2LEzUgQF7ngg0pPcahGwEHjEP0o1b8VULEI8hBiLEPG4y0FKe4eLB01KBYNEQ5tKUfBycGn/bvfO7oPbt2/fWpnlpYh+
StL9HJPvqkgFglTUqnTYASBWBB5kp4Kj6cT5PpA02xNhtWbUMRLOQ2uD7VhISweN6ZVGDBojydLBMz9U21S9XzMmHvjB3io04GpL
NlfxIQcnCV1rBs6QV5GbjjWrHF+zShF3VfG7wmhImO2MPhnbtGoz758OhkcXT65nYzZt+WA3J4PtjWUEYwiS42FpYVNH3I/D3qBr
Bs4CQ42rxJMi/5pUD8TSEV3xcBm9Ki6mgAuExWBjXzrYURQpLhy1wCi0ZiwCUVWj/WvNcBn1KFo6iZwXrWjtKpqlJUgCu7VQLE6o
XMNdDAlBAbGzQE6kxmc9Ek3DYuXFqLKoE8tjgXS5ZFiup5VIFDG9GvReOaL7htqxuU2IqqUiuXJEAz08OWqgjtQGAa+q33/EpSxz
yFcUTCsZdMw0SOkGJr+VOjpE+c0a4Eo1ruVKytCR9oqkUkL9FvPSFa2srtMkoM+SLCwiXE/K+hStNHfZUrVS44hJLypd2uZ0SHza
23mi3i1cvm/WxD82ayakFkGk0jSelNQ6zWDO0VqnxYNOQEjEcYrUqlSI4xSp+hM/h9Jx50AkFtDSpASPxuKVg+6aZhVuVvMgctgj
2hSF7a4ZkTnkh3qmR9MKKcQg0bYiyaDOl47K4EfNlyDRtAwI7XFXvLZjAGk6puyp4SrhVZyjMXMmfRw4pgUUXJWfHAZESpp6IYxG
EHl44o7N57M4ksG40qHHEmFbrKXmQaQkmchTdkgSKyNKRzwzklTLSJLpEmVJ4y2HWLL4lLJYEgxtpuRpKKohsc20ULuVsvjI8vIt
eSw0Zo6IBGEF7ad0eOoOfaXjNHJZ8ShnlKiVNByc1PRocYkDuNvSljPfOTDpDEH01jg0ZB6eHEoXG9LSKpRvwjSWrHRnlAw65hyV
Yafr7tYUlqwqWFLynhySjGjYYxUJT/L5BUkNAVRVBKjqdWZvmpWdB5ujTE2U71JlcyQ+7acS98MKTBSMqlPQa9gNLCNJsulkC4Ct
OKVXIUlYgSQdjySJEZQUXAO5yiIIK5GkQWuZdgR7CIokHUGSrnTweWbG15gaxNMXLzOvLgVijHREXQoknBSLWac8QV068BhMohXA
0caKub7uoA2RKR3v/EAS5oHkmYcbA0ktrKIZ+sS13j85NOM0BBFov8pbAfNIUnZsI2az2I9ozXBjLEm8BKRS8Iomp2uGG0mebQ3Y
sAQk9WxC/mQ5iTXD8dYwQaxhUhFjmFC0UC7pthSPVvGZYWVbqg4scrHv5AG3ffqSEYOnx6KKjqDIVDcZpKwUMxyMDJwOvy7pt2Rk
dhcCoD/cClK+K5CeS84Rk6RXzlS/PKP1bgmvYB2MrCxOte5m3Y1JElYCSSW63Ps2FdknhfVzNEoyuwOmybjqE8ZftnRhTEBfZ1xV
g5cNn/Yuz9RMUTlSoc17vea86UIgycVboAKh1mkQegWQRPu3MlzEm8IYU+VW4Sq93iCZWToyFo2MRSNj0aH8hRlWZvjQgYn2cScl
/NeMycFIjOVMHa4UMlBywHPls3TAVfGrzoNIx4HIqb46a4YbY0jHeFH4Lp3FJ4jT7dGm5FC5N4udTtYNd8PYlBSpWaiIKfLw5IgG
JKIBJ2zxq3IWBTPfg49AuuLDyrm0GdUj0CgI1D3WRZnoWQy5tcder3dsb+Ht04soUmhFajopSaItNYGRhiQXWe+KBcqEMXKf9e5J
Ii54jUXy89N+a13UXg2Q1CLrwFYKGQ5GOtp+mgvboq2ZZMXNnQOSxnu0ChsnSDq8J+2xrKBl2qTz1e8/YlqBiSVcUTatZNA59gG0
4RnjtIGa+8VjyW2toHo9kgwWkWQOHTv+fvA2aktyiKqP94TaFt/IKYokhcFGaPg0F1DrvRPHphTOrQkmmc0eQvGPt82RViS0TRMx
B6RgTlMxNA8pUy0OzP8mXdgEKdsmiNJ4WnJl5egz0FIG4pD11CMbSDtTQZ3tyUVSMYkxo7N4/i3qb5ZUa8O4aWvWuF1mMoYQE0X5
rzFfyxwlzfF3158EeVRdXESA1leMPYaaGvVVwCjtdaVcVw46BpwtFWzXbu3YGkDa3ZGYWIUcVZGYWLUqJnY6OYiIL3CKFBm1wZGk
WIXuNHAG6+uAczVvzVgxNa81DYIbJvurrBx2jEVpchCa/MDSXmACk4NUxaAj0XUOkDKXE6RIvkU+WbTpKjHBwCCovWZgFhhZh23l
jTJoHD88OQKjQEoFga4ZfgKkxnfxEkiJQ6Ex+z0/zA5zabL1RSXV1kK94GJgqkWcRvr8WURqFqGaDVXMbNbgaRQxO2tJIqf7J4ft
sDQSNtiWJRhzOc16NElsnKUmIdUw+Dyz0XwkFdEK5Sq1cCnhRzEJP5PFNqbq8RW+OSfCTZcbRUQRju2cDr8f2zkFbOckqkT4PIqN
grtv/6cO3eZB5wRkfJj7VZD7qE0KGB0cxm/tPohz+vHuw937t1+/2H379hu7d2/fWpXwYwiy9/O1P6c97IXjDwCtIdEOqESz2Zro
7Uq+moYJjDGtoWnZx2Xg2zuR7hG2aRaZLRqSDeQL3O+iYXC8GXQDHFhSlERhvAkQV7zVRDIPOysXTiKjXIP6u16oRzrVpqdwaBgN
rfPQwLg+J2vxlG76/TXj8kWnVdO4NyfjRpYjBSbDK0kMrqTvvSbe+obxswwynP9tXMNPEVdzyxHLiJccMTUyTk6km564/gqHPgo+
w5ls5nz5ummnCfA1GLNgZzo2bHXGMtQyg3yXUe8E0ZHsNJKcNrTIlh4wImaRpaXug5BLTyWnvse2OFHhkSrHGchU+8CSvol63960
Xc5dMoze5UNI20NZrvqTkk0XYISDKQNwVhMG4Iha0pfdOXaBk2g4N876Ldbk/sfDf0eYmNwRNQp92f6OnIBiwg3VqgqDuknqj3Ex
lf4SpX9E/xb9ttiuyfeRp8e6jmqLk8nf3MSSVVbTTLyMACHfXNU5jXnwkPIJA+khFdyw98vyzNQiQJaIDCeaZAwqtomG8XmAHFG3
Z8rmGFZj9KVMQ60EyEk3WWoplQ+ra1oGhm9qNo9D06pFgrQhbRicuyK+U6CY5lKucyw+jgIltMxhjI99XnzDRAA0YRc1D4/XQ4iW
ccdiweVDz3VGpQqBbBx5DJAtwcdgSIdUMA4D44LAZpKpm0PLjq+DyWo5ALRw3DFMdkwxAy7WJJSKQzWLkRXyuNnX1U2bzWNkPSP/
J1vwFo7MwbAoybDkLgTaibfv3HnQxKz3RBkrRctqES2H+JI6ZAVcdkrIo5E4xKVXkG3EInI/BejZTI30vGkXeRxW9mgUIXLXs/W9
23jAHFY+9Ao7RCf1daIONmOiLR9KKR1jTKzbYkkYqOxnoLIhcWAboBAGKgNzSTlF1jXxhmmofBcccQ4qh5QFRWpIYLChswgGPYk2
lEFtsRMsUgaRGhhbEhBhdEbKERNCb9Xr7YdxKtKgFzFF5oRCpAjLtmTk4rQZgFhTxKlw/AmoLDs9ngDlGYY6V1oWYA4qa1QpyUUl
9T1RjGrftAwsVNaMiWFVKlfh4NwVSaWkKNtmq2AMtcli5ARLtuRxN6KwjSkZlrCyXoWVbSmHgnmsbBnTJlfYPTSdeA4pa0XqiYMz
XB8clIsto3MIWa+CjNC0y2OEXFnuq3DcVQgZtrAUwXorMhfWO9k3oHBk1oqsNAksBUO6ttJWasN29sK1rPuEPTm1oJSaNG+1MhgS
SHGo59k7bm1ExSFbA2yETUardmF3yXoOLWMYgW0MI7AWJFutSJyLBxpYwZn8/UkXotr1uP/ciqPqNj+q0wiZAKBZnmSaOPE0Ql6r
RG9xGnljsrAEfljNly7QxJVcbEyGAoisO+hr2/TX1h4dP/0tSfUtdS6IFbEaaG9JQSyjS93cehkj442dsOY6as31DeNPYGR0OdCA
D2AxcjFA1GsxsslnlZiTDaoKRtDq/y2zYLgmAiZpJ/pek8rHLYPz5uQgNKNGelKGlTTSSm06QDXMYQyRpcgL4JdaiEHTGRyDZFr9
cgySJ5PZSjedAcl2xoS+3bgMSJaaxFsEjXlKWkrPAGVjQ8t+rxNLPHZsefc5qOwXoLJoGJeRg3aVnaqNr/BQWcxY6BqMyXoJKpO6
XEQL80Dc+d6TJlvWypYjxqNk3XmpsD5XSMXkqLgFq7JJKjkU+0ILvSlCd7IYJeuVKNmiTYS0ijRrmm5VToLfIiD5ikCLpxlF+g9Q
W7JTpbZkvRYpj8MuOFvFaT/FygnMqe9yW4+PXgOV9exFNcRUs8XbMyhA7cNwMHXdka5cOeoyFWcgbbmM2+Jo8nc3pYx5m+u0R/QV
hCTZrBZzsVN0VbC0sh5AmQ9oNs0OpQf6Ozwyco8CxOO19aUiZD7dDueA4UEecarHmxLQ9x584WFZkXRHLKskfN6S0EmSPZFCYPK9
9aUBlNPZd8g0SBwCGpctAkeLwNFC62owt8ahgT91kDuql051uaVN/GxJww7ZOIkxeAZ0kYNl23fhkdCliGo6OW/ByUG9LESBs62b
MJYcyizV5EOcA6L1WvJWZ0+tzst4um0KY9m1FEs83YG5fPQxqgaiweMBoI29ENGf9lErH38kOlO3jnGwvmIbFOvW48eB7HgBiOBc
6O/VzAVZJGdd4OC2IzDOCtrky2poPIR8tKTvvEef9L5JNxq4UmqucI70+oqIm/b6MgCbCE0Gc1tSDIREcFna7iuLilI713LWH90s
5QjspukOxnrM+jts3rHtF9iNVmbMPwVrgBNsYItqPb7TAFxqNhjdbOI7WMoKNAvt7MWafuf1q8FDC5vLmBxahO6xBXS5jkv8nEsq
R2wBYqvNmbrcQWPtNBFvicO+7Mn1A+hTj4xeKVI/KHICoU8aWMZp7f7t3u7v4yw/irN8b2VzMEksl2I2vEJ6gj9PGysWjD7A4nc/
/BiGSy5v3rJ585bK3vo5ZI4qG2vcFI2K90Ly8e+kZxjWf+7shLG8aOyMtGVjPamiUWE0amXp1bLdvX86qjS0MTzHjadzQAuHvjkd
etAI+vilz8s7fHH0SchPqieRpRGZxFxaIAlwb3j1DKXlOgO1n4qoLBr1KPLu9IAR3Hyn42YAJkm6GfY6zV8O8fRA7NNAozjSl/rT
RcQpHjAR1Sas9EK+9oAjf3XDp4E8bRVol2Oufmf3fgSRJS3BSHbI0B0ydIcM3SHTskMjdCwbawJXjT7mQnchcU6QsGyswV315pyc
h86TKgTeYzAR5gECRjfvP7dvAXtVU3Cwxw7ix29G0G9u8Czgs5JJqUXQS7psLatromxv1DLopaW4JFuKS9I0iPrhZ0CvxjlonIOW
rMWzZQ53yx7VEvL1pKOfJ8FMfhDLTOtiWF09gTH8PRPgV/Pw90wsUM3D37sZ9WYk+TBskTSu4IpgVI96Z1BXzUPdlvJLRUOPBNuZ
RLpawrtnAvdqCe9ylS/Mklui7EzxANd1KthcXAq/HuuDHr664dNAnrYKLwbgKsW6i9RELclm+ckCXNACDfaHLwcVpP/i6JOQn7Sv
x1jRv6t7eOcoV32MKFetRLkK06MOHw/l9zSWO9MTvqO61eeuqk9x0uZoVsNvew6Vv7nBs4DPSiYFH6tVF1YA3DPyB1gLcPFO0uIW
eCc13kktGubAANy7Wn0e4GoCcDUBuJoA3EG0vq6ewBjgSktCWthySAMPVP27j1HumTAJLBh5z4P8YAnlCmMs6fZrMIn98MXRJyE/
qZ7EWOacyfgI86B3vprFZD5Q0ah3Zd+FJbx7Fy/LgSvtSFndw5eDJuVIWd3Dl5Cf1J8uTqaGlM0DR5SH3/Zac/7mBs8CPmuVapcf
I2ufQ753vjnTyPdM+hisA713cTumQO9dcIQ5067rM4sOPlxDClhhxtHhczj+3r4FvGnXpA4xWJkxf+2Nu/mrGz4N5GnJzPQy/D0j
CNErghpIaTO+XoOm4XT1w88FNSD8BYS/gPAXxGREZ9EcGB5Jupm45VS8+qF5+OtJvRL0ePhB5UP0ePhCj4degL+KrWqnJuJI6999
DH/NbBKeIXnKDZt9V749vQR/pXQIfw9fehl4+OLok5CfVE9iLH3u4tUZm+9dMZe7knj6Y8LAehEDa0lgliYtjQ9fHH0S8pP6I8ba
lXTndM4ExG97pJW/ucGzgM9aRdvlx8jf5zDwIeeWScA9fHH0SchP2tfjrvxceh0GvgsuNIWBz2T70Gsx8LmduXo1Bnamk8Idwxvw
26EdyeGbGzwL+KxgUrN5djgfPBYOz4XDg+HwZLiyozGfZUdXROByCNKaBRdC4Co0zWAGB0tNuJUm7IrUbjt+DsffW+Yy5pcGnf4G
vf4G3f4G/f4mtI3O3JJjgLsiwe6HIPhsJxwEwafPLXMYo2KN0IFWzwO2Ewu0rQDcpUl6OrVOcs2hNNscatCMp3F0JhpCWCwjcPhy
QDH9F0efhPykZR53ZamZTqxD0Bpor8UMW/HyQaBN8ZpGH8tHIiAlW5RF0prXTYNzyJmmdRq2MNGgLlDT+DyCDhRBB4qgA0XQgSLo
oJrO3oQh2bsAaEg+fDsYkg/f3OBZwGcbSEcGR38MMpo3JwtFOYSiHEJRDqEoh1B+k1UZ80vk1gq5tbJsawS7zaowqPp8IZpLaXR3
YUtakUSHx1SjR/n42Qj87Mjv4fj7JrvCR0SlqoMK21/mr33ITf7qhk8DeTrsAPju7ZtxRu/ufhI/fSP+7we7H9++vSqJjnRPJs2T
SRfUQFonk2ak4rQZadkUJppIDzqxKr4TKyYThtPmzWVzmGklPaylYMa1FMBhR3N32tK8bBrjdqQS6KUxTPVM7KcuoW10piVlauaL
sc3Y5cbkPlkJYUpEm/K0F2nZFMbNpdtacpaNPu4yXd3KvnDfRz1YDRODZpZbTpUPzPWaxuhQkMqQVtOk9x6QWqnaNV2+cRPY9XX0
m95+3HNaYmhsIEwHg2OlJExHtTGdUftZ0gMaBycdoEm3+tO2t2Ujcx2o7+iosz2ogyflD7xGxHZ4cihXISU6P2wjp+F736aW7c5k
lpekDZCOBflpz3njkZQWua/tYBieXCkJma7UVOc3rM4/XQ60dhbcTmmvJAnVAKIB9U8O3CFINFPrMLRSVq/K/Y+JO083q6Z8yjDG
aryuvu26TrSrRohEmBXyKsKq2vjFTNPq1BuZVKoFg2fTk8ANFUjghlObbAd7gwWkKuyZifsuBEygzc/coVqFcNmfkiqmu1A8NbWM
pj1ukyNo2uNGBUTTqc5eyxR4NK2wq7oCnIPCrurKEERv2uYwg6ap6RTvDDWd6nxr9EkJ8tJpjNkoSZScqEumaCPvprPA3BVFaqAm
DcrhBLLXMzU7D0QTd7ZlEgyeRpAhBTmNCDNS0R1sw9NBy/hjRI1qnUOehUqdQ57linmWmsfUjnFsriivXD4wg6m9Ie1s+8qUB1Dd
PzmGOGI6n1FOtkxjLK2oTmGYiBqTI2p8y8AMqiaCikoqycNq2bbxM7BaEtaXxwY87tDG9zhYfUeLzoE1sAENoaAEKWrRPzkcOhmA
mqubDh0nk0NqMqaPe+BN/CaPXjt8trdWp2f5IEThLUNQG0hDHlGz9iWLZ9KSM2nbzuQcpraGNCiyfZfDY2cSie3IrUHeYI2BTZbl
DiG1Wgupz6cDqzWg+o5XgIPUvgPSw0+QOBCgzXwxDOSk8Xn1brCIOvWiF9nxJVWUy5gKSJ722CEZhhV2qocuQOk2wTKktqwrwa6p
T1g6gQlA7VlXJOYKADpUoFz9g7VwmhQJR0+0JmXCcRrGFANaWDBOC4olWXMZ6akUPzedBO6m6C5335K6yyGNx98PcNZrUn7D2JZJ
qDsFtLDWOl0UwVq88SMpgdqkxndGXVLjO+u2d+awtFYOK0FpCRgqfXhy9MKTRPkIa2TLPMZyCjVJ9EqgHonsR7St/RhMn8/mBAso
uiows3RcDkPf2XFj4ZlQ5Lgpi5G5hyeH1nHOIz7zXrcdN960FaWLdEf7q4+32mHNKXx4tE0HYwKJAgwyqA3kIAOlW9Lja2fB7lTP
Cw5uBG3Qc3V4cgDSXnrSvES5TVZlxCPv6J5OI2lq7jBMjBqKJd92YZeM09TMRIA8mhu7TbaAwweuA49aXpAYRQCk1583CKUlbHJL
+Btsoct6sNQy4fZsnD4+69F9VIRB5kCPpAlrUcpS9DKONgjgiFn48PPRdE9WB1qmMBHogRGNpFS+DBOdj0TLDObCPFDIgMKlABQz
YImhyLUtBYOlHfUSsPEuBErKciypl7A0vSuRTzj+rljiyPHFxmk9j6Wr+2KXDj3G0p7A+Lz76KIhHppyB42eR9Pni3bQS1h6kHYo
jPcTaYfOYrSHktK3zGMsqc6nx+h5LB24aI9ANh1tkW33bQZNb40O9LJFmkHTyPEtVVhNy+BsAIHASA+wJM21f3AMLrLEIm1s09JP
maSFNMdKHC4KNi0BTdKHZz3TkZHnCTRi+PjVbCEHF4zShshjy8tj18aJZgM9NOlIogVxWB2eHHbKkzAwE3nGJuvC2BvGuqdmouFc
sYVLr8PSgOAA7wwQF5JDs3Bou7A8mr4jfjEf6IHNAG0nSd9sR/osgCen02yzHSyWjoA1+CBJJ0/nMHAan7pD1yElUB1OIctQKExn
ExXztCyKFEtgnEWh4oir00EnG2fBg2pPkBQGkHpSmheztIIrFasrEhaJikHCXwz1/hoSAGMNxbaieVXGrNVKLmuR9gVEnGF1KR+Z
SVzEGfj+0hzOhj4C7OPv+0n4zh8DkuKZUbJ1HmOMTTPIPHq4BhlkBjPIfOMExkibmouw/iw1F6GT7bS+Zc1RGEmTcwdkTecxUue8
IA0nhEAg1D84FHsVhvSbELZ1ImPBRsLr8TDQ6HrZkXDn1r0YI3AgkpVEZwCRrSQwLYXYNk5hJFkDSeU0nM5J4lJa35+D47QEm+VK
sFFfpmxmTCzYs95SEzdB5f2TQ1loKUj7OTCtrGHCyi27IHK0VAid7b/1BzM/7DdHJ7GevXsiajHemG0kKgPPHcoQizLEoQwhzW1d
uwyZw+fGKNLeUeVQwvzk2FAdMBDbCglbrc19Jq6Q5OkYtuemwTo+eqOJMBytoR1D1ZXmAbtnMsQ92xnBbrQUHOSIciz3C0yt+hCO
5kzY+NkqhBxebbY1fIZF1LX7lIo+eiMyfZ3NAORpf2R0F1Su+pg6BnqnB/f7L26/fvvm7Vu7d+J/H+0+vH1zVaYjEOMzbSxmufI/
1Ht1cl7Wjz6A6UAS2snwCgOOaE67Rzh2YvVcP/wYoOc5aEO7XxFISGtLIyQ8jSJZP4fMT4FcVTI0Lf81KL5E7B2+bmi8GkDM3dIi
n8g1qAbqbOpVizDQisrhMwgHvqWoYSspENxROS6MxtWDvS7JLCzY5/unow5OOeABYwpm+Np3zRgbCK4Bg0JShoABsEqRwrLGGocO
YTc0Yq2fQRZF0BhKuH7IjKcBnb5kg00e1ZIwdzORLbt+2KOwQz42uMeMgWrSibF+UIKd+ePMN2vAyJDKNc7oC2i/ZkvS4Po4i1ET
FU860dmTTnQFp4pIUcI+oiIkMmgP0AW0x0VQLElLTFBxdKwZEFm5i/+8SYpdlrByzdYlcrVSZASKybY4ILWBjCWR1CRzOUiBHUKF
GCZZVCzF/TGPJdyOD/qfqF1aPvqY51Qaikvu4QDw3t3AY5wLA9+1JZUKgkckQcV58MSqIxuPIH81TUK0pE1hSiEGDDyxKQI1F/Rw
vpMSmwO60IExevW01DLAtYN7KVnNLEwmva0ffQLgAg3uMEzdGcyAiUC3cuwZdAthgHhQImrFphzaghupFtEt1S2QEwz71xN0W7nx
3IVIyh6OTRunBFI5OnS5UlX6rCsnwODbQUvdwNbtmLYsrh95jHDNEBKYjRCumke4MOxPPY7iMCTcrnrUMcJVmsZvOIkST2piBoL+
ycFyF4FD3QzG0kby7y3ZuODKURdALgY3egbklugTahbjAh2UC4HWRKxXMjIe454FTqgljKsFzR4ESbIHjSOlHqwMJGdVVy43K0tT
FKQI2Z+qUjGOo0KjOycMaZeiPWRWlnR7HWyTIOMwrh+cdnQXkGDFYafdximwqodzjtYgJLVSgFTxtl4B0WlVaFyM+7OWBMXo9JNR
GeWjcyiX2s1KuvSVXMa7A7pqLdCNUlrQI5hr5SR4pdBimfl9KvcgW9ef10HT7fSY2B3XwMsMdBOy8dmVmQIwvJOOBGAobddfD1gD
dMU81HK0k5GqG3oC5Q54tALWjDso81j55nNmXDEwsnm2RSApS1xiD4EFoOtpAX3DNpEiMeSV4/Jqnx4wZJoqSzJlFUmVVaFy7RmQ
O+BCk+X5BJbnE3VDL6Bcvj/KRGGBgo1mUC7L84HlurZ21DHKjRoyaY6igUQp94HJh+YIMmBijzppFrF+BmNpcyYIBvMo19PV5tuR
1aBcWLDkDgA9Z8lFi6quG5RDuTB0EWxlyYUllAsQEE4pE0i1QysJygUS+nDae67gaHGCNMhEXmF5DOWz5dL7TtoM64LpAmDBObF3
6bsmScbB3ME1n2Brg6Y8tm0KrClXASmKYUkVSkdSLklwg5c2NK7EvCUXmGYotNB74yIs6NbESchATai/ijzGJWUFGGi/wV2cN+Y6
UlEOgwM8aQUYNS5iytWydfVZhJtshArnkoyZ2Jc9FcJQId9NaSIL07nrRDyNXZRA6/mELjXlEpBr2XadoascfQLkDmqnGaZ2Ghr0
VIlBT6815Q7eH/D9Ad8fPO1+UPn+DEskTcrwQgyblJHgJl83Ln8XrME9t12uuLlRnIKeh7hhth9qoL7kyg3nsG1VxayC/WWA7faO
er1ou5WSoFpPW35FjIuoVguSe2ds5aHm5IuYES7EOl/70pzhdlyBrxXP6hqrrWahu60blMOzlfW71w/KmmwDKd2trSJJW5acKCMN
SdrS1SeKBbPQQW4z6FPQujqeax8PEvYxE6m4s8sZOSnICXybwOKwrF0KBq10wOm1QFY70tgPtNZkTyyp7oahIr5xEe7PGgsUG35l
+GjU8tG5IKiF0evQpF5nqmXNxJuy9nkkmwNSkjeYdE8gotsTf4FqPIETllrXhaxgSqGSPzDjCGXjw1weJcUrBIAMY5Nv3KzXMWeT
5IBkQlEHnuTSsWizIL/+VM7nxw3YFPFPYwmngJaOgFw6+PWsYUVyHGFPQweqobXdiSy2BNbHhVHVcxlzSIvYx+INsXhFrKZ5cfUb
wdwPJzul8tunQHSNXaNyTUGXbEOYVSFF/VaMAe6wHrXYLDhkOg+OVah4X+KgGLWv3/T7i6EpYjOj/XTy2zBsAYGKUgKFonIWwxaU
J5VpVbD1bIhBv4MIlcnWMkTBqV+CMRAm7ewVScLEKw+Klo+sv3UjkSgHHMezXfwMjdOsv20cLk4ikTK8zjAh/6hXnvYiKxqfhWNg
0KQGwZDmcKAHhbhJ8cCTIsRlJ49FyaYLUuYm2FEC54KqQaaUlxyhFYW1RzNnynLXen0o8XJuG5UDVCajSkhFAdFbXMN9mMPLVpIO
I4NeMBZcINWQSXyDtSK0L8mYSwZq9eGqzAINXhTNUxhzqcBf1sB23GwBBlMIWg7cyqEwRbhqFTigEFlgyIfTJZmBDZhz0z7b24f3
OMF1suHSzsPpuNdeOPT7RiZmRY7DSOxaIZyGLljIcQ+gI/S3g/v77TihlFX3fsq0u4hf30wtBYvS2TZqmlc5k4kCyXIp8pdEypz4
LyonMlfdDVP+gRTWAUz51xLTzPVJafHK+cwXebuz/eFruZBSLoqUcrFoFwkWQYAcVnqrnAnTjQTYQt6rHI2VkxgXpKguL1t7LpgK
cJVF76tnwJSCQ0ROm/5JQ6rAodlQbjGNuUpwnLVmMrykcvxxOQo0UBumRUyg9QfVBhNgiqZW2eQrh5+vDme46nCkTo/Z5CiyxQ2E
JsXHIGA9isOTY96md6S4gduCWU7UpNCp50HAmhROZgRCHvZoaB+IlCV+iqZwm9yWmbJxhnVPrrJyN06G3T5Ny4lYUlTy8OQA4ZVH
fmKd3XJaYxZrybl2mBaLZfoVJmi4DWcy5nGk0CzORJMbpvCGuU1u+VRJubsVOnO15UznSTK+JX2GPUEkAkijYVBb7hLf+SQp30aS
WhUhYCofPu2BStpNl3MQTDJf1Z7pFU0FSUSCZCMSLDXfiQ3mMYHzWSvmsKUeqea1xUTmqzhnAIFZZ6BpcTOBKH+L2TDsWPIt9ki7
M1L9XKraW66Wu6ME0h1FA56S3P4rWeAAq+F4u8FU5nqkWK5Hiqd1ds0GM+AwflVAW+2ZmOmYApyOQWCV3WQF2JJzNiCsUl5b2g+O
1HQ14EmyoN7iSDBlms6NdBd6ElYGaFSO/3EIX7XcVkXMBFxPBsZUzoAH+jSZzZDA/sOTI9CXJGAlwBYngi8LHc+dEBrLQkuX4z3x
2bH0nFdYTiByca+c3k7MLlaHVnx1aEekvtlwOmzdwBBIGRTCTw4PDrtncgWvvUV/yzW6/9zHeqanQb5hDAkr8kWqb/dUL0OGywBn
2TKbHpbl8tF8HKvrFFFT7ZZHZaIanewUKJ3RGkS9AqeGT3tbfqolIS1B+DJIWTdHWGPLJ8Car4xG3LRig3lMNGjxNLOSK1BK4hR0
rXVhbcdDwJsFGLABhkX4cpPZMJy4IR68chJ8n1BS4QMC7owjaRISvQky2A1mopjeKeMSuivCeivHH4N7y4QVWxbLbXFLxkLn7FgS
frGgPSxB+8rOX5UzGIP7OzD2zTdLPLu6Cb8AdnxYritNnEpaKyCxz8SvpL3wJNhGwRancgLeC5Ehqw9HQH+A94dne09PVMWkO84q
QBcnL7eTsBy4V6xBj1TpUjTMwm84GVYzs6S+tNakPv3hybHMCClQ74x3W67RLJ8tyAxqnMec7YIreUQicjaRN1Pg/m55zHyXRZKv
SUvvKYSqmiB7Y9SW+/P/MfcuybbrSJbYVGIENACOb7PmcPt3DjGC+MiqOmlZlUpZtKokqyyVTD1FpTIkZfxyCufMSOQhCV/cdJIg
AO4Tjfds4+DZcxA/X/Dfkl1246QYPut6yjbh4KXcGZYIHuu5KiTFwVdDpXuki1YmXQSmwel91GEkl0E6hiNgt0E6wNzaY0rOoD2r
bQI7MXmkYAzABZE6jEe4ifnVrOHZrPndrOHhrKtfzvYa3isgL42QB62ggh9x+oiq9YnZK3zP5nuBG3Hlfl+KVVOHMewx/uN3/wVJ
YyWbbrX8PcZfWB4Wk7mGWn9Lz2JuI6shu0/1OCRnGB8pnXhLwAHpckwljK9Oylyz99N3WRDBtFbFWVkp/orHMUg8joanoIcKkWAi
wbOTIkH1FHx1KiyeolyPwYhKP/qBg+rHayLXLFw7vswR4wm1YT0iMU4t6qdhf4oKRYojLyoS1TgYkSY9AaG9J7AVLD2BX+4LsI/W
95yh3eWKduog2ak5NLcnGvr5/ffaMbh/PEjIlsP7TMc0oVEgUWcmpjhwkPt4pHzsuUpHpvuYAle2npirQgTT/dK7huGT99lbZ6cU
tFiH2wp5IK2YVmkPcmlMn6EckEFyiE6AugscohM5sHa8KW2fsZxxQlq0n0MOmgULugO2b0eVb48CYkgHxkI4a45fHQ7YAL2qPPaF
DJG50N20KTwzROYE7amuUM6GnUK1+4xG4IlkxEtsYCDGvJazI63ptT574K+ZUXVNjljrnXLNXbYoq8rrsIQ3MiFTn5SwqzGlRfUa
h2DuJwW1mRWwvy09S9lvx0mMFFzsM6C90gzC2yxIVYJ0r0nZPwcMa0wDRJKGdaYBs7+JvY7xPoRfYfSfgYQ5eC8HyCvuNBDpiWB4
LAbGYhSenQBnp9NYZGfAJtZHI8kkxPrYSJh5S52UtOwPGO+1ubCjmW3+WqUc7pP7Frrk5PRqwkxuiIp8V53980ozGVkzcdyg7nXz
nfsFoF6m85Ceu/QsRohgIWIrBNt5rk7v5ijdzZBUrtMQ+45HeEds3DlKpMOEN5/vdQfIrwnNl6OGnaT5ctRIwF0bgnqLmjIA1vHM
v7f+fWHDznXMx98vde06LJtMLD/un5loYF69iUk+5OAK7g1LTcjxT/lhYcabwcWNp/C/jIP8t8/ffP52HOrvP39dlNCrJQO9Fg30
GsmBVY3gzaNBCxUqiF0DxI8GYnxOqk7y/omQxTuga+cPd/zhjj/cVX54vnF1D+7YUql8FjQWxbAc65Wr4Ky/A/w9rX+vkZ1Rvm6M
UywVSDuBQDh+WHRFHTAUFi/sj1eplRFC5QJ/+SJQDZbBDiMdhjmMcWoEZoWj995ge4cOo1RgBtx6Xxo0CXUpkOC4SuCqw971gQCf
33U4Mr7SiK+w5PPcWPGVx4h4z6mvU6NqE4H64/M5YSZmlYLmrP5yM2x7E/Q2aKGf+0sZMszQSa0OOHdadOBPaUk81KrygV0cHkgl
fcBUVtc0AbvLq9LVdFfw2++UF7Sq9wQ7Es0WMoo1fa+kiNMcmL5ULeFI8OV3gL+n9e9Ncy4ewYljOyQHVQGX5oxGcjNsexP0Fg7K
XCNQyKjkXWB4GxiL0d+qRrCMQB8696YQgYKN2sL2t1gZENh0qsQLCLQh6r9UqrTxIyDQCAg0AgKNgEBjOQI15wg0QPC4EgIrmGGh
6mPfhkDNuxGouUKg5CCGc2msDH5IDjk3Uu6pkb9XHv21ljkHpPG0Vn0EivkqgW9QU+YKkD78iSIgdRE4tR0W83AxQpJfhMyCsVG1
iURtaCavD3uguTlb1HIzbHsT9DYoJQGQYkJUPCidqg5qL9yVLi2J1fBGWBrL+1LDG2FppNzTNAu7C63/+8+UodKHD8ERKu0Phk0p
Kg1zLbYFFORySOvvAH9P69+b5lx+GOpRO3CZFm46tWmGbW+C3sJB0TUqhcsXbl9zELisagS/FZVSKSrlD7dGrFdtDguclIp/2xOc
rlCph+fYNmeYn2MenmO+/DlG56gU/TdG9N8YLLxa9cV7aNr/aqNzVNofpNEVKsU6g3KZQagy6Gqk7lVGf/BN7zaO0ikWrcxIKpUn
YdH++pCusChWjbdgibNgibNgibOhbv+I6o/s5PRP2Y2am05tmmHbm6C3QQtJMFSLLmYt1obWTTpQhKHRAyvr0pjXY2kE7Em5p2kW
djdY/1cYlcHQx4/cu2AolcLQcRNDObH8Elx/B/j7WvggNs350Tm0imEoN9dzuDTDtjdBb+Gg7CUMxagTvui3USfAx6hqBB/A0GcC
A2wpDGVAaJVINc6A0Ko68cKN15BIUSpVhqEJStfw+8vD+wvrZfvy95c9h6H9T7s9R6D9gYs9R6CPC9wj0G0CTIKyBJsEmKWRck+N
/HfoDXuORfsb2O2bsai9wqL98b29tIvuqhdyUC0HxqvBVe0a2QAz0RFHznjJTac2zbDtTdDboIAkBPo+9ScbQm2AKJu5sbwI5kbA
npR7mmbhx5vv6He65+33uOdtKQKNUEMubpjBkVCR41Gm301zfmQITVYRGEKX5moIXZph25ugt2xQp6lmGrJeFTNJKU55zXYjj+ld
qlK2jEMT3/aJr/vE933iCz/pauEnUFQHdJmjzxyd5ug1D9XD2F+BnmPEPAeJeaj+xWFi3lULFg6DT4PzTJ2YCcTW3wH+nta/V4rf
A1MTkcAuXwNQFJCtlSZWf/ceoGJaGi82pKURLzU1rPTe0PBIIMxxupc+qO62NITqbksj5Z7KIQiY1YolIO0BTU/tt++xKxEmIebV
BhM12KipWvBOx3U3xJ+la/EWA/ob/lpDSDinDhIMb4gW4awhoN0zpKG4CEEtX0NQXOSlJtqdLSZ7+mmqR5AVKzdn30tuhm1vgt42
5SZgXM9bwfNe8LwZgJvXx1btKsJcFwheGwGqbCyNgD0p97TOxY/3BD5cZ1U9HB13lUb1sFOwIGuK3/JqyKpg/e0U/w7w97T+vXUZ
jqJySGcHADTXqJylGba9CXq39Ix///n3H3/8/A/j+P748efx3//y8f98/OEm/aHFympimhtW7nphc7s7hINqyV6m+PNY3Q1T7ZrG
cFZNTYuFkoFimJE57Wh27w3jvELyXcRwV7pcbdwAnxAvAoQNRKA3bNqL5wWR9/ABjZpNgs8qITvBEIa2vibBZ6U5jVCejcujpcaN
JhVDcIGhiw4Rah/PPWtdNAdscdY0nbuOhYPuit6XPuDyfJFFR4ldJ7aJFugK3/TRcv2zqoqAd0WLWfBsdYxGS/ZfHRLg5Nh02uTa
BRMJclrDEaIbyOZkfO77um7coHWmM0/jPWR9B73XmYewdhRiqQnCQBETNJC8Y6iIhUWyzpous3Ja3OwWCUftECTKwSpWlPun9Ipp
ECuzcFVXrstC2xoWtTMgUwxmYp8JHFrYnZazu6M9iiWsXg25yMDEWOq4UMwoLVpmisu9sw154hryHEwbpso+dwd3j1PwLqvF3QEc
gOdNhSXBoYRhvdQ2gjPoTFj7wAlWJ9YurzbGu8N4O3Q212WHDThZdOBnvgEnCwHPlFW2ZRB7/FwZFHBX8FmV4VsZ37cX/YS3ygjF
3B1Us7Vtgvf42UQXZPy89CzW5WRQWwbdMo6OpCF3Re/xc1vFrrvyT+r433Lj3pXblSTkrnCxdjCANBpxND/bEKMpDdwgI6pu2nWS
Kg5+8IbWwMYwQgOfEw24L8x8gClmi3P8KmwcOihBAUlDeN9SuGpfmB9KDdoeoxCRtNObCAtOIV16Vl4hi3eDT11mZXdJGiGyyIjZ
Cr7HAPaXVHWd4vuH9aCQVx1rau0MSPAgDfmwTFngCUokRMgGCpwOZGOX/Sk/gyeYnl1XcVyBkMumct98h081ihm2THZBRXdXiW4Z
oaGe7LbWmgH6iKYhHPB6GLnuIxDMwrPT+EG3jOEMSYO33kreeowTvr9pr2j64OGJBiqehnWFFiztWuQfUHgYdlgZwNLKQMBS4CA5
Y1rGcMbd0d84SOWc232tDhd8fJXe+/uC92DaY/0mqKyPxZsCEZBmja2WQeyV1HNa8oJ1L0g0McjybnirNV16p1RYD+60riwcd4WL
8AwrAZBxciEApyJBsbDG/SYSb0zSszE6pImexnCdeWJjdNSDtjHvAz14Sj0U4E/RjvKsN7iYTM87wyjaa80oeunJCQeMooMOusu8
/Pima/kYRRvJIA7QBJBJ25pc02KTZJEHTEB9NsUBaR4UdVABWPOgqIMFf4nztsuCyHQaaYg26gyXxtvaao7nyL3zudHTky+jFhvH
C+e2j8s+SoF9dwCHQLq6zsTdEZzBaDaIEltEKUmk15TahiFcpEG0yoeDsK+mfSCD6AQvzgggOgEPnkpAdO2aBrFH0Zh5yAbabeah
48zD2CL9HEjfYPq8vfKCqmBwY/YsRKA+72vPS747EzSXLxsvIzYQLj1hpb/g/HIyWreM471xHRcMd0x8klg0055A7f/UJvokrqO/
Mf6S1y4IyXxBfL01HXI5cEBhZS3gVVl61sCBCJW1XHJNW04E1GmuiL/wVCuv1tiOtedrB6jx1suxnmkasgodlOBPUQdwPVLYE/y8
02nDwkM9xiGu0Rxxs1YBIb4Wlp5ljZLWwCNyGzoVMtdBMckwOKGWpEGXeQ+EcmqW7o9nbZlZ+k2X5LlhOrHLxAAFSGKztEOumy5b
4iC8Qw9qjtiYvYph8hVpSBFcetfwDu9yXoWeQlPvRicWktLxZnUMIz3vVs8hweNP3ziGazY6dm0hGx2nT48/Y+MgTmnooBIsJYh4
gfMLLGfT79bhCCw/fLO6AF5YvlnHn+yHDUPr1hDOTtAzqdySVpB5U9e/f3k/R1W/QuwwXi6xdSoEyjk2VhNMBbG5mgD0WHXXXl/G
NgdPTonjjfeD0XcdrwUkc0aj6yaA69fxk5MV7W38VcYvFywkHLoEvL5zzwK9IxR9pRRaL4y9ZvNSgAKnUgT0XrRuhj0AJw6nJ44H
In+Qa9r6+Xu12lDhq0K+yCDXkJJYMQQR8Bnei1aBjdsAxTQlQHs2UetOFEH5iJ6CWwcztqLO0XHctxCqaspnY/JKk099tKmAzTFF
MYgpimwd8qmXWpfp4ZRKEBvvITZ+7lnD452H8Hhre83Nj3P2UCOzhwa40V2noQiWBIbpGgzfmmR/hGseyhFaZy0bJJdw2sS69ZoP
GXXkR26YqMgZdeQn7og6HDHq0L1W5wi1z0xvM2gfj0pyWOB47lwxu/KGIJiEwrZC1P/58afPv/v468dfPn/98cdffPzPEz3d57//
+NPH7z/++SYJXEP13apRyKU+WrJ/q4ZxUvSDeCzAH008FuKxUI+xCMWPvGgV95jsojiZoH0IcmW6AJXpAiQ6BqhMF6Ay3RakVQ1E
qJbcELJeNYR9XZDKqkR1m6FXDahK6fsKIYbttRAFYDiO1nBpAOOah7DPzH/n9Hej/qiS3qveXZXwjix1VfLF8hSkGaAvjQXzzY2A
PSn3tO/CI/46Ih2Av25prvx1SzNsexP09lKgP/8GtMZp/bwRwsDTyoMpfW4E7Em5p9/89KJJahpFt8p6lUe6a5m9ppmQIQYzg3lg
BvPADOaBGcy/MIO1LY14xNWUlpaDvqE5O19yM2x7E/TWjNA8SktSNYoDpN4QN1M1jDOk3hC6UjWWvnxNVUOQS1dGDGKJPAounh6h
eHp8KZ5eNRABqTeU0K4aAvWinK7bDD/e/E64JOXTWkO1krmxUKDMjYA9Kfc0D+btWu6Crq+lmFbVEHpRKVQJ/wbgXkDt5zxQ+zkP
1H7OA7UfGmud8+1bUQbufnCGef646dSmGba9CXp76dPOta6bhiKtHCULcf9zY/HPzY2APSn39Jufb71Xvw24P0kU2DQTMpc1I44E
iCMB4kiAOFIPxHEK3KMbnOZkzdz6MtfkVtj0Je6rGdw1leA70GEJr+Abng70BnaXqrF8j53kin5wimI07HjJkYbr7wB/X6MeTfNA
TFfC6qoh7DH7w2aZZ1gKK6ULmD1CqYqlsWD2CLUqlkbKPc2D6cYeUyX97XZ2+k64Tg+RzFTJF621Hpi4l8ZirfXAxL00Uu5p34Wi
Lo+Di2ZFodz62hK5FTZ9ift6adGffwOq49TE/l2LdozU33mIj5H6O+/0bzKxUzmtTY4qW38vGNRBlQhnGIU6029p5Md4HEJSDtJB
l+b8GM/NsO1N0FszQvs3YWIv4V9siIusGsYZXIdKuGxJsmxKsmxLsq7DWPpy0VYNQYLrmmka199L6HSmaVx/p/XvzQMR4Pq752IP
1xWUiVJZ46isblTWNR02Qy9Kx0rpe7gedWS39tJYiInnRsCelHuaB7PXcu9ciLfD9XMayEraySrhElx/58dLyC+YwCyjS2Ne/qUR
sCflnvZdKPvMpyepYZ95bs62hNwM294Evb1UqYDZW8i1moYiWtexMGVwYl1Kgqc+Bddvat4dEVPINfne03xAv/Os7aOUhdIojohZ
fy9mwRwRs/5O69/7Lc0RJaUlq4GScmmuoetLM2x7E/RWjLCQn5IRe2TEHhmxR0bssQ6xF5JVQuppZH28TT1VnHraZSRnzJUeXOAe
fOAenOAevOC+0+wILGd8tjwfLq8PMoO7jEKitwpAbxWA3ioAvVUAeqvwQm9VOxaBeZ3xg2UAYRlBWIYQNvaZEfqOcJ0C6ss3uMiK
eDBNchw7szSWSPe5EbAn5Z4e43m3Siwhx4T0VC+mpzI6IN9nFHul/Ly1tYw783kEWUak6YCTfGmsQTXASb40Uu7pskdl5O+m9FNG
/rk5I//cDNveBL0dVbIA/gPbrwLbrwLQh7H9KriuAOEoLD5gWHzAsPiAYfEBw+JD11naB9g8b/i8w8T5DSrh6F3wrL3pFj2nG3zk
rewj72UfeTP7yLvZx64LdXQDBJVtYtBcb4ClGba9CXpxkP/54w+fv/r89efvPv7yi49/GG5ydDboi1LBB7UcCQ+QE0327oBcqFRy
YQ3HyKRCm0B4x4HwVeLPKdn4bkVKNr5btaubbrm6KUHhc80pCDlyZiqvYjkJwYeqPSZxB1VWzSsVKbEGVdKoFK+rQCq35wtC9ykX
/KrbSFI9Rg6B8aAT2aBusK65cds606WSO1Y3KxXZkVGzVKRQwbyy0mapRKm4i4WZDXuhQFPxwlJRKlQs54IV8i3uJItF8pdGyj1V
m0mkDDGDpbiC9KQH4zPJPPd93RU0eJ2riyY70EtR7psK6ed5bd27qVp3pYulyQlLk0Pu+NKxVCanhNy5tm0SBF4foXygkU74C0XH
XdHfcLlcc2LeLIJ+95tl7awZgzko5ql513moPJ5i06yLxVbIjtiXTdY0zn/gZyn3zlauEQlmq9/USuRLx1RAfBlFkMSPGaPxaVUl
+IDw0slkQQwedACyoDhQjexzqst8AJCeFo4A6B5yxefviuTSIzkDfzjXdtPAFzLl5dfIlXlgLVDDc9rw8vcFE+eiQ9PvbZhIqfAz
WnjPhS/hoQpll2yNROrH/lO8sj/eAYbNNQe84zBLUiYBB7yDMEuFhZtfaEtL5e+ViBPqsztw0bpsCa4SKDBWNhFblcrtV+a3VOI5
SO2HjC+5KZ2H4I1l1yxWXL8J3uDMyGi0qdpOcikyn99baXwoBwqQCJnvKfoihs6vaDtqSWVMg3r6Kd7UoBydaLBxbLBpUY5yvW/r
2elDMcF7wXoNJEdQ4siRp6ZZOIWotyjM74p+D0Q1j9Lk3P3mS4iaSISoOalgsiuFpn0vvxQnqgnDEQOJMlbOXV+aOw1eRYdOE12s
WwoYJRtY0UsFH8BTqPrv5ar/cQNPQ43sM3jKMJHg+ieGiRQ3FaF1zQDOL72OVuorvsiporThEOZ86a9/X94iOav7tV5pqWiB4KYh
aLlUaje69eJ1PQGnVgKnnvdSqJtZEZxaYnBqPNRZWHoWcGp0OswBKpUvEdhURc2VCpRoazjw1gnuODa/V33hKfnjDXakUnkSJq3m
gS4VKpPTQFzqtlrbhjF9k0r2koVavI3E+tdxAqLETOkhWAcpf3Pf7EQ0k0LOBhs9MR1q16CYfp67tuzgRNcWVFNuUYsyMA0E1fOM
5WikpWcBpsYA+6bTTeMQClzXUjXdFX12s9wiIy8/eDIsfeJBSOUkM9pDqh4U1dCei2oEiP1VsWnFj1wY2uZK1dGNTyHLHoyl6+u2
jUN0XN9yPJYxlAJTW0J1LpVwsiJKGuF6jeBDX77j9DsDznzHzvwA3vxUI/uc2NwxNQQD0yBTlehiYGq/BZjaa6spsxmNNyzcth4K
TfODzA+efI3wM6vprUSeUoH0boFnd3g/u9olvaJKgEpViKy+lp6lqIVKFgrRxVAjXzJq7FOvSSzmULWLOpKTl4oUPHwMvJ2QXc56
OdbIk8nIuUSGE95TbHmvEnmJSjFbagNK0TRXjEivaRLTeB5sCBxx57gWGvetZWBJqQA56lob16CRfp47tUApoFMLGIt0HJrki3Zr
5qdRKyBf7NbMTjM1AkfIuxBN0zzs7jMPz3onpKy5nLIWWwTvb5VKavPyI9eZ/vDuF0uaeVuf2Rq5PnMECoiUmtZbdlhMlODKUnZm
j8pxjiCZIWHundkgprdjLio++S1CSoW4tIzqMLCC8UAjS2Kqly2FpqUUh5DxpsWMN87aGX/6SuFn1IYOH8dQr9LB8xhyurS31cMQ
CKnYzObBue/B0OZZ346/Q6VoiVTIznv9SyHogbKqX/7+FYIdZ9PqkudnQ+2nCxSG8C4AoIHvAoAaFKs/fQ9bDeNWw8DVQBQLQ1dT
jF0LKAvRP+FE/wTUQnGDrRYsxKQmdtMaEz0khLGXlhwmayevKwcgWEAMBBsHyQ0PMZTTzVspeQ9nF1rSpaw6m0EU20FA7VhTCjUK
SAkJQ5ZEGjOMPaDqj5YA7jZSywmRWkDPaar3uBy3GoB3miD5cOlZeacT0JzoWLvbZbw7KnWd2VZGhBuSyqGruW/Gu1NWN5RwsMOo
jkObmpO4B/nse750PJ/9wHp24nhrHICIebmk7HSXW479gYKyX6ZXsMMGS61zsb8HSbwNSI7lS9VH8pRiEAxyTkwXhpMZq0/mAQ5+
KCzpFqmgHaxl/a8s639rOYU8rkaN8Fr2uGYhjtgEiVRGxOMJTMpASYalc3auTu9CxuhGj2d5W2j9P42D+t3nr28yBzZUKbsSeEA9
0lAm/UriWQk0RtyWDUqWEbflEAlr7okVqp010N1dSXuCgudK5r76QWVu+5UgepegbmV0rgXtyxRYEz04BgGVLo2APSn33JHbraDm
laBuVcWuBO3Seh/6IKl4wEOipORyH2BzLI0l4ybA5lgaKffc2hxiFvD08DQZPEJzxgi5Gba9CXor9MPP916fZ3n975j64/z9ylKA
pQK7lQq/Pj9daS9Kv0/Siw25c8XTelRuK3gboNzW0lyx3dIM294EvReDueaa64y0CmjlWnioryQ+xCB3Jfbxy6mAF46xXQRsFwHb
RcB28RrbXVDAVQZdXgnaY7vKuhyXa/YgtjPfhO3MI+zKV4L22O6hJXtQWZiH6sBeiZKZkF0AJmQHWUNzI2BPyj23Nodc4UUPiTTb
GnJzrvCSm2Hbm6C3Qj/8fO/1eYbt3C4SkS3hm/CDqs/88dyztZDY7KEb4AjMPaQjzquqqgjoIgK8iIAvIkZWVs2tdIC+MnhyZCe3
vu6/3AqbvsR9F+OgR1kNrgT25wi+kvg+MEdvBXN0Deaqi6JdiTMPwit6hLD3crkefCJfUno9Iqhb/f8rQXvg1g8C05uA2yXf1kOi
xFr9EYDb0lhq9UcAbksj5Z5bm0MEbuNlbozJRjluzo6z3Azb3gS9FbpAAm4NheJKpcoEWe3loIq/uhezYKnAJ630RQxXD3+f7Kzi
yn+eSTPX3wH+nta/V82t/BCya3W61ekKxeqgGba9CXovBnNNUdX5KJWwUbFEwxINSzQOI7BuSXyIJ/ZK7OM31BWdlE5gmUtgmUtg
mUtgmUvXljn7Lq/rBT9UP5uIfZdl7pL1icvYccltrmDHOSy31ujJh7h9lynOvgnR2YcYVK9EiRw+nJPC2fYcEcrhoLd2g6RywqjA
QvYFcevr8OZW2PQl7qu4/DtznpZKFV2qxsnVIw0b3zwU5vHGVX3wuy6Wd3hT7YPUpKXf9x367hy4mTCYxNEJ3JwfQrkZtr0Jes8H
U0ZWFPjaDXzvBohn5Zs3XF69pbxEWswX4e0VeX9FfVvoKQURcMt7IJf3wC7vgV7eq9vi97eVY6DjGOk4hjqOsY4LtwVKhfkNEAsZ
IBYyQCxkgFjIvBALFYjd47mWjIwCgXtcZ3j3Gt69BjIxePea+7u3LzNQkcA9ziONpR61Y0btpRGwJ+Wem6K78f0UyBKoffj+J0iQ
Zg1ArALI3Ba4Z/Hp+rYqI+zhS8/wpWe0WHT3/qUnOgmDwuxYZSA7VmF2rIIUjrFxd+vIlok4EOWaHdCcLRO5Gba9CXrrlI0AFx3f
E47vCcf3hON7wlVrOdlR6wMf36WRK2vy8V0aKffUfvuPZ6HyDcqc5+6PQ8LMru6vW0Q4NGdefH0tQRbG8jvA39P699qZlkP01BBT
ZEbM3HRq0wzb3gS9OJ5/+vz7z19NlDz/9PGHjz+N//zx468ffy7KxiCAA55tk4ZrQBODLeKySeO9bhuGsIGaJMTwLVUcy2tJ3hzA
HnYSkCvCTFjO5LNwBfFAXNtA8v1HSE9DuZKZ9tJEnIRP3pTP54MAN2XehSkrzLP/PUYYADy2wraezs0xZHBKUsVHAP5APcolSrRp
WgDaCfc4//IL/qBIxN21//Eq2qJoulMm5bboX76I/qr+yN4uHS0jEOOhZhxZlyBj2bmGYWRFxHfA5hoS/UIH9EE3RWdcS2x4BtFB
KA+p4a35ki56U/iqDXnDo2wtWlUOSvLclAxol9gRB8LlOpUHhtabwjPmgl0XnPNARQ8FSMdLhTFXCJ4Noy0TgBqZp396yaeUAbCl
gTJBnh41gtM2418/nn9KOUQxjPe166AMfwqqYHMasIxdUdZs5TjEVdKW4HWi2Y3hrUl5laLR7MxIL6HFtdOyuyUN3pLmju+4cgj7
a6rS4nr7sG6ws6wmZM/IQWHZygmQUMJ4YgzvzVFpRsYJOZt5xQyLjUXbHgdFPsIjDIkcrZImYcrDEzZZh4w93mUTrZs2VKB7IzOX
kHoyQnucISGQERGtu7tXTAGk3tztSC1pD/zvvmEMZ6iaNnuWT42FynqMbmxomgzhKt2cGXxxFtmRbsqXzsuIIzQPIPK7c8LbDKsd
BBFbFRsGIcBqtZkEQNYKQQ5Q/bTI3yPrymrYdxd/pzDcFawOEFVgmkTvYTVFD3WACMreLYNajEraRHBFKtcwDAlWiyqzC74057B6
85yVrUtc175FtKAm8SEvhgdXv6bMFaj2p+Uue204Cas5C0GLRkHOCe43j0GLwW5Dy+7uN1EhT0VFUlxXPo4KwHDegpvKuuasH6NH
hMkwNgw+hXZd+FM0rqBODjdzUyvHIa5SQFszr5HxsEYmwZ2QdOoxJ+dGh85o0pTB6e07R6hISzAE23RSZURNWw3R1exgChH1qPAN
BDlYyDIynGSUd8TEtWK7LIcMp8OQguJg1Yl0M2U8PWKWqHV+Eps4BJWpBL4IMsnfRJJ0Dajj9eHFGmv67lahWhs15pmDJr1tLaFC
NE2bmSArUbZbfgVb1zQTdwwTBg0Tig0TLZtBtlE74HC3YKN2AKY3GXm6ZQwimOYgSNcZSNM5kH7SSkrnWDoKnsIohn76Jrl7IO3J
poMM5tkgtVgKlbGQCKMbRiHZfUBPSMwiGP3WMgGSefodEy9oxs1Wc0I9fAbRLbeMBKI3DzfXF8HTFY62zAA9NpyzRxnzwNvq7N2H
GxUAaTOOJUeHpInrz3JwRhp0chaK9VEkjWlYyYV2NShhab9RAjLNUUQsHTuMQ8TSBl48Fl88eDF4LG3gX4ps1s7KjxIbR0FCaqX8
/Q1Fp8G5DTiavtsyTaWW6TQE4EEKDBIBRjvLKJGs7bEUR2c3xvwgj2nKisSzG5PKkR4TSLHMJG/8ZC+/eXbtNYrevHUAumgxyuj2
C9gWgGiDxZqcUPqTPY7G3tVnthREb82CbJImwFFxAJ6k2DAQ0SStuFi1EyOgoB5ti3D5lBjw3wBf2PTehKoWQBdmg2kYxB5Cp+1O
VOWEjDdFS6bo52yD9hxBBxnURDHOQA+mSfgeRjudLJDiJK4V7rVDGB3APuhJNwxD0lLPhRzYcxxdmbt2U+5OMzr8Ytc3qsheG6M3
W27Pe4MRni3XrYTOiCJQzLoIcUVYwICSAYrZEbm1bDhJF/s46Miz8FUOf73z3LghZ9bShYPDzdbpr40w7k+vXbsWlDD0RgehGnBi
sZvYdB2cYuiYEtQ6gFVyVvMq+Qh8FSF0mZUf/+7QdyLZNlgl6zaVfNseXUACd/uwFuBoyR7tIIuwx5aQEQKWfyR2V+fwDqPmsI+V
7/sukLY37NF2fMmCPXo8sDZXOJywsmfD34jltKYcnzriuYmL4tbQTjPy2CazcaiHIObH8V6Jd6HEeY4eD2OjXCLvl8gKNbKbNd41
ERVk7YFVAHMYCIJe4H0RgIA36NZJ2V+rLm4UHhvqt2kkjtNIGocgHJ/AaUthPM45y8DlzP/xZ67EGcbHHjWuyh5gW7XBWfzY27IP
ABVP4zzsgfaxx3WTiAfEg627YR8v6MTawFFMbKDWsyFhbktAxz3FmgEdt+XKrM6Cv9clm9pGIpivlQB+8eGv63XrceYgo34o+SM/
t1mt0V29dphISBuiYDRYSglWAbeibRuBBMXNxoUA1bExu5DNdLejAo+TDRHsObZsO4y53oSIOA9Yz8XG7XBgHxsvBG0SqHVnmJxn
7LUmYCpU0Jx3Mnkho7JdNKqAzf0maMozDPS8YT2vlKdOyv0oXAScX94nTA+FJYvO4A2iO03Ojwvjv5EeLlhynVq38TFY31iRYm9X
7FVCI84IHmxggPMHdUt9nxmRsMdmdcZWrikQBiCaynF3489OgzkA76MysVACesq2cATB2WbUg/mU2wmuW2BDDtpvdMF///j9x7+N
IxrH9vl3v/j4nz5/9fnbz1+Pf/zTx/9XmPho8328sH0uPFT57jFMjGwYKdJLjabKkWxgvAWzeOar3ioGx6HauWCNCduC0ZVD2UP5
PB7LpjLL3j3LE+M40cH5LhMjWEg2UP4AP/ICmT4LdJAWqXEgXD4wbPgQogEaPRM7jEYIPtlaKLQYgGMxnL3HAdqje7c1o+8dnh6C
e7tsECEb6KyMLQRBDbGL/D2611BoSiqkq9FmRh0GcZk26cpToiqHsAf3jckFleO4yqR05fWpKkdwmVFp7uTxVg5CztnTHKnkvYLC
wRQ1ZFY6zqwMacs3XbtBRRRgp3vSxGzsToMBZiJk/3Zp8CkxObGZvFbUT/H+7J523zgeEfJ7xX4Tm9BRpxwEiVsLYS0+2p7T9OPU
b2oExx1f+LHrDEl3njqx0Pe+8I5gf30oZuOEyBHk484A24Jhlz7b7PWgIUI6+Z6LJKP/CRN5B4xkXhEzko3AQOdM71GXTvE4bLsf
1afSdUM01+BfZ3IWo5gWGRG3AcQ94vMOIzkA//gOsTAWx2MJWJEtdhjKGfgP/BixfBtaRt028dSMv3tMzeW1fMQGwt6ELvMie8NS
wPxNB0Uoc3mMyYgXebu8mPQrR9M9kbNyHHv4H6DSqStP7KvdHSeOYimGJED0kO8iX8jutIljT41iHW0IdLS1GsxyLwyLlaORgtOf
h76XyZ4wBC8keyYsoWI6jKG0lEpBiH7lCPoGrlcOQgxgtwZKTacI9BXWQvnx6Dnwxmlje2xO2dg/7ggTLTNZJZOrnk2Wf50CVj2L
nBI6GTBSdP20r/QC2NZrEEuGYnpZnzN09gLwI1Zio/8chrMY/eNhKq/2oec0nSfTSyHeHmLJTMeh7C88exVF6CCmznY56ge2/3ff
OedhO9tCE8jbGjw/ZY0HO2XPQcmHn76s/Dkwn8a3h8+pGlPKaIq53PFUgMkqx9HPcUrAqTQA0OUzQAP21oax9wIsLwq+Vo7k4Bmg
YChp8ySB2rNwB9kOQzl5BpDPLyRimEleDB+hLqMRM422xVuc8AiAGJLUZRyyD0BtSvrEyE4+Lo44ZedCccRaHwDdiJ6P4hNAYyC5
7zCKsyeAE9MS4Q3SZXPsPdFvQL2XuamjouagehWJKwdrHTjAR1vD7wDtfOownJ7FXiqHIDoD1GmG0WZn9liV85IO0ixYyMUzHUYg
vQMqKx5VjkCMvk/gBSBiCEnRWPmBOmpp12NjSjAgTCA7p+5PIflmPSvRDIniOqY0bk1Kq6UyjdAlJNdP7cqx+AchLSSZmozqsmvO
Y/IVxORH4B/cxuQ7BxnvhkLPaTrxt8qXbd+r5awE45vv/UIPwCOuCCqvyugsz4obNIdlYfHmTBI9eQNM14WSXQBu3L8Jq8loDS4A
PQTSEapBpehz2UgzXhKUKsdor30Ahn0AGpNRHSejQjSm7TCOw/AfJ5BOEEb/5B3thx4Tcmb/56BpC2FREP0zPsrY/h9r7f+2thp6
Ufh45TgOfGtYnQlca8mDa83wwTLadxiLBPwx20MMeOwB/O0V8BfLZRdQNdZujh+nQ3ByxW6IZ7ddBiFE93vNJaxNMsDxGAjUNHmO
7g9kexwYySZ2xqp7WKylUr6A/I1Yul+LBatdl41RWs2tILu4cgRyzq24Ob3gnopd1kIO9CcojRKh+ChzILoUGPuPWLLHdSGjgFHD
LeXC10AA40K2Tk6Wt8g0euMpTswyROPFHoPrp3xlB4BAGIa13RSi/9hxMKL7xni+VmhOz12uFWW5Mr51UO/KRt1zin7048ZoHErP
EgLVh/z7IoBseQRQiIz/05BTZNfIhAWn5MTD8UyS6bpS0tFPU72+tDqz0ggoiYtJTm5BxVk+4+PAueCBX8zGUGcpO83izVA3aQ51
iZaxbuKTnyDERNU+J8+zedkboRSPJ7GHLcV8aWsFiQBTGSHqM6CT94BO/DzRIfIs6cjjMmrxnswIWPca1/6uDhs3ZJCZSdkYEWtd
EyWZvuP/PPGjOyQupTPlX68GrXGaaN3UUw03Z/oMSMj63dzSFgAQsJhybpbTlbdSQfYvkn4Cp5hGrkhIve20RoLRSh1lXRqxcqvp
NSUi0VLAgigmBSBa0glYQhOkDbhOcyMY0jZkW04k21JMttXp7tu/HOwmM5Yg5pC3i+XgApt6DWXvzt+kzcsOjM2c+D4DEfOF4T2V
xILDm2ThTiORqywCQ6lNHt4S2ltIYjeaw4lc6DSiA7uiGYzKYfKTWpyLX+bSmC4GKI2pbMq4Yqo5H63uqsmFR0XY8h+wOo9Q0YZQ
UcW+Q5JTC7TBwCIgOnUE9QiCx3oEKnSerP01ba54d+GO7qYuTnwMJF6NWizR5/tdAwdRRkfXkhILOJje8yNBHzv49fTFEdSse9sN
zjPuSQB7VO9Fky0NegRYlhlWA5fT11OWb2I6tymqPkW+FkYwHbe8O//4NZ7ff/5m/Gcc2+evxva/FmUZR0g04CzBKQ4LLnIx1eDF
IXxvCJt3RYTDxWNY44wWDCS7GEz9EPYviQhUUHkYSARlD4igUv0o8nUcwcnruNpagKoFjl28BtwJVC+dz0sEd8Yc1bp4M3IlmIlu
AeKHEocPUcMq5LdBBLDHmwDiyzTvQ+RWtS3fTzvp4/8uv2xh9jFg07D9yDcs/I+9aP5uOi1M3Cj4ly+CJ1QPtTt0NAycjOM8THKQ
hzkVDKofRFZkvOxD5NICYkQQ+8obzlsG9RFcNHnaHa84RoiBIWOL6O/JXtVlhDCo/M1yFJRhu1zDkgNqh90Ww363YVq84dyYhiOW
wR1sN5d8gOo+hnNSXHIG3VFQydFuC7ze3G+ggyOm/S0VNtfSPtay1Wi6ZXyywCCTHJOqTnohtQxpB8gjlF0GVYw6YFN52YCTitrH
IS0UmPe9BteMN8lAHV4I7o/adhiLdDnyKTXSvbzZr6l9CNINxacV/IaMTPC0qpbTugHXcFvgFEiV8znYoceGkMEBQ9ZxxjWEoQM2
0OEg2KFyKcTD+3UedT68GrJjJ19cIpvt9eNZ9Tpmygk3nhev7t2n5hJAo28O9AgyP2l8GlK9fBk9T5PNKtRhHArkCBgYgqsfwgl6
poCqBXz8bNu1GuGzqR+GCJ+9nCohl9BLLQshHZEp79ZBDR5YBgO5LU6zn5YaZkAA0JCgkcALAvhZIyO3qhe+x88OLkm52k71NW3O
4TM/mza43QmpttQid4+eCfjhDQUH9KlQxcQ6Aqtj1A2D2OsmI6JnjKoxTqZ1uSd5D59TnvaCOK+Gb96pQxLB88EnN1z2Enh2ANCs
CJ659nXDN4vYWQcLVEIROJ8C8Nz7HD874ehtxuzNzSZq38mNOsJhZo3xIQ9lckWnyDUxaULZTBoTJoO4adZ+EnAOCJw3dOZOoHuK
HXSwCJttgFwbSpBrY6PnG8GZCDn3LuoOc7K7FvGYaL4YOWZ3c05S+xCky0l85bp23GpKcDMdvLKdcFhV+/dLmCBNxRbYtsq4OXJd
vrUCzhKam7ztsBbS2U0jTpz35BIMEChXCknTcy5ly/OU1B75zWvcVCbn3jala8uzwRI3Uk1J4Fp6SZy8J/7A6qwGdx5UB7an8Yap
l39uclbsV2b8wgDGcm0de/dZRXcwM1Sd3ticoS4s1Us/MDk7uLYJnpUcbjnzpC2I+SVj9d4QRMgM+GnjRwObswHQ3DADktFZsjnL
cCI1LLxgVRFsKg/IFSzOMbAJUCfPMIaUY/8uxQihoy8Vvu+N4tygc1USoWG9JZMzwxQ47OE6MeGe5J0q1HjORRNWly+WIbM30sOQ
32fG9fhm0YxplZeY7GdHx1qiRHMkQdxmKN7caLK5KgxWoanZaZ3L85rMHTxfNTRE5dm8O56+lNr13k/x5j80NR+4G1X7OMR3DRJQ
WG1hmUxwsEwKKsm4DnNyYkgwwtMO6ddSu/hTtLzB6x3MvFRmZo5PwXUqNzMjCUayggt6MpsBIPDBd1gM2Us0nZH86J6Org4mH93J
1mE5hGs85oE4KWTKUaSbXiJ7jZixovNBqOiGnzTVD+DQ1qykW0MuR2NuwlVbammO6DLnDAuKaGkOYGpW9eO4hs0HsRpHfAr3xMum
5rkE8cp/ZvhZifSowH+mla0fwh43J5iAIHBVrqboBTQ3nIM9aObKijDzXrT3tqz6Tj0EQI0Oakq6rCCCO4i7vSt6D5udIrAsGQIt
qSNnXIE/tmEAkm6SXZCibmiYdcnMzJIhOC2hTwMMzQ0XnqAQrxRyrWndXoPmCN6U4E6S2e4aM+0VaLYeSAMo4U4jAwG4HKIxFf1u
2G2ynXkyB2Rv1nivQT328UbxzjCn4rT7OHqaJko7S82aT0bNQQTN/uDeb9e/sp3ZKlgiUpB/6YkxM4eUT36BFg1wBpqlm8HBrVgf
RGVLgfM5cMVsS2o5p9dmZi1HDqoe98QZbh5vRgO+Z2c4CAHtzAbszCZ22A+ynXmCwn4NFwlTIpFZMyqTHwVH8BBFpy1kUgflbr3t
TtMo84giXOVRIwepEzlIVdMQZOQc2GEawOgWYAwQpBHTPehckC/Ju8UbqOToNvcY2D09eMxCaJyT/V3qMQjSccqAZzucByrheO8u
PUmO5BWh2SGz7NIc5jYuz2oeWulRFyoyFdsWZQ+kLYa/UoDrg6M2rNtUs7FNQ9jjaXSDeDl+CfK4bsKN4wTIKDAdHDxp6yP/j5Me
N2EcqE21Y21KDqKgrVEQBW19aBrIXo1FvqMun1S68TjucfYIomAjQlwF309IgXozsuIwsRE9MRqAvvD9m9iltstAgtwH2xBs5AQJ
5L7x+0VLqMOa2spBhIdPUKxHBXaNeBPbpuLA9DW+clJW4VPdAzOXCJrxt53cwZi9aIPjOI8p8VGHHir0p6g2wIfCbmPPgQ2oNlzj
dXVquTYaKti4AMSnColPnQMWJG1dn5nZ3aNT7Ty0Ssm2OXOUkVs9kv1NlgBtDeKLPUAaj2o8yjIq10YOhoTccQfDiJ32iQwzcnk2
P85CYJiRndwjzshvtQln2NhnaeSn9VRgHShP9RSkkg+3shPtsYGSJ0FxwcPpsaM3N85/Hgfzl3FIf5iGVpR6qPOVy8ZkEyX2cxOR
VkjdF7uB4iyXK0MbVrWG3bBIZxRq5O7RdxbuuHqhA/oN4N+A96KrEZ4vTS2FqAO7ehF9XJlM3vcaC1sYju8hKEvIpETLbzEeukxy
xtRaqhqhD7jkj0o8l8mknUzeyVGoNX7o1i9c0B+v4ui0lN5hVddScb98ETdVBMdEQdR1SyNgT8o996VndaL3ZX3NRX5gxcdmJKz3
/gW5ithB0GCZuFVPvefjAOe+5/MyTIKt43HreNw6HreOx63j67YOqDu+DaZX9UzpPZttc3POGMrNsO1N0FuteH4KdzGhB1UJhbz5
+GpqUHo/pcWICsKvlsa8GEsjYE/KPQ0zsLu8WPsGgYYrAEdPw8fvrxHQCkLMF6iEupO2gaF6X/PU3in2ee9bBeU7afEIKj+Czodi
24zspt8Nsy2evCnbS6UcQMFNpzbNsO1N0Fs0JHONNHkdDC+EsQdhE+q+WBlpYslZIzKhHxewKpN7gjTJY2ixQJBDfNuTrxEu3G5R
KlS6rYitDmLKymRKmz0Ombdh/b0mQEIYMZsopt/3Je+RZgIqkb3FCiPGKj6U3ituf0/3xrTmCmQu3H9ripO1B6yATMI7N+5L32sH
J5AFuWsXXpm4PcjUDpOGIe4erBPqoBZ1mczHkaa5QpqP7h8J3Lxx/4j6zkz1NVPmkeOmU5tm2PYm6K1WPBLS5Psfc+IgHZrvf+0b
lJ6INCGIxDgoK5OXwXAAiXEN3/3ee+sYX/aGeuY78KUpx5esctMm4IBVbgKVm6pU7im+9HryO2TTam4tXuSlFTZ9ifuKRkPX0JL5
xDCLbVOaXbGJS90XewAtnzjd9J3Qki6g5eMfLO3zMPCFvvxeI9I9h3zxPT79vi/5eWhJ59Cy95uczqHlo18nQUuHQY1LY/Wtgk5a
Gin33Jf+PLSkc2j56NzuFFBvSzS9235JV6jSSRlkvGN4u1TsFRFGTvUyXPTARbI0V0aypRm2vQl6q5WMBCOfsGFQKYx8cPbfZ6Gk
MgT56EVxhCAfFXqGIB1YKB1YKB1YKB1YKF2phZLKLZTKDxqLdOfmyv6xNMO2N0Fv0ZDs91go7TWMJN7rFKSSXcTbnUKN3BMY6dkm
79kojxGgbJb3uka4cJ3xB+uAlWKUGMFfIVP2hYMrHDzh4Ag3B+ytZULf5Aa35zCyt3K25zCy991lr2CkSVYDn57lAqZLI2BPyj33
pT/vv7LnMLK3ZcG+1Thpr2Dko1tHBDIOqFWWxkpsABl8SyPXhqrQLweoMg5Ra2a4y835VsjNsO1N0Futc36+KSTJFrvBfWRL8dJY
3OA+sqV4aaTc0zADu8vr0QP2Pje4LQGZvV+JthBkGjOE9SG1/l6yC4LnKk8BKOeCb5ht8eT5qZ50LpkHzdl+lZth25ugt2RIp6lQ
bB1gFZ1YRydW0om1dCpU0+cZUCyaL/nEt3ziaz7xPZ9MpegTtDnRwbFzGr3T6J5G/3SsHMT+ugtsJQ5sJQ5sJQ5sJQ6qUqxwAAI8
ZeEtC49ZeM2GSrl77Gn5cy1/ruXPtfy5tvZz9/CTWCyxWGKxxGKpVuz+JjcMJgyjCePEDCZXLXYPR8kFhqNLY81cCsjfEDTQtQVd
NQBBlTwQynacqsSrDETSFhOV8iqz8iRbKXanwjqbVc9ykt67sUSwylzIubGAVQvFZpdGyj11G0t2ptvBBw7b5OZsfczNsO1N0Nui
uwTI6vl2wcQjvl083y5etalOEbV6DfQZS2OJpNU+YAlgTDjyoW0e9tedR+bMvCsZQhiPCZtt8yDcOg/YNa7Sih51QhYkE7E+90OO
cVl/f2n05XeAv6f1720LIJ/Or+TYbDjl5lrWbmmGbW+CXhzV//Hxh49/Gf/558/fTVlOv/j4Xz/+9Pm7jz9//Nv4h/94L6tIs3rA
kqYbChV1kH5WOY4DD32DeqwcyAn6bUl6qhxNX19S5SDkeFEIF4VoUQgWhVhR02EU3Rz8lfIFrFz3JK/dCr1st9XyhXBTVKcW1alF
dWpRndoXdVo5nG5mmUr5e1Bd6fKslN/LC1kpXsLalea4yhFI8A5ZpjZsIlAGlaDsFm3LoNZuRUm/j7AhGdIMNJbWAjWWVtj0Je7r
p1J//k3okDNQbl0kqAIQmSxtaQTsyQWtqecc/fjey/0Yq1e6daqPtQza33u3ncU8RE6JfqnbSYA9CMAHmZ4LJJ51NaS5uPJinVxa
i4FyaYVNX+K+uuEVpGzpK8JtKDRbrRUfzuGqHMhZ5C0cbTjbcLjhdFOX0QhXcEP2SuUgpNPkGcZ7hvGeYbxnGO9rYbz5ZhhvHom4
qN0K74fxl1ljRERQDBPYSpdGwJ6UezoM5xtgvPlmGH+eY8YWUjaQsn2UzaOqi/iOcR2VIxBt5x5Jjj08K5dGwJ6Ue3rsx6O0NG82
aWlLc01LW5ph25ugt59y7RwK0jgYaf10tAY40a3hUidzI2BPyj095+jH917zx4D+vfdL56y3xtmQAb23DOgzmxHyu62/0/r3ngsk
5sXZgUIu88Gtrz2TW2HTl7ivbnjXiXKalYMGGm2FFS7VQanoynEclP+CV6kRC44apLDoMZAzQM+jwbKbQO/No6E+oxGu4AafWuUg
nqjjUDkU0ysdrlL+HtU/fsHRI6UfquXvUX0wgVH90pi18NII2JNyT4fh7LXe47CWHilJVin//eqWvt04f5kFqJLiiJilsWyDuRGw
J+WeHvtRUvKOJg6gVclz6+u9k1th05e4r59y7Vx4onEw0tKFqBnQL43lKpkbAXtS7uk5R7sb9r1XSvdw8Oqz3TUJsXE25ESt6moO
rWsjp20AakbYjLgZgXM9crZ/I5E1BUmL71ml0izGlnjUytF8l0n+Ks9xYsazPAx+Dzt4Dzt4D7vq97D9Zru8/Wa7/EVq5LvlC+E1
7BVnlzj7w9kZ3kF+t2pMlfL3kP3xF9Qj+ZSV4jvW6agcgViBDA3xBg3xBg3xBg3xptoQf52CGdPgdM4f5pZT2AqbvsR9/bTptwXV
2EerfrROyjffod8YRfNkDmfjbPwNgYyLiPgwxBCy+ZKb80HKzbDtTdBbNcaytE/HeMgxIHKMiBxDIleLiQrzQD1bVz2bVz3bVz0b
WH3oNZazxFDIaNGQ0qIhp0VDUos2vteo9jex43Pm+KA5PmmOj5qjXuMQjpmPg1udxutvp/h3gL+n9e99RrOH8295YxVkmL7FV12S
ctpQLqhlHALMdyZC/KuJEP9qIsS/mgjxryb2GdFeWT6ttAsSVFt4jerHsc9Ve4dTryiP9XtOjRhjTzkQJjfWYLEcBJMbKfd02q3i
s2DcifmizY2vR8HaCNiTck9XJS08Cr4VNMjB9t5zwvvSWImvPVB1zo2UezrP1I9vsb7dyZF9x3OuPGn2W47/WRatGTxzMXsm5FsP
2/IzLX/tvFzSLZDMYKJ3XB9laS0lUpZW2PQl7nulC/zzx//4/O3nrz7+75uMfAw+DYNPQ0gFnDUX7akbr8UeVCFsCIIpk/sQI1+Z
8L5VCMtkygHzuQbZ+nvxzDIn+/I7rX+/L7lvKcIymd0M8IULugfkdTaaUnECI58OjGSWxhKmqgMjmaWRcs996d0MUmXiupUiLBO3
11Z1HvEyaR2DW8oEynVEYOt43Doet47HrePrto7MyOcH47ggGjfn+gi5Gba9CXqrFU/ncJV7ssVShJD1GyHrN0LWb4Ss3xhcw8f3
yt+5J7Zb2EnpITvAoU8KPY8Sr643cXOixfMW3BBNTmqA5ow5cjNsexP0Fg3pOqmz5Z1UJvYAX/LSE6898eITrz7pGrkn+NLy/WL5
frF8v1i+X6yvES7caQ2hOmUyZVIgdswkcMwkcMwkcMykF8dMmWQBXzY8d8tk7vFlbwxkHonnKBUn2HVNZJCwNBZTi4kMEpZGyj33
pXeLty4Tt8eXvXWgeYIxpUxax5TIMoGi9z5oDv5ZGosPf24E7Em5p2LrHBQvoWADFy9ZWkvxkqUVNn2J+6p1jgAtWwIV78kWcT5w
HiHlETIeIeHRC9/RzY9/75V1DC0fPl8ytHxUG5zFK2tWuevvpU5GVrnr77T+vWG2ZZ5nGt9kXNiEmzPmyM2w7U3QWzQkejRCuUxs
/5TCMrkPkfGVCX+T/bKAjI/tlwHslwHslwHsl6HUfnmRDvjIs4Xeiy/pvfZL+lb75TOMfGXiBLLnJ8zf9EQ9vTJp7zBiXibk9aDi
Kdw/IsiM4/8trbcRt75AZm6FTV/ivmrF8y77ZTlBXyLgtZkbC9ifGwF7Uu5pmIFeVCr3xHbLMSg9ZO9Bmjf4+vKFvf5etGC+rtff
af17w2zLnoM07R32HOTm7DnIzbDtTdBbNKSCXLgnDt3DqW9lct+NNO1ztM9lMmVLpgVLpgVLpgVLpgVLprX3JQtIs6FISpnM5z3l
F5lpvb0wl4loCu0gCg0hCi0hCk0hqtQWYt/rKbffgDTtW5GmfTfSvMwjsx6Y+5bGGnQIzH1LI+Weiv0jxoeagZRddzC3viJEcyts
+hL3VSseyZzZUJP1nmwxdjcC2VEErqMIVEcRmI6ibvj4H2+9JruTQpeer/eATFtefoHNmQrMmQrMmQrMmarUnFmeyZX8EFMyTJK3
tBaevKUVNn2J+0pGU5az5TmNwHMagec0As9pBD6WrURhipZWYPFSYPJSYPNSYPRSrlL+WVpWAAKVAAwqAShUAnCoBKochMD1xFrG
s5rxrGc8KxpvK8VKkck0ZGbS9feXSWH5HeDvaf17lXCBtI/3m+X9Znm/Wd5vtna/CSlVDwD8ggyqJ6BSUcKUDporSS2NxfAZNFeS
Whop91QNYK9PKgtjFEsU+PoeCL4pyH56wBVUluz0gGukKLfJRgs1N6KFshvRQuWNaKH4RrR118dREoNKGjhekwaW16SB5zVpYHpN
ukV3CUAVNAcoDtAboDaqtcap610ZoOqbG4vzfW4E7Em5p20efrwDsN9JQ2pgQLtxBruyftz+bJmlzwBLnwGWPgMsfQZY+oxtm3vZ
XBqGYDKFZm7Ni7C2wqYvcR8O6B+moRTlEhH4xKE0P26/TZay4yzlWC5xA1sJrl7PRV/GHeX2t+84EBBpykXukSoJdMfEYNkaMcIz
bisenQvNVxv1gEvnsngjE0Qu68gTGgbyHL0cwRJKhzGd5zIzCiXpzjC8gEV6+1wW7WTR6TOX8kaxt5bsx6sYeMfxBxUBzStJv3yR
hJVzrPJc6dJjOoHHfAL/klBwLjLf8bz38ZTLFjmexhsnPONIAsQOohJPpcI4X1e1N1btQY1RZOdSAC7S3qzk8hehWclls9KNycuQ
RN4ZiA2hrJJ2HGrhzJ1dAdqHl+tLAWSbnR5vD+1iAo7Y5HMl+qmkSiSjswaaOGK35FJFCuGncGHajVKwvG+K4nfKxEoTbr0Gg6UC
Zjg6JALx0dd89O7K8XhUeGMVVKArk7i/BzbaXve7Tl/gHRxQRBf5C5vvnT2sQ21oHX/jiCk4iQYq1Kchp88Y9VITp3B2pfOUaNC5
vlSIQ4gZzKVxKBxhqcYBJJ398CMIGA/a+YybSzQ3MVfAKXLSw/YOtDLXaG6KVQWRfOFjPjoASDfc+MhzNAdbi/jCAEe3tQjoYrlc
4X4ym880B+8kx++kG+KkXTwuWOTbMCd2jH9H17YHE3vy5RL3iM7h91nBdxLyx8U7S7jHc0Fw0QTxyrvxQfvLNQgqO4omtFtiBDBH
ATyPxmg5kcIHhSQIKZRLvbzHnXCPM/a58X17PLe5UYI4k45DHW5sDEFXgCRqwqjmGs+BLC9aHly2PNyQJUI6DdWMLOTZWG8JQu8Q
7ocLuG+ugd0ExDUFqFXiQgIP0TgsB5UObciGDm0nh902CLNINUi4zh3hOn9gavX3xYqzbolxnVNwKr12cCqNMpBaQzUf/aPkhkNQ
xxdpvC9Ougxgihlk1TyETAmoq/Q7l32ejOhUlK0bHNSYBg3WDfI1Mys/kNSofW2OZhxxm7FgoBvxXsrvBE3jdemIX7d6SCac72i6
ttGBEyzyXa/Eej9XVyMVGeiUYJ1zWAnKcZ6MKZd3gucoisWFWKhlq4ENFzuLLq1zBzAZLqWI1a5ufKO0gUekplldx4E4UlF7jpiA
/C9tbblIwT63hauMEDDdmtAMc0PcHtJB3c+mNx6VQjoSrDAYdGnufI8E6qzb8EsCHQ2rbA/1Pm6c8YsLXEZzpa9yOkdzCSVFwbWM
Zs5bR1vQEt0gKl3BOQfRhk4Iv3I5/OrG/hOtRYl4W1hrOCxgvPXRWmSBNjYafWN3iEpITQZ75zjcyWt+Ek62sxBzRP34xNYxsZFO
j6N2dFspyGBO8hQ6EcnFO7vn1EJnE4diUIocGWyROnpcGrCLxljxwT/299rFk8SDg7Licx992FEJlttYb7iQP/ie8+GhO1fPGZzz
A0XUvoG1L9vnAuC58TGma2ZXPkpmvK0VcSnXYHI5ZK3Gq1zlIo6TNTwFlTNVzbinfTqfBXttpEubNeY3tY6YH+IOMpHPJRYa6YTy
nggir4x0ttRIpzEVJT9M+JltCUHdDaGXoE72ut54adprC11QR15XFaFoFPAKE90QKsA6jYknspfSsZcylsuic4OgE4OpS71e9grW
bYwxTrAJ5kKqt/aIhOs8YXh6gvj0yGF9TvkERU/Cjc+TfK+iZdWK4eE3tseV7xV3iBGdr+bOuj0J8Ow1wNvM4R7jebDX3dgiIt4g
DtAjnyCRZPbVCwF6ji70ky2CeMrn2sJajYeak5W1ioMOpPMxH7+Xywvr6X4n62+rCQnjRayiIIXmKAxGi/dlitY6BXni4wJEyfON
lYi8tlTxuYILVp3Au4hpfPa+vItADCMELSCijHcOzZWtTvb2Vp3Qc1sdbbSii1BJMUIpcQNZmb5mbmWrdxwBHPOTq3G+nWdFOV5+
LmeGTPEL4z8Rzd6BTif9NCeEfSabeziwSRZiWyMvdLpwFp6ngxAUO8hmZz65gcFBZPQzvSBviTwBedptAS1FCOS2chJIujhOx8kf
fHJ5a3ve2x5OEmNLf3eGhZ09LiM8W8Zp1ZFL0uc9NaV65KL0drDm3mfuIZ/dqFnLV4YlRNHZz31h4T9O9cD3CQJpsDhrMYhA35zb
/SVsNhKxNhMkrbMVZ3xVmJsSBSAYHXCMWA79thpCv10MEPaTTLgld3/9Y76DEfMdPL5Q7n3mHhDi+wuShZT4/nI3r4S9LeEQfuqD
Gp/3BErA0GwjjPiqNYwNjcOX7b3jIQLECK5Fa1WQU42djmCRsvcOyYF2G2cv5JyFCSMmMg6cutqwL2pyj1GKQIeWVAg1ikaAiX4z
716J2XcBMx+rVJxsEIwpoEOdDjzqDqsmKmfrPn1/W1ksI5f3GyBk/m4Tbu63Y+SotzYcJUYgbaKu754uGUEmyNDYR39sWM1v3lcn
MHJUr8awTSBmdmKfNro252a4cVfU7bCD11gcvOX8Yjek6LPXTvlBzybE+UqbKlFkPTFZCUcdsjH5/9Pnrz7+8PGXz998/LUoPcOA
UcuxRctA8IRjhZEDR6YgwrtiN1jSQIq3YwAQwOfsuFZiPl7TJXxX7h5QZuEUpEQNsBoSR2+NuszcFZ1vMwNny+VNHkBtOQ49zl87
PWruiuQtbiCFItduHk9RCnycc+XmzeLGyeZwV3AGlIYDulw26hk+0GzzBpE0hLsCaScQ8zek2BT2L8TbK/njVRifFzgufFpgPivO
SsaRBsrkkGHvlA9ZGeWeuV4OJXZnjMrI+7uys0LgM8pHFE5o/lTiHVuxjBlLGtYDuboJz2zifaNg3+j7J2TVO6YxyqhEFgBJYeM4
nk3eOIFnM9yfzYxiDKOY4AHFOAMkh0vPUmhJpVwcbdCObm9a1G+8b0e0n0hnO8l40n3iUq3cO7/c46jvKMeqp/ECVKQrNc1P4fpV
sLcCZAxxgCJcvxXKZocmYR2iIg+lESNHeiw96zpEWAcy9Z+/u7Mqc2DvCN3fHpUFvsqO1wZJvufC2oNJg2ELEbIVIwQuQEViF4Hd
MFbPtHjc/KgBQsxP48nXqtm4z72zcTSNDzcfs//XjLOlSwZkLvHkFtcZGdcFyA/Rd8XKeBLzYY0YA5zQoOTvSj1Dk3yzEyAe4rud
An8yxSLVZS7wpBfNPF50orzEqpVIlDb5+PSzHO/nDdiWLMPJnDEyhTZoc1fwHk5qI+ECbeQHSiq6u80VoMwLCusJ9wivZhHGM+eQ
ksUZvrZYnOVryxZdW+YcUk45lRB7aDSUslt6Vg8k0B+7oG7fEHul4PKnWv5UB+jZMHq+/al7SKl5ajXMrebJ1TC7umJ6T1Cl5LFm
37gZbk+nhCotTKjhyFjeO4H3zu2tKqMZA2kPAQrwLT2LdjeewYxXty8gWctNZbNtrqkxRx7mkt/QO+uZ8YtTdr4ZN6pESrFS3cig
cg+rMAud0FZWLVhcBGALFMkCN/AquOqP3t1VCLGMBLFgv1WLPXuKGkksQor72/waTj4g9AxOjkeGI6jBWKT50tKOE9XrN7V8ypyf
AhjzOo9vNZOyDwJ6Z+UbB0sKiQIpFL1d6BpL8uVm4HYzfL0Zu8GS6a7YAyzJFiwdAHKADSuBXFW00akQTVq2ZlnWIOiLtehRs3dF
n95jUaymoTGg67ZA+b3ElslxLyd+L7FlcnwvEVsmX8iBSgTvoWSQbHchr2rkRU3313QPI3vfH1RqmexjYKIrGBk8gYEpWAMGJmIX
d5qLJSzKyN8WfWZakALKYbsWaV26MkvuT0cSkpPTfWE7jVNZ1LpElgQenWSpZzTueIO6+xtUwi0pOI6OjUkzdFl6FuziOLtk0CrR
7S0j20imvQhFD8YvDdkqCr3z9eMHMlyJJo7/sbWxUsFI6JFvhyQ/a2H6dcWJPcWP7FlXq9N3gZCc8YIYUld/+I+33orHCLKvk4i+
Az9SKX6kjB/TkJgXhfHjqP54Z48PBF89z/Izbcp098jP4r3OfivunQ1V432d2NNr9mVuDwZkS4yRSqgbo8QsGFW08Paua7uTCdSW
wkcWTuBptizcghnclmFmew4gk/AOLkr4LBEnWyI1WCJ1YEtkLocypXfBFo/ptuA9fHRCdnBBgkCJsD14rMx4LlpBATpWVZ8rEyYB
x8T2R++ALGzpmYFjjJbtjyOKvC1bQo6C+QKiMEDl3j8Ze+yID44gPTgMV+oId8WdoMfuW0ZCj329ivYSO2rYNMGA0XrpmRdwhIts
eTSmYtPIKs0N5FLePPRlW+Tkldw7xy67aYAcvDRqYVepW35evNeVWP0SqimbWsFiSMFczGZ5eVFk3Lj0LC8+pSO/+LQK1R//4989
+hqzZdixknyl7GDJ2LG3C8qWYkcze/CWoDGONlr+vpgfbQJ+P1U/1eJZS2Ewc7GD+Z02XiBAvZY7ZwPgqGxNyHVNwni5vqREyOM5
zbMxkGfjoNRiDmZjjRHBfxSLXsnnyTZ8u/LjPICvLPDzPIC3LBS5ywqSbsBGB0V6HNIdA4WBS1BkuihQ8jjvhtWY7E+3bMpy4E/3
RX6Gk/wbPuVhWJ8s465KGeHNf17izfOtPuXh1HzvHlYS+H/xsBNCdnarpKpNtgeY2wrb92K9S9d2//iHNzE6kOBVTJu3kasSu4ec
3juooRgCo4elZ4GcjjgdJ41dFeIF1KnA1iOGn2nwQuuqo7wHn8QBlQSXF3FI5RTb7DjOuUbsTnFpI24rc5CgoypkSmDUyIfIkLyx
qOoQiaY0a3hjeZvAlGY5HXVQBP5wp2ukHxg1aSANtpYp1VmzUTP3zm4r/xVZmOOg3HiHaarXWQI6Bf8VkZRL6Ngt7lLV+p9aNjWE
tnjrINhSQ2hLimQYoUbrW+Zgf8tpkZ0XshsNRme1CBcuG3jjKjnXAINdTdWxP4Kt2/oh6lZ63s0vlzT4eHVadp7m99n69y8dbmZQ
uyaNWdsy+/JbcVRbJpcbmMyfSTuuEZ57Zwtkmo5kyrthfLikLbD43z5+P47odx9/nf/9+euiDB8m/2HjjvFS4qvhjFDzwhRSLlom
EGzh/CqXfUIeCCDWsZUCIKxjO4ULtQMQ6KobtGC5XImyeimCsFqfxeIIy++1ilqsk76nDqwklSsXuacNrCTouLG2P75nbfecgZUB
9eUiu/FWl4vc0wRW2lrKRe4Ywh7fQhI94OMzK9HRyVGDQAkEFcK3QYM3NpFIQWZHyEGZ8BKas98mN8O2N0Fvk14SyAGx1qQTa00e
s8Hcl79fjseP8wkf4NOb74QL8OmvPuIBfPyUnxAB6jBEqC/OaXDL7wB/T+vfG2dePIaBpjwgDWzKxBHc0Azb3gS9xcMyl8i0hXu3
XLSMTA3XMjLsRTFQIJNLc5lYK/sEmRKDJWK0REnKQadUOwDh2lPiC02JBRdUrVz5EATPhyAA4UyAoM8A+UPbuJRy6XtkWumrLBf5
HmRqzpFpZfjgHZF7ImsgG+YyKZxDxBzDdSLfo0LM+0GpOQWlj29YCZR6gVbEC4UDfa1Qka/64f0jKUI/1TXKRlNufX1wboVNX+K+
JkX0TihqSomqucIEF6zlarVcqrb1099/Zb0XipoSKPq2b5a0cAItnEALpw3VDWvhVK2Fz6Go1UNQnFDMTac2zbDtTdBbPCy6hqJP
nUAqgKJQ7sse0GMD+06l7O+AonQBRd8y6fIh4DSA5fdarB4ryB7WayiXvoeij5kP6f14lM7x6BMghi7xKNJNWMxtsMi3a5Fvd2rU
jeA9NlM6h6dP4DZ6u82UruDp49tJQkaE24lwO23omwm3E9VvJ0lXhjh47S0Hoi2tr+/PrbDpS9zXpKsktAolZq1YYnYTONSmK0W0
6kP0UBItcmWrpRGwJ+WexpnY3W+Pn7tj8PqGA/g+8ErldtR8DNbfiwnJQ+KHB16f1u0vg9eJAtpZrkORm3MIXW6GbW+C3uJh2WsP
v8F6sYpj6uRwE1Un+s0eflvq4eeT4PgoODneiGoHINyCEB4bJBr0LRVQpVzhEBjNL7j19xJelF9w6++0/r1OugBen7Ie23Pw+oTq
t+83ptor8GoIEqSXxqzZlkbAnpR76kbwHpuIPQevjy/tW1SY/Q7waq/Aq4sR8nUilJ5cGgF7Uu6p3E6irlRTTpZJUFJ+ac43R26G
bW+C3iaFJSFYqJ4n07gBSYJvVJgygsUKTktjQbAR4jGWRso9jTOxu+SikEoWBXai2DoJ33DVHCHYJx7othDBTknghkODc1To+jvA
39fiQqZx5sVT6aa6YLmSGzRnE1xuhm1vgt7SYZ2mWvH9yLdy4Gs58L0c+GION27m83QriEWAVy0+a/Fdiw9b3TKGEzyrI/CjRUjw
jMCJEIE3LMaGgezvRsfz4HgaHJAJ8iS4ljkQTkfwHCez/naKfwf4e1r/Xj2APbpFinieeeKJJ553apn2PcB9SiUd52TxncTQxDA2
KaMbuSl6D3jdrpDThpGPUUm11L3mecBsc5yOpSFDSXGGkuIEJYkflxpE7xTeU0/Vs8Ssx40Tx8lZuLtMiFACH7L+lkbAnpR76jeb
pGzTlKhLibnVl5ZT2AqbvsR9rZpOAMCen/HIrMTPeM/PeK/ala2IgaMGy/rSWFO1wLK+NFLuaZ+P/RWoxIQtJZbT7jEf33MpHSFi
TAUmMRWYsKRE++dLut8NWfXPP5d096z4558LGVNsXwIZF6uBUk4VhuaMi3MzbHsT9L6QRH38y+dvP//Dx1+mcX3+5l7uVkuEbLno
g7CEBlhQLvssLKEBjZUPQLAOvGXS5dwtdm9EcG9EcG9EcG/EYcdrWyh9j30rQxzLRe4xb+WVd2Ntf3yDSAHjWgcwd24sKGRuBOxJ
uaduBHvVoqBcqsqyVRasstTKj96D3UoTa7nInQ57fP92NOqWC5UjEoBPaGmsEQnkISKBPEQkkK/cTpKanBL3dbbpcsspbIVNX+K+
JjUl2XMb8Mt9+SKWjZbXZGksWDZaL1e0iy9lBypmopfT6r7obnFPdw6gjF4fP/nnEQltSS0VMy/j1nFambc3t2bUurbCpi9xX/GI
zKPlBspFH0DWBltDuewTyGoZslqgw2bIahmy2lg7gL4hWeVy5UhaA5G0BiJpDUTSGoikNbFOet9ghHK5e9z6xCVrnqs5cEfuHrxG
F5jvZ2ksmm1uBOxJuaduBO9RKxfZXu/ZV738kuUSJQT7hPK8TPnyCvbU0li83wr21NJIuadyT8khfHbcpjZCQdelOUPH3Azb3gS9
TapLgLHvUV9nMNaqxAuzNJbY+bkRsCflnsaZ6BWWcF/0/r55/EAcGmHr4iHuf7McU8hhCRrCEjSEJWgIS9AvYQkVMy8HC021Pph6
ipvzDZybYduboLd4WAVZYQ2Os3LRbza/lmaFWYA5gHMsElEp5hGoHEDfAgXlciUs68D86sD86sD86sD86u6YXy+ywp4wTNL7YSy9
3/x6mRAGqSKMMxhkMMKoE/kee+tFBtgT2PGRDLByie+yt15mgKHxHkz3YLgHs32l+jnQhnHwiQuhc3O+EXMzbHsT9DappM6hs/fl
PxU6WzET77/L3otRn0z+uv/NTyR/Vcy8bGodrzSvTUYKuTmDhdwM294EvcXDKkj+aghdKRd9gFHfIvsMo/IALA/A8gAsD8BWD0C4
BRMSiykoHwDVA6B4QJ1c6aE2vnHYRGDYRGDYRGDYRFC50QRT61MG5ou8ryduuou8rzeIFPK+PHGiztJY8r7mRsCelHvqRtAtnbhc
5B6yVhYLLhf5Fu1lv8PAepn3ZQMw8y6NxY4XgJl3aaTcU7mdxKDX8bpNKvNQ5tbXMudW2PQl7mtSU+/ErcUpX9ZHYCDxwF+2NAL2
pNzTOBNvwa2FKV9vOIDvCxGw5RW3GLcmwK0JcGsC3Jru4NYbKV9KT+y+kIc5t9Y0zMjEv7mVuK90RM9ne92SLgPXxIcg8SlIfAwS
n4PkG8SfJXoFMHQGsHQGMHUGsHUG1TCQvolet0TLwd45bGb9vYR757CZ9Xda/149ACHRC6hc+UYgvhKI7wRyDd8uJHo99HQoSfRi
0YZFGxZtWLRpEy0EwQIicQBIHOARJN9y5WikIOXrAQVQkvLFYg3LNSzYsGTTInrvTHzovVaW8tVQ0PCWcNFaW5NQeG93iXA3DD7k
irLc+tIruRU2fYn7WvWblOPFxnLP1nLP5nLP9nIf2lXsT7kAhdFQgMJoKEBhNBSgMBoKUBjdPh/72+9du/Kk5MFDj5DyRK+GiptV
cyAneTvPSd7Oc5K385zk7TwneTvfvhgiJB4vXUo6BwFxc46uy82w7U3Qu2EP+/j9OKK//8pG+7tfTGlp4zD/l9z4+NePP439vylk
8JqUYY6pNSmQkD0cJL4Csi8E9M3D2qBoGJcdlwaLUUkqDmrHqqHzdO3hNQ/OjdK8TftwXMdRNc5hmqvpO7h8JVtwYTg+/8w/DwEl
QAlvXpRU+4j4NFrxdRqHTF85/j2yEd8GDvzRofM8ZXhuIZ/DcQYezxOzxGrLXEXj9KW+I6LdiCzYRzMhMfMRHzKAt2+iH69D8QBl
89QwN3IAUugXascOg/nly2DUQC6yhjcWKOHJGrZcW6cB2xvqvLOzrrVs0Q4Cg/pb1iw/BKxIAGxkAmDezzTovgNacQBMjslsucAM
DAb/TAvcW5XAg0FeK8n9kOfG9Z6aDFotWs/ZF0MxgPHcOMCsKkGAAZnO84SABfbRNBe5an7Sg7OejXhTfQSu2jRlummfXc8jQNAp
0iPA4Keg6zxCTd7zfFNpuKp097tq9yqBBQ4avG0WvW05gOQw3rbjnO2udjgHm2OZ+boHPgfPILyf+4vU8F2BgcmZVhuo6Xsfzpdn
DO8tnid5SIyhntpWEpQKg3a85SPAO83wDpL3X6oAd1xE8e6YcK+2XIhxPIk5d3jqdIq535IbUsygb/xPtbHti2sKnjrjDa8yorh+
UtgXDunmYR09dWiAF5g8rIjWU9t3WGcvHTvRU2gNxNfrS0eJxvzQ4RlmLl86V7d/75eOuXrphCHXYBrHAZm6GsLHEvH5jLHzIgov
HS1EQWCpEUIzbedl2z9z8G0RnOAczkvmO1yspvShw28uD0lALgeI2N5DEZ45wcEzR3umySLl0XsPmNCGznfT6SvHAHLmR45hxNB5
wYRXDuplJW5nQAu6B1wwp88cPQiEx4/hBHP9zFHCWl2GP7QPRsLALgIGnh/nuTI0Fxr3TFwwNVzna1oGKjQ+Tw2HvZvxnRDyK0zT
CK1iTEANFslnq6wde0NvqHLyyklsqQq8xzAEFl72hoaHRiY+Y/0mmVNDwJEzzLDodALLjHP00Nzt7vfNw4KnLh/Pw6jDfoOSXjvi
mNQzrx1T8tqB1P/hW2ZJNhsnAFO5sPz4d4ocipwzYyZPgY0PLaJsKRm3F8cfxlFpj/fCmkocB+LKW1pPtkomurbjhDvVftNRwWuH
Bh/yTYfPChLzPFOHJaaS1844eUHRPk4Fk9E9EmTbvuM6fe6Mb1PrkhD3DxEdbL9xvgNupovnTuRrI2JQj7tOhGgfj3Q+x1EQP8Cc
5lKCmVVzenX5g3IC7YMS3joG3jrhyqnjO1yvdPXacXtPSoAXRoaHsQM8pPPXjkMbeJ4dB34dHoztPRjhvaMTE9Ho5JibmUgRuHWs
gWhl3Xma9vrZMWiwgxMiuQJ7vzpvIOHBA2hBHTzfYUObvuMRgAIAKgkp5P1Mvfez9Nyxg7Ch7VNvU7p+8IQgJ/LCe8clDyVnbEyd
97MIVeyEKnV+75AZQUHK9ima7qHsSdV2xC4mHzo9hS3YoB9BBbJXxzghgshjSYy+Xh0qfe84T5hgkfjKGnsSPGkjc2cFT/TQ3Anv
HcEaYcC5Y/hp8dC07e/T97pSqOy9IzicDHiDn7vGzr07EYJ3vGNElRP0JsBnII3D24eWUbxGRr0TTS6slSaSHkZZo6qy0XNBq3EW
CSqZjM1k262VtujJ42yu8INpw1Z68pDqgHZswZNnerk6bYW7LYkMHb4DYLbFTx4zxKTXJw95sRxPPEzXaB/ceSwbOAu2sWwBPDy6
74jkI4rKybNy4kpucdAEDzHTd1DCo0dtbOLsCduUk3dcurHzbpdePXlAfnBCEZWu8WP2ysdj3M7Jw48eeILpHj5CW+DmgTxsQ2wz
JSSMsSEZgIm68zydPnucmLXZ1TFnL589uKdByfBTvvPDx5Y/fESrLSOGzudLevh4MdnI5bXi89V5MKIXQKnE0WwuQTQb8tc7taFA
0qnzjpYRy6jWDAWo9mcjlxCezKEmZEcP6YE4N1pTGqxN9Ag2kB4+FkvM85bfpE7wZeV6n8Kzhw+Ba5qgYick2FhHULEz6qfmbXfB
g0oOYinPrjY2W/boIcnJg57YwLep7n1fyI+e90bk2tJHjx+ShYwFBnmZaGJCVBwyQ848tIjSBZLs4CjP21Q5QjNX0xQNawLntLuB
si97unkotnuzT5PdGb6nhO6UxFohQdqvQnLRDpky56nwMLgpalitlpMIg2OImqDOqXIPjO3s8aOnheX4Nh0wbVpDrSnwsYx3dP9B
7rVC4Lh+D1oBkiwDBK/HDhD2JLueRzVu7nW2whSttIJ8PSShvEJoD6E8TrlnazBoAvJgEIZoJR6U1+2g8TgZ34IxgpEjOKUMD+qk
zG6fTbW3E1okCuUxbdLG2SzXAWocJ/AjyrCOzasmGczz4WgSQgO6VV53H5tgw4xysGnEOnRsijBD/0HtH0mW7SPWoL0m73aFEbn9
d/sOYiTe63D+0lN2iLNSAbLJxm/MSBAozFzXDxxA0ZUwx3QuroSkeLPbBAQFTm8KVrQn15zWGoD393jBK36Cj7s76RzBp8f7Yzyq
FtKWI3tptfFD7GCwuC5EwM9hcG17EisRJHDb6geO56nbyESuEG3BL8gkZpNfUEMcJGl6bv52+kDD/K0hEvsz4tjvYB8bnHD3ApPw
ICU2HtcZ6XSpyO8puOwiTNmmirxjEs/npkzCZ3rIwRUecq5H3Kah+lGuGhueuPYufElTYFp+sIybLtqcZjhOl1a5nMl4SoaUciDR
FJ6rw2Y6//fP3378j4mb9xcf//D5q48/jOP754//t6wEggNQFCNGyjkxASfb+V5qn94cw/ap5CA5iUL2ER0OIvuuXlIHbw5CeBM5
KG5gmJjJgp3RSMUNXGwaSb7THRgGosC4GA4q//oG4XyAHIbbZQ7MJaduQXzMujfZB8HT07Il82vGCddeEKqXIhvXy0v0pmjaieaZ
l+o3IjdOy6xnNeQ4nMfuvPJWqBv5Ep52W+4vX+ROsWiGI3cUa2hDDiLRvMbUG99y/LOa4/Xm7S6GJBzEYN+Um18RsNh2VzHTCYud
mjbZqj3hc/3J5yKdRMsegweBKysRCuVamu73jANhl7ngrFQ7LGjIlUBOAR9dyx4DvcvzPr39U+SCRCMyiZSvty/cn3K0iaERr3pk
JdN+m75Rp/Z+ivf9ReGogK+02GEQ0hoFghAqUnGftz/+TBaYCHXwPSZkdxu604LZUD106CH/7EJSFxdS2wndoGsGQoIuoO7S9xAa
EUDkvCuC9BML4e25WLr5es71WAjx1LpxNbTOV6abIqZz6IB2aUkLnc/JCP5ng8IcIekHHf1NjWkK8LIbxicrCVA1iHn0t6GqKcHL
48PBZppwTCM5Au2+YRAXeNkDXg4cawIp8oR4uWUk4v3phftzEywF92eLcOm0JKb3W3+/lqAejwicFtuyIwW4rPnCUCJ97TFB+03h
e8DsMmXGee3XMLR89F5FxNNC6xH5ek2T4D1itjGxoiR2iqoBsQyYtIKLDUPYqydimpKLVOeWrb7Hy1plwVFMyNA47S3fLOhEf0LB
dpjdfVOshJldlmzvsKLdlCxiZi+W2yWCDHLlONzeBxVa9pmMmc2QMnTXk5axuaD4ZO42moMVJutMIIbMNIynwLRrPgkyh6tir8eF
visHIS4RuSDTTrhN/FeIiZfJeOoxJw2oucd03L2WDksC3j6jMmo+r/9AcEPEDl8v4wDgs81bAVMYNqhZD1r3OBzyuZ2y50wOao8j
jCefkclUBDGijZlSdiPYsWFu6g0qgswEkBneVXxmSWHpKdUwhiPIPIVEZofUkkG8T49wBxnhNwdxBpltHDSDd2u47mkSq0rdfunS
BWRm3IYl7OVi2eHukaECE7OJUOEADouDfJ9MUTGBZtUyiFMjc+xtZKZSzCzZHZEkqGnRf+yN23u5voe9ky4hs9IBCjyBmdnO5/GV
QMBr4xrGIJl0HjK60jlmDoKNOR6wbbdIFhCzOWG8R31oGsRKiJnyVNMFEVXLB0twzCbPVmYKCJodoLGkGI25lJo22pHyjTYwG9SU
Chs4dtZMEQ9c+X4yaBmC0Hsd2lWfhJnBrBcvyHGMatoWp5jZWQchIYYziT1i5qAcr1JQscuc7K5Ec2rpbfCIUBlmtgJml86Kazyl
V5bmc2Jz6vL1MhCwuVIabso4GPkhZ2OPdZCfunEIwXDCjEOu0CmLN1qfR+Un+K5zcL5309P3ponLFsDm8fnNQRGIWMVaQfcjM+xt
2GzYGoGFlBwWUvINgziFzS6fV6jECqWJLBu83e0nhL3AzEF004lmZqOGls0gPy99gKPCOxMfmJpz5F+rDt0cwh4xp4wowOKiROsf
NX39HjGHvOxSnlOqN23Yc8RMWe45c69vlLtHzIYCB05q8qwldfDsjzWzn3aJCNeGGoax104ShjQ9MKQ9B80p6yV/utzaNu0zwYZk
CpFry3LLkRmmKDLDNV2tImaeK/etbMaBJGK3GZitiNmlll0m6d40fr+yuZ5gmgzba+BmmtIYgT7VD3Ee5fxq+urV7UpPNjILfkUs
P3pceK5yEPIK4atmXoq1uhJ7AvjprJLpMR3CZeh3WLmA27hSvnQbxSILLzWdkTv2ZekudE1X0rl9maIV3P0K/MwasvSm/7zHQohg
ecr4jJkrJcSp0ldGJMqOD16dDcxTjH1MzMDjpxIP995z57mgDnJBLQcyJ9ZaCYKpMGpe38WqF4mfDkiQNdu7Y5J4kKdJPMxKvT+S
M+A8cU5wTC++/iFYRHH1BKPv6vTjhE4HuSpeSFWxEmlmMDfP0UnupgN2ROOAGjmuqCqEIeXkgMgYevydVOM49kDaKraQiQTJFgqs
6tZ12KNpZM61kiMGyTZi6wERjC4Q4WYlvUpwmVO7fCHomS1QOlrIttSGVapNBhhtbWy8sfYKje1xTrINbzwQjZtgj7HHic3GH74P
CFQa42x79yl5mC8JEN/vqTnUQShJ48dLgFuzYR4DCEKXAILjNEi0gxqfAHh7z7DOekiDdI6xnQuUGnehbLCegp2zcXiimrAhQXW7
ybaVVyYNZi6XtmrQYEzsokQFDO7hLcg7JfBOCbxTgm5dp1O7dbBQ4tREA7EejrkYPVGEMHbjO83MDyHq0Z4XC424gzutkOBZ42jp
dE2G7luPsozOJ/QkxED2sU4VpB8yxBgxTg5WDpTr5I1/z4n8EwyJjDCs7zScA7P2qDyiyeBv3CvkrYU6UDHzLk4m8MixSpom+qa0
S9/814+/rOmR//TxXz/+8Rcf//jx18/fTmzRH/99HPkfCzmYJx8YE4BZC7l/RorM8IeopHVUGxAPw5oWM9dGQWporji2TUnsOqo9
oOeh2bRqhi+jEZTr5jpxGHPt9jS+N4dmipZxYY95faNaMYVTNc+XKVlFu3oWd5trs4w5/WgqhtZzXKfrGDfr6MXYea4d4QQekptD
o4J1NDhjS/77vnSL43U0PQd1so5wGnEdSTyOtu9cnS7jF41SZNqMHM/lMQWCS5+p5vNoC9ZRYRlZAwRKebqgHJvxzcfRliwjTYAR
MiaMOy9S3L7lbfEy2sHw2AiKFsE6Wiio4XTr1j+1bPHIJjZivifQnsT3fWL7jVbNuujc0gUjS7j10ebGDC2TmZ4tXb517xeYvnh8
U4WzOAfN761fAFOX4a6crJuooP9rHN4fP3/3i4//NCKyP5aBG0zUUhv9I2Mbd4BtCmVvV2eTqsYfbzjQdGvuAdVnK4QL84/ZYUqA
KZvUMHcQ53os3hTMe8BSX5KZ6zhFr1Dw0aSHDeJwPOkRg53dwQ1XKPx00jeYx/HnO0jOYg3+WhHueARUMO9mIDxvFzErJpSuOZVN
vYP9DrEqCc2Leepj1YefTf2oHiJP/QGis7jlY9kIbMHUE8o+oDzzeNpNheyjqY+cAALuUph3xrFkqySfbnmD1yxUbbFR3PKlIzhX
0GzQHBEgi0c16OUypCrUDuBg/rVKePCSg5JRMAYDduXXDIfyMZytxKQ+OckCNqGCXbjcinOk0ktZrf/2+e8nA8j4z58/f/f5H1dD
SKldwWQbN6REkeRoIYCZLzdBxRg262LKw8aOw24rBrFfGIMnZNUIRixqxFvDvZj8K0aSjbkGTJYirQUQl58QHFUMgQ10BtzoFsKg
ExQ3iBDWcZx5UjGM7AI0+1w9mWiNybtaZdNOttbDAWmGk5ikaQuFazbCj9chVLOXVUn/5Yv0ia6UCJgooOpRhOxasprDnij41pFk
2zkvRd4GDFQgn5qfc83bIDsADXvghBp++GTb0IO0fvtqsTfnxLt7UgnT/Ong/jP7IvXEJ4CAzY+NgdvHaoX87EyCDZi0Z2dShEyV
pWPegAnZ1FLcsqnVbEBwEPAejOMCUGDC4fHO8S7fi9w766owl7xen8leN49q5/wTVWcQVWfALEXVaSDSglmtIwThBfbWLj0rFQRk
rdhobK+52V2h79zBO+fft4zixfdn9pF5dBEX2GuDSLBi4f5d9qqPAifwUkxxLUOjfa9lkY51osFmFvE4PnZCjnbKXWEhhgnMUj8H
1FaAUHOJxpHUTK60a0DnjI/X1DgKGY8bpoo0yBVp0CiOFeCpcRgniNyybYz4uWbZUGDZUGDD0LosV4jcyIicsWCswYLmCpHHmeJ+
zQELfHQikj+GI2q5imHsEbnEK+XFOi++UbaAyJWMyBWisQBorPVkSOrkwUvUXOFxHT2zLOnAoTVLx5LsohQTwxkVWxfiDI9LSdMQ
QtI8C3s8zifR8xbwwGwGvEmt6/9eNG6u0LgT6pIWFESokC5Cu+ggEI+A2njpWQLxyHBJL+dVq1qQtbaeqGPzw2zc5l4Re/qWvlVv
m7l276q3Q4quj8r8FixuSrG4WBZLKCLRbTJ2t2UcmDAyn9XIt7Xe2E9Cp3Gcwe8Nr9k163XVwZXhN1iTQHXxXMBU9EJ0BwA8AACP
GsqWW0YRHlDES9hqy8qID+uvjB1vIfIuBWZg5N7VLJ8rp031+MfH4/3B0SUIx9h/EkuGBoyyax3DAQTnPbMJ4wFqdNg2dS8BKoXg
rHet3cR7OM4p52OUap4DdAHC5WJ2RfVLKwYgHR7IdAxc8X/+63K5R77cU+u22MPvdF6N6bhOToX0PQCPgvQolv1Xrau/0yVGuj+N
ZEKxNRqNLrG3B1Okmu3fC/jGeiyQC9N6AvZ67FH4See4OwpkE3JtgdZ9927VSVe4+2E3DF0hb+8cu2G8AaPq0rOkVMT/n7l32bWl
Oc7EXuWfecRCZkbkbag36MGZn3fgE1ASIHjQgEh3mwO55YsAGx4ZlGwaLdKSX2GfN+qqvWplfLUqqiorM9c+BCFg10mKGStv8cXt
Cwf0mS+ETS3nT9XWbKfw6Orr1o7IicVJJaOPdsR2Ih9KX56lx5zJNEZZauA7qeA7HVQrhUGCqNZScEDfZIHydB15trwOBN3bnB+1
Nrv3MyqOm6iRRZhRaEYB4sjj5BXmCK+zjzbd4ysYjk3jBVCBS7P7Lp/DcA/NAxj6RfoEzI4SQ5pRcBq1LzoMj1NMIpRZ8lXFPy+j
DxgeJpdZKmDs0uO94WLzNRIX3++2mt+oxfyhU4YDJE6IxNUWrdgONUy9S3GGxAmbemo9bBkN2iaLli+Q+CY1IuqpEQ4aefaKoN2g
MJG45DmKS54EjFsSNB5azFg+h+PQJVd5zyQ/JbQkbvE5GI+T0klVgkTHXe1bjsC3nzv7Ho6HZCTQ7KMVX/g68mzEY8QbmXxwnaJ8
MSbnK0wuuSmKfw1qtVILMOZTVP5mpxpfofK3RqL4CpN7HzdlzkBp/xhZ8V0KwNZrqPcM6Kp7UT6Srbe8t4+udI/S1efgo8fJrMdD
SduzvLA3JjdGaX4/D5wGPXAKERsbuo/pqUPceqFm8ibCnj1G1j1zBBEMl+yo1Tl5RME/Lo8oFFDZl47EHWJoT5jcZMW1AlmOrnt/
jnD5V8pwhsuzFMjnR3rrmvYKWa8eenrmxKP2RUfls/Eq3cmXLNMcC9Ypg0/X+MoTnNd7b02DzXCayO8gj/8pVQavSIZ+0MZiamRL
fux5Sr+oQXHQJNBFSVw0KQAeiQ3ovCKxH2w7wAQROlIn6DWVInTLNtOAbdo/vkFguof60SAwPWCtrW94Yk7opOSVE37MKBhp/ddP
Xe2mpyUZaSLXL8UeqTuA6ua64J77ZdgjdpI7Q5DXT3JnKANqMiPO6V7pYH4ZquRNJ/cIrdz9CCm05PIsySzOA1HMOrJml0suC1Ea
IMte/yGS9Apjo1hxA67pHsWT6D4C5UcQ3AD1Ry3675BoSg0s+V1gSWp+84g10OA89gUSEba0+/6Ap6dNCtXVHrIkuYRAGVj3syS5
hOiBvSjmAc+Frv6Jp5RNAm+7MeLrkNEHtl+Y2YP0KLILAAjDtK4C74NYoCHKixbEBg3wos1/x3HSqF73yMjS6yRIt47EJ42YWGUc
40CplMClmEAOsr+dmEAub0hVBkqjpO6J2W5BGiuGu6VNVk4ecd11zC+HJ2rui7zJ5RyHHjXcH2fwA3RUnISPygbho0pPdB3jknoy
cKN08D8jeiYGw37WkwyEc+toXEeDKwI+vADbh+l/+/jjj7/9+PMvH/+4UEn8+JvKOlGCfBTku4mqN94jQ0NqnH7LFqvPjzBOkk+d
5OC6l9LlG/MrHLHgfofqYTZqewUB9/xSOnxDiPLcEvq8gTIg3MwvvDG3XBEC3JiACzLI1KWvwjMVfdXRsXX5C3ynzuyDG3OSMics
drwVNL2zy99e592wSklcoSLN/Na0v36ZdklkAepGy9HqzRRsSoLJbfbUKEJRTvDL8Xir/OHiU2592goAh42GBdfyqg7h741pn9rv
aJu1agIpCmudFZC2vGA4MavtqA/qqm9MXNAZITWoddDSKh+0gQ0JYjaRU+uNRp0qr9gnk6MQUMwYKJceMTlN7KSz4oKnDYtz2k0p
Otun0r4rr3oH80GDAOrOGIadiSZoCePkobvFvJe9C7F7+DYHk+6Y4g3T71+fjWLX3t3Dnny3LuQGC1Nt34T+d2+PfgmTwh0khQfA
deD3dsD2QKF7A9TrOSML64oaWhzehsRKmS8vJ4G8y13OJYFqcYBbvrE87hLx2nwEuBJQtCDibJ3+APHSRkk4lagIWoqFO6+Dq0e8
JWtLtCNbhLuFD8XcuRnuAu7GzeoDUc+Gvdkf9Ja5Mff1nSDAuwkytALciehS4/x7wGvNxta6JGu2zeu+B75hQ4qmvMNi3qTmA3/x
/GvJFAFgmGuedo97yTjI7A6Q2Y1tDlx20Hd323z0hgCa3nnD0+/OIS+0dc7Af4WVBOB7aT3WCubdqFhTny9zY1YN8zYSCd2YVYNV
M8YFWGU89HM2QbjwGZNEPHHrTdY1qpl/HeenCznZ+VRLiJSm7AXw2qWpLpcKliUCZYLv02jfVTcG3nR415P2rtvUvCmnmDdGiT95
StIvLBjMqncZmoYx297l2L178TT8hP3GXefcFxb3+y6kjnnt1YOP/cZN52/XVHyeDHCceSdBpwC1kOXd/4TDvUfxwCqdnyXK0n1g
4XdwRR6agW2iQkWyhH1EG302CrPhhm6ga0+vebmhmqc3HLayvTH9oacXCjm8Mjkj6G397Wegd3M4SSKSHFU3b7xzO+gC96YXZlgz
8H2ka9wLd6IEQp53Zb0TxRRcLKTQOr0Ce2HjAfNatd2Taz51e8y7MTWC4uxNAD5bz9v+7aetx3P/BDK4PEzztIqvNxGCXgONc1H7
OS4lKEv+hXONIuw1D+MvJzV7/qCb641pT4Gvluhw0lDrxrSKd2fzY280zb0xq4Z76bRAYcSsKu5FaEUJwvDeRUnO9zmCSUUptx4u
HffODxnbEr+Zb62RMiszo5lkYtGofnI5FuDrefE792k0DfduLVzZEWwDFfBdN50yaJsTDEEWdIBCZp8NpLg8Oms/G3T50Lscir1v
Tm7EiOt/jHstXkiNVBShZ2i+kDrudVe4t9Hmp1rcu9QAiE+z1BPPf4trK0nm3WIWpl45Dt292UpY2fDyQkjXrSW+6Eo9F5kpc2Bo
u0X5Dh7nGtyL+jirnPxwNG6ZJXyNe2ekjUdDfBXbAD9QP7f+/BPoS5IPSRIeJD6gA2+VQMW9235K93qJ35hbx70ZYyBpn4G4nTs0
zq2C3gNfLySMy5tgbz0KfIV7TQG9vp7S484+7x7+l8fXFAW0f3zpDgbjK9Br0um0T4fvypzXOPG5g9ePCvPzJc6FM6X7eIE42zUf
qau8hmHpK3yd17Cnx+vGunyZ1ODQx5sJCwEF6UZw8ObWg6XD3CX/KkrZ/rymOeSjFvPzsY4ljP9ZTpS4T4NpONdvtKjshvUqzo1T
rwyqEcJONoYeTcFW5zsBe713DNV+MfeuhmLiv/v2n7l3N0jTKO4FQdmu+S5epzT4t9zHM4w7A8skPz1MJclnE8BNkxVry5Hr3QHt
iuZ5jy1DhVpIFCXjyPlyGpcGNNGy3E87xWzrgcZp6Z6ku0l5WhQfSAKtL76udAffnZfrXcwfBWclORvpTgprRZEenA8GNjkwxQMg
7QC5tKFLkv0jGTcKOgoeiRabdntp2t2xEcoFifzorrxmvBfCxfnvEhCc/zulMM/fSm84rshDD/tBIjVklkj4iemGljguxpOncZPa
4yTKQEbtxOV6plcwsFMd3U7tB37nbT6uvUPvbzRAJM3Q2cVGEh+QzRD7dNbbdikUF8xWOVnVG8Zo9XQcvz1Aps3+k7j8CVqYijeY
uef670OfmziHV+IcAM67rr6GlI8D8PmAKq3jt+v5pgnSTYFK2rskeRGewybjtGMVDrxPn8AwgfcpmFKHCi3g4xqTDSYW3bzwAHiK
3cpRwc9hk/UXIBNWbkaQ5yF0Hc3TFAmXpBpy5ch42jYMJCfeW9gnzwMW5duFc8irUWHIVhiyKPtXK2+2RnGZYICSe3TGYdbExkcp
lGiW0FlkpIJvwDpo+ME97+eKX7IU9hf8QpI8MQMIE92ALdFvMs2ayxf/3WwUZx8LBc6Co1maMy1kYDGL72MhjN9Wlv/dZ0flP/74
zce//fLxP33W9f3fyx9LB+b5426DRSdNxBx0EXPSRoyAOZZe2oi1C3NAZgfc1uJhRm5rqOF7SYhrl+Ws4aLwd3ioSfVAseqhl/hL
0+l2kc7piljvvAjMZvaF5bpdEp1rOgHZtLR7Wf99FSaAMHbbR6tdHKc1Q2zGK+1ykNIQsnAfwkHRun/EYafk218dtiAB5psAdcvu
qAdNjxS/VjpyBCEzdSFBtXuCer1oWMgKEvlBmzOwy1m7EHs2jb4mQe2SjOoV0y7BwJYx7UJorAyRjbAyRAIYuY6s55S9w7o/O+ic
qqX1YfFnQBR83o9cMv9h9OEHSlPyqRBrxOVXxLE6+vtfjjI6I9gI3gE/ikV+lMfImjySrTw50W4b5Q1Yq2/HeoE0vQDN2fJkxwqz
fwex9Y7Teu9ANCiPuvs6xcZPEOWc9TpDC7sQJK81Qwu7cnY+y+/C4O3SeXbcRMmU58DFJS0lQv71OvqwrhYiELl4NMN4m9vFvG4J
6ZzKge0OagDTGFEOqbDNeT0iYz3koGU548MGHWdVNg7xZHk78RiB1PY0wKsWdV41B8l9cYwkOotlBhpLC+69DPW8yQpvOplBp+bc
drBqTvemjtGMkYP+wzFs9xpsj7I5cdTeDGzp3iOEQr7nrSSBOpsFkq0ja/r3I1K5pn8bGiPQsL4Q7SIM5NJuF2KnrX/C4bgyGm4U
S7YLoZI1B8zPteB6DpC9EayBLGUatDE6RjBpokDPuFle6JBygQgy+AwWpCiNbxYu3sRhqF7WLAZGi8FBQM1LQC1CRC2PlUi1GBwh
UX+C5kWPkXUjA/YuCm7wUp20wxEnflQK3ONgNKV2p2xKEOq587qx0JhA178mVw1y0mSt3iDHoPlLg0/NAdk+LW1tpYpjwU5cgoYy
+ggGzW+kLVUui11ho20Xs6JxJavwnNWyzZdW6e2iHJgKrHfNkefIAehyPXq1uo2lqHkGwlIWRc9508YyjhFJeas3fLfmVrZ5uxz6
TQtQ8Cn3DKIMSayWPOjIKHbC12mviz6XX4PD3tXwskcIpe8lQy2fiQSpQo+R1VJwAMRc4EHL8rOCDHQZZDDn1DybTqxmjBwnIQad
uuGAPaFdAt1aUJqYeLgsErkdtSGqvRCBTt5bD3n4EejkA0OJR0hm0M05gAr5mSjj1rwDNqnk5svo422bVaQpzId2Vkk5xTxUP3//
qV6r+t6ZGbITmRxk7j9GnulhEdLDknOD10qJ+zZ1nOoX5azr2Jfd/KN+mmZXXgCePXkFRwly3rzHskCZEg9//vtqMzDYDIFGb5X+
ECyNb4N06Fp4+7N06Cqjj/DCEk/gwmrIYf4vp/aH4LrFZl+TrnZhrrveQ3NJ7HqPXVpdx4tU23KThH6aIFFckv5IKiZeK2LbBbp6
rO+2bmmXRLlr83aUu7aEErJsU7lri2mcJcTgeYw0iuUAWBkMTeuwqQEE7EYdGS0/6f2uU77KTjI7GpugpnyYUSLsrQaSsDNB3Jkk
8EwQeaZBD8rP8a3xX0xCEl9EF5qa1bZL8LNiC5fNOzfl52SDUP8RloYSnNAw6IRq8CDOioRSKcaKU3xYLJ8+6TIWH/VyLqfim04z
lmEaqow1EyHpJoL4zBxkLy+P/liJVGvPYdE1hC8fA08DoaCoR1vWwSv17edkp/L9sIKp5zLvue4/L6xQ299zPp0lA2nxSTuBLSUD
aaERdwIvHQ/eKe0JWOqA2UJHzewK47SMPZTF/Hr7XPwExk/Wm2bbt7Lbp9ibSeJlWSzODKQ0xrS/3ZUdP7OcqSwZA1lOVZacgcwD
xTlt+xnAow/WcAK28gwWce448BXdP6Pgjgi4IwruiIA7YgfuqGoCGiWeFxfC0qc4UeJ5EYyXha/NjzvVe5uBJf7CwO6ZMW0M4kDj
jtDeaPg6i7emPajUKlmoVbJZ1/ZmqDBKolIMCWocvPQkWkeeiUpJWDxcbjc1K9qFWmAOAG4TVku3R95xpXWobBZBh1/K2FFWAiJ5
4B3fewI3RrjTjXAM3/lhsmiGhjN4jh1Ui8KlilAuaoeJowJWhkpaAxbHY+BJiGkZegyEcbt12FzcBelyMW9VDpagu/g6+khlmBVt
lF4rC5gNfrSe13qMQqqOGK5A04EsGTy54SLp2WkBok0+WGxlBSGKGNyQEMWdfqOsJotDBojkQ5Mb+DwdmyLQJRbas6P/w0HxIA18
F3SLRF5NeDSz1nh05It51n50XowgnDgl4vX89xUdpSDoyPD4jdMfiTw91K5bW0RwlITyMvjsH5Go5AEsn5nSS3vAf/j47cdvq8qn
WSk5sLBhWHMAprV72bOrKTcWhsy5Zgc5qYdeS7aBHMYdde29mnNvRpSJvZGJWeKMXoJqHn6sT5O9mNhdr++aheVgJf0jP8xIohbk
afGdKQ/WF3LD5FXH1DAhiHPx3oxnq2thdYUV0Ysl6xmye/hyW+l6dZ1MaTOsrgAjKLinHanp+ZQHq0twegmVAEzpoMj/1s88WV+W
UCIDCGTG/Lp4Y335en1hS23Qesk42pAr5DtTHqyvwf4UUWMJhmPkbv7Ms/WFH0vQNZvlMGE3Nnd1cU7dPwzuljJrFEyRgc5xE8OY
gaS9Oa++zllcKhlMLQMZBga4TgxdXZ4Khw6LJQxPsY0ESZri0XFo1Djrtz/8P3/88eNfPv7wy8f/NSvkT5KT+R/+nyrtZ4uNJ743
Ak8usIW7Qzx3c/7NJogAErMhUUgkm09Sfbj82SHAfjesZgZIM2YwA4Cs75Wr71oKd70NkulC4twji3jACBzomV/fBiy8JLXwElqf
U5cAJ9vgxbrwQCnAqD+NPO83paDLbUDePKfy5m0axPTM/zO2gSq34b13kq+3oaOm8+b8+jZg+nzW0udP+OtuCnC2DYIDSPKAkUFS
8oBfkd2lFKe6uYgAUYckCR8Qc0iS8pFCrwgHm2EwHxoTojEjGlOiu9fiZFO2gZigB2JgVezrsvzviwi/+vinWYY/3dPVPaeyataD
NwkYvaBXm1f5AXzDtCerzRDehfguBHj5kETzaO5rbWxhi2GHA/KLC3Vvw6wHC+3U6qqq5iBV0569NUlrw7xNQoS2HFVzV+jbjmBH
1awHmBN6TouWJUjdgi4C1DDt2YkWpMmCNLd9/4z0/aua+1qjOjnRTk60C2rhXmiY9eBEM5IBat7iDZ3H/WnPsKRoDS9Kw4vO8AYb
i9fMXaczo7yXUd7LKO9llPcy+raJjzQl3CeDJi3cKANXynDb9Gf6MYKWAnPeAgX78+/8/HeU4R9+/PUsxZ9++fjdj998upr/9PHv
tykxA6osrx53f0Bpc3P+6+JU0otTT1xkN0U4I7+UOKWHNA0vcUoPrhX/4r66luOaXWhbiapTQUNOeeqY/2ArDOY/qKFaWALXtwRn
af4QTYfUB4imM0Sq2Nw9Eveqtwm94vAag1cxdcyvbwWqe6eqe8BVL+XjNwWorLfwWr0FS3t6dncfp+uCmOb44M3JD64DgEyjgkxG
NO86BDjbA0hKgqgFSVISoVudt2HbSznq8g6jRI4jJHFEiR1vcsbSzd2ozDZM4uxIQn6bxNmRJAMh2Zu38laKoYfkLA9n8zkQ1wFx
xfuXE/qPsxD/8vFfP/71lxlDLNHhf56RxPzxDz9+v8SMb6vxrtS1dmEObo7URzkIjriIhBMRwoqDluaMcKIjzfCWQBV6vsNn1y6K
vlMQ1HBZDWqA/zKMWpazfRKDE1vRQMJ8QKOvXaAKEACZa1mHQxZOsdsm0rULc12OiUQUVpfGdtypahIXpzI+QlczcQ7Mf4Zmgbjm
/YNAMCyOBtjCKFFqmDm9GpnwEplIY2Q52ydW7xS4BhPa99wqUCWWEGkCtB0TaaK8xMm1L08lqNh6AuyBJwAakJk0UKYzkBET+Fah
c2hMG/e5mIR2e65/+/HnxZk/452b3vOeaErNpAceMLnCDlokwhUWtejs/VnPHGAytbdqqpdF6trbU5dkWvH9QCod3WwdUzOj5DZa
6AYqPOseuo2vf0f49/z897vzlvoO20lMUTMZ7SfriEpXbeS31xkbi+rqJvv1y2QL92uMUjH7+FgrZh8fEUdyGbk7d8kd/op9LPUU
Q8LaNTM+M5FtJ49LzVxQFGE7uQFqpitZ8l+xc5ipbDFfi7wvfYfl8+HPLJ9xO5phtFGlfN+/tD2hwjsz79d9RlIGuLDWj5VAyQAX
1vqRy0jzr989T288aLsyA7vnFvBKsxMPrY4bLtOmkuAr7u++XAAhS2DBLIEFtAQgXwkssCVw8zprdy3PwJRL0rh8fTpKy1fcjGUZ
q5DlOgegJ3RZM+kBXMya5wLqAqC8xuX7s57BRQFvXsCbF/DmBbx5uj+18oi9eYm1s+2nQoXx/HvNZik0GM+/nx3q0t153VcgOPfl
mNGdY8axj5W7wowmAcvK+vHQQutHxJFcRu7OvVcEb/yhe8wo1rJk5UAvdUm+uz/ZTt1AVozC6QEpMffn0uDiWIXqzuEitirSOhUJ
Gc9t9aGrMLf0IntaNfK1MlusX3EzlmWsUZt8DVJ0tUjROyd23fpRiHSgw+/jI5eR5l+/e5TeeH2OkeJYeOpqkKL4tsS1JZ4tcWyZ
9t+pOV/yJF1x17+fdA7Qcas03vv8u3mdVatsCXuFUt0Kn95sPuN2NMNohUD01pT1mkm/DC5W5zEa1f1uMIvCSDLL7akV72LEaIjk
lUFaGWaP3J5RP+DiXczgXczgXczgXcx13kW6gIsdBOA1M+7hoiytrGxUyYfvb+SbvYt0hRQdeUGK68dK8kJekOL6kcvI3bnfjRTp
HCm+cQvfjBTpax2LdIUUXWZxcK0f64F5fEQcyWXk9oHRNJr3UyikROXjExs/PyKO5DLSqFI0zGjUHDFz0D2kVZnpmBFwOgB1QOoA
1X3zb949SqJJoQRZGN/K783tP/fdPkWqQYpvfHzPfIoL+Y+YIYVS8/l3hH/Pz39vXmcVKbowkQnShLV8erP5jNvRDKMVAvHPQIr8
3sLSmlnPSoug/6KcOZJDR3LqiO9PrSBFuVdWLpYFQgG5Wtbfn1FDiglMoQSmUAJTKIEplOpMIT5HimM1I5+DxLFPJJ+DxLGWO1+B
RI+kzevHqnqQtHn9yGXk7tzvBon8lSCRvzD6zFcgcSwm5SuQKAhR4KFgQwGG9vYJUVFhniiXB0a+PgFD+YqbsSxjjTpEg4bj85m4
GhoO9OByHSp84xk+RoVvvKRHqHBsNgXX+g8hGmchGmchGmchGmcro3F8w3/IU6JQsjrk8+E/LJ9xO5ph9FqgyrpRUeZRtHmEHlCi
z2OVQq8uG4VgoYFooYFwoYF4oXFNs59WjUJaaISOlZgWGhP491KTCPsHLQiqCQJrguCaIMAm5KZJlZMfszDFP//2Rv6O8O/5+e8N
U+8R43iX8TEH9Rudxcdc0wjJtR6omGcMLB2NkyruxmAycEqbLN6jx0fEkVxGGqZ/s+o4pogewqRQOeleVw33jp9RPr/RKXZM7YzZ
dj4I7lk/1my7x0fEkVxGWo6TGsie7eNcGFrl60nRlwtBq3xlGWtXUgr4TGLXJ7Hrk9j1Sez65Hs0pJ74iH5hRr8wo1+Y0S/MVX7h
awLmIWw8NydXbNXhXqQrVuUvetlOQGoyWJ9bDp3c+yT3Ppm+Bdcdl9CXz0kZmIXCUwsMup9/oxD/88efPot2/rCIcq9mpsfaq5v2
wFspu+1ku7eMwdBnvmHes8h2wqJaI03clcIyTi2TKx7LDsapujl1n2WAjqYBu7OzYKSS6Pv59/2Z3ahoc910w/yWlXs5Kju9djql
fCZCP8QI/RDjQUPE2PA+DKudqZtuDzXfuqyjHCJ1sw3MhqybcJg3rfKsqHrMuimIbQSf3mw+43Y0w2izflHAZA/33L25tdWPjx7M
jw1YPx57sH5EHMllpGMFRoVY7k37xe/GEXxsjA7c+63vSJG8udrqzaM0kc2xlOaWz0d5bvmM29EMo1UiVVTVdFgvddOOp9asm/dN
lI91kyuvWwdXbN2cuj+fnPjzyYk/n5z488mJP5/c/ZnHJkzWzfl+VHlRXjP6AbsssLE+ih9z/VjDnI+PiCO5jNyffVheVN10e2j5
1q0clatfN9tXQMvrUpuYoB/d42MFmI+PiCO5jDQcHVXfBZps8sW0lc+HUV0+43Y0w2iz4vn+RcZ9dRFO8EG6Wa8fj81YPyKO5DLS
sQKjMr7vTfvFz8hXIU1XH0zvY0K5udo60pzha6Zi48nnA2mWz7gdzTBaJRL9HKRJPwlp0s9EmvS+ZMu6OfXD/uZJh+Va1k33fnx5
UZQzGpVcluW8cbr3+xwuanFGKx76Ul8lfTWgvKzISZidmzA7N2F2bsLs3PSSnVt5dHRAGaYQQnGgyOcDyJXPuB3NMNqsX77/jCf3
NA1TVqJ8lO7EUNUta/D46FiBb++OIFXW64z2mNLPAJS1NTsLQ18COr8EfH4JCP0SMPqljtXWb16abPZy88rn4+aVz7gdzTBaJVJF
3U5H+XHdtAfZmR31QnXzngBKJyfeyZF30P1UDr1LLZN/0et2Wb/D4Kdn8NMz+OkZ/PRc66fnr42FX9TwjH7B3lPFUzvde+p46mYf
lnxZN51CJ/mOACa/o+67bjYNaY52HF2W9RCCG0JwQwhuCMEN1YIbrgiSz4cjZgmSl8/Ho18+43Y0w2iz4tGQZkcO9b251RorAj/y
+rGGIAj8yOtHLiMdK/DtS9/n4RxCtTdtKN/kvd+qKd8ISDMC0oyANCMgzViLNOsrgQJPbNxTCvn6XPjyFTdjWcZqpKlsuQo6E5Qm
aE1Qm5VXrrIQKMkBSHICkhyBJGcgucapT6uAII4bIZAbIZIbIZQbTaMQSh2QoM4gqDMI6gyCOoNvnFY5+4EmLqdu/ftz5vXvCP+e
n//eNPkee/ZQ4ldPu8egPe2l63f42199gQFXWRFkPVQEWQ8VQdZDRZD1UBFkfZMAiu/zDTixpjSog/ahetqdChvsmqyqDHpHrk1V
bZAP2YEnLjvwxGUHnrjswBOXXdvBOkotC07qZ+XzmVq2fsbtaIbRHgWmQNYgRkoQKyWImRLETgnUpz+/vz2x9k5lUEeP39vTK7VB
HfVxN67fQf36G5xkd4qEZkTALPCBWfADswAIZkEQzH0boNcK+YmM0GHK56NiqHzG7WiG0dd+wP//x79//PHjzz/+4y8fv5v/+OOP
3/zq4/+c/6gtJvJwQNL+mcRiIofFY2mEIBuc6yFGH5QyNj7ogTlEkj3s9dAdjpXqaF/THa5VnPJues3tyqrb9bzpfYsUcpM8KtPo
9KokIYVIQNFupmCGrEkByF5i8G5n/GNW6fGj0ioD7WTw5dZ4xRNw6IBoPhjfXgWgIoBTfHAIqcIYAX79IsDSFrnwhCweHvDqOihx
8hHqbkeIUpScnE55whQPyaEDtFWAgqq9ONR55xfSskLmUzrkTjw1rTxW5UrwxWkYsgKAt+FGhJMbgb6xIQeyYDw4kLN1liDM4MVN
HOBEhgRVd5H8kDMJel+O5dK+L7jybM67z6GU6dv5cXJOHs/5eU9JKjwWVp1o7UB1+13RLmFySlAvqo7kvG073SuNtoGekZk+JmCm
z6znMMxaxg9dpN1LG5WDnQ5qGMJIUfYPnVMeOlICL37MVr3gehAjVInBY16bPcJHXEIM1dKJxcJxYuGwoBKOQ0+xfvEXJOYk73dp
5ZJ88Z8vclqbgSWAUipEAbSErpqFdJeAf14yVoJHWbUIXbPKdhWA3xdJEPD7GkdnqyRngJ/L7WJSGQZEkXMeI47yJAuWwLKEfGCP
DTknysVaOk9rHPtLx/cgB0WogWm2CdIIYfZ4HxuYWtVVGtBFPEQKDfHziQ8TSODG7MleD7ESB9Si7nHQ0dQQv01BEH+yoqCdNeAz
B1YD9n6ENHtdyPJ4KEoI92PIgdiDfi0V/NAMDSNkUGB/2sF+fx2Pbp3/PuxHQ3jIEmiokSNLQgJZL6eSIzjcAxzLkMIQxaaq/+wn
60q2SDbLy/hMksg8ERVy9RyWL1vwY5gR8DY1slPjaphfLg72oGM14JbGPGZnmD84dnpZYohApRZNAs6FsM0N614lBfTvb1ZS77Yb
u0Qa6A9V3g0/BqsdgX4qpiJdgP4wcj000D9DJKloD1BUF6cI9A0JYH90aeguqbCfZmidTQm7zSZ1dpEhTcwFyZwJs+liTMlcjWE+
79skqRtC0i3Y/0ZcSRWw34ifX6cvO64BbJXkDPYDC6qIw3C04Wy3H2669POz4okJNWkQrVLo9nRiScfIAPulE6uTnK/5b2IeIcwe
9qdyTIKS4zfmCaZzzB9OvfxY7TfmWHzbC7DPtI1qpvYgAfaYny1lvVEXbwqLoEx9hCiaHuSTPEj08g9ZC83Ln07S5jde/iHnscrL
74fGfOgK7r/f4qBLuJ+Ae5eS48sD6cacSBXr58VbJ1A/50LlMA8FiqUMytili2pJll0wSYo0UM1qWD9eFe1kzP5II6VR/fsGIoYc
KYKlFg/qQIKlsaukOFbCyZkeEculWqTvKjO9B91vHenXl3eEkeuheyEf9/uZnpIU974V9gzH00tfle5NUi/9bBangvOzW7jYCvv2
bMObWHhsFodQcFIMsvx/utZl4+tsHlKpgUnNoWl3yXAFyo9XNYry9JBtNl25EuV7QPnirvOyMF50l2934fI5yl+P7UnsZcPUPWR7
9GuVSzfLp0f/1QiDGmK3tFQZsiSKcx/yz4yaI+gwD3rIOSElpYhPoN0YVy6fQ315/DXuDYztpjEC7KG+tTmf8n+sXv7VtW+jHyHJ
Xg9KbhOdKqBBS6G59vcZZirSd2NOg+JsC5UrMORa6o59V4n0hyyBDhcNC1y0yGmh09UF5/KQI6lH9ecr6lxxxhqacgqSzjO/o1ID
6mfLhEsSL4XJekoDta0O9p3i4VOTeZwZ85KeOvYp2YNsnpAkLJNMlGyexMkOXSXFo8JVaWpp7AJpYP+MH2iEO4Pr3Pr8Zrc+14J9
u9xxAftOWNqcgJLS92BBsnmoTPqtnyFrdlZqv3mKvjSZsUvSeqQC932aIolTf2EZsrHxRJ8WpnqoDn2uWSK1+07AwlQeJIsO+rNE
g5NozgxebAOpGqYZZ1eUr3psYlfexQgRT6ignc03yHNsTls7Lmj1WmM70hrbRY/tBEetj3LdkpHimHkmetoAyT7+jo8T5Z73LXF7
Rv9x1as4HaxAHsA8Vuvv4t2whdmbArPhykqXmzQ2K+24MtZrBLV0UcDozLAVUfN+yEMMwATI9OeQxDyIhbF3MQ9CHCSSkvHvFIMN
jUbGHIZRp3ZvKJCE7ciq5bQRg4ejBNlpcSuRKt3Ruy3zC4Pk0OyGWY278/DuMdVqhyi6/ZChHsCh/UBOjq33CQiv/ajFOagJcFM0
BU7kPEVrC3OaWbCxJXAeBptLZvDiezDWjdXgiiERJMgTZAeD7GCQpzjaYYfpNHAQPBTwQo8/Hwk2MsUEcQNjRy/V3mkj5sQBn3HA
vq2jl2r/MGYlHS+rxQp+GBY8Mi2spAyha49VyrhxWOfExFg6ARYTYwY9pSJ42ziwVN/OmKe5XqCqRhicC3mWRvpuLzFkptJVcPF8
PsJlj5J9J40FZ7TEFDZuj/8yC/mn+T9/+PjnH7//5eOfHtL+6uO3H//y4/cf/1JVILwE/dczPgO/0mITn3SvNf579eW3yrIxMEAY
N80WGfbuUbi8/WEbwlZp9vaFiMRLBlqhfJgRq98zfHtbNsy/WNKtIpV3e0HrxQM0KwkLdPNe8QJlyc9+7Q7ZKotcNRAmTzY7DHck
gbAOaghQGj/k7BTrQqTJWIEfZWGy6uSnMctCeznilESOcJVG4occlG87KWjKUbZGVgPCYbItacxiFLtCxFjeuCzEkPPRDcIll0Gx
k09JZ+holaZoTpHGXS4K+t2HSFEMCjyofmfYZGkpu8l6GyLDU23j6YC74k69gmHM2QA7QsTgpYbuLBDkR4tRkCce0eCg49Jah/os
KnaQns7ixY45jTmkgBTgYV08ksaWtIjFLMhe2nnM54KTsOMuROdG2ov4+XNbqNWpob8fKCJhiTwoLc7YpNiPlEjdR5+gSoQzSzgi
mKKO5o9cFmupDo9DBVOe4vnVsUBuuk/7GX7Kd8YDChNg2y6qDfyYm7+xHfARcr7qCfJjz84BlDHCYDxDmUIWN/9dCtNnKFN66M0n
epsw0b1Z+lOw3CdTiNgXbwKJbAsANNJlaH4YHBWuI+tnJJ1N60vgKuwGWhgMkhgO/spwMCOEOTEcjARvk5oc5QtKd3mMNKeGg58R
R/HBM1R1SsDEi8XuqdmycleGw3yksZOGxim8Qeo0ZnXUyzZfqc1lc1mEkdOcofZ4yRIcIY5iN3y64IpJpRd6OlmY2XLwIyRRLQfv
/VlFBNh1dtBdUvTVbBhEBZVp+bJ5zFqohgMZ56EPJngEXTJCFcLOSXYIhxRGCKQoTp78SW32YQpxqwSK3bBYi6lcG7BwMSpCcER4
zO7sNfeMaJi5UncPWY4D8yHKQdXrJcpiDNoWHXdSwiSmKPaD98YK7nz4/VfcOf/3hpzUAwNiKVA2EQwIScW3jicbJZNp1kUchB00
xCUJMA9U1Kr9EBa/q1ML4ryazTQSOhzYD0jrysFADwvPDPZDZDAETRi6VPv3mFEtOLn2rOY0+ZHSqAYEpUoDYtD702pAjDCr3A0D
gigCmT7yJyKfCkNhZRp6dI4MiPkVtOX85E9XGxQsxwdt0UN5zOjY5lLA5O0SFrCNMlKF/eAWQjdW2GtpsKufauwHO2VHVrZQDzyA
/cAjpDm1H3hRLl4pXg6YNQOBBzdCJO29ni1N59w5UemGSs6PkOTAfIgs5kOEplHbqwamOg9ZF8V6SNAvpLzMaTBip0vDAfYmXoUc
hkih6inwrpOshmY4pFFSaIYDsWQlOwukpWRiBsPBQwHpCGlUbQkox11oyyFLohgOCY+HRl+6jTkMEaMp5lDuih8jxIHVAF1NtaBD
GH1KdbCZoWEaR5sh4SUK2MSQAw95Uo8shtlEKO0yliTe7OVdXahKU5R+AvNbR8Jn6D+NDTdQNR9YDAXufaZonzcSGaSALgyGGPQ6
dc8+QOQoyEuTmk0rqrYXKEiKgHLRRvhvqNpcgLfnXa4CqjQXjLhP3EUpxNB1OUAxHjCMA3OheLcWF6gVEJPD0M06MBdmuSIHqXye
bd8MjQyM8eUdAN6cxViw5FoRMVclKfkHe7KTFNPXYEPCxG03QphjY8HEkK+MhQGmC98INkCzbCQ49ZilVK6bHyOSHhyOXstRSgcU
tEMEOTDLU8JQgyTc5SkJvwCLLDGMEEYPNBhx5KSDSmh03w5ZFtVeCJn53SlKfGEvBEyUwlI9v5fCDjqrmr3ALoO9ECDQQAdFqB3O
W74wGAiDUe5NKTF8HWkACJGua/aHHBEtzhC9Yjv5gdCBL+0Fj7CBL4iBh4ihY02JuHzarlDokJABFXpxRDZD9PGRxeCm7IoGNJ/Z
zQwxBh+cLRbDjHCctDzyn40A7ED9fJCjBArxQBFtS6zMSJH0fTSQbDZvUNAq3jHCEIKjoQulRnw1kM5FpABqaZBCOLEZaCK6AOpF
nDBGmhOj4Su8FnwjzJCFjWyTOpEfIYcnYVIERJOGbpj+GiweAwrPQ5RpskaIUU1c2qRIimueKBoL/Qi9i41o/bSGWsBxnsXjFPcN
fjO4DwxUq5hm9/55JbWItAjhLUnj3+I+hsZCKJGd0iCJzswIm6AhWfSTvJmY2ckSBzG+1X16XE0NKDFjwDHAIx5SeZ+ioJEYW1+E
k5JqgM5Lkk4RJ9LjonkjxUZrhVFJFp6XkwYJpBgV892B5cGGr1FwksMK1VGro9gVJN0gobCabJEEGww3g6XjwmpAjnaTXgDJMmif
B7TP/TBptJiEdYLenHei9clZgW8Ug1RBsB11rRQdawltQGy6QWrTjTxstxR7gzKmOGBvaqmLYI+9GEcJs1f4S925pGtjVaNTqxrH
yaJaH9jj+JLUm4fJcuDuhlpri9kx2HvNEmDXmEcp0SMjhBeTHfjYQwolvWHxmWYIW8wGizfSBnnJpWY/VslrdkjY5N4GeaODaNQo
5YNx3JE6t0QYohfeOuhi7W2GDXXSxTpQCKMX7JtquEGVI03+nEcxTKN3UXsxrcSfrp134/TJoWmyvNTyauJLRfhSiUxp+DIdIKTo
yubF+X6SIKTiqkizgoYabMejN/AgukELt0B5LwzjwV9qsOcFLdGXpSw6hQxt215t8Zda8X96/nGnDJuBlsaf9xGP5Y2gczaYakE2
pgnD/fMKL414Zp1FpRNGSLI3SRgCG0apvSbBBhLR9GaMOOUZZ6XPOoZ7DGpiqJ7gEVLI/dqIQVIcH0AQq2UZLtyvaYQsxfhg7UU8
QJC0KZyII+SgnRxJ3MRqRa3ZeK+GCFGUFosPDXzVYqOqrSTMGBF+/SLC/DszlEwEK05il5OQrzNZsDKiG7IiRWPCipTTwZPfRXci
JucPEaFYFgydsIyCIrycjere6tVSPBU1q1EMpea8tnt1tQRgSGjHE5sbKg7eMGYzCuRk7N4c5XxyAgsiRyi0xuMZcx5yOgEOyOmg
uDRkLiGMWdVm6TTllr4PqRCysPnkdS8RjDgb6ykP1MHf9yrHMjywkFGJBDviiLvg+7ktj7aDHK34MYgytIoxGdo7sMVOfJ6GrtPZ
2ysvDR7uKEwXdqQo+0fPTHDEZcNkv+S9GXXZN+YBQ7bT+aNTLvwgvby3CXhTXQ0xiyRhuCRhOO+Bvs66ofukXf/82VdS7BT2qXDG
LHl65IDSND36oj5bPETrWu+auzYEANDMf5dYrxNMU89QWS2Kbgo4ONA2oSEOtkC5Xkt3jhGynBkDQp/C4NbljAXVRaClBH6EQMrb
7Dc5TkbFFoI727WquzAIlmBa0AOCpbJnuVHAwTRkSfb2gOxM1mH4URFCqwh7UwByhDFSBKCzrtVP/cH4thfBl+Ryv29qN34VNENg
1sbi23POQMPcRyOb1bdnkE5xtgtGyLNXilbcCUCE7pWuR0MWZG8HpCKAWO75nXuiGAHOn3Z+EuaFMS+FbgQITIr+PUaAuzICogGE
FCyJ2zmalKBpAFlImk9D5NHNgEV3RFtCPgtPdSi8kXZhpc6FPmgpdhDAPcNJTmmg9tWMADm8Nmq0Hasurmj6cFsa1QTAsAHlLPvH
BL24OUVhWfLWxaGrpJgAXquQVpwNLzUPvaLsXzsG5K0lLYrXw4256EcGgKyI4P+9GH7skdFdlgx8ns6Ky7Lw0C/kQZCylKwbuknq
vTdLsm05zCksKYql7cM86IiLeeI+cytL7HCG4tlwq7qkaxOAAF1mLHUAC8BBNCCOEOXABLAbWSSTAlQ4TWAC8AhZzuMBXiiVymVP
apWDm9IIcZR3GYIkBzysCSwjMw05LboBIKxX8dGiaDUAignr1tZFTzIlHrJHmgkAMVKgYB1rBNBVPMAJ3HP+vLphzOnY6SNNCejU
FIMEUIIBZCRfwzpKwLxqpYuuy0kUNaUw5JjutSIAYHCMacn8POZcaFaAWqsv5SYRavWHLMO5GeDeEwugW7EA9x4jma7MAB+BE5Oz
kaqbYKFzQ8gZyHZMoiGnU4UDsxKjXK7M0iOIKUE9Q7K2wAHPkxXGYhtmpWfIDdTBmiGAmIAw3OolGIDlYyMxgWoK+BQYiHMT9H+L
uIdkMCXM2aHrpDhgzN7KDVD9PCDuSHWmACIn1fMxwOqnu8ZA1KyBIsjgY6PhlcVqFrySgpi0JMipdGyZ/w4pDd0n9fov0NZIOgPN
FqyVcqZFnXJxAixNQXyykBnk2bTmnfC1NQBxQAe85Y4xl9lLLjOPEOXAGoCuhvlIFDnW7Y4krrUGwOfL0M4c8t/zJiCQRwikPM8Z
84OilrKI/KpmzMIcBAQgZBOdSMLAZGwiOI2azW2+TBGSmICsiRS7vVQ8+xFi7C0CMdWCAB45IGCn2XY7jc8tAgdOcH/KA5TGCLC3
CCxF4CGPkh5k0fNqoxWDwL0wqraKoxkEUVuQ4t+AOztmS/YGwcZa9e9KHOMLi8DsG1IMDgzwX0JggC8tggTp6AEqnIMxUF/gCLpB
U3NqCVcYBH6i4BIUF0RfwhWfhEmS82LJTmTFImCz1BuGgXr4IjSw6cCg5em6Me/KeXaQh8iAlWZ8DLXNj9LeNS6QiYYukRIXcAry
RUo6X9UO+rYo+weP1DxAjdnBN5vdXGcMuP3LixXNUW67HbkmemxAKDfXHKA1NsDAz1J61M5/pzT0IB/EBpay09LDccZzJISuS1FF
4AjGwHzWpawozm9UaLz5p9XMIB0YBDlIVo4B8uSNM6DdG3BezczQpPqpxyI6XUWJZNCmmScaJM6JZWATUsZFSABOAEAzISpvhqDH
tcyCQsH7FySoE8UhEOHuzcuYB4mi3LuYcL9KtHuZ9GnlxvRgVv0UJk+2tYzguIyZoTW0RnJr1RLm5iyR4xJmeRoh3c1FPfUOfLRh
GnXP95rLQhbghu8wqCRazWHc4wJmjCUIDlrCB0ZUu3OMsQSoX6YQR11yxbOGrjUjG2XlzFhCiqJRh3dvRVAWUUj4Elgqgxj9a+Ne
4p1i/xJ75qxuGb1EqsPKYcgYm82MEkYvO4C+YD75CKEGkqT1wA7c1L7VJXJauQzOfFpIT0rS0eI/kwZvfqEWLcEG/mRdFC+Jn99o
78aqc8W6CFJ2F6S4NAB2lucnmmGP8qmBkQ10aGAW0lx+FGo+bQwnuWM+jF6pvWsHfDty8ZDEPh4yQQ4QSHkdDb4EzmtdJsHr5oa9
BLrBYdHiwLIIeQ62jaLT4BXSANCMdJ4AaGE3FgAUnhdvBj2FxGU+33n0QTqwPJYmDL54HcwS9iSxPdwUoHfLbOvPd0LSkhZ9vI2v
/e7j32bR/vWXj/+0tLH+8XezlL//8fdVlclIQgFEatinwao8FC99S2+KsDEwhNHFI4mCg+i5g7pk4Nsn2yXF3q4I0JUBKMSZRBT2
WJMsDkzTJUp5mj3kHrGSExrVVgxpC1JvTq70WF9IyErHoyfv2JNTFUoOIrRvy1sqwptCFIvBA01LUEqPN52wjio7b05Ou8l5Skot
5Z4PMnStfNEzHshBT5hBwxEz6O15f/0y7/yqGMjWsA8K6Cf+hx655KEZ2/LRIUbRal5zlZ91UqCun18AvgdYfdod0CBPreuY+qk4
vVbEZ06ShV7Mm5vTAoCHmVMlCW3PDy4YD44a+4gQTzj1ORrh1PeQW+57fjwqYdnxGRFwCUfksJS6SmawnVIxG+zSPjKUlqsLjEm2
X/l9V978pL75+YDJ1A0QQtsen208aJbNDliDfAS+ZO/CiDVRHsTa1gJpwPzai1TX26XvRXqB1Q2/fsRhUICA+6zMUWChgT5M81sN
Lc/Y8Yid0C6tnfU8+VQqDeZ7an2wYFQboCOO1mZgAsu3AYq7CZttUGmsMMDLd4+Jq8DNUo/oxAOOrELSP+w+ZHa1kDkgemev9SZg
v0nT6ZFEeT2D4MWAlEY6Yu7ZB/2iWKyAucLL7sVrcVOCPVxORZ1Hrf/TBsGkjpn3WDmWmS8bDbieHVd0gzlBMJhuHrrm3YNlmzxB
4r3XsbJ73Ic1zSZzz6/fayZbln0wVnbnWDlNZ33oEpAg9RyznSrk083usIzcFVDmgsr49Jj5ruOtIjFHgJRNAP7GsCHktOIMDS72
rLuud42bKJaUwxSWPNryxBl+vvxPim1nSyuNJefeBe5XehpehnzMdNDuHvByGCCEtkuBLbCmPvgaV7ycoanw2iXr2bhhhDTak8i7
vnek5p+MmH//KGn+iyFw3dXAZafMrj+JI3697jfjQssAVZ0vhXrksFAvjdgJ9dougexUtAQtxV8l/WZJZ1tb3jwA8xLjtMJ+5aaX
iu5ryegSLztQ2VLXeZRZHu8+5VQBlgk9u5v8dgjmyEGl2zCV6gEz+pgxYUwq1rz4QTx1LYfyfHpxN5AKmCOSYIeOyfWrEr1Tr0qC
LhcB0sJi4g4Z9pBZ1Ee68vqZrrXfY+ZQll4MpaiAqNS17Hv94BWnXzj4xX0z70EzWyNduKyNoif5QWT5TP/14FaKPW+Q5s6Rcrbz
9pA9v34Pmq1RzlrWjCS6iyLpAja7Ktgcu6bVYXOoaqkVuq6WCsgcZoMcoLGQAY3lLenz3WOm+6rmjSbpH2kWzhkqlXAUFtUs3JPu
GXN5BOznxePYr/y+t7hJIrpJ0gAhVNMGLWhsdQa9MkKErrox+zBiQZQn0e3OaVBvyIi1qLPjIegOoLXvhtZDZhoBmakeMgfBqEkS
VZfsyABtcoOkklEYshVH19ZZoYqyn+RqDCUhnKjwx8+Q2norJSHzlTd3LTu+Rs0W7VynA9YIBZmmQ4ID2OywY9u2Ga4XNzPy1fQI
cQabN73jKKipGR4YK2fc7DtEUcN07oquqz1Mx9eOZhcVT/MmImOmUqe0cPj4Hhk03JxOo+XQONRNPbeBFC/3WVPbhE7unok1x0pt
tDx1zaskZtjodNhsM1TNuRixsZS1HWJofp10oiY68mH4AjeDdtQbNR2ngN+c/AQ6awC2w5fGV9A5KBEVaCQE1mHPD1ZTM7C+z4Ar
0z/YhZ7dSwlIRcn1PC96dsZ8kxN76eXoudQaLg5maPm0tDkjSc9Y3M0vHDJtClDPz1A6dWF/ECjG7HuAzqAzx4T9ZbM92KQIBO0z
rBmxJAp4DlXgOY/AJN97Esb67ugReOb3+Jv5RnpGgDzNohyW4krVi7bEakbshB4mskvdQilbXJqxSeHibOEGwgaxWRovzdCRMt+7
taeFlAUupg1cPChcTJBUnO9m8p6XTxZBckR3L/SCzVDPZITjZXnJTJ8gJyjaJovpGjZAF9iISNYDlPW9K7N/UkMsD0gQT2hIagXK
3dTLkypJAZRL/D7J8QhQJlA6lkQB9Ilvu6GPqyMFapiCNUg0C8kyUMbemJ3z72G1k2xzB705RbNhSeTd1+y4EhKdUk4JRXi19oBu
op7j8kdE2czgnM5WQu0LnJD0Z0qiVZeQeJ8kijfIAIXhBd7tfiD2eJskE5rEuGOZnyElufsg7BQqhqSc6pzcNKztnF6D3sv/6t64
3/aAgyqYTglUz6g98IxmqcL1HrrveH8zH/m0fnGT9JGkqXGmGXWXqgBrFoRuMEfauhCAjGA2QcfoUQWIR4F9QcIpUTYpgts29T6X
p25sbLD6eELW7eJgwV4ywGiTjR20MN/OqoZIPb0G31EzRo79KyaXKKoeIouPGPfeYh2cL8y/+/AuHbykYcxS6AgjlR6uT/YFbwpB
w9or9WlQp9ir3i7823FWYtIWfAZ1NhW+dvvZGztviE9c6dKwUB7R/D+A4v2njz8sEn38+ZeP//Hzj3//+Leq0kPSWfV0Uj3obMPN
s29wOQEfBFY+QnGvSEAO0U9qlmAPyAmaHoEYLLrXG2x7BPWG7WJIA2vQu+jeZ2ykaNRGiqF5dugojEQYoHWTg+5h0D4Mqw1dav/9
BX8TvBXw+6EJ1qb3sz94s25NTrvJpQhfsvcrQvT3tvzb66wMs5pTpqrQMeuvX2ZdkqKjZEs6AxRi7sHbsAJtxjrDGG2zENLoG2Ke
cOOnGxzdtyYu6Jo0enKNtB6oO9oPWGkjLjuNAbPJv2mzAU3D1LDSdCcB5dbUBZnRBkaLQcePBVihmeCylAmaY22LWe8dMmxGju1H
Cg9yWloQAT/qJ7VhiW/P72z2UOU/a2XKtlfVfd+/9IszUK1YQpAou+Jyx4nYAWfYnsgsyDlkkmdA3oAYsUGUDbl/Pb7tH1/1hPrr
goYmAS4fIy0HQ9p4po7buUHJckY3x0HtWybspP2/X1H+sx1iwwHVQHFhz/tRcqbT4rHr3wn1xhI6IGYcYK105LXzvBkYeyzka8WF
zufWm+2ugXHagDKvYmOLHelNswA6Nnabw4nEN4zEN77xsXC12FjyX1iCwEwHTqn2VVCBMa7AETD2kibdfgg0YJwfDunzqkIHuJjv
4WJ3iYuN0vAT8oWhn0r7L9cwMSy7n/xJaDH0bLiiD+DOaTFNbJ0SOibeI2MK0M3HYR0RQ889LyWFnu7BYncFi80JfXmzHnBXmBh2
+rp3Z/uS71QfbVW/P60gbT9jOi5WITlml3i9UurW1Brw4kByyigCv1qE0lXvHWR4GNtxzjQ9m93khAFricGl4jvN82+PkmK5ZHu4
UDTtbIbbWepeRfddfeqNxB78ef14vmetuFpYzI9fvu4OeXEoU0pAfxcy9tuJqX89FAeBej3wCfZ60luTAF+JjF0NMj6YviK/o+n3
a9p/1jVeAGCYpHwgPnDy81iyJOy7AVuhQ+P5V/pyOhcqDnbFbLM+L5nQrkCiJSvLFXfhLLoz/o5gdA2P4b6KwrKs3Vd306FCFdiY
sEuOl/adOjA2zdOfAGMShxJB2F6uKBskVW5fAuW1DBvjJKgUdZv6wfYV0D3Gwk+XNMvo2TNzfa7Zhub5FVy8NZvNFT0d31MXdA6O
Ny/zed5d6Fj2vUrYPInyo42CmELPaVPAsfUhn0y8NO/yBA7j0P7a7DXRgZNEq9JxU/s508AxJq8dFKn6g1ant+ZW0LE59ZA3ql66
hsZviUvQFS6mDAlA5D0AY+kKuWRVWqkbZJ/af7yqZWOaYi4ZncHOr1wJzsZZpyZ+qthMk81cXjjDUzYm9Sq57+pLbxQfVFD9H2Hq
F0HbneAIvPnegjffQ2OrGJyYxyl5278e347DdfrT2wiKqQ4U2+0T7JWnSG4md9xMHRRrlEtOeRW4H2991/1hHlBHkoTnp+f4ybAB
LrGYXf9GqJB4wRiuRHjyfAUpS+h4hiXZC8U7Pet1Hvwaccp8z2PM15A4vniMzdBsCq5AxRts4uCQkNrIMt17xbnWYyy+LBJ0iB7j
gPm0qVkG5blkdCOYoe5ivkLF86HaVGkyGIwlm3ntwvpkjub2X78HxnHrt92X6CFecffAKZ+j4o1Bcl43FTsm3iuEoDtKvMbu0HPj
NFjsH9RfT38RRrm9Fy7diAxhyXC7EBduGr3xZaNW4Fue4/2GZ4xOtO/4ueNY04MMIKB93hsJFV7BHz1vi5pQ4RAgM/b+I+yxEuCo
xRzbXxfdc7z0XotPQRLPmp5KV5UlbyZ4C60TyGRJqpgv3wuTdIvG+66++XgHSPVGeUTJ3XpXR8kW+AJ9wNcgJeDVeODn1Xuc7uWZ
cB1Kxhbtvp7gqGn2r0TJXOc6BlV0xathun//ARKIGyTgBAmU/m8W+r8tHjU/YC903/EMPyQJeilts77wittPEgdfUiv8Uq8otYFh
seTsHXfOaWWgnNBNxm2QSxsh2AuFX/e8K+dVgYQlk0Z6UEp6GOSpG3g5zD3oUlESCEfGbFPDgM06wjsWEb+mPnH2T2mwGxgHXank
8ga5veHe9T2pCBQIu/CXitONS9bF8mdpHOQlG3lekHsg+rggkKAOCxW9oBvsOSfLwNy3D3s0/ZInmFVTkg6bE94+CDsFYtNRoqKt
qcW6L4CSkGECJGREQDwkz9XiawhYE+i6jsJelyFLoVeY43xTeuJxLaCcwW2CBDQSBmzvsVVmlwD71MQtqoKYPKlFCbFvfg1u21iR
LbshGe66hmpGgM/g+XxwCK+YziRIZXYmSGKQu5UhcVoPiDkbPhRirLQ0NwxP5J0/6/cLrFh67BaIuZQK2sQDdKeCvMPG4xUYW+4W
JS7vZOxUF+cu6iSvRTAWaKA5W7CPApB1x0BDFuab8n7TxlOq9YUAR1wYgm4UKL55xoOW32WRLsr1XWEdj2+UaQaUVcP12roUGrRY
vXBrS0JbejJ7SXWeIWdxXie7ZGUN2Rcdli89D0u/zQWWZ3KpFJ3PeCeWKtYl/4O4PDKW0hRjPmv/O3/+apby/5u//1RVEsjgRTbS
0cScEjQ7f9pLskoIveH5VoqL+xNPmzNXSXHS55wl0YPlDmFpoOSBe9cvymUitD8ozvd1HT6rhDiI/0Rk73ZBYVh/KRU0rns9NEf3
nqYBk4R98W90T74H5wc5szWlZE2nQckNVLOkdXY00z+/UkYYKUCyNANfxyOQv2JzQ5AVYlL3odR8TaholeQQaMTbfyn2MH0T9/Dv
PoyKM3zj6/L1pNot0w/MGGmZXsV9HvziwdkklWyWCUIwSGZstwC96SSqKt0t0NsSZml6lpj0DLtYWHXdbDOS5FbPFhyZ7AdpU81P
Hl+UiFeSBjPGp80oYdSdQ7ZDnwMBYo+A2DNw/sT5eRm2QCdec1K85settHqk2L9ofHWlPQSZTP+VPko1aS3P6FmNgyIsyDXNTitQ
NBMD5Eg0bnt05L5cVWHxoNlwIqaDUkU3xVQ6kS2JKDHZlkPsLoG7lVaE21o5LynB/oAauUWIA+BukVEkgs/Gi+dS+LJnQ6xXjhPo
TrEsCckpJulozpIRwdS/JFcZKZoTzSOzR7cEero2ewxH5aTjdoKUbUrdh/SK6UP3BdhDGukWGfbwnfQg/jjI5K4SVjbofV/ZgzXv
/QdSQe+cDZQ52XzQCYYjQyuY3CvJXtM1pjO2TL4H7i9eqXvJ3S0iKMwgaiFBRRliy/QadpdkWv8GK7KiGNIB37VL0uqaMKfFkwOf
rbfUfRKPiiJjiTu6xXaNQkY63xL0r8fgAyS2OCI7SJlqpCGmIhh4zETQI4xeJAkBkRSBc88CReIDwj8pEg0PW55vfzWmVKJHiP1T
FuAm3ajVbLvIFWkveibgQcVkz1LocIOgRUfQ6cSkPnGGAWnY+Tiwy5c2HKmA46WvQRTaPelW/uzFOL84AWj35mPeYpfTtat9bcm9
pmtHYPjw+0JGasPsdM/ZbjdiqK0hwxR75Thzt2dZFIb88YxUfO6ok3mLMKqv5KDAsqoivUWIgzyyhAHwlPT2TJxEkGy7RdkD97xl
6DPvwO105XY3J52vM4Dm7snPlYzWhVqgauqfXKm/xHzKTcdzaEEjbjLXfSPOwfp4NXtRiRkV9Yr5rJKFw72T16aTVnCztEx/DtNv
UVS2TK86aikDWZzFDts+YL9ACzTLqQmmX5doLs62EEx5mWcNGnwsqtzN7yI432hpKZATON9MmyVbX615UALmDzo8jlLoBxnpGcjM
ebNzMUNwxBH0RudxC3RO8KTRmvSGyyoLObdt7O6wm7Rd6iMnuyrFOCc71UN2YTqLE7gHM3S1g77pxOM254gK0FiPfCePFLvHLY8L
d23BP7yAnigp67SYny0Ila8BuxFsepAV6wEo+6lbDB2wW8DIm66U4GXfpOeaXjnOALsRJ7sgZZY1YfFkcWi61nyZH2Mk8mFu0We3
SHBwhTym/GFyjEk6L8pLK9UWUS6weryX89ciwQV/oPauRUBs3Ztxhta1jBAGv2bqn1zJjnFW1K55kCKshCmPZXlmxzhJGmZm2yvL
RXaMxtxy2OC4ZX4Nt1+knx4m0LfM//VB7MuK0aB42JNaiNc9vVo1+uiz/iTuUEi45zcrAfYL3YfgkIs7UOmskxaHevFl5DjN6lrq
RsPinIvQnM5bYwep0e+XuZV6jLY7cF5dRcqP5kJPJhwjQZG12eQzKAI9iIJpwux8Py1mn9+2SYtJo6TQEPuFTul1wnBDWsybpbhw
shs5tdDY5pNVQlyUCUpMsxu3QUfX3FMulCwzaJe8roV0X0Kr87GxpcfEZ8eqGd/fF+602LQA1CT+oyTwMEFulxQg5KbgyHnBqQiS
BLYnFkd7yuVJzmJ55qZ0kIqqUzEjIgQgbATXv43o6Ib6jDRmfZRiImgrJm624LDwtPDbNRk0VfWnm6j8/GWlSiRA+enT5bZUog6Q
RKlBhWd3k7lUPBTi9mQzZDmUOtRwVYSRMDo1QgilFjUcFT94TUHbPEgOrXsOFJk5hsyZ9WP1xCeHae/BDxDnorQrKfzwndi6pkZ1
k5lAYmBgjSqYGGNurVaqqln/UDXPSAY6QgYN6zujtVJzDiu2vXSC5gFiqJgfOjx5HwD05yBcfcEk8NbbYEecUZ2mzS38gcLWumQH
AE/bkgsaApDGcOJCqLy4uZwdp44V+B/NQQ1rAHeFbFzMQ3TymQXgH6ST6w66BBk2CTnioXvSvM155Cp9U3KQtvFwjXK5tyDvRl3r
S3xev/W50wFaXeFq6Yp1DnJ8/RBRzupcecosEIaEphG5VRal+jzTC9N5GLlfehLOYhw4By79mMSlb2cZHNbGGEpS7kpTIrO5eH/3
4zdL+e2P3//y8Q8/fv/jN//dj7//8Te/+vjtj7/5+MPNeldJWHcGXOkWXeni03dbn36THIdZOF6ycLwU3nrlhr28Q01inPn0fZGF
jZgFDJZukix+b6fcLc05+S2r5UoJY6mhWwS9/gS5zknLwFmiQkHKpnK/ICrpudnnzVd1yGyS4KxTJt/pmdJ2Fs7Zz92dWGmjAEoW
TsrimpuhioTTV/rzJ/xngP+W+p8LzbFvTvjVDmtem2a/KHoNd3iYmwSo9cJVpK43za/n5LS2yGoSQcWJCdLnfbTQ/jxh+nyG4xiM
6X+oD3x/S26HdULHbE2SpkKZJfl3Sb8NtGWGdKN06i2CyKoAcZc0alpOZIKa5SzV82HD5ZnFQosv5Vh9K3Teh5PVtJyDvmtdgoxs
Dtx4s3VAr7HMe/Vqj1sMHXtwRsINhupXD4V7HvJ/gx24QUdV7tFRBoveBk9g0acoyTnkF2bLYtHzZw18022/LoB1bvL7lO31Xz+B
WsbeuqZbigMcT1N5mR2AZ6TQ8Uj+xN2CnCP56JUGnh7T6QXJUxuSd5f59EZx4CaVeMAMWBCt1a2diLGOymndC9xUDuvipOuXpD6Z
3mKGDpBnp24Z6JwqXqPvPiQJbjsR55mg/k4bpUYBFIZJhjwdF8Cdv6k/ZBukrQnPq9YtTDWNe0VRcpMAF72Ogkp/cUTn3iTBeTXs
LW71pvn1ZJ1NttKNfJ0mEfTiygAJO4+D9+yDBKljHukm/ZY0pu1EqiWx+emM/3wk7JSJn+koCzlekDyD5SV30pGblsRhE0fpVw3T
5wOHNOiUbVFsGCaNaoxhqj0/uoWuxtimltnDW+LjyCWqIoGvSLTvkuIi096pOXgHGTON91pH9AfF9hVMVV3rocEQwsydBZSIw1my
7RlgSJgcp3FbpF71GdLHQhGf7eRy4cecv7xJz9SzPL+SrpCozk9EyE2H+Lo41qZNrr3z52Xvru0Y0008f1Cl65ES03QLcobnobuS
cAGwRAlYNCXHAYtyySQGppZRO12ZtotNFa55ECOJebyEdJKckezgkLj+FbnTknSTheAlC6F/RbQmTObUKdsJqOkK0ZvT99XrzUkb
Z1eyc3IVnF+b56zueQ79r5fG4WZOs847Nd5FteyWZQkPYxW5TZMUisI1J025uh30N4pm338a1Y6mEdO4gSH1MbCexewl+55Nov6z
qGn3tBBc+NK2yS6dOkvqaVr48kr2KU/OFarMJTM/uDxKrerJ93hUddf8ccFslzQHyVOYe+P1kEogKJqYraBxC/TttNu0U25SBJt0
3NpoHorWrkqNV/oIx3/dy3rumTebaj+rIqDCy/GZITBOooNMuyXDPwv7rJ9ShvZscXKFMXMJulkupQEzwmXfJh/fw/EMOD5hKxmo
mXXdchzgeCGDfBLv+EeyGqQzR8hntt2CnOH4IFYFk+aX5wQ0N2by3dJcInlocj40yeaqeHbB5UksLGIJ3ZTGqcvfDElYvv+YaK55
lds9q575xneOa3luwi1u+bYT8e2vBnYMaRRBY6i0DtGT6GGOJG7QGeNLO/j4kmjfJMxlRNrU06o1CVBNVJkPuqf0i3DVcVULTkju
V/+Z1Fnm1ZYHAzX/ZS0tWyuJ9RQg+4uSY4kWsZUzyXHEmVQddktabxEo5smG+HQk5sXLAXyV8xv/6DPziN6EKXEIo3SsBurTkVJJ
at8uO07jHyTUG3lNvCFA9QEItHxKwDNqiMct0TmspzuUkV2CXPRq1dgSgEVkwNsynm6+az10KFK6RC1530GgCDO0XmSJLI07KAe4
fmlQxZI4P791LL3ll9HAHiporCeprI2zTRBbNq6ytFYyUaPsW5Lc+QTMfNy0VJW1tTkVdJ8FTmdIMTEE4Ck0PTr3imuzpCPB4iwl
tGW7EgC67MbIpNTWpqMip4xFTkaqWodslHK9lt8OyS/znE/P2Py3fWL9+e/0vGBpaYQ2ZF32cJ9tRaNXIC9KTW9xXaEtYiyv1toC
CUUas0Fq50IUxExa8hphXD2NEURJz7FAekkG6cEtUE2TA+ZLavJWVlTbbn3q6S2WYU3F7RY4OLXiVgI9NOic7CveglrxRug0gEZC
Q86IWnP7elwv2hW2JU9VtopNkLpDLOn4zEaOK0cC1qdMNOS8HsT0U2nTNr+iZKIYCIGKY2UhtHdAaD8fZMcuDtTU31WNVNwcEqyM
8uxGcTNEO2bXzh3/UNvjQ4Cq2xgYiikw8ypnO3SVtNfYSL6qXlJu+lKc7hTdmtNm2P2OofqC23id/QUvkB2kt0/Mhvl+QW/MhSQF
cE2pu0/+kZfwiWv8jMyHbthR+X0M3kGyvjNWOHlmUz37WHbQLw0cCcrvnaeNB+Oz7+2Pv/vxtz9+88vHf/n4w8ef5v/84eOfF6n/
8eOPi6zzT/iPP/6+qgbXgcvgPD/cWSzmTmNF2hgURSYnMiEgE5lIYuiv1Bb9Mu0NCweRAyMJ/cVWDmo/Wh69WOUtd1pyf7gZQeiX
Ry6jw3CCkwS7AA8nO3k5C1/18nRuwwn9YhVjwwFLgBfMFIFmzkuGkIMUoThWItpJJAKRhOVEHGjT9lIVMuAUfXsVxoMwUeLcvsS5
nSSs0mhhfv0izBJ1cNDi1kFCLwlPwaeHl7DRzeDHsihhOdvAberruej7RSnGiAM84CXJyEEzNy+hiQjxqsFn6AkLXGeFer8kYJo4
ScbzJYrmxMHri4M3iod38MIUgAvnmcnKeaZHb7fVPCEr59lbyEiacfDgm4YYpQgXZ3hm8zPvZrYMLPsnepKx1ZnIhfYmL/kNkd8C
Bb4rWi5hcWEExOmBIh2MXnqPZNrmBhMkNrWELqSi2CDpkzUOSJ/cmyTcv+2SeBHlLkR426PU9uT3CLV/SeX1gscLuXGj+JXy6Pdi
Y9fA416OGOSngEJ2opDje9ZJg1M0QXJGZnGxJChXCNDRNYY3baL2fHwyCRF091p8McK+LqOPxKfF5SwMY0svq9wvrbu0cFzA5CMH
/l3o9BXBwevHCnVg4xBCd83XK/EJeol09ct0ZuNI7IThLrAQpXpAhMEOuKHuwsrJSO+n92AEE9WNlUe7lvZR979aOZyBuAqKH0rB
2dLZzQ3eQcXKERIrC0DeSq2MhQ4Z1g94x9y5lcOalSPieARigy+dpgmbOlmNEGVv47AzUDGRI7isHyPPZp4EXOM2DpZMs3FEDTpN
DUZJ9hm8Y4qV44B4q1x7d9C/O40VR0EI5rQVmSRrmLGSaCZOY0+4fmE0DOxDZMjyQTqkx0jJ3ZfynxBp8PnREcoCmwxboD5cYgfl
LJXRJ/ehMyWBg+bB7N6CBb6rSk5rK2QzGjkRjJz3oBTVyPHYtTxk6DxscHsTdC23+U0L9+3YlQauPajJFJXHozXesYXztf49d8/C
cdrT7uBwTW/aPA1KmakcroUNxkE1B0M5R4ByDvumS3lg47ilCLxEc91CuSSlHWXw4UV2Tya2h1JaMlRN/6GjWzYO6zbOSYFFv1AH
Ns4miGqUlm2ECQJurExnNo64RRj8IiyOEQaTkEe4RujCxpENtFAhY2UDLVhd89+DT5VyNZd2hAE6IEcJL8mJT1MC+n9mHivW3soR
NZll46CJh2xbHr1rewMnKmXfFUROAw7QTg3K+WHN6Z024QA7WhrNwnEOasKRT+cxslo4MUJBZ7SDJTvTzF8BF+jCwrFy68HgspLE
ay1snB19on+qkUPXRs5Xgii6MnMCs8BgaOjLDCA4QXOHaP3ga68jlFk/UWYp9Jzxm6RRwejDZrZTcqVLivU8cUz5LWjg+3nCglcT
FoAn0oxGBKc2TiYwYh91Rc/uHQQbbJjBI2PftHC7591qjhl4JuCVmN60m/u39CcK9Zdh5VCllfNZlwUubAhhBmzzDA5jR+FN23hk
5dhU4tRLJur8jgQwc9bRh06a32NnJG11qRnoB358beXA7uL2OuwXEYGXKo8VSrdySJxxBN44cMcRRnLSWJlOrBwviT0eAl9eUns8
uODnvwcvlxrJN6fFifD82zwANPN1JCdDvpoPwAgH5e8uAY1VHiyWEslxmK8mL6uccwvn3A4/53tDB3qSaCliAQKCA1Az1+ar4anW
wkrjj7SWr2YCQbUMift7HXk2nc4MzSkGOLH43NL52ugbn1s6X2kp86mJE5Q4YEX1b78kmomTIa//Fr9Avzh6JMcHIGEyBJEcHzCb
CVrYUbKDj7KKUIybElnhvp/XxkDmYxl9LNzSV0EcsWa2gXxKb0EDmpVDaNar+cYYiQ8DEta41s6JD9PmscUxQkl+RDKFmIFMIebg
37R23050X9ZVHzxcQ/ygXGvsNDX+GfFu6KaOVzrcV/D8jVsjPWHNCEK3SS1hLY2xHc+K0L1pA3VXyaz/HJVwMMXJ5lDMHBktrhLO
xczxCxuw77ZzTqv7HbCMlGsAjsosT0gGV2Ue4Ks8L/Z32G2oYHgDtfUGeANNgHr/OI1fsxOrxwIdgc0b9mJMGMZg+qw87HAR92oh
St5riLKnUTJfI+SZLs0khwulXNlEU+lbu3AEeKmei6V6jidjhRXAchou2d4K8lL06KW82BvsnmKEpM0MF2lvBjmou4AddLKDDnbQ
vWMH9xoTm1cecGSDfeZiP9o4JhHYlPEQ8LBaxjIeQiJWB6xN1o8XTlHliDGijjEQno3fxb2BxIIvSOjvWBAGA9FW6DcGDpkGHDDE
as29qzq3DZFIM5icw5axRqnLZuSDd8OFUs0mihnMJuegBUXM0IICCAh8eMNJ1127djbYNjn6tpSpyODDsTtbUTagY3f+AeFdyEKx
nEJSVWQQzR0y+Jry+2CPXu1jE9jHMUL/QJvAPpZtDjOCfN/67XWBx+uhtQrPWDnt3iba/tEN4LWL+wQFyG6Jb9jVIyMKXa1W1wWg
0S298cBp2ExcnZ9kTE9oJs7OtNTYPJHZQhqe3releoLtbNqZnMoaLp2LfAk+y+jDax0WDqrSMtGlKXDaHMIXIf/XWfz/fuFm+NXH
nz/++KuP/+PH75evKooDAtSGnBVJbZoB9bsvPHedIm3MKGRqRJmiyqNxHAzplGlvPgnP0YZCksRHSuKuZam34ZfSpE7BLikthf4E
6n+2CzVy8zSawqVmEVuDOqccJztZ6Eue/MjdU5qg0AuZv8YLRciSlQbKozUsV1sMVlSE956fHZGPRVE0z90hzWa3KHtiNROBD9W4
IP51g9QG1nihNrAURq6QxrPclPPVKYdGt3zF83YcDukURmFFNaeN0w4oUTvFuORfpjsNVTqF0fl8w/4qffbuyJAhBSlwwYU88vAe
NEv02UZhIKUkvReM2XgljSkmmDWzTjU+jlf13yv6dnmVqwdTqd8gl7qlBvteBrUcf972DL0T/TuWbPdwb1Wa8jJ56KqV3iCS1lGx
qequ+1W4bshi6pnhB62OjowsbZqms9qbJQIycja8Y+vUd2I+RLkEkFOacijMptlNmQpfR+b5h8QswZXZouk89O7SStFzo7YrB6mx
ZqA8ByZK2NxAIT2CznooE/WqZFdromx0ohOoQKCiDZooIwW787i/yZa76re+FJDmK+MyTSUfaSkyNTxQJsVC2djgwqiJ5dwOWSxH
nm+tCTtIE+80nO49PTs9Z/DwiEsbcIssy9iDrBkoWOOwF2X+K0Ab5RlRxYECnZkm9k7/l045LkwTrRNMBvqINFCUndJ3W1Tkv2xR
9Lbt+wyUoPTHi0MlUSmgMxgmzMwaihUqDT9yj3SwkaYgHeVTnmwuJNCLURIehsjDLplHoYX7wkVJicfrd70LJMT4tBaQYl5aP/Ro
nxklJKRSZnIMjBLko4WOPx5ix3mbhzhqzb6dOrdY4ZLBg/8GmHZqmGhPAkMqmBv6JNQbJm9x/dU2fl+c7YxtkXISQJQSIKIEbuR3
3MAjw8RKVmYwz9RV/8gCst55qLyxKcXixw1LLm6noHRtmvgNwNUfCkj5itNIkXTrBBMBrNp84LhXRqdAZ6bJxhQgyXWnrJom3W7M
q37xdG3D4RMfhu6cdhXDJnrCj+u3SiRXMYBxkl5ZDTulcoo1sOk9qcG70kmp90Wnq8gJiKJx6co1G3vLLiIn9r2hgcsW89ZZsJGi
FYBn2YnzeX5HBeU5s82x6RRrr3ZfTDctK1xMt5ErtLdQGpPBO+U4jZtoPkmJm4y80Xrc5ItAyGU/eh+MdDbnKHmQ3mUoLHdA/xwS
xZEnV0UdYaEXfEq2dIiz5VYt+eTgDo1TjCWc/RlucX68gq83T8JBCqJ5g1DajobkJd8PQyYhRSAx5ih5YHHsjp6FTTb9HG/0ax0k
0shOl93vgm6guKucAILn4R2HSgNG8VFlvFJjJCeHPQLbEoOF4q17x/6pzWxcmu3wUiptlyhOZEnzigvzSWlrRX42zcmWfhqzxs6h
0zjn6zSvrW8sq0zBkB3cbZxzVZrXgeUUMfUMMpZHynRiqGC6uRTxkLysLC+rd0OlUnO89GS4d0W/+CrHy06lDfqSR5tEnnIZF+Ln
JIHM3HsZ+SKAsmmHbo0aQWG05Eau0YWhogWgI9hMPPL47FRe0B1yHkTxo8w3vuydGZH3DLAB9Ltmn6B1TSIaKJMWRMFoqWrHHbSu
6RRlb6VY+3KIjepxNuJxHrld5zhAM1gwmDJSkstEL36vEcfXLTWB8BcOMfRjtG5cPIUvrZUlectSgoAKBTFXaMnXoOIkDbMtw5Ls
+snCmsbr+u8VmkzDIBtNFt4gl7qj7Dx2IQLGecMJthUyvVJ8h3T7B/wlx9ErOY5y9N8hkvZoYndb9WkYlu3FNTbLV9p1XJvwtfSr
FZjkhDxJUNKsXFjc8OYt91D3b5gpS/vNpfO6c0/fS5wVXRZvs0lTitJ8k5b3L/eds9Mqf3E8b6BLlG2NkIkivS5TLxA/r/EXsTYh
Tr1LcBKrZX6M3VixTgyXhZUBMQSBnceQ3hQcxKPsaAHP2yqL9g4AQmUfY++TcVLXj5tY7PX576Is57+fhvpy/J6XM8alO9NYqfZW
DItbliGzHlzE8tTPFujgXdtbMY7VNs/gmhIzhnqV9XEtPyQawdnWmLIwSc6Pfg3UwAtHaXBnEvA4WxMjBF6SuK8dxcEnaa+gGwnF
+kXZWzW0eS6RAxI6FkuGGJtp8Orsu3RvIlN68xyxbSyNXiPNvnlxhaklEAiN7eg3Uq9+YIjNZCjI8gYr9y1Bb047/HTrFs8nfcDz
8uUZ+T44V0sKWbQW+M78I0Ka1+biTM6+BSsoVk+QRMAg8D3I6Q9y+uPw039q9WQjBftsE1g93sP+spesVp9zfNPK7WPsYXMpvOoQ
iVgX9B4AqEVtjJ70au1BxvbwN0Q3hCyobKcWunrMRHrTemmAS1TSDKxKAlecnrbQDOqfMZH5zzfdgqNsVGeSMAlwjIVLYN5ZL006
F6bEQNJx4DNzbhPs/R9m2X7z429+/O2Pv/7l43ezWH9eRDttIoq2j4Wjr/Wur4rtNsiwsXSsdpiA44bRpW6kL3yvEHu7pkhCwnZF
4hclMbjIIBLtlaQ84larP0oqN7OAKpu6BZALZLEXlMLtUxUDapCgWCc/6zTQTgCwjlQPz0HZccv+f3ud2wC/pVr5URRk/9y//mlz
Fz33E+YuloPtzNpqmPupRm1n1X/D1GAM2H2eGN0JuzTMXlChRV84VDltqkUYm7msH7mM9J49UM1FmiTnL8kBTAb9f8WWNaPU4Hfl
/TfA6ymPnzmgXxgliLY/XiIVHurR9VIeP2pFdi/i193OHQr/Ca/TC96WA6pQlFfUSXasgg4NnBJoMMWVuf6p9Q3q2Q/txs7GLJnw
TACUL2/wK27GsozdF81dYmlwYzpBDo7VvKcm8OCusTRWZDiVmtFhm/ZeIU6wNItpwWJasJrcxKFbEuUtfbtt4y6w9IzQYxA4H4Pg
eUwnLBmqn393iqEA6rfrFHcOqIPiJcZmD8Ig0n0Ivn0pmHdXgBqi/XoSEhZK+E4B9nrrrfDSnaPqrLSNr6jlaZh7pyzfimvdT0XV
7gpVBzhym6R4OHQBTl3oP3aael6SkTPZkkstn89U5fUzbkczjI7RjxrIjmoDrIjOXnPgvOwQRAXZj+qTJ2lxZr0uZf3IZWTU2vzE
9/IYbn/ds3UEt79OghO4/TPQwynmtkt5RDRCM1s+HyGJ8hm3oxlG78tH18BbdI2DDucZ+ZbLjjXpG6oA3oJynMAcJzjHCdBxoVuI
E+DtwWQGmxmMZrCafbckyvvq1OBQFclQgwDa1Zl3OUENSoIE0ARkUMg00Hs098AbQlFKdsQhQUXD3HvM/Va3CZ1j7q/73RrmXlL5
oYcwQTbB4yPiiJ7+3yDKF3uN6Bx9v1WN0yn6/rqpNfT9VluTrtB3ftSUPPZ8/Xhs+/oRcSSXkd6zp2prb6ecpZuwfD5SG8pn3I5m
GB2jKTUI/vYgH9VC8GCiZB+tH6vNZIDmaP3IZWTU2uyez7f6eqkOgr/ViKUaCP51EpxB8DiV5PXn36ut6B1kBjmxFr0btSnqpeY0
sRG7Wj6fudfrZ9yOZhi9Lx9fQ/C3Zw1wBQQH4gs6YCMCotReIc583yIJQ5kIZFgTMqP2SqK8r0BGdtSIADJPewXQg0Ueq3chXOQh
XuQhYNRydfgcgr/VC8k/MY+EzyH4W3UIX0HwYL14tdaPZ4cmD2wr1otXa/noFGWvxr5uC/YQ3Epxv7Vq4iYQE9luAb42ZMxXOPzr
ZlfdqkhduX48e8ElyGB4fOQy0nsANZWdl4ZQhapGvj6fovIVN2NZxsZoSgWC/xyVrUJwrIplLHTGcn0IWrDzo5blJ76cx+j7rc4b
/unom2vzTSxACAsQwgKEsAAhbD+EOEffJkwpxlJZLp+PiHr5jNvRDKO35TstYpVjK0gnCdRJgnVSxlJR0y+GjsGTnN0khzfJ6U1y
fJMfIccJDLcR3AwR/AwRHA0RPA2RR0i0f2u97I6X3fGyO152xw/ZHeU6xTTR0xv0/Hstcyaojii9eT7/7pdkD8pZNA+L5mFzQM46
YDX24NxmrDk1t9rZNZ6Kb38JMuzBOgFVDkHX3MffEf49P/+9X5C9znun4/a4fPTrYoiH9aJfl5ZwViH6dbDwuCTUHvT4XT+0Hr+P
j1xGBpxMPfJNS381LvZc+fRm8xm3oxlGh2lcBcsHsTWD2JpBbM0gtmawA7W/ntRiAyS1PD5W6+vxEXEkl5GBK7R7a7Ens+AS7Mns
DwkA+2TZP3VvT8i8KuP8uhTZitJNxCdG4IlCwhLF+opp6BbpYD9NOZYaFfj0ZvMZt6MZRlHE/+XjDz9+8/FfP/7fexWawBHJcrXZ
qn2AX6729ZQHjvSOTPbrOU8Ae5BbGrzG3xKg36C/O7HiA+mIGFzPp5m0NOMnmZGcTElO5iQnk5K7N6uSAN4RDLiej4YmnFfs4u5t
b3RW1Ey1h8uWjCi39WOlYnl8RBzJZeTezMOSIq+nGpa6fT3VXv8orcUqiPuvZ9KQbaNX7XoyDQMlFySrYP14HJP1I+JILiM3j4mq
uKKfkuUSI5bPhz1RPuN2NMNokxYZnGpdP68KQrUWyYI9BXg2/tRRWW/1U37hm3CEFxuN9PrfqKlMllzO599POJIAjyQAJI3nV4/f
LHRgwgBUvrzBr7gZyzJ2KYl7a3LF9ZTjOTmu5zyj4ICeITIxNJUlmZhuTzy2SvB6Pj03uZjHz79Xv18xjp9/PzvV8r1Zh+VBXE9F
/+Gd7+BFcd/I5+iylo8iUPdGUSoE3UYef+fnv9+bVHEUDIRH7uswn3sH58X1TAPTga8ney/kqCiwM3aSRHf5elhaz6+4Gcsy1qQi
vo/lK6qfV1tsMl48xOvHeiEfHxFHchlp/OXfvuyZO4Z7b5vyCO6NtJiqq956S3durLB2sSItLzewQa9f3uBX3IxlGbuU5LqcjeSF
JCA9h+7S8khSzStZU702GmJWF6vJj/XyY738WO8OW+BcT6w8WNDMzN/sfHE9n2a+zEIzMHIwUHIwcHIwkHLwvVnd+z1w76k2q9jA
b2/H6xcFZTO22vRNC4V8uXxEHMll5N7M77Xu6evQH30Z+qOv9Phd1n5BvFqC1RKpljD1zXOhl4H4pZd9qe2Sz4eVXj7jdjTDaJPa
+AoXX3XlFkm/7vKxQsDHR8SRXEYaf/lbIWBlXVZjQnvN9Xk/BKT6vE/x+Fnw+Fnw+Fnw+NkaCFhfZJV5SlGI28vX2pR1/YqbsSxj
l5Lw13v8+L3MYddznhVLOWxHUU61TMwyMd+eeCw9wfV8mkkTgMgjAJEHdhUNQOQRXog8rmd1o6qPrqfao7/GlPmKvfv2Zb9K9fhh
f9r1Y9UvBEwC60cuI/dmfq/b76JI6W1bN4oW4Hqmd7v9LouNNu3nQ4DmLEHsg/KRy8jNY6LnJeb5xNlinsjnIy+xfMbtaIbRJt0x
uJ6oft43lA/d+NFvjUBUFge9bcrDJMF3PUfnhfcSJIsQJIsQJIsQJIs1QbIblT6Op2hD0dry+cAK5TNuRzOMXolTV9hjMXsL07cw
fwsTuGoyuGqLeTposaumPavdyfC7M/zuDL87w+/OLb9bSaWG+wUXDG4YXLHQMKVy2pMVgPj8+3OloYfj8+/8/PfbE+8xIslvJfmt
0H+c5LdSy2/dY8WeDht1O7pP/e4o7aicco8dmYDYcf1YtRIBseP6kcvI7cnfqiAqCmdIPAQkLgISHwGJk4B8w5Q7fTTQWVlXETPY
MK4rgHkTqKysd6GlV1E5wfD5yBwun3E7mmG0VQcp0DLKDkTZgSg7EGUHomtXfzq6NE56L64fz57NTno2rx+5jLT//v0LJp5GJ65G
F7UGmS52/H7FFu2oUay8cQeoc7Bb+U4lyqxfUxJ9nJLo49Kv+Pl3fv57+6IfsEbEEjcsHytnRCyRw/KRywjK8btZjr/++NNnLczv
7/YGc+IccOIdcOIecOIfcC8OgtqJddzZg0tqZz6rQmG1i7B4K4K4KwK3Ta+4JTtuWu2sB8528LUrzXcttgNtmnZsgLp2VhqVBVm9
p6PaedVP+OuvnXBYkLp2wmHdumonHOW3rJ1vYNi6dkoN5EQKAnLWj8fZWT8ijuQy0nSIdMRJ6G+HzwfiJPS5w2eG0Q4F9H1sptDd
2fdbMibt5PYqjGq/dXfiYUGQ+ns3lPjo7u/VEx+d6H8n+t+J/nei/13XcqtXMMWSwvz8+5HNG0sa8/Pv/Pz3SincWzvR1k58UOrS
YWnUznxW8AJHDs4cHDo4ddQ2/VgC5tpZdb9+ksJEsayef0f49/z895a5hwXEayekr8K4FzUx4wHZZWWMZZelJPrxsZZEPz4ijuQy
0jL/MKrq2gkVYs8vuDqjKi9r59NgaGN6W+2Uwypoqo+O2vk1T5xKq1P5Wtnz1q+4Gcsy1qGFBofQ786uLb8QgQkNmJCACQWY6/rZ
395vlru3htLrL9VXYUxX38E1BzlhOcgRy0HOWA5yyHLoWnP1ztnJSu9W+VpdrFZ6t8pXlrFKea4LbXp4CGsnPoCc7/H1VRfddNBR
1k4/tlawdla9ACc6id9FJwG86CSCF52E8KJrmXss8U7trMOKcao39tvP+Jka7owOqCvXj9Vp5YC8cv3IZaRl/q/wfNL7COVrZx2V
/FU738DkzdopVbYeLCZOWEycsJg4YTFxeikmrj5JOmfP/Gt8KO5P+XxYx+UzbkczjHaop8FlPXdnV8PubOVyrx9r2P3xEXEkl5Gu
Vfj2fqcB/Sz3Z1W5zxf9Xj0OWXLgHOTAOciBc5AD515z4G6vuQpN40Q2PA+efH1C0/IVN2NZxirluS4AetPt4/f2Ma2d+U2dk2qn
/zJvKF9D0wC14QFqwwPUhgeoDQ/cMvcemo5Xsu8pEqre029f/AvVpE8HrNXrx7MUAVir149cRlrm/wpAelE8NN5B+JaORrXzfY0P
9LKc6EsPkcqAMhuGqfQxla/P316+4mYsy1iHKvo+lp7t7uxqHi7UFgWoLQpQWxSgtig437UA374oisRvrTevv3I6DH3ztTuDoTSx
8DOz0DOzsDOzkDNz13LrJX12mv/PFku/fD6s/fIZt6MZRuuEqqw/MmCKGDBGDJgjBgwSU22SVFYhZXn3szz8WV7+LE9/Ts2Tn9Yi
ob8WHbbosUWXbW4WQ6lI6iBFvzGxlgfNU2HYeP79qXTXvyP8e37+e+P0SnXSWxIjamqUOgi47uz0Pnb/llzcyi5AZKENEFnoA0QW
GgGRhU5A1PrU7DXN8CBgTe2S6BkSRUOiaUhUDcXmifdO1beEf+qKmTr6oNyYWoVS2JQzYFPOgE05AzblDC9NOe8csaOC+ZRLvB8+
nwXz62fcjmYY7VNuWvceUXBBFFwQBRdEwYXUq111tnRvIffi8bGmXzw+Io7kMtK7Fvsn8C1FJ3cKnjrYYW7dzfFlTw2/X1P6XlqU
PP9eeziWFiXPv/Pz33u3Qb+rPKUA5Bbl83FXy2fcjmYYRbn+84+//fjnjz98/Ou9MqieHbme8iA/4L8x9y5L2iNHduCr1G5WhMX9
suRidmNjWuRCu/8d6glI9phsFjK1qSVVm0nUiNMjyWZZbDbVI7JZ/QqZbzRAAgg/+OAA4gZkWRmLiAx2u39x8xMe7n4ansqvZZ4F
ozZk814L7vvKdC2Pj7KG4utQex1Kr0PldVMmsG8IwLW8bqTzGXPX1TORI28PUm0I9Ca4NBYrNTc89sTUUya5Gx/ztai+waXX8nqF
ol1L4rBnJcS/FsbGMmpLxbuWxhLRqC0V71oaMfUUrpWDR0SnQqRHxKW1PCIuLb/pi9RXZUC+9a1zly+XdWFLKF6wNBYXtoTiBUsj
pp7KX74/mRoiZfLldnu0z9lNXZ/r838jb0EDmNAANjSAEQ1gRYOpHGE2M34K4rKJjye1Pk+T1PKbvkh9l5qoWxPjr0UePMs3MBVe
yzyLFaVFbWlVW1rWlta1laWC+8bAX8vj1zNV54xQnTNCdc4I1TnjS+DJtVQGFvY+m9Utr+8ZE/h2JwK9TD2qyli4FtXtcf1a1B4C
VgYuXIvamY/KagHXkjpmtV8L61aMPWNd8MXY5eBFKtYCzRm6p6bf9kborTIRnZPW8+XyFQSgWK7HYrkei+V6LJbrX5L2C375c8fO
M3DvzsSh/N/ImccA5jGAeQxgHgOYx1BsHi8qcarBCZkio6k5W8zU9NveCL2X6lznCrW8m1yL7F+R/VrmWRAmeeUMeeUMeeUMeeWM
LxXcN8jnWt4daejXUhnM1/tWek8eUMYE9ko6zxHFpJtDSJaEkCwJIVkSQrLkNiTrWui9J/09VDzXonpR8VxLuhv9XabzAFMnEHUC
TyfQdPrSxcFXL4pDUClmK7VmX//a8pu+SH1VxoLLH2+IH8mXe1fQasEvvxX93UvFk7N9ePTX87amvxT96YI3YD8oT4SQ1JxtZ2r6
bW+E3kt1MtJxer/H5iTi9EacT6TgXAt+4A3YXKM/WtQeFrWHRe1hUfucRX2RcNPTA2duKXOZMXe3nruX6TW3PemZWyoMXYvqllJz
LepW9HdXGs21sLsyhzOWyRErt7bSQLz80lxD5pem3/ZG6K2yHd/61hPOl8tOgNfw/Ds3lgmYGx57Yuqp/OX3P/+aW/m5c7ZUV2ae
/N944DoBZp4AzDwBmHkCMPOEyjV94G0fb/TapAheas7e9tT0294IvVfq5GbGYAAURkBhCBTGQGUg88ycGCkgTkhAoJCASCEBoULC
VMg+AYkKFFCggAIFFCigqhRgqCDIVejJVejJVejJVehdhUg+DSaVB1m/lzSYVBxk/Y7r34sFMxUrO7thM1JfWrhy8mZ07zuEhQTr
CJYRrKKqRcSlu0hL9a6XxpLuIi3Vu14aMfUUC+8bYJ4lkslzocuHptuHpuuHpvuHDhUi95apc2xoVm6LJnuoySBqsoiaTKKWFULZ
zAkZCewsjSUmdW547Impp3wl8f6RMCgbU+Foas6QJzX9tjdCb61J+vYFp/NpDovUEB0sNUQHSw3RwVJDdLDUDb9/d5a1JJQViWaO
lc5vaNkZKx29DYVpKomXaf1e0lQSK9P6Hde/1483/zQth2jTZROa84memn7bG6EXNfpPHz9MSTO/eP+b8ePP4z9//PjVp4p/KSTu
aXg3rtChf8BihRJn/kxypBpypBpypBpypBrbrEnfoJwKBbhLmxkMxSfN30tegaHIpPk7rn9vVKNbBfYK2brr43vNItg7Jeru6lWy
GVYgpBMUSCcokE5QIJ3gK5lIhSrdXs0qZHdzrlbI7hW5VSG6Y0XNCul3UVnWrD3WZFs1aGvS6w815yM5Nf22N0JvH3PZudZRgyJs
BIUMwAclA/BByQB8UDIAH5Tcxki1jM0eWT+2gY4hduWTSdUm7prv0zAKfMKEBqYiDVRFGriKNJAVadVrUo6KOghDT7rUXIs6LE2/
7Y3QW67fdZ5QC9d1hQ4HnuLbwaf62eDwG1OLKhTgi9ymc1UBsawCYlkFxLLqlVi2Qo2+OekVCuzB+K1A7B5+pCrZTOwqRa5S3CpF
rVLMaqPsh43XFWNSg8u7QoFn0fddPEoV0vvGwtYsOz441k9VCdOLLTXnYyc1/bY3Qm8f0/itLylJgyLcTEm8J0m8J0m8J0m8J8mq
e1JmSlVlyFqDDt1Kb1Rt3a5htw2jwGFuN1hFd0Or6HJoFd0OraLroVW9JuUoCz9GSMKPEXLwY4QU/BghAz/Gcq10RihG/XNjhQ5f
5fHWD1CKVmjSN6S3QgE+TOlJDb4CZOtbKkHVzH8v3vsq2d9/mexukcIVsh/2buuvw9eXqWXPLXS2zlQ0dKlbGktQ+tzw2BNTT+va
44yyDUNQegXa1PoETanlN32R+vpYxW9f8QCanahGCJvgNWFrAta+12C8fd25/EWebH1jKYOGUbgjEahlUviIkjhYQcHN1JwdJqnp
t70Resv1u85/ux/a5iTENZTdqlDipppYFZowB+rt4NJkeLLBkQ1+bHBjgxe7UQP1rA/5IsHu1vf0i4y7+x8xLvLw5GCpkHtqzOEs
lgq5p0ZMPY2qPGzGzC11Gipk90rdqxD9uDv7OrnPQaH6pbHkljkoVL80YuppXXt8tp+dLG5CENScvU+p6be9EXr7WMrO6X8NirCB
2cpR3u7SWAKzlaO83aURU0+vsXn7OtvRvXxE1Sb+Aghu8mlfBQAJAUhCAJQQgCVE6DUp/GuVGi/DPtUbo+YMsVLTb3sj9Bbrl5dw
GOgaHegaHegaHegaHWqu0ZnJh4GOlkBHS6CjJdDREmIPPc64uTxkQnrIhPSQCekhE9KbHhr1pemq04HZTt4OKR9h/V54G1Mmwvod
17+3a8KQdzVkwtTpwCQz3n1hzMluvPvmnJfuOMJvSnecG0u649zw2BNTT7s2TKLS3ZH4OQmRDTQndTrsMyQfHwaWDqwh+bZOC559
CpanxeVpcXlaXJ62z/Lk40n1IK1KEICa8xSlpt/2RujtZns784W16cKnXRpMgDWYAGswAdZgAqzxHUdof+renQFXkph5t68un1vs
7iiJkvzNEYa5ddOv359IZfn28Pe4/r3jPB0FkktvJQSSL801kHxp+m1vhF5U8befmpXnbzZUfrsWeeBcJ5+mIqemIq+mIremcqUy
b/KlXwvuG/B3LY+74QrKvVy/F5dtyr1cv+P69zKpfWsNX8vrFtGdMYG9EilzRDGhJNZSXdOlsQSUWEt1TZdGTD1lkrt5tq9F7bFw
ZZTntahe2UvXkjoWHL4WdleKY8YyOaqPrwVVAKHmWh9/afptb4TeKgPy7emz9QyEOp341lIjcQtH5BaOyC0cK3/52y+fWm3dI6Nz
dhMPH2//jbzxVGQ7FZlORZZTkeGsXMkHGysMXqqYNlZqzhsrNf22N0LvpTr3ko1dizzIGWxIRL6WeZYiCCYCbAQYCbASplRw38C0
a3kH6znAgg6wogMs6QBrOpRJfQAMqlsijzMm8K0nj1qOvD0iJEcVeanIRUX+qcJZuxcC3sM4di2qV+jCtaS7IeBlnp1FCGgRAlqE
gBYhoM2BgDkJduO57zzRT6Tm7GlKTb/tjdBbZTY617DIl8tGhggHZZ/nxhIZMjc89sTUU/nL33751CY6hoC3iTyCgLf/Rp6Cwhm6
VThD1wpn6F7hDF0snKkcYZbN2Q5Bu0gP70treXtfWn7TF6nvUpMM7rEGWpdrkf3ZJ65lnkFAWmOGFpmhVWZomRlXKpg5sHqDo8vM
NAupnBZSOS2kclpI5bQvqZzXUtWt/jn9HPq7SDPr6bq6zCrzFBSUGkv+CAUEpUZMPWWS78WB+jlX4C0ZY9eS7saBl/lgTidrlRqr
JyrZq9SIqadwmfA40E7uPApdS80ZB6am3/ZG6K2yHU+4ArMzvGwAItqlsQDxAES0SyOmnspf/vbY+XeMA2/cTV1rIuT/Rr4EAiVr
OUjWcpCs5SBZy70kaxWMME88EQfpUqwDNOdKqqnpt70Rei/VMbcW970WeQAGG3JOrmWegUH6sYZ+rKEfa+jHmuIf2zdo/loeX0sv
OpIYHYmMjmRGR0KjK5P6gD/wIl2qJ6y4yI6qzMrJEfX9LTxH15LvvfZfZDm1BC9dy7sVFpqbquJeC+NQiXaBYOHSWCJwXSBYuDRi
6ilcK3xUkxiUEsmHQs1526em3/ZG6K2yIt+ePmDPYGEPxpeCX/727KH7zDNxVlpRz/uqyX8mDhBjFSDGKkCMVYAYq1C5pvlqWHY0
y+niRa1P85NaftMXqe9Kk7z0oEgHWqQTLdKRFulMixmHWm42EPmWAjmXAnmXArmXgq8Qe5b8E8A+BTBQASxUABMVdIUCTAQ2iXUk
1ZFQRzJdjUguYNalUIjlc+HWUcSto4hbRxXLZAgbCPcbwv2GcL8h3G9Exc9kcnYa8mXyJnNfWryhTHOmSA474muVx9cqj69VHl+r
vCuf1b4EZFki++bbZIncx7rf/SPZbJrODzV5yTNU+Tw1Fn8XVT1PjZh6ylcSHyU1mrBgks+RmvN4p6bf9kborTVEHAEZzYCnGfA0
A55mwMt6G8i/QAdLQYhLY9nTc8NjT0w99b9/n/jSOYy/JM+lM7TPT2vpfEctymJxFNy/fi9mOQX3r99x/Xv9oB9lqwVpAmSrLc01
W21p+m1vhN6NRlNWzajRn95/KktbaSm5mSOUh59N1jtD6pl3soGfNUd05+iaDIl8+EX9Q06GyM7uyQyJe9BZe0fOmcNetVHzhH3/
nDDGAlRm5mQI6xezmCGsm1syQ1bHgqY54lgAqSXwpmoJvKlaAm+qlsCbqrdlI7IWDOs5mZgGkwml1qfrIrX8pi9SX6UNYZBjE3Yv
kMzGDEhIXpGQuyIhdUVC5oqM1T+8F7NWidB+XsmsHdXXL1nwO+8gyyoaZ57yzg7Kx1Sxj5pz1Flq+m1vhN4MhdStRURzhB7ksrRc
SzKk3vV6nSF6f4jVvrRmCDuIyABSCg+kFB5IKTyQUnhVKnePD/tiGnXP23XO9L31LOuaJ3FfzNMpquU5f8+wbf728Pe4/r1U7N1I
Ud0T1ZghrFuKc4as+5HiZY6LC8AmtTQWpBKAUWppxNRTvGCOaKSsjQFopJbmSiO1NP22N0JvpUnhHrKh6hwUnYOac1ByzlQbMx4u
+kAe36WxTMPc8NgTU0/1r78ZM2Zmu9x4aHTmWy35nbzLhTBjAMwYADMGwIwhDzPm57xIoQcvZdpr1JwrBqWm3/ZG6M1QSH8FZtRf
ghn112FGfV8t+RyJ/AJPWdDr97LAUxb0+h3Xv5fKVTe7+fSTwFHfmAadJXHvWIxKE/HN0pit0NLw2BNTT6nsfhkxGcK60SrlCLvZ
RaGfRY/6idLtWQvm6J06KOHgnXppru/US9NveyP0VhqXb32ZqEsks9MAkwBTABMAw2+rf3Mv4uUSoXf7GfXzmFHn+xkTs/j6vYCW
xCy+fsf179XjfBADab02FAO5tJYYyKXlN32R+jJ0MV8BF829xJ85Uu/Kls4QvT+6bnhvv+Qa0rS21+9lhNPaXr/j+vdSucw7dP/H
9otEmRucCldsQl0P5cuMGYFFUgQWSRFYJEVgkRTxWiQlQ/beEPSF4+Y+vvscifsQqK6ukcvMmdosqwxxbOqGhdr1ULke6tZD1fri
tcIaMzdoaSKF0y+tJaJ+aflNX6S+SrvyrXN5igLJvLtXBnD3ygDuXhnA3SsDuHtlqP71+3ttQ8XqEsl3OxvN88DR5DsboboOFNeB
2jpQWsdUD/HRrUzAlUzAfUzAZUzATSzjx2fS6ZB9D2TfA9n3QPY9uCq5PF6MtLgira5IyyvS+oqySvApYQ7GaWKgJkZqYqimr1KB
yZohwY7kOhLrSKqrE8qH6Cb0uH4vIboJPa7fcf17hWiGA4dggCYYoAkGaIIBWlX93j2GbEq1yJzZPYykmVU0s4pmVtHMqsrlxCXR
CA0V3zQUfNNQ701DuTddIblzBk2eUCaHBh7d4NUNnt3g3U1XCe3KGJEps3seTZ5Y3jNmISfLWsjJshZysqyFnCxrVc2a4iP55RCc
Sj5Kas6YLzX9tjdCb72R4ohmIMgWomwhzBbibJssJF/JJwqKi1waS2Dq3PDYE1NPyxi8/fL2+PeSnJr7D5rDrJruvoiivBo7ROCx
i8BjF4HHLgKPXVQtQ39EBxOcBjaY4DSQwQSngQsmOA1UMGFLX/qvPn41KfTxw3fv//n9jx+//sX7f5/pa95/zEqyUVD+x5KbN0Xg
LH9eCjOmcC3lhtiqxgbDKg5mmMGyMMMSzAitWuwBbVLF+jQkdvntn25+n4bELiP1eU7JQbYqk05K1WO31GhAW0ZhDQR6U9JQWRx4
VJUis2m2wZg1aiTMqygtw6a0jDQVLs2Eg4loX5p6J56WgifxtBA8ifft4pOxYMQ72pskPtDODINvF//9i/gpL8NCSaKgJORyWKhC
HSKhYqekadUlmS4FNPHJgQJhpRRU6immtHkmElpWjQ9/NcJX08nIppMRZdsku/lYBPSs9n4kzfiR8Ke7VvEJtamj6EQn/EF0YoQC
mF745kMZTTitQhEHI2N6ldRDDPQ2Sp0zljJDSItytGpSBN3Jen5jTEakHRLotFj+vBJBWIoC8L104WYtUP14MaUEk589UO14MUQZ
KaY0Wqm6jc/uLNVpeBSNjk6DY2hsTL+h2Z9jpIamc4zU0HSO6fZz7AWMk22HsUhKAOQjJVQ3wPeNxxhURywODgh5BHLZOVi1/dYI
v8MnCmsHlUHjIIRMfmzqne9qctBaU1zheBrq7ZNRpoLqGqlH1idG72RaovdPtSrB43QsFyoGy5QL1ckejeefadXiDKcD0yOpsmV6
tMT0GFpV2R+5ywgoOFuX1QtHrocjN7bqcIDTgdMq3SjXv6/+LUlIfbRMrYrskXpL3G+NBs+DdXUO1jUT+6DZ9zbXLnwP1aUJ9Loq
YiBv49IzW1wZI+EkpWLztuBMHA2EZbI5yMA1C98DdUl7MsJFgbakxKvCqGCrCoxxpadey8TAkGlt/vkcXL8b56hLwG4MeFkBrhso
Vx8lRXF7GXXzKmRtuR4xBXjcVBiCsMmWU++cqjwRKFFyqR2NbLShkxnlALtn48H8gSfY9dKEmzOrJMF1oyUU/Zp71qIBAjLxnFHd
RucUrituGXu69tpeauzPssqAtrrNzIP1Rzd0PlhXhgfrVBRVvqSRNc0Mu7/Hs9yFhNVjHLwx8Ly19M2XcT1Co1RxROrx2Im+Bgbp
S6Q+blYLaeyW3njYm/g4TK1qHGJ1S6X9Fefbx5te3VVPZ6J1Q2hdU5SfIbRuNFJxqlZV2PPWUj0xggaEECVAROnbh4PZRFNuZSAt
ooRcTEhVI79RHKRuHgwGrTe8xNdo8Dxa1+do/V6Upi/RulQQuRygfPjSs6B1HSlyWVrbPBDPO6T0OV4PTD5cYGN/XavwE8v6wE/n
sPq9nn19hdSN1UAiEAKmb2kC61YCiYC17VaTNeaj9Y4mAYw4WiOfqk5S33w0uSEKjFQZN5DpZEO/BqjrbKC+YaMy0vNsVFRhZfy2
3YZmd4pWxhi3KMFh9Hu92frngNJ1JkqfojgcAQwPDEAJDI+QMxVvnmBftxV7gNLHy4KzVPxhiq1JETjQOzsJ1eAE1Qcy4/oVpmZz
m5zYFwgLtCz1FaDj0KrEAUq3mO1nGWZUR1r4oXkozjA6FKC1bAFassdWVGF0c4HRGypl1GjAXXOnwEfyp1NI1PL3ZWEkXPZ5sXSt
ijAIXWGmjmVj9cCP2bw6OYROJVEtU6qA8Hmz8L1luTvwx1whdG0UucS0ChQtuvSsuWJWA8OSbt4UT/vTzYU/XYI/nXwZEHmjBziu
q3wZ5gKl729q+K5CILn5bOQw+t33VHOF0rHIhSCIbqGSbiS0Z3ssQT70RY030HQ0RjMQ6xj0zQB9vK1pTfGrZrzVBtnJfjII/Yss
OY/QBTKxaA9R5gKYWMZrFED08WrfbXROjlJLO9jBCla0gn0vNfbHmE1qGFLDwomu6ET37Tv5KPLlTn++yXWly9FUQNh1IIyhEWMY
rEDQbdXy23u6FIT0dDehdG81wY7Uu5ZDDDIZPjlCRKN1hYKnmZXkGlBcZmWkeYtwARey5gZ+nmoJh6AhJ7LASNSlYz7+0I88noW+
gz4nsF1GB0H8GmP/aJfLGDH6ryoW5TgnU3GEORSUsyXMsUSY02PJMPsrAIYPYkjYLQCGD6MdWBdw0IOpiSs7TtrEdw9LnGdpagwB
KKPAN6C6LJY9llcUHaoCeEkoOlRDqJKWNW+qx4md7P0uDJa938HTR+iixB7bKweFZ5VTlOq59Mw2WnlF0anK2x5Tw0S2N5RkqFRi
D/Mx/0azab6UcavD0GOv7GyzJB+aBIe81HjXoINVdznCOMgP5dSW9bgvpwZRfT1WKYsknUNfr6Y3oqVnTb4I9EbkhJI9VinvyBtR
vqB8UuHGpqQMttQ7L9owNik0dgIQwth+lpm5BniCmN7AZY1Apge/gfdDT3XYUCgJPgTrAlJxgA/BWWuw4onuOUpvZ+/gjn8HD7DN
YtdRYo6/r9n1/PUAvB0SInrR3aE2yvQcGx7IKMIxhmCMIhSTaoqowYSeU8VfE8IQhUrXhMn7qWJM14TUO0/cePH3Bqumj7eEzer+
t+8/vv/9+48fv/r44eOvy/jhWope5ok9KM/XwImYJ/emAn15wpmaUVBUB2rqQEkdqKhja2TyFNwawto1pIakouXrd1z/Xi6ZKdPX
kCOfJ5MpslJX/z5zQt8eFrcvrEIOMYpKJ1cYVb+omMBupbbyxHUr6JwnbldGoTIPNU8aVzvl1rXCVomjZE9CSwSVCCdVrBW2NF8Y
jEqFWqj16WJKLb/pi9RXbVU6l3Euk80OvAIyFgVcLMrycSVb5p7CH78vjHLntj0uiVIZwZK7ofhiKLduqvOqzgZCuClYOhK98Pod
1783jDbPu6OnMO2YMH9qzrA/Nf22N0JvlkoZDHIQekHHKYZe0IGqck/UDA65++WeIEgrMG0y2RA6ZKzAII8K4X15hvNkPoogLwjl
KmsD54nTz4q7HzyqK/AokSVMIkuYRJYwiSxh8oUlLE/63jpUcgPmietGQpwn7nYrpJ7Gko8QzGUuHb4c7fhLdGIpgeZ8pUxNv+2N
0FttczpzEpfJZsvvGe/Ane0hx9B4QvhLI6aehhHoVZ++TOz+GOl9R8wim+t9VVRPUYcUjvZRIWhvjYVa0EtzLQe9NP22N0JvlkrX
lHMtdVbzxPZnEcmTewYyyZJZMmWWbJklY2ZdjXAGZGIdajZNA4tQ18jkFrsHbkUP3IoeuBU9cCv6XJB5QT7X8mqcJ3OPNHvDE/0s
0rykoJPeAdKcGwvSnBsee2LqKZfejYQuT9weafY2DvoOPpE8aR2p6PIEcuAmSuC2iBK4LaIEbosogdsivrCaZi4d1t7FKQxYppOI
mrPlTU2/7Y3QW214vj30ZJRNSeeRFsgjLZBHWiCPtED+hRaocATefnnzzTGToO7WY+QppKnzuUaqa2kXDjQPMv3gvU9Yl5ozyExN
v+2N0JulUgZR3R03O/NFnkzzlW/hF3R1tzzRmGt+4zYC7zzJD4HMC9K6SjKwzAm9HWRectYBaS1Q1gJhLdDVZrp/LrjqesP2K7a6
2xfN7Z7MS8a63j73S846WjRsBhg5zCoWDM9VYAanyI9DTSs2Tb/tjdBbbWS+9WWVL5PNAkotCN0vjQVQzg2PPTH1NIzA26PnYnfi
utytdUAWcqfQM0DphkiVf+fv5aUyOnqqTDWrP78bRpvdeVMylUhl96Fpxabpt70RenNUyiO1m4K8BQR8U6UoeLwV8HorMq1IJred
3FgTNCdoT9CgyEr5J0BTgWtRgW9RgXNRgXdR5boXM0jugF0PyPWAWw+o9XytWGYrBJfoP5bPhc0wEJthIDbDUCV3jzixBIVnS1DQ
zzW1P5cJwGxgoc6f3LcnbhNZTHdqLsy6xGRildal4bEnpp4qBRijcoM1z6G8M0gylGaZHIOaPIPaVIrtynyVLZWlvXtidbHP7kIT
0drSWJ7d54bHnph66lYX75pRg5SRXDOpObtmUtNveyP0tlgxBsF6WgaeloGnZeBpGXjTZkT5+E4PXD5LYyVSAS6fpRFTT9s4vD3h
Gi6hwHtmNx7g2hueKEto8IJO4dXLpxXp09Nfl5xo1zb2fPSnGry3Mp3+qTlbgNT0294IvVtevomT70/jv38/5Tn9bvz48ZOoL5cJ
T18wW2S5Mqu02ABeDc/2jp4vlCpmv63QZA99kzrGojqGChBZelWw9KxgVQd10tmpYd+AFnhseNaYvRQWqNKCto8Gg2piqr/4ycrI
xZTKuR42T4tXpUvCyRqcbFonVSCVXiCGsnSOumYl9E4JhzrwvkUq6dI+I8mWaHp4gmVBxfMME1oVuijw/YsC47R7rSBmAEh+NzED
4yoArGOFb1YmmTbd6AWtkp7AtSaXKMwFLcgMF1WVAqtthV8Pi1Gx5XQP3iGr5APgpsMSVeADD6gIaPtyTCBPY82ACOwb0iL0DgDy
DFS2cE7F9uUI1l6j98jHYCm5ygUFPuQQ0mVzqhcQHVX1nU5OoWUvK/uNMSqbFcvnJaBRUdv6ck3acFM3zglM3XxerMwpFpzMUpO7
31uj+g3R7nxVuKC5Oul4qLhuiuyPNnNard1CKUrXYWdvwDu/uzljc0iF2TQYPA4RilZuhBdhrPJL2ZIT55fot3r5vT4R+BlHBcT8
uJ+BbMcNTtNu1yN4UiYheyMHr02VUVTXyN5tIDXk7NOTnaI3OxXqzIPKAPeGvWNItnRA7KDGCbK3YoPsqcSrIZNt6S5sfQd1mEPY
bo4YzSaHBaz5q5q1YHbUdI/Y2AIFMRaUGhnSjpo+o2tWZQ/s4xbXc8wcm+fzdh04XL+P53JsukK79L3RkbguaZdIhpxD15lkdQXq
qWiPgERowcZoKNmswt7ePTwGezS/uXA75sF5yw/Tvg525lZt0BmLPQhMtw8Bj+fFiblvBvPqCswHoynswM2UF0vYQZBUP2M0s44e
xL1rPyF5A29Hmy0kJW6PWEOo5FM3U31JnQy8n2JAXErcDuP/beWtV2XC+Y034oCo49i92qQNexPzlibPhEQhPFHqRWBCNB6ZEL3r
N0S7k1Wfomiw9R2M7DGW32xsvqJ5o6dC5cB5eXW+HAZLN40HBz5GfGpp5U5OLQJBFGTlibRjRB8v8VVtc8RudzNaFuEJmskhWpfw
vLGDoXUsrR2CiykA1alB6rrbu77E8y18PFVaHIB5v/E4OdZTH6hQZKyzUjrXU7/1PumDgGhLAdGhWR0Wz+O24l/YSAtZedfSV3h+
HHdDaizkegueB/r2mSJ70STY9qXCAXpx6hS1ySnaPh8clIfp4NB8AK+sal8QjAtpH8rX0X+kL6G8URSLIDyY4aVniXQZjzuIdPHt
RxcH6sUJoqe62LLDStgjesdwzTmGzSV0OKYY53wVN3qVcA7J26tt0PxSpa/AvImSFqKxShIe1BB+Ya2G8gfO6PaFyFn3OC43Kz1F
+LmQmCriROibinjGaexSXHOMU1WK2Mukfrt86yW/FOZ1bViguhl43i3vNdTwjVpCDV+gxvZKUR1DH0LHITp/9uSYIx04xE03RTgo
v3lzE/k8qpU7+wjKb54pxM1n6xmUj4N0gDqAfY+yb4cUajYFFJh+Ch3i+AAeVvGJzJNbfrR2IugE46eFTI9wbjyjVZV+5torv5m1
w6xF2xjTYTKAPM8elFUpo0qNMxS/ecaGeBsDjifwPJkOo8KeweI02AYOYCXqcKO5hvB4K/ZDAJe8w2qFUEgmBtOsCxNsI5h0L6wI
vIlWds0a7GG8Z1L4PJtv1i49z2+UwcBWKZ3B8NaARz6EgxgbpSW50pTTsVmZC+cVB+PRi9a+Nc9gvGZgPLkbXknbq8Sf4vgCKsIq
4RyOd9sImz3y8XCfbv/9rF83QpEYE0U88Ou64CBMo/2IPHh0N0N0PiWHT08xwSXrPkFmRdy6E52pJyfiFNGigu9lVb+VvO7qA2+Q
66YNO3daAJYXUBjE6gDptNZ7uINF1W+IdmdrwBGyjIdIAnjtAT6O0bxljpaOdwpTHGRzSwidyYXy4eU9KUJtLyxkZ6GArBWm3wSx
G35CqNZo2PDKOwVZkcE6rDkSg7bAv+FsVdTCabIorWWCKYHmLoC3k0IXoqqCjedpo+TeIE0iWaoISSZCgKmsM1cZKaSwmiCEzUO8
dkCvBmz1WAckjzNKyYZu3AcOvIGw2cEfVrfdT3JMNaREpZKg47dxlBQVVkvq7bCCuvFT2S6Dsof4Om5CyQnmGwC6BPNN3Ul8nImq
OXYZzVbUIYynRac1so/B2Xh2IlvFw+JbuumjB4P9A5CgSR0DYH8J/ntlHfrvYw999ibyxVrvi11hWEynNbK/A6iNGopQAyYsK6wq
2uWo3fvd1FWlF4XYpcsa4W4FU/YNmOrA5rqCK9L0mRjWWyy0PgjiRoTprNSQDhL6LFf+huAGIylqR04sYTaR9qnpHUwYyIRVQnmo
QW2M6GmnmUuCo2I6jmy1MyxlrOszdacOf4mXBC+JHNTGgIE7mNHjou86SvsTOb7AUY7lXOMaNz0VYl4zAdAI1ivmD7PL6zf/wQPA
Npjn6CwC2i7Xc3B4aCN9mi6vhlTufPwmcGOGFDMzLm3pQtc5Y88DPQ6JFRoJ/Og4GA2IUSrdH0YcJDVxeE7sxG57w/nvH796//v3
P04qfff+t5+f//j+D5m5t13KFxZr0L9cdrEKZ0UNoXQt3H833n9B3v82PZg6Ww3p8cXi+eLxQL0rWIJoePptk696VbgultyNxaV8
yntVqq2Q/P0XSe5W1axY8r4GTWVl6mLJu9oTTy3tjuW4i2WztfvIWw/8L8D+4jGdsm2d8ZUUp5cBhx67pTkDlNT0294IvT3sXecq
3dVqsKVp8E3NYcFuhwW7HRbsdi8Fu+vHpVdN2GoN9ufTUxocFap5yhqd1KqZ4kMUgUKib5ObUhtE3yZfih7XTwe7jYUfgo8R0luX
5pyel5p+2xuht1S7awbDu+GSupcOu1iFM+BMh5mhw8zQYWboMDO2UQ/mLH1yJrjtosiXruC6uX57+Htc/96kRN9y4cXiH4XQF5yI
N+KbS3pEJQUhnKWxlHKUkLG6NGLqaVLkUYt1QZpYWRi5WHIv2otiwR2pFItlc2jNKEkO0qUxr7il4bEnpp62FccZ5aAH41OVR2ot
Ve2Wlt/0RerrYQ070yxWq8EWjZ/TXZInG1JQ54bHnph6+ozLk7A6k4fxxjtwFiXjU3biDFa3siLXT8cRh060Cjl0lubKobM0/bY3
Qm+pdtecjQrqffvCCozFGhwUPm+gzC5W4QxW067FnNKsqIhiPfry7RSL56l3jKGHZWP4sjDGUAijMU1KPA+r9S2EPOWTvy8VfJ+X
9pILEkC1YEvAAJwuB9MXZJA32okLXsi733/OeSIr2euKBXOI+kY0f8keSbzYVOKFGLGhUmPbOuPpRtwIzhO9DTTniU9Nv+2N0NvD
/HHu6QZ6v2o1uOnRwtIMLY15kpaGx56YevqMy9uDj5T6y93TWYSTT50T5yznVOtw+V48fpic5gzkKZs+08HehOVgXFKIWp834dTy
m75IfaWKXTNS3v22lEFOeff5kctT2cIAX6zH88eouaZIp53iYad42CkedoofGhckA6EhUlWxwWqb6LC2MdBPhpVc0Fs+J3kPoSFe
MTUWszk3PPbE1NOkyN5Y3fgscMWE2UBzVCz+Sff0JT8mkJYx+cpYQbRNNovWvmTZsUZ5HFqdbPLa+Fx/a8NjT0w9Pawhh6hv9iRk
E2oaEeEJYW4sTwhzw2NPTD19xuXtZzEuj8Jq88Xu6QIGzmDI7xkMVEI05PkMhlyfoc9uOYj6mCowkkbUXEkwl6bf9kboLdQuj5cz
EKAJBGgCAZpAgCaUA5pMds5AKzbQkg20ZgMt2qDbtTgB2dLBhcPBjQOyGtfvuP69VR+G5Y1OEk8niZeYzJKMn2zXgNlC3tMWWr8X
zlBKPPC0hT6/W/VgmDzJc2/Ic2/Ic2/Ic29E+0gwpJ40F5rmQtNcaJoL3WEuHrcrWVSf0poIjPMmAue8iVDR3ETgnTexVRcm6Pre
J60MAtCn1+SeAPDeu3gWLejNkWpZDKFWQkWfpbFEJswNjz0x9TQvSt7I20EbqqlAzdnIp6bf9kbo7WRfGcwOdg3MGubqk1HrYdNO
YXtUGmrfKLpaLQ2PPTH1dBudPYPyvWlbJcSi9zo88zlGHz5iTsC8D4NcY03W708ksnx7+Htc/95tho7wvDIuAJ5fmiueX5p+2xuh
FxX8L5+EqL95//9GvX58/0th1iM8Y8A7hsKcdWAfEzWCD2K2G+JbciWfxZSQBTJkgYB3yJAFMrFOfN+wvFypfARJ9X7MFcs4vBtc
j7lSu0VgZ89pr9zFfIHfPytwb1Yqvbi5AhmntWDz+sUB80+V1F4PurnyOuYk5opkUa8CPp+lsZYO1MTwuDRi6qlaSXy0pRrUXKha
2U3Tik3Tb3sj9DZYIc4j3XAnLJXeNzu0+Lf3iocuFdwt+zl/t3WN2ij9vbwNTkHP6/dihVPQ8/od1783jTkfauUHDSmE1JxDrVLT
b3sj9GYqdZ05qMhWq1DIiJkr+AB83oPD1NeCz6v0wIbbYa5U/jnFwXOKg+cUB88pDp5TtrEXubL3CLQ/cLlI/7sJuKhbAizyBe4R
KOX0UEIPZfNQKk/VRO6tRf9j+yKhrzJ4IFfg3mPS3Rzelb6XK5JFOlhjwWONBY81FjzWWPAvNRayFxFr/lwY1MwRNx9HqTmfSKnp
t70Rehus0HPgMzs/z1mnKPLbOqprvzQ89sTU0zQKvbIsSgU/frAcwdCb994ZDPWQTOQhmchDMpGHZCL/kkxUPOb8JXA0DRJS7lJz
rQi5NP22N0JvplL3ZtrlCj6AocDFTh4ARS4ART4AZeokn8BQS9vO0r6ztPEs7Twr68T3Le+WK5Vf+HT/8nD/8nD/8nD/8vn3r4vk
uf4uSX1LrG/2nD7gRtDXNSeMhpoTRkPNCaMPSBXl9pU0V/4TduMiKe7mNdQrDi9XHgdIb161HPoJ1lMMwNKYF9HS8NgTU0/VImIN
4URaHFIgKjTnu3lq+m1vhN4Ge8QB0ntebXQ+IEXCeQsO6qXhsSemnqZReHv4yH7SL6q/xi+q8wGpBbtswS5bsMsW7LJVTWPO7kPt
hkhl/qE5r/7U9NveCL2ZSmXkrd1zJzT3ViDOlXyWpUZ+UU1+UU1+UU1+UR3rxD/mFzXXZYbbsjdzZate5RRyBT4BSM3TLlHztEvU
3FLsLFdgtxpnuQJ75Y7lyutYfyFXJJsphi5RjS5RjS5RjS5Rne8SzUgQ65e1XWh7OPAJpK+eJX31yPjTYvn4amUzsecSHYEsn0vD
Y09MPU2j0KuUY6ngJ2615mu8oSa/gkKqWLp+LzY4VSxdv+P696Yx58tzu0FHigil5pyqkZp+2xuhN0+pvMSuSBMSaUYiTUmkOYnZ
k5KZzxXpEI50CgNtXqRzONpq4WdpXAHDJTFeEgMmMWIyVKvRN9a+QDAXKi1TEaDlc8kbc5Q35ihvzFVK3sNRDayf9JM1/WRNP1nX
/2QmO4ue6zU91yODGT3Xa1E/yW/PXDmycrGUFuRfWRqL31SLAETWgvwrU6NSBcbD0VAHo0DwHsBqsquaDKsmy6rJtGpTLXj/tv/M
D2bzrG4J/M5Kr3LS0o1oaSx+vLnhsSemntp1xj8zxin9mYimUnOegdT0294IvW3GjUumUsg1lw54Wg6OloNTrdb1gOtCauC6AHq5
peGxJ6ae1rF4eyb6oSBjqrufID9H6pZQsaLUqBG6pSTt5fvT4i/fHv4e17+3zgDvfB1/n7OJO4+aKyHc0vTb3gi9qNffjNr8eUrZ
+vj1xw8ff/0cD1yu4P4ZUbmSz4JSyVYY4B4mW2HIVphQJ76v8zVXKh+NTRe/ABe/ABe/ABe/8HLxy5Xdjd8tV2C34rnZc9rLf5Ev
kEG16EZT6EZT6EZT6EZTL260XPnd/Ca5AruFp+YK7BUNkCuvY25UrkgWFmHVLodVuxxW7XJYtcu9VO3KXkS8I0gO0RAxKjVnR1Bq
+m1vhN4Ge9S5/m2pdHZKooFbRDRwi4gGbhHRwC0ibiMWi0ehV3hTqeBujzz5++4Am9YFIZT+3jui9IrH/ImonBOlMrKkGt7kcwX3
jwbIlXwGSAkpGYJKhrCSIbBkXJ34vk9QuVJ5QArZgZAcCLmBkBqoasR2S5DKFai7FsHNntO3h38mC0gjVBdaGgsgjVBdaGnE1FMj
/wlAqp4GpLfkS+XK4wBpZUhgrkjeh+qBlkF6YGaQHsgZpAd+Bpl9Uc3Il1Ju0F6nmzI1522bmn7bG6G3wR5968txWSr9LuK54lHo
FZ5aKribvzR/33WNECj9vXyEAIWnRghPjRCeGiE8Nb6EpxaPOX8x9ENQPqXtU3O+GKam3/ZG6M1USt9aMypXcH82hVzJZ4C0oXRj
rvi++VK5UrmFbyA81UB4qoHwVAPhqeYlPDVXtuqaRZ8rdQ9N+zuc9C0xqvkCmRhVFRxYr+DAegUH1is4sF7B1cjvVkcqV2DfOlK5
Uh+wVXdlTuWKZKGph5W0NBZo6mElLY2YeqpWEu+jGS9MnpxF1Jx9NKnpt70Rehss07evOahPg1e1pTeQpbEEr2pLbyBLI6aeplF4
wFeqbw1ezd93T0FTnc9TQMvfga/Uga/Uga/UDW1L/gCabpwd0Jyh6cbhAc0IvZlKmVsrSuUKPoCm97xjmwcocnPFP3YZv8qcUprY
BdbvxTOd2AXW77j+vUa2esp/eZE+1d+7Zu4jDciX+j3DFh8isMWHCGzxIQJbfIjAFh9ijfwnLMgVOdcDq6lXqm+uvI6lpnJFsmCI
lhGtIVpAtHqqls6Rj8aY6MFHszRXH83S9NveCL0N9ujbY09Z2dlU2irazUtjyW2bGx57YuppGoUHopHMVwHSOxm0Sn8vZ5cFBdWt
358/WlBQ3fod1783jTm3D70alE21Yaj1aTRTy2/6IvXl6ZOZSEVzEWkyIs1GpOmI2fPxADFWgfCzRCqPBxCeQHgE4Rnkq9Xom0hV
IJgLrLZUY239/lx9lmqsrd9x/XuleCabqoHqt0Awk01FB56mE0/TkafpzNOqfqb3BCzk/1LkAFPkAVPkAlOuQXCnpP8CmUz61C3Q
PCd9ikZZ0yhrGmVNo6zrR7lrokSBXA633pMLkJc+Rc5V8qySW5V8qrULi3/uN4MTivKlUnO+l6Sm3/ZG6G2zYVy+FPl0HPl0HPl0
HPl0XGg1ot8euEyUJEjd8uJZQinVUA6saPcdRKI+c9qdZ0mlesHr95IllSoFr99x/XvrNPD3yxGgBm3TcKTmPCSp6be9EXpRr9+O
ev1p/OfH999//PDd+7/6+NX7T6OS4+fvxr/90/uPH7/5xfu/y6aU0nALAPprWLV4bmqsa6o6q7XBwxrrLVGWiLRpLyvB1lt1W6Pd
Qa89VNZMaMGIyVbNDI2YIZhsX079DpqlU1fj4wLMoj04eyydPaGzTrQXUantFHJKbVzOvQcqwW3NwJN4cEDRLuw+SHqnz2beXIm7
uMc6entVp5Ksqosu3/9sdEl2lU7Jzfais4j1ivZexQnca8oR3Kwau0sTRE+07r6MV7sPwwMbXbFOY5uGp7c2cBXg54v359nkz+ut
UMKksJSdifBUoj0kq1kfkUuBbqFBy95LCaEKnYnKD957org0g6ScTKnwVVQc06H3gwLfSgycQ/84GDhzk27c7FpnFLDcQpixUxYK
NHgoghv1XRruj3UJR6ll4u/TsHXfC7urCujkMAt4r5aCM8N0PzM2F5ifx7mxv9ZolgfCDhF4IEIgKCWh8llUt80le4iMdsZTPb4Y
B0XVoHcVeSf1ZHpOG4Gz25ZuqdNWXd92NscIsjjEgypSrrNa/G1HbXYDeq+y0vc66HV224FSExAKSpvCkmZWdx8xxhj4jTHwhZ6W
DjoxW/ST7BKnMJUEHHuko2B7D4EuIoTOmrFXHryGwVGb5ZfpoBN37QGVCMDedu1R59eezXoyTLRzJBOpejhA1NXVx+gIdd+IcXWK
hPVA/OoBL3pre+897iYkTq5Bnc2jOr8GBZy2UFItpYM6O+ygcU1r9jHfpsf83oPDXYI2q9qXZJ51UIi9BElFFTustI6lz/WAkv1L
cYgeS5pHL2EUTBBLjFvfpEjF6bFEW6rpqu2gImGtEb4Eq+w9QIG/C+GB7p43f6d3IaEclMiGu9DSWE8uDSeXEOGu4dsd9oF3uUXW
GPY/7Y9vRGZ7gFgmNcSm1JD+Bwh/I7Jb0/isUic3InzsFZToOf45RdNPGMzQe2/Qt00le6SI0SD4lAQa/aBEKtonxXhxA+rmaZ1Z
Ss6To4EIqsONSJe//1j25mEhOLuzWgc3oq3LNZJemwp7lirs9dbr7Ea0cXhqqL8nsdyJJVLo3soxVmGzU/H117Iesi6ISF89Afkh
erypBcXaKmmgWIoynRXb34kCE6KfF6PSQZ39dchfvQLR010ceo/O3kI6JgDcMaB6HJvu65q7C5lAIR3KK8iUMxbcq3PPgijMtqZT
B9U4F+bmVaggariDOvvrkN9eqveraDNzqrM+O+SwGR3uCQ/iC7trw92HKhOaO2jD4mRvaXCMCPQiZLAGpguewu2DEq73qj5y5iqq
ZiOFHG9iRLyt7KcmUOI3ap8My1SMXXh/D0r4xho+LpHXHlSldjcpxs8xvAtZiZW7vYOTCwlqXLT6rrE7OekJ2lvGCRe6b9Dja9DL
Y7LNz0Dqcm7w1yC9fRjaK4WZGnctMQ5gjbdZQb6AEUg4ehqKQIqT+CEmUhx722zy78tikIHiVIUdpHDJCz3CZyd1ugoZMxhPZUrG
XxHblTXX9yDFhW7iZYMwmOoCwkzGJQhC5clmwYO8cngD6q1U5psQHB1gTsGedjGo5vJNSFx5xPqGv5mrByEi6FGCKvKMC81B2nPa
l+O3Nr1n8CIAzrJXH9sV3purq8++GEpg7j39p4yzhiyAduyrlOuuzv7mo6MjjKg9vAJtCwBpR1HtxkjdWbVLG12QuNhBnf3NR26c
WPHghbNvYKc5vfyYq4BBB9iq967nLj+bwAfL+HHvXNssNlaebvXGKSzfDsUDnAlQK9wa2Xtt877bcdVaqSgh01mZglkmt6OVwC69
OHbXxyAd7wEH386tHk0qxsQFjK9WNynGPvYJAxGPRtBjH/psjPRw83H+rpE7f/K3TARCb2eJqbr93B5Ka/LC4sSJl6tz7LPJvfqM
57yji4+BWp+JGU9NtLdU3DKEcNdEssfIeDZ4E6gerx2CCKnYkhl7rUg5QHac+lQF1Lopfq59O5xmvpM7nCoDeXoBimA9BcJC12E/
nKfFo2awHwLdvCMgewEPtqJH4ElG2jwcwWH7HA+Jb4EelZUgAKJEj8eX46R6MvpQahDebuHxlh6EXA8ccpJvTydvWD2cn+vN0eNt
kLRzx2+/OsXGaZfiBuX2tyO9mUvtydETD8rGuv5q7S9JamOtIGs+j/y00/La2VAlWM8rLHrb12dynOSPlyYRIHRORAidC5BqoSM4
1o3rANqOywEcxZ7Y3YUXXx7lHcfs/vKk7WbNU3wo1oKIWI/uBrX2r0eKrSOgsTynpQp9N6wtNq9IYgk9y6TKKgyVuEEr/pnBBUg/
8QGuUgJol4KiSgRe3qDc4WNSAB+10EPUQgIrgbFEzzOdXSKm2LpxDUodw21wg7lTYSUazVai8VjB4D4oxD8pWXgMN94puFYJeDbU
jk43F4K7cQT3dmFzicGN4tjYnVuQx0nm0SbZIrBeUINPhrccLvxNCws4QKS+YR8z/Z2Lj0NuDh+bvBlUqpXkZ7T26UiCx6bxO8Qb
195hMK/VFMs7tqyiVKTxNJLag/tGWnq7nlw78ZU9/f0vk1ofP3z+99+Oyv/08evnaGpLFThgBxMYLymYYgqEPJRo0+Cs8i0NhJFs
2BwNhGkciL4lFkulcyWg9ZCg6fq9oBrK4lu+19TtprXYl6yhVHq3yrjFE7+vOVZX3LFc8L7mmERWKImsUBJZoSSyQskXVqhSPbqR
OZQK7kvqUCq9F+9YqVy2GNkzK+4uHrLiFceTPoxX2JjuJtCcI7RS0297I/R2sITf+vKY12rBThWmwDskwnabIEAkwnYvRNjVo7I7
JSsrrtQq0K0Gb/l+5cuXPfz7eYRQXb+tehr4rTtV4XfE15Ka80ZJTb/tjdBbqNw1t25Leb1SBQ5QdAOZWqkGN1GblarxOIq+5N71
QHXmgerMA9WZB6oz/0J1VqrDHkVHpohYRtZxqeA9gK5kkiye817sZ+WCv7+lLn2pHnvTVFmnvVTwHkA/NOW9iCdK5XLY+b7LyiWH
75csNtYAi3FgjaCiXKlpxabpt70RejvYv29fe/yflvmlakOpsXjC54bHnph6uozKgx4G9dXYOYvz96E9e85nkbxr6/fi40zetfU7
rn/vMhf8/rWDdlhQQhNSSS2/6YvUV6iX/moPdA4d8L3nRi4tsCaDConbmkyqJpuqQ5sazOHp2DQFd5Cs1iSdJyEEDkKgIAQGQiAg
VC3iH3c+XzAG3wdhL5iD7zsaLxmEtVT0Wrw0FjgjoTLl0oipp0WPvWmqJBQrFbzHzvfd0/QX+Z0vyYXvu6dckgxbLQGQzY2V0VYC
IJsbMfU0LTaeGCMOTkWKu03NOeIkNf22N0JvB/vHYecGgoRaLbip8spTYtHSWCpxzQ2PPTH1dBmVPW3GM9vlhDbjqf3KY+f7oHs2
ObGC+rKKOLDWbw9/j+vfu8wFu3+1G5y1MXnCU3MtJrs0/bY3Qm+hctdkxTdv3RzS4nuxYy55sSIcrzybKU04Xvk2NR73Pphr5zOF
cHgI4cBCzB5COHz5JfOC1Pg+h+AFr/F9aMp8VfSG+ZlEb5ivit64ojy+9538nPr4oV/NoeiHljoHzZQzFBKwNJb0hbnhsSemnqYV
x1lhJwerZEpWSK3Pe3Nq+U1fpL4O9o8B0IrukkoXVkmv1YKbJWkV3XWWxnIuzA2PPTH1dBmVty86mY8B9EM26ZB37jb8nkumPBEu
OAp4p8tm3FThiUjQ4LrMBbd1w7jYnEiVsVPrc1pSy2/6IvWV6ZVHruxpjjxNkqdZ8jRNvnieMkmWPZkUTzbFQ8UrsireNitxRrZs
AMwbQPMG4LwBPG9cszoMWSdUXyEw7Tyb6uKbFWD2jbOpSM7yuSTbOKJ9dkT77Bo1YPiXaQg0DYGmIdA0BLp9CBgeZuASFIWJnTWL
4O3ZWMgsXmaldAS2A02v7kvDY09MPY2qMI6gW2+XGbTNhu46Burq0G3H0HXHqGYFdpb0NiSRx+B8a5JKFpOzMQLKchoBZTmNIM/o
0oipp3Up8m/IYVBaeshkWprzu1lq+m1vhN4+dpQjeob3NHhQgxc1eFJzvQw6C8aNB2/20ljmzIM3e2nE1NNrbPZHKiQv0y6WiuVn
VN3Gpi8fdNWe5vH5bdeDEmrocRGqQCs2ZQSu3x7+Hte/95oZniJaj3fCVOENmrNnOTX9tjdCL+r3f73/edTsn97/efz3j1VZig2e
3RLhB7HV9HSJxGyGpXs29dJPULmlS4qlS4qlS4qlS4qV9Sowfo0Gy1cimXdpe6ij5KGAbPJwrt9x/XutfCYopOGsLJHcLZi6aKJ3
pqHyfb5MKIOwCV4TtiZgTai6Vuje9lTG25QI7ZtuWCK51xtuiUwOMVd6vkrEsnlrlLRGGWuUrka5atXLiU9KGkdPEaH82prvq2vL
b/oi9TWaKy6cw7AFkAxbR8I0G0wW/AYlTZqNpWEFNDz2xNTTPBq9wjhqhHcLfy7bfzy6rQxvq/ndnPm2G2pgCt2wELphIXTDvoRu
VI0/vz2n/9eClEnNWZ/U9NveCL0Fil3nC2pakZqWpKY1qWlR6rJVmZEreKMpUg9U2yhRgTkZG1I1SyTzeLbah1ci+lEoq26JbS6a
417p4mVCv39eaLcA5hKhDJRtCDgrkfwQlFU3RS6XiGXx08xKuuAn5CsNyFe6NGLqqV5drOkMClMPoTkfmwrTD6EZobfRen3rm6xS
owEbDhMgqHxprNU8Iah8acTU0zwavaLQaoR/yQnUObOv5ndz1jxAwGWAgMsAAZcBAi7DS8Bl1fhze3S8REbn1rOCWp/jkVp+0xep
r0AnfWsljBLhB55akq6AKykgV5Kg4o7V0s+QLTCX0dEEwRMQO2FcvQrMoXgfyNPXyNZA5QsDlS8MVL4wUPmi6Kp3kb53z0GgbykZ
VzTHD3kyLrP1REC+zLmx4Nu54bEnpp5aHZ4yM/dk6ZUI7fUiWSKzY2WLErFsutdML7qkeyHX6NLw2BNTT/XC4qMZJlJRCWWgVQLb
1PKbvkh9jSaMg7cNj+M1GrC3D6qxJbDClsD6WgKra4ldba2q0Xj7goP2aXirb4wdrvndnEk3A9V4Wb7XN4QAjwgBXhFC8/iz2zOO
yNlpCmJOzTmOOTX9tjdCb4Fi1wl3N8YDZCTb3eg7zk20s2AtwFxkcWyWqPBoNMJVgt3E0mSoxF4A2sxgqMheIoX7/K6Vz7hwG4Jh
SyTrh9D1RWbdjT7Gy9Q6S5UprOXI4SzVpLC2VvJTcQnmy+ISzpPo7vGqXibQPTDIbCQolDsxUO3EQLETA7VOjKpfWEdVn4QnhxU1
17pPS9NveyP0NtoxLmmuoWppjQZscZAogA5ybiw7fm547Impp3k03n75/NI8BrqPbMcnwxRyE+SUpqqt6/eSuol02KkO2ud38/jz
1SXkILVOYUTUnK9/qem3vRF68xXLy46b+BXJPqCBEBsORjIRRTYiMz0u0hKJtEYiLZJIqyTqJgXOUuNADQl6SFBEgiayUZX9eelp
MjzNhaep8DQTvm0imJ0yJR4ayFI0kKZoIE/RQKKiaVBhD4INnFJwTME5BQeVafr9exys6HwGokNkjqIjWrm2mX977sUzMw8uOMiD
Cw7y4IDvfWnE1NOgBpMzctvbUk7+W0Nl80Lhe0/Qc5PPhvI+J54N6dUWqCjmxkoiaJFYzgIVxdhoWXu8edaDmungZ/Ocmit919L0
294Ive2WkUt4A15Ay/IC0iXd2R7mmacMkdYAu4s1wO5igexxbsTU02NMvvK8PElyu+1FNjvB7UEVTjC2V4AcFCAHBchBAXJQbcjh
HGarOIhoIhRhXJprEcal6be9EXpRt/82avTP7z+9/3HU8V9/NzZ/ev/zRB2YleKmA9yFlRSWXtGTlwDvwwAx4jYnpliNLa95gHrM
2iVPp4ygx0wUvni4vaWEO92kCMNhnrQxerA+RCqEkLQxBLpNxOS3NmWIdtXBWeJVqqqD+RQ2DYiMh0Wqi1UAgkuHZVy0dQR/fKRC
LlBEPNExjVqY0KQGkZA7cBbKmCI9AusulENyY01MyE0q6L0KEyu7DztXhYMHQZIvGuUTg6wjP4n3Sb5iXCX+0F1aIf77V/Fi0CqS
mZVCEaeyshLKAJhAL+a67cAiploHMYc0DZIJO9RHHqNi6UQM7vaBGYEJzDh5GigWnnhwHXmrpE4HkmIcVuboRaRYOBJ9wwLUIcnn
fHX2yFFYLJ9onh36S4HR2wCjt5TkynZekrfUW9e2/JBLl86iEa4EncIeQxiiIiKDsVNGqsRqByNiMuJGDlLFHsbzG28uoiSSactm
egQ4pWLjKbVn5cbpMhoo2HWU8KClKY7DGb2hYNd9Rmd/fmocHMUu3zQyoXH97pm24fwyMEcX55dt3MVbRu2vOEkY7myEFtoRvAlD
quz42UPYIqHgdmzB0mQ7LC8TdVIpukEZmSjHJ5Z0mZ6qJmBkUkjztNO932bU5ih3mLI3OdhT6IEJzu4TuDD6AAI/7JadvViLLTz3
8FwhQ1QEz7lSkwELc7WpwYBzD+Dc+eScAdIS9MgCNteNutBha+GwDVcnLfnjy3fRQU4fKBAHb7wBv2QAYG6Iw8Yoesh8KeNfrAhB
cwvQXI/H9i5eFaG5wqjBNg30XgM37o+0GCyLzC0fMFu+EN520s3gSPo5Lm+X/v2r9BF8R4v1mAPAcgV1V42hZC3rnG5ShOyaJbsW
aBjm+/qxTWlbAoTLLeByku6Z+GWE5q1LMBlUSwbV+1RzQPWGxkd5giTfjhOftqBhgqXcUbBqsXzCerAGTRS0BrUDcL7cmZdQBg9h
+l5uI5jLFyGacLuJN3EpuDvYwWqtIKhaG5d8J2pysKQ0uQmev1JJ11rPb7zJsDRNaDUcZzWUaFyne3hu8Tbl6DY/4gvJVS3Rjg4N
70MfdZgzdFy/3P7B9WsbfBvnGYV4hllDGacsNrf1uPgstRCPEmO9PMHm/Y4SBptv0IUAcEEX3LjlDlQEzYPts0AONvaEg6VI2NyO
8NsnyKOnYuIEzqeLNnkCpB3PRFV8mdIZ2FwN2lN0rmBBsT0sEVesxRE2V+ONzkoGmytWjTCEJjVOsfl4bZXaMkU2HIvNK+4J+hqb
C4obQL+5Y0texMZZOdhAUlAlQz9A4gDEd4khuT8mqO5VkyIMPI+p0NmBx9I2wCJ9icwDzYNlkmgC+qzbxPPQnDtLDWNWQrN0Bppr
AT5LqSJk6kdDLnMdLbnMnbK+SRPGrskhknk99zm1LgIGm4dBXWBz0eQ211fYXGR5ukyj5ANgrmgFcsDctwBzfQnMbYgEzCn42wkv
wWXuKPzbuxjblh9vvMcDMFrigxhxuE/bQYIjS2o1xKBTvvZoTYJRvofRPIDkgiYILcWRI8d1UYWfLMj/MMYY/hLlAk6XVrrP2HCg
3HCODcu4VZoP71NQbu5C5DoLkY8Lkm5t3b3lOhuRh3FjEMYb9xNAcsxCA0RuZKdpOdjU4+8VdnWDxk+aUAkPYUJSseZxX+tA6ZHG
D+PeLkY7JstdHrQPjLtcHmBh0aTFESQfR19Ycw7JTRMkN9mQXA3K6BUHaqgOAovYYyiLadKFP2jBS4zJ5Rvfh20IZTFXkFx9hi1E
2EHgMU+FnUfRiZNonKAX5rpiRVhIbgmOuYNYlhZYbi5huSc/Bw/LW9wt5gKWu3Hrq3gSSROb8KC5xOVGCHqaXmLfFg5wqik2vWBT
hQFnXWxShLdsgD0Y49YETs2ly3wDyy2bgGlpHtpWIYfLwdmm2NSvBtNuMqC5PveZ+6OSJ8XyebQnFMRHCAfpf0qDzxxL23ghbdsq
PPCZjyhbxBV9xk+QqSBFc9wjqZrS5Mqg2htSj4ZLeN/DhLIA3WMKCEZUA0CPTT5zkw3QTXQepkwBQoc3DmChdKbx1DiF55v9Y3fr
N8C7Y59BYc6we93VJtNlHu56ejO56HwC3jLxJE1RVoaiNBKZyogtJJRp8FL1mZeDXe3GC4BOcWp+CN7SMR+HEEWqLzQiHRWEhayR
ETCWrprTzE4KrQ7jlKjkGvVhjuj+9AkBBoBA7ym3U7XqchB1PoqMPqW+BkVx3oFOmOhQmfJ1lJHsSRpJD6kTzpJl9lBFf1SONIqx
fXjo7PWY3pO4BZxkE3zo6PWu+NZwkvtJWvgRp+tkmzxZgHHhrJg9SCp/EiafW/PKJdieFJksHV0sNUEGTfdvQ6eN8e3LRO+1UA7D
sLB4Ac2KolkZlW5fGm87LT6fBxUTanpwye4wFoTk/SYmHZG8pDJuWgYIfhGRPOzGFj8FHueIkjLjHpSpctmVm9u0jwdBelqh44iH
VE9Ogzmmk0OHw0TlGi2SNfZQQ87YtF+xho5hc+BC+7GB+N4zFZsgaRdjJhXGTDbvEkKMHl26hkpp2znwd0GMxkkAjUHCC1BxdNZp
HilMzXSNc1SKZfLQpaxWOa5PZXSKjRkHTVuyMSoMshjmX+eSgrUZDzYR1tFymiNQREZk275u9mAfp45C/KfLmof7mVZ4PxNA7iOj
7jZE7KmLrk0IIIos23voN0TMgSc/EyzCaU6WxtJgoX2fb28AeObAwYdnjmLLlLl+A3MAWSTV0B9Xanr7H3v8apamIvXr0glmXPz9
Jutg90/Pe/QOPG7/SHfYiT5xnM5UznAKFlbBQvqp03ZzSE4JsaNu73/4+OHjV+N/fv3d+398//Hjrz6ZIEt5FhsqG1bqcVDIu4Hw
sVKRm9gXK7XpWxS2Ugm+6j1kiiiqwaEgs4kwRY+VqnqV3q6U343UpnYpvHWqyFgtf18PRgcDkcZzYwH8c8NjT0w9HdTZl1moZN6p
lL8vDfPscuhVPq1SfMfi4ZUasHXENTwlLY2ljrgWhBKXRkw9PdbjUSm34BKMhuZaym1p+m1vhN5+xpWrvEgmTUnWbSAPCxo0KsPX
jolQOiZCUnGEwjEREX7P4enFs9CoR7eqjNV7u2sl8sbR4NGHgWh44hkJwDMSgGckvPCMtE4Qu9fNeI9RUCw9NWd3S2r6bW+E3jod
r4kmFR3NyrI+SoveF9FBjwNA/7wiZ4CeIIMlzGAJNFiBkfE9tGEAfQMdZKUS/JaygOgtQHoLmN4CqK+9f16QVFayrVTK78a6Xrse
3r5a/h7VGw/5g0tjKVM9Nzz2xNTTQZ1uzBuV8ruR/1TKf97m3kV7WakBW5rbhgCl+CGneml47Impp8d65FG9GSIUX6DmjOpT0297
I/T2s7AccVCEAsWcWxpidGTsae750upWQdrD3FjnTzmYP+Vg/pTrOUbPO08yiTOf1eMI2ldyhzaOxpc5FlVBMfYpVzCtX2jOez01
/bY3Qm+djtckmy20CZV6HKD6hpLYlYqcUW8SgjJQi4YwlIFH99BFm69C9ZeEnHZIiQbr91KxMhX9Wr/j+vcOqjC8RQ38MpVK6Ocf
DC5oO2/HspcMntI7ympbGrMpXhoee2Lq6aDOF5i+C17P229a+g4a+0rxHLR/dj2yfEiPXjUzOECnWmZCUwpqas5OvNT0294Ivf3M
bGfK+0Zl2AeXmdZ+eXCZG8uDy9zw2BNTT88xet5rn0ka+vQG56H97fd3/fPz2hewihoxxEhRXNScN3xq+m1vhN46HTMIRh+B1Rlc
o89cNMzPCt9fMJA+42K5JCOVc6bs8n7rHT3gYgath6I23nVQ5QticS7YSVsILmoXxc7w3H7UXnKWAnXESy7t0vDYE1NPB3W+4NX6
itP0kVC5c3rT2+84l0ynt99yLklPnfUQ1mEhdntpeOyJqafHomRt/nggm2BTeiA159Db1PTb3gi9/WwtB/IbSJ4blWH990zNIkix
JZ99z0HpSgXdqMwXQHvzs/Da55KnrjWuFzgUDFP7ev2O6997ThC/zcUQApV0oOaaYLM0/bY3Qm+VjnmUqpEgUyTMFCHNFUpV18Km
TGLVZnrXeoVOiVbBiQ5jJCNWG8DKkd2GiWHXg/Bex2VDOYjwdb304EjUKMF++fzE2jrRPy2fcflrH0X2iF+D9xDch+A/BAei7jUg
+meix94+QYYjVp4Vm2pgVBijmx5MVH6qfLB+LzH5QlNIvqA6JuN3H2X6Uh/W67G/CmjavVBwS9Pu1bR7dbfdu4/meQLDZJG5fs1q
ZdFlDAAw58aCMeeGx56Yejot2aOwfaOpXDA117D9pem3vRF6u5pr5o7gaSl5Wkpestm5sjN8OAjeV0DQOzfW8H0FBL1zI6aeziO1
vzg08EW368OciU+4JbKZYe+ObyzihzVDiqpbvz9X9fLt4e9x/Xvn6TqifDbEcATNlfJ5afptb4ReVPM/vf/xF++/HXXNI4dVyYrB
oR2J1IUObU3X8fEzFMrc3A0U3O0s8I1autpZMhWJq238Lv2p+xtAkm1ItgEGWkOyx89UFWH8joWyGd9Lw2mRIZBz4gtw4gtw4gtw
4gtw4osXJ36GWNXVYZ4hsFu2bM4k9nLEZ8linO7ECpwaDOfT0oipp1A0Y0nqHrozZDHO9AbomiGwl28rQ1THSJgMaWzUBC4YjwvG
44LxuGB8xYLhH7zjoBVRSFNzfvBOTb/tjdBbZ1c6O7oLBHNzoLWGW/LcWK7Jc8NjT0w9tb/97bnz8AR63nZgdI4+KfiVvEV1wALk
oPgjcr04A3VKTe0gc5ss+EGJqKnW49JaChkuLb/pi9R3rYq6xIrKJ+ikLMG25c9LXToFaLFUKo8WWxI8M4SeYEWr0g+2gn7w8ufP
11ZDP9ja4h/MZM8Dm4lmyaag/JMu/rHcytZDyjBev1eZAYQGkBoKxTJYsSHoMENgt0DtnEl8e1IW55s16Jw16J016J416J81pWu1
W7B1hqxuOZMZsu6FieqmWIoMaaxDTQRwqM2NlTEogENtbsTUU7pWjnIctSTGDGquOY5L0297I/TWGZRvfZNpCgTzTk3vwKnpHTg1
PXBzSo91B3Pu55l5ivcdT92jG7I21QMwMTfXcApVDRDXGiCwNUBka4DQ1lA7yHyEgh6USiW0oLnWAFyaftsbofdaH32NFTVhRQVY
URNWBOg0fvtCqQdY0ZBYQ9695c+LWA9iY6HYM88iBBF7NoiYzJR5qdifIblv3b4MgVxQjie/4vq9nJvp5Fq/4/r3QrGqV/mMDFl7
mNgV0+gHXYr6GiZ6DzDRe4CJ3gNM9B5goi/dl92K6GXI6lYwL0NWr+J4GaJu9ybqJ4re5ayVo/QZ7yymzyzNNX1mafptb4TeOlvy
CEzMzoNzmFfrMK/WYV7thtzTveTVlvz2ffRRT8imvwAm6hsr0BX8Ss6OuiHVn1m/F/9Hqj6zfsf177WDzG+yMNWjTuEo1Jw3WWr6
bW+E3mt9TA5MZHKe0N0F1ZjjYAplXj8/ewKJ+PwsASTqQRaKPQOJ5FI04FI05FI0gItNlkvxIq+spSBnhkA+h+xOiYw3sfujj3nw
5fkiBawrfrpM97pLVrdcrgxZTzw1n+do3ffbOHDY1aFzmXtlhEFeQ0PgcGl47Impp3TBcHbLy8EQWyC1lkDFpeU3fZH66sxH53Jo
BYLvKn1W8tvvdR9mJkd1vf5kJULdJ/F592FBgpMIg4sUlkjNOXcnNf22N0LvpT6nyUwKOUspgQj8WwJopgTY9fHblcvmIWIICaYF
B/SFIcG0CBGKUWbAtIw0JUCoBBNllBgiSUBRRk0qTI2KkU8nWxLtbZLsLYHj5c9+/nMCx+N3LJdKK58GXAyJ6C1MBaLWEV/+vpAA
pnTeicfRV0x3wo6EyOk8M4QdjefctvaFpjJPpt7JVLS8lAM3NS0vLWiQtcq4gRynEynOmcHz7G7iUF2NzO9fZE5hiVg+QEH5AAXV
A7SH6gHGlotONkOBexxI6mgL0ROEtLCBXM0eTkiTlhMlN2uIbTaU32wUXPB0jdTVWCmAt/BT6UYr6afCjXb8rti0ADwVR8FH0ePb
lB8I4w7lUhMCUviAagM8YmPyv7QEQJ3S8IitgqlYVGAjabCnbaroHWSqkyMoCZh611SeEfsSA8d4fjthqu3Ut/2J7egwGT/T5Ds6
TFygyR+/Y710bjqsU4BHkQHRUeDp+HcKPLW+aQT2B5tjH8LcAYF4qJe9P2UAKACzcBp6GPlQt+s28JR+smF/sjnweNQDo2+csZ7W
NBlrvwYbr39fjDWR38npBtIw6uw2FHGECcR16Qdjogf2a0lFtuSIIkSKzZxYMqPZxt3+7uOHj99k09px/kslB8v4Lw0dgC+Eltci
D964m57Wr6WeeS8deS/heX3581IoKy34ES3JMtH7ww0v2mGw7EXbQgmIMnnM2lYLaeNa/wquYNrRFSwacg+rUCh2D0LxKV+xT/kG
uYVdmbw9AHXwzG0Z/2WyFqUjuj+caclAzDAtGE9L1ZeuVA5uah0iPHMHDMIPVBlIuzSDE3+cL/yZezsAFzbYk4TnITC69GfuQSYZ
HQ+XU7qbwtW0VNjO0sAJR/mB4Aihn1U4hhygrAy/uBbGARcTNREAG23Ijxm1REcaUTfb4G3hSuFtl0ai32hG8TYm25U6U5CWjhSk
FYYQRKiyJAx6xORERwds4A7YETvVWTAeN0aLwDFilr4l5Og0MGdbVanA/lyiSDEN1zSK7IfA/tI9dIwWFfParFivoivdTDxOBKxA
9pqQAgTD2cHW/UrOiMohvRFMD3xQRDK9EIzfkdyYcosPC8b4CBsGQ+xTYgQrwiI4XHrnDTYCs2A2G8yFa3SYkTDj+IQZsn4KzJ/K
sX8Z+TKSrIM04Pcg+6DUBh/aMqln+JDOb0Ol3YxFX1pab2pQZYKZwwvefQ377rvJInFl8piFPa0NCh2gHMeQSiVNzg9IAvPBlAlV
/+IYQAHkdtwNU+YAbnWODBFaKA5aoLTClbM/hD2AQ/I1s9Jk6W/j0KFxAXJlDNicpWex+VS2bTI6pXN477mvzrGhJAekFOALJAek
FJtJjGUSdyZGcSaGkK+CUJpSWRxEpO3guO3gwNVVKo5DKT5A+VuvoJbQ0jOvmKDhuTU4HQpXDGvFzHjnDEAUNP6gAPSfqXe+IPrp
sEk+joktXllbZVA4mEjDLsGUSRp4CaZM1poyFigazG3XsGUxs91EQxc6q32o/OVvZ1flfagPPne7OpHnp4VloiBpP4XS/cSjxMpw
lfzfyGNECRnV6She/75gxJQrP2FHYyqHmNteYdxOIq0sF4dxBaWM6tTnl+fnQPhwHAVtrte3zim+Y6n4joLqO8ltIDfPgLZM6GP+
w9wUGUsLzZKby9JSs4TXrM/Ah/ph76G+8h5OziR6EUuPIevf1+DHSIvaSlcmdo8QPZhEv0dRgWYzlE7mHiDiW6Pg3xrB5E+vrYWz
uH/eMfxDruFtT8gw+/oKKHqIvYc3aw8WH16sY+Fv3B/5fb2k+gIiSmROs2w0JODuwm2xf6sC1yitF5o8WC2lU8f7EJk7hQXHDrg8
CsVdohMjpTzAJ1j1JcOO6Qx4GKeXZg8x/E5oC2V3lt75OXCUamVEeChcnT35dnURt2wAtqXnyFAnlw0GcILS2mwQ5MVdepZUJSMl
0EVFXfnLd4eThThBu/NY00uKH0ydSCaw5c79dD9A1JkAcbSfHqoERrLeHgruSLKk0rrKIebDPD5vUMmajwehU4HCPFLvHOYxbXdN
5QpGHOOu4YvJcSFyxW8cmyCTcwUxWUnUUIMGcqi5R22TsZlzk2MshRMZCieyFE5kKZzIilLB528fpYUZr+XxSzqSwFS/afnzsqRT
6Y3pxmMLF5FiHHoUrm2Z7GmChoXjyUBDAVVaLEvICn6n0ul7e1jcHhRaayH0XnlkjbcQeh/A1BfuSA4Z0tFrGf8DzZ8pE8XgQgWZ
9pap3wlGvPisOX2a8tzTFEVJZkBek/+0bBhDTb/Mlq4UFhR6BIV6Q0IKoNBKHcFpFWLhYmHtllZDiC5xlEw+lEh8o9Q7ew3N2Iuo
UAsRqowIhwo9iwo9+7YsS8/XU1SoDVSAtxZCROeeBRVKeFz2TqjKX747mnoGTZhcp2HPABSTgwp7Yl+T6zYcT3IFpY0jeViCAiOK
YceudpDZ/TVFjYdAFXbG00uJhE6pd75I+1EVp4DqS8drdXIzZCDuVUBEv4DIVwEvoMJd77HcBBnKoQ5wzQ6UQx013UPiS3B9luDT
/Bi7yY+BHG4L+THwfKiEqlCBSY9p4EjIEsmH2xoIt7UQbptYboOCcFs1RKmKJe+ho7bIg5POMYsUJ+lQydjkGZkxyKpCJgNZVQjx
6FAhkjmr6clbQzidhmQcLGuaEVGXlRdjx/simSgtIW5+7knhZwHyOGP5vDIOBM+my/uDePlykUxaDEFLTfcQQ9jS0FKa6vIXi9x7
HelxJYLTWGBKjAIvZ/msshkxDbzGWULZSDoq+DY1FKEdS8XepoaDJ1J1HcmYmQ+jp1wuTfkwU0RKSkuG3jUfxkXjMR/Ge1Nrlbh0
GAgQAMvkIETAb9JhbLVwdi6sh6hGg+kwc0/Kh/GQD6Nj/Qi8/fL2ALCifJiqYm2ZG+4gGQaCEyAoBIMTICxEZsSFFGXDaHrIHk2v
WXff+vfP360hddUMMtSPOB/v6Kd3RAXp0soYCihOvX7xZPoAgVXjFdBurz5/9/Gr9x/f/3nKz/n4ISsjprWAXp5EnmpS05rTtOg0
rTpNy06HGrln1PF01TOUiGXosmcousvoGuFM9YkGopU8mVzNqoYaO3lC93V9Kp8d8sTpXoUmM6exV2GfXHHfPymuW9nHPHH7Cj+3
iutVHSRPGlfiRyKbLtz00dlN9/waqVytGSUgXFZhMtXS8NgTU0/F+uGqkYyXTuBUodbCxgqsKtSK1FdtX759xZF7XPOnskRh4U9+
++XjB353Vpnc/cWX/Ln1CDmrBukHKKoMNZWhpLI7eHcpHGhul6lx8qKIMuXrpOYcjpmaftsboTdHpTwO8874MZOt/A4IWcJLfgeK
zGAgv+NUyyEcv+NoyeAX7wwnM5jEOyPKDM7wzodYFjv4vRJvh5YZZN93S7wbYGYReN+BMbP4uu+EmXn03A8gzQIi7keO5cfwZgmv
9hN24THUmU2Yfffp8jD2LGPC7g0//9v7T+9///7HUafffPzVx2++e/+3o4L/MLF4jx1/ef+xzJ3ZUui1WhMeo7ZwEVSrcoJZ8amW
gg/wqZZCD7TrpA9zWBo2B9wclM7po8Yd3DjVyjD10BssebUauhferl8cb1+vAcPlHSxlKCyNBd0ESxkKSyOmni4KdTNy1RrswXTl
pahag17guloBDmxX+jmqdWBLiWtYmUtjeavWsDKXRkw9fVYmX396ynFXENQ2t9awthgUBLbNrUh9Pa0vh8sbSGub1WEn0EFxXYuJ
ikvDY09MPX3HqRdzULMmX3/YdSajbB4RDqdE4B6KwD0UgXsoAvdQfOEeap8mlt/cTtHwkmJKl9ancUgtv+mL1FeroLq8AWA2FE3e
NhtKUCyN6KIJfwNQdBlRwJZElxEVkS2pjyonNwBLhtSSJbVkSi3ZUqs76fN1x/IVk/qUmehoZ9ENIMINIMINIDbcAC741StJOqo1
2IP/Ssbu+nXxFeD/kpwdi99D6XuofA91720XNbqx6lVrwPAvPXQhVXewMlUr0JGrqVoHllpS8dUtEDMiZHS9ViYL+T/TXFO9DWiu
edZL0297I/T2tLmdWUCb1WFxfwgQZBzSFSk1PPbE1NN3nN5++XPY5V8K/tWN7v3mEeEIpsQANbcM5c4ZSp0zB9V+22eIZyKdir+6
lBZPzbXI3dL0294IvbVa6p8N+NcZ7v+HUK7ODVkRWOtVEM9Nsqe0so3opM/XgX99Df6dgcpfhsC/MwT+iYRi+u6iDOP+1yw5CLAJ
0L1M6k5jsr8BPOBm1ec3gEqa1gYN9jcADaXEtCewpaGWGLJf6Xo4o8+vAU/Px9ddA/RXXwP01TXgaR14979UmIEKKHJueOyJqafP
8uTvAqNBj5FSCVNzvgukpt/2RujtaYK5u8BXnK2ndwEd4A1AQ4HipeGxJ5UasX3H6SveAHTeNeAB+K1zrgEP3Px1/jUgvQGs32sR
DShkn94APr/7ThPL7uw+efrWUUqtz5FKLb/pi9RXq6C5jgJ6KMbEZFwDHgpIMrlRQBprRAiKmRQUMCkoWrKPPl8XBWSurwECXtcE
vK4JeF0T8LomQhdlVK/g82oN9jeAB/C3Ob8BPHDomqsbAJhnss1kmMkqd9HgS9z/5hz3P/AWZL4a8pufgeffXEF+rVKpgtRYrqMq
FStIjZh6+qxM1ubr6ca7Rt5T63N0Ustv+iL19bS0HNoHHzI4kTeVsgXVkOlq+Hm0/+hZcuLqb0hRaFbnS1z95mcS52Py43wUIBEF
SEQBElGARFToO03cfo9hsCYVLafWp0lILb/pi9RXqWBe7mmkIzrSGR3pkI50Ssf6YzozJzVSiGikGNFIQaKRokSj7afNCdiXEbZ9
hH0fYeNH2PlR9dNrfzo7QriOIK4jjOsI5LrQTxNmt4VAN+r1+xNvB7pRr99x/XsvfZhMV6h2S2sHifho7ZiOa2d/BUCGbLofIkM2
3Q+V7rha9h4p2t2KdrfyB6XUO2rCPApET9SzS2PBYXPDY09MPb102lvQ+y+LGWm2D70pXqffPhTxkZWWqyxLAUB7WdFeVh33Mgs/
VTRQaTAaqDQYDVQajFhpMJpu65ZPGoiDnPmT55el1FzZXpem3/ZG6O1sypmrhCNs6AgcOkKHjuChM93BBXubcEJS4cilsYSCzQ2P
PTH1dB+t/WENvBeO5b0g75Fy/Udrf0Den7CTnS+s6JhUdEwqOiaVwPqt3QeHwT9RUI3i9fsTrgqqUbx+x/Xv3SeNPx3GGTGUUrS2
5rNhbflNX6Q+1HFU6TPl+S/vfxj//T+/e/93Hz+MjT++/37683fv//fU94v3345//cP41z9//OusZOOpOGoKN/JEJ6WwkDvVtVZQ
iublot9Rwc2dBDTU43EZ4p41RQHjq4NyyfdpuL+nkJrGD8HJFWIZGEhDvHgWKCbsC65p01NlTLXFgVQSg8tS3Wnaxtrcpt/RTKsh
EosFvDVhMWOcadlzplX2TIdBSpppoBIxVO/WEmeq67ogdcZEu4m31HCMm8ADH2+aaZ0z0wZnGi2KRQUVEJHepeHZTFs5GOk8XWsT
5QR5+qwlNZ3uOdMma6bFTAq0O72hKILEmb5LwaOZtpu1uLUvgp/qu1S8mmpakQaKWONcw/uh7rgiT32FpKMUaqpknkCPxiLz8O68
eSHoOePnbkRSNDq0hVGgY9MCwzEUwxd3DufZxE+cHcCyJUOAUY2Q6ChBX/Vawvzv3v/z+99991mf5qe5Ps37H2al/5BZ/2VzONqr
kxHiGzrocYy8Ih1/aI81v3VfS4NVKXNhercHcnLY0ERZqDRvbaVGOXBKrYRse7RiyU9CayZUTlUmdArCYiooR6II2En5HtqczpXF
AmdGoU84Ma7B6rGqUiOdBX0V3HIUe72HXSV7qHKMfTQdjgjDNfqF8T4j2pU5nam4Egguto/Ds2YDc+o0MsUzJdj7qBabQ8e2K3OM
XsA2bJA0fztWPUbmFKeMI+LTqQMnoKVXeeuPIWmuRpmQZDToju6UUaalE4HSVOgNHhFd9DmBHoEUUtwrphQbHh7ZaYTOUYbZWItg
EWXAEY2qqYmsEFX77fsfP1HRn9//9P7T5K/6aUQ+v9nSyh0XmNNg142TlrFcknuZ0i987MVabB+bNZfpxp+C4bC2RLEOzBOzBtcN
jIem8TC0iC25yu0L71iOLiprViCwHmfFsO+FoXxEVM6sjAtQi3DuhDl+/ilW43RixmuYi4qBEAZZvAQhCFWqi86amHF1GAaSa3a7
iPIR0XnbZRwLzbxsW27H6Bci4GI1TicmDFqkkBpDu9c4dIwcXZZzdDF5EwMjoi4mZvwMTVocT8yohTwHlx6PMtOkxsXEwFo1jj3K
oESOK94x53FNSZHRpsEKCZGNJoIXbRHaNTmYn1GVYBJT1dYbQZAAQoiELZ6hnJAmTWxpcZDjWQ9ggMKaIIYwgv2bXD+g0//+v/2v
//K799+9//TxV+//8N3734wfvxr/+WH67/d/zMIDBraRB4dDqs6hiBl7a3dCuyKbqSJNAlWvUeCCUwGdmZacmapdk/1EJXUshZ4t
XJEL0IYjLuFsWTks6np+/GCZigCbU//ABVOpxsHsWFgnRM6Kp1zYuDt8uyZns6NpsRg6X45c83KQVero6/kJMC7A/6iAzpVOF20r
N7LO2j+sInTObW7OZsNPXKnJ2Qx5foYgWhI8UuN33YIx1zNk2ZoaFt+PwTFm2tU43EGQSYcHLewgD3u5w4CczQ/EvhlECuD0Sd65
WHewnMIFs4mYWA000oob4M4lDnspXd3Rfw4ZzIbcMoVvINm3ALeGNOhFcJurer1GJ/OlpAIOcg8LCPgKFG7xqeG3UObH9z+8/+UT
wfzw8dffTUEwI5T5Pz//nedH0EycLUYsBvQ3w1O966DH1ukDV1bn6J2CTDXqYhGAqw66MO4euLeCQnhvJYUMuZvHz8rBUdeTpDBw
y7LlhICXN3TQ42CSLI7J8kZyAmRElzE5myS/mSTPvSVtJ0nUKaSvJyngU4miBawlC3xl7TzpjHlym3kyXNFXLVAX00GXk3myEAZJ
O8nSCrYerVTlqjE5kyROYTiOiq6dIZMxQ/6K8D3iHUl0UCRvetDdQJDTUUiZ85XDcv4aQWZb4j5agg3WQu501af1Mtlz00cjfq4m
lQQ9RqTXEYjWEFAtbIox6KPOyYxNvgQYJhkxoFdsAh4o3uHFfv/Xz7jZH97/9PHr797/6//y8W+mIIz3f5ooxT9+M7lLMsGE4YIe
FB8kBITq416UvdTJuOyagQug86BP7KfPOWCnCxXv28PgB7e535UqpXIcR4LzGx3EP8R+6hzMmefXkEcjitcr2Uuf8zlLUxZokNBN
AYFFdnvnK1Uqw1khsXiht+exB9Ma76XO4T6DbaYsezfvte9zHReLt+81xgguxmTT7AsYK1TJ5MyYt1xgj2JLgtp+6hzMmIGTkQ0i
RzfgazR+iz7ZM5aWkYk4ZR5CjRqWUZ5LI9JdHZ8caJQwszu0TFq2UwO8T0Jg7KiFqMxIwzSFRvQcp5P5kxE8LpHsrYwA3GI8fA/5
REp/Gv/z+T7z51GvP00f+UBEcUaNf1Hc1N9tVmIzWapHFeAaLfZTo+CefA7wDT1TTc+PFaqo61lRaNpZqKHATKghtqrBz4sGPcCD
q0kPbQFi+L33rVSPk5k5uBlDiiDNzHgzDhWq6MuZ0QQoFPCYg+PC4NtHsxIH+yXyLx90/msN8/KCjGv0OJkXR5BvxXZWpD9/3onh
7uBfjFGmMuZ6zzjM4rZsQqulhNbQqsTBzMCblGcfxxSYnFe4UKPH2Y6hW6aFG4Ilq2zhRcr6mmVyCg8U5nJYclSAMVZgjP02RUJ2
0Iafpc88kVR2Ck7YtWN+8XAIDkIXfU5mS0FNBAXVmxVURVBoCNVLEMl//Ezk/fvP8I1/eP8fI2aZkknef//5h/8wwZgqV8UWkFsm
eFJvrlC6q04HqFyRa+AgohOeq9T2tapdqTNoTlBGBwotp9uLAf/ti9+0Ri+VM4HsWIHvC9wor6k+zTodwHNwOEmwGsgrQM8j8iWI
rFmpkwnUnmVdhpgcyJx4eSWp0UuXRSxRzDkGLAEKG3eg7KoTP4FacVR5WrBlad0WJzerdOaAgoVl6QkQYKKDF2On2le7yZlAy7xK
6oyAjHaNLh1RWGPAsDE8HY6pbM8GVTI1YJcxCg2ssg3Naz3TvQHRV1EeRHemVRVfLkI91OKnMUIwFoZIQCbn5PygByDdDBnK/BxQ
zi5ETALBwE9MjR2H9CU/ZVRocre8/+W79/9nSn39+M2cFPvv3//H+48f/34qaDLFnkx6/2MBwMHADnyOFteZCOE2BQ+eQNWLhpdO
UXWbhmchBpuYB0Mxk9sQA0shBl0HMiM6ZBMDcXBLjJhTcJ+Ch+EIONNq4IPmjkIjump49goOZVgD+zBH+2U0saanjhnxJZKr4L/F
u/bAm9hVu6w5Nix4EhgaFG7T8DxgCBOa9cC/5SWPm+l76GREqMTNOHp2HDco5j4F+ZlG6gbFUjcYrFXiblPvZJrdxrY4sn7OYtyy
oLDlnmrWRbqEAR9z2NAScbOeR/EvAR6d1MAH0m4je8Odap7GxWz3j4DHTWBbUjJCXIzenuWL52sEkD/NL2Kf6UIfvx7/meqX/KE6
whavuVkYp1qTrBjbyNphCCW1ncYlP8oW0JXlskPNywYoUekaSclNCKeE0LMNTw74b0IXZQ7PWdBG0xmmCcFrSFkNnbQ5Q0jyJdyW
z6cSh9Xx8lXSpbhXc8HIJ+Gc1cpkhUbLg+DSI06jam3OZmuzgKAkCMyWowFyrn6AMnCNe7mlMCHJEPOqYidlDmdLnF9HwHcV6s8c
07yxJJeoaEO1ncjDJnGzmkPkg17lJt22m0ZHKGSjkwTLvhYnWYNfKBRXVU9dGeiQr8G4ZDlgTQnIBX4t//13n7BnBEUf/8fHD+N/
/9MnPBqR0t++/3MZ4JDxwIoh5oC7onoFaxWKHOyysNnzlBKBucBwQqsOQ3KGNQKmypVe/PJ0UTl4cJtDI1hIaA+qVFcpcn0IgvXc
MBvYg7oLVWqczw0egJbzslm69r6+Meaqo0t3T2BfYSmNR8kOihzgP7FZKHQ9xCSe4/eMKk3OTBQkmno2U5tugPuSfXm6mBzQh9c7
yxJyRHQltiuS477Gy5Rgr3bu1WFYocnZBooHmXBWsTNkqtTJQw9ho0ugKjcB2PvQrSGqDtts4OC2wMHwHgwBNzsp+gzPyYRNUbGb
kwavm1iTH5gWxUvW676C6ghoFgX/+P6XvJJimwddn4j+Doq0RwwElD2U2VZ82SQTgTbxqlbFqye0Vh2m6IvBquFGRabAMOYVQH6K
fcmXKVBKZUzYeATZxAiEQ8RX1X8tN1arzdGMjUiS1Nk+6HEJV69FtmrVOZ2x8V4ijaYiV8qyFUwxraBWKZ0xY2ZwtIB4L8G2/pfv
oczRhGkcHCWPqlkcFfCrVed0wjwuInQIen6H1Z9CJmu+jApsSV57tKZDD3WOZwx3PHtEo5fp5ZWyVpuLCfPjjBH/FnJmMNB9epqp
0+m8apnhCBAhqEUhAaLH5GHRSZ2DOYv0rBwgOy6KTRl1C8xSvpNCZ9MmAzDCewiymYqnk7Mb8/aE32Zk/G6U/z8n3X76xfv/+/l2
85fv3v/D+5+XVKKfvnv/L/P/oiibOMLe21YRFmw+MZyV6h7tNvMaD2thMGlGyNqg71JvP8sRLgVQSNcQELeaC2p4LYrapKO6nmC9
Kcx8VZPzteZiP+0OJnjLwGK4kpBbj766R72TCeYv5egz0Ti7oZuC+np2HRa5Rja2wIYl6a2rtp92B7PrN6dLhnrmHvXOZpcNSIKq
KViW48Xl0qSguZ7dDX8SnH4HYSqm4+yarNnFo8Veze5t6p3N7jawECaYzhcnsf5ut+17iqTihqqGNIwBHlzkphTL4YNLVyX5qZ7c
O2Dk1reglcPziFDnvrE8mfLJgbMxeYAaJEbySciHenlZm/T7/USi88l1OAXSfPxVpqtnwnjJy+2oggQfTAipyS0KbHPVLNavN+es
NP7QOBSqwKSnJT0mUGQotYkrTu6xTkShHup6LqZRN+eTYQ8Zpgs1OJoMOWjHXCp5FULTIJxOhhzMPksJ08ygULwoVUNn7QsaCKzo
fBCtJRo0OJ4LZ84r95tDAoFCFS7mIoV5ArSkg99orKsXytQwWfsicPtCHuwL16DBwVxMGXrmfDkofOFRDSqczsV4t12R9P/f3rU1
x21c6b+i8su+rFzobgDdyBsvI4qiKEq8itxKqTi8iKRIineK3Noqy/I6fthyvJtEqXKcZJONt/ZR8foSS778hZl/tMDwggbQBzjd
6MaMNn5wQo3GBoj+cK7f+Q4LVON7fgAqUVbeR/mw8/VN8Cg9D546LpECU6QRVUTq3gRwJCKIvf3VXaSPQjB1Z4lpAhMz2pzCIw66
+fWq00DSvOWetPJd2pwjgiKBN9lW/F0ysXSj83lPov7Lztc6a4ADib7opZU4wpRyrlSmd9q4l2zTLZCqN0JRGwkwdWXDO1E029I3
iMawuX42skGjqimevKPH3xFFnFPsZNKyu2RXZH/LZNNGLdwLdE7sXaosJ6tKbMzKnVScU5A+G8DxhOCsOf6OWPU5JcchvU9c6YJY
6YCawb3A5+RJzTX1rqVA5k/auJnSo2JyM1SOnQPlUXmmL7mPeqVCqS8SAkcVAJGb4b1AR0WrXikf5Lka3knpORG5oybLNzBMjoO+
o3JuSBpGcBk2YfpW8TSqEilsImbrdoDT6s3pXj8eiYUhD/PK2+880yPDsENSRMchRBp6StLsoVTTEL40jKBIy77r/Nj5ovsSHUtw
eeNz2ves3lNFzK6dORIuvcxUsruBkllFZLtrdPHi0+fyzjKJNeAB0YK0gclH3gGtfvRpSR9IObx3g5L0D3Nh9XOX9OpkLgnHSLtj
r1z+0FUaMxLBkID0tZLLM8wTl05bJkwTIO03ujbw0NVjvWq35aOPm2EfOpF/d3ltouQc0t/dD4r5FHAHfvVzZxkjQ5RzGIFMQfZM
rg089yBDVlJOD8n858jsFy979BkzJy2nkifOiRzqIt+4UkfMJZKmxK8TKVtLJmmKjNszvb76BCJPDgREegIyVUMKBNDHj/Czqa+J
o47My0+lJF4iQgjpSUSFSbuvOt8mlYRkiPXTpGvQ+SvK3QbycFRRdkjmeROZ00ONr545CamOkXYDlFUDIi+xyw+Faly+eBByNkGV
a8YVIaov9B4BrTyAZJSzGLLLdRNP7ocI46sDByDSy0s+nwPLWkPjy5ccQOxzr0DAUhCwSFnv1XwErPoN8N8NFW4wUjItPL0nwKoP
IKGuKngoVFnnLUy1aly+7A3g7yp2fsqTtQTcklt1D371AUhxiK+EYCbc94yvDrwBPM2ZpcvLwtHpATA99Pn4N+DqHpi0DdezYIJK
XXIgyY2rmItpLCiJZxdEr3TvQH0OPLgGoZyapwEZT5+BiLRggHDLkkfiUvYrgMK6tO6V57LfT3q9+h96SuffazrjIHXGtEK2Oy9V
hLss8Ar4Si8s7beViBY5LhzuumXGx0uxz5UNJiJXiCjq4tV+N37AfnlJE56gwl0WjHh8hbkJlPaemfy65XZGKOwMVTbzcpVK8OIo
F+spCjrQmu1Q/7JgcKPqoYa1zhjtVFOAAcmtFNVw5MuM8KZp8U6OJ4iyfkVNLgviWuXFfVUkn5cvwV23DNfpSyU3qQOlWgnSgKEd
53UiK5TThukbFUW41xnpMMPUYQa+SmaKp/0z4RleusxTJrTe6yOXQyfp0EPJc3BPYVF6NeveVpAbshvXdJyVWzeY/KLXvwvQ3tBi
IyjbA5fevbD2bZSZH6J0q5GytiYU6EDcC8rLhuWMokBOrvzadwEcjEQ65MrKNgH3oRjdRrlfoAq/IDWZKchuwt4LM31jKNBe9mvf
RbWH9qtYTswMpQz/xvjptKbqYCqCI8S9+DpvjByXUkxBxugugLIEq0pFMgw4v/ZtlB1MmBbofNUKIXmfmpmPQTr71KZK0gLSyHxE
M1myb+NOgGw5LRvyUNnIlrJlbgQTrViAy4sB5axZamVLCTwRORWyz7svOj8i5/RCyayTzOZuXtGv1bxo5tGHUqQrJPWEtHtApEVo
kjwAzWniV121+JhDKc69blRKEgBMWlEpNUqjyuvSykdMSU7SiUrKhIGSFqx1UeARZ7RW5FhKuqhUFyeV50qxTzijBsRSBgBLZRZ8
ab1HMiZecWmGecipqrv6CUvrB/MBbNUl1Y+YevLCAvV7A/U/q65Y8nx9cn1ZOSxNJ+z9FMBB9YvjV9uIrNZJoF6NIrVbPb2LAgAO
MpZJUq6VcyNZpl/vqmUIJuodcyQd00+5Oyysum6pZwwlapVkE7mk58C5iluVL+Airqt+0GEkXzhIySM8lSXgIn15BNG+cMmzJn5G
CZRI7BHiS5uvJGVSErJ8hPKm+283O5/35AG/63zV/dc6iXDazJDDNqHcIsPyYYHZjQABZBoryXUotUIVt3MnZTGkULW4JL6DTKoi
+cgNfTsUUwulCl6vmvUcGEMFkRTHZ1GRacCrSk3vpOyEUtq+zOgNlU3IQsECfTsM04Xx0PMDwviEGOqEaLkcYGZW0rdxJ2UnFKR5
GFPOK8KrXDVux9d5h6rLfcL4ufgoK1cxaxKAU1qmd1KeKXuKzeOh0sqxfG0Jezu4ZDnR1ruyL1EkTUbLOr4yy8v4VUKmzHHKeXVW
QhoiS+OTKDUwEbX2cMqyZuGn5S8u859k9oEUO+Q7CH+4UOKJ7+f9i5/qhA6pKo+0+yMrbgk5JsP7AK2eh+5g5ejfhjdS9kpFyldK
XXwKTR8LxbBkRLleTSYJpxbuAzweX5mYq02esHAjyGq6crukVE3PU5fwd4Oqp/vl9VIuNzo8C/dRfTy+ZuHW8EbKuZVCMVEXKsfq
qeljQVXVRflCOhtvj4/qRNHyWTEbxs3Hvz2K3awy7ywEh7vRd4OsradrYqP0VqJQyQgnxNCuYEOF1AUJ9bx5+hIJYQhbvfo6l2Zz
IoApHmaY4qpo6rPOq+57vT3b33Rfdn+pGSH4ypSeA1sdaY3rg1RNUc6dqoy2sTeAfGnUU8K8pB1VdRd6kQABVnVLHUJR4/qgDfPQ
Nowp0y/sDZTzBlWuRb28ieuCQa9MQCocbX5yQvP6Wr3zEJhUrHMD5R5eRXNTH4PQBQOK76aop8mcG1LjbcB5dFEuYERB/rjmDZS/
DZ5iaJSoWLS+Lhr1u+OSB2fqmS6iaRjxLPKr5xCm6ZrE3ZbGqaLa91DKIw+lSE/K7UPJY/Mo0xHPnMrvO3/qfKI5O03luT55rNHH
bNkrvyIwzOjJviCokhMR+OuVzS9mfk+Wjs2qa5FB7pwL10WMRvuZJ6suOoagYGz5FYFZdJHZQqAezspo6gj8JcsfrizawOUOuKdo
0uaSlMJ1EVPQmUlgYB03A0OK8iuCg/7qhRNyHYOBxJHyS5YPPcvIjao4+GHFdatHnZMoQdKRT99RGUYEpKWWX7FysLz+YWJnmhNa
d/p7SsQNOZvNjJOXXhc5yRzKyBU+UFUOgKpy5UXVzzf2GtL4MveV9C/JzXmav2rZ0HKYGV4n0n7j7IBUOjxPhKcMev4zca6dL7sv
UT5O3ovAg9QOA1sRgvLuT/W1M0/el6s3QVEdRCbqMIn4xbJ6+uiLFw/AV0lVSGQsiWwQkPQOknUV2Dug1Y8+kh69L2/FCdJh8XQf
RY4Jhr428OhFevGrn/OhhifPintGFy979OKaWHPFAMttc8lsB9H49Vn1oyfSo/eU65F8CXg+/tgZ6tEH5RmOJ2+l840uXfbgvfSt
YxI/TrbsEmee4c2NX/3gA+nBE1XRK4N5ll0Jg762+sFTmv7iwFIjw1P38ZiX0krVyqkQVPkvuX6pd/VlFXAvlQBPiXueRK3yJKZt
oq5NTG9BfQRcMjtccjdS3VWkUYUIzB9BySkkLjQ1fpykt3Glw3XRKWYZTS7vX37+j+8sH69uHj1aWT5cO3znZ//0z+8cHG+vPdpc
ja83Nd+afvTg0czs9NS9sUceyXujw6fHBytrj9Y3t5OPvPHzndnWxGZyx71//dmIvzCxcrQef5D8to+WL+awX3X/tfuLzlfdj5Ie
ebIr5kbnv3o08y/evfpmO/nmZxfs8+7zzqv488zvfrnHe/nk8cUHj/ZWjt75WRxEBIlA8sZm/Iejg+O1+BuPLmFzKdHQvv7zRelt
+dH64+VH2+34P/nO7fGx5Gm3ix8tP9p7enh4+dnk+GjvW9JH119bP7r6N+Mnt/go/S9ef353aqH3zaOnuf/c9QdX31hbf5z7SvrJ
9QUvESnd8tWXi3+TXjp5YqrblP+m92ECVGtwONnfvz2zfcAycPhzb8jgr51vepOGzzs/dn7ovpeBQRVgssC4YB0VgBFQRi+Bsb68
fZhBxkXWISGDIJBx8ZgzwLh4vlXAuPpW+sn1l3IQaOdBUo6Jq+/oQEK6dHryKSKzn9kFw53bc+3RqXYWDPHhfxAf90c5ACRH/7fO
6+77l3+XswW+2haIxAQqbcHFuy+deJQ/8etTKrw9xdNtF63DcvF4HZy4/PLWNwVKK5Bewurpn995wqef7U5nTv8vlz3R92JzcPnK
a+HgootQxEEv5VfjIP/m+3kcXD3CcpeQtaLAX2C8w9Wpw97BlSVIr9yIc9jYaq3tTi0dXCOixenm/e2ZjQwikgjgda/M/tfuyxvJ
ntuLYOmTnq7aFxl0/Cb+q5fdD7svEveQ+xd/3+u0J9HEv/WK9Tn7wdW4YR6BYglWhZvcAZdZERA8GrhBGJTrr1iPK5ROpEnsbE/t
jAYHrSCDnc/j846DiaulhJ/HP75JYowMaApfenUzBlJCyXhdNC+ROrIIggiMLMIcTli1fbHpZ6qDTRQqYHtyhTKcPXERWBTA0GbP
7uzN3HlSZkiSZOujBAw3Y0x8dbPz37E1qYRGip8cLtSpCI9zLRAX+fjDq8aFIuDU8Tgoo9GYt1GaDKW10MMHRYQeT4dPVzYfDxsm
pZ8mnd8kDImz9deKxNSHnImAnMlFutHOGQ0rYDBMP4xgUOI6YCuBx4Gm16D9rU4AIAh5BIEg5ykYMYtEccUJswiiZmEiGxE0HUYg
ANGen5k5XZk7yAEin3RIIWd8+EnK8iZJWm7Ef3wvdhCvpSJGzjSo89QwSjZq4RwFiQyLVg5rE87qVejiRG0ojJPTrcdbT6bT6sQZ
2Xh2Z345A4X/6n6YBADxP98lxdpk0fV7SaTYGw1OUfGX7os4ge1tW7j4wredLzrf4MPJ6wpmEQ2sylPgClU6ccO12VHmIHoVjOvr
vi3BQwEX57OnG9EEGc3g4sPeGs6kbhEf+e96EPnf5IdkU2evqGkCj0Bdx+ReCTyEQYXbZrpRqF5gklAFKKrqWjbiibrQKOQdG0OH
J9He+Ghp3hH/EEcNF/8UoonySkcOHmrrEXlMQBGG0I8wkLbDEjjcljxVCam1kicCH5O31/bCoeGtbHSRlKuSEFLZ+cCnoOp4k/uc
o0sThgGnxRo4ohNmyYOgsWDdSmwdHk4uTt8n+qWqD3vR5feJdfg0SUv+ofvL7vuxgUjiUmxwEUYc9h5eVU/EHBBXBzkA7dGqLATd
GdGDBqtOPw6G7009eLI17bY9StXAiJMQCBgkrIo6i80yheNQZKb2Y86aFqMEF27qmQhYjN8bfnJrb/9uBhaf9IKILxND0X2RZJ4f
xGf/XK9bpvYZF4sKARMR5bJRYVTOdtPwKBiKLGaa6JdZyE0RgJhq7Qzt7t+ZQ5cpPuvlJq9j9/JN/NMHvZLFBdEoF1MK3aoV8037
YA3XKFC5qFHG4bRvyqoDiqlnk0tnM8PtWgFF0h99ft0GUcUT6h5HEAl0rYKahRM4UBQ9hvvClWYQYcE4+NXGYeju1hynZ2uGRW3Q
r+QshafON1jA0Xjg1VFE1oKXxxLFnANyHUAskX9pNUAC+48+RZsIoBy19u5stw6IRjpaGVGoUREKQdBJh2dW0VQAophzWCpKDHYO
ijj6xaXRnbnNgx10APHn7sdJAHFZlurpW/3Q+Q7b3xAk5FBNilUVIYpGwbo9MCpEOKTk2ggc/OrAITxuD41sTZfyJKQQ4T97Cy1f
JSWHXyX1S7NCFUTb9kJmTNs2LluaWQgjcJhXtLF5hkHV8mq8oMJcbE9sLy8/aAnDkKLKa5BI3emI0kQ0j4koV8qOzDDhjrVpmb9d
EkDkgxlMEuIOKkP39463V4eO0J5Fr3h1uSlIkYZQkIJ3MWWbYkV4tqpXtaoVxUDTGUSwyYhWnIEEhEOODSXaJBtCcnxv4pkOAbkv
XxXdSvY/VjdHdVm4QKLD4gwIERRISRlsGnLVq0jUSEnrcSpc8/4br1UgEeCquE1CgF2TzoIU4RBVeQqomKkTW+DzEteY0GRRNOUy
xvnmzO7Bs8fFwoQWAqhHAYZEBE2FRfn2hseMYoWatGyrva6Bo0qoUVDITsfCW7d39x+MZVCQJJ5v4ijhzY2LaR9zzv6lqpbCXYDe
gocG6KgROmBK2rbLlY1XKpFoaIg1QYQAaBM+WMGMvKogwpw2YYgKdCVbz144rGXfGppoPerBZPju1MiE5jDpUBQsPj59Ou0o62Rq
U+FDSYZfVbBCORFF36tYryr2Qu3wZ+w2NHCRQz0MDIi2QOWA8U/SAs6h0FrZXN2/E941rVd2P+5+HJuNX1zt8Ygzka97k+g5s0CU
YIDq2HkWFQmsIcHILmCo2VffcTHHkV7fHRLa0xvHd86GQ7RjqBrug9oXZR1OUdX3tsmDKB58dW3aXR5RW3Cm3vE7I0oxAfgBUG/E
qyLO4SYzcFXoYow4cG8/ssBU7/zHZ+/eOxg9HbYhMQNM4/iCYvuWZqP/KGKDSRjoehKnBj9S69QLI1qrC7t7x7x1mDv1V/G5f3wh
AXKj88f4kF90f339h863SerYfT9XZSwd+MsBhAImAaI+kBz1gUXW/IKrPEHLIDihUmsho1BFOF95Msw32qGxxIxWMQFSpeIwZ5IR
fZuBpMOguA4moLDWj6olBVAPGBti9elZa/J+Bhh/6FUYv+v8GP9vApEfcsZBe2JPnS+EHgXVqbi+OpWpC0HEDEbtB/N5vabyxwIa
VtfX59enl3f0i41olZlcOKluZLOQQLOchBvQJnHQKPanTJxHERpvWS5ZQMXuDluiy7Nb+qhIYJBElc+xJkGkJqFaOSi0dvJGRqHP
ahAW/AMdlMk8QEbKC3ykZgyJ3j7/YK0/WT+3QADhaC9a9sVooClbaaHbQNPRikKmmYsaRQ3agmE2YVJkKmLh6js6UID5j7WdAgIP
U+HS4q3h9ooGk15tNHAaYh6Bx62I/pCuQ1vgqAWpV2psxCQsLD7xTs/8YuMpNgoxAH7QKjMxdc5I/dBHKhmbCRmjXntMwoihtVot
M6nOXfXmJ5/VOPZCOHhy53DBfzh8aFxLQE7aATM1AWUEiYhiVKDhD1yLOWDyBJPmI+wWameQCHBMjc8vnT8ULRuFpsrRXGAWM/Dg
pnSgz1AwbkZgetJFq4GrTsMBgw2aim2TcXC6cLy8QQxUiv89iS87397o/CoJM7sf9pSuf4nVPw95RLDzd8zy/H5zOaVdYSgXoUMB
EBMz84fRs2dr9QtNOVkxnBYQC9CKtGY6H01SFSzp0brhsNFBkXEAulTC88C5fZIjRTNmcWLGlXnQKjYa6dI6KTwVS9F0lQ4dnt3O
oEKuJxpQn5naS1BCBTaiJLYYjZhis6URfWuxQhPZxcCoxFURWX4SibPoJhDaT3R5mWwH2+PG3Maivgduh44fwOuzuJUytDWCm9mO
i8FqWmPkAneOTo+j1Ww6EUcAnR+773dfJO6/MDNnQyjQA6fyC1xn7++pXe2w4oAAw8be7ubhSThX5jA+iF/9F8nLH7uNv/XkW4x7
FOqSAwH3rUX6Mi8OkYEpUlpqVzsJIBB4GBtZW14kc7fwIzGXA5a5g1b3JFlvn7mamWBns42zUkLNk9ZktdXuQQ22HGSRwPaTHmRt
WQUtNBQ5jkvh1Njq2UR55pDZcmZOaYPyBx8dInB31sEkUmx2pLp+kREBiJntx2MnO7O3jXsRf7woK8ZB5cve///2mguJGopgNIyw
LLa3sa5gS/rPgasoDtVvLS3NzJ1md2z+LqEmflZD0s3TrilWN6BQZsAwHrSj6Na43F+9o3dCXwQ6j4HHsfbfsoC4LdLCQFQMtHIC
hOzn+Nn22dpEyzesHSGpS6E+dcnXV4p26AYMKSwDFBX4A1BHBLoMPiU+shdN2GA1HWt2GeqDQCsaQGBgor06tHXe2s5goHKloXHN
CJBuLFHr86vICfa4bXZmHFySGeu3F/yGyYyAK+Bl6y6JFWaS4ZHbqhvYZTNaiA8RJ3/ydHWvfXdsS6+K1H2eJIU3Ov+O7h9weEbe
N5iRb7J0WDcobHxE3h+MPadqIFCwWVDdRzIGwkDUBSzwDWra/WJh4Hxp7Whmb924SAQo86HY7YSCc46FmNBwP4BpJdlSEGArKqzd
OUJAoTU9tcXHQm7aSbRUQ+acU/TqSmZNOsGOgUCVkY0ISTWWhtSDRVPzDkBNKQzgKfnAit6e4Sg0onwA7aGD5LxdbqOzDYtmJx0I
IMITYnceU9bfmLLBhpObgTgEJJrdgw3MR/kBntzeaFkBYS9wsy8mVsKJOCcCEuf3R5+Iw6UhG6uGtGVXgLF6HoRYQiMVDhMQRH9q
ALfcWrcae1uHk2zyIFuQ7qUS3Q+TYvON/MEn3uL9Cz2v3iiUvWVU6VLLqnSEUVsNa0ge3iT6tKPsZoHpBKnD10OKi5kISIiF+BSt
64lIQsDlIYaZCKKT2aypqF+rRC6JaESRhXCq3czk+dVk/oBlI+6KV05yU+zSkCZ62yRQdzSY8KCmZpQLNsPA4fYxTEcDUcCC8tJB
6W0iIRHy+97Z/Pye203IBAgiKNzoDvMb6axN12Kiy+qcFKpXFInS7tVhdaqbSFw02/YmTO1DCCzpEuZm7DgdsAQVAoiWI2lW0gWJ
jS02MhmMjC8Xtw4l6WiP7vKq5yq+rQMJQBmWRpBgh6D6iDDlS7oLKkz8h4uKBRILjrS9SBgCRBiwMy5C/RWmxrLA6G10fd491RQK
XE3XEA5JyoOOIb+fjqOnayyq9xjxJGwVtd0QJbA5hr11AiQCCg6RD1Yl86kE8bz+bpAx64XbjRm1thLqIqBQejqaicbmAm+3zixN
970YE3+LjcMr1VpCYAFdAItzRHkBQNbXIRozUAxKyxOJg0b04kkEbBsTDBy+9bwqMPydKwI7AMPQ/YcPvdbCRi5CABTbTHaVBmrq
NPFACh3PNSqE3+SAlfPdIno+wk3u2C81N3BHKSOhsYvQiBsbdBSNC8c7sA7PJk52W3O7j41DBo2VxgQg2oL5RM5bCFJDKdbZoqqa
ig2wrbAiHT53b9RwO9ni7DB/fD79xNHWWl+NBl+AOvIFYozhUpGa688RPQt8hdoIGHVVfGpggk3ztcm1pzlrAVUUjGexAAnAgIVI
bDjdXoqZ2TUSdjGpRtec1TfHwk9rbP+frbE1h8JIa/zBzslKyxAKf+xRpV4n2emNzp+ggV1gOIOHMAUm15Yg1Nq6OndbTe2sHUFj
AUg0zLEwfodvLt15cC+DheLeYsuyf0HA0XxaxIwOSIRC8V7MuhOY/MKaSGwj0YKjDhUQHZC0QVUUiQ7srLlXhY4oU2BGqy46BiMA
9M8YPJjl4s7pyI4WAtJZ7ty5M+DcYb23vIITCc08ACjogksZLCk8wodvJBHejDu4N/zk1t5+NkoEuUsaw/uAH4CVHwvS0BRhA+o5
gsHbO1OzDm2Og/b8zMzpytwB2hJUMp7qGwdhkCoY8xZclaHfRqMw88CbXDi+l1N/BRNCjf1kUIkRJDvmRVwIN+pIqXwDKjBEaHrU
DAyckRzR598PVQdw3z2IBEasqD0aI6EGsVEHDv33DQU4TIn5jf2V2Vu5GOFNfLAJmzGrAfvnOHL4ITYUvcmqGz3Gwovur6//EMPh
Tfz37+O3Gfs8AjeXelWQcFtAsKME2YdJTDQWCv0oujx3n57To7IuZewFkqDgZud/4h++qDugC+g7lIgEB1Ui0si4AcwqjKYwbVcU
qsIHPSFp60BZ3F8fO3g4NmWrcakAFUouLBAeOI5pZ/FIvV00JnJB7joTjdiQ9eVgJFiYGbcFjT9f/QDOcDNgKQ3HQoOGRgQY43qU
0eCVFmfe/dYyc4CMbZzvL+4P32vOdnjA0vQI1qBmDSYmkPHIUxDqmBBbGYqLjlbRhNy7//jh6CbNIKTHjo2zz69rBRwhIPjgg81t
n1QFHKZsOYj2YGQwoOzETkdLb39RKcXaHBb9UIEAdtwQTghY2aLOZIXwQhCWzIZJ5NFYZavoWcJbt3f3H4wVB/SK4oMWdSB8H8uH
YEENGYiazCmjES1b649qS9eag2J1fX1+fXp5pz7zNrYzHyVfupmUxa+lyLBa1iwEhSwLq04MuXWGuhDOWZYDwJopOpP1hw8nJsbH
nO1bBygzHoeHuAqa5p61vMSsCNpYZ6zmVHcNl+GPMHE3mNWc6Ub7DECBLiBg2TO/UruoWFlvjhNfvEDEEM5GNRpj2TaiIAX0yxkV
kc1+OdIW2A0a8G0Rk2CyWRp+Hwc0IIgwERpT64xnMyzJizkLG6ys0zVHxfb2fNQendt3jwqIbUdg/WvC3ErPvQ3kipqDfgAyKEJW
6njT374ttk1p+RUMC6pul3np2E5lTElsp6AKR4LuhhjFl9ak0RtBxMn+/u2Z7QPmVIsQqHD7lFHkGg3UPhWb1BujDuog0/AQWBja
X9hYb93fMLQOnyaEzc7XNzu/v3Qfr1XNDgJQb/BxhFdj0FOjq+7WSrxVwGhosg8wEywEV3IWKlPMXhpiFGTaXtnen8wDAYmt5dW1
2dXjrN/47x4v86skokwihBuyrEiuw6E588m05vwKA5/2ZruM2DdGcgEuWXrWw4hGhvzU0aXPwKHwPIkblYKalqncVagGh69J+z7i
ByxqpxSe9MxXLItMK4uFiKI1AGgStqd8nC1sN4fCQvuUscXpcfRYB1TYzkHAA3relKGzzLDRWlT/Qka9KMG6W2itzYVtf+4WGgE4
RwAoQPgc3oriOQOAJbJ2PQRY2Ipiv7Rw5/Zce3SqbUFyEjpxGgp0s5Jam9iopQdjONvrUovauuG3O9kbAYtuQvQy9tDaikUzaQej
yQzL7Fkdodka0Z+riV6ggmhcF6gnDmYnC8TE/oPDrUec/lFr785264DkRjdfJLu2eyGdqpBs1FIgjMHBHmlwF7crutIAMRsxBcKl
0Z25zQMdg/9xUjy87EP2hnp/6O3iRdl+5vkCu4g9sNc4MORCozz+W17/OXm6ute+O7alpzTffR4H+sms7rvoYW1SMrIfOav62Xnv
++D1G3n7547Hz1sT5xt4Mcjui86XsVv4RTLRHYPg/aK/B3RifRohO4bFoSkNl2+vbWiZ1q736lsYtqSIBalL4dTY6tlEOVX11c3O
b3rCDa9rDchAHoGHUP+wQDgJnUu7WFp2+JYVAArAaHG6eX97ZsOYh6QzPIMTBAw8mN3s6Q9SORUEtMw+qfIfbjjOCJi0yZkXDp/3
fyKXR+YTuQ3wnasBUrOSZEOFvHZJoQCPk9HJyYOd9d1Ca/F5L2n4vtBa1nYpDBAUBgtMBR1ZPpAzl666zFj9D+uW4uTO4YL/cPjQ
2FIg92JCUxGUYWU/ij3GmtGnTTPRfLPJRfWpgI6p8fml84eiZYyOD3syct9fRRfPr2foFDy2QF2QDDyY0hhUDc6gNmPa26aLqkto
yTrYmLC0H17shAcTQw9atQcsTXMTmu6/KgadoT4bXgEKHQlSc16jjVGa/teuiovR1v396aGVUlWpD7rvxVnI8x4/4W+9QqXJDA0H
hu1g7iutshluJy4xRqPmjF3/KUwFQOzusCW6PLulbzHkObycv1CHFIL50BosmjcNwiwfNVcmRoSYzirZKn8BC4jZzzpmViOxtr/I
jYpa+PkpCgzMlHiMoCr3wO3TbXQK2x0mGqlQbMwczT94ODNWFlnmF5/ZG8ZmFKxtUt9AmWHQduVZK0i44LYUN+W1Du7NPG77tcNJ
KdtIsPNdwnu82aNAfo9mvRLCzVmvDUiE2YofBqcXyqrbYc9G/IWJlaN1Qy78Z72i9ssk2CxGEeoIUqRC5gWV2soIEqlHigwj7MzI
ZPE0wJEEAg5Dd7fmOD1bM4QDtmQF7NstaZgGVTQZjclbW8RoBDYswcJRvxQBiJDf987m5/fcAgLqgJXNYvOq0NJ4y6pRNGEmdu++
cmndY/R3Wx4lHkevwWCNcmrsaJwPTq6BwMLi07mRsdmF24ZYwBLrPH1aZVDFrLMxh20WUlhfmNU4yRKBjKlny3zheHmmyK61ItIR
eQHaMVCXagxmnEpXKsMuJ6hYXyjVEJce3pJIAoMkwnh6yohWh+9NZL85KP1NNhiLcdTQCDlYpI7c7U7EqwnXGLHVyib0Ri6cZBMF
VOyeTo7uk9MHGVT8KTYBLxJJ4USK5dvLhTfaaACChFJP4dkpSBqmEIOlzeLCQRTOf2JrLlo9HM82KsADlskOz2P/kEi4xF/6Xc+G
/G/yQ7Ikqaf6VN84MM9UuaffvEr9+lO/ZvLYwGzQgtR7OAGZ2Xllp6hZxm0/+5tNVaeLbP2VJ8N8ox3aYEl9miSi/9D9ZdK4+CSJ
NrFyX4yKADvAxZ22vdFsmKa38rphP7AB3H9D1U0NL+XfVo942pvwdlWWGmQhcjbQsxwAbcpnEXoJH60TeIC1KqOpQESHQ8fD2Ag5
rIPF4YIkYHVryKMA2wH1bJH1/98EGbVjziJj5mjhcJXNRsbm4te9DEYl/wSxKEt2clLTlQUO1SD6OhjaSCBxsDkZ3Du4t2eMgb/0
NEWTFPWvV/0PLDOG8TDCFjNZNYGuXt8bvzsvz0sYEEKddWiEDw/H19Ynl3PQ6I1aJB2rekt7KaQTGWEFgyLXVAg7ZYpmx7mso2Bj
6PAk2utdyYBWqQ0LDyJERNiWpzF5SpFn2KHH9J1faT123J7aGQ0OWoE+v7KcnY0KIpjwCFaGPHRZrKpDqByMuoR1YCwcLB0uDZ1O
Gy9/r70Ui0XUeClWTY2JOps1becbA1u5mjk4mnxA767omw6c8CxgNMKQMOyaLG6RfYufFQeizT7vUWwEEy7WpQUCYGELyD4wZmAf
nK4zMeNWVnEnTGg0yja5yYpmBBrGz/aXN+9uDTewL03tQ3zCAqxYfWhtYANHssFaCWd6tf1reRwNHUSCtGrHnJW9DoCm73uBj407
DTctamy+QVSv3vZtSAhQLD5dXB8RtwP3ugHg+k0usBVNzzCuGNANJza6GnqA8KupmOdPh09XNh8PG6/Hyu/AqLsZq5B1GNJydcjY
RrYB5TEGaQ8vAg2OVymyAFC8FyFWS8R0z4lhUxwRObga9XQ5voOAwslKa/v23g4xhML/JN7hctPi6xud/0j4Ncna9th/KHxFAKjh
U2xxG98MN1YSab791XRQiTEQwyOji2SDFZfmZWSPLa9EI1EJa8azIjaEo0L0QRxXDwT1q9kIEGzs7W4enoRzptpCmgsU1YUJTiiW
bVcUPTVXE6mzx705aUtHHAkENo72omVfjGYTDHXP29hIQDK416M+1boRgfN8E61kCLTHnRUkXM5+IQDiny8sPl0ZmkUtUNKyE0CD
gxMfSeUnvtE2JY3OhlWFOoeaAY3EEp4/uzF+f3rK8iY9ql6r5TEBxxDM2UC4M/WpZtcoWjcDYyNry4tEY4ne1UwwbocawS/U4Jbn
d5pbo2Ry4P1LKPeOdu8/mBm5VWhTpLpA+EoCMLYlCEev0hD2aC/4RrbrbEG3xmhDLghx9lPh0uKt4faK1u68V3G0+M1lCUFreXIQ
54vGW/RwqYGzUc7ie29/bM8NNw4BA7srFAN94ZdQnxJpc4eirT7kAEl8IE69PT8zc7oyd4A+9d/EL3ySML5JUsZEDCqZzXsNWwEg
Bgg9juwxmMqN2hufaJgj7YbvhsDCzANvcuH4Xi7sA6WeNFoLkbb+BzdoNIH0NpxJGPyAoBE/0J69Nce2dkayNQBI5Kn+Xs10s6LB
Xk37tUIET75oCvBaMNZWGbjwDoVB/tWF3b1j3jrM1YPwA/spPEplY1Adp5DD45dcX1nSKpPJLHiwJAWDDRlrpw79XcQHNSKDkCAF
7YshpXEogSodmYziulx/4sJ/9H+aH2g5eEIE5rqjGuNXfaa42aXNu3ArgzTaH6jzEREEEPON+VWhqM18BD+Gg29N6RcpNO2KNXFr
BHYWFja8vbu728bY+Y+kVRWHJy+6zxPjckmnVOtAQNxZJrDUF8/h+pzm1cj6T3opjuBsPx472Zm9bYyHP15MesaIeNn7/98muY0i
KgVEY3SWNpK/dxWA+qPeCERMPZtcOpsZbuvTqHXW8QG2gcHClbRKuNKRmjGmxGU2nDNIIjEIYKyur8+vTy/v1ObXJ9NcHyVfupnU
Qq5RgpU/Jx6DKbXUDqXWtPrputs9iOGoi+1rPtD9Yhy9rdMa4cVu2dtGE0xPGsLC1k5/cFZuqasYHK8lxKxtAq8jHNNYHKE5x2k9
nNjeno/ao3P7fZvgDDxqTpKzyI3qV8e8YXbc5NDsyO3qNsn42fbZ2kTLN96ageFQEKBcwYQHQSLK9UzCwKEbMZrJcbt8q5+YGNpf
2Fhv3d8wntd7033R+fpmT0M9sRiqoT0CCcmwAOqo81xkwanrPNQZjXJgElEkIhbZNF+bXHtaHMmJD/er+O3/KlmmFUcTX3Zf1qBX
E8B1EA6WwMM8KFzO7lnqjdgpVbkh2yAB0dDiLQJQLkgIE7BEznMEFkmYdraxuWLc9BURTc5ykpACaSnoPEQuxBT0JynTZqIJSyx8
Aq38LlmqEnqmEYNLfco+s6+aMghWibgkUNepKYvAvXs81x3nYbNbMkyESW32xft6+q4JuVBKGYATuyInExWKgSNmu9rJ6WiHDjYy
eLq61747toWGQnHWP58mMN00QeTYD0HkMihEbDNoWlvSCfdSDYD+U6kIMLrlhddCpJVpQxj1Qd3c9czeIKOkj2wqEgGigwLkRHi5
dkZEzbyJTgccPe5v27g03gZHQsbhuhQSqX1MIBgDt7LlWFPCG7i1bBDDv9mRHwtpZv9ZEQRaBe4xUCMkpFUxqHEBws5oaE2+zGDU
porkutVIrO0vciN+P7rvScIAWJQAx6U8l5QI7pBoWUO61pU2RD8DjoaYEiRU17CpT8BglOeGRyOv37LXZqR+I2nKBhtdBUw8mzjZ
bc3tPnbPkiBczaqjHGyJ50vYlnm4dvJYZ7DQ05RyAY7mKDSgzaCwwJDI0bRFnfVctnpflktctvLXHjp+/i//BwUkQIxU0hEA'''
_EMBEDDED_ZONES_CACHE: Optional[dict[str, Any]] = None


def _valid_zone_table(data: Any) -> bool:
    """Accept only the compact team-relative table used by the 3-of-5 gate."""
    return (
        isinstance(data, dict)
        and isinstance(data.get('team_relative_zone_thresholds'), list)
        and len(data['team_relative_zone_thresholds']) > 0
    )


def _zone_table_metadata(data: Any, *, source: str, path: str | None=None) -> dict[str, Any]:
    valid = _valid_zone_table(data)
    metadata = {
        'status': 'LOADED' if valid else 'NOT_LOADED',
        'source': source,
        'schema_version': data.get('schema_version') if isinstance(data, dict) else None,
        'threshold_rows': len(data.get('team_relative_zone_thresholds') or []) if isinstance(data, dict) else 0,
        'matrix_rows': len(data.get('zone_hit_matrix') or []) if isinstance(data, dict) else 0,
    }
    if path:
        metadata['path'] = path
    if not valid:
        metadata['fallback'] = 'MATCH_LAST35_PERCENTILES'
    return metadata


def _load_embedded_zone_table() -> dict[str, Any]:
    global _EMBEDDED_ZONES_CACHE
    if _EMBEDDED_ZONES_CACHE is not None:
        return _EMBEDDED_ZONES_CACHE
    if EMBEDDED_ZONES_GZIP_B64 == 'PENDING':
        return {}
    compressed = base64.b64decode(''.join(EMBEDDED_ZONES_GZIP_B64.split()))
    raw = gzip.decompress(compressed)
    checksum = hashlib.sha256(raw).hexdigest()
    if checksum != EMBEDDED_ZONES_SHA256:
        raise ValueError('Embedded team-relative zones checksum mismatch')
    decoded = json.loads(raw.decode('utf-8'))
    if not _valid_zone_table(decoded):
        raise ValueError('Embedded team-relative zones table is invalid')
    _EMBEDDED_ZONES_CACHE = decoded
    return decoded


def resolve_team_relative_zones(
    source: Optional[dict[str, Any]]=None,
    *,
    zones_path: str | Path | None=None,
) -> tuple[dict[str, Any], dict[str, Any]]:
    """Resolve the stat-zone table deterministically.

    Priority: explicit CLI/API path, SUPER_BASKET_ZONES, the table embedded by
    the parser, a file beside this script, then the current working directory.
    If none is available, the existing last35 percentile fallback remains
    active and is disclosed in ``data_gate.team_relative_zones``.
    """
    explicit_candidates: list[tuple[str, Path]] = []
    if zones_path:
        explicit_candidates.append(('EXPLICIT_FILE', Path(zones_path).expanduser().resolve()))
    env_path = os.getenv('SUPER_BASKET_ZONES')
    if env_path:
        explicit_candidates.append(('ENV_FILE', Path(env_path).expanduser().resolve()))
    for source_name, candidate in explicit_candidates:
        if not candidate.is_file():
            raise FileNotFoundError(f'{source_name} zones file not found: {candidate}')
        data = load_json(candidate)
        if not _valid_zone_table(data):
            raise ValueError(f'{source_name} is not a valid compact team-relative zones table: {candidate}')
        return data, _zone_table_metadata(data, source=source_name, path=str(candidate))

    embedded = (source or {}).get('team_relative_stat_zones')
    if _valid_zone_table(embedded):
        return embedded, _zone_table_metadata(embedded, source='EMBEDDED_PARSER')

    candidates = [
        ('SCRIPT_SIBLING', Path(__file__).resolve().with_name(DEFAULT_ZONES_FILENAME)),
        ('CWD_FILE', Path.cwd().resolve() / DEFAULT_ZONES_FILENAME),
    ]
    seen: set[Path] = set()
    for source_name, candidate in candidates:
        if candidate in seen:
            continue
        seen.add(candidate)
        if not candidate.is_file():
            continue
        data = load_json(candidate)
        if _valid_zone_table(data):
            return data, _zone_table_metadata(data, source=source_name, path=str(candidate))

    built_in = _load_embedded_zone_table()
    if _valid_zone_table(built_in):
        return built_in, _zone_table_metadata(built_in, source='EMBEDDED_CALCULATOR')

    return {}, _zone_table_metadata({}, source='LAST35_FALLBACK')

# ===== EMBEDDED CONFIG AND SIMPLE INTEGRATION API =====
DEFAULT_CONFIG = json.loads(r"""{
  "engine_version": "6.0.0",
  "calibration_status": "calibration_default_not_backtested",
  "odds_min": 1.44,
  "odds_max": 10.0,
  "dispatch_threshold": 0.60,
  "signal_gates": {
    "history_zone_min": 0.75,
    "live_edge_min_points": 3.0,
    "scenario_direction_min": 0.50,
    "play_min": 0.75,
    "risk_min": 0.60,
    "allow_live_reversal": true,
    "live_reversal_p_live_min": 0.80,
    "live_reversal_p_scenario_min": 0.68
  },
  "smoothing": {"alpha": 1.0, "beta": 1.0},
  "credibility": {
    "h2h_k": 8.0,
    "form_k": 8.0,
    "pattern_k": 12.0,
    "scenario_k": 18.0,
    "pattern_min_sample": 3,
    "normal_min_sample": 10,
    "zone_matrix_k": 15.0,
    "zone_matrix_min_cases": 10,
    "zone_matrix_low_threshold": 0.55
  },
  "history_weights": {
    "exact": 0.45,
    "form": 0.15,
    "h2h": 0.10,
    "distribution": 0.20,
    "scored_allowed": 0.10
  },
  "stage_weights": {
    "PRE_MATCH": {"hist": 0.65, "scenario": 0.35, "live": 0.00},
    "EARLY_LIVE": {"hist": 0.32, "scenario": 0.23, "live": 0.45},
    "HT": {"hist": 0.27, "scenario": 0.20, "live": 0.53},
    "AFTER_3Q": {"hist": 0.16, "scenario": 0.17, "live": 0.67},
    "Q4_CONFIRMATION": {"hist": 0.10, "scenario": 0.15, "live": 0.75},
    "CURRENT_Q1_Q3": {"hist": 0.23, "scenario": 0.23, "live": 0.56}
  },
  "sigma": {
    "MATCH_TOTAL": {"PRE_MATCH": 16.0, "EARLY_LIVE": 15.0, "HT": 12.0, "AFTER_3Q": 8.0, "Q4_CONFIRMATION": 7.0, "default": 14.0},
    "TEAM_IT_MATCH": {"PRE_MATCH": 9.0, "EARLY_LIVE": 8.0, "HT": 7.0, "AFTER_3Q": 5.0, "Q4_CONFIRMATION": 4.5, "default": 8.0},
    "H1_TOTAL": {"EARLY_LIVE": 9.0, "CURRENT_Q1_Q3": 8.0, "default": 10.0},
    "H2_TOTAL": {"HT": 10.0, "EARLY_LIVE": 10.0, "default": 10.0},
    "TEAM_IT_HALF": {"HT": 7.0, "CURRENT_Q1_Q3": 6.0, "default": 7.0},
    "QUARTER_TOTAL": {"CURRENT_Q1_Q3": 8.0, "Q4_CONFIRMATION": 7.0, "default": 8.0},
    "QUARTER_TEAM_IT": {"CURRENT_Q1_Q3": 5.0, "Q4_CONFIRMATION": 4.5, "default": 5.0}
  },
  "strong_live_edge": {
    "MATCH_TOTAL": 10.0,
    "H1_TOTAL": 7.0,
    "H2_TOTAL": 7.0,
    "CURRENT_QUARTER_TOTAL": 5.0,
    "TEAM_IT_MATCH": 7.0,
    "TEAM_IT_H1": 4.5,
    "TEAM_IT_H2": 4.5,
    "CURRENT_QUARTER_TEAM_IT": 4.5
  },
  "live_dominance": {
    "max_live_weight": 0.80,
    "edge_thresholds": {
      "MATCH_TOTAL": 7.0,
      "H1_TOTAL": 5.0,
      "H2_TOTAL": 5.0,
      "CURRENT_QUARTER_TOTAL": 4.0,
      "TEAM_IT_MATCH": 5.0,
      "TEAM_IT_H1": 3.5,
      "TEAM_IT_H2": 3.5
    }
  },
  "projection": {
    "minimum_segment_seconds": 120,
    "scenario_projection_span": 20.0,
    "simple_information_weight": 0.05,
    "weights": {
      "history": 0.25,
      "scenario": 0.15,
      "segment": 0.25,
      "stat_adjusted": 0.25,
      "control": 0.10
    },
    "regression": {
      "current_pace": 0.45,
      "history_pace": 0.40,
      "scenario_pace": 0.15,
      "current_ppp": 0.30,
      "historical_offense_ppp": 0.30,
      "opponent_allowed_ppp": 0.30,
      "scenario_ppp": 0.10
    },
    "adjustments": {
      "efg_very_high_no_volume": -0.04,
      "low_efg_high_volume_bounce": 0.03,
      "ftr_high": 0.025,
      "orb_high": 0.02,
      "to_high": -0.03,
      "opponent_allows": 0.02,
      "opponent_suppresses": -0.02
    }
  },
  "caps": {
    "stat_limited": 0.79,
    "stat_off": 0.79,
    "fake_over": 0.74,
    "fake_under": 0.74,
    "small_sample": 0.74,
    "team_it_weak": 0.55,
    "team_it_70_74": 0.74,
    "team_it_75_79": 0.79,
    "q4_danger": 0.68,
    "zone_matrix_low": 0.74,
    "ot_tail_under": 0.68,
    "q4_low_foul_conversion": 0.74
  },
  "team_it": {
    "own_weight": 0.50,
    "opponent_allowed_weight": 0.35,
    "h2h_weight": 0.15,
    "unrealistic_points_per_minute": 4.0
  },
  "patterns": {
    "margin_bucket_size": 5,
    "total_bucket_size": 5,
    "team_score_bucket_size": 5,
    "specificity": {
      "PATTERN_01": 0.70, "PATTERN_02": 0.75, "PATTERN_03": 0.80,
      "PATTERN_04": 0.75, "PATTERN_05": 1.00, "PATTERN_06": 0.75,
      "PATTERN_07": 0.70, "PATTERN_08": 0.70, "PATTERN_09": 0.80,
      "PATTERN_10": 0.80, "PATTERN_11": 0.80, "PATTERN_12": 0.85,
      "PATTERN_13": 0.85, "PATTERN_14": 0.85, "PATTERN_15": 0.85,
      "PATTERN_16": 0.70, "PATTERN_17": 0.70, "PATTERN_18": 0.70,
      "PATTERN_19": 0.80, "PATTERN_20": 0.80,
      "PATTERN_21": 0.90, "PATTERN_22": 0.90,
      "PATTERN_23": 0.85, "PATTERN_24": 1.00,
      "PATTERN_25": 0.90, "PATTERN_26": 0.90,
      "PATTERN_27": 1.00, "PATTERN_28": 1.00,
      "PATTERN_29": 0.80, "PATTERN_30": 0.85
    },
    "independence_factor": 0.90
  },
  "history": {"last_n": 35, "h2h_last_n": 5, "scored_allowed_sigma_floor": 8.0},
  "match_format": {
    "default_quarters": 4,
    "default_quarter_minutes": 10,
    "ten_minute_league_patterns": ["SUMMER LEAGUE", "WNBA", "EUROLEAGUE", "FIBA"],
    "twelve_minute_league_patterns": ["\\bNBA\\b", "NBA G LEAGUE"]
  },
  "q4": {
    "epsilon": 0.000001,
    "bonus_team_fouls": 4,
    "thresholds": {
      "pre_fouls_low": 26,
      "pre_fouls_high": 34,
      "pre_fouls_very_high": 38,
      "q3_fouls_high": 12,
      "pre_fta_high": 35,
      "pre_fta_low": 24,
      "q3_fta_high": 13,
      "total_after_3q_high": 137,
      "close_margin": 5,
      "chase_margin_low": 6,
      "chase_margin_high": 10,
      "blowout_margin": 21
    },
    "under_weights": {"hist": 0.18, "scenario": 0.18, "live": 0.24, "dry": 0.18, "no_foul_tail": 0.12, "no_kill_chase": 0.10},
    "over_weights": {"hist": 0.16, "scenario": 0.16, "live": 0.28, "foul_tail": 0.14, "kill_chase": 0.14, "volume": 0.12}
  },
  "aliases": {
    "FGA": ["FGA", "fga", "field_goal_attempts", "shots_attempted", "fieldGoalsAttempted"],
    "FGM": ["FGM", "fgm", "field_goals_made", "fieldGoalsMade"],
    "2PA": ["2PA", "two_point_attempts", "twoPointsAttempted"],
    "2PM": ["2PM", "two_point_made", "twoPointsMade"],
    "3PA": ["3PA", "three_point_attempts", "threePointsAttempted"],
    "3PM": ["3PM", "three_point_made", "threePointsMade"],
    "FTA": ["FTA", "fta", "free_throw_attempts", "freeThrowsAttempted"],
    "FTM": ["FTM", "ftm", "free_throws_made", "freeThrowsMade"],
    "ORB": ["ORB", "orb", "offensive_rebounds", "offensiveRebounds"],
    "DRB": ["DRB", "drb", "defensive_rebounds", "defensiveRebounds"],
    "TO": ["TO", "TOV", "to", "turnovers"],
    "FOULS": ["fouls", "FOULS", "PF", "personal_fouls"],
    "CLOCK": ["clock", "time", "time_remaining", "seconds_remaining", "quarter_clock"],
    "OVER_ODDS": ["over_odd", "overOdd", "over_odds", "overOdds"],
    "UNDER_ODDS": ["under_odd", "underOdd", "under_odds", "underOdds"],
    "LINE": ["line", "total", "value", "handicap"]
  }
}
""")


# ===== VPS ORCHESTRATION, AUDIT, LEARNING, GPT AND TELEGRAM =====
SYSTEM_VERSION = '9.1.0'

def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat(timespec='seconds')

# Кожен виклик process_vps_match_file() (тобто кожен чекпоінт кожного матчу)
# дописує один JSON-рядок сюди — незалежно від того, PLAY це, RISK чи PASS.
# Шлях налаштовується через VERDICT_LOG_FILE (в docker-compose природно
# покласти поруч з SUPER_BASKET_DB, напр. /app/state/verdicts.log), щоб файл
# лежав у volume 'state' і переживав рестарт контейнера.
VERDICT_LOG_FILE = os.getenv('VERDICT_LOG_FILE', 'verdicts.log')
EXCEL_AUDIT_FILE = os.getenv('EXCEL_AUDIT_FILE', 'super_basket_detailed_log.xlsx')
ENABLE_EXCEL_AUDIT = os.getenv('SUPER_BASKET_EXCEL_AUDIT', 'false').strip().lower() in {'1', 'true', 'yes', 'on'}
PUBLIC_FILES_BASE_URL = os.getenv('PUBLIC_FILES_BASE_URL', '').rstrip('/')

def _public_file_url(file_path: Any) -> Optional[str]:
    if not file_path or not PUBLIC_FILES_BASE_URL:
        return None
    try:
        relative = Path(str(file_path)).resolve().relative_to(Path('/app/state'))
    except (OSError, ValueError):
        return None
    return f"{PUBLIC_FILES_BASE_URL}/{urllib.parse.quote(relative.as_posix())}"

def append_verdict_log(entry: dict[str, Any], path: str | Path | None = None) -> None:
    """Дописує один JSON-рядок (JSON Lines) з підсумком чекпоінта.

    Ніколи не кидає виняток назовні — збій запису логу не повинен зривати сам
    аналіз/сигнал; помилка лише друкується в stderr.
    """
    target = Path(path or VERDICT_LOG_FILE).expanduser()
    try:
        target.parent.mkdir(parents=True, exist_ok=True)
        with target.open('a', encoding='utf-8') as fh:
            fh.write(json.dumps(entry, ensure_ascii=False) + '\n')
    except OSError as exc:
        print(f'WARNING: could not write verdict log to {target}: {exc}', file=sys.stderr)

def _excel_scalar_rows(value: Any, prefix: str = '') -> Iterable[tuple[str, Any]]:
    """Flatten every scalar in a calculation payload without dropping detail."""
    if isinstance(value, dict):
        for key, child in value.items():
            path = f'{prefix}.{key}' if prefix else str(key)
            yield from _excel_scalar_rows(child, path)
    elif isinstance(value, list):
        for index, child in enumerate(value):
            yield from _excel_scalar_rows(child, f'{prefix}[{index}]')
    else:
        if value is None or isinstance(value, (str, int, float, bool)):
            cell_value = value
        else:
            cell_value = str(value)
        if isinstance(cell_value, str) and len(cell_value) > 32767:
            cell_value = cell_value[:32740] + '... [TRUNCATED]'
        yield prefix, cell_value

def append_excel_audit(core_result: dict[str, Any], path: str | Path | None = None) -> None:
    """Append one summary row plus long-form calculation details to an XLSX.

    Excel is limited to 16,384 columns, so arbitrary JSON paths must be stored as
    rows (json_path/value), not as an ever-growing set of columns.

    A sidecar flock plus atomic replace makes concurrent Docker workers safe.
    Excel is diagnostic only: export failures never interrupt a betting job.
    """
    target = Path(path or EXCEL_AUDIT_FILE).expanduser()
    lock_path = target.with_suffix(target.suffix + '.lock')
    tmp_path: Optional[Path] = None
    try:
        import fcntl
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Font, PatternFill

        target.parent.mkdir(parents=True, exist_ok=True)
        with lock_path.open('a+b') as lock_handle:
            fcntl.flock(lock_handle.fileno(), fcntl.LOCK_EX)
            workbook = load_workbook(target) if target.exists() else Workbook()
            for stale_name in ('Market evaluations',):
                if stale_name in workbook.sheetnames:
                    workbook.remove(workbook[stale_name])
            if 'Runs' in workbook.sheetnames:
                ws = workbook['Runs']
            elif workbook.sheetnames == ['Sheet']:
                ws = workbook['Sheet']
                ws.title = 'Runs'
            else:
                ws = workbook.create_sheet('Runs')

            calculation = core_result.get('super_basket_calculation') or {}
            system = core_result.get('super_basket_system') or {}
            decision = system.get('decision') or {}
            snapshot = calculation.get('canonical_snapshot') or {}
            market = decision.get('market') or {}
            gpt_review = system.get('gpt_review') or {}
            source_file = (system.get('files') or {}).get('source')
            result_file = (system.get('files') or {}).get('result')
            run_id = f"{system.get('processed_at', utc_now())}|{snapshot.get('match_id', '')}|{snapshot.get('trigger_checkpoint', '')}"

            def _codes_summary(items: Any) -> str:
                if not items:
                    return ''
                return '; '.join(str(code) for code in items)

            def _rule_summary(items: Any) -> str:
                if not items:
                    return ''
                parts = []
                for entry in items:
                    if isinstance(entry, dict):
                        rule_id = entry.get('rule_id', '')
                        reason = entry.get('reason', '')
                        parts.append(f'{rule_id}: {reason}' if reason else str(rule_id))
                    else:
                        parts.append(str(entry))
                return ' | '.join(parts)

            row_data: dict[str, Any] = {
                'run_id': run_id,
                'timestamp_utc': system.get('processed_at'),
                'match_id': snapshot.get('match_id'),
                'match_name': snapshot.get('name'),
                'trigger_checkpoint': snapshot.get('trigger_checkpoint'),
                'stage': snapshot.get('stage'),
                'clock': snapshot.get('clock'),
                'score_home': (snapshot.get('score') or {}).get('home'),
                'score_away': (snapshot.get('score') or {}).get('away'),
                'action': decision.get('action'),
                'status': decision.get('status'),
                'stake': decision.get('stake'),
                'p_final': (decision.get('probabilities') or {}).get('p_final'),
                'market_type': market.get('market_type'),
                'segment': market.get('segment'),
                'team': market.get('team'),
                'side': market.get('side'),
                'line': market.get('line'),
                'odds': market.get('odds'),
                'signal_id': decision.get('signal_id'),
                'reason_codes': _codes_summary(decision.get('reason_codes')),
                'explanation_uk': decision.get('explanation_uk'),
                'main_risk_uk': decision.get('main_risk_uk'),
                'trigger_uk': decision.get('trigger_uk'),
                'blockers_summary': _rule_summary(decision.get('blockers')),
                'caps_summary': _rule_summary(decision.get('caps')),
                'gpt_status': gpt_review.get('status'),
                'gpt_explanation_uk': gpt_review.get('explanation_uk'),
                'gpt_main_risk_uk': gpt_review.get('main_risk_uk'),
                'telegram_status': (system.get('telegram_delivery') or {}).get('status'),
                'input_hash': calculation.get('input_snapshot_hash'),
                'source_file': source_file,
                'result_file': result_file,
                'source_url': _public_file_url(source_file),
                'result_url': _public_file_url(result_file),
            }
            audit_payload = {'super_basket_calculation': calculation, 'super_basket_system': system}

            # Migrate workbooks created by the old wide exporter. Keeping those
            # 30k+ columns in memory would make even a summary-only save fail.
            existing_headers = {
                ws.cell(1, column).value: column
                for column in range(1, min(ws.max_column, 16384) + 1)
            }
            if ws.max_column > 16384 or any(
                isinstance(header, str) and header.startswith('raw.')
                for header in existing_headers
            ):
                old_ws = ws
                old_index = workbook.index(old_ws)
                preserved_rows = [
                    [old_ws.cell(row, existing_headers[header]).value if header in existing_headers else None
                     for header in row_data]
                    for row in range(2, old_ws.max_row + 1)
                ]
                workbook.remove(old_ws)
                ws = workbook.create_sheet('Runs', old_index)
                ws.append(list(row_data))
                for preserved_row in preserved_rows:
                    ws.append(preserved_row)

            if ws.max_row == 1 and ws.cell(1, 1).value is None:
                header_index: dict[str, int] = {}
                for column, header in enumerate(row_data.keys(), 1):
                    ws.cell(1, column, header)
                    header_index[header] = column
            else:
                header_index = {ws.cell(1, column).value: column for column in range(1, ws.max_column + 1)}
                for header in row_data.keys():
                    if header not in header_index:
                        new_column = ws.max_column + 1
                        ws.cell(1, new_column, header)
                        header_index[header] = new_column

            for cell in ws[1]:
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill('solid', fgColor='1F4E78')
            ws.freeze_panes = 'B2'

            new_row = ws.max_row + 1
            for header, value in row_data.items():
                cell_value = value
                if isinstance(cell_value, str) and len(cell_value) > 32767:
                    cell_value = cell_value[:32740] + '... [TRUNCATED]'
                ws.cell(new_row, header_index[header], cell_value)
            for name in ('source_url', 'result_url'):
                column = header_index.get(name)
                if column:
                    cell = ws.cell(new_row, column)
                    if cell.value:
                        cell.hyperlink = cell.value
                        cell.style = 'Hyperlink'

            detail_headers = ['run_id', 'json_path', 'value', 'value_type']
            detail_rows = list(_excel_scalar_rows(audit_payload))
            detail_sheet_number = 1
            while True:
                detail_name = 'All calculations' if detail_sheet_number == 1 else f'All calculations {detail_sheet_number}'
                if detail_name in workbook.sheetnames:
                    details = workbook[detail_name]
                else:
                    details = workbook.create_sheet(detail_name)
                    details.append(detail_headers)
                if details.max_row + len(detail_rows) <= 1048576:
                    break
                detail_sheet_number += 1
            for json_path, value in detail_rows:
                details.append([run_id, json_path, value, type(value).__name__])
            for cell in details[1]:
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill('solid', fgColor='1F4E78')
            details.freeze_panes = 'A2'
            details.auto_filter.ref = details.dimensions
            details.column_dimensions['A'].width = 42
            details.column_dimensions['B'].width = 72
            details.column_dimensions['C'].width = 60

            ws.auto_filter.ref = ws.dimensions
            ws.column_dimensions['A'].width = 42
            ws.column_dimensions['B'].width = 24
            ws.column_dimensions[ws.cell(1, header_index['explanation_uk']).column_letter].width = 60
            tmp_path = target.with_name(f'.{target.name}.{os.getpid()}.tmp')
            workbook.save(tmp_path)
            os.replace(tmp_path, target)
            fcntl.flock(lock_handle.fileno(), fcntl.LOCK_UN)
    except Exception as exc:
        if tmp_path is not None:
            try:
                tmp_path.unlink(missing_ok=True)
            except OSError:
                pass
        print(f'WARNING: could not write Excel audit to {target}: {exc}', file=sys.stderr)

def env_bool(name: str, default: bool) -> bool:
    value = os.getenv(name)
    if value is None:
        return default
    return value.strip().lower() in {'1', 'true', 'yes', 'on'}

def normalized_action(
    probability: float,
    blockers: list[dict[str, Any]],
    mode: str,
    p_hist: Optional[float]=None,
) -> tuple[str, str, str]:
    if blockers:
        return 'PASS', 'PASS', '0%'
    if mode.upper() == 'STRICT' and probability < 0.75:
        return 'PASS', 'TRIGGER ONLY', '0%'
    if probability < 0.60:
        return 'PASS', 'PASS', '0%'
    if probability < 0.75:
        return 'RISK', 'RISK ENTRY', '10-15% live-limit'
    if probability < 0.80:
        return 'PLAY', 'LOW PLAY', '15-20% live-limit'
    if probability < 0.85:
        return 'PLAY', 'MAIN PLAY', '30-35% live-limit'
    return 'PLAY', 'STRONG PLAY', '40-50% live-limit'


def build_budget_recommendation(action: str, status: str, stake: str) -> dict[str, Any]:
    match = re.search(r'(\d+(?:\.\d+)?)\s*[-–]\s*(\d+(?:\.\d+)?)%', stake or '')
    pct_min = float(match.group(1)) if match else 0.0
    pct_max = float(match.group(2)) if match else 0.0
    live_limit = to_number(os.getenv('SUPER_BASKET_LIVE_LIMIT'))
    bankroll = to_number(os.getenv('SUPER_BASKET_BANKROLL'))
    base_amount = live_limit if live_limit is not None else bankroll
    base_type = 'LIVE_LIMIT' if live_limit is not None else 'BANKROLL_FALLBACK' if bankroll is not None else 'PERCENT_ONLY'
    currency = os.getenv('SUPER_BASKET_CURRENCY', 'USDT')
    return {
        'action': action,
        'status': status,
        'base_type': base_type,
        'base_amount': base_amount,
        'currency': currency,
        'percent_min': pct_min,
        'percent_max': pct_max,
        'amount_min': round(base_amount * pct_min / 100.0, 2) if base_amount is not None else None,
        'amount_max': round(base_amount * pct_max / 100.0, 2) if base_amount is not None else None,
        'text': stake,
        'educational_note': 'Budget range is a configurable live-limit allocation, not a guarantee of outcome.',
    }

RISK_POST_FILTER_P_LIVE_THRESHOLD = 0.85
RISK_POST_FILTER_P_FINAL_UPPER_BOUND = 0.65


def apply_risk_post_filter(decision: dict[str, Any]) -> dict[str, Any]:
    """v6: no second hidden threshold. A clean deterministic RISK starts at 60%."""
    return {
        'enabled': False,
        'applicable': False,
        'p_final': to_number((decision.get('probabilities') or {}).get('p_final')),
        'p_live': to_number((decision.get('probabilities') or {}).get('p_live')),
        'passed': True,
        'filtered': False,
        'reason_code': None,
        'policy': 'RISK_60_TO_74_99_AFTER_ALL_HARD_GATES',
    }

def _precomputed_line_reconciliation(source: dict[str, Any], evaluation: Optional[dict[str, Any]], data_gate: dict[str, Any]) -> dict[str, Any]:
    if not evaluation:
        return {'status': 'NO_MARKET_SELECTED'}
    if data_gate.get('cross_format_team_a_n') or data_gate.get('cross_format_team_b_n'):
        return {
            'status': 'NOT_COMPARABLE_PRECOMPUTED_MIXES_FORMATS',
            'reason': 'P_hist was recomputed from same-duration raw history; parser table may contain 4x10 and 4x12 together',
        }
    root = source.get('history_by_exact_line')
    if not isinstance(root, dict):
        return {'status': 'BLOCK_NOT_PROVIDED'}
    market_type = evaluation.get('market_type')
    line = float(evaluation.get('line'))
    side_key = 'over_rate' if evaluation.get('side') == 'OVER' else 'under_rate'
    comparisons: dict[str, Any] = {}

    def find_array(rows: Any) -> Optional[dict[str, Any]]:
        if not isinstance(rows, list):
            return None
        return next((item for item in rows if to_number(item.get('line')) is not None and abs(float(item['line']) - line) < 1e-9), None)

    if market_type == 'MATCH_TOTAL':
        block = root.get('match_total', {})
        for parser_key, history_key in (('team_a', 'team_a'), ('team_b', 'team_b'), ('pooled70', 'pooled'), ('h2h', 'h2h')):
            row = find_array(block.get(parser_key) if isinstance(block, dict) else None)
            derived = evaluation.get('history', {}).get(history_key, {}).get('raw_pct')
            if row and derived is not None:
                comparisons[history_key] = {'parser': to_number(row.get(side_key)), 'recomputed': derived}
    elif market_type in {'H1_TOTAL', 'H2_TOTAL'}:
        block = root.get('half_total', {})
        row = block.get(str(line)) or block.get(f'{line:.1f}') if isinstance(block, dict) else None
        for parser_key, history_key in (('team_a', 'team_a'), ('team_b', 'team_b'), ('pooled70', 'pooled'), ('h2h', 'h2h')):
            parsed = row.get(parser_key) if isinstance(row, dict) else None
            derived = evaluation.get('history', {}).get(history_key, {}).get('raw_pct')
            if isinstance(parsed, dict) and derived is not None:
                comparisons[history_key] = {'parser': to_number(parsed.get(side_key)), 'recomputed': derived}
    elif market_type == 'CURRENT_QUARTER_TOTAL':
        block = root.get('quarter_total', {})
        for parser_key, history_key in (('team_a', 'team_a'), ('team_b', 'team_b'), ('pooled70', 'pooled')):
            row = find_array(block.get(parser_key) if isinstance(block, dict) else None)
            derived = evaluation.get('history', {}).get(history_key, {}).get('raw_pct')
            if row and derived is not None:
                comparisons[history_key] = {'parser': to_number(row.get(side_key)), 'recomputed': derived}
    elif market_type and ('TEAM_IT' in market_type):
        block = root.get('team_it', {})
        row = block.get(str(line)) or block.get(f'{line:.1f}') if isinstance(block, dict) else None
        team_key = 'team_a' if evaluation.get('team') == (source.get('match', {}) or {}).get('home_team') else 'team_b'
        parsed = row.get(f'{team_key}_{str(evaluation.get("side") or "").lower()}') if isinstance(row, dict) else None
        if isinstance(parsed, dict):
            comparisons['own_scored'] = {'parser': to_number(parsed.get('own_scored_rate')), 'recomputed': evaluation.get('history', {}).get('own_scored', {}).get('raw_pct')}
            comparisons['opponent_allowed'] = {'parser': to_number(parsed.get('opponent_allowed_rate')), 'recomputed': evaluation.get('history', {}).get('opponent_allowed', {}).get('raw_pct')}
    if not comparisons:
        return {'status': 'NO_EXACT_PRECOMPUTED_LINE', 'line': line}
    differences = [abs(float(item['parser']) - float(item['recomputed'])) for item in comparisons.values() if item.get('parser') is not None and item.get('recomputed') is not None]
    max_difference = max(differences, default=0.0)
    return {
        'status': 'MATCH' if max_difference <= 0.02 else 'DATA_CONFLICT',
        'tolerance': 0.02,
        'max_absolute_difference': max_difference,
        'comparisons': comparisons,
    }

def build_input_usage(source: dict[str, Any], calculation: dict[str, Any], evaluation: Optional[dict[str, Any]]) -> dict[str, Any]:
    data_gate = calculation['data_gate']
    blocks = {
        'match/raw_data.main_match': 'PRIMARY: score, time, teams and duration',
        'bookmaker_lines': 'PRIMARY: only real supported bookmaker lines',
        'raw_data.team_a_hist': 'PRIMARY: recomputed exact-line same-format history',
        'raw_data.team_b_hist': 'PRIMARY: recomputed exact-line same-format history',
        'raw_data.h2h_hist': 'PRIMARY: recomputed H2H with shrinkage',
        'history_by_exact_line': 'VALIDATION/FALLBACK: checked against raw history, never double-counted',
        'scenario_patterns_by_line': 'VALIDATION: raw historical pattern matches are recomputed',
        'checkpoint_matrices': 'VALIDATION: scenario condition support',
        'quarter_result_profile': 'VALIDATION: quarter scenario support',
        'stat_conditioned_line_profiles': 'USED: timestamp validation and live-calibrated projection component',
        'projections': 'USED: conservative projection components; stale/extreme values rejected',
        'stat_alignment/stat_zones': 'VALIDATION: stat-gate is recomputed from live stats and team-relative zones',
        'history_zones': 'IGNORED: old global zones are forbidden by the supplied rules',
        'line_evaluations/markets_evaluation/final_verdict': 'AUDIT ONLY: never fed back into P to avoid circular probability',
    }
    availability = {
        key: any(part in source for part in key.split('/'))
        for key in (
            'history_by_exact_line', 'scenario_patterns_by_line', 'checkpoint_matrices',
            'quarter_result_profile', 'stat_conditioned_line_profiles', 'projections',
            'stat_alignment', 'stat_zones', 'history_zones', 'line_evaluations',
            'markets_evaluation', 'final_verdict',
        )
    }
    reconciliation = _precomputed_line_reconciliation(source, evaluation, data_gate)
    return {
        'rules': blocks,
        'availability': availability,
        'exact_line_reconciliation': reconciliation,
        'data_conflict': reconciliation.get('status') == 'DATA_CONFLICT',
        'cross_format_policy': {
            'exact_line_hits': 'same regulation duration only',
            'different_duration_games': 'normalized pace/PPP baseline only, max 25% influence',
            'current_match_excluded': True,
        },
    }

class LearningStore:
    def __init__(self, path: str | Path) -> None:
        self.path = Path(path).expanduser().resolve()
        self.path.parent.mkdir(parents=True, exist_ok=True)
        self.connection = sqlite3.connect(self.path)
        self.connection.row_factory = sqlite3.Row
        self.connection.execute('PRAGMA journal_mode=WAL')
        self.connection.execute('PRAGMA foreign_keys=ON')
        self._create_schema()

    def close(self) -> None:
        self.connection.close()

    def _create_schema(self) -> None:
        self.connection.executescript('''
        CREATE TABLE IF NOT EXISTS processed_snapshots (
            input_hash TEXT PRIMARY KEY,
            source_path TEXT,
            output_path TEXT,
            status TEXT NOT NULL,
            processed_at TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS signals (
            signal_id TEXT PRIMARY KEY,
            input_hash TEXT NOT NULL,
            match_id TEXT NOT NULL,
            match_name TEXT,
            stage TEXT NOT NULL,
            format_key TEXT NOT NULL,
            market_type TEXT NOT NULL,
            team TEXT,
            segment TEXT NOT NULL,
            side TEXT NOT NULL,
            line REAL NOT NULL,
            odds REAL NOT NULL,
            bookmaker TEXT,
            p_hist REAL,
            p_scenario REAL,
            p_live REAL,
            p_raw REAL,
            p_rule REAL NOT NULL,
            p_calibrated REAL NOT NULL,
            p_final REAL NOT NULL,
            deterministic_action TEXT NOT NULL,
            final_action TEXT NOT NULL,
            gpt_status TEXT,
            telegram_status TEXT,
            telegram_message_id TEXT,
            result TEXT,
            outcome_value REAL,
            profit_units REAL,
            created_at TEXT NOT NULL,
            settled_at TEXT
        );
        CREATE UNIQUE INDEX IF NOT EXISTS idx_signal_snapshot_market
        ON signals(input_hash, market_type, IFNULL(team, ''), segment, side, line);
        CREATE INDEX IF NOT EXISTS idx_signal_calibration
        ON signals(format_key, market_type, stage, side, result);
        ''')
        self.connection.commit()

    def calibration(self, evaluation: dict[str, Any], stage: str, format_key: str) -> dict[str, Any]:
        p_rule = float(evaluation['p_final'])
        scopes = [
            ('market_stage_side', 'market_type=? AND stage=? AND side=?', (evaluation['market_type'], stage, evaluation['side'])),
            ('market_side', 'market_type=? AND side=?', (evaluation['market_type'], evaluation['side'])),
            ('market', 'market_type=?', (evaluation['market_type'],)),
        ]
        selected: Optional[dict[str, Any]] = None
        for scope, clause, values in scopes:
            row = self.connection.execute(
                f'''SELECT COUNT(*) AS n,
                           SUM(CASE WHEN result='WIN' THEN 1 ELSE 0 END) AS wins,
                           AVG((p_rule - CASE WHEN result='WIN' THEN 1.0 ELSE 0.0 END) *
                               (p_rule - CASE WHEN result='WIN' THEN 1.0 ELSE 0.0 END)) AS brier
                    FROM signals
                    WHERE format_key=? AND result IN ('WIN','LOSS') AND {clause}''',
                (format_key, *values),
            ).fetchone()
            stats = {'scope': scope, 'samples': int(row['n'] or 0), 'wins': int(row['wins'] or 0), 'brier_score': row['brier']}
            if selected is None:
                selected = stats
            if stats['samples'] >= 50:
                selected = stats
                break
        assert selected is not None
        if selected['samples'] < 50:
            return {**selected, 'status': 'WAITING_FOR_50_SETTLED', 'weight': 0.0, 'posterior_hit_rate': None, 'p_rule': p_rule, 'p_calibrated': p_rule, 'delta': 0.0}
        alpha = beta = 2.0
        posterior = (selected['wins'] + alpha) / (selected['samples'] + alpha + beta)
        weight = min(0.20, 0.05 + 0.15 * min(1.0, (selected['samples'] - 50) / 200.0))
        raw_calibrated = (1 - weight) * p_rule + weight * posterior
        delta = max(-0.05, min(0.05, raw_calibrated - p_rule))
        active_cap = min((float(item['cap']) for item in evaluation.get('caps', [])), default=1.0)
        calibrated = max(0.01, min(0.99, p_rule + delta, active_cap))
        return {**selected, 'status': 'ACTIVE', 'weight': weight, 'posterior_hit_rate': posterior, 'p_rule': p_rule, 'p_calibrated': calibrated, 'delta': calibrated - p_rule}

    def get_signal(self, signal_id: str) -> Optional[dict[str, Any]]:
        row = self.connection.execute('SELECT * FROM signals WHERE signal_id=?', (signal_id,)).fetchone()
        return dict(row) if row else None

    def record_signal(self, decision: dict[str, Any], calculation: dict[str, Any]) -> tuple[dict[str, Any], bool]:
        market = decision['market']
        probabilities = decision['probabilities']
        snapshot = calculation['canonical_snapshot']
        values = (
            decision['signal_id'], calculation['input_snapshot_hash'], snapshot['match_id'], snapshot['name'], snapshot['stage'],
            snapshot.get('format', {}).get('format_key') or 'UNKNOWN', market['market_type'], market.get('team'), market['segment'], market['side'],
            market['line'], market['odds'], market.get('bookmaker'), probabilities['p_hist'], probabilities['p_scenario'], probabilities['p_live'],
            probabilities['p_raw'], probabilities['p_rule'], probabilities['p_calibrated'], probabilities['p_final'], decision['deterministic_action'],
            decision['action'], decision.get('gpt_status'), decision.get('telegram_status'), utc_now(),
        )
        existing = self.get_signal(decision['signal_id'])
        if existing:
            return existing, True
        self.connection.execute('''INSERT INTO signals (
            signal_id,input_hash,match_id,match_name,stage,format_key,market_type,team,segment,side,line,odds,bookmaker,
            p_hist,p_scenario,p_live,p_raw,p_rule,p_calibrated,p_final,deterministic_action,final_action,gpt_status,telegram_status,created_at
        ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''', values)
        self.connection.commit()
        return self.get_signal(decision['signal_id']) or {}, False

    def update_delivery(self, signal_id: str, final_action: str, gpt_status: str, telegram_status: str, message_id: Optional[str]) -> None:
        self.connection.execute('''UPDATE signals SET final_action=?, gpt_status=?, telegram_status=?, telegram_message_id=? WHERE signal_id=?''', (final_action, gpt_status, telegram_status, message_id, signal_id))
        self.connection.commit()

    def mark_processed(self, input_hash: str, source_path: str, output_path: str, status: str) -> None:
        self.connection.execute('''INSERT INTO processed_snapshots(input_hash,source_path,output_path,status,processed_at)
            VALUES(?,?,?,?,?) ON CONFLICT(input_hash) DO UPDATE SET output_path=excluded.output_path,status=excluded.status,processed_at=excluded.processed_at''',
            (input_hash, source_path, output_path, status, utc_now()))
        self.connection.commit()

    def settle(self, signal_id: str, result: str, outcome_value: Optional[float]=None) -> dict[str, Any]:
        normalized = result.strip().upper()
        if normalized not in {'WIN', 'LOSS', 'PUSH'}:
            raise ValueError('result must be win, loss or push')
        signal = self.get_signal(signal_id)
        if not signal:
            raise ValueError(f'Unknown signal_id: {signal_id}')
        profit = float(signal['odds']) - 1.0 if normalized == 'WIN' else -1.0 if normalized == 'LOSS' else 0.0
        self.connection.execute('''UPDATE signals SET result=?, outcome_value=?, profit_units=?, settled_at=? WHERE signal_id=?''', (normalized, outcome_value, profit, utc_now(), signal_id))
        self.connection.commit()
        return self.get_signal(signal_id) or {}

    def report(self) -> dict[str, Any]:
        row = self.connection.execute('''SELECT COUNT(*) AS signals,
            SUM(CASE WHEN result='WIN' THEN 1 ELSE 0 END) AS wins,
            SUM(CASE WHEN result='LOSS' THEN 1 ELSE 0 END) AS losses,
            SUM(CASE WHEN result='PUSH' THEN 1 ELSE 0 END) AS pushes,
            SUM(CASE WHEN result IS NULL THEN 1 ELSE 0 END) AS unsettled,
            SUM(COALESCE(profit_units,0)) AS profit_units,
            AVG(CASE WHEN result IN ('WIN','LOSS') THEN
                (p_rule - CASE WHEN result='WIN' THEN 1.0 ELSE 0.0 END) *
                (p_rule - CASE WHEN result='WIN' THEN 1.0 ELSE 0.0 END) END) AS brier_score
            FROM signals''').fetchone()
        return {key: row[key] for key in row.keys()}

def apply_learning_to_evaluation(evaluation: dict[str, Any], store: LearningStore, calculation: dict[str, Any], mode: str) -> dict[str, Any]:
    item = deepcopy(evaluation)
    calibration = store.calibration(item, calculation['canonical_snapshot']['stage'], calculation['canonical_snapshot'].get('format', {}).get('format_key') or 'UNKNOWN')
    p_rule = float(item['p_final'])
    p_calibrated = float(calibration['p_calibrated'])
    p_hist = item.get('history', {}).get('p_hist')
    action, status, stake = normalized_action(
        p_calibrated,
        item.get('blockers', []),
        mode,
        p_hist,
    )
    if action != 'PASS' and item.get('stat_comparison', {}).get('stat_support') == 'OFF':
        action, status, stake = 'RISK', 'RISK PLAY — NO-STATS FALLBACK', '10-15% live-limit'
    item['p_rule'] = p_rule
    item['p_calibrated'] = p_calibrated
    item['p_final_system'] = p_calibrated
    item['system_action'] = action
    item['system_status'] = status
    item['stake'] = stake
    item['system_reason_codes'] = []
    item['calibration'] = calibration
    item.setdefault('p_trace', []).append(_trace_step(
        'CALIBRATION', calibration['status'] == 'ACTIVE',
        'Beta-Binomial empirical calibration after >=50 settled predictions; max weight 20%, max delta 5pp, caps reapplied',
        calibration, p_rule, p_calibrated, [calibration['status']],
    ))
    item['p_trace'].append(_trace_step(
        'P_FINAL', True, 'mode threshold + hard blockers after rule probability and calibration',
        {'mode': mode.upper(), 'action': action, 'status': status}, p_calibrated, p_calibrated,
        [action, status],
    ))
    return item

def summarize_line_evaluation(item: dict[str, Any]) -> dict[str, Any]:
    """Compact, bot-safe result for every exact bookmaker line and both directions."""
    return {
        'market_id': item.get('market_id'),
        'math_market_key': item.get('math_market_key'),
        'market_type': item.get('market_type'),
        'team': item.get('team'),
        'segment': item.get('segment'),
        'side': item.get('side'),
        'line': item.get('line'),
        'odds': item.get('odds'),
        'bookmaker': item.get('bookmaker'),
        'action': item.get('system_action', 'PASS'),
        'status': item.get('system_status', 'PASS'),
        'stake': item.get('stake', '0%'),
        'p_hist': item.get('history', {}).get('p_hist'),
        'history_zone_rate': item.get('history', {}).get('history_zone_rate'),
        'history_zone_source': item.get('history', {}).get('history_zone_source'),
        'p_scenario': item.get('scenario', {}).get('p_scenario'),
        'p_live': item.get('live', {}).get('p_live'),
        'p_raw': item.get('p_raw'),
        'p_final': item.get('p_final_system', item.get('p_final')),
        'projection_used': item.get('live', {}).get('projection_used'),
        'line_edge': item.get('live', {}).get('line_edge'),
        'live_reversal_active': item.get('live_reversal', {}).get('active', False),
        'stat_gate_status': item.get('stat_comparison', {}).get('stat_gate_status'),
        'stat_values_projected_to_scope_end': item.get('stat_comparison', {}).get('values_projected_to_scope_end', False),
        'router_status': item.get('router', {}).get('status'),
        'router_reason': item.get('router', {}).get('reason'),
        'parser_issues': list(item.get('parser_issues', [])),
        'caps': [entry.get('rule_id') for entry in item.get('caps', [])],
        'blockers': [entry.get('rule_id') for entry in item.get('blockers', [])],
        'offers': deepcopy(item.get('offers', [])),
    }


def _candidate_pool(evaluations: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        item for item in evaluations
        if item.get('odds') is not None
        and float(item['odds']) >= float(DEFAULT_CONFIG['odds_min'])
        and not any(issue not in {'ODDS_BELOW_MINIMUM'} for issue in item.get('parser_issues', []))
    ]

def select_one_decision(evaluations: list[dict[str, Any]], mode: str) -> tuple[Optional[dict[str, Any]], Optional[dict[str, Any]]]:
    active = _candidate_pool(evaluations)
    eligible = [item for item in active if item['system_action'] != 'PASS']
    eligible.sort(key=lambda item: (float(item['p_final_system']), float(item.get('odds') or 0.0)), reverse=True)
    # PASS fallback must describe the strongest rejected candidate. The old
    # key put ``-len(blockers)`` first and therefore preferred a weak,
    # blocker-free side (for example OVER 30.81%) over the genuinely stronger
    # blocked opposite side (UNDER 69.19%). Probability is authoritative;
    # fewer blockers and better odds are only tie-breakers.
    active.sort(
        key=lambda item: (
            float(item['p_final_system']),
            -len(item.get('blockers', [])),
            float(item.get('odds') or 0.0),
        ),
        reverse=True,
    )
    closest = active[0] if active else (evaluations[0] if evaluations else None)
    return (eligible[0] if eligible else None, closest)

def _reason_codes(evaluation: Optional[dict[str, Any]], *, pass_fallback: bool=False) -> list[str]:
    if not evaluation:
        return ['NO_SUPPORTED_REAL_LINES']
    codes = [item['rule_id'] for item in evaluation.get('blockers', [])]
    codes.extend(item['rule_id'] for item in evaluation.get('caps', []))
    codes.extend(evaluation.get('system_reason_codes', []))
    if not codes:
        codes.append('P_FINAL_BELOW_ACTION_THRESHOLD' if pass_fallback else 'NO_HARD_CONFLICT')
    return list(dict.fromkeys(codes))

def deterministic_explanation(evaluation: Optional[dict[str, Any]], action: str, mode: str) -> tuple[str, str, str]:
    if not evaluation:
        return (
            'У JSON немає підтримуваної актуальної лінії букмекера з коефіцієнтом не нижче 1.44.',
            'Без реальної лінії система не створює ставку.',
            'Додати актуальну Match/H1/H2/Quarter Total або Team IT лінію з часом і рахунком.',
        )
    probability = float(evaluation.get('p_final_system', evaluation.get('p_final', 0.0)))
    live = evaluation.get('live', {})
    stat = evaluation.get('stat_comparison', {})
    history = evaluation.get('history', {})
    edge = float(live.get('line_edge') or 0.0)
    direction = 'вище' if evaluation.get('side') == 'OVER' else 'нижче'
    if action == 'PASS':
        codes = _reason_codes(evaluation, pass_fallback=True)
        side_label = str(evaluation.get('side') or '').title()
        line_value = evaluation.get('line')
        odds_value = evaluation.get('odds')
        line_label = f'{float(line_value):g}' if line_value is not None else '?'
        odds_label = f' @{float(odds_value):g}' if odds_value is not None else ''
        explanation = (
            f'Найсильніший відхилений варіант: {side_label} {line_label}{odds_label}. '
            f'P_final: {probability:.2%}. Причина: {" + ".join(codes)}.'
        )
        risk = 'Наявні дані або гейти не дозволяють безпечно перетворити цей варіант на активну ставку.'
        trigger = (
            'Потрібні P_final не нижче 75% та відсутність hard blocker.'
            if mode.upper() == 'STRICT'
            else (
                'Потрібні P_final ≥60%, історична зона ≥75%, live edge ≥3 '
                'та відсутність hard blocker.'
            )
        )
        return explanation, risk, trigger
    explanation = (
        f'P_final {probability:.1%}: проєкція {live.get("projection_used"):.1f} очка, '
        f'{direction} лінії {float(evaluation["line"]):.1f} на {abs(edge):.1f}; '
        f'P_hist {float(history.get("p_hist") or 0):.1%}, '
        f'історична зона {float(history.get("history_zone_rate") or 0):.1%}, '
        f'stat-gate {stat.get("stat_gate_status")}.'
    )
    if action == 'RISK':
        risk = 'P_final перебуває в зоні 60–74.99%, тому це RISK, а не clean PLAY; усі hard-gates уже пройдені.'
    else:
        risk = 'Лінія та коефіцієнт можуть змінитися; сигнал чинний лише для вказаного snapshot.'
    trigger = 'Брати тільки якщо та сама лінія ще доступна, odds >=1.44 і рахунок/час не змінилися суттєво.'
    return explanation, risk, trigger

def build_decision(selected: Optional[dict[str, Any]], closest: Optional[dict[str, Any]], calculation: dict[str, Any], mode: str) -> dict[str, Any]:
    evaluation = selected or closest
    action = selected['system_action'] if selected else 'PASS'
    status = selected['system_status'] if selected else 'PASS'
    stake = selected['stake'] if selected else '0%'
    explanation, main_risk, trigger = deterministic_explanation(evaluation, action, mode)
    if evaluation:
        market = {
            'market_type': evaluation.get('market_type'),
            'team': evaluation.get('team'),
            'segment': evaluation.get('segment'),
            'side': evaluation.get('side'),
            'line': evaluation.get('line'),
            'odds': evaluation.get('odds'),
            'bookmaker': evaluation.get('bookmaker'),
            'offers': evaluation.get('offers', []),
        }
        probabilities = {
            'p_hist': evaluation.get('history', {}).get('p_hist'),
            'p_scenario': evaluation.get('scenario', {}).get('p_scenario'),
            'p_live': evaluation.get('live', {}).get('p_live'),
            'p_raw': evaluation.get('p_raw'),
            'p_rule': evaluation.get('p_rule', evaluation.get('p_final')),
            'p_calibrated': evaluation.get('p_calibrated', evaluation.get('p_final')),
            'p_final': evaluation.get('p_final_system', evaluation.get('p_final')),
        }
    else:
        market = None
        probabilities = {'p_hist': None, 'p_scenario': None, 'p_live': None, 'p_raw': None, 'p_rule': None, 'p_calibrated': None, 'p_final': None}
    signal_id = None
    if selected and market:
        key = '|'.join(str(value) for value in (
            calculation['input_snapshot_hash'], market['market_type'], market.get('team'), market['segment'], market['side'], market['line'],
        ))
        signal_id = 'SB-' + hashlib.sha256(key.encode('utf-8')).hexdigest()[:16].upper()
    return {
        'action': action,
        'deterministic_action': action,
        'status': status,
        'signal_id': signal_id,
        'market': market,
        'probabilities': probabilities,
        'stake': stake,
        'budget_recommendation': build_budget_recommendation(action, status, stake),
        'explanation_uk': explanation,
        'main_risk_uk': main_risk,
        'trigger_uk': trigger,
        'reason_codes': _reason_codes(evaluation, pass_fallback=(action == 'PASS')),
        'caps': evaluation.get('caps', []) if evaluation else [],
        'blockers': evaluation.get('blockers', []) if evaluation else [],
        'p_trace': evaluation.get('p_trace', []) if evaluation else [],
        '_evaluation': evaluation,
    }

def gpt_review_decision(decision: dict[str, Any], calculation: dict[str, Any], *, api_key: Optional[str]=None, model: Optional[str]=None) -> dict[str, Any]:
    api_key = api_key or os.getenv('OPENAI_API_KEY')
    if not api_key:
        return {'status': 'SKIPPED_NO_API_KEY', 'approved': False, 'action': 'PASS', 'explanation_uk': '', 'main_risk_uk': '', 'telegram_text_uk': ''}
    try:
        from openai import OpenAI
        from pydantic import BaseModel, Field
    except ImportError as exc:
        return {'status': 'ERROR_OPENAI_PACKAGE_MISSING', 'approved': False, 'action': 'PASS', 'error': str(exc), 'explanation_uk': '', 'main_risk_uk': '', 'telegram_text_uk': ''}

    class GPTDecisionReview(BaseModel):
        approved: bool
        action: Literal['PLAY', 'RISK', 'PASS']
        explanation_uk: str = Field(min_length=1, max_length=500)
        main_risk_uk: str = Field(min_length=1, max_length=350)
        telegram_text_uk: str = Field(min_length=1, max_length=1000)

    compact = {
        'match': calculation['canonical_snapshot'],
        'decision': {key: value for key, value in decision.items() if key not in {'_evaluation', 'p_trace'}},
        'p_trace': decision.get('p_trace', []),
    }
    instructions = (
        'Ти контролер готового детермінованого баскетбольного сигналу. '
        'Не перераховуй P, не змінюй market/team/segment/side/line/odds і не вигадуй нову ставку. '
        'Ти можеш підтвердити дію, понизити PLAY до RISK або понизити будь-яку дію до PASS. '
        'Ніколи не підвищуй RISK до PLAY і PASS до активної дії. '
        'Якщо p_trace суперечливий, є hard blocker або дані stale — поверни PASS. Відповідай українською.'
    )
    try:
        client = OpenAI(api_key=api_key, timeout=30.0, max_retries=2)
        response = client.responses.parse(
            model=model or os.getenv('OPENAI_MODEL', 'gpt-5.6'),
            input=[
                {'role': 'system', 'content': instructions},
                {'role': 'user', 'content': json.dumps(compact, ensure_ascii=False, separators=(',', ':'))},
            ],
            text_format=GPTDecisionReview,
            store=False,
        )
        parsed = response.output_parsed
        if parsed is None:
            return {'status': 'ERROR_EMPTY_STRUCTURED_OUTPUT', 'approved': False, 'action': 'PASS', 'explanation_uk': '', 'main_risk_uk': '', 'telegram_text_uk': ''}
        review = parsed.model_dump()
        ranks = {'PASS': 0, 'RISK': 1, 'PLAY': 2}
        deterministic = decision['deterministic_action']
        if ranks[review['action']] > ranks[deterministic]:
            review['action'] = deterministic
            review['approved'] = False
            review['status'] = 'DOWNGRADE_ENFORCEMENT_MODEL_ATTEMPTED_UPGRADE'
        else:
            review['status'] = 'APPROVED' if review['approved'] and review['action'] != 'PASS' else 'DOWNGRADED_TO_PASS' if review['action'] == 'PASS' else 'NOT_APPROVED'
        return review
    except Exception as exc:  # network/API failures fail closed by design
        return {'status': 'ERROR_GPT_REVIEW_FAILED', 'approved': False, 'action': 'PASS', 'error': f'{type(exc).__name__}: {exc}', 'explanation_uk': '', 'main_risk_uk': '', 'telegram_text_uk': ''}

def build_telegram_message(decision: dict[str, Any], calculation: dict[str, Any], review: dict[str, Any]) -> str:
    market = decision['market'] or {}
    probability = decision['probabilities'].get('p_final')
    icon = '🚨' if decision['action'] == 'PLAY' else '⚠️'
    explanation = review.get('explanation_uk') or decision['explanation_uk']
    risk = review.get('main_risk_uk') or decision['main_risk_uk']
    name = calculation['canonical_snapshot']['name']
    team = market.get('team')
    market_line = f'<b>Ринок:</b> {html.escape(str(market.get("market_type")))} / {html.escape(str(market.get("segment")))}'
    if team:
        market_line += f' ({html.escape(str(team))})'
    lines = [
        f'<b>{icon} {html.escape(decision["action"])}</b>',
        f'<b>Матч:</b> {html.escape(str(name))}',
        market_line,
        f'<b>Сторона:</b> {html.escape(str(market.get("side")))}',
        f'<b>Лінія:</b> {float(market.get("line")):.1f}',
        f'<b>Коефіцієнт:</b> {float(market.get("odds")):.2f} ({html.escape(str(market.get("bookmaker") or ""))})',
        f'<b>P_final:</b> {float(probability):.1%}',
        f'<b>Статус:</b> {html.escape(decision["status"])}',
        f'<b>Stake:</b> {html.escape(decision["stake"])}',
        (
            f'<b>Бюджет:</b> {decision["budget_recommendation"]["amount_min"]:.2f}–'
            f'{decision["budget_recommendation"]["amount_max"]:.2f} '
            f'{html.escape(str(decision["budget_recommendation"]["currency"]))}'
            if decision.get('budget_recommendation', {}).get('amount_min') is not None
            else f'<b>Бюджет:</b> {html.escape(decision["stake"])}'
        ),
        f'<b>Пояснення:</b> {html.escape(explanation)}',
        f'<b>Головний ризик:</b> {html.escape(risk)}',
        '<i>Сигнал чинний лише для вказаних лінії, коефіцієнта, рахунку та часу.</i>',
        f'<code>{html.escape(str(decision.get("signal_id") or ""))}</code>',
    ]
    return '\n'.join(lines)

def _load_telegram_chat_ids(chats_file: Optional[str] = None) -> list[str]:
    """Read the {"offset":..., "chatIds":[...]} file the bot maintains and
    return the chat ids as strings, in order, de-duplicated."""
    path_value = chats_file or os.getenv('TELEGRAM_CHATS_FILE')
    if not path_value:
        return []
    path = Path(path_value).expanduser()
    try:
        data = json.loads(path.read_text(encoding='utf-8'))
    except (OSError, json.JSONDecodeError):
        return []
    raw_ids = data.get('chatIds') if isinstance(data, dict) else None
    if not isinstance(raw_ids, list):
        return []
    return list(dict.fromkeys(str(item) for item in raw_ids if item is not None))

def send_telegram_message(text_message: str, *, token: Optional[str]=None, chat_id: Optional[str]=None, chat_ids: Optional[list[str]]=None, chats_file: Optional[str]=None, retries: int=3) -> dict[str, Any]:
    token = token or os.getenv('TELEGRAM_BOT_TOKEN')
    if not token:
        return {'status': 'SKIPPED_MISSING_TELEGRAM_CONFIG', 'sent': False, 'message_id': None}
    targets: list[str] = []
    if chat_id:
        targets.append(str(chat_id))
    for value in (chat_ids or _load_telegram_chat_ids(chats_file)):
        if str(value) not in targets:
            targets.append(str(value))
    if not targets:
        env_chat_id = os.getenv('TELEGRAM_CHAT_ID')
        if env_chat_id:
            targets.append(env_chat_id)
    if not targets:
        return {'status': 'SKIPPED_MISSING_TELEGRAM_CONFIG', 'sent': False, 'message_id': None}
    url = f'https://api.telegram.org/bot{token}/sendMessage'
    per_chat: list[dict[str, Any]] = []
    for target in targets:
        payload = json.dumps({'chat_id': target, 'text': text_message[:4096], 'parse_mode': 'HTML'}).encode('utf-8')
        last_error = ''
        outcome: Optional[dict[str, Any]] = None
        for attempt in range(1, retries + 1):
            request = urllib.request.Request(url, data=payload, headers={'Content-Type': 'application/json'}, method='POST')
            try:
                with urllib.request.urlopen(request, timeout=20) as response:
                    body = json.loads(response.read().decode('utf-8'))
                if body.get('ok'):
                    outcome = {'chat_id': target, 'status': 'SENT', 'sent': True, 'message_id': str((body.get('result') or {}).get('message_id')), 'attempts': attempt}
                    break
                last_error = str(body.get('description') or 'Telegram returned ok=false')
            except (urllib.error.URLError, urllib.error.HTTPError, TimeoutError, json.JSONDecodeError) as exc:
                last_error = f'{type(exc).__name__}: {exc}'
            if attempt < retries:
                time.sleep(min(4, 2 ** (attempt - 1)))
        per_chat.append(outcome or {'chat_id': target, 'status': 'ERROR_TELEGRAM_SEND_FAILED', 'sent': False, 'message_id': None, 'attempts': retries, 'error': last_error})
    any_sent = any(item['sent'] for item in per_chat)
    first_message_id = next((item['message_id'] for item in per_chat if item['sent']), None)
    return {
        'status': 'SENT' if any_sent else 'ERROR_TELEGRAM_SEND_FAILED',
        'sent': any_sent,
        'message_id': first_message_id,
        'chats_attempted': len(targets),
        'chats_sent': sum(1 for item in per_chat if item['sent']),
        'per_chat': per_chat,
    }

def format_gate(calculation: dict[str, Any]) -> dict[str, Any]:
    snapshot_format = calculation['canonical_snapshot'].get('format', {})
    data = calculation['data_gate']
    warnings = list(snapshot_format.get('warnings', []))
    same_format_n = int(data.get('pooled_n') or 0)
    if same_format_n < 20:
        warnings.append('SAME_FORMAT_SAMPLE_BELOW_20')
    if same_format_n <= 8:
        warnings.append('VERY_SMALL_SAME_FORMAT_SAMPLE')
    if data.get('history_format_override_games'):
        warnings.append('LEGACY_HISTORY_FORMAT_OVERRIDE_USED')
    return {
        'current_format': snapshot_format.get('format_key'),
        'quarter_minutes': snapshot_format.get('quarter_minutes'),
        'regulation_minutes': snapshot_format.get('regulation_minutes'),
        'format_source': snapshot_format.get('source'),
        'same_format_history_team_a_n': data.get('history_team_a_n'),
        'same_format_history_team_b_n': data.get('history_team_b_n'),
        'same_format_pooled_n': same_format_n,
        'cross_format_history_team_a_n': data.get('cross_format_team_a_n'),
        'cross_format_history_team_b_n': data.get('cross_format_team_b_n'),
        'cross_format_exact_hits_used': False,
        'cross_format_normalized_baseline_used': bool(data.get('cross_format_normalized_baseline_allowed')),
        'small_sample_cap': DEFAULT_CONFIG['caps']['small_sample'] if same_format_n < 20 else None,
        'warnings': list(dict.fromkeys(warnings)),
    }

def process_vps_match_file(
    match_path: str | Path,
    *,
    output_path: str | Path | None = None,
    zones_path: str | Path | None = None,
    db_path: str | Path = 'super_basket.sqlite3',
    mode: str = 'ACTION',
    require_gpt: bool = False,
    enable_gpt: bool = False,
    enable_telegram: bool = True,
    dry_run: bool = False,
    strict_schema: bool = False,
    checkpoint: Optional[int] = None,
    gpt_reviewer: Optional[Callable[[dict[str, Any], dict[str, Any]], dict[str, Any]]] = None,
    telegram_sender: Optional[Callable[[str], dict[str, Any]]] = None,
) -> dict[str, Any]:
    mode = mode.upper()
    if mode not in {'ACTION', 'STRICT'}:
        raise ValueError('mode must be ACTION or STRICT')
    source_path = Path(match_path).expanduser().resolve()
    source = load_json(source_path)
    if checkpoint is None:
        checkpoint = _v11_checkpoint_from_filename(source_path)
    if checkpoint is not None:
        checkpoint = int(checkpoint)
        if checkpoint not in {1, 2, 3}:
            raise ValueError('checkpoint must be 1, 2 or 3')
        context = source.get('analysis_context') if isinstance(source.get('analysis_context'), dict) else {}
        source['analysis_context'] = {**context, 'trigger_checkpoint': checkpoint}
    zones, zones_metadata = resolve_team_relative_zones(source, zones_path=zones_path)
    core_result = SuperBasketCalculator(deepcopy(DEFAULT_CONFIG), zones, zones_metadata).calculate(
        source,
        dispatch_threshold=float(DEFAULT_CONFIG['dispatch_threshold']),
        strict_schema=strict_schema,
    )
    calculation = core_result['super_basket_calculation']
    target = Path(output_path).expanduser().resolve() if output_path else source_path.with_name(source_path.stem + '_result.json')
    store = LearningStore(db_path)
    try:
        evaluations = [apply_learning_to_evaluation(item, store, calculation, mode) for item in calculation['market_evaluations']]
        calculation['market_evaluations'] = evaluations
        line_recommendations = [summarize_line_evaluation(item) for item in evaluations]
        line_recommendations.sort(key=lambda item: (str(item.get('market_type')), str(item.get('team')), str(item.get('segment')), float(item.get('line') or 0.0), str(item.get('side'))))
        calculation['line_recommendations'] = line_recommendations
        calculation['active_line_recommendations'] = [item for item in line_recommendations if item.get('action') in {'PLAY', 'RISK'}]
        selected, closest = select_one_decision(evaluations, mode)
        decision = build_decision(selected, closest, calculation, mode)
        input_usage = build_input_usage(source, calculation, selected or closest)
        if selected and input_usage['data_conflict']:
            decision['action'] = 'PASS'
            decision['deterministic_action'] = 'PASS'
            decision['status'] = 'PASS'
            decision['stake'] = '0%'
            decision['reason_codes'].insert(0, 'DATA_CONFLICT_PRECOMPUTED_VS_RAW')
            decision['blockers'].append(_blocker('DATA_CONFLICT_PRECOMPUTED_VS_RAW', 'Parser table and recomputed raw history disagree above tolerance'))
            decision['explanation_uk'] = 'Рішення PASS: готова таблиця парсера не збігається з повторним розрахунком raw history.'
            decision['main_risk_uk'] = 'Неможливо визначити, яке джерело історії є актуальним.'
        deterministic_action = decision['action']
        existing_before_review = store.get_signal(decision['signal_id']) if decision.get('signal_id') else None
        duplicate_already_sent = bool(existing_before_review and existing_before_review.get('telegram_status') == 'SENT')
        review: dict[str, Any]
        if deterministic_action == 'PASS':
            review = {'status': 'SKIPPED_DETERMINISTIC_PASS', 'approved': False, 'action': 'PASS', 'explanation_uk': '', 'main_risk_uk': '', 'telegram_text_uk': ''}
        elif duplicate_already_sent:
            previous_action = existing_before_review.get('final_action') or deterministic_action
            review = {'status': 'SKIPPED_DUPLICATE_ALREADY_SENT', 'approved': previous_action in {'PLAY', 'RISK'}, 'action': previous_action, 'explanation_uk': '', 'main_risk_uk': '', 'telegram_text_uk': ''}
        elif dry_run:
            review = {'status': 'DRY_RUN_NOT_CALLED', 'approved': True, 'action': deterministic_action, 'explanation_uk': decision['explanation_uk'], 'main_risk_uk': decision['main_risk_uk'], 'telegram_text_uk': ''}
        elif not enable_gpt:
            if require_gpt:
                review = {'status': 'REQUIRED_BUT_DISABLED', 'approved': False, 'action': 'PASS', 'explanation_uk': '', 'main_risk_uk': '', 'telegram_text_uk': ''}
            else:
                review = {'status': 'BYPASSED_BY_CONFIGURATION', 'approved': True, 'action': deterministic_action, 'explanation_uk': decision['explanation_uk'], 'main_risk_uk': decision['main_risk_uk'], 'telegram_text_uk': ''}
        else:
            reviewer = gpt_reviewer or (lambda d, c: gpt_review_decision(d, c))
            review = reviewer(decision, calculation)
        ranks = {'PASS': 0, 'RISK': 1, 'PLAY': 2}
        reviewed_action = review.get('action', 'PASS')
        if ranks.get(reviewed_action, 0) > ranks.get(deterministic_action, 0):
            reviewed_action = deterministic_action
            review['status'] = 'MODEL_UPGRADE_BLOCKED'
            review['approved'] = False
        # GPT review is recorded for context (gpt_status / explanation text) but never
        # blocks or downgrades dispatch anymore — the deterministic action always stands,
        # so RISK/PLAY signals go out to Telegram regardless of GPT approval or failure.
        if deterministic_action != 'PASS':
            decision['action'] = deterministic_action
        risk_post_filter = apply_risk_post_filter(decision)
        decision['risk_post_filter'] = risk_post_filter
        decision['gpt_status'] = review.get('status')
        decision['telegram_status'] = 'NOT_ATTEMPTED'
        delivery = {
            'status': (
                'SKIPPED_RISK_POSTFILTER_P_LIVE_NOT_ABOVE_85'
                if risk_post_filter['filtered']
                else 'SKIPPED_PASS'
            ),
            'sent': False,
            'message_id': None,
        }
        duplicate = False
        if selected and decision.get('signal_id'):
            existing, duplicate = store.record_signal(decision, calculation)
            already_sent = existing.get('telegram_status') == 'SENT'
            if already_sent:
                delivery = {'status': 'SKIPPED_DUPLICATE_ALREADY_SENT', 'sent': False, 'message_id': existing.get('telegram_message_id')}
            elif risk_post_filter['filtered']:
                delivery = {
                    'status': 'SKIPPED_RISK_POSTFILTER_P_LIVE_NOT_ABOVE_85',
                    'sent': False,
                    'message_id': None,
                }
            elif decision['action'] in {'PLAY', 'RISK'} and dry_run:
                delivery = {'status': 'DRY_RUN_NOT_SENT', 'sent': False, 'message_id': None}
            elif decision['action'] in {'PLAY', 'RISK'} and not enable_telegram:
                delivery = {'status': 'SKIPPED_TELEGRAM_DISABLED', 'sent': False, 'message_id': None}
            elif decision['action'] in {'PLAY', 'RISK'}:
                message = build_telegram_message(decision, calculation, review)
                sender = telegram_sender or (lambda value: send_telegram_message(value))
                delivery = sender(message)
            decision['telegram_status'] = delivery['status']
            if not already_sent:
                store.update_delivery(decision['signal_id'], decision['action'], review.get('status', ''), delivery['status'], delivery.get('message_id'))
        market_audit = deepcopy(calculation.get('market_audit', {}))
        market_audit.update({
            'evaluated_unique_market_sides': len(evaluations),
            'system_eligible_count': sum(item['system_action'] != 'PASS' for item in evaluations),
            'one_signal_selected': selected is not None,
            'all_unique_line_sides_evaluated': bool(calculation.get('line_coverage', {}).get('all_unique_line_sides_evaluated')),
            'distinct_line_variants_evaluated': calculation.get('line_coverage', {}).get('distinct_line_variants_evaluated'),
            'line_recommendation_count': len(line_recommendations),
            'active_line_recommendation_count': sum(item.get('action') in {'PLAY', 'RISK'} for item in line_recommendations),
        })
        evaluation_for_output = decision.pop('_evaluation', None)
        decision_probability = decision['probabilities'].get('p_final')
        probability_text = f'{float(decision_probability):.1%}' if decision_probability is not None else 'n/a'
        system = {
            'version': SYSTEM_VERSION,
            'processed_at': utc_now(),
            'input_hash': calculation['input_snapshot_hash'],
            'mode': mode,
            'status': 'OK' if not input_usage['data_conflict'] else 'DATA_CONFLICT',
            'data_gate': calculation['data_gate'],
            'format_gate': format_gate(calculation),
            'input_usage': input_usage,
            'market_audit': market_audit,
            'line_coverage': deepcopy(calculation.get('line_coverage') or {}),
            'line_recommendations': deepcopy(line_recommendations),
            'active_line_recommendations': deepcopy(calculation.get('active_line_recommendations') or []),
            'decision': decision,
            'decision_text': f"{decision['action']} | {decision['status']} | P_final {probability_text}",
            'gpt_review': review,
            'risk_post_filter': risk_post_filter,
            'telegram_delivery': {**delivery, 'duplicate_signal': duplicate},
            'learning': evaluation_for_output.get('calibration') if evaluation_for_output else {'status': 'NO_MARKET'},
            'files': {'source': str(source_path), 'result': str(target)},
        }
        core_result['super_basket_system'] = system
        snapshot = calculation['canonical_snapshot']
        snapshot_quarters = snapshot.get('quarters') or []
        completed_quarters = sum(
            1 for quarter in snapshot_quarters
            if isinstance(quarter, dict)
            and quarter.get('home') is not None
            and quarter.get('away') is not None
        )
        line_reason = next((
            code for code in decision['reason_codes']
            if code in {
                'NO_SUPPORTED_REAL_LINES',
                'NO_LINE',
                'NO_ODDS',
                'ODDS_BELOW_MINIMUM',
                'SYNTHETIC_LINE',
                'UNSUPPORTED_MARKET',
            }
        ), None)
        append_verdict_log({
            'timestamp':    system['processed_at'],          # utc_now(), напр. 2026-07-19T10:15:00+00:00
            'match_id':     snapshot['match_id'],
            'match_name':   snapshot['name'],
            'checkpoint':   snapshot['stage'],                # computed live stage
            'trigger_checkpoint': snapshot.get('trigger_checkpoint'), # 1/Q1, 2/HT, 3/Q3 queue source
            'explicit_stage': snapshot.get('explicit_stage'), # сирий статус з фіда, для звірки
            'verdict':      decision['action'],                # PASS / RISK / PLAY (фінальне рішення після GPT-гейту)
            'verdict_status': decision['status'],              # людський статус, напр. "RISK ENTRY — GPT DOWNGRADE"
            'deterministic_verdict': decision['deterministic_action'],  # рішення ДО GPT-огляду (чисті формули)
            'p_final':      decision['probabilities'].get('p_final'),
            'market':       decision.get('market'),
            'description':  decision['explanation_uk'],
            'main_risk':    decision['main_risk_uk'],
            'reason_codes': decision['reason_codes'],
            'input_hash':   calculation['input_snapshot_hash'],
            'gpt_status':   system['gpt_review']['status'],
            'telegram_status': system['telegram_delivery']['status'],
            'stage_context': {
                'current_quarter': snapshot.get('current_quarter'),
                'completed_quarters': completed_quarters,
                'clock': snapshot.get('clock'),
                'elapsed_game_seconds': snapshot.get('elapsed_game_seconds'),
                'remaining_game_seconds': snapshot.get('remaining_game_seconds'),
                'score': snapshot.get('score'),
                'quarters': snapshot_quarters,
                'time_reliable': calculation.get('data_gate', {}).get('time_reliable'),
            },
            'line_diagnostics': {
                'detected_market_sides': len(calculation.get('markets_detected') or []),
                'evaluated_market_sides': len(calculation.get('market_evaluations') or []),
                'eligible_candidates': len(calculation.get('candidates') or []),
                'offers_before_deduplication': market_audit.get('offer_sides_before_deduplication'),
                'unique_market_sides': market_audit.get('unique_market_sides'),
                'duplicate_offers_removed': market_audit.get('duplicate_offers_removed'),
                'selected_market_source': (decision.get('market') or {}).get('bookmaker'),
                'empty_or_rejected_reason': line_reason,
            },
            'probabilities': deepcopy(decision.get('probabilities') or {}),
            'budget_recommendation': deepcopy(decision.get('budget_recommendation') or {}),
            'coursework_forecast': deepcopy(calculation.get('coursework_forecast') or {}),
            'gates': {
                'caps': deepcopy(decision.get('caps') or []),
                'blockers': deepcopy(decision.get('blockers') or []),
            },
            'files': {
                'source': str(source_path),
                'result': str(target),
            },
        })
        save_json(target, core_result)
        if ENABLE_EXCEL_AUDIT:
            append_excel_audit(core_result)
        store.mark_processed(calculation['input_snapshot_hash'], str(source_path), str(target), system['status'])
        return core_result
    finally:
        store.close()

def _finished_match(canonical: dict[str, Any]) -> bool:
    status = str(canonical.get('explicit_stage') or '').upper()
    long_markers = {'FINAL', 'FINISHED', 'ENDED', 'ЗАВЕРШЕНО', 'КІНЕЦЬ'}
    return canonical['elapsed_game_seconds'] >= canonical['full_game_seconds'] or bool(re.search(r'\bFT\b', status)) or any(marker in status for marker in long_markers)

def _signal_outcome_value(signal: dict[str, Any], canonical: dict[str, Any]) -> Optional[float]:
    market_type = signal['market_type']
    segment = signal['segment']
    team = signal.get('team')
    team_side = 'home' if team == canonical['home_team'] else 'away' if team == canonical['away_team'] else None
    if market_type == 'MATCH_TOTAL':
        return float(canonical['score']['total'])
    if market_type == 'TEAM_IT_MATCH' and team_side:
        return float(canonical['score'][team_side])
    if segment in {'H1', 'H2'}:
        quarters = canonical['quarters'][:2] if segment == 'H1' else canonical['quarters'][2:]
    elif segment.startswith('Q') and segment[1:].isdigit():
        quarters = [canonical['quarters'][int(segment[1:]) - 1]]
    else:
        return None
    key = team_side if team_side else 'total'
    values = [to_number(quarter.get(key)) for quarter in quarters]
    return sum(float(value) for value in values) if values and all(value is not None for value in values) else None

def settle_finished_match_file(match_path: str | Path, db_path: str | Path) -> dict[str, Any]:
    source = load_json(Path(match_path).expanduser().resolve())
    canonical = adapt_match(source, deepcopy(DEFAULT_CONFIG))
    if not _finished_match(canonical):
        raise ValueError('Match is not marked finished; settlement refused')
    store = LearningStore(db_path)
    settled: list[dict[str, Any]] = []
    try:
        rows = store.connection.execute("SELECT * FROM signals WHERE match_id=? AND result IS NULL AND final_action IN ('PLAY','RISK')", (canonical['match_id'],)).fetchall()
        for row in rows:
            signal = dict(row)
            value = _signal_outcome_value(signal, canonical)
            if value is None:
                continue
            line = float(signal['line'])
            if abs(value - line) < 1e-9:
                result = 'PUSH'
            elif signal['side'] == 'OVER':
                result = 'WIN' if value > line else 'LOSS'
            else:
                result = 'WIN' if value < line else 'LOSS'
            settled.append(store.settle(signal['signal_id'], result, value))
        return {'match_id': canonical['match_id'], 'settled_count': len(settled), 'settled': settled, 'report': store.report()}
    finally:
        store.close()

def watch_inbox(
    inbox: str | Path,
    outbox: str | Path,
    *,
    zones_path: str | Path | None,
    db_path: str | Path,
    mode: str,
    require_gpt: bool,
    enable_gpt: bool,
    enable_telegram: bool,
    poll_seconds: float,
) -> None:
    inbox_path = Path(inbox).expanduser().resolve()
    outbox_path = Path(outbox).expanduser().resolve()
    inbox_path.mkdir(parents=True, exist_ok=True)
    outbox_path.mkdir(parents=True, exist_ok=True)
    signatures: dict[str, tuple[int, int]] = {}
    stable: dict[str, int] = {}
    processed: dict[str, tuple[int, int]] = {}
    print(f'WATCHING {inbox_path} -> {outbox_path}', flush=True)
    while True:
        for path in sorted(inbox_path.glob('*.json')):
            if path.name.endswith(('_result.json', '_calculated.json')):
                continue
            try:
                stat_result = path.stat()
                signature = (stat_result.st_size, stat_result.st_mtime_ns)
            except OSError:
                continue
            key = str(path)
            if signatures.get(key) == signature:
                stable[key] = stable.get(key, 0) + 1
            else:
                signatures[key] = signature
                stable[key] = 0
            if stable[key] < 1 or processed.get(key) == signature:
                continue
            output = outbox_path / f'{path.stem}_result.json'
            try:
                result = process_vps_match_file(path, output_path=output, zones_path=zones_path, db_path=db_path, mode=mode, require_gpt=require_gpt, enable_gpt=enable_gpt, enable_telegram=enable_telegram)
                decision = result['super_basket_system']['decision']
                print(f"{utc_now()} {path.name}: {decision['action']} {decision['status']}", flush=True)
                processed[key] = signature
            except (OSError, ValueError, KeyError, json.JSONDecodeError, sqlite3.Error) as exc:
                print(f'{utc_now()} ERROR {path.name}: {type(exc).__name__}: {exc}', file=sys.stderr, flush=True)
        time.sleep(max(0.5, poll_seconds))




def calculate_match_file(
    match_path: str | Path,
    *,
    zones_path: str | Path | None = None,
    output_path: str | Path | None = None,
    in_place: bool = True,
    dispatch_threshold: float | None = None,
    strict_schema: bool = False,
) -> dict[str, Any]:
    """Calculate one parser JSON and write the calculation block.

    `zones_path` is optional. Resolution order is explicit path, environment,
    parser-embedded table, script sibling, current directory, then last35.
    """
    match_path = Path(match_path).expanduser().resolve()
    source = load_json(match_path)
    zones, zones_metadata = resolve_team_relative_zones(source, zones_path=zones_path)
    result = SuperBasketCalculator(deepcopy(DEFAULT_CONFIG), zones, zones_metadata).calculate(
        source,
        dispatch_threshold=dispatch_threshold,
        strict_schema=strict_schema,
    )
    target = match_path if in_place else Path(output_path).expanduser().resolve() if output_path else match_path.with_name(match_path.stem + '_calculated.json')
    save_json(target, result)
    return result


def _add_runtime_switches(parser: argparse.ArgumentParser) -> None:
    parser.add_argument('--mode', choices=['action', 'strict'], default=os.getenv('SUPER_BASKET_MODE', 'action').lower())
    parser.add_argument('--require-gpt', dest='require_gpt', action='store_true', default=env_bool('SUPER_BASKET_REQUIRE_GPT', False))
    parser.add_argument('--no-require-gpt', dest='require_gpt', action='store_false')
    parser.add_argument('--gpt', dest='enable_gpt', action='store_true', default=env_bool('SUPER_BASKET_ENABLE_GPT', False))
    parser.add_argument('--no-gpt', dest='enable_gpt', action='store_false')
    parser.add_argument('--telegram', dest='enable_telegram', action='store_true', default=env_bool('SUPER_BASKET_ENABLE_TELEGRAM', True))
    parser.add_argument('--no-telegram', dest='enable_telegram', action='store_false')

def _single_file_cli(argv: list[str] | None = None) -> int:
    argv = list(sys.argv[1:] if argv is None else argv)
    if argv and argv[0].startswith('--') and '--match' in argv:
        argv.insert(0, 'run')
    parser = argparse.ArgumentParser(description=f'SUPER_BASKET VPS STANDALONE ACTION SYSTEM v{SYSTEM_VERSION}')
    subparsers = parser.add_subparsers(dest='command', required=True)

    run = subparsers.add_parser('run', help='Process one parser JSON')
    run.add_argument('--match', required=True)
    run.add_argument('--output')
    run.add_argument('--zones')
    run.add_argument('--db', default=os.getenv('SUPER_BASKET_DB', 'super_basket.sqlite3'))
    run.add_argument('--dry-run', action='store_true', help='Calculate without external GPT/Telegram calls')
    run.add_argument('--strict-schema', action='store_true')
    run.add_argument('--checkpoint', type=int, choices=[1, 2, 3, 4, 5], help='Advisor checkpoint: 1=prematch, 2=early_live_q1 (2 min Q1), 3=early_live_q2 (2 min Q2), 4=HT, 5=q4_confirmation (4 min Q4)')
    _add_runtime_switches(run)

    watch = subparsers.add_parser('watch', help='Continuously process stable JSON files in an inbox')
    watch.add_argument('--inbox', required=True)
    watch.add_argument('--outbox', required=True)
    watch.add_argument('--zones')
    watch.add_argument('--db', default=os.getenv('SUPER_BASKET_DB', 'super_basket.sqlite3'))
    watch.add_argument('--poll-seconds', type=float, default=2.0)
    _add_runtime_switches(watch)

    settle = subparsers.add_parser('settle', help='Settle one signal manually')
    settle.add_argument('--signal-id', required=True)
    settle.add_argument('--result', required=True, choices=['win', 'loss', 'push'])
    settle.add_argument('--db', default=os.getenv('SUPER_BASKET_DB', 'super_basket.sqlite3'))

    settle_match = subparsers.add_parser('settle-match', help='Settle all active signals from a finished match JSON')
    settle_match.add_argument('--match', required=True)
    settle_match.add_argument('--db', default=os.getenv('SUPER_BASKET_DB', 'super_basket.sqlite3'))

    report = subparsers.add_parser('report', help='Print SQLite learning/performance report')
    report.add_argument('--db', default=os.getenv('SUPER_BASKET_DB', 'super_basket.sqlite3'))

    check = subparsers.add_parser('check-config', help='Check deployment configuration without printing secrets')
    check.add_argument('--db', default=os.getenv('SUPER_BASKET_DB', 'super_basket.sqlite3'))

    args = parser.parse_args(argv)
    try:
        if args.command == 'run':
            result = process_vps_match_file(
                args.match,
                output_path=args.output,
                zones_path=args.zones,
                db_path=args.db,
                mode=args.mode,
                require_gpt=args.require_gpt,
                enable_gpt=args.enable_gpt,
                enable_telegram=args.enable_telegram,
                dry_run=args.dry_run,
                strict_schema=args.strict_schema,
                checkpoint=args.checkpoint,
            )
            system = result['super_basket_system']
            summary = {
                'output_status': system['status'],
                'match_id': result['super_basket_calculation']['canonical_snapshot']['match_id'],
                'stage': result['super_basket_calculation']['canonical_snapshot']['stage'],
                'trigger_checkpoint': result['super_basket_calculation']['canonical_snapshot'].get('trigger_checkpoint'),
                'format': system['format_gate']['current_format'],
                'decision': deepcopy(system['decision']),
                'gpt_status': system['gpt_review']['status'],
                'telegram_status': system['telegram_delivery']['status'],
            }
            summary['decision'].pop('p_trace', None)
            summary['decision'].pop('caps', None)
            summary['decision'].pop('blockers', None)
            print(json.dumps(summary, ensure_ascii=False, indent=2))
        elif args.command == 'watch':
            watch_inbox(args.inbox, args.outbox, zones_path=args.zones, db_path=args.db, mode=args.mode, require_gpt=args.require_gpt, enable_gpt=args.enable_gpt, enable_telegram=args.enable_telegram, poll_seconds=args.poll_seconds)
        elif args.command == 'settle':
            store = LearningStore(args.db)
            try:
                settled = store.settle(args.signal_id, args.result)
                output = {'settled': settled, 'report': store.report()}
            finally:
                store.close()
            print(json.dumps(output, ensure_ascii=False, indent=2))
        elif args.command == 'settle-match':
            print(json.dumps(settle_finished_match_file(args.match, args.db), ensure_ascii=False, indent=2))
        elif args.command == 'report':
            store = LearningStore(args.db)
            try:
                print(json.dumps(store.report(), ensure_ascii=False, indent=2))
            finally:
                store.close()
        elif args.command == 'check-config':
            store = LearningStore(args.db)
            store.close()
            print(json.dumps({
                'python': sys.version.split()[0],
                'database_ready': True,
                'openai_api_key_set': bool(os.getenv('OPENAI_API_KEY')),
                'openai_model': os.getenv('OPENAI_MODEL', 'gpt-5.6'),
                'telegram_bot_token_set': bool(os.getenv('TELEGRAM_BOT_TOKEN')),
                'telegram_chat_id_set': bool(os.getenv('TELEGRAM_CHAT_ID')),
                'telegram_chats_file': os.getenv('TELEGRAM_CHATS_FILE'),
                'telegram_chats_file_chat_count': len(_load_telegram_chat_ids()),
                'require_gpt': env_bool('SUPER_BASKET_REQUIRE_GPT', False),
                'live_limit_set': to_number(os.getenv('SUPER_BASKET_LIVE_LIMIT')) is not None,
                'bankroll_set': to_number(os.getenv('SUPER_BASKET_BANKROLL')) is not None,
                'budget_currency': os.getenv('SUPER_BASKET_CURRENCY', 'USDT'),
                'excel_audit_enabled': ENABLE_EXCEL_AUDIT,
            }, ensure_ascii=False, indent=2))
    except KeyboardInterrupt:
        print('STOPPED', file=sys.stderr)
        return 130
    except (OSError, ValueError, KeyError, json.JSONDecodeError, sqlite3.Error) as exc:
        print(f'ERROR: {type(exc).__name__}: {exc}', file=sys.stderr)
        return 1
    return 0




# ===== v9.0 FINAL HYBRID FULL-PARITY OVERRIDES =====
# Implemented from Basketball Hybrid Master v4.0 plus the project NO-STAT,
# Team-IT, production-router and Q4 rules.  The public CLI/API remains intact.

_V9_ADAPT_MATCH_BASE = adapt_match
_V9_CALCULATE_SCENARIO_BASE = calculate_scenario
_V9_CALCULATE_LIVE_BASE = calculate_live_projection
_V9_CALCULATE_STAT_BASE = calculate_stat_gate
_V9_EVALUATE_BASE = SuperBasketCalculator.evaluate_market
_V9_CALCULATE_BASE = SuperBasketCalculator.calculate
_V9_APPLY_LEARNING_BASE = apply_learning_to_evaluation
_V9_SUMMARIZE_BASE = summarize_line_evaluation

DEFAULT_CONFIG['engine_version'] = '9.1.0-FINAL-HYBRID-PARITY'
DEFAULT_CONFIG['dispatch_threshold'] = 0.60
DEFAULT_CONFIG.setdefault('signal_gates', {}).update({
    'history_zone_min': 0.75,
    'live_edge_min_points': 3.0,
    'scenario_direction_min': 0.50,
    'risk_min': 0.60,
    'play_min': 0.75,
    'allow_live_reversal': True,
    'live_reversal_p_live_min': 0.80,
    'live_reversal_p_scenario_min': 0.68,
})
# Correct the old 1.02 sum for current-quarter weights.
DEFAULT_CONFIG.setdefault('stage_weights', {})['CURRENT_Q1_Q3'] = {
    'hist': 0.225, 'scenario': 0.225, 'live': 0.55,
}
DEFAULT_CONFIG.setdefault('caps', {}).update({
    'partial_stat_3plus': 0.84,
    'partial_stat_under3': 0.79,
    'stat_neutral_live': 0.74,
    'no_stat_support_5_6': 0.79,
    'no_stat_support_4': 0.74,
    'no_stat_support_3': 0.72,
    'no_stat_support_0_2': 0.67,
    'full_stat_reversal': 0.79,
    'no_stat_reversal': 0.74,
})

_V9_FORMULA_REGISTRY = {
    'history_total': (
        'P_hist = normalized weighted blend of exact-line pooled probability, '
        'last5 shrunk probability, H2H shrunk modifier, distribution probability, '
        'and scored/allowed interaction. Mandatory signal zone is raw pooled exact-line >=75%.'
    ),
    'history_team_it': (
        'P_hist_IT = 0.50*Own_scored_exact + 0.35*Opponent_allowed_exact '
        '+ 0.15*H2H_IT_exact; weakest own/allowed gate controls cap/block.'
    ),
    'smoothing': 'p_smoothed = (hits + 1) / (N + 2).',
    'scenario_full_partial': (
        'Use only current-state patterns (quarter result, score/margin/total bucket, sequence); '
        'P_scenario = credibility*P_state_smoothed + (1-credibility)*0.50, '
        'credibility=min(1,N_state/8). Future-result/broad-history patterns are excluded.'
    ),
    'scenario_no_stat': (
        'N_state<3 => scenario OFF. Otherwise P_scenario = credibility*P_state_smoothed '
        '+ (1-credibility)*P_hist, credibility=min(1,N_state/8).'
    ),
    'projection_full_stat': (
        'Conservative multi-component projection: regressed segment pace + history + scenario '
        '+ stat-adjusted possession/efficiency projection + control; simple pace informational only.'
    ),
    'projection_partial_stat': (
        'Same conservative engine using only available stat groups; missing values remain N/A; '
        'PARTIAL_STAT cap84 with >=3 independent confirmations, otherwise cap79.'
    ),
    'projection_no_stat_n5': (
        'Projection_used = 0.35*Projection_regressed + 0.35*Projection_history '
        '+ 0.30*Projection_scenario when N_state>=5.'
    ),
    'projection_no_stat_n3_4': (
        'Projection_used = 0.40*Projection_regressed + 0.45*Projection_history '
        '+ 0.15*Projection_scenario when N_state=3-4.'
    ),
    'projection_no_stat_n0_2': (
        'Projection_used = 0.50*Projection_regressed + 0.50*Projection_history when N_state<3.'
    ),
    'p_live': (
        'OVER: Phi((Projection_used-Line)/sigma); UNDER: Phi((Line-Projection_used)/sigma). '
        'NO_STAT sigma = 1.20*sigma_base.'
    ),
    'p_raw': 'P_raw = w_hist*P_hist + w_scenario*P_scenario + w_live*P_live.',
    'p_final': 'P_final = min(P_raw after context, every active cap); any hard blocker => PASS.',
    'verdict': 'P_final<60% PASS; 60-74.99% RISK; >=75% PLAY, subject to gates/blockers.',
    'normal_signal_gates': (
        'Exact-line historical zone >=75%; live projection at least 3 points beyond line in signal '
        'direction; scenario not against; real odds >=1.44; router allowed; hard conflict OFF.'
    ),
    'reversal': (
        'Confirmed live reversal weights = 0.115*P_hist_live_side + 0.085*P_scenario_live_side '
        '+ 0.80*P_live_live_side. FULL/PARTIAL cap79; NO_STAT requires support score>=5 and cap74.'
    ),
}


def _v9_stat_channels(canonical: dict[str, Any]) -> dict[str, Any]:
    home = canonical.get('live_stats', {}).get('home', {})
    away = canonical.get('live_stats', {}).get('away', {})
    groups = {
        'FGA_POSS': bool((home.get('FGA') is not None or home.get('Poss') is not None)
                         and (away.get('FGA') is not None or away.get('Poss') is not None)),
        'FTA_FTR': bool((home.get('FTA') is not None or home.get('FTr') is not None)
                        and (away.get('FTA') is not None or away.get('FTr') is not None)),
        'ORB': home.get('ORB') is not None and away.get('ORB') is not None,
        'TO': home.get('TO') is not None and away.get('TO') is not None,
        'EFG': home.get('eFG') is not None and away.get('eFG') is not None,
    }
    count = sum(bool(value) for value in groups.values())
    if count >= 4:
        mode, support = 'FULL_STAT', 'ON'
    elif count >= 2:
        mode, support = 'PARTIAL_STAT', 'LIMITED'
    else:
        mode, support = 'SCORE_TIME_HISTORY', 'N/A_NO_STATS'
    if canonical.get('stage') != 'PRE_MATCH':
        gate = canonical.get('data_gate', {})
        if gate.get('schema_errors') or not gate.get('time_reliable', True):
            mode, support = 'DATA_OFF', 'OFF'
    return {
        'data_mode': mode,
        'stat_support': support,
        'groups': groups,
        'group_count': count,
        'missing_groups': [name for name, present in groups.items() if not present],
    }


def adapt_match(source: dict[str, Any], config: dict[str, Any], strict: bool=False) -> dict[str, Any]:
    canonical = _V9_ADAPT_MATCH_BASE(source, config, strict)
    classification = _v9_stat_channels(canonical)
    canonical['data_mode'] = classification['data_mode']
    canonical['stat_support'] = classification['stat_support']
    canonical['stat_channels'] = classification['groups']
    gate = canonical.setdefault('data_gate', {})
    gate['data_mode'] = classification['data_mode']
    gate['stat_support'] = classification['stat_support']
    gate['stat_channel_count'] = classification['group_count']
    gate['stat_channels'] = classification['groups']
    gate['missing_stat_groups'] = classification['missing_groups']
    gate['stats_found'] = classification['group_count'] > 0
    return canonical


_V9_STATE_GROUPS = {
    'quarter_result', 'score_state', 'margin_state', 'total_state',
    'sequence_state', 'line_threshold', 'time_state',
}


def _v9_state_sample(patterns: list[dict[str, Any]]) -> int:
    # Correlated patterns from one team cannot be added as independent rows.
    per_team: dict[str, int] = {}
    for item in patterns:
        team = str(item.get('team') or '')
        per_team[team] = max(per_team.get(team, 0), int(item.get('matched_games') or 0))
    return sum(per_team.values())


def calculate_scenario(
    market: dict[str, Any],
    canonical: dict[str, Any],
    history: dict[str, Any],
    config: dict[str, Any],
) -> dict[str, Any]:
    result = _V9_CALCULATE_SCENARIO_BASE(market, canonical, history, config)
    candidates = [
        item for item in result.get('patterns_found', [])
        if item.get('pattern_group') in _V9_STATE_GROUPS
        and item.get('pattern_group') != 'match_result'
        and int(item.get('matched_games') or 0) >= 3
    ]
    # One best pattern per team/state family prevents double counting.
    best: dict[tuple[str, str], dict[str, Any]] = {}
    for item in candidates:
        key = (str(item.get('team') or ''), str(item.get('pattern_group') or ''))
        if key not in best or float(item.get('pattern_rank') or 0.0) > float(best[key].get('pattern_rank') or 0.0):
            best[key] = item
    used = list(best.values())
    for item in used:
        item['used_in_scenario_v9'] = True
        item['pattern_weight_v9'] = max(
            0.0001,
            float(item.get('credibility') or 0.0)
            * float(item.get('sample_quality') or 0.0)
            * float(item.get('specificity') or 0.0),
        )
    denom = sum(float(item['pattern_weight_v9']) for item in used)
    if denom > 0:
        p_state = sum(
            float(item['pattern_weight_v9']) * float(item.get('smoothed_probability') or 0.50)
            for item in used
        ) / denom
        outcome_items = [
            (float(item['outcome_median']), float(item['pattern_weight_v9']))
            for item in used if item.get('outcome_median') is not None
        ]
        outcome_denom = sum(weight for _, weight in outcome_items)
        outcome_center = (
            sum(value * weight for value, weight in outcome_items) / outcome_denom
            if outcome_denom > 0 else None
        )
    else:
        p_state, outcome_center = 0.50, None
    n_state = _v9_state_sample(used)
    credibility = min(1.0, n_state / 8.0) if n_state > 0 else 0.0
    mode = canonical.get('data_mode') or canonical.get('data_gate', {}).get('data_mode')
    if n_state < 3:
        support = 'OFF'
        active = False
        probability = 0.50
        formula = 'N_state<3 => scenario OFF; scenario weight is renormalized away.'
    elif mode == 'SCORE_TIME_HISTORY':
        support = 'SMALL_SAMPLE' if n_state < 5 else 'ON'
        active = True
        probability = credibility * p_state + (1.0 - credibility) * float(history.get('p_hist') or 0.50)
        formula = 'credibility*P_state_smoothed + (1-credibility)*P_hist'
    else:
        support = 'SMALL_SAMPLE' if n_state < 5 else 'ON'
        active = True
        # FULL/PARTIAL scenario remains independent from broad history.
        probability = credibility * p_state + (1.0 - credibility) * 0.50
        formula = 'credibility*P_state_smoothed + (1-credibility)*0.50'
    result.update({
        'patterns_used': used,
        'state_patterns_used': used,
        'p_state_smoothed': p_state,
        'n_state': n_state,
        'effective_sample': float(n_state),
        'scenario_credibility': credibility,
        'p_scenario': max(0.0, min(1.0, probability)),
        'p_scenario_effective': max(0.0, min(1.0, probability)) if active else None,
        'scenario_support': support,
        'scenario_active': active,
        'outcome_center': outcome_center,
        'formula_v9': formula,
        'excluded_pattern_groups': ['quarter_strength', 'allowed_threshold', 'match_result'],
        'leakage_guard': 'Historical final-result patterns are excluded from P_scenario.',
    })
    return result


def _v9_segment_indices(market: dict[str, Any]) -> list[int]:
    segment = str(market.get('segment') or 'MATCH').upper()
    if segment == 'MATCH':
        return [0, 1, 2, 3]
    if segment == 'H1':
        return [0, 1]
    if segment == 'H2':
        return [2, 3]
    if segment.startswith('Q') and segment[1:].isdigit():
        idx = int(segment[1:]) - 1
        return [idx] if 0 <= idx <= 3 else []
    return []


def _v9_remaining_from_values(
    values: list[Optional[float]],
    indices: list[int],
    elapsed_game_seconds: float,
    quarter_seconds: float,
) -> Optional[float]:
    if not indices or quarter_seconds <= 0:
        return None
    total = 0.0
    for idx in indices:
        if idx >= len(values) or values[idx] is None:
            return None
        start = idx * quarter_seconds
        end = start + quarter_seconds
        if elapsed_game_seconds <= start:
            fraction = 1.0
        elif elapsed_game_seconds >= end:
            fraction = 0.0
        else:
            fraction = (end - elapsed_game_seconds) / quarter_seconds
        total += float(values[idx]) * max(0.0, min(1.0, fraction))
    return total


def _v9_game_total_quarters(game: dict[str, Any]) -> list[Optional[float]]:
    return [to_number(item.get('total')) for item in (game.get('quarters') or [])[:4]]


def _v9_game_team_quarters(game: dict[str, Any], team: str) -> list[Optional[float]]:
    quarters = game.get('quarters') or []
    if str(game.get('home_team')) == str(team):
        return [to_number(item.get('home')) for item in quarters[:4]]
    if str(game.get('away_team')) == str(team):
        return [to_number(item.get('away')) for item in quarters[:4]]
    return []


def _v9_required_probability(values: list[float], target: float, side: str) -> dict[str, Any]:
    valid = [float(value) for value in values if value is not None]
    if not valid:
        return {'available': False, 'p_required_history': None, 'hits': 0, 'n': 0, 'target': target}
    if side == 'OVER':
        hits = sum(value >= target for value in valid)
    else:
        hits = sum(value <= target for value in valid)
    p = (hits + 1.0) / (len(valid) + 2.0)
    return {
        'available': True,
        'p_required_history': p,
        'hits': int(hits),
        'n': len(valid),
        'target': target,
        'raw_rate': hits / len(valid),
        'median_remaining': statistics.median(valid),
        'mean_remaining': statistics.fmean(valid),
    }


def _v9_required_history(
    market: dict[str, Any], canonical: dict[str, Any]
) -> dict[str, Any]:
    indices = _v9_segment_indices(market)
    elapsed = float(canonical.get('elapsed_game_seconds') or 0.0)
    q_seconds = float(canonical.get('quarter_seconds') or 600.0)
    try:
        clock = _segment_clock(market, canonical)
        current = float(clock.get('current_points') or 0.0)
    except Exception:
        current = 0.0
    line = float(market['line'])
    side = str(market['side']).upper()
    threshold = math.floor(line) + 1 if side == 'OVER' else math.floor(line)
    target = float(threshold) - current
    if side == 'OVER' and target <= 0:
        return {'available': True, 'p_required_history': 1.0, 'hits': 1, 'n': 1, 'target': target, 'already_reached': True}
    if side == 'UNDER' and target < 0:
        return {'available': True, 'p_required_history': 0.0, 'hits': 0, 'n': 1, 'target': target, 'already_failed': True}

    market_type = str(market.get('market_type') or '')
    if market_type.startswith('TEAM_IT') or market_type == 'CURRENT_QUARTER_TEAM_IT':
        team = str(market.get('team') or '')
        own_pool = canonical['history']['team_a'] if team == canonical['home_team'] else canonical['history']['team_b']
        opponent_pool = canonical['history']['team_b'] if team == canonical['home_team'] else canonical['history']['team_a']
        own_values = []
        for game in own_pool:
            remaining = _v9_remaining_from_values(
                [to_number(v) for v in (game.get('team_quarters') or [])[:4]],
                indices, elapsed, q_seconds,
            )
            if remaining is not None:
                own_values.append(remaining)
        allowed_values = []
        for game in opponent_pool:
            remaining = _v9_remaining_from_values(
                [to_number(v) for v in (game.get('opponent_quarters') or [])[:4]],
                indices, elapsed, q_seconds,
            )
            if remaining is not None:
                allowed_values.append(remaining)
        h2h_values = []
        for game in canonical['history'].get('h2h', []):
            quarters = _v9_game_team_quarters(game, team)
            remaining = _v9_remaining_from_values(quarters, indices, elapsed, q_seconds) if quarters else None
            if remaining is not None:
                h2h_values.append(remaining)
        own = _v9_required_probability(own_values, target, side)
        allowed = _v9_required_probability(allowed_values, target, side)
        h2h = _v9_required_probability(h2h_values, target, side)
        components = {
            'own_remaining': own.get('p_required_history'),
            'opponent_allowed_remaining': allowed.get('p_required_history'),
            'h2h_remaining': h2h.get('p_required_history'),
        }
        probability, weights = _weighted_available(components, {
            'own_remaining': 0.50,
            'opponent_allowed_remaining': 0.35,
            'h2h_remaining': 0.15,
        })
        return {
            'available': own['available'] and allowed['available'],
            'p_required_history': probability,
            'target': target,
            'side': side,
            'current_points': current,
            'components': components,
            'component_weights': weights,
            'own': own,
            'opponent_allowed': allowed,
            'h2h': h2h,
            'method': 'fractional-current-quarter historical remaining; Team IT own/allowed/H2H blend',
        }

    values: list[float] = []
    for pool_name in ('team_a', 'team_b'):
        for game in canonical['history'].get(pool_name, []):
            remaining = _v9_remaining_from_values(
                _v9_game_total_quarters(game), indices, elapsed, q_seconds,
            )
            if remaining is not None:
                values.append(remaining)
    result = _v9_required_probability(values, target, side)
    result.update({
        'side': side,
        'current_points': current,
        'method': 'fractional-current-quarter pooled historical remaining',
    })
    return result


def calculate_stat_gate(
    market: dict[str, Any],
    canonical: dict[str, Any],
    zones_data: Optional[dict[str, Any]],
    *,
    project_counts_to_scope_end: bool=True,
) -> dict[str, Any]:
    result = _V9_CALCULATE_STAT_BASE(
        market, canonical, zones_data,
        project_counts_to_scope_end=project_counts_to_scope_end,
    )
    classification = _v9_stat_channels(canonical)
    mode = classification['data_mode']
    result['data_mode'] = mode
    result['stat_channels'] = classification['groups']
    result['stat_channel_count'] = classification['group_count']
    if mode in {'SCORE_TIME_HISTORY', 'DATA_OFF'}:
        result['stat_support'] = 'OFF' if mode == 'DATA_OFF' else 'N/A_NO_STATS'
        result['stat_gate_status'] = 'OFF'
        result['fake_over'] = False
        result['fake_under'] = False
        result['stat_unknown_not_zero'] = True
        return result
    result['stat_support'] = 'ON' if mode == 'FULL_STAT' else 'LIMITED'
    if mode == 'PARTIAL_STAT':
        over_count = len(set(result.get('over_positive_channels') or []))
        under_count = len(set(result.get('under_positive_channels') or []))
        evaluated = over_count if market.get('side') == 'OVER' else under_count
        opposite = under_count if market.get('side') == 'OVER' else over_count
        if evaluated >= 3:
            status = 'CONFIRMED'
        elif opposite >= 3 and evaluated < 3:
            status = 'AGAINST'
        else:
            status = 'NEUTRAL'
        result['stat_gate_status'] = status
        result['partial_independent_confirmations'] = evaluated
        result['partial_opposite_confirmations'] = opposite
    return result


def _v9_no_stat_edge_min(market: dict[str, Any], canonical: dict[str, Any]) -> float:
    market_type = str(market.get('market_type') or '')
    stage = str(canonical.get('stage') or '')
    if market_type == 'H1_TOTAL':
        return 5.0
    if market_type == 'H2_TOTAL':
        return 5.0
    if market_type == 'MATCH_TOTAL':
        if stage == 'EARLY_LIVE':
            return 10.0
        if stage == 'HT':
            return 7.0
        if stage in {'AFTER_3Q', 'Q4_CONFIRMATION'}:
            return 5.0
        return 10.0
    if market_type == 'TEAM_IT_MATCH':
        return 5.0
    if market_type == 'TEAM_IT_HALF':
        return 4.0
    if market_type == 'CURRENT_QUARTER_TOTAL':
        return 5.0
    if market_type == 'CURRENT_QUARTER_TEAM_IT':
        return 4.0
    return 5.0


def _v9_projection_alignment(live: dict[str, Any], scenario_active: bool, side: str, line: float) -> dict[str, Any]:
    values = {
        'regressed': to_number(live.get('projection_regressed') or live.get('projection_segment')),
        'history': to_number(live.get('projection_history')),
        'scenario': to_number(live.get('projection_scenario')) if scenario_active else None,
    }
    def aligned(value: Optional[float]) -> bool:
        if value is None:
            return False
        return value > line if side == 'OVER' else value < line
    flags = {key: aligned(value) for key, value in values.items() if value is not None}
    needed = 2 if scenario_active else len(flags)
    count = sum(flags.values())
    return {
        'values': values,
        'aligned_flags': flags,
        'aligned_count': count,
        'required_count': needed,
        'passed': bool(flags) and count >= needed,
    }


def _v9_no_stat_support(
    market: dict[str, Any],
    canonical: dict[str, Any],
    history: dict[str, Any],
    scenario: dict[str, Any],
    live: dict[str, Any],
) -> dict[str, Any]:
    scenario_active = bool(scenario.get('scenario_active'))
    n_state = int(scenario.get('n_state') or 0)
    small_patterns = [
        item for item in scenario.get('state_patterns_used', [])
        if 3 <= int(item.get('matched_games') or 0) <= 4
        and float(item.get('smoothed_probability') or 0.0) >= 0.68
    ]
    scenario_point = bool(
        (scenario_active and n_state >= 5 and float(scenario.get('p_scenario') or 0.0) >= 0.68)
        or len({(item.get('team'), item.get('pattern_group')) for item in small_patterns}) >= 2
    )
    edge_min = _v9_no_stat_edge_min(market, canonical)
    alignment = _v9_projection_alignment(
        live, scenario_active, str(market['side']), float(market['line'])
    )
    required = live.get('required_history') or {}
    checks = {
        'P_HIST_GE_68': float(history.get('p_hist') or 0.0) >= 0.68,
        'P_SCENARIO_GE_68_OR_TWO_SMALL': scenario_point,
        'P_LIVE_GE_70': float(live.get('p_live') or 0.0) >= 0.70,
        'MIN_STAGE_MARKET_EDGE': float(live.get('line_edge') or -999.0) >= edge_min,
        'REQUIRED_HISTORY_GE_68': float(required.get('p_required_history') or 0.0) >= 0.68,
        'PROJECTIONS_ALIGNED': alignment['passed'],
    }
    score = sum(bool(value) for value in checks.values())
    if score >= 5:
        cap = 0.79
    elif score == 4:
        cap = 0.74
    elif score == 3:
        cap = 0.72
    else:
        cap = 0.67
    return {
        'score': score,
        'max_score': 6,
        'checks': checks,
        'cap': cap,
        'edge_min': edge_min,
        'projection_alignment': alignment,
        'n_state': n_state,
        'scenario_active': scenario_active,
    }


def calculate_live_projection(
    market: dict[str, Any],
    canonical: dict[str, Any],
    history: dict[str, Any],
    scenario: dict[str, Any],
    config: dict[str, Any],
    stat: Optional[dict[str, Any]]=None,
) -> dict[str, Any]:
    result = _V9_CALCULATE_LIVE_BASE(market, canonical, history, scenario, config, stat)
    mode = canonical.get('data_mode') or canonical.get('data_gate', {}).get('data_mode')
    result['data_mode'] = mode
    result['projection_regressed'] = result.get('projection_segment')
    result['required_history'] = _v9_required_history(market, canonical)
    result['projection_formula_mode'] = 'FULL_STAT_CONSERVATIVE_BLEND'
    if mode == 'PARTIAL_STAT':
        result['projection_formula_mode'] = 'PARTIAL_STAT_AVAILABLE_CHANNEL_BLEND'
    elif mode == 'SCORE_TIME_HISTORY':
        regressed = to_number(result.get('projection_regressed'))
        history_projection = to_number(result.get('projection_history'))
        scenario_projection = to_number(result.get('projection_scenario')) if scenario.get('scenario_active') else None
        n_state = int(scenario.get('n_state') or 0)
        if n_state >= 5 and scenario_projection is not None:
            configured = {'regressed': 0.35, 'history': 0.35, 'scenario': 0.30}
            formula = '0.35*regressed + 0.35*history + 0.30*scenario'
        elif 3 <= n_state <= 4 and scenario_projection is not None:
            configured = {'regressed': 0.40, 'history': 0.45, 'scenario': 0.15}
            formula = '0.40*regressed + 0.45*history + 0.15*scenario'
        else:
            configured = {'regressed': 0.50, 'history': 0.50}
            formula = '0.50*regressed + 0.50*history'
        values = {
            'regressed': regressed,
            'history': history_projection,
            'scenario': scenario_projection,
        }
        available = {
            key: value for key, value in values.items()
            if key in configured and value is not None
        }
        denom = sum(configured[key] for key in available)
        if denom > 0:
            projection = sum(float(available[key]) * configured[key] for key in available) / denom
            weights = {key: configured[key] / denom for key in available}
        else:
            projection = float(result.get('projection_used') or market['line'])
            weights = {}
            formula += ' [fallback: no valid components]'
        sigma_base = _stage_sigma(market['market_type'], canonical['stage'], config)
        sigma = sigma_base * 1.20
        line = float(market['line'])
        edge = projection - line if market['side'] == 'OVER' else line - projection
        result.update({
            'projection_used': projection,
            'Projection_used': projection,
            'line_edge': edge,
            'line_edge_over': projection - line,
            'line_edge_under': line - projection,
            'sigma_base': sigma_base,
            'sigma': sigma,
            'z_score': edge / sigma if sigma > 0 else 0.0,
            'p_live': normal_cdf(edge / sigma) if sigma > 0 else 0.50,
            'projection_formula_mode': 'SCORE_TIME_HISTORY_NO_STAT',
            'projection_formula': formula,
            'projection_formula_weights': weights,
            'no_stat_components': values,
        })
    return result


def _v9_dedup_rules(items: list[dict[str, Any]], *, cap: bool=False) -> list[dict[str, Any]]:
    output: dict[str, dict[str, Any]] = {}
    for item in items:
        key = str(item.get('rule_id'))
        if key not in output:
            output[key] = item
        elif cap and float(item.get('cap') or 1.0) < float(output[key].get('cap') or 1.0):
            output[key] = item
    return list(output.values())


def _v9_stage_weights(market: dict[str, Any], canonical: dict[str, Any], config: dict[str, Any]) -> dict[str, float]:
    if market.get('market_type') in {'CURRENT_QUARTER_TOTAL', 'CURRENT_QUARTER_TEAM_IT'}:
        return {'hist': 0.225, 'scenario': 0.225, 'live': 0.55}
    raw = config.get('stage_weights', {}).get(
        canonical.get('stage'), config.get('stage_weights', {}).get('EARLY_LIVE', {})
    )
    normalized, _ = _normalize_weights(raw)
    return normalized


def _v9_reversal_conditions(
    market: dict[str, Any], canonical: dict[str, Any], evaluation: dict[str, Any]
) -> dict[str, Any]:
    history = evaluation['history']
    scenario = evaluation['scenario']
    live = evaluation['live']
    stat = evaluation['stat_comparison']
    zone = to_number(history.get('history_zone_rate'))
    opposite_zone = 1.0 - zone if zone is not None else None
    strong_edge = _strong_edge_threshold(market['market_type'], DEFAULT_CONFIG)
    common = bool(
        canonical.get('stage') != 'PRE_MATCH'
        and opposite_zone is not None and opposite_zone >= 0.90
        and float(live.get('p_live') or 0.0) >= 0.80
        and bool(scenario.get('scenario_active'))
        and float(scenario.get('p_scenario') or 0.0) >= 0.68
        and float(live.get('line_edge') or -999.0) >= strong_edge
        and not ((market['side'] == 'OVER' and stat.get('fake_over'))
                 or (market['side'] == 'UNDER' and stat.get('fake_under')))
    )
    mode = canonical.get('data_mode')
    if mode == 'SCORE_TIME_HISTORY':
        support = _v9_no_stat_support(market, canonical, history, scenario, live)
        active = bool(
            common
            and support['score'] >= 5
            and float((live.get('required_history') or {}).get('p_required_history') or 0.0) >= 0.68
        )
        cap = 0.74
    else:
        support = None
        active = bool(common and stat.get('stat_gate_status') == 'CONFIRMED')
        cap = 0.79
    return {
        'active': active,
        'opposite_history_zone': opposite_zone,
        'strong_edge_required': strong_edge,
        'mode': mode,
        'cap': cap,
        'no_stat_support': support,
    }


def _v9_evaluate_market(self: SuperBasketCalculator, market: dict[str, Any], canonical: dict[str, Any]) -> dict[str, Any]:
    evaluation = _V9_EVALUATE_BASE(self, market, canonical)
    evaluation['data_mode'] = canonical.get('data_mode')
    evaluation['stat_channels'] = deepcopy(canonical.get('stat_channels') or {})
    # Early return for parser/router-invalid markets.
    if 'history' not in evaluation or not evaluation.get('live') or evaluation.get('router', {}).get('hard_block'):
        return evaluation

    history = evaluation['history']
    scenario = evaluation['scenario']
    live = evaluation['live']
    stat = evaluation['stat_comparison']
    mode = canonical.get('data_mode')
    weights = _v9_stage_weights(market, canonical, self.config)
    scenario_active = bool(scenario.get('scenario_active'))
    if not scenario_active:
        weights['scenario'] = 0.0
        weights, _ = _normalize_weights(weights)
    reversal = _v9_reversal_conditions(market, canonical, evaluation)
    if reversal['active']:
        weights = {'hist': 0.115, 'scenario': 0.085, 'live': 0.80}
    p_scenario = float(scenario.get('p_scenario') or 0.50)
    p_raw = (
        weights['hist'] * float(history.get('p_hist') or 0.50)
        + weights['scenario'] * p_scenario
        + weights['live'] * float(live.get('p_live') or 0.50)
    )

    remove_caps = {'NO_STATS_FALLBACK', 'STAT_SUPPORT_LIMITED'}
    caps = [item for item in evaluation.get('caps', []) if item.get('rule_id') not in remove_caps]
    blockers = list(evaluation.get('blockers', []))
    if not scenario_active:
        blockers = [item for item in blockers if item.get('rule_id') != 'SCENARIO_DIRECTION_CONFLICT']
    if reversal['active']:
        blockers = [
            item for item in blockers
            if item.get('rule_id') not in {'HISTORY_ZONE_BELOW_75', 'STRONG_HISTORY_LIVE_CONFLICT'}
        ]
        caps.append(_cap(
            'CONFIRMED_LIVE_REVERSAL_CAP', reversal['cap'],
            'Strong current live direction overrides opposite 90%+ history only under reversal rules',
            reversal,
        ))

    # Re-assert mandatory normal gates after old-engine adjustments.
    zone = to_number(history.get('history_zone_rate'))
    if (zone is None or zone < 0.75) and not reversal['active']:
        blockers.append(_blocker(
            'HISTORY_ZONE_BELOW_75',
            'Exact-line history zone in signal direction must be at least 75%',
            {'history_zone_rate': zone, 'required': 0.75},
        ))
    if canonical.get('stage') != 'PRE_MATCH' and float(live.get('line_edge') or -999.0) < 3.0:
        blockers.append(_blocker(
            'LIVE_EDGE_BELOW_3',
            'Projection_used must be at least 3 points beyond line in signal direction',
            {'line_edge': live.get('line_edge'), 'required': 3.0},
        ))
    if scenario_active and p_scenario < 0.50:
        blockers.append(_blocker(
            'SCENARIO_DIRECTION_CONFLICT',
            'Current-state scenario points against the candidate',
            {'p_scenario': p_scenario},
        ))

    no_stat_support = None
    if mode == 'FULL_STAT':
        if canonical.get('stage') != 'PRE_MATCH' and stat.get('stat_gate_status') not in {'CONFIRMED'}:
            caps.append(_cap(
                'FULL_STAT_CONFIRMATION_NOT_ON', 0.74,
                'Live edge without 3/5 stat confirmation cannot be a clean PLAY',
                {'stat_gate_status': stat.get('stat_gate_status')},
            ))
    elif mode == 'PARTIAL_STAT':
        confirmations = int(stat.get('partial_independent_confirmations') or 0)
        cap = 0.84 if confirmations >= 3 else 0.79
        caps.append(_cap(
            'PARTIAL_STAT_CAP_84' if confirmations >= 3 else 'PARTIAL_STAT_CAP_79',
            cap,
            'PARTIAL_STAT uses only available groups; STRONG PLAY is forbidden',
            {'independent_confirmations': confirmations},
        ))
        if stat.get('stat_gate_status') == 'AGAINST':
            blockers.append(_blocker(
                'PARTIAL_STAT_GATE_AGAINST',
                'Available partial-stat channels directly oppose the candidate',
            ))
    elif mode == 'SCORE_TIME_HISTORY':
        no_stat_support = _v9_no_stat_support(market, canonical, history, scenario, live)
        caps.append(_cap(
            f"NO_STAT_SUPPORT_{no_stat_support['score']}_OF_6",
            float(no_stat_support['cap']),
            'NO_STAT_SUPPORT_SCORE cap',
            no_stat_support,
        ))
        if no_stat_support['score'] <= 2:
            blockers.append(_blocker(
                'NO_STAT_SUPPORT_TOO_LOW',
                'NO_STAT_SUPPORT_SCORE 0-2/6 cannot produce a signal',
                no_stat_support,
            ))
    elif mode == 'DATA_OFF':
        blockers.append(_blocker('DATA_OFF', 'Critical score/stage/time data are unavailable'))

    # Required-history is mandatory evidence in NO_STAT and is always logged.
    evaluation['required_history'] = deepcopy(live.get('required_history') or {})
    if mode == 'SCORE_TIME_HISTORY' and not (live.get('required_history') or {}).get('available'):
        blockers.append(_blocker(
            'REQUIRED_HISTORY_UNAVAILABLE',
            'Historical remaining-points distribution could not be calculated',
        ))

    # Keep Q4 harmonic/context reduction from the base engine conservatively.
    context_probability = p_raw
    if evaluation.get('q4_context', {}).get('applicable'):
        old_raw = float(evaluation.get('p_raw') or p_raw)
        old_final = float(evaluation.get('p_final') or old_raw)
        old_cap = min((float(item.get('cap') or 1.0) for item in evaluation.get('caps', [])), default=1.0)
        old_context = min(old_raw, old_cap)
        if old_final < old_context:
            context_probability = min(context_probability, old_final)

    caps = _v9_dedup_rules(caps, cap=True)
    blockers = _v9_dedup_rules(blockers)
    active_cap = min((float(item['cap']) for item in caps), default=1.0)
    p_final = max(0.0, min(1.0, context_probability, active_cap))
    clean = bool(
        not blockers
        and mode == 'FULL_STAT'
        and stat.get('stat_gate_status') == 'CONFIRMED'
        and not caps
        and zone is not None and zone >= 0.75
        and float(live.get('line_edge') or 0.0) >= 3.0
    )
    verdict = _verdict(p_final, blockers, clean, history.get('p_hist'))

    evaluation.update({
        'weights': {
            **(evaluation.get('weights') or {}),
            'normalized': weights,
            'scenario_off_renormalized': not scenario_active,
            'formula_sum': sum(weights.values()),
        },
        'p_raw': p_raw,
        'caps': caps,
        'blockers': blockers,
        'hard_conflict': bool(blockers),
        'p_final': p_final,
        'verdict': verdict,
        'live_reversal': {**(evaluation.get('live_reversal') or {}), **reversal},
        'no_stat_support': no_stat_support,
        'formula_registry_version': 'v9.1-FINAL-HYBRID-PARITY',
    })
    evaluation.setdefault('p_trace', []).append(_trace_step(
        'V9_FINAL_HYBRID_PARITY', True,
        'Separate FULL_STAT/PARTIAL_STAT/NO_STAT formulas; exact history zone; state scenario; RequiredHistoryP; support score; final caps',
        {
            'data_mode': mode,
            'weights': weights,
            'scenario_active': scenario_active,
            'required_history': live.get('required_history'),
            'no_stat_support': no_stat_support,
            'reversal': reversal,
            'active_cap': active_cap,
        },
        p_raw, p_final,
        [item.get('rule_id') for item in caps + blockers],
    ))
    return evaluation


SuperBasketCalculator.evaluate_market = _v9_evaluate_market


def _v9_calculate(self: SuperBasketCalculator, source: dict[str, Any], dispatch_threshold: Optional[float]=None, strict_schema: bool=False) -> dict[str, Any]:
    output = _V9_CALCULATE_BASE(self, source, dispatch_threshold, strict_schema)
    calculation = output.get('super_basket_calculation', {})
    calculation['engine_version'] = '9.1.0-FINAL-HYBRID-PARITY'
    snapshot = calculation.get('canonical_snapshot', {})
    evaluations = calculation.get('market_evaluations') or []
    if evaluations:
        snapshot['data_mode'] = evaluations[0].get('data_mode')
        snapshot['stat_channels'] = evaluations[0].get('stat_channels')
    calculation['formula_registry'] = deepcopy(_V9_FORMULA_REGISTRY)
    calculation['verdict_thresholds'] = {
        'PASS': '<60% or any hard blocker',
        'RISK': '60-74.99%',
        'PLAY': '>=75%',
    }
    output['super_basket_calculation'] = calculation
    return output


SuperBasketCalculator.calculate = _v9_calculate


def apply_learning_to_evaluation(evaluation: dict[str, Any], store: LearningStore, calculation: dict[str, Any], mode: str) -> dict[str, Any]:
    item = _V9_APPLY_LEARNING_BASE(evaluation, store, calculation, mode)
    probability = float(item.get('p_final_system', item.get('p_final') or 0.0))
    action, status, stake = normalized_action(
        probability, item.get('blockers', []), mode, item.get('history', {}).get('p_hist')
    )
    # STRONG is reserved for clean FULL_STAT only. No-stat may still be a normal
    # PLAY at 75-79 when support score is 5-6/6, exactly as requested.
    if status == 'STRONG PLAY':
        clean = bool(
            item.get('data_mode') == 'FULL_STAT'
            and item.get('stat_comparison', {}).get('stat_gate_status') == 'CONFIRMED'
            and not item.get('caps')
            and not item.get('blockers')
        )
        if not clean:
            status, stake = 'MAIN PLAY', '30-35% live-limit'
    item['system_action'] = action
    item['system_status'] = status
    item['stake'] = stake
    return item


def summarize_line_evaluation(item: dict[str, Any]) -> dict[str, Any]:
    summary = _V9_SUMMARIZE_BASE(item)
    summary.update({
        'data_mode': item.get('data_mode'),
        'stat_channels': deepcopy(item.get('stat_channels') or {}),
        'n_state': item.get('scenario', {}).get('n_state'),
        'scenario_active': item.get('scenario', {}).get('scenario_active'),
        'required_history_p': item.get('live', {}).get('required_history', {}).get('p_required_history'),
        'required_history_hits_n': [
            item.get('live', {}).get('required_history', {}).get('hits'),
            item.get('live', {}).get('required_history', {}).get('n'),
        ],
        'no_stat_support_score': (item.get('no_stat_support') or {}).get('score'),
        'no_stat_support_checks': deepcopy((item.get('no_stat_support') or {}).get('checks') or {}),
        'projection_formula_mode': item.get('live', {}).get('projection_formula_mode'),
        'projection_formula': item.get('live', {}).get('projection_formula'),
        'weights': deepcopy(item.get('weights', {}).get('normalized') or {}),
    })
    return summary



# ===== v9.1 DIRECTIONAL STAT-GATE RESOLUTION =====
# A 3/5 rule can occasionally mark both Over and Under evidence as present
# (for example high FTA together with high TO and low ORB).  Such a snapshot
# is mixed evidence, not two simultaneous confirmations.  Resolve the gate
# per candidate side before caps/blockers are applied.
_V91_CALCULATE_STAT_GATE_BASE = calculate_stat_gate


def calculate_stat_gate(
    market: dict[str, Any],
    canonical: dict[str, Any],
    zones_data: Optional[dict[str, Any]],
    *,
    project_counts_to_scope_end: bool=True,
) -> dict[str, Any]:
    result = _V91_CALCULATE_STAT_GATE_BASE(
        market, canonical, zones_data,
        project_counts_to_scope_end=project_counts_to_scope_end,
    )
    mode = str(result.get('data_mode') or canonical.get('data_mode') or '')
    if mode in {'SCORE_TIME_HISTORY', 'DATA_OFF'}:
        return result

    over_count = len(set(result.get('over_positive_channels') or []))
    under_count = len(set(result.get('under_positive_channels') or []))
    side = str(market.get('side') or '').upper()
    candidate_count = over_count if side == 'OVER' else under_count
    opposite_count = under_count if side == 'OVER' else over_count
    fake_candidate = bool(result.get('fake_over')) if side == 'OVER' else bool(result.get('fake_under'))

    if fake_candidate:
        status = 'AGAINST'
        reason = 'FAKE_DIRECTION_PROFILE'
    elif candidate_count >= 3 and opposite_count < 3:
        status = 'CONFIRMED'
        reason = 'CANDIDATE_3_OF_5_ONLY'
    elif opposite_count >= 3 and candidate_count < 3:
        status = 'AGAINST'
        reason = 'OPPOSITE_3_OF_5_ONLY'
    elif candidate_count >= 3 and opposite_count >= 3:
        status = 'CONFLICT'
        reason = 'BOTH_DIRECTIONS_3_OF_5_MIXED'
    else:
        status = 'NEUTRAL'
        reason = 'NO_DIRECTION_REACHES_3_OF_5'

    result['stat_gate_status'] = status
    result['stat_direction_resolution'] = {
        'candidate_side': side,
        'candidate_confirmations': candidate_count,
        'opposite_confirmations': opposite_count,
        'over_confirmations': over_count,
        'under_confirmations': under_count,
        'fake_candidate': fake_candidate,
        'reason': reason,
    }
    if mode == 'PARTIAL_STAT':
        result['partial_independent_confirmations'] = candidate_count
        result['partial_opposite_confirmations'] = opposite_count
    return result


# Publish the final version in both CLI metadata and calculation output.
DEFAULT_CONFIG['engine_version'] = '9.1.0-FINAL-HYBRID-PARITY'
_V9_FORMULA_REGISTRY['stat_gate_resolution'] = (
    'Candidate CONFIRMED only when its direction has >=3 independent channels and the opposite has <3; '
    'opposite-only => AGAINST; both >=3 => CONFLICT; neither => NEUTRAL; fake candidate => AGAINST.'
)



# ===== v10.0 CORE LIVE-PROJECTION PARITY OVERRIDES =====
# Replaces the earlier possessions*PPP approximation with the full project CORE:
# rho_stage -> live/pre shot-volume regression -> remaining attempts -> remaining
# points -> tail adjustment -> PreFinal + rho*(LiveRaw-PreFinal).
# NO_STAT keeps its own score/time/history/scenario formula.

SYSTEM_VERSION = '10.0.0'
DEFAULT_CONFIG['engine_version'] = '10.0.0-CORE-LIVE-PROJECTION'
DEFAULT_CONFIG.setdefault('projection', {}).setdefault('core_live', {}).update({
    'gamma_last5': 0.25,
    'k_mean': 10.0,
    'k_rate': 12.0,
    'own_weight': 0.55,
    'opponent_allowed_weight': 0.45,
    'orb_bonus_factor': 0.70,
    'to_drag_factor': 0.80,
    'partial_sigma_multiplier': 1.10,
    'no_stat_sigma_multiplier': 1.20,
    'projection_conflict_sigma_multiplier': 1.20,
    'match_projection_conflict_points': 15.0,
    'team_it_projection_conflict_points': 7.0,
})

_V10_CALCULATE_BASE = SuperBasketCalculator.calculate
_V10_EVALUATE_BASE = SuperBasketCalculator.evaluate_market
_V10_SUMMARIZE_BASE = summarize_line_evaluation


def _v10_scope_indices(market: dict[str, Any]) -> list[int]:
    segment = str(market.get('segment') or 'MATCH').upper()
    if segment == 'H1':
        return [0, 1]
    if segment == 'H2':
        return [2, 3]
    if segment.startswith('Q') and segment[1:].isdigit():
        q = int(segment[1:])
        return [q - 1] if 1 <= q <= 4 else []
    return [0, 1, 2, 3]


def _v10_scope_minutes(market: dict[str, Any], canonical: dict[str, Any], game: Optional[dict[str, Any]]=None) -> float:
    indices = _v10_scope_indices(market)
    q_minutes = to_number(((game or {}).get('format') or {}).get('quarter_minutes'))
    if q_minutes is None:
        q_minutes = float(canonical.get('quarter_minutes') or 10.0)
    return max(1.0, len(indices) * float(q_minutes))


def _v10_segment_stats(game: dict[str, Any], side: str, market: dict[str, Any]) -> dict[str, Optional[float]]:
    indices = _v10_scope_indices(market)
    metrics = ('FGA', 'FGM', '2PA', '2PM', '3PA', '3PM', 'FTA', 'FTM', 'ORB', 'DRB', 'TO', 'FOULS')
    if indices == [0, 1, 2, 3]:
        source = (game.get('stats') or {}).get(side) or {}
        return {metric: to_number(source.get(metric)) for metric in metrics}
    rows = ((game.get('quarter_stats') or {}).get(side) or [])
    output: dict[str, Optional[float]] = {}
    for metric in metrics:
        values = [to_number(rows[index].get(metric)) for index in indices if index < len(rows)]
        output[metric] = sum(float(value) for value in values) if values and all(value is not None for value in values) else None
    return output


def _v10_decay_weights(n: int, gamma: float=0.25) -> list[float]:
    if n <= 0:
        return []
    raw = [math.exp(-gamma * index) for index in range(n)]
    total = sum(raw)
    return [value / total for value in raw] if total else [1.0 / n] * n


def _v10_form_adjusted(values: list[float], *, k: float, gamma: float=0.25) -> tuple[Optional[float], dict[str, Any]]:
    clean = [float(value) for value in values if value is not None and math.isfinite(float(value))]
    if not clean:
        return None, {'n': 0, 'long_mean': None, 'last5_weighted': None, 'credibility': 0.0}
    long_mean = statistics.mean(clean)
    last = clean[:5]
    weights = _v10_decay_weights(len(last), gamma)
    last5_weighted = sum(value * weight for value, weight in zip(last, weights))
    n_eff = 1.0 / sum(weight * weight for weight in weights) if weights else 0.0
    credibility = n_eff / (n_eff + float(k)) if n_eff > 0 else 0.0
    adjusted = long_mean + credibility * (last5_weighted - long_mean)
    return adjusted, {
        'n': len(clean),
        'long_mean': long_mean,
        'last5_weighted': last5_weighted,
        'last5_weights': weights,
        'n_eff': n_eff,
        'credibility': credibility,
        'k': float(k),
    }


def _v10_row_profile(stats: dict[str, Optional[float]], duration_minutes: float) -> dict[str, Optional[float]]:
    fga = to_number(stats.get('FGA'))
    fgm = to_number(stats.get('FGM'))
    two_pa = to_number(stats.get('2PA'))
    two_pm = to_number(stats.get('2PM'))
    three_pa = to_number(stats.get('3PA'))
    three_pm = to_number(stats.get('3PM'))
    fta = to_number(stats.get('FTA'))
    ftm = to_number(stats.get('FTM'))
    orb = to_number(stats.get('ORB'))
    drb = to_number(stats.get('DRB'))
    turnovers = to_number(stats.get('TO'))
    fouls = to_number(stats.get('FOULS'))
    if two_pa is None and fga is not None and three_pa is not None:
        two_pa = max(0.0, fga - three_pa)
    if two_pm is None and fgm is not None and three_pm is not None:
        two_pm = max(0.0, fgm - three_pm)
    return {
        'FGApm': safe_div(fga, duration_minutes),
        'FTApm': safe_div(fta, duration_minutes),
        'ORBpm': safe_div(orb, duration_minutes),
        'DRBpm': safe_div(drb, duration_minutes),
        'TOpm': safe_div(turnovers, duration_minutes),
        'FOULpm': safe_div(fouls, duration_minutes),
        'Share2': safe_div(two_pa, fga),
        'Share3': safe_div(three_pa, fga),
        'P2': safe_div(two_pm, two_pa),
        'P3': safe_div(three_pm, three_pa),
        'PFT': safe_div(ftm, fta),
        'FTr': safe_div(fta, fga),
    }


def _v10_history_stat_profile(
    pool: list[dict[str, Any]],
    market: dict[str, Any],
    canonical: dict[str, Any],
    *,
    allowed: bool=False,
) -> dict[str, Any]:
    cfg = DEFAULT_CONFIG['projection']['core_live']
    rows: list[dict[str, Optional[float]]] = []
    for game in pool:
        perspective = str(game.get('perspective_side') or '')
        if perspective not in {'home', 'away'}:
            continue
        side = ('away' if perspective == 'home' else 'home') if allowed else perspective
        stats = _v10_segment_stats(game, side, market)
        rows.append(_v10_row_profile(stats, _v10_scope_minutes(market, canonical, game)))
    rate_metrics = {'Share2', 'Share3', 'P2', 'P3', 'PFT', 'FTr'}
    adjusted: dict[str, Optional[float]] = {}
    details: dict[str, Any] = {}
    for metric in ('FGApm', 'FTApm', 'ORBpm', 'DRBpm', 'TOpm', 'FOULpm', 'Share2', 'Share3', 'P2', 'P3', 'PFT', 'FTr'):
        values = [row[metric] for row in rows if row.get(metric) is not None]
        adjusted[metric], details[metric] = _v10_form_adjusted(
            [float(value) for value in values],
            k=float(cfg['k_rate'] if metric in rate_metrics else cfg['k_mean']),
            gamma=float(cfg['gamma_last5']),
        )
    return {'values': adjusted, 'details': details, 'n_rows': len(rows), 'allowed_mode': allowed}


def _v10_pre_stat_team(
    side: str,
    market: dict[str, Any],
    canonical: dict[str, Any],
) -> dict[str, Any]:
    cfg = DEFAULT_CONFIG['projection']['core_live']
    own_pool = canonical['history']['team_a'] if side == 'home' else canonical['history']['team_b']
    opponent_pool = canonical['history']['team_b'] if side == 'home' else canonical['history']['team_a']
    own = _v10_history_stat_profile(own_pool, market, canonical, allowed=False)
    allowed = _v10_history_stat_profile(opponent_pool, market, canonical, allowed=True)
    opponent_own = _v10_history_stat_profile(opponent_pool, market, canonical, allowed=False)
    own_w = float(cfg['own_weight'])
    allowed_w = float(cfg['opponent_allowed_weight'])

    def blend(metric: str) -> Optional[float]:
        a = to_number(own['values'].get(metric))
        b = to_number(allowed['values'].get(metric))
        available = [(a, own_w), (b, allowed_w)]
        available = [(value, weight) for value, weight in available if value is not None]
        if not available:
            return None
        denominator = sum(weight for _, weight in available)
        return sum(float(value) * weight for value, weight in available) / denominator

    pre = {metric: blend(metric) for metric in ('FGApm', 'FTApm', 'ORBpm', 'DRBpm', 'TOpm', 'FOULpm', 'Share2', 'Share3', 'P2', 'P3', 'PFT', 'FTr')}
    if pre['Share2'] is not None and pre['Share3'] is not None:
        total_share = pre['Share2'] + pre['Share3']
        if total_share > 0:
            pre['Share2'] /= total_share
            pre['Share3'] /= total_share
    duration = _v10_scope_minutes(market, canonical)
    pre_fga = pre['FGApm'] * duration if pre['FGApm'] is not None else None
    pre_fta = pre['FTApm'] * duration if pre['FTApm'] is not None else None
    pre_2pa = pre_fga * pre['Share2'] if pre_fga is not None and pre['Share2'] is not None else None
    pre_3pa = pre_fga * pre['Share3'] if pre_fga is not None and pre['Share3'] is not None else None
    shot_points = (
        2.0 * pre_2pa * pre['P2'] + 3.0 * pre_3pa * pre['P3']
        if None not in (pre_2pa, pre_3pa, pre['P2'], pre['P3']) else None
    )
    ft_points = pre_fta * pre['PFT'] if pre_fta is not None and pre['PFT'] is not None else None

    pre_orb = pre['ORBpm'] * duration if pre['ORBpm'] is not None else None
    pre_to = pre['TOpm'] * duration if pre['TOpm'] is not None else None
    opp_drb_allowed = allowed['values'].get('DRBpm')
    opp_drb_allowed = float(opp_drb_allowed) * duration if opp_drb_allowed is not None else None
    own_to_baseline = own['values'].get('TOpm')
    own_to_baseline = float(own_to_baseline) * duration if own_to_baseline is not None else None
    orb_bonus = (
        max(0.0, pre_orb - opp_drb_allowed) * float(cfg['orb_bonus_factor'])
        if pre_orb is not None and opp_drb_allowed is not None else 0.0
    )
    to_drag = (
        max(0.0, pre_to - own_to_baseline) * float(cfg['to_drag_factor'])
        if pre_to is not None and own_to_baseline is not None else 0.0
    )
    opp_fouls = opponent_own['values'].get('FOULpm')
    own_fouls = own['values'].get('FOULpm')
    foul_pressure = 0.0
    if opp_fouls is not None and pre['FTr'] is not None and pre['PFT'] is not None:
        baseline_candidates = [float(value) for value in (opp_fouls, own_fouls) if value is not None]
        foul_baseline = statistics.mean(baseline_candidates) if baseline_candidates else float(opp_fouls)
        foul_pressure = max(0.0, (float(opp_fouls) - foul_baseline) * duration) * pre['FTr'] * pre['PFT']

    pre_final = None
    if shot_points is not None and ft_points is not None:
        pre_final = shot_points + ft_points + orb_bonus - to_drag + foul_pressure
    return {
        'side': side,
        'duration_minutes': duration,
        'own_profile': own,
        'opponent_allowed_profile': allowed,
        'opponent_own_profile': opponent_own,
        'pre': pre,
        'Pre_FGA': pre_fga,
        'Pre_FTA': pre_fta,
        'Pre_2PA': pre_2pa,
        'Pre_3PA': pre_3pa,
        'ShotPts': shot_points,
        'FTPts': ft_points,
        'ORB_bonus': orb_bonus,
        'TO_drag': to_drag,
        'FoulPressureAdj': foul_pressure,
        'HCA': None,
        'RestAdj': None,
        'PreFinal': pre_final,
        'formula': 'PreFinal = ShotPts + FTPts + ORB_bonus - TO_drag + FoulPressureAdj; unavailable context terms stay OFF',
    }


def _v10_stage_trust(canonical: dict[str, Any]) -> tuple[float, float, float]:
    minutes_played = max(0.0, float(canonical.get('elapsed_game_seconds') or 0.0) / 60.0)
    stage = str(canonical.get('stage') or '')
    if stage == 'Q4_CONFIRMATION':
        k = 6.0
    elif stage in {'HT', 'AFTER_3Q'} or minutes_played >= 20.0:
        k = 10.0
    else:
        k = 18.0
    rho = minutes_played / (minutes_played + k) if minutes_played > 0 else 0.0
    return rho, k, minutes_played


def _v10_live_stat_team(
    side: str,
    market: dict[str, Any],
    canonical: dict[str, Any],
    pre_profile: dict[str, Any],
    rho: float,
    remaining_minutes: float,
) -> dict[str, Any]:
    cfg = DEFAULT_CONFIG['projection']['core_live']
    stats = canonical.get('live_stats', {}).get(side, {}) or {}
    elapsed_minutes = max(0.0, float(canonical.get('elapsed_game_seconds') or 0.0) / 60.0)
    raw = _v10_row_profile({metric: to_number(stats.get(metric)) for metric in ('FGA','FGM','2PA','2PM','3PA','3PM','FTA','FTM','ORB','DRB','TO','FOULS')}, max(elapsed_minutes, 1e-9))
    pre = pre_profile['pre']

    def reg(metric: str) -> tuple[Optional[float], float]:
        live_value = to_number(raw.get(metric))
        pre_value = to_number(pre.get(metric))
        if live_value is None and pre_value is None:
            return None, 0.0
        if live_value is None:
            return pre_value, 0.0
        if pre_value is None:
            return live_value, 1.0
        return rho * live_value + (1.0 - rho) * pre_value, rho

    adjusted: dict[str, Optional[float]] = {}
    metric_rho: dict[str, float] = {}
    for metric in ('FGApm','FTApm','ORBpm','TOpm','Share2','Share3','P2','P3','PFT'):
        adjusted[metric], metric_rho[metric] = reg(metric)
    if adjusted['Share2'] is not None and adjusted['Share3'] is not None:
        total_share = adjusted['Share2'] + adjusted['Share3']
        if total_share > 0:
            adjusted['Share2'] /= total_share
            adjusted['Share3'] /= total_share

    rem_fga = adjusted['FGApm'] * remaining_minutes if adjusted['FGApm'] is not None else None
    rem_fta = adjusted['FTApm'] * remaining_minutes if adjusted['FTApm'] is not None else None
    rem_2pa = rem_fga * adjusted['Share2'] if rem_fga is not None and adjusted['Share2'] is not None else None
    rem_3pa = rem_fga * adjusted['Share3'] if rem_fga is not None and adjusted['Share3'] is not None else None
    rem_pts = (
        2.0 * rem_2pa * adjusted['P2']
        + 3.0 * rem_3pa * adjusted['P3']
        + rem_fta * adjusted['PFT']
        if None not in (rem_2pa, rem_3pa, rem_fta, adjusted['P2'], adjusted['P3'], adjusted['PFT']) else None
    )

    pre_orb_pm = to_number(pre.get('ORBpm'))
    pre_to_pm = to_number(pre.get('TOpm'))
    orb_bonus_live = (
        max(0.0, float(adjusted['ORBpm']) - float(pre_orb_pm)) * remaining_minutes * float(cfg['orb_bonus_factor'])
        if adjusted['ORBpm'] is not None and pre_orb_pm is not None else 0.0
    )
    to_drag_live = (
        max(0.0, float(adjusted['TOpm']) - float(pre_to_pm)) * remaining_minutes * float(cfg['to_drag_factor'])
        if adjusted['TOpm'] is not None and pre_to_pm is not None else 0.0
    )
    # FTApm regression already carries the foul path. Extra tail terms are OFF
    # unless a source supplies an explicit coefficient; this avoids double count.
    foul_tail_adj = 0.0
    endgame_ft_adj = 0.0
    tail_adj = orb_bonus_live - to_drag_live + foul_tail_adj + endgame_ft_adj

    team_name = canonical['home_team'] if side == 'home' else canonical['away_team']
    current_team_points = _current_team_score(canonical, team_name, str(market.get('segment') or 'MATCH'))
    live_raw = current_team_points + rem_pts + tail_adj if rem_pts is not None else None
    pre_final = to_number(pre_profile.get('PreFinal'))
    live_projection_stat = (
        pre_final + rho * (live_raw - pre_final)
        if pre_final is not None and live_raw is not None else live_raw if live_raw is not None else pre_final
    )
    return {
        'side': side,
        'team': team_name,
        'rho_stage': rho,
        'live_raw_metrics': raw,
        'pre_metrics': pre,
        'adjusted_metrics': adjusted,
        'metric_rho': metric_rho,
        'RemFGA': rem_fga,
        'RemFTA': rem_fta,
        'Rem2PA': rem_2pa,
        'Rem3PA': rem_3pa,
        'RemPts': rem_pts,
        'ORB_bonus_live': orb_bonus_live,
        'TO_drag_live': to_drag_live,
        'FoulTailAdj': foul_tail_adj,
        'EndgameFTAdj': endgame_ft_adj,
        'TailAdj': tail_adj,
        'CurrentTeamPoints': current_team_points,
        'LiveRaw_Team': live_raw,
        'PreFinal_Team': pre_final,
        'LiveProjection_stat_Team': live_projection_stat,
        'tail_policy': 'Only residual ORB/TO anomalies are applied; FTA path is already in RemFTA; unavailable explicit tail coefficients stay OFF.',
    }


def _v10_historical_remaining_projection(
    market: dict[str, Any],
    canonical: dict[str, Any],
) -> tuple[Optional[float], dict[str, Any]]:
    indices = _v9_segment_indices(market)
    elapsed = float(canonical.get('elapsed_game_seconds') or 0.0)
    q_seconds = float(canonical.get('quarter_seconds') or 600.0)
    clock = _segment_clock(market, canonical)
    current = float(clock.get('current_points') or 0.0)
    market_type = str(market.get('market_type') or '')
    if market_type.startswith('TEAM_IT') or market_type == 'CURRENT_QUARTER_TEAM_IT':
        team = str(market.get('team') or '')
        own_pool = canonical['history']['team_a'] if team == canonical['home_team'] else canonical['history']['team_b']
        opponent_pool = canonical['history']['team_b'] if team == canonical['home_team'] else canonical['history']['team_a']
        own_values: list[float] = []
        for game in own_pool:
            remaining = _v9_remaining_from_values(
                [to_number(value) for value in (game.get('team_quarters') or [])[:4]],
                indices, elapsed, q_seconds,
            )
            if remaining is not None:
                own_values.append(float(remaining))
        allowed_values: list[float] = []
        for game in opponent_pool:
            remaining = _v9_remaining_from_values(
                [to_number(value) for value in (game.get('opponent_quarters') or [])[:4]],
                indices, elapsed, q_seconds,
            )
            if remaining is not None:
                allowed_values.append(float(remaining))
        h2h_values: list[float] = []
        for game in canonical['history'].get('h2h', []):
            quarters = _v9_game_team_quarters(game, team)
            remaining = _v9_remaining_from_values(quarters, indices, elapsed, q_seconds) if quarters else None
            if remaining is not None:
                h2h_values.append(float(remaining))
        medians = {
            'own': statistics.median(own_values) if own_values else None,
            'opponent_allowed': statistics.median(allowed_values) if allowed_values else None,
            'h2h': statistics.median(h2h_values) if h2h_values else None,
        }
        remaining, weights = _weighted_available(medians, {'own': 0.50, 'opponent_allowed': 0.35, 'h2h': 0.15})
        return (
            current + remaining if remaining is not None else None,
            {'current': current, 'remaining_medians': medians, 'weights': weights, 'method': 'Team IT historical median remaining 50/35/15'},
        )
    values: list[float] = []
    for pool_name in ('team_a', 'team_b'):
        for game in canonical['history'].get(pool_name, []):
            remaining = _v9_remaining_from_values(
                _v9_game_total_quarters(game), indices, elapsed, q_seconds,
            )
            if remaining is not None:
                values.append(float(remaining))
    median_remaining = statistics.median(values) if values else None
    return (
        current + median_remaining if median_remaining is not None else None,
        {'current': current, 'median_remaining': median_remaining, 'n': len(values), 'method': 'Pooled historical median remaining'},
    )



def _v10_weighted_median(values: list[tuple[float, float]]) -> Optional[float]:
    clean = sorted((float(value), max(0.0, float(weight))) for value, weight in values if value is not None and weight is not None)
    if not clean:
        return None
    total = sum(weight for _, weight in clean)
    if total <= 0:
        return statistics.median(value for value, _ in clean)
    threshold = total / 2.0
    running = 0.0
    for value, weight in clean:
        running += weight
        if running >= threshold:
            return value
    return clean[-1][0]


def _v10_scenario_projection_center(scenario: dict[str, Any]) -> tuple[Optional[float], dict[str, Any]]:
    """Return a candidate-side-independent outcome center for the current state.

    Probabilities in P_scenario are side-specific; the projection of final points must not be.
    Therefore the center uses only state-pattern outcome medians and weights that do not
    depend on whether the candidate is OVER or UNDER.
    """
    projection_groups = {
        'quarter_result', 'score_state', 'margin_state', 'total_state',
        'sequence_state', 'time_state',
    }
    selected: dict[tuple[str, str], dict[str, Any]] = {}
    for item in scenario.get('patterns_found', []) or []:
        group = str(item.get('pattern_group') or '')
        if group not in projection_groups:
            continue
        median = to_number(item.get('outcome_median'))
        n = int(item.get('matched_games') or 0)
        if median is None or n < 3:
            continue
        team = str(item.get('team') or '')
        specificity = max(0.10, float(item.get('specificity') or 0.50))
        distance = max(0.10, float(item.get('distance_match_quality') or 1.00))
        weight = n * specificity * distance
        key = (team, group)
        current = selected.get(key)
        if current is None or weight > float(current['weight']):
            selected[key] = {
                'team': team,
                'group': group,
                'outcome_median': float(median),
                'matched_games': n,
                'weight': weight,
                'pattern_id': item.get('pattern_id'),
            }
    weighted = [(row['outcome_median'], row['weight']) for row in selected.values()]
    center = _v10_weighted_median(weighted)
    return center, {
        'method': 'candidate-independent weighted median of matched state outcome medians',
        'selected_patterns': list(selected.values()),
        'n_selected': len(selected),
        'excluded_groups': sorted(set(_V9_STATE_GROUPS) - projection_groups),
    }


def _v10_candidate_independent_baselines(
    market: dict[str, Any],
    canonical: dict[str, Any],
    scenario: dict[str, Any],
) -> dict[str, Any]:
    clock = _segment_clock(market, canonical)
    current = float(clock['current_points'])
    simple = current / clock['elapsed_seconds'] * clock['full_seconds'] if clock['elapsed_seconds'] > 0 else None
    history_values = _history_values(market, canonical)
    historical_total_median = statistics.median(history_values) if history_values else None
    historical_rate = historical_total_median / clock['full_seconds'] if historical_total_median is not None and clock['full_seconds'] else None
    previous_rate = _previous_quarter_pace(market, canonical, clock)
    current_rate = current / clock['elapsed_seconds'] if clock['elapsed_seconds'] > 0 else None
    rates = [('current', current_rate, 0.45), ('previous', previous_rate, 0.25), ('history', historical_rate, 0.30)]
    available = [(name, value, weight) for name, value, weight in rates if value is not None]
    denominator = sum(weight for _, _, weight in available)
    regressed_rate = sum(float(value) * weight for _, value, weight in available) / denominator if denominator else None
    projection_regressed = current + regressed_rate * clock['remaining_seconds'] if regressed_rate is not None else None

    projection_history, history_remaining_details = _v10_historical_remaining_projection(market, canonical)
    scenario_active = bool(scenario.get('scenario_active'))
    scenario_center, scenario_center_details = _v10_scenario_projection_center(scenario) if scenario_active else (None, {'method': 'OFF'})
    projection_scenario = scenario_center if scenario_center is not None and scenario_center >= current else None

    control_candidates = [value for value in (projection_history, projection_scenario, projection_regressed) if value is not None]
    projection_control = statistics.median(control_candidates) if control_candidates else simple
    anchor_candidates = [value for value in (projection_history, projection_scenario) if value is not None]
    history_scenario_anchor = statistics.median(anchor_candidates) if anchor_candidates else projection_history
    return {
        'clock': clock,
        'projection_simple': simple,
        'projection_regressed': projection_regressed,
        'projection_segment': projection_regressed,
        'projection_history': projection_history,
        'projection_scenario': projection_scenario,
        'projection_control': projection_control,
        'projection_history_scenario_anchor': history_scenario_anchor,
        'scenario_active': scenario_active,
        'scenario_projection_method': 'CANDIDATE_INDEPENDENT_MATCHED_STATE_CENTER' if projection_scenario is not None else 'OFF',
        'scenario_projection_details': scenario_center_details,
        'regressed_rate_details': {
            'rates': {name: value for name, value, _ in rates},
            'normalized_weights': {name: weight / denominator for name, value, weight in available} if denominator else {},
        },
        'history_remaining_details': history_remaining_details,
    }


def calculate_live_projection(
    market: dict[str, Any],
    canonical: dict[str, Any],
    history: dict[str, Any],
    scenario: dict[str, Any],
    config: dict[str, Any],
    stat: Optional[dict[str, Any]]=None,
) -> dict[str, Any]:
    mode = str(canonical.get('data_mode') or canonical.get('data_gate', {}).get('data_mode') or 'DATA_OFF')
    base = _v10_candidate_independent_baselines(market, canonical, scenario)
    clock = base['clock']
    line = float(market['line'])
    side = str(market['side']).upper()
    sigma_base = _stage_sigma(market['market_type'], canonical['stage'], config)
    projection_stat = None
    core_stat: dict[str, Any] = {}
    projection_conflict = False
    conflict_threshold = None
    no_stat_components: dict[str, Optional[float]] = {}
    no_stat_weights: dict[str, float] = {}

    if mode in {'FULL_STAT', 'PARTIAL_STAT'}:
        rho, k_stage, minutes_played = _v10_stage_trust(canonical)
        remaining_minutes = float(clock['remaining_seconds']) / 60.0
        pre_home = _v10_pre_stat_team('home', market, canonical)
        pre_away = _v10_pre_stat_team('away', market, canonical)
        live_home = _v10_live_stat_team('home', market, canonical, pre_home, rho, remaining_minutes)
        live_away = _v10_live_stat_team('away', market, canonical, pre_away, rho, remaining_minutes)
        if market.get('team'):
            selected = live_home if market['team'] == canonical['home_team'] else live_away
            projection_stat = to_number(selected.get('LiveProjection_stat_Team'))
        else:
            a = to_number(live_home.get('LiveProjection_stat_Team'))
            b = to_number(live_away.get('LiveProjection_stat_Team'))
            projection_stat = a + b if a is not None and b is not None else None
        core_stat = {
            'rho_stage': rho,
            'K_stage': k_stage,
            'minutes_played': minutes_played,
            'minutes_left_scope': remaining_minutes,
            'pre_home': pre_home,
            'pre_away': pre_away,
            'live_home': live_home,
            'live_away': live_away,
            'LiveProjection_stat_Total': (
                to_number(live_home.get('LiveProjection_stat_Team')) + to_number(live_away.get('LiveProjection_stat_Team'))
                if to_number(live_home.get('LiveProjection_stat_Team')) is not None
                and to_number(live_away.get('LiveProjection_stat_Team')) is not None else None
            ),
            'LiveProjection_stat_Margin': (
                to_number(live_home.get('LiveProjection_stat_Team')) - to_number(live_away.get('LiveProjection_stat_Team'))
                if to_number(live_home.get('LiveProjection_stat_Team')) is not None
                and to_number(live_away.get('LiveProjection_stat_Team')) is not None else None
            ),
            'formula': 'PreFinal_Team + rho_stage*(LiveRaw_Team-PreFinal_Team)',
        }
        conservative_components = [
            value for value in (
                projection_stat,
                base.get('projection_control'),
                base.get('projection_history_scenario_anchor'),
            ) if value is not None
        ]
        projection_used = statistics.median(conservative_components) if conservative_components else line
        conflict_threshold = (
            float(DEFAULT_CONFIG['projection']['core_live']['team_it_projection_conflict_points'])
            if market.get('team') else
            float(DEFAULT_CONFIG['projection']['core_live']['match_projection_conflict_points'])
        )
        control = to_number(base.get('projection_control'))
        if projection_stat is not None and control is not None:
            projection_conflict = abs(projection_stat - control) > conflict_threshold
        sigma = sigma_base * (
            float(DEFAULT_CONFIG['projection']['core_live']['partial_sigma_multiplier'])
            if mode == 'PARTIAL_STAT' else 1.0
        )
        if projection_conflict and (stat or {}).get('stat_gate_status') != 'CONFIRMED':
            sigma *= float(DEFAULT_CONFIG['projection']['core_live']['projection_conflict_sigma_multiplier'])
        formula_mode = 'FULL_STAT_CORE_SHOT_VOLUME' if mode == 'FULL_STAT' else 'PARTIAL_STAT_CORE_AVAILABLE_FIELDS'
        formula_text = 'median(LiveProjection_stat, Projection_control, HistoryScenarioAnchor); simple pace diagnostic only'
    elif mode == 'SCORE_TIME_HISTORY':
        regressed = to_number(base.get('projection_regressed'))
        history_projection = to_number(base.get('projection_history'))
        scenario_projection = to_number(base.get('projection_scenario')) if base.get('scenario_active') else None
        n_state = int(scenario.get('n_state') or 0)
        if n_state >= 5 and scenario_projection is not None:
            configured = {'regressed': 0.35, 'history': 0.35, 'scenario': 0.30}
            formula_text = '0.35*Projection_regressed + 0.35*Projection_history + 0.30*Projection_scenario'
        elif 3 <= n_state <= 4 and scenario_projection is not None:
            configured = {'regressed': 0.40, 'history': 0.45, 'scenario': 0.15}
            formula_text = '0.40*Projection_regressed + 0.45*Projection_history + 0.15*Projection_scenario'
        else:
            configured = {'regressed': 0.50, 'history': 0.50}
            formula_text = '0.50*Projection_regressed + 0.50*Projection_history'
        values = {'regressed': regressed, 'history': history_projection, 'scenario': scenario_projection}
        no_stat_components = dict(values)
        no_stat_weights = dict(configured)
        available = {key: value for key, value in values.items() if key in configured and value is not None}
        denominator = sum(configured[key] for key in available)
        projection_used = (
            sum(float(available[key]) * configured[key] for key in available) / denominator
            if denominator > 0 else line
        )
        sigma = sigma_base * float(DEFAULT_CONFIG['projection']['core_live']['no_stat_sigma_multiplier'])
        formula_mode = 'SCORE_TIME_HISTORY_NO_STAT'
        core_stat = {'status': 'OFF_NO_STATS', 'unknown_fields_not_zero': True}
    else:
        projection_used = line
        sigma = sigma_base
        formula_mode = 'DATA_OFF'
        formula_text = 'DATA_OFF'
        core_stat = {'status': 'OFF_DATA'}

    line_edge_over = projection_used - line
    line_edge_under = line - projection_used
    line_edge = line_edge_over if side == 'OVER' else line_edge_under
    z_score = line_edge / sigma if sigma > 0 else 0.0
    p_live = normal_cdf(z_score) if sigma > 0 else 0.50
    parser_components = _parser_projection_components(market, canonical, clock)
    components = {
        'projection_simple': {'value': base.get('projection_simple'), 'included': False, 'role': 'diagnostic_only'},
        'projection_regressed': {'value': base.get('projection_regressed'), 'included': mode == 'SCORE_TIME_HISTORY'},
        'projection_history': {'value': base.get('projection_history'), 'included': True},
        'projection_scenario': {'value': base.get('projection_scenario'), 'included': bool(base.get('scenario_active'))},
        'projection_control': {'value': base.get('projection_control'), 'included': mode in {'FULL_STAT', 'PARTIAL_STAT'}},
        'projection_history_scenario_anchor': {'value': base.get('projection_history_scenario_anchor'), 'included': mode in {'FULL_STAT', 'PARTIAL_STAT'}},
        'projection_stat_adjusted': {'value': projection_stat, 'included': mode in {'FULL_STAT', 'PARTIAL_STAT'}},
    }
    for key, value in parser_components.items():
        components[key] = {**value, 'included': False, 'role': 'audit_only_not_used_in_projection'}
    return {
        'clock': canonical.get('clock'),
        'elapsed_seconds': clock['elapsed_seconds'],
        'remaining_seconds': clock['remaining_seconds'],
        'elapsed_game_seconds': canonical['elapsed_game_seconds'],
        'remaining_game_seconds': canonical['remaining_game_seconds'],
        'current_points': clock['current_points'],
        'data_mode': mode,
        'components': components,
        'projection_simple': base.get('projection_simple'),
        'projection_regressed': base.get('projection_regressed'),
        'projection_segment': base.get('projection_segment'),
        'projection_model_live': base.get('projection_regressed'),
        'projection_history': base.get('projection_history'),
        'projection_scenario': base.get('projection_scenario'),
        'scenario_projection_method': base.get('scenario_projection_method'),
        'projection_stat_adjusted': projection_stat,
        'projection_control': base.get('projection_control'),
        'projection_history_scenario_anchor': base.get('projection_history_scenario_anchor'),
        'projection_used': projection_used,
        'Projection_used': projection_used,
        'line': line,
        'line_edge': line_edge,
        'line_edge_over': line_edge_over,
        'line_edge_under': line_edge_under,
        'sigma_base': sigma_base,
        'sigma': sigma,
        'z_score': z_score,
        'p_live': p_live,
        'required_history': _v9_required_history(market, canonical),
        'projection_formula_mode': formula_mode,
        'projection_formula': formula_text,
        'no_stat_components': no_stat_components,
        'no_stat_configured_weights': no_stat_weights,
        'projection_conflict': projection_conflict,
        'projection_conflict_threshold': conflict_threshold,
        'stat_projection_details': {
            'core_stat_projection': core_stat,
            'baseline_details': base,
            'parser_projection_audit': parser_components,
        },
        'core_live_formula_version': 'v3.4_CORE_FULL_WORK_FORMULAS',
    }


def _v10_evaluate_market(self: SuperBasketCalculator, market: dict[str, Any], canonical: dict[str, Any]) -> dict[str, Any]:
    item = _V10_EVALUATE_BASE(self, market, canonical)
    live = item.get('live') or {}
    if live.get('projection_conflict') and item.get('stat_comparison', {}).get('stat_gate_status') == 'CONFIRMED':
        cap = 0.79
        if not any(rule.get('rule_id') == 'CORE_PROJECTION_DIVERGENCE_CAP' for rule in item.get('caps', [])):
            item.setdefault('caps', []).append(_cap(
                'CORE_PROJECTION_DIVERGENCE_CAP',
                cap,
                'Stat and checkpoint projections diverge beyond the project threshold; conservative P_live is required',
                {
                    'threshold': live.get('projection_conflict_threshold'),
                    'projection_stat': live.get('projection_stat_adjusted'),
                    'projection_control': live.get('projection_control'),
                },
            ))
        item['p_final'] = min(float(item.get('p_final') or 0.0), cap)
        clean = bool(
            not item.get('blockers')
            and item.get('data_mode') == 'FULL_STAT'
            and item.get('stat_comparison', {}).get('stat_gate_status') == 'CONFIRMED'
            and not item.get('caps')
        )
        item['verdict'] = _verdict(
            float(item['p_final']), item.get('blockers', []), clean, item.get('history', {}).get('p_hist')
        )
    item['core_live_projection_version'] = '10.0.0'
    return item


SuperBasketCalculator.evaluate_market = _v10_evaluate_market


_V10_FORMULA_REGISTRY = deepcopy(_V9_FORMULA_REGISTRY)
_V10_FORMULA_REGISTRY.update({
    'core_stat_stage_trust': 'rho_stage = minutes_played/(minutes_played+K_stage), K=18 early/1H, 10 HT/late Q3, 6 Q4.',
    'core_stat_live_regression': (
        'FGApm/FTApm/ORBpm/TOpm, Share2/Share3 and 2P/3P/FT percentages are blended: '
        'adjusted = rho*live + (1-rho)*pre.'
    ),
    'core_stat_remaining': (
        'RemFGA=FGApm_adj*MinutesLeft; RemFTA=FTApm_adj*MinutesLeft; '
        'Rem2PA=RemFGA*Share2_adj; Rem3PA=RemFGA*Share3_adj; '
        'RemPts=2*Rem2PA*P2_adj+3*Rem3PA*P3_adj+RemFTA*PFT_adj.'
    ),
    'core_stat_final': (
        'LiveRaw_Team=CurrentTeamPoints+RemPts+TailAdj; '
        'LiveProjection_stat_Team=PreFinal_Team+rho_stage*(LiveRaw_Team-PreFinal_Team).'
    ),
    'core_projection_used': (
        'FULL/PARTIAL: median(LiveProjection_stat, Projection_control, HistoryScenarioAnchor). '
        'Simple pace and parser projections are audit-only. NO_STAT uses its separate 35/35/30, 40/45/15 or 50/50 formula.'
    ),
})


def _v10_calculate(self: SuperBasketCalculator, source: dict[str, Any], dispatch_threshold: Optional[float]=None, strict_schema: bool=False) -> dict[str, Any]:
    output = _V10_CALCULATE_BASE(self, source, dispatch_threshold, strict_schema)
    calculation = output.get('super_basket_calculation', {})
    calculation['engine_version'] = '10.0.0-CORE-LIVE-PROJECTION'
    calculation['formula_registry'] = deepcopy(_V10_FORMULA_REGISTRY)
    calculation['live_projection_parity'] = {
        'FULL_STAT': 'FULL CORE shot/volume formula',
        'PARTIAL_STAT': 'same formula with missing live values regressed to pre baseline and partial caps',
        'NO_STAT': 'score/time/history/scenario formula with sigma*1.20',
        'candidate_independent_projection': True,
        'simple_pace_used_for_signal': False,
        'parser_projection_used_for_signal': False,
    }
    output['super_basket_calculation'] = calculation
    return output


SuperBasketCalculator.calculate = _v10_calculate


def summarize_line_evaluation(item: dict[str, Any]) -> dict[str, Any]:
    summary = _V10_SUMMARIZE_BASE(item)
    live = item.get('live', {})
    core = (live.get('stat_projection_details') or {}).get('core_stat_projection') or {}
    summary.update({
        'core_live_formula_version': live.get('core_live_formula_version'),
        'rho_stage': core.get('rho_stage'),
        'K_stage': core.get('K_stage'),
        'projection_history_scenario_anchor': live.get('projection_history_scenario_anchor'),
        'projection_stat_adjusted': live.get('projection_stat_adjusted'),
        'projection_control': live.get('projection_control'),
        'projection_regressed': live.get('projection_regressed'),
        'projection_history': live.get('projection_history'),
        'projection_scenario': live.get('projection_scenario'),
        'no_stat_components': live.get('no_stat_components'),
        'no_stat_configured_weights': live.get('no_stat_configured_weights'),
        'projection_conflict': live.get('projection_conflict'),
        'projection_conflict_threshold': live.get('projection_conflict_threshold'),
        'simple_pace_included': bool((live.get('components') or {}).get('projection_simple', {}).get('included')),
        'parser_projection_included': any(
            bool(value.get('included'))
            for key, value in (live.get('components') or {}).items()
            if key.startswith('projection_parser_')
        ),
    })
    return summary



# ===== v10.1 FRESH-HISTORY + CONFLICT-SAFE SIGNAL POLICY =====
# User policy:
# - PLAY from 75%; RISK from 65%; below 65% PASS.
# - normal signal requires raw exact-line history zone >=75% and live edge >=3 points.
# - serious global conflicts are hard PASS, never RISK.
# - history older than 30 days is not discarded; its probability and projection influence
#   decay smoothly toward neutral/current-live values.

_V101_ADAPT_BASE = adapt_match
_V101_HISTORY_BASE = calculate_history
_V101_SCENARIO_BASE = calculate_scenario
_V101_LIVE_BASE = calculate_live_projection
_V101_EVALUATE_BASE = SuperBasketCalculator.evaluate_market
_V101_CALCULATE_BASE = SuperBasketCalculator.calculate
_V101_SUMMARIZE_BASE = summarize_line_evaluation

DEFAULT_CONFIG['engine_version'] = '10.1.0-FRESH-HISTORY-CONFLICT-SAFE'
DEFAULT_CONFIG['dispatch_threshold'] = 0.65
DEFAULT_CONFIG.setdefault('signal_gates', {}).update({
    'history_zone_min': 0.75,
    'live_edge_min_points': 3.0,
    'risk_min': 0.65,
    'play_min': 0.75,
    'serious_conflicts_hard_block': True,
})
DEFAULT_CONFIG.setdefault('history_freshness', {}).update({
    'grace_days': 30.0,
    'half_life_days': 90.0,
    'minimum_factor': 0.05,
    'live_sigma_max_extra': 0.25,
})


def _v101_datetime_from_source(source: dict[str, Any]) -> Optional[datetime]:
    candidates = [
        (source.get('meta') or {}).get('generated_at'),
        source.get('generated_at'),
        source.get('snapshot_at'),
        (source.get('match') or {}).get('date'),
        ((source.get('bookmaker_lines') or {}).get('_source_meta') or {}).get('fetchedAt'),
        (source.get('_source_meta') or {}).get('fetchedAt'),
    ]
    for value in candidates:
        parsed = _parse_history_date(value)
        if parsed is not None:
            return parsed
    return None


def _v101_parse_date(value: Any) -> Optional[datetime]:
    parsed = _parse_history_date(value)
    if parsed is not None:
        return parsed
    if value in (None, ''):
        return None
    text = str(value).strip()
    for fmt in ('%d.%m.%Y %H:%M:%S', '%d-%m-%Y %H:%M:%S'):
        try:
            return datetime.strptime(text, fmt).replace(tzinfo=timezone.utc)
        except ValueError:
            continue
    return None


def _v101_latest_history_date(games: list[dict[str, Any]]) -> Optional[datetime]:
    dates = [_v101_parse_date(game.get('date')) for game in games or []]
    valid = [value for value in dates if value is not None]
    return max(valid) if valid else None


def _v101_freshness_factor(gap_days: Optional[float], config: dict[str, Any]) -> float:
    cfg = config.get('history_freshness', {})
    if gap_days is None:
        return 1.0
    grace = max(0.0, float(cfg.get('grace_days', 30.0)))
    half_life = max(1.0, float(cfg.get('half_life_days', 90.0)))
    floor = max(0.0, min(1.0, float(cfg.get('minimum_factor', 0.05))))
    if gap_days <= grace:
        return 1.0
    factor = math.exp(-math.log(2.0) * (float(gap_days) - grace) / half_life)
    return max(floor, min(1.0, factor))


def _v101_history_freshness(canonical: dict[str, Any], config: dict[str, Any]) -> dict[str, Any]:
    snapshot_dt = canonical.get('snapshot_datetime')
    pools = canonical.get('history') or {}
    latest_a = _v101_latest_history_date(pools.get('team_a') or [])
    latest_b = _v101_latest_history_date(pools.get('team_b') or [])
    latest_h2h = _v101_latest_history_date(pools.get('h2h') or [])

    def gap(latest: Optional[datetime]) -> Optional[float]:
        if snapshot_dt is None or latest is None:
            return None
        return max(0.0, (snapshot_dt - latest).total_seconds() / 86400.0)

    gap_a, gap_b, gap_h2h = gap(latest_a), gap(latest_b), gap(latest_h2h)
    factor_a = _v101_freshness_factor(gap_a, config)
    factor_b = _v101_freshness_factor(gap_b, config)
    factor_h2h = _v101_freshness_factor(gap_h2h, config)
    known_team_factors = [
        factor for factor, latest in ((factor_a, latest_a), (factor_b, latest_b))
        if latest is not None
    ]
    if len(known_team_factors) == 2:
        match_factor = math.sqrt(known_team_factors[0] * known_team_factors[1])
    elif known_team_factors:
        match_factor = known_team_factors[0]
    else:
        match_factor = 1.0
    return {
        'snapshot_datetime': snapshot_dt.isoformat() if isinstance(snapshot_dt, datetime) else None,
        'team_a_latest': latest_a.isoformat() if latest_a else None,
        'team_b_latest': latest_b.isoformat() if latest_b else None,
        'h2h_latest': latest_h2h.isoformat() if latest_h2h else None,
        'team_a_gap_days': gap_a,
        'team_b_gap_days': gap_b,
        'h2h_gap_days': gap_h2h,
        'team_a_factor': factor_a,
        'team_b_factor': factor_b,
        'h2h_factor': factor_h2h,
        'match_factor': match_factor,
        'formula': 'factor=1 for <=30d; afterwards exp(-ln(2)*(gap-30)/90), floor 0.05',
        'history_not_discarded': True,
    }


def adapt_match(source: dict[str, Any], config: dict[str, Any], strict: bool=False) -> dict[str, Any]:
    canonical = _V101_ADAPT_BASE(source, config, strict)
    canonical['snapshot_datetime'] = _v101_datetime_from_source(source)
    freshness = _v101_history_freshness(canonical, config)
    canonical['history_freshness'] = freshness
    gate = canonical.setdefault('data_gate', {})
    gate['history_freshness'] = deepcopy(freshness)
    gate['history_gap_over_30_days'] = any(
        value is not None and value > 30.0
        for value in (freshness.get('team_a_gap_days'), freshness.get('team_b_gap_days'))
    )
    return canonical


def _v101_shrink_probability(probability: Optional[float], factor: float) -> Optional[float]:
    if probability is None:
        return None
    return max(0.0, min(1.0, 0.50 + float(factor) * (float(probability) - 0.50)))


def calculate_history(market: dict[str, Any], canonical: dict[str, Any], config: dict[str, Any]) -> dict[str, Any]:
    result = _V101_HISTORY_BASE(market, canonical, config)
    freshness = canonical.get('history_freshness') or _v101_history_freshness(canonical, config)
    factor = float(freshness.get('match_factor', 1.0))
    raw_p_hist = to_number(result.get('p_hist'))
    adjusted = _v101_shrink_probability(raw_p_hist, factor)
    result['p_hist_raw_before_freshness'] = raw_p_hist
    result['p_hist'] = adjusted if adjusted is not None else result.get('p_hist')
    if 'p_hist_IT' in result:
        result['p_hist_IT_raw_before_freshness'] = to_number(result.get('p_hist_IT'))
        result['p_hist_IT'] = result['p_hist']
    raw_zone = to_number(result.get('history_zone_rate'))
    result['history_zone_rate_raw'] = raw_zone
    result['history_zone_rate_freshness_adjusted'] = _v101_shrink_probability(raw_zone, factor)
    result['history_freshness'] = deepcopy(freshness)
    result['history_freshness_factor'] = factor
    result['history_freshness_formula'] = 'P_hist_fresh=0.50+factor*(P_hist_raw-0.50); raw exact-line zone remains the mandatory >=75% gate.'
    return result


def calculate_scenario(
    market: dict[str, Any],
    canonical: dict[str, Any],
    history: dict[str, Any],
    config: dict[str, Any],
) -> dict[str, Any]:
    result = _V101_SCENARIO_BASE(market, canonical, history, config)
    freshness = canonical.get('history_freshness') or {}
    factor = float(freshness.get('match_factor', 1.0))
    raw_probability = to_number(result.get('p_scenario'))
    adjusted = _v101_shrink_probability(raw_probability, factor)
    result['p_scenario_raw_before_freshness'] = raw_probability
    if adjusted is not None:
        result['p_scenario'] = adjusted
    result['scenario_freshness_factor'] = factor
    result['scenario_freshness_formula'] = 'P_scenario_fresh=0.50+factor*(P_scenario_raw-0.50).'
    return result


def _v101_recompute_live_probability(
    result: dict[str, Any],
    market: dict[str, Any],
    projection_used: float,
    freshness_factor: float,
    config: dict[str, Any],
) -> None:
    line = float(market['line'])
    side = market['side']
    base_sigma = float(result.get('sigma') or result.get('sigma_base') or 1.0)
    extra_max = float(config.get('history_freshness', {}).get('live_sigma_max_extra', 0.25))
    sigma = base_sigma * (1.0 + max(0.0, 1.0 - freshness_factor) * extra_max)
    edge_over = projection_used - line
    edge_under = line - projection_used
    edge = edge_over if side == 'OVER' else edge_under
    z_score = edge / sigma if sigma > 0 else 0.0
    result.update({
        'projection_used_before_freshness': result.get('projection_used'),
        'projection_used': projection_used,
        'Projection_used': projection_used,
        'line_edge': edge,
        'line_edge_over': edge_over,
        'line_edge_under': edge_under,
        'sigma_before_freshness': base_sigma,
        'sigma': sigma,
        'z_score': z_score,
        'p_live': normal_cdf(z_score) if sigma > 0 else 0.50,
        'history_freshness_factor': freshness_factor,
    })


def calculate_live_projection(
    market: dict[str, Any],
    canonical: dict[str, Any],
    history: dict[str, Any],
    scenario: dict[str, Any],
    config: dict[str, Any],
    stat: Optional[dict[str, Any]]=None,
) -> dict[str, Any]:
    result = _V101_LIVE_BASE(market, canonical, history, scenario, config, stat)
    if canonical.get('stage') == 'PRE_MATCH':
        return result
    freshness = canonical.get('history_freshness') or {}
    factor = float(freshness.get('match_factor', 1.0))
    mode = str(result.get('data_mode') or canonical.get('data_mode') or '')

    if mode == 'SCORE_TIME_HISTORY':
        components = result.get('no_stat_components') or {}
        base_weights = result.get('no_stat_configured_weights') or {}
        regressed = to_number(components.get('regressed'))
        history_projection = to_number(components.get('history'))
        scenario_projection = to_number(components.get('scenario'))
        base_reg = float(base_weights.get('regressed', 0.0))
        base_hist = float(base_weights.get('history', 0.0))
        base_scen = float(base_weights.get('scenario', 0.0))
        adjusted_weights = {
            'regressed': base_reg + (1.0 - factor) * (base_hist + base_scen),
            'history': base_hist * factor,
            'scenario': base_scen * factor,
        }
        values = {
            'regressed': regressed,
            'history': history_projection,
            'scenario': scenario_projection,
        }
        available = {
            key: value for key, value in values.items()
            if value is not None and adjusted_weights.get(key, 0.0) > 0
        }
        denominator = sum(adjusted_weights[key] for key in available)
        projection_used = (
            sum(float(available[key]) * adjusted_weights[key] for key in available) / denominator
            if denominator > 0 else float(result.get('projection_used') or market['line'])
        )
        result['no_stat_freshness_weights'] = adjusted_weights
        result['projection_formula_before_freshness'] = result.get('projection_formula')
        result['projection_formula'] = (
            'NO_STAT age-aware: history/scenario base weights multiplied by freshness; '
            'lost weight transferred to current score-time regressed projection.'
        )
        _v101_recompute_live_probability(result, market, projection_used, factor, config)
    elif mode in {'FULL_STAT', 'PARTIAL_STAT'} and factor < 0.999999:
        original = to_number(result.get('projection_used'))
        regressed = to_number(result.get('projection_regressed'))
        stat_projection = to_number(result.get('projection_stat_adjusted'))
        current_live_values = [value for value in (regressed, stat_projection) if value is not None]
        if original is not None and current_live_values:
            current_live_anchor = statistics.median(current_live_values)
            projection_used = factor * original + (1.0 - factor) * current_live_anchor
            result['current_live_anchor_for_freshness'] = current_live_anchor
            result['projection_formula_before_freshness'] = result.get('projection_formula')
            result['projection_formula'] = (
                'FULL/PARTIAL age-aware: freshness*CORE_projection + '
                '(1-freshness)*median(score-time regressed, CORE stat projection).'
            )
            _v101_recompute_live_probability(result, market, projection_used, factor, config)
    return result


def _verdict(
    probability: float,
    blockers: list[dict[str, Any]],
    strong_clean: bool,
    p_hist: Optional[float]=None,
) -> str:
    if blockers or probability < 0.65:
        return 'PASS'
    if probability < 0.75:
        return 'RISK PLAY'
    if probability < 0.80:
        return 'LIVE PLAY'
    return 'PLAY' if strong_clean else 'LIVE PLAY'


def normalized_action(
    probability: float,
    blockers: list[dict[str, Any]],
    mode: str,
    p_hist: Optional[float]=None,
) -> tuple[str, str, str]:
    if blockers:
        return 'PASS', 'PASS', '0%'
    if mode.upper() == 'STRICT' and probability < 0.75:
        return 'PASS', 'TRIGGER ONLY', '0%'
    if probability < 0.65:
        return 'PASS', 'PASS', '0%'
    if probability < 0.75:
        return 'RISK', 'RISK ENTRY', '10-15% live-limit'
    if probability < 0.80:
        return 'PLAY', 'LOW PLAY', '15-20% live-limit'
    if probability < 0.85:
        return 'PLAY', 'MAIN PLAY', '30-35% live-limit'
    return 'PLAY', 'STRONG PLAY', '40-50% live-limit'


def apply_risk_post_filter(decision: dict[str, Any]) -> dict[str, Any]:
    return {
        'enabled': False,
        'applicable': False,
        'p_final': to_number((decision.get('probabilities') or {}).get('p_final')),
        'p_live': to_number((decision.get('probabilities') or {}).get('p_live')),
        'passed': True,
        'filtered': False,
        'reason_code': None,
        'policy': 'RISK_65_TO_74_99_ONLY_AFTER_75_HISTORY_3_EDGE_AND_NO_SERIOUS_CONFLICT',
    }


def _v101_serious_conflicts(item: dict[str, Any]) -> list[dict[str, Any]]:
    stat = item.get('stat_comparison') or {}
    live = item.get('live') or {}
    side = item.get('side')
    conflicts: list[dict[str, Any]] = []
    status = str(stat.get('stat_gate_status') or 'OFF')
    if status in {'AGAINST', 'CONFLICT'}:
        conflicts.append(_blocker(
            'GLOBAL_SERIOUS_STAT_CONFLICT',
            'Serious global stat conflict: candidate direction is opposed or simultaneously contradicted by independent live channels',
            {'stat_gate_status': status, 'over_score': stat.get('over_gate_score'), 'under_score': stat.get('under_gate_score')},
        ))
    fake_candidate = bool(
        (side == 'OVER' and stat.get('fake_over'))
        or (side == 'UNDER' and stat.get('fake_under'))
    )
    if fake_candidate:
        conflicts.append(_blocker(
            'GLOBAL_SERIOUS_FAKE_PROFILE',
            'Candidate is a fake Over/Under profile; RISK is forbidden',
            {'side': side, 'fake_over': stat.get('fake_over'), 'fake_under': stat.get('fake_under')},
        ))
    if live.get('projection_conflict'):
        conflicts.append(_blocker(
            'GLOBAL_SERIOUS_PROJECTION_CONFLICT',
            'CORE stat projection and checkpoint/control projection diverge beyond the project threshold',
            {
                'threshold': live.get('projection_conflict_threshold'),
                'projection_stat': live.get('projection_stat_adjusted'),
                'projection_control': live.get('projection_control'),
            },
        ))
    return conflicts


def _v101_evaluate_market(self: SuperBasketCalculator, market: dict[str, Any], canonical: dict[str, Any]) -> dict[str, Any]:
    item = _V101_EVALUATE_BASE(self, market, canonical)
    existing = {entry.get('rule_id') for entry in item.get('blockers', [])}
    serious = [entry for entry in _v101_serious_conflicts(item) if entry.get('rule_id') not in existing]
    if serious:
        item.setdefault('blockers', []).extend(serious)
        item['hard_conflict'] = True
        item['verdict'] = 'PASS'
        item.setdefault('p_trace', []).append(_trace_step(
            'GLOBAL_SERIOUS_CONFLICT_GATE',
            True,
            'AGAINST/CONFLICT stat profile, candidate-side fake profile, or major projection divergence => hard PASS; never RISK.',
            {'conflicts': serious},
            item.get('p_final'),
            item.get('p_final'),
            [entry['rule_id'] for entry in serious],
        ))
    item['history_freshness'] = deepcopy(canonical.get('history_freshness') or {})
    item['signal_policy_version'] = 'RISK_65_NO_SERIOUS_CONFLICT_V10_1'
    return item


SuperBasketCalculator.evaluate_market = _v101_evaluate_market


_V101_FORMULA_REGISTRY = deepcopy(_V10_FORMULA_REGISTRY)
_V101_FORMULA_REGISTRY.update({
    'history_freshness': (
        'After a 30-day grace period, history trust decays continuously with a 90-day half-life. '
        'P_hist_fresh=0.50+freshness*(P_hist_raw-0.50). Raw exact-line zone remains the >=75% entry gate.'
    ),
    'no_stat_freshness_projection': (
        'NO_STAT history/scenario weights are multiplied by freshness; removed weight is transferred to '
        'the current score-time regressed projection. History is never hard-deleted.'
    ),
    'signal_thresholds_v10_1': 'P_final<65% PASS; 65-74.99% RISK; >=75% PLAY.',
    'serious_conflict_policy': (
        'Stat AGAINST/CONFLICT, candidate-side FAKE profile, or major CORE projection divergence '
        'is a hard PASS and can never be emitted as RISK.'
    ),
})


def _v101_calculate(self: SuperBasketCalculator, source: dict[str, Any], dispatch_threshold: Optional[float]=None, strict_schema: bool=False) -> dict[str, Any]:
    output = _V101_CALCULATE_BASE(self, source, dispatch_threshold, strict_schema)
    calculation = output.get('super_basket_calculation', {})
    calculation['engine_version'] = '10.1.0-FRESH-HISTORY-CONFLICT-SAFE'
    calculation['formula_registry'] = deepcopy(_V101_FORMULA_REGISTRY)
    calculation['signal_policy'] = {
        'risk_min': 0.65,
        'play_min': 0.75,
        'raw_history_zone_min': 0.75,
        'live_edge_min_points': 3.0,
        'serious_conflicts_hard_pass': True,
        'history_freshness_enabled': True,
    }
    output['super_basket_calculation'] = calculation
    return output


SuperBasketCalculator.calculate = _v101_calculate


def summarize_line_evaluation(item: dict[str, Any]) -> dict[str, Any]:
    summary = _V101_SUMMARIZE_BASE(item)
    history = item.get('history') or {}
    live = item.get('live') or {}
    summary.update({
        'p_hist_raw_before_freshness': history.get('p_hist_raw_before_freshness'),
        'history_freshness_factor': history.get('history_freshness_factor'),
        'history_zone_rate_raw': history.get('history_zone_rate_raw'),
        'history_zone_rate_freshness_adjusted': history.get('history_zone_rate_freshness_adjusted'),
        'projection_used_before_freshness': live.get('projection_used_before_freshness'),
        'no_stat_freshness_weights': live.get('no_stat_freshness_weights'),
        'serious_conflict_codes': [
            entry.get('rule_id') for entry in item.get('blockers', [])
            if str(entry.get('rule_id', '')).startswith('GLOBAL_SERIOUS_')
        ],
    })
    return summary



_V101_DETERMINISTIC_EXPLANATION_BASE = deterministic_explanation


def deterministic_explanation(evaluation: Optional[dict[str, Any]], action: str, mode: str) -> tuple[str, str, str]:
    explanation, risk, trigger = _V101_DETERMINISTIC_EXPLANATION_BASE(evaluation, action, mode)
    trigger = trigger.replace('P_final ≥60%', 'P_final ≥65%')
    risk = risk.replace('зоні 60–74.99%', 'зоні 65–74.99%')
    return explanation, risk, trigger


SYSTEM_VERSION = '10.1.0'
DEFAULT_CONFIG['engine_version'] = '10.1.0-FRESH-HISTORY-CONFLICT-SAFE'


# ===== v10.2 EDGE-TIERED HISTORY GATE =====
# User policy:
# - P_final thresholds stay unchanged: <65% PASS; 65-74.99% RISK; >=75% PLAY.
# - serious global conflicts remain hard PASS.
# - required raw exact-line history zone depends on live projection edge:
#     0.50 <= edge < 3.00  -> history >= 75%
#     3.00 <= edge < 5.00  -> history >= 70%
#     edge >= 5.00         -> history >= 65%
#     edge < 0.50          -> PASS
# - freshness adjustment of P_hist/P_scenario/projection remains enabled.

_V102_EVALUATE_BASE = SuperBasketCalculator.evaluate_market
_V102_CALCULATE_BASE = SuperBasketCalculator.calculate
_V102_SUMMARIZE_BASE = summarize_line_evaluation
_V102_DETERMINISTIC_EXPLANATION_BASE = deterministic_explanation

DEFAULT_CONFIG['engine_version'] = '10.2.0-EDGE-TIERED-HISTORY'
DEFAULT_CONFIG['dispatch_threshold'] = 0.65
DEFAULT_CONFIG.setdefault('signal_gates', {}).update({
    'history_zone_min': 0.65,
    'live_edge_min_points': 0.50,
    'risk_min': 0.65,
    'play_min': 0.75,
    'serious_conflicts_hard_block': True,
})
DEFAULT_CONFIG.setdefault('edge_tiered_history_gate', {}).update({
    'minimum_live_edge': 0.50,
    'tier_1_edge_min': 0.50,
    'tier_1_edge_max_exclusive': 3.00,
    'tier_1_history_min': 0.75,
    'tier_2_edge_min': 3.00,
    'tier_2_edge_max_exclusive': 5.00,
    'tier_2_history_min': 0.70,
    'tier_3_edge_min': 5.00,
    'tier_3_history_min': 0.65,
})


def _v102_required_history_zone(line_edge: Optional[float]) -> tuple[Optional[float], str]:
    edge = to_number(line_edge)
    if edge is None:
        return None, 'EDGE_MISSING'
    if edge < 0.50:
        return None, 'EDGE_BELOW_0_50'
    if edge < 3.00:
        return 0.75, 'EDGE_0_50_TO_2_99_HISTORY_75'
    if edge < 5.00:
        return 0.70, 'EDGE_3_00_TO_4_99_HISTORY_70'
    return 0.65, 'EDGE_5_PLUS_HISTORY_65'


def _v102_sanitize_trace(
    item: dict[str, Any],
    required_history: Optional[float],
    tier: str,
    raw_zone: Optional[float],
    edge: Optional[float],
) -> None:
    obsolete = {'HISTORY_ZONE_BELOW_75', 'LIVE_EDGE_BELOW_3'}
    current_blockers = deepcopy(item.get('blockers') or [])
    current_codes = [entry.get('rule_id') for entry in current_blockers]
    for step in item.get('p_trace') or []:
        if isinstance(step.get('reason_codes'), list):
            step['reason_codes'] = [
                code for code in step['reason_codes'] if code not in obsolete
            ]
        inputs = step.get('inputs')
        if isinstance(inputs, dict) and isinstance(inputs.get('blockers'), list):
            inputs['blockers'] = [
                entry for entry in inputs['blockers']
                if entry.get('rule_id') not in obsolete
            ]
        if step.get('step') == 'ALIGNED_SIGNAL_GATES':
            step['formula'] = (
                'Edge-tiered raw history gate: edge 0.50-2.99 requires 75%; '
                'edge 3.00-4.99 requires 70%; edge 5.00+ requires 65%.'
            )
            step['inputs'] = {
                'history_zone_rate_raw': raw_zone,
                'required_history_zone': required_history,
                'line_edge': edge,
                'edge_tier': tier,
                'p_scenario': (item.get('scenario') or {}).get('p_scenario'),
            }
            step['reason_codes'] = [
                code for code in current_codes
                if code in {'LIVE_EDGE_BELOW_0_50', 'HISTORY_ZONE_BELOW_DYNAMIC_MIN', 'SCENARIO_DIRECTION_CONFLICT'}
            ]
        elif step.get('step') == 'HARD_BLOCKERS':
            step['applied'] = bool(current_blockers)
            step['inputs'] = {'blockers': current_blockers}
            step['reason_codes'] = current_codes
        elif step.get('step') == 'P_FINAL_RULE':
            step['reason_codes'] = [item.get('verdict')]


def _v102_evaluate_market(
    self: SuperBasketCalculator,
    market: dict[str, Any],
    canonical: dict[str, Any],
) -> dict[str, Any]:
    item = _V102_EVALUATE_BASE(self, market, canonical)

    obsolete = {'HISTORY_ZONE_BELOW_75', 'LIVE_EDGE_BELOW_3'}
    item['blockers'] = [
        entry for entry in (item.get('blockers') or [])
        if entry.get('rule_id') not in obsolete
    ]

    history = item.get('history') or {}
    live = item.get('live') or {}
    raw_zone = to_number(history.get('history_zone_rate_raw'))
    if raw_zone is None:
        raw_zone = to_number(history.get('history_zone_rate'))
    edge = to_number(live.get('line_edge'))

    if canonical.get('stage') == 'PRE_MATCH':
        required_history, tier = 0.75, 'PRE_MATCH_HISTORY_75'
        if raw_zone is None or raw_zone < required_history:
            item['blockers'].append(_blocker(
                'HISTORY_ZONE_BELOW_DYNAMIC_MIN',
                'Pre-match signal requires at least a 75% raw exact-line historical zone',
                {
                    'history_zone_rate_raw': raw_zone,
                    'required': required_history,
                    'tier': tier,
                },
            ))
    else:
        required_history, tier = _v102_required_history_zone(edge)
        if required_history is None:
            item['blockers'].append(_blocker(
                'LIVE_EDGE_BELOW_0_50',
                'Live projection edge must be at least 0.50 point in the candidate direction',
                {
                    'line_edge': edge,
                    'required': 0.50,
                    'projection_used': live.get('projection_used'),
                    'line': market.get('line'),
                    'tier': tier,
                },
            ))
        elif raw_zone is None or raw_zone < required_history:
            item['blockers'].append(_blocker(
                'HISTORY_ZONE_BELOW_DYNAMIC_MIN',
                'Raw exact-line historical zone is below the minimum required for this live edge tier',
                {
                    'history_zone_rate_raw': raw_zone,
                    'required': required_history,
                    'line_edge': edge,
                    'tier': tier,
                    'source': history.get('history_zone_source'),
                },
            ))

    deduped: list[dict[str, Any]] = []
    seen: set[tuple[Any, Any]] = set()
    for entry in item.get('blockers') or []:
        key = (
            entry.get('rule_id'),
            json.dumps(entry.get('inputs') or {}, sort_keys=True, ensure_ascii=False, default=str),
        )
        if key not in seen:
            seen.add(key)
            deduped.append(entry)
    item['blockers'] = deduped

    stat = item.get('stat_comparison') or {}
    strong_clean = bool(
        not item['blockers']
        and not item.get('caps')
        and item.get('data_mode') == 'FULL_STAT'
        and stat.get('stat_gate_status') == 'CONFIRMED'
    )
    item['verdict'] = _verdict(
        float(item.get('p_final') or 0.0),
        item['blockers'],
        strong_clean,
        history.get('p_hist'),
    )
    item['dynamic_history_gate'] = {
        'line_edge': edge,
        'history_zone_rate_raw': raw_zone,
        'required_history_zone': required_history,
        'tier': tier,
        'passed': (
            required_history is not None
            and raw_zone is not None
            and raw_zone >= required_history
        ),
        'minimum_live_edge': 0.50,
        'formula': (
            '0.50<=edge<3 => history>=75%; '
            '3<=edge<5 => history>=70%; '
            'edge>=5 => history>=65%; edge<0.50 => PASS.'
        ),
    }
    item['signal_policy_version'] = 'EDGE_TIERED_HISTORY_V10_2'
    _v102_sanitize_trace(item, required_history, tier, raw_zone, edge)
    item.setdefault('p_trace', []).append(_trace_step(
        'EDGE_TIERED_HISTORY_GATE',
        True,
        (
            '0.50<=edge<3 requires raw history>=75%; '
            '3<=edge<5 requires >=70%; edge>=5 requires >=65%; '
            'serious conflicts remain hard PASS.'
        ),
        deepcopy(item['dynamic_history_gate']),
        item.get('p_final'),
        item.get('p_final'),
        [
            entry.get('rule_id') for entry in item['blockers']
            if entry.get('rule_id') in {
                'LIVE_EDGE_BELOW_0_50',
                'HISTORY_ZONE_BELOW_DYNAMIC_MIN',
            }
        ],
    ))
    return item


SuperBasketCalculator.evaluate_market = _v102_evaluate_market


_V102_FORMULA_REGISTRY = deepcopy(_V101_FORMULA_REGISTRY)
_V102_FORMULA_REGISTRY.update({
    'edge_tiered_history_gate': (
        'Raw exact-line history threshold depends on candidate-side live edge: '
        '0.50<=edge<3.00 -> 75%; 3.00<=edge<5.00 -> 70%; edge>=5.00 -> 65%. '
        'Edge<0.50 is PASS.'
    ),
    'signal_thresholds_v10_2': (
        'P_final<65% PASS; 65-74.99% RISK; >=75% PLAY. '
        'Serious global conflicts remain hard PASS.'
    ),
})


def _v102_calculate(
    self: SuperBasketCalculator,
    source: dict[str, Any],
    dispatch_threshold: Optional[float]=None,
    strict_schema: bool=False,
) -> dict[str, Any]:
    output = _V102_CALCULATE_BASE(self, source, dispatch_threshold, strict_schema)
    calculation = output.get('super_basket_calculation', {})
    calculation['engine_version'] = '10.2.0-EDGE-TIERED-HISTORY'
    calculation['formula_registry'] = deepcopy(_V102_FORMULA_REGISTRY)
    calculation['signal_policy'] = {
        'risk_min': 0.65,
        'play_min': 0.75,
        'edge_history_tiers': [
            {'edge_min': 0.50, 'edge_max_exclusive': 3.00, 'raw_history_zone_min': 0.75},
            {'edge_min': 3.00, 'edge_max_exclusive': 5.00, 'raw_history_zone_min': 0.70},
            {'edge_min': 5.00, 'edge_max_exclusive': None, 'raw_history_zone_min': 0.65},
        ],
        'edge_below_0_50': 'PASS',
        'serious_conflicts_hard_pass': True,
        'history_freshness_enabled': True,
        'odds_min': float(self.config.get('odds_min', 1.44)),
    }
    output['super_basket_calculation'] = calculation
    return output


SuperBasketCalculator.calculate = _v102_calculate


def summarize_line_evaluation(item: dict[str, Any]) -> dict[str, Any]:
    summary = _V102_SUMMARIZE_BASE(item)
    gate = item.get('dynamic_history_gate') or {}
    summary.update({
        'dynamic_history_gate_tier': gate.get('tier'),
        'dynamic_required_history_zone': gate.get('required_history_zone'),
        'dynamic_history_gate_passed': gate.get('passed'),
        'minimum_live_edge_v10_2': gate.get('minimum_live_edge'),
    })
    return summary


def apply_risk_post_filter(decision: dict[str, Any]) -> dict[str, Any]:
    return {
        'enabled': False,
        'applicable': False,
        'p_final': to_number((decision.get('probabilities') or {}).get('p_final')),
        'p_live': to_number((decision.get('probabilities') or {}).get('p_live')),
        'passed': True,
        'filtered': False,
        'reason_code': None,
        'policy': (
            'RISK_65_TO_74_99_AFTER_EDGE_TIERED_HISTORY_GATE_'
            'AND_NO_SERIOUS_CONFLICT'
        ),
    }


def deterministic_explanation(
    evaluation: Optional[dict[str, Any]],
    action: str,
    mode: str,
) -> tuple[str, str, str]:
    explanation, risk, trigger = _V102_DETERMINISTIC_EXPLANATION_BASE(
        evaluation, action, mode
    )
    tier_text = (
        'Історичний gate залежить від live edge: '
        '0.5–2.99 очка → історія від 75%; '
        '3–4.99 → від 70%; 5+ → від 65%.'
    )
    if tier_text not in trigger:
        trigger = f'{trigger} {tier_text}'.strip()
    return explanation, risk, trigger


SYSTEM_VERSION = '10.2.0'
DEFAULT_CONFIG['engine_version'] = '10.2.0-EDGE-TIERED-HISTORY'


# ===== v10.3 CONFLICTS-AS-WARNINGS SIGNAL POLICY =====
# User policy:
# - P_final thresholds stay unchanged: <65% PASS; 65-74.99% RISK; >=75% PLAY.
# - dynamic raw history gate stays unchanged:
#     0.50 <= edge < 3.00 -> history >=75%
#     3.00 <= edge < 5.00 -> history >=70%
#     edge >=5.00 -> history >=65%
# - directional/model conflicts are warnings, not blockers.
# - P_final is NOT recomputed after removing conflict blockers; existing caps remain.
# - structural/data/router/market blockers remain active.

_V103_EVALUATE_BASE = SuperBasketCalculator.evaluate_market
_V103_CALCULATE_BASE = SuperBasketCalculator.calculate
_V103_SUMMARIZE_BASE = summarize_line_evaluation
_V103_DETERMINISTIC_EXPLANATION_BASE = deterministic_explanation

DEFAULT_CONFIG['engine_version'] = '10.3.0-CONFLICTS-AS-WARNINGS'
DEFAULT_CONFIG['dispatch_threshold'] = 0.65
DEFAULT_CONFIG.setdefault('signal_gates', {}).update({
    'serious_conflicts_hard_block': False,
})
DEFAULT_CONFIG.setdefault('conflict_policy', {}).update({
    'directional_conflicts_are_warnings': True,
    'keep_existing_probability_caps': True,
    'keep_structural_blockers': True,
})

_V103_CONFLICT_BLOCKER_IDS = {
    'GLOBAL_SERIOUS_STAT_CONFLICT',
    'GLOBAL_SERIOUS_FAKE_PROFILE',
    'GLOBAL_SERIOUS_PROJECTION_CONFLICT',
    'STAT_GATE_DIRECTLY_AGAINST',
    'PARTIAL_STAT_GATE_AGAINST',
    'STRONG_HISTORY_LIVE_CONFLICT',
    'SCENARIO_DIRECTION_CONFLICT',
}


def _v103_is_conflict_blocker(entry: dict[str, Any]) -> bool:
    rule_id = str(entry.get('rule_id') or '')
    return (
        rule_id in _V103_CONFLICT_BLOCKER_IDS
        or rule_id.startswith('GLOBAL_SERIOUS_')
    )


def _v103_clean_trace(item: dict[str, Any]) -> None:
    remaining = deepcopy(item.get('blockers') or [])
    remaining_codes = [entry.get('rule_id') for entry in remaining]
    warnings = deepcopy(item.get('conflict_warnings') or [])
    warning_codes = [entry.get('rule_id') for entry in warnings]

    for step in item.get('p_trace') or []:
        if isinstance(step.get('reason_codes'), list):
            step['reason_codes'] = [
                code for code in step['reason_codes']
                if code not in warning_codes
            ]
        inputs = step.get('inputs')
        if isinstance(inputs, dict) and isinstance(inputs.get('blockers'), list):
            inputs['blockers'] = [
                entry for entry in inputs['blockers']
                if not _v103_is_conflict_blocker(entry)
            ]
        if step.get('step') in {
            'GLOBAL_SERIOUS_CONFLICT_GATE',
            'LIVE_HISTORY_CONFLICT',
        }:
            step['applied'] = False
            step['formula'] = (
                'v10.3: directional/model conflict is audit warning only; '
                'it does not force PASS.'
            )
            step['inputs'] = {'conflict_warnings': warnings}
            step['reason_codes'] = []
        elif step.get('step') == 'HARD_BLOCKERS':
            step['applied'] = bool(remaining)
            step['inputs'] = {'blockers': remaining}
            step['reason_codes'] = remaining_codes
        elif step.get('step') == 'P_FINAL_RULE':
            step['reason_codes'] = [item.get('verdict')]


def _v103_evaluate_market(
    self: SuperBasketCalculator,
    market: dict[str, Any],
    canonical: dict[str, Any],
) -> dict[str, Any]:
    item = _V103_EVALUATE_BASE(self, market, canonical)

    original_blockers = deepcopy(item.get('blockers') or [])
    conflict_warnings = [
        entry for entry in original_blockers
        if _v103_is_conflict_blocker(entry)
    ]
    remaining_blockers = [
        entry for entry in original_blockers
        if not _v103_is_conflict_blocker(entry)
    ]

    item['blockers'] = remaining_blockers
    item['conflict_warnings'] = conflict_warnings
    item['conflicts_block_signal'] = False
    item['hard_conflict'] = bool(remaining_blockers)

    history = item.get('history') or {}
    stat = item.get('stat_comparison') or {}
    strong_clean = bool(
        not remaining_blockers
        and not item.get('caps')
        and item.get('data_mode') == 'FULL_STAT'
        and stat.get('stat_gate_status') == 'CONFIRMED'
    )
    item['verdict'] = _verdict(
        float(item.get('p_final') or 0.0),
        remaining_blockers,
        strong_clean,
        history.get('p_hist'),
    )
    item['signal_policy_version'] = 'CONFLICTS_AS_WARNINGS_V10_3'
    item['conflict_policy'] = {
        'directional_conflicts_are_warnings': True,
        'p_final_recomputed': False,
        'existing_caps_preserved': True,
        'removed_blocker_ids': [
            entry.get('rule_id') for entry in conflict_warnings
        ],
        'remaining_blocker_ids': [
            entry.get('rule_id') for entry in remaining_blockers
        ],
    }

    _v103_clean_trace(item)
    item.setdefault('p_trace', []).append(_trace_step(
        'CONFLICTS_AS_WARNINGS_GATE',
        bool(conflict_warnings),
        (
            'Directional/model conflicts are logged as warnings only. '
            'Signal eligibility is controlled by P_final, dynamic raw-history gate, '
            'live edge, odds and remaining structural/data/router blockers.'
        ),
        {
            'conflict_warnings': conflict_warnings,
            'remaining_blockers': remaining_blockers,
            'p_final_unchanged': item.get('p_final'),
            'dynamic_history_gate': item.get('dynamic_history_gate'),
        },
        item.get('p_final'),
        item.get('p_final'),
        [],
    ))
    return item


SuperBasketCalculator.evaluate_market = _v103_evaluate_market


_V103_FORMULA_REGISTRY = deepcopy(_V102_FORMULA_REGISTRY)
_V103_FORMULA_REGISTRY.update({
    'conflict_policy_v10_3': (
        'Directional/model conflicts (stat AGAINST/CONFLICT, fake profile, '
        'projection divergence, scenario opposition, strong history-live conflict) '
        'are audit warnings only. Existing P_final and probability caps remain. '
        'Structural/data/router blockers remain hard blockers.'
    ),
    'signal_thresholds_v10_3': (
        'P_final<65% PASS; 65-74.99% RISK; >=75% PLAY. '
        'Dynamic history gate by live edge remains unchanged.'
    ),
})


def _v103_calculate(
    self: SuperBasketCalculator,
    source: dict[str, Any],
    dispatch_threshold: Optional[float]=None,
    strict_schema: bool=False,
) -> dict[str, Any]:
    output = _V103_CALCULATE_BASE(self, source, dispatch_threshold, strict_schema)
    calculation = output.get('super_basket_calculation', {})
    calculation['engine_version'] = '10.3.0-CONFLICTS-AS-WARNINGS'
    calculation['formula_registry'] = deepcopy(_V103_FORMULA_REGISTRY)
    calculation['signal_policy'] = {
        'risk_min': 0.65,
        'play_min': 0.75,
        'edge_history_tiers': [
            {'edge_min': 0.50, 'edge_max_exclusive': 3.00, 'raw_history_zone_min': 0.75},
            {'edge_min': 3.00, 'edge_max_exclusive': 5.00, 'raw_history_zone_min': 0.70},
            {'edge_min': 5.00, 'edge_max_exclusive': None, 'raw_history_zone_min': 0.65},
        ],
        'edge_below_0_50': 'PASS',
        'directional_conflicts_block_signal': False,
        'directional_conflicts_logged_as_warnings': True,
        'existing_probability_caps_preserved': True,
        'structural_data_router_blockers_preserved': True,
        'history_freshness_enabled': True,
        'odds_min': float(self.config.get('odds_min', 1.44)),
    }
    output['super_basket_calculation'] = calculation
    return output


SuperBasketCalculator.calculate = _v103_calculate


def summarize_line_evaluation(item: dict[str, Any]) -> dict[str, Any]:
    summary = _V103_SUMMARIZE_BASE(item)
    summary.update({
        'conflict_warning_codes': [
            entry.get('rule_id')
            for entry in item.get('conflict_warnings', [])
        ],
        'conflicts_block_signal_v10_3': False,
        'remaining_hard_blocker_codes': [
            entry.get('rule_id')
            for entry in item.get('blockers', [])
        ],
    })
    return summary


def apply_risk_post_filter(decision: dict[str, Any]) -> dict[str, Any]:
    return {
        'enabled': False,
        'applicable': False,
        'p_final': to_number((decision.get('probabilities') or {}).get('p_final')),
        'p_live': to_number((decision.get('probabilities') or {}).get('p_live')),
        'passed': True,
        'filtered': False,
        'reason_code': None,
        'policy': (
            'RISK_65_TO_74_99_AFTER_DYNAMIC_EDGE_HISTORY_GATE; '
            'DIRECTIONAL_CONFLICTS_ARE_WARNINGS'
        ),
    }


def deterministic_explanation(
    evaluation: Optional[dict[str, Any]],
    action: str,
    mode: str,
) -> tuple[str, str, str]:
    explanation, risk, trigger = _V103_DETERMINISTIC_EXPLANATION_BASE(
        evaluation, action, mode
    )
    if evaluation and evaluation.get('conflict_warnings'):
        warning_codes = ', '.join(
            str(entry.get('rule_id'))
            for entry in evaluation.get('conflict_warnings', [])
        )
        risk = (
            f'{risk} Конфлікти не блокують сигнал у v10.3; '
            f'вони залишені як попередження: {warning_codes}.'
        ).strip()
    return explanation, risk, trigger


SYSTEM_VERSION = '10.3.0-TELEGRAM-READY'
DEFAULT_CONFIG['engine_version'] = '10.3.0-CONFLICTS-AS-WARNINGS'


# ===== v10.4 EDGE 3+ / RAW HISTORY 60% =====
# User policy:
# - P_final thresholds stay unchanged:
#     P_final <65%       -> PASS
#     65-74.99%          -> RISK
#     >=75%              -> PLAY
# - minimum live edge stays 0.50.
# - raw exact-line history gate:
#     0.50 <= edge <3.00 -> history >=75%
#     edge >=3.00        -> history >=60%
# - PRE_MATCH history gate stays 75%.
# - directional/model conflicts stay warnings, not hard blockers.
# - odds minimum stays 1.44.

_V104_EVALUATE_BASE = SuperBasketCalculator.evaluate_market
_V104_CALCULATE_BASE = SuperBasketCalculator.calculate
_V104_SUMMARIZE_BASE = summarize_line_evaluation
_V104_DETERMINISTIC_EXPLANATION_BASE = deterministic_explanation

DEFAULT_CONFIG['engine_version'] = '10.4.0-EDGE3-HISTORY60'
DEFAULT_CONFIG['dispatch_threshold'] = 0.65
DEFAULT_CONFIG.setdefault('signal_gates', {}).update({
    'history_zone_min': 0.60,
    'live_edge_min_points': 0.50,
    'risk_min': 0.65,
    'play_min': 0.75,
    'serious_conflicts_hard_block': False,
})
DEFAULT_CONFIG['edge_tiered_history_gate'] = {
    'minimum_live_edge': 0.50,
    'tier_1_edge_min': 0.50,
    'tier_1_edge_max_exclusive': 3.00,
    'tier_1_history_min': 0.75,
    'tier_2_edge_min': 3.00,
    'tier_2_edge_max_exclusive': None,
    'tier_2_history_min': 0.60,
}


def _v102_required_history_zone(line_edge: Optional[float]) -> tuple[Optional[float], str]:
    """v10.4 policy override used by the existing v10.2/v10.3 evaluation chain."""
    edge = to_number(line_edge)
    if edge is None:
        return None, 'EDGE_MISSING'
    if edge < 0.50:
        return None, 'EDGE_BELOW_0_50'
    if edge < 3.00:
        return 0.75, 'EDGE_0_50_TO_2_99_HISTORY_75'
    return 0.60, 'EDGE_3_PLUS_HISTORY_60'


def _v102_sanitize_trace(
    item: dict[str, Any],
    required_history: Optional[float],
    tier: str,
    raw_zone: Optional[float],
    edge: Optional[float],
) -> None:
    """v10.4 trace override so audit text matches the active gate."""
    obsolete = {'HISTORY_ZONE_BELOW_75', 'LIVE_EDGE_BELOW_3'}
    current_blockers = deepcopy(item.get('blockers') or [])
    current_codes = [entry.get('rule_id') for entry in current_blockers]
    for step in item.get('p_trace') or []:
        if isinstance(step.get('reason_codes'), list):
            step['reason_codes'] = [
                code for code in step['reason_codes'] if code not in obsolete
            ]
        inputs = step.get('inputs')
        if isinstance(inputs, dict) and isinstance(inputs.get('blockers'), list):
            inputs['blockers'] = [
                entry for entry in inputs['blockers']
                if entry.get('rule_id') not in obsolete
            ]
        if step.get('step') == 'ALIGNED_SIGNAL_GATES':
            step['formula'] = (
                'v10.4 edge-tiered raw history gate: '
                'edge 0.50-2.99 requires 75%; edge 3.00+ requires 60%.'
            )
            step['inputs'] = {
                'history_zone_rate_raw': raw_zone,
                'required_history_zone': required_history,
                'line_edge': edge,
                'edge_tier': tier,
                'p_scenario': (item.get('scenario') or {}).get('p_scenario'),
            }
            step['reason_codes'] = [
                code for code in current_codes
                if code in {
                    'LIVE_EDGE_BELOW_0_50',
                    'HISTORY_ZONE_BELOW_DYNAMIC_MIN',
                    'SCENARIO_DIRECTION_CONFLICT',
                }
            ]
        elif step.get('step') == 'HARD_BLOCKERS':
            step['applied'] = bool(current_blockers)
            step['inputs'] = {'blockers': current_blockers}
            step['reason_codes'] = current_codes
        elif step.get('step') == 'P_FINAL_RULE':
            step['reason_codes'] = [item.get('verdict')]


def _v104_evaluate_market(
    self: SuperBasketCalculator,
    market: dict[str, Any],
    canonical: dict[str, Any],
) -> dict[str, Any]:
    item = _V104_EVALUATE_BASE(self, market, canonical)

    gate = item.get('dynamic_history_gate') or {}
    gate['formula'] = (
        '0.50<=edge<3 => history>=75%; '
        'edge>=3 => history>=60%; edge<0.50 => PASS.'
    )
    item['dynamic_history_gate'] = gate
    item['signal_policy_version'] = 'EDGE3_HISTORY60_CONFLICT_WARNINGS_V10_4'

    for step in item.get('p_trace') or []:
        if step.get('step') == 'EDGE_TIERED_HISTORY_GATE':
            step['formula'] = (
                'v10.4: 0.50<=edge<3 requires raw history>=75%; '
                'edge>=3 requires raw history>=60%; edge<0.50 is PASS. '
                'Directional/model conflicts are warnings.'
            )
            if isinstance(step.get('inputs'), dict):
                step['inputs']['formula_v10_4'] = gate['formula']
        elif step.get('step') == 'CONFLICTS_AS_WARNINGS_GATE':
            step['formula'] = (
                'v10.4: directional/model conflicts are warnings only. '
                'Signal eligibility uses P_final, edge/history gate, odds '
                'and remaining structural/data/router blockers.'
            )

    item.setdefault('p_trace', []).append(_trace_step(
        'EDGE_3_PLUS_HISTORY_60_GATE',
        True,
        (
            'From v10.4, any candidate-side live edge of 3.00+ points '
            'requires raw exact-line history of at least 60%.'
        ),
        {
            'line_edge': gate.get('line_edge'),
            'history_zone_rate_raw': gate.get('history_zone_rate_raw'),
            'required_history_zone': gate.get('required_history_zone'),
            'passed': gate.get('passed'),
        },
        item.get('p_final'),
        item.get('p_final'),
        [
            entry.get('rule_id') for entry in item.get('blockers', [])
            if entry.get('rule_id') in {
                'LIVE_EDGE_BELOW_0_50',
                'HISTORY_ZONE_BELOW_DYNAMIC_MIN',
            }
        ],
    ))
    return item


SuperBasketCalculator.evaluate_market = _v104_evaluate_market


_V104_FORMULA_REGISTRY = deepcopy(_V103_FORMULA_REGISTRY)
_V104_FORMULA_REGISTRY.update({
    'edge_history_gate_v10_4': (
        'Raw exact-line history threshold: 0.50<=edge<3.00 -> 75%; '
        'edge>=3.00 -> 60%; edge<0.50 -> PASS.'
    ),
    'signal_thresholds_v10_4': (
        'P_final<65% PASS; 65-74.99% RISK; >=75% PLAY; odds>=1.44. '
        'Directional/model conflicts are warnings; structural blockers remain active.'
    ),
})


def _v104_calculate(
    self: SuperBasketCalculator,
    source: dict[str, Any],
    dispatch_threshold: Optional[float]=None,
    strict_schema: bool=False,
) -> dict[str, Any]:
    output = _V104_CALCULATE_BASE(self, source, dispatch_threshold, strict_schema)
    calculation = output.get('super_basket_calculation', {})
    calculation['engine_version'] = '10.4.0-EDGE3-HISTORY60'
    calculation['formula_registry'] = deepcopy(_V104_FORMULA_REGISTRY)
    calculation['signal_policy'] = {
        'risk_min': 0.65,
        'play_min': 0.75,
        'edge_history_tiers': [
            {
                'edge_min': 0.50,
                'edge_max_exclusive': 3.00,
                'raw_history_zone_min': 0.75,
            },
            {
                'edge_min': 3.00,
                'edge_max_exclusive': None,
                'raw_history_zone_min': 0.60,
            },
        ],
        'edge_below_0_50': 'PASS',
        'pre_match_raw_history_zone_min': 0.75,
        'directional_conflicts_block_signal': False,
        'directional_conflicts_logged_as_warnings': True,
        'existing_probability_caps_preserved': True,
        'structural_data_router_blockers_preserved': True,
        'history_freshness_enabled': True,
        'odds_min': float(self.config.get('odds_min', 1.44)),
    }
    output['super_basket_calculation'] = calculation
    return output


SuperBasketCalculator.calculate = _v104_calculate


def summarize_line_evaluation(item: dict[str, Any]) -> dict[str, Any]:
    summary = _V104_SUMMARIZE_BASE(item)
    gate = item.get('dynamic_history_gate') or {}
    summary.update({
        'signal_policy_version': item.get('signal_policy_version'),
        'minimum_live_edge_v10_4': gate.get('minimum_live_edge'),
        'raw_history_required_v10_4': gate.get('required_history_zone'),
        'edge_3_plus_history_min_v10_4': 0.60,
    })
    return summary


def apply_risk_post_filter(decision: dict[str, Any]) -> dict[str, Any]:
    return {
        'enabled': False,
        'applicable': False,
        'p_final': to_number((decision.get('probabilities') or {}).get('p_final')),
        'p_live': to_number((decision.get('probabilities') or {}).get('p_live')),
        'passed': True,
        'filtered': False,
        'reason_code': None,
        'policy': (
            'RISK_65_TO_74_99_AFTER_V10_4_EDGE_HISTORY_GATE; '
            'EDGE_3_PLUS_HISTORY_60; DIRECTIONAL_CONFLICTS_ARE_WARNINGS'
        ),
    }


def deterministic_explanation(
    evaluation: Optional[dict[str, Any]],
    action: str,
    mode: str,
) -> tuple[str, str, str]:
    explanation, risk, trigger = _V104_DETERMINISTIC_EXPLANATION_BASE(
        evaluation, action, mode
    )
    old_text = (
        'Історичний gate залежить від live edge: '
        '0.5–2.99 очка → історія від 75%; '
        '3–4.99 → від 70%; 5+ → від 65%.'
    )
    new_text = (
        'Історичний gate v10.4: '
        '0.5–2.99 очка → історія від 75%; '
        '3+ очка → історія від 60%.'
    )
    trigger = trigger.replace(old_text, new_text)
    risk = risk.replace('у v10.3', 'у v10.4')
    if new_text not in trigger:
        trigger = f'{trigger} {new_text}'.strip()
    return explanation, risk, trigger


SYSTEM_VERSION = '10.4.0-TELEGRAM-READY'
DEFAULT_CONFIG['engine_version'] = '10.4.0-EDGE3-HISTORY60'


# =====================================================================
# v10.5 — CLEAN LIVE PROJECTION CHANNEL
# =====================================================================
# Critical fix:
# 1) P_live is calculated from a live-only projection.
# 2) Exact-line P_hist and P_scenario never shift Projection_used.
# 3) FULL/PARTIAL_STAT:
#       Current points + expected remaining points from live shot/FTA/
#       ORB/TO rates, with shrinkage only to pre-game STAT rates.
#       The previous second shrink to PreFinal and the
#       HistoryScenarioAnchor median are removed.
# 4) SCORE_TIME_HISTORY:
#       Current segment points / elapsed segment time * full segment time.
# 5) P_hist and P_scenario remain separate and enter only P_raw/P_final.
# =====================================================================

_V105_EVALUATE_BASE = SuperBasketCalculator.evaluate_market
_V105_CALCULATE_BASE = SuperBasketCalculator.calculate
_V105_FORMULA_REGISTRY = deepcopy(_V104_FORMULA_REGISTRY)
_V105_FORMULA_REGISTRY.update({
    'live_projection_no_stat_v10_5': (
        'Projection_live = current segment points / elapsed segment seconds '
        '* full segment seconds. P_hist and P_scenario are excluded.'
    ),
    'live_projection_full_stat_v10_5': (
        'Projection_live = current points + expected remaining points from '
        'live FGA/FTA/2PA/3PA/ORB/TO rates regressed only to pre-game '
        'box-score rates. No exact-line history or scenario outcome anchor.'
    ),
    'live_edge_v10_5': (
        'OVER edge = Projection_live-Line; UNDER edge = Line-Projection_live.'
    ),
    'p_live_v10_5': 'P_live = Phi(candidate-side live edge / sigma).',
    'p_raw_v10_5': (
        'P_raw = w_hist*P_hist + w_scenario*P_scenario + w_live*P_live.'
    ),
})


def _v105_team_quarter_points(
    canonical: dict[str, Any],
    market: dict[str, Any],
    quarter_index: int,
) -> Optional[float]:
    quarters = canonical.get('quarters') or []
    if quarter_index < 1 or quarter_index > len(quarters):
        return None
    row = quarters[quarter_index - 1] or {}
    team = market.get('team')
    if team:
        key = 'home' if team == canonical.get('home_team') else 'away'
        return to_number(row.get(key))
    return to_number(row.get('total'))


def _v105_score_time_controls(
    market: dict[str, Any],
    canonical: dict[str, Any],
    clock: dict[str, float],
) -> dict[str, Any]:
    elapsed = float(clock.get('elapsed_seconds') or 0.0)
    full = float(clock.get('full_seconds') or 0.0)
    current = float(clock.get('current_points') or 0.0)
    primary = current / elapsed * full if elapsed > 0 and full > 0 else None
    quarter_seconds = float(canonical.get('quarter_seconds') or 600.0)
    current_quarter = int(canonical.get('current_quarter') or 0)
    segment = str(market.get('segment') or 'MATCH')
    controls: dict[str, Optional[float]] = {'primary_scope_pace': primary}

    # A half during its second quarter: actual first quarter + projected
    # current quarter. This is independent from the whole-half pace channel.
    if segment == 'H1' and current_quarter == 2:
        q1 = _v105_team_quarter_points(canonical, market, 1)
        q2 = _v105_team_quarter_points(canonical, market, 2)
        q2_elapsed = max(
            0.0,
            quarter_seconds - float(canonical.get('quarter_seconds_remaining') or 0.0),
        )
        q2_projection = q2 / q2_elapsed * quarter_seconds if q2 is not None and q2_elapsed > 0 else None
        controls['current_quarter_projection'] = q2_projection
        controls['half_by_quarters_projection'] = (
            q1 + q2_projection if q1 is not None and q2_projection is not None else None
        )
    elif segment == 'H2' and current_quarter == 4:
        q3 = _v105_team_quarter_points(canonical, market, 3)
        q4 = _v105_team_quarter_points(canonical, market, 4)
        q4_elapsed = max(
            0.0,
            quarter_seconds - float(canonical.get('quarter_seconds_remaining') or 0.0),
        )
        q4_projection = q4 / q4_elapsed * quarter_seconds if q4 is not None and q4_elapsed > 0 else None
        controls['current_quarter_projection'] = q4_projection
        controls['half_by_quarters_projection'] = (
            q3 + q4_projection if q3 is not None and q4_projection is not None else None
        )

    # At HT provide quarter-equivalent full-game pace checks.
    if canonical.get('stage') == 'HT' and segment == 'MATCH':
        q1 = _v105_team_quarter_points(canonical, market, 1)
        q2 = _v105_team_quarter_points(canonical, market, 2)
        game_quarters = max(1.0, float(canonical.get('full_game_seconds') or 2400.0) / quarter_seconds)
        controls['q1_full_game_equivalent'] = q1 * game_quarters if q1 is not None else None
        controls['q2_full_game_equivalent'] = q2 * game_quarters if q2 is not None else None

    return controls


def _v9_projection_alignment(
    live: dict[str, Any],
    scenario_active: bool,
    side: str,
    line: float,
) -> dict[str, Any]:
    """v10.5 NO_STAT support uses live-only channels.

    It no longer asks history/scenario projections to agree with P_live.
    For a live half, both whole-half pace and quarter-control pace must
    support the candidate. At HT, at least two live pace channels must
    support the candidate. A lone early current-quarter pace does not
    satisfy the old 2-of-3 alignment point by itself.
    """
    controls = live.get('live_projection_controls') or {}
    values = {
        key: to_number(value)
        for key, value in controls.items()
        if to_number(value) is not None
    }
    side = str(side).upper()

    def aligned(value: float) -> bool:
        return value > line if side == 'OVER' else value < line

    flags = {key: aligned(float(value)) for key, value in values.items()}
    segment = str(live.get('segment') or '')
    market_type = str(live.get('market_type') or '')

    if 'half_by_quarters_projection' in values:
        selected_keys = [
            key for key in ('primary_scope_pace', 'half_by_quarters_projection')
            if key in values
        ]
        required = 2
    elif 'q1_full_game_equivalent' in values or 'q2_full_game_equivalent' in values:
        selected_keys = [
            key for key in (
                'primary_scope_pace',
                'q1_full_game_equivalent',
                'q2_full_game_equivalent',
            ) if key in values
        ]
        required = min(2, len(selected_keys)) if len(selected_keys) >= 2 else 2
    elif market_type in {'CURRENT_QUARTER_TOTAL', 'CURRENT_QUARTER_TEAM_IT'}:
        selected_keys = list(values)
        required = 2
    else:
        selected_keys = list(values)
        required = 1

    selected_flags = {key: flags[key] for key in selected_keys}
    count = sum(bool(value) for value in selected_flags.values())
    return {
        'values': {key: values[key] for key in selected_keys},
        'aligned_flags': selected_flags,
        'aligned_count': count,
        'required_count': required,
        'passed': bool(selected_flags) and count >= required,
        'policy': 'v10.5 live-only multi-clock alignment; no history/scenario projection',
    }


def calculate_live_projection(
    market: dict[str, Any],
    canonical: dict[str, Any],
    history: dict[str, Any],
    scenario: dict[str, Any],
    config: dict[str, Any],
    stat: Optional[dict[str, Any]]=None,
) -> dict[str, Any]:
    mode = str(
        canonical.get('data_mode')
        or canonical.get('data_gate', {}).get('data_mode')
        or 'DATA_OFF'
    )
    clock = _segment_clock(market, canonical)
    line = float(market['line'])
    side = str(market['side']).upper()
    sigma_base = _stage_sigma(market['market_type'], canonical['stage'], config)
    controls = _v105_score_time_controls(market, canonical, clock)
    projection_score_time = to_number(controls.get('primary_scope_pace'))
    projection_stat = None
    stat_details: dict[str, Any] = {}
    fallback_to_score_time = False

    if mode in {'FULL_STAT', 'PARTIAL_STAT'}:
        # Current-quarter stats in the source are cumulative for the game,
        # so a quarter market uses quarter score/time rather than pretending
        # cumulative FGA/FTA belong only to that quarter.
        if market['market_type'] in {
            'CURRENT_QUARTER_TOTAL',
            'CURRENT_QUARTER_TEAM_IT',
        }:
            projection_used = projection_score_time
            formula_mode = f'{mode}_CURRENT_QUARTER_SCORE_TIME'
            formula_text = (
                'Current-quarter points / elapsed quarter time * full quarter time; '
                'cumulative game box score is not reassigned to one quarter.'
            )
            fallback_to_score_time = True
        else:
            rho, k_stage, minutes_played = _v10_stage_trust(canonical)
            remaining_minutes = float(clock['remaining_seconds']) / 60.0
            pre_home = _v10_pre_stat_team('home', market, canonical)
            pre_away = _v10_pre_stat_team('away', market, canonical)
            live_home = _v10_live_stat_team(
                'home', market, canonical, pre_home, rho, remaining_minutes
            )
            live_away = _v10_live_stat_team(
                'away', market, canonical, pre_away, rho, remaining_minutes
            )
            if market.get('team'):
                selected = (
                    live_home
                    if market['team'] == canonical['home_team']
                    else live_away
                )
                # LiveRaw_Team already equals current score plus expected
                # remaining points from the regressed live stat rates.
                projection_stat = to_number(selected.get('LiveRaw_Team'))
            else:
                home_value = to_number(live_home.get('LiveRaw_Team'))
                away_value = to_number(live_away.get('LiveRaw_Team'))
                projection_stat = (
                    home_value + away_value
                    if home_value is not None and away_value is not None
                    else None
                )
            projection_used = (
                projection_stat
                if projection_stat is not None
                else projection_score_time
            )
            fallback_to_score_time = projection_stat is None
            formula_mode = (
                'FULL_STAT_LIVE_ONLY'
                if mode == 'FULL_STAT'
                else 'PARTIAL_STAT_LIVE_ONLY'
            )
            formula_text = (
                'Current points + expected remaining points from live '
                'FGA/FTA/2PA/3PA/ORB/TO rates. Live rates are shrunk once '
                'to pre-game box-score rates; exact-line P_hist and '
                'P_scenario do not enter Projection_used.'
            )
            stat_details = {
                'rho_stage': rho,
                'K_stage': k_stage,
                'minutes_played': minutes_played,
                'remaining_minutes_scope': remaining_minutes,
                'pre_home': pre_home,
                'pre_away': pre_away,
                'live_home': live_home,
                'live_away': live_away,
                'stat_projection_selected': projection_stat,
                'important_change': (
                    'Uses LiveRaw_Team. Removed second PreFinal shrink and '
                    'removed Projection_control/HistoryScenarioAnchor median.'
                ),
            }
        sigma_multiplier = (
            float(
                config.get('projection', {})
                .get('core_live', {})
                .get('partial_sigma_multiplier', 1.15)
            )
            if mode == 'PARTIAL_STAT'
            else 1.0
        )
        sigma = sigma_base * sigma_multiplier
    elif mode == 'SCORE_TIME_HISTORY':
        projection_used = projection_score_time
        sigma = sigma_base * float(
            config.get('projection', {})
            .get('core_live', {})
            .get('no_stat_sigma_multiplier', 1.20)
        )
        formula_mode = 'SCORE_TIME_LIVE_ONLY'
        formula_text = (
            'Current segment points / elapsed segment time * full segment time. '
            'P_hist, historical remaining median and P_scenario are excluded.'
        )
    else:
        projection_used = line
        sigma = sigma_base
        formula_mode = 'DATA_OFF'
        formula_text = 'DATA_OFF'

    if projection_used is None:
        projection_used = line
        fallback_to_score_time = True

    line_edge_over = projection_used - line
    line_edge_under = line - projection_used
    line_edge = line_edge_over if side == 'OVER' else line_edge_under
    z_score = line_edge / sigma if sigma > 0 else 0.0
    p_live = normal_cdf(z_score) if sigma > 0 else 0.50

    components = {
        'projection_score_time': {
            'value': projection_score_time,
            'included': mode == 'SCORE_TIME_HISTORY' or fallback_to_score_time,
            'role': 'live_only',
        },
        'projection_stat_live_only': {
            'value': projection_stat,
            'included': (
                mode in {'FULL_STAT', 'PARTIAL_STAT'}
                and projection_stat is not None
                and not fallback_to_score_time
            ),
            'role': 'live_only',
        },
        'projection_history': {
            'value': None,
            'included': False,
            'role': 'separate_P_hist_only',
        },
        'projection_scenario': {
            'value': None,
            'included': False,
            'role': 'separate_P_scenario_only',
        },
    }

    return {
        'clock': canonical.get('clock'),
        'elapsed_seconds': clock['elapsed_seconds'],
        'remaining_seconds': clock['remaining_seconds'],
        'elapsed_game_seconds': canonical['elapsed_game_seconds'],
        'remaining_game_seconds': canonical['remaining_game_seconds'],
        'current_points': clock['current_points'],
        'market_type': market.get('market_type'),
        'segment': market.get('segment'),
        'team': market.get('team'),
        'data_mode': mode,
        'components': components,
        'projection_simple': projection_score_time,
        'projection_score_time': projection_score_time,
        'projection_regressed': None,
        'projection_segment': projection_score_time,
        'projection_model_live': projection_used,
        'projection_history': None,
        'projection_scenario': None,
        'scenario_projection_method': 'SEPARATE_P_SCENARIO_ONLY',
        'projection_stat_adjusted': projection_stat,
        'projection_stat_live_only': projection_stat,
        'projection_control': None,
        'projection_history_scenario_anchor': None,
        'projection_used': projection_used,
        'Projection_used': projection_used,
        'line': line,
        'line_edge': line_edge,
        'line_edge_over': line_edge_over,
        'line_edge_under': line_edge_under,
        'projection_minus_line': projection_used - line,
        'sigma_base': sigma_base,
        'sigma': sigma,
        'z_score': z_score,
        'p_live': p_live,
        'required_history': _v9_required_history(market, canonical),
        'projection_formula_mode': formula_mode,
        'projection_formula': formula_text,
        'live_projection_controls': controls,
        'live_projection_independent_from_p_hist_scenario': True,
        'history_scenario_used_in_live_projection': False,
        'fallback_to_score_time': fallback_to_score_time,
        'no_stat_components': {
            'score_time_primary': projection_score_time,
            'half_by_quarters': controls.get('half_by_quarters_projection'),
            'q1_full_game_equivalent': controls.get('q1_full_game_equivalent'),
            'q2_full_game_equivalent': controls.get('q2_full_game_equivalent'),
        } if mode == 'SCORE_TIME_HISTORY' else {},
        'no_stat_configured_weights': {
            'score_time_primary': 1.0,
            'history': 0.0,
            'scenario': 0.0,
        } if mode == 'SCORE_TIME_HISTORY' else {},
        'projection_conflict': False,
        'projection_conflict_threshold': None,
        'stat_projection_details': stat_details,
        'core_live_formula_version': 'v10.5_CLEAN_LIVE_CHANNEL',
    }


def _v105_evaluate_market(
    self: SuperBasketCalculator,
    market: dict[str, Any],
    canonical: dict[str, Any],
) -> dict[str, Any]:
    item = _V105_EVALUATE_BASE(self, market, canonical)
    item['signal_policy_version'] = 'V10_5_CLEAN_LIVE_PROJECTION'
    live = item.get('live') or {}

    for step in item.get('p_trace') or []:
        if step.get('step') == 'P_LIVE':
            step['formula'] = (
                'v10.5 P_live = Phi(candidate-side edge / sigma), where edge '
                'comes only from clean live projection. P_hist/P_scenario do '
                'not shift Projection_used.'
            )
            step['inputs'] = {
                'data_mode': live.get('data_mode'),
                'projection_formula_mode': live.get('projection_formula_mode'),
                'projection_score_time': live.get('projection_score_time'),
                'projection_stat_live_only': live.get('projection_stat_live_only'),
                'projection_used': live.get('projection_used'),
                'line': item.get('line'),
                'projection_minus_line': live.get('projection_minus_line'),
                'candidate_side_edge': live.get('line_edge'),
                'sigma': live.get('sigma'),
            }
        elif step.get('step') == 'STAGE_WEIGHTS':
            step['formula'] = (
                'P_raw = w_hist*P_hist + w_scenario*P_scenario + '
                'w_live*P_live. The three channels stay separate.'
            )

    item.setdefault('p_trace', []).append(_trace_step(
        'CLEAN_LIVE_PROJECTION_V10_5',
        True,
        (
            'Projection_used is live-only. Exact-line history and matched '
            'scenario outcomes enter P_final separately and cannot create '
            'or reverse live edge.'
        ),
        {
            'data_mode': live.get('data_mode'),
            'formula_mode': live.get('projection_formula_mode'),
            'score_time_projection': live.get('projection_score_time'),
            'stat_projection': live.get('projection_stat_live_only'),
            'projection_used': live.get('projection_used'),
            'line': item.get('line'),
            'projection_minus_line': live.get('projection_minus_line'),
            'candidate_edge': live.get('line_edge'),
            'p_live': live.get('p_live'),
            'history_scenario_used_in_live_projection': False,
        },
        item.get('p_raw'),
        item.get('p_final'),
        [],
    ))
    return item


SuperBasketCalculator.evaluate_market = _v105_evaluate_market


def _v105_calculate(
    self: SuperBasketCalculator,
    source: dict[str, Any],
    dispatch_threshold: Optional[float]=None,
    strict_schema: bool=False,
) -> dict[str, Any]:
    output = _V105_CALCULATE_BASE(
        self, source, dispatch_threshold, strict_schema
    )
    calculation = output.get('super_basket_calculation', {})
    calculation['engine_version'] = '10.5.0-CLEAN-LIVE-PROJECTION'
    calculation['formula_registry'] = deepcopy(_V105_FORMULA_REGISTRY)
    calculation['live_projection_policy'] = {
        'channel_separation': True,
        'p_hist_used_in_projection': False,
        'p_scenario_used_in_projection': False,
        'no_stat_formula': (
            'current segment points / elapsed segment time * full segment time'
        ),
        'full_stat_formula': (
            'current points + expected remaining points from live box-score '
            'rates, with one regression to pre-game stat rates'
        ),
        'full_stat_removed_operations': [
            'second PreFinal shrink',
            'median with Projection_control',
            'median with HistoryScenarioAnchor',
        ],
        'p_live_formula': 'Phi(candidate-side edge / sigma)',
        'p_raw_formula': (
            'w_hist*P_hist + w_scenario*P_scenario + w_live*P_live'
        ),
    }
    output['super_basket_calculation'] = calculation
    return output


SuperBasketCalculator.calculate = _v105_calculate


def apply_risk_post_filter(decision: dict[str, Any]) -> dict[str, Any]:
    return {
        'enabled': False,
        'applicable': False,
        'p_final': to_number((decision.get('probabilities') or {}).get('p_final')),
        'p_live': to_number((decision.get('probabilities') or {}).get('p_live')),
        'passed': True,
        'filtered': False,
        'reason_code': None,
        'policy': (
            'V10_5_CLEAN_LIVE_PROJECTION; RISK_65_TO_74_99; '
            'PLAY_75_PLUS; EDGE_3_PLUS_HISTORY_60'
        ),
    }


SYSTEM_VERSION = '10.5.0-TELEGRAM-READY'
DEFAULT_CONFIG['engine_version'] = '10.5.0-CLEAN-LIVE-PROJECTION'


# =====================================================================
# v10.6 — LIVE DIRECTION + HISTORY/SCENARIO CONFIRMATION
# =====================================================================
# Live determines the only eligible side.
# History and scenario may confirm or reject that side, but never alter
# Projection_used or create a live edge.
#
# Dynamic weights:
#   edge 0.50–3.99  -> history 40%, scenario 15%, live 45%
#   edge 4.00–7.99  -> history 30%, scenario 15%, live 55%
#   edge 8.00–11.99 -> history 25%, scenario 10%, live 65%
#   edge 12.00+     -> history 15%, scenario 10%, live 75%
#
# Dynamic confirmation:
#   edge 0.50–3.99  -> P_hist >=65%, P_scenario >=60%
#   edge 4.00–7.99  -> P_hist >=60%, P_scenario >=55%
#   edge 8.00–11.99 -> P_hist >=55%, P_scenario >=55%
#   edge 12.00+     -> P_hist >=50%, P_scenario >=50%
# =====================================================================

_V106_EVALUATE_BASE = SuperBasketCalculator.evaluate_market
_V106_CALCULATE_BASE = SuperBasketCalculator.calculate
_V106_FORMULA_REGISTRY = deepcopy(_V105_FORMULA_REGISTRY)
_V106_FORMULA_REGISTRY.update({
    'live_direction_lock_v10_6': (
        'Clean live projection determines the only eligible side. '
        'Candidate-side edge must be >=0.50.'
    ),
    'dynamic_weights_v10_6': (
        '0.50-3.99 H40/S15/L45; 4-7.99 H30/S15/L55; '
        '8-11.99 H25/S10/L65; 12+ H15/S10/L75.'
    ),
    'dynamic_confirmation_v10_6': (
        'History/scenario confirmation minima decline as clean live edge grows: '
        '65/60, 60/55, 55/55, 50/50.'
    ),
})


def _v106_dynamic_policy(edge: Optional[float]) -> dict[str, Any]:
    value = to_number(edge)
    if value is None or value < 0.50:
        return {
            'tier': 'EDGE_BELOW_0_50',
            'weights': {'hist': 0.40, 'scenario': 0.15, 'live': 0.45},
            'hist_min': None,
            'scenario_min': None,
        }
    if value < 4.00:
        return {
            'tier': 'EDGE_0_50_TO_3_99',
            'weights': {'hist': 0.40, 'scenario': 0.15, 'live': 0.45},
            'hist_min': 0.65,
            'scenario_min': 0.60,
        }
    if value < 8.00:
        return {
            'tier': 'EDGE_4_00_TO_7_99',
            'weights': {'hist': 0.30, 'scenario': 0.15, 'live': 0.55},
            'hist_min': 0.60,
            'scenario_min': 0.55,
        }
    if value < 12.00:
        return {
            'tier': 'EDGE_8_00_TO_11_99',
            'weights': {'hist': 0.25, 'scenario': 0.10, 'live': 0.65},
            'hist_min': 0.55,
            'scenario_min': 0.55,
        }
    return {
        'tier': 'EDGE_12_PLUS',
        'weights': {'hist': 0.15, 'scenario': 0.10, 'live': 0.75},
        'hist_min': 0.50,
        'scenario_min': 0.50,
    }


def _v106_drop_old_direction_gates(
    blockers: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    obsolete = {
        'HISTORY_ZONE_BELOW_75',
        'HISTORY_ZONE_BELOW_DYNAMIC_MIN',
        'LIVE_EDGE_BELOW_3',
        'LIVE_EDGE_BELOW_0_50',
        'SCENARIO_DIRECTION_CONFLICT',
    }
    return [
        entry for entry in blockers
        if entry.get('rule_id') not in obsolete
    ]


def _v106_evaluate_market(
    self: SuperBasketCalculator,
    market: dict[str, Any],
    canonical: dict[str, Any],
) -> dict[str, Any]:
    item = _V106_EVALUATE_BASE(self, market, canonical)
    live = item.get('live') or {}
    history = item.get('history') or {}
    scenario = item.get('scenario') or {}

    # Empty/unsupported market evaluations stay untouched.
    if live.get('projection_used') is None or not history or not scenario:
        item['signal_policy_version'] = 'V10_6_LIVE_DIRECTION_CONFIRMATION'
        return item

    edge = to_number(live.get('line_edge'))
    p_live = float(live.get('p_live') or 0.50)
    p_hist = float(history.get('p_hist') or 0.50)
    p_scenario = float(scenario.get('p_scenario') or 0.50)
    policy = _v106_dynamic_policy(edge)
    weights = policy['weights']

    p_raw = (
        weights['hist'] * p_hist
        + weights['scenario'] * p_scenario
        + weights['live'] * p_live
    )

    blockers = _v106_drop_old_direction_gates(
        deepcopy(item.get('blockers') or [])
    )
    caps = deepcopy(item.get('caps') or [])

    if canonical.get('stage') != 'PRE_MATCH':
        if edge is None or edge < 0.50:
            blockers.append(_blocker(
                'LIVE_DIRECTION_OR_EDGE_FAILED',
                (
                    'Clean live projection does not support this side by '
                    'at least 0.50 points. History/scenario cannot create edge.'
                ),
                {
                    'candidate_side': market.get('side'),
                    'projection_used': live.get('projection_used'),
                    'line': market.get('line'),
                    'candidate_side_edge': edge,
                },
            ))
        else:
            hist_min = float(policy['hist_min'])
            scenario_min = float(policy['scenario_min'])
            if p_hist < hist_min:
                blockers.append(_blocker(
                    'HISTORY_CONFIRMATION_BELOW_DYNAMIC_MIN',
                    'History does not confirm the clean live direction',
                    {
                        'p_hist': p_hist,
                        'required': hist_min,
                        'edge_tier': policy['tier'],
                    },
                ))
            if p_scenario < scenario_min:
                blockers.append(_blocker(
                    'SCENARIO_CONFIRMATION_BELOW_DYNAMIC_MIN',
                    'Scenario does not confirm the clean live direction',
                    {
                        'p_scenario': p_scenario,
                        'required': scenario_min,
                        'edge_tier': policy['tier'],
                    },
                ))

    # Remove duplicates while preserving all structural/stat/router gates.
    deduped = []
    seen = set()
    for entry in blockers:
        code = str(entry.get('rule_id'))
        if code in seen:
            continue
        seen.add(code)
        deduped.append(entry)
    blockers = deduped

    active_cap = min(
        (float(entry.get('cap', 1.0)) for entry in caps),
        default=1.0,
    )
    p_final = min(p_raw, active_cap)

    stat = item.get('stat_comparison') or {}
    strong_clean = bool(
        not blockers
        and not caps
        and stat.get('stat_gate_status') == 'CONFIRMED'
        and p_live >= 0.75
        and p_hist >= 0.75
        and p_scenario >= 0.68
    )
    verdict = _verdict(
        p_final,
        blockers,
        strong_clean,
        p_hist,
    )

    item['signal_policy_version'] = 'V10_6_LIVE_DIRECTION_CONFIRMATION'
    item['weights']['normalized'] = deepcopy(weights)
    item['weights']['dynamic_policy_v10_6'] = deepcopy(policy)
    item['p_raw'] = p_raw
    item['p_final'] = p_final
    item['caps'] = caps
    item['blockers'] = blockers
    item['hard_conflict'] = bool(blockers)
    item['verdict'] = verdict
    item['confirmation_policy_v10_6'] = {
        'live_defines_direction': True,
        'history_changes_projection': False,
        'scenario_changes_projection': False,
        'candidate_side_edge': edge,
        **deepcopy(policy),
    }

    for step in item.get('p_trace') or []:
        if step.get('step') == 'STAGE_WEIGHTS':
            step['formula'] = (
                'v10.6 dynamic weights by clean candidate-side live edge.'
            )
            step['inputs'] = {
                'edge': edge,
                'tier': policy['tier'],
                'weights': weights,
                'p_hist': p_hist,
                'p_scenario': p_scenario,
                'p_live': p_live,
            }
            step['probability_after'] = p_raw
        elif step.get('step') == 'P_FINAL_RULE':
            step['inputs'] = {
                'p_raw_v10_6': p_raw,
                'active_cap': active_cap,
                'blockers_v10_6': blockers,
            }
            step['probability_after'] = p_final
            step['reason_codes'] = [verdict]

    item.setdefault('p_trace', []).append(_trace_step(
        'LIVE_DIRECTION_CONFIRMATION_V10_6',
        True,
        (
            'Live determines side; history and scenario only confirm/reject. '
            'They cannot modify Projection_used.'
        ),
        {
            'projection_used': live.get('projection_used'),
            'line': market.get('line'),
            'candidate_side_edge': edge,
            'p_live': p_live,
            'p_hist': p_hist,
            'p_scenario': p_scenario,
            'policy': policy,
        },
        p_raw,
        p_final,
        [entry.get('rule_id') for entry in blockers],
    ))
    return item


SuperBasketCalculator.evaluate_market = _v106_evaluate_market


def _v106_calculate(
    self: SuperBasketCalculator,
    source: dict[str, Any],
    dispatch_threshold: Optional[float]=None,
    strict_schema: bool=False,
) -> dict[str, Any]:
    output = _V106_CALCULATE_BASE(
        self,
        source,
        dispatch_threshold,
        strict_schema,
    )
    calculation = output.get('super_basket_calculation', {})
    calculation['engine_version'] = '10.6.0-LIVE-DIRECTION-CONFIRMATION'
    calculation['formula_registry'] = deepcopy(_V106_FORMULA_REGISTRY)
    calculation['signal_policy'] = {
        'risk_min': 0.65,
        'play_min': 0.75,
        'minimum_live_edge': 0.50,
        'live_defines_direction': True,
        'history_and_scenario_are_confirmation_only': True,
        'dynamic_tiers': [
            {
                'edge': '0.50-3.99',
                'weights': {'hist': 0.40, 'scenario': 0.15, 'live': 0.45},
                'hist_min': 0.65,
                'scenario_min': 0.60,
            },
            {
                'edge': '4.00-7.99',
                'weights': {'hist': 0.30, 'scenario': 0.15, 'live': 0.55},
                'hist_min': 0.60,
                'scenario_min': 0.55,
            },
            {
                'edge': '8.00-11.99',
                'weights': {'hist': 0.25, 'scenario': 0.10, 'live': 0.65},
                'hist_min': 0.55,
                'scenario_min': 0.55,
            },
            {
                'edge': '12.00+',
                'weights': {'hist': 0.15, 'scenario': 0.10, 'live': 0.75},
                'hist_min': 0.50,
                'scenario_min': 0.50,
            },
        ],
        'telegram': 'PLAY/RISK only; PASS is never sent',
    }
    output['super_basket_calculation'] = calculation
    return output


SuperBasketCalculator.calculate = _v106_calculate


def apply_risk_post_filter(decision: dict[str, Any]) -> dict[str, Any]:
    return {
        'enabled': False,
        'applicable': False,
        'p_final': to_number((decision.get('probabilities') or {}).get('p_final')),
        'p_live': to_number((decision.get('probabilities') or {}).get('p_live')),
        'passed': True,
        'filtered': False,
        'reason_code': None,
        'policy': (
            'V10_6_LIVE_DIRECTION_CONFIRMATION; '
            'RISK_65_TO_74_99; PLAY_75_PLUS'
        ),
    }


SYSTEM_VERSION = '10.6.0-TELEGRAM-READY'
DEFAULT_CONFIG['engine_version'] = '10.6.0-LIVE-DIRECTION-CONFIRMATION'
DEFAULT_CONFIG['dispatch_threshold'] = 0.65


# =====================================================================
# v10.9 — RESULT-PATTERN DECISION SELECTOR
# =====================================================================
# This layer does not change P_hist, P_scenario, P_live, Projection_used,
# P_raw or P_final. It only filters/ranks already calculated real lines.
# Seed patterns were mined from completed line outcomes in tmpmatches_3d.
# Every new match and every new line is evaluated; no bucket is permanently
# disabled. Rolling results are a small ranking modifier only.
# =====================================================================

_V109_APPLY_LEARNING_BASE = apply_learning_to_evaluation
_V109_SUMMARIZE_LINE_BASE = summarize_line_evaluation

DECISION_FILTER_VERSION = '10.9.0-RESULT-PATTERN-SELECTOR'
DECISION_FILTER_PROFILE = os.getenv(
    'SUPER_BASKET_DECISION_PROFILE', 'RESULT_PATTERN_60_73'
).strip().upper()
DECISION_FILTER_RISK_P_FINAL = float(os.getenv(
    'SUPER_BASKET_DECISION_RISK_P_FINAL', '0.60'
))
DECISION_FILTER_PLAY_P_FINAL = float(os.getenv(
    'SUPER_BASKET_DECISION_PLAY_P_FINAL', '0.73'
))
DECISION_FILTER_RISK_PATTERN_SCORE = float(os.getenv(
    'SUPER_BASKET_DECISION_RISK_PATTERN_SCORE', '0.74'
))
DECISION_FILTER_PLAY_PATTERN_SCORE = float(os.getenv(
    'SUPER_BASKET_DECISION_PLAY_PATTERN_SCORE', '0.84'
))
DECISION_FILTER_ROLLING_LIMIT = int(os.getenv(
    'SUPER_BASKET_DECISION_ROLLING_LIMIT', '60'
))
DECISION_FILTER_EMPIRICAL_MIN_SAMPLES = int(os.getenv(
    'SUPER_BASKET_DECISION_EMPIRICAL_MIN_SAMPLES', '20'
))

# Seed pattern outcome counts from the supplied completed archive.
# Laplace smoothing is applied at runtime: (wins+1)/(samples+2).
_V109_PATTERN_SEEDS = {
    'OVER_SCENARIO_75': {'wins': 9, 'samples': 9, 'unique_matches': 6},
    'OVER_LIVE_75': {'wins': 11, 'samples': 11, 'unique_matches': 4},
    'OVER_REQUIRED_HISTORY_55': {'wins': 12, 'samples': 12, 'unique_matches': 6},
    'UNDER_SCENARIO_75_EDGE_POSITIVE': {'wins': 3, 'samples': 3, 'unique_matches': 2},
}

_V109_SOFT_BLOCKERS = {
    'HISTORY_ZONE_BELOW_75',
    'LIVE_EDGE_BELOW_3',
    'LIVE_EDGE_BELOW_0_50',
    'SCENARIO_DIRECTION_CONFLICT',
    'NO_STAT_SUPPORT_TOO_LOW',
    'LIVE_DIRECTION_OR_EDGE_FAILED',
    'HISTORY_CONFIRMATION_BELOW_DYNAMIC_MIN',
    'SCENARIO_CONFIRMATION_BELOW_DYNAMIC_MIN',
    'REQUIRED_HISTORY_UNAVAILABLE',
}


def _v109_checkpoint(calculation: dict[str, Any]) -> Optional[int]:
    snapshot = calculation.get('canonical_snapshot') or {}
    trigger = to_int(snapshot.get('trigger_checkpoint'))
    if trigger in (1, 2, 3):
        return trigger
    stage = str(snapshot.get('stage') or '').upper()
    if stage == 'HT':
        return 2
    if stage in {'AFTER_3Q', 'Q4_CONFIRMATION'}:
        return 3
    if stage in {'EARLY_LIVE', 'CURRENT_Q1_Q3'}:
        return 1
    return None


def _v109_allowed_market(checkpoint: Optional[int], market_type: str) -> bool:
    if checkpoint == 1:
        return market_type in {'H1_TOTAL', 'TEAM_IT_H1'}
    if checkpoint == 2:
        return market_type in {'MATCH_TOTAL', 'H2_TOTAL', 'TEAM_IT_MATCH', 'TEAM_IT_H2'}
    if checkpoint == 3:
        return market_type in {'MATCH_TOTAL', 'CURRENT_QUARTER_TEAM_IT'}
    return False


def _v109_blocker_codes(item: dict[str, Any]) -> list[str]:
    output: list[str] = []
    for row in item.get('blockers') or []:
        code = str(row.get('rule_id') if isinstance(row, dict) else row or '').strip()
        if code:
            output.append(code)
    return list(dict.fromkeys(output))


def _v109_empirical_context(store: LearningStore, item: dict[str, Any], calculation: dict[str, Any]) -> dict[str, Any]:
    snapshot = calculation.get('canonical_snapshot') or {}
    stage = str(snapshot.get('stage') or 'UNKNOWN')
    market_type = str(item.get('market_type') or 'UNKNOWN')
    side = str(item.get('side') or 'UNKNOWN')
    results: list[str] = []
    try:
        rows = store.connection.execute(
            """SELECT result FROM signals
               WHERE result IN ('WIN','LOSS')
                 AND stage=? AND market_type=? AND side=?
               ORDER BY COALESCE(settled_at, created_at) DESC
               LIMIT ?""",
            (stage, market_type, side, DECISION_FILTER_ROLLING_LIMIT),
        ).fetchall()
        results = [str(row['result']) for row in rows]
    except Exception:
        results = []
    samples = len(results)
    wins = sum(result == 'WIN' for result in results)
    rate = wins / samples if samples else None
    adjustment = 0.0
    label = 'INSUFFICIENT_SAMPLE'
    if samples >= DECISION_FILTER_EMPIRICAL_MIN_SAMPLES and rate is not None:
        if rate >= 0.73:
            adjustment = 0.02
            label = 'ROLLING_73_PLUS'
        elif rate >= 0.60:
            label = 'ROLLING_60_TO_72'
        elif rate >= 0.50:
            adjustment = -0.01
            label = 'ROLLING_50_TO_59'
        else:
            adjustment = -0.02
            label = 'ROLLING_BELOW_50'
    return {
        'scope': 'rolling stage+market_type+side',
        'samples': samples,
        'wins': wins,
        'win_rate': rate,
        'label': label,
        'pattern_score_adjustment': adjustment,
        'hard_block': False,
    }


def _v109_seed_pattern_matches(
    checkpoint: Optional[int],
    market_type: str,
    side: str,
    p_final: float,
    p_live: float,
    p_scenario: float,
    required_history: Optional[float],
    edge: float,
) -> list[dict[str, Any]]:
    matches: list[dict[str, Any]] = []

    def add(pattern_id: str, condition: bool, description: str) -> None:
        if not condition:
            return
        seed = _V109_PATTERN_SEEDS[pattern_id]
        posterior = (float(seed['wins']) + 1.0) / (float(seed['samples']) + 2.0)
        matches.append({
            'pattern_id': pattern_id,
            'description': description,
            'seed_wins': seed['wins'],
            'seed_samples': seed['samples'],
            'seed_unique_matches': seed['unique_matches'],
            'laplace_reliability': posterior,
        })

    add(
        'OVER_SCENARIO_75',
        side == 'OVER' and p_final >= 0.60 and p_scenario >= 0.75,
        'OVER: P_final >=60% and P_scenario >=75%',
    )
    add(
        'OVER_LIVE_75',
        side == 'OVER' and p_final >= 0.60 and p_live >= 0.75,
        'OVER: P_final >=60% and P_live >=75%',
    )
    add(
        'OVER_REQUIRED_HISTORY_55',
        side == 'OVER' and p_final >= 0.60 and required_history is not None and required_history >= 0.55,
        'OVER: P_final >=60% and RequiredHistoryP >=55%',
    )
    add(
        'UNDER_SCENARIO_75_EDGE_POSITIVE',
        side == 'UNDER' and p_final >= 0.60 and p_scenario >= 0.75 and edge >= 0.0,
        'UNDER: P_final >=60%, P_scenario >=75%, projection supports UNDER',
    )
    return matches


def _v109_decision_gate(item: dict[str, Any], store: LearningStore, calculation: dict[str, Any]) -> dict[str, Any]:
    checkpoint = _v109_checkpoint(calculation)
    market_type = str(item.get('market_type') or '')
    side = str(item.get('side') or '')
    p_final = float(item.get('p_final') or 0.0)
    history = item.get('history') or {}
    scenario = item.get('scenario') or {}
    live = item.get('live') or {}
    p_hist = float(history.get('p_hist') or 0.50)
    p_scenario = float(scenario.get('p_scenario') or 0.50)
    p_live = float(live.get('p_live') or 0.50)
    edge = float(live.get('line_edge') or 0.0)
    projection = to_number(live.get('projection_used'))
    required_history = to_number((live.get('required_history') or {}).get('p_required_history'))
    odds = to_number(item.get('odds'))
    data_mode = str(item.get('data_mode') or '')
    stat = item.get('stat_comparison') or {}
    stat_status = str(stat.get('stat_gate_status') or 'OFF')
    router = item.get('router') or {}
    router_status = str(router.get('status') or '')
    parser_issues = list(item.get('parser_issues') or [])
    snapshot = calculation.get('canonical_snapshot') or {}
    data_gate = calculation.get('data_gate') or snapshot.get('data_gate') or {}
    blocker_codes = _v109_blocker_codes(item)
    hard_blockers = [code for code in blocker_codes if code not in _V109_SOFT_BLOCKERS]
    soft_blockers = [code for code in blocker_codes if code in _V109_SOFT_BLOCKERS]

    if DECISION_FILTER_PROFILE in {'LEGACY', 'OFF', 'DISABLED'}:
        return {
            'version': DECISION_FILTER_VERSION,
            'profile': DECISION_FILTER_PROFILE,
            'enabled': False,
            'passed': item.get('system_action') != 'PASS',
            'recommended_action': item.get('system_action') or 'PASS',
            'pattern_score': p_final,
            'reason_codes': [],
        }

    reasons: list[str] = []
    warnings: list[str] = []
    if parser_issues:
        reasons.append('RP60_PARSER_OR_LINE_ISSUE')
    if odds is None or odds < float(DEFAULT_CONFIG.get('odds_min', 1.44)):
        reasons.append('RP60_ODDS_FAILED')
    if router_status not in {'ALLOW', 'PRIORITY', 'DOWNGRADE'}:
        reasons.append('RP60_ROUTER_NOT_ALLOWED')
    if not _v109_allowed_market(checkpoint, market_type):
        reasons.append('RP60_STAGE_MARKET_NOT_ALLOWED')
    if projection is None:
        reasons.append('RP60_PROJECTION_UNAVAILABLE')
    if not bool(data_gate.get('time_reliable', True)):
        reasons.append('RP60_TIME_NOT_RELIABLE')
    if hard_blockers:
        reasons.append('RP60_HARD_BLOCKER_PRESENT')
    if p_final < DECISION_FILTER_RISK_P_FINAL:
        reasons.append('RP60_P_FINAL_BELOW_60')
    if data_mode == 'FULL_STAT' and stat_status == 'AGAINST':
        reasons.append('RP60_FULL_STAT_AGAINST')

    matched = _v109_seed_pattern_matches(
        checkpoint, market_type, side, p_final, p_live, p_scenario,
        required_history, edge,
    )
    if not matched:
        reasons.append('RP60_NO_VALIDATED_RESULT_PATTERN')

    pattern_reliability = (
        sum(float(row['laplace_reliability']) for row in matched) / len(matched)
        if matched else 0.50
    )
    # Decision score, not a recalculated match probability.
    pattern_score = 0.65 * pattern_reliability + 0.35 * p_final
    if data_mode == 'FULL_STAT':
        if stat_status == 'CONFIRMED':
            pattern_score += 0.02
        elif stat_status == 'CONFLICT':
            pattern_score -= 0.01
            warnings.append('RP60_FULL_STAT_CONFLICT')
    empirical = _v109_empirical_context(store, item, calculation)
    pattern_score += float(empirical.get('pattern_score_adjustment') or 0.0)
    pattern_score = max(0.0, min(1.0, pattern_score))

    if pattern_score < DECISION_FILTER_RISK_PATTERN_SCORE:
        reasons.append('RP60_PATTERN_SCORE_BELOW_RISK_MIN')

    passed = not reasons
    recommended_action = 'PASS'
    if passed:
        play_ready = (
            p_final >= DECISION_FILTER_PLAY_P_FINAL
            and pattern_score >= DECISION_FILTER_PLAY_PATTERN_SCORE
            and len(matched) >= 2
            and not (data_mode == 'FULL_STAT' and stat_status == 'CONFLICT')
        )
        recommended_action = 'PLAY' if play_ready else 'RISK'

    return {
        'version': DECISION_FILTER_VERSION,
        'profile': DECISION_FILTER_PROFILE,
        'enabled': True,
        'checkpoint': checkpoint,
        'market_type': market_type,
        'side': side,
        'data_mode': data_mode,
        'stat_gate_status': stat_status,
        'p_final_unchanged': p_final,
        'p_live_unchanged': p_live,
        'p_hist_unchanged': p_hist,
        'p_scenario_unchanged': p_scenario,
        'projection_used_unchanged': projection,
        'required_history_p': required_history,
        'line_edge_unchanged': edge,
        'matched_result_patterns': matched,
        'matched_pattern_count': len(matched),
        'pattern_reliability': pattern_reliability,
        'pattern_score': pattern_score,
        'pattern_score_is_probability': False,
        'empirical_results': empirical,
        'soft_blockers_reconsidered': soft_blockers,
        'hard_blockers_preserved': hard_blockers,
        'risk_p_final_minimum': DECISION_FILTER_RISK_P_FINAL,
        'play_p_final_minimum': DECISION_FILTER_PLAY_P_FINAL,
        'risk_pattern_score_minimum': DECISION_FILTER_RISK_PATTERN_SCORE,
        'play_pattern_score_minimum': DECISION_FILTER_PLAY_PATTERN_SCORE,
        'passed': passed,
        'recommended_action': recommended_action,
        'reason_codes': list(dict.fromkeys(reasons)),
        'warning_codes': list(dict.fromkeys(warnings)),
    }


def apply_learning_to_evaluation(
    evaluation: dict[str, Any],
    store: LearningStore,
    calculation: dict[str, Any],
    mode: str,
) -> dict[str, Any]:
    item = _V109_APPLY_LEARNING_BASE(evaluation, store, calculation, mode)
    gate = _v109_decision_gate(item, store, calculation)
    item['decision_filter'] = gate
    item['decision_pattern_score'] = gate.get('pattern_score', 0.0)
    item['decision_quality_score'] = gate.get('pattern_score', 0.0)
    item.setdefault('system_reason_codes', [])

    if not gate.get('enabled'):
        return item
    if not gate.get('passed'):
        item['system_action'] = 'PASS'
        item['system_status'] = 'PASS — RESULT PATTERN FILTER'
        item['stake'] = '0%'
        item['system_reason_codes'].extend(gate.get('reason_codes') or [])
    else:
        action = str(gate.get('recommended_action') or 'RISK')
        item['system_action'] = action
        if action == 'PLAY':
            item['system_status'] = 'PLAY — RESULT PATTERN VERIFIED'
            item['stake'] = '15-20% live-limit'
        else:
            item['system_status'] = 'RISK PLAY — RESULT PATTERN VERIFIED'
            item['stake'] = '10-15% live-limit'
        item['system_reason_codes'].append('RP60_PATTERN_GATE_PASSED')
        item['system_reason_codes'].extend(gate.get('warning_codes') or [])

    item.setdefault('p_trace', []).append(_trace_step(
        'DECISION_FILTER_V10_9_RESULT_PATTERNS',
        bool(gate.get('passed')),
        (
            'Decision-only result-pattern selector. P_hist, P_scenario, P_live, '
            'Projection_used, P_raw and P_final remain unchanged.'
        ),
        deepcopy(gate),
        item.get('p_final'),
        item.get('p_final'),
        gate.get('reason_codes') or ['RP60_PATTERN_GATE_PASSED'],
    ))
    return item


def select_one_decision(
    evaluations: list[dict[str, Any]],
    mode: str,
) -> tuple[Optional[dict[str, Any]], Optional[dict[str, Any]]]:
    active = _candidate_pool(evaluations)
    eligible = [item for item in active if item.get('system_action') != 'PASS']

    def rank(item: dict[str, Any]) -> tuple[float, ...]:
        action_rank = 2.0 if item.get('system_action') == 'PLAY' else 1.0
        gate = item.get('decision_filter') or {}
        return (
            action_rank,
            float(item.get('decision_pattern_score') or 0.0),
            float(gate.get('matched_pattern_count') or 0.0),
            float(item.get('p_final_system') or item.get('p_final') or 0.0),
            float((item.get('live') or {}).get('p_live') or 0.0),
            float(item.get('odds') or 0.0),
        )

    eligible.sort(key=rank, reverse=True)
    active.sort(
        key=lambda item: (
            float(item.get('decision_pattern_score') or 0.0),
            float(item.get('p_final_system') or item.get('p_final') or 0.0),
            -len(item.get('blockers') or []),
            float(item.get('odds') or 0.0),
        ),
        reverse=True,
    )
    closest = active[0] if active else (evaluations[0] if evaluations else None)
    return (eligible[0] if eligible else None, closest)


def summarize_line_evaluation(item: dict[str, Any]) -> dict[str, Any]:
    output = _V109_SUMMARIZE_LINE_BASE(item)
    output['decision_pattern_score'] = item.get('decision_pattern_score')
    output['decision_filter'] = deepcopy(item.get('decision_filter') or {})
    return output


SYSTEM_VERSION = '10.9.0-RESULT-PATTERN-SELECTOR-TELEGRAM-READY'
DEFAULT_CONFIG['engine_version'] = SYSTEM_VERSION


# ===== v11.0 TELEGRAM ADVISOR + SCENARIO MINER OVERRIDES =====
# This layer preserves the v10.6 calculation chain and replaces only routing,
# adviser selection, rich scenario explanations, theoretical line search and
# Telegram delivery policy.

ADVISOR_VERSION = '11.0.0-LIVE-ADVISOR-SCENARIO-MINER'
ADVISOR_HISTORY_ZONE_MIN = float(os.getenv('SUPER_BASKET_ADVISOR_HISTORY_ZONE_MIN', '0.75'))
ADVISOR_EXCEPTIONAL_EDGE_MIN = float(os.getenv('SUPER_BASKET_ADVISOR_EXCEPTIONAL_EDGE_MIN', '15.0'))
ADVISOR_PLAY_MIN = float(os.getenv('SUPER_BASKET_ADVISOR_PLAY_MIN', '0.75'))
ADVISOR_RISK_MIN = float(os.getenv('SUPER_BASKET_ADVISOR_RISK_MIN', '0.60'))
ADVISOR_MAX_PRIMARY = max(1, int(os.getenv('SUPER_BASKET_ADVISOR_MAX_PRIMARY', '3')))
ADVISOR_MODEL_ODDS = float(os.getenv('SUPER_BASKET_ADVISOR_MODEL_ODDS', '1.90'))
ADVISOR_MODEL_OFFSETS = (0.5, 1.5, 2.5, 3.5, 5.5, 7.5, 10.5, 15.5, 20.5)

_V11_ROUTER_BASE = _router


def _router(market: dict[str, Any], canonical: dict[str, Any]) -> dict[str, Any]:
    """Advisor router.

    Q2 and Q3 totals/IT are full calculation markets. Match totals/IT are
    calculated at every checkpoint. Completed segments remain blocked.
    """
    market_type = str(market.get('market_type') or '')
    segment = str(market.get('segment') or '')
    elapsed = float(canonical.get('elapsed_game_seconds') or 0.0)
    full = float(canonical.get('full_game_seconds') or 2400.0)
    half = full / 2.0
    current = to_int(canonical.get('current_quarter'))
    trigger = to_int(canonical.get('trigger_checkpoint'))

    if market_type in {'MATCH_TOTAL', 'TEAM_IT_MATCH'}:
        if canonical.get('stage') == 'AFTER_3Q':
            return {'status': 'PRIORITY', 'reason': 'ADVISOR_MATCH_MARKET_AFTER_Q3', 'cap': None, 'hard_block': False}
        return {'status': 'ALLOW', 'reason': 'ADVISOR_MATCH_MARKET_ALL_CHECKPOINTS', 'cap': None, 'hard_block': False}

    if market_type in {'H1_TOTAL', 'TEAM_IT_H1'}:
        if elapsed >= half:
            return {'status': 'BLOCK', 'reason': 'H1_ALREADY_COMPLETE', 'cap': None, 'hard_block': True}
        return {'status': 'PRIORITY' if trigger == 1 else 'ALLOW', 'reason': 'ADVISOR_H1_ACTIVE', 'cap': None, 'hard_block': False}

    if market_type in {'H2_TOTAL', 'TEAM_IT_H2'}:
        if elapsed >= full:
            return {'status': 'BLOCK', 'reason': 'MATCH_ALREADY_COMPLETE', 'cap': None, 'hard_block': True}
        if elapsed < half:
            # Calculate and report as forward-looking RISK/model context, but do
            # not permit a clean play before half-time.
            return {'status': 'DOWNGRADE', 'reason': 'H2_EARLY_FORWARD_LOOK', 'cap': 0.74, 'hard_block': False}
        return {'status': 'PRIORITY', 'reason': 'ADVISOR_H2_ACTIVE', 'cap': None, 'hard_block': False}

    if market_type in {'CURRENT_QUARTER_TOTAL', 'CURRENT_QUARTER_TEAM_IT'}:
        target = to_int(segment[1:]) if segment.startswith('Q') else None
        expected = (trigger + 1) if trigger in {1, 2, 3} else current
        if target is None:
            return {'status': 'BLOCK', 'reason': 'UNKNOWN_QUARTER', 'cap': None, 'hard_block': True}
        if expected is not None and target != expected:
            return {'status': 'BLOCK', 'reason': 'QUARTER_NOT_ACTIVE_FOR_CHECKPOINT', 'cap': None, 'hard_block': True}
        if target < 1 or target > 4:
            return {'status': 'BLOCK', 'reason': 'INVALID_QUARTER', 'cap': None, 'hard_block': True}
        if target == 4:
            return {'status': 'CONTEXT_GATE', 'reason': 'Q4_REQUIRES_CONTEXT_GATE', 'cap': None, 'hard_block': False}
        return {'status': 'PRIORITY', 'reason': f'ADVISOR_Q{target}_STANDALONE_ALLOWED', 'cap': None, 'hard_block': False}

    return _V11_ROUTER_BASE(market, canonical)


def _v11_round_half(value: float) -> float:
    return math.floor(float(value)) + 0.5


def _v11_completed_quarters(canonical: dict[str, Any]) -> int:
    trigger = to_int(canonical.get('trigger_checkpoint'))
    if trigger in {1, 2, 3}:
        return trigger
    count = 0
    for quarter in canonical.get('quarters') or []:
        if quarter.get('home') is None or quarter.get('away') is None:
            break
        count += 1
    # A quarter with a partial score can appear as known; do not treat it as
    # completed unless the elapsed clock crossed the boundary.
    elapsed = float(canonical.get('elapsed_game_seconds') or 0.0)
    qsec = float(canonical.get('quarter_seconds') or 600.0)
    return min(count, int(elapsed // qsec))


def _v11_relevant_market_specs(canonical: dict[str, Any]) -> list[dict[str, Any]]:
    cp = _v11_completed_quarters(canonical)
    home, away = canonical['home_team'], canonical['away_team']
    specs: list[dict[str, Any]] = []

    def add(market_type: str, segment: str, team: Optional[str] = None) -> None:
        specs.append({'market_type': market_type, 'segment': segment, 'team': team})

    next_q = cp + 1 if cp in {1, 2, 3} else canonical.get('current_quarter')
    if next_q in {2, 3, 4}:
        add('CURRENT_QUARTER_TOTAL', f'Q{next_q}')
        add('CURRENT_QUARTER_TEAM_IT', f'Q{next_q}', home)
        add('CURRENT_QUARTER_TEAM_IT', f'Q{next_q}', away)
    if cp <= 1:
        add('H1_TOTAL', 'H1')
        add('TEAM_IT_H1', 'H1', home)
        add('TEAM_IT_H1', 'H1', away)
    if cp <= 2:
        add('H2_TOTAL', 'H2')
        add('TEAM_IT_H2', 'H2', home)
        add('TEAM_IT_H2', 'H2', away)
    add('MATCH_TOTAL', 'MATCH')
    add('TEAM_IT_MATCH', 'MATCH', home)
    add('TEAM_IT_MATCH', 'MATCH', away)
    return specs


def _v11_market_key(item: dict[str, Any]) -> tuple[str, Optional[str], str]:
    return (str(item.get('market_type') or ''), item.get('team'), str(item.get('segment') or ''))


def _v11_market_label(item: dict[str, Any]) -> str:
    labels = {
        'MATCH_TOTAL': 'Тотал матчу',
        'H1_TOTAL': 'Тотал 1-ї половини',
        'H2_TOTAL': 'Тотал 2-ї половини',
        'CURRENT_QUARTER_TOTAL': f"Тотал {item.get('segment')}",
        'TEAM_IT_MATCH': 'Індивідуальний тотал матчу',
        'TEAM_IT_H1': 'Індивідуальний тотал 1H',
        'TEAM_IT_H2': 'Індивідуальний тотал 2H',
        'CURRENT_QUARTER_TEAM_IT': f"Індивідуальний тотал {item.get('segment')}",
    }
    label = labels.get(str(item.get('market_type')), str(item.get('market_type')))
    if item.get('team'):
        label += f" — {item['team']}"
    return label


def _v11_segment_indices(segment: str) -> list[int]:
    if segment == 'MATCH':
        return [0, 1, 2, 3]
    if segment == 'H1':
        return [0, 1]
    if segment == 'H2':
        return [2, 3]
    if segment.startswith('Q') and segment[1:].isdigit():
        idx = int(segment[1:]) - 1
        return [idx] if 0 <= idx <= 3 else []
    return []


def _v11_view_game(game: dict[str, Any], perspective: str = 'team') -> dict[str, Any]:
    if perspective == 'team':
        tq = list(game.get('team_quarters') or [])
        oq = list(game.get('opponent_quarters') or [])
        team_score = to_number(game.get('team_score'))
        opp_score = to_number(game.get('opponent_score'))
    else:
        tq = list(game.get('opponent_quarters') or [])
        oq = list(game.get('team_quarters') or [])
        team_score = to_number(game.get('opponent_score'))
        opp_score = to_number(game.get('team_score'))
    return {
        'id': game.get('id'),
        'team_q': tq,
        'opp_q': oq,
        'team_score': team_score,
        'opp_score': opp_score,
        'total': to_number(game.get('total')),
    }


def _v11_sum(values: list[Optional[float]], indices: list[int]) -> Optional[float]:
    selected = [to_number(values[i]) if i < len(values) else None for i in indices]
    if not selected or any(v is None for v in selected):
        return None
    return float(sum(selected))


def _v11_outcome_from_view(
    view: dict[str, Any],
    market: dict[str, Any],
    source_team: str,
    target_team: Optional[str],
) -> Optional[float]:
    indices = _v11_segment_indices(str(market.get('segment') or 'MATCH'))
    is_team_market = str(market.get('market_type') or '').startswith('TEAM_IT') or market.get('market_type') == 'CURRENT_QUARTER_TEAM_IT'
    if not is_team_market:
        tq = _v11_sum(view['team_q'], indices)
        oq = _v11_sum(view['opp_q'], indices)
        return None if tq is None or oq is None else tq + oq
    target_is_source = target_team == source_team
    values = view['team_q'] if target_is_source else view['opp_q']
    if str(market.get('segment')) == 'MATCH':
        return view['team_score'] if target_is_source else view['opp_score']
    return _v11_sum(values, indices)


def _v11_hit(value: Optional[float], line: float, side: str) -> Optional[bool]:
    if value is None:
        return None
    if value == line:
        return None
    return value > line if side == 'OVER' else value < line


def _v11_margin_bucket(value: float) -> tuple[int, int]:
    absolute = abs(value)
    if absolute <= 3:
        return (0, 3)
    if absolute <= 5:
        return (4, 5)
    if absolute <= 10:
        return (6, 10)
    if absolute <= 15:
        return (11, 15)
    if absolute <= 20:
        return (16, 20)
    return (21, 999)


def _v11_active_scenario_conditions(canonical: dict[str, Any], source_team: str) -> list[dict[str, Any]]:
    home = source_team == canonical['home_team']
    side = 'home' if home else 'away'
    opp_side = 'away' if home else 'home'
    completed = _v11_completed_quarters(canonical)
    qrows = canonical.get('quarters') or []
    tq = [to_number(q.get(side)) for q in qrows]
    oq = [to_number(q.get(opp_side)) for q in qrows]
    conditions: list[dict[str, Any]] = []

    def add(pid: str, title: str, group: str, matcher: Callable[[dict[str, Any]], bool], description: str) -> None:
        conditions.append({'scenario_id': pid, 'title': title, 'group': group, 'matcher': matcher, 'description': description})

    for i in range(min(completed, 3)):
        if tq[i] is None or oq[i] is None:
            continue
        qn = i + 1
        team_points = float(tq[i])
        opp_points = float(oq[i])
        if team_points > opp_points:
            add(
                f'WON_Q{qn}', f'{source_team} виграла Q{qn}', f'Q{qn}_RESULT',
                lambda v, idx=i: v['team_q'][idx] is not None and v['opp_q'][idx] is not None and v['team_q'][idx] > v['opp_q'][idx],
                f'{source_team} виграла {qn}-ту чверть {team_points:g}:{opp_points:g}.',
            )
        elif team_points < opp_points:
            add(
                f'LOST_Q{qn}', f'{source_team} програла Q{qn}', f'Q{qn}_RESULT',
                lambda v, idx=i: v['team_q'][idx] is not None and v['opp_q'][idx] is not None and v['team_q'][idx] < v['opp_q'][idx],
                f'{source_team} програла {qn}-ту чверть {team_points:g}:{opp_points:g}.',
            )
        else:
            add(
                f'TIED_Q{qn}', f'{source_team} зіграла Q{qn} внічию', f'Q{qn}_RESULT',
                lambda v, idx=i: v['team_q'][idx] is not None and v['opp_q'][idx] is not None and v['team_q'][idx] == v['opp_q'][idx],
                f'{source_team} завершила {qn}-ту чверть унічию.',
            )
        for threshold in (18, 21, 24, 27):
            if team_points >= threshold:
                add(
                    f'Q{qn}_SCORED_{threshold}_PLUS', f'{source_team} набрала {threshold}+ у Q{qn}', f'Q{qn}_POINTS',
                    lambda v, idx=i, t=threshold: v['team_q'][idx] is not None and v['team_q'][idx] >= t,
                    f'{source_team} набрала {team_points:g} у Q{qn}, тобто {threshold}+.',
                )
        if team_points <= 18:
            add(
                f'Q{qn}_SCORED_18_OR_LESS', f'{source_team} набрала ≤18 у Q{qn}', f'Q{qn}_POINTS_LOW',
                lambda v, idx=i: v['team_q'][idx] is not None and v['team_q'][idx] <= 18,
                f'{source_team} набрала лише {team_points:g} у Q{qn}.',
            )
        qtotal = team_points + opp_points
        if qtotal >= 45:
            add(
                f'Q{qn}_TOTAL_45_PLUS', f'Q{qn} завершилася на 45+', f'Q{qn}_TOTAL',
                lambda v, idx=i: None not in (v['team_q'][idx], v['opp_q'][idx]) and v['team_q'][idx] + v['opp_q'][idx] >= 45,
                f'Загальний тотал Q{qn} становив {qtotal:g}.',
            )
        if qtotal <= 39:
            add(
                f'Q{qn}_TOTAL_39_OR_LESS', f'Q{qn} завершилася на ≤39', f'Q{qn}_TOTAL',
                lambda v, idx=i: None not in (v['team_q'][idx], v['opp_q'][idx]) and v['team_q'][idx] + v['opp_q'][idx] <= 39,
                f'Загальний тотал Q{qn} становив лише {qtotal:g}.',
            )
        if qn == 1 and team_points > opp_points and team_points >= 24:
            add(
                'WON_Q1_AND_SCORED_24_PLUS', f'{source_team} виграла Q1 і набрала 24+', 'Q1_COMBO',
                lambda v: None not in (v['team_q'][0], v['opp_q'][0]) and v['team_q'][0] > v['opp_q'][0] and v['team_q'][0] >= 24,
                f'{source_team} виграла Q1 та набрала {team_points:g} очок.',
            )
        if qn == 1 and team_points < opp_points and team_points <= 18:
            add(
                'LOST_Q1_AND_SCORED_18_OR_LESS', f'{source_team} програла Q1 і набрала ≤18', 'Q1_COMBO_LOW',
                lambda v: None not in (v['team_q'][0], v['opp_q'][0]) and v['team_q'][0] < v['opp_q'][0] and v['team_q'][0] <= 18,
                f'{source_team} програла Q1 та набрала лише {team_points:g}.',
            )

    if completed >= 2 and all(v is not None for v in tq[:2] + oq[:2]):
        wins = [tq[i] > oq[i] for i in range(2)]
        if all(wins):
            add('LEADS_2_0', f'{source_team} веде по чвертях 2–0', 'SEQUENCE_2Q', lambda v: all(None not in (v['team_q'][i], v['opp_q'][i]) and v['team_q'][i] > v['opp_q'][i] for i in range(2)), f'{source_team} виграла Q1 і Q2.')
        elif not any(wins):
            add('TRAILS_0_2', f'{source_team} програє по чвертях 0–2', 'SEQUENCE_2Q', lambda v: all(None not in (v['team_q'][i], v['opp_q'][i]) and v['team_q'][i] < v['opp_q'][i] for i in range(2)), f'{source_team} програла Q1 і Q2.')
        elif wins == [True, False]:
            add('WON_Q1_LOST_Q2', f'{source_team} виграла Q1, але програла Q2', 'SEQUENCE_2Q', lambda v: None not in (v['team_q'][0], v['opp_q'][0], v['team_q'][1], v['opp_q'][1]) and v['team_q'][0] > v['opp_q'][0] and v['team_q'][1] < v['opp_q'][1], 'Після виграної Q1 команда віддала Q2.')
        else:
            add('LOST_Q1_WON_Q2', f'{source_team} програла Q1, але виграла Q2', 'SEQUENCE_2Q', lambda v: None not in (v['team_q'][0], v['opp_q'][0], v['team_q'][1], v['opp_q'][1]) and v['team_q'][0] < v['opp_q'][0] and v['team_q'][1] > v['opp_q'][1], 'Після програної Q1 команда виграла Q2.')
        if all(float(v) < 21 for v in tq[:2]):
            add('NO_21_IN_Q1_Q2', f'{source_team} не набрала 21 у Q1 і Q2', 'POINT_SEQUENCE_2Q', lambda v: all(v['team_q'][i] is not None and v['team_q'][i] < 21 for i in range(2)), f'{source_team} не дійшла до 21 очка у двох перших чвертях.')

    if completed >= 3 and all(v is not None for v in tq[:3] + oq[:3]):
        won_count = sum(tq[i] > oq[i] for i in range(3))
        if won_count == 3:
            add('LEADS_3_0', f'{source_team} веде по чвертях 3–0', 'SEQUENCE_3Q', lambda v: all(None not in (v['team_q'][i], v['opp_q'][i]) and v['team_q'][i] > v['opp_q'][i] for i in range(3)), f'{source_team} виграла перші три чверті.')
        elif won_count == 0:
            add('TRAILS_0_3', f'{source_team} програє по чвертях 0–3', 'SEQUENCE_3Q', lambda v: all(None not in (v['team_q'][i], v['opp_q'][i]) and v['team_q'][i] < v['opp_q'][i] for i in range(3)), f'{source_team} програла перші три чверті.')
        elif won_count == 2:
            add('WON_2_OF_3', f'{source_team} виграла 2 із 3 чвертей', 'SEQUENCE_3Q', lambda v: sum(None not in (v['team_q'][i], v['opp_q'][i]) and v['team_q'][i] > v['opp_q'][i] for i in range(3)) == 2, f'{source_team} виграла дві з трьох завершених чвертей.')
        else:
            add('LOST_2_OF_3', f'{source_team} програла 2 із 3 чвертей', 'SEQUENCE_3Q', lambda v: sum(None not in (v['team_q'][i], v['opp_q'][i]) and v['team_q'][i] < v['opp_q'][i] for i in range(3)) == 2, f'{source_team} програла дві з трьох завершених чвертей.')
        if all(float(v) < 21 for v in tq[:3]):
            add('NO_21_IN_FIRST_3Q', f'{source_team} не набрала 21 у жодній із Q1–Q3', 'POINT_SEQUENCE_3Q', lambda v: all(v['team_q'][i] is not None and v['team_q'][i] < 21 for i in range(3)), f'{source_team} жодного разу не набрала 21 у перших трьох чвертях.')

    if completed >= 1 and all(v is not None for v in tq[:completed] + oq[:completed]):
        margin = float(sum(tq[:completed]) - sum(oq[:completed]))
        low, high = _v11_margin_bucket(margin)
        sign = 1 if margin >= 0 else -1
        add(
            f'MARGIN_{"LEAD" if sign > 0 else "TRAIL"}_{low}_{high}',
            f'{source_team}: відрив {abs(margin):g} після {completed}Q',
            f'MARGIN_{completed}Q',
            lambda v, n=completed, lo=low, hi=high, s=sign: (
                all(None not in (v['team_q'][i], v['opp_q'][i]) for i in range(n))
                and (1 if sum(v['team_q'][:n]) - sum(v['opp_q'][:n]) >= 0 else -1) == s
                and lo <= abs(sum(v['team_q'][:n]) - sum(v['opp_q'][:n])) <= hi
            ),
            f'Після {completed} чвертей різниця для {source_team}: {margin:+g}.',
        )

    # Keep one exact id per logical pattern.
    unique: dict[str, dict[str, Any]] = {}
    for condition in conditions:
        unique.setdefault(condition['scenario_id'], condition)
    return list(unique.values())


def _v11_scenario_sample_stats(
    games: list[dict[str, Any]],
    perspective: str,
    condition: dict[str, Any],
    market: dict[str, Any],
    source_team: str,
    target_team: Optional[str],
) -> dict[str, Any]:
    matched: list[tuple[dict[str, Any], dict[str, Any]]] = []
    for game in games:
        view = _v11_view_game(game, perspective)
        try:
            ok = bool(condition['matcher'](view))
        except (IndexError, TypeError, ValueError):
            ok = False
        if ok:
            matched.append((game, view))
    line = float(market['line'])
    side = str(market['side'])
    outcomes: list[float] = []
    hits = 0
    pushes = 0
    for _, view in matched:
        value = _v11_outcome_from_view(view, market, source_team, target_team)
        if value is None:
            continue
        outcomes.append(float(value))
        hit = _v11_hit(float(value), line, side)
        if hit is True:
            hits += 1
        elif hit is None:
            pushes += 1
    n = len(outcomes)
    raw = hits / n if n else None
    smooth = smoothed_probability(hits, n) if n else None

    completed = None
    scenario_id = str(condition.get('scenario_id') or '')
    if '3_0' in scenario_id or '3Q' in str(condition.get('group')):
        completed = 3
    elif '2_0' in scenario_id or '2Q' in str(condition.get('group')):
        completed = 2
    elif 'Q1' in scenario_id:
        completed = 1
    next_q_win = None
    swept_4_0 = None
    no_21_all_4 = None
    match_win = None
    if matched:
        match_values = []
        next_values = []
        sweep_values = []
        no21_values = []
        for _, view in matched:
            if view['team_score'] is not None and view['opp_score'] is not None:
                match_values.append(view['team_score'] > view['opp_score'])
            if completed and completed < 4 and len(view['team_q']) > completed and len(view['opp_q']) > completed:
                a, b = view['team_q'][completed], view['opp_q'][completed]
                if a is not None and b is not None:
                    next_values.append(a > b)
            if len(view['team_q']) >= 4 and len(view['opp_q']) >= 4 and all(None not in (view['team_q'][i], view['opp_q'][i]) for i in range(4)):
                sweep_values.append(all(view['team_q'][i] > view['opp_q'][i] for i in range(4)))
                no21_values.append(all(view['team_q'][i] < 21 for i in range(4)))
        match_win = sum(match_values) / len(match_values) if match_values else None
        next_q_win = sum(next_values) / len(next_values) if next_values else None
        swept_4_0 = sum(sweep_values) / len(sweep_values) if sweep_values else None
        no_21_all_4 = sum(no21_values) / len(no21_values) if no21_values else None

    return {
        'matched_games': len(matched),
        'n': n,
        'hits': hits,
        'pushes': pushes,
        'raw_rate': raw,
        'smoothed_rate': smooth,
        'outcome_median': statistics.median(outcomes) if outcomes else None,
        'team_won_match_rate': match_win,
        'won_next_quarter_rate': next_q_win,
        'won_all_4_quarters_rate': swept_4_0,
        'under_21_all_4_quarters_rate': no_21_all_4,
    }


def _v11_sample_label(n: int) -> str:
    if n < 5:
        return 'INSUFFICIENT'
    if n < 8:
        return 'SMALL_SAMPLE'
    if n < 15:
        return 'NORMAL'
    if n < 25:
        return 'RELIABLE'
    return 'STRONG_SAMPLE'


def _v11_effect(rate: Optional[float], n: int) -> str:
    if rate is None or n < 5:
        return 'INSUFFICIENT'
    if rate >= 0.80 and n >= 8:
        return 'STRONG_SUPPORT'
    if rate >= 0.70:
        return 'SUPPORT'
    if rate <= 0.30:
        return 'STRONG_CONFLICT'
    if rate < 0.45:
        return 'WEAKEN'
    return 'NEUTRAL'


def _v11_mine_scenarios(market: dict[str, Any], canonical: dict[str, Any]) -> dict[str, Any]:
    rows: list[dict[str, Any]] = []
    target_team = market.get('team')
    for source_team, own_key, opp_key in (
        (canonical['home_team'], 'team_a', 'team_b'),
        (canonical['away_team'], 'team_b', 'team_a'),
    ):
        conditions = _v11_active_scenario_conditions(canonical, source_team)
        own_pool = canonical.get('history', {}).get(own_key) or []
        opponent_pool = canonical.get('history', {}).get(opp_key) or []
        for condition in conditions:
            own = _v11_scenario_sample_stats(own_pool, 'team', condition, market, source_team, target_team)
            allowed = _v11_scenario_sample_stats(opponent_pool, 'opponent', condition, market, source_team, target_team)
            total_n = int(own['n']) + int(allowed['n'])
            if total_n < 5:
                continue
            own_rate = own.get('smoothed_rate')
            allowed_rate = allowed.get('smoothed_rate')
            weighted = None
            if own_rate is not None or allowed_rate is not None:
                numerator = (float(own_rate or 0.0) * int(own['n'])) + (float(allowed_rate or 0.0) * int(allowed['n']))
                denominator = int(own['n']) + int(allowed['n'])
                weighted = numerator / denominator if denominator else None
            if own_rate is not None and allowed_rate is not None:
                same_direction = (own_rate >= 0.5) == (allowed_rate >= 0.5)
                intersection = 'ALIGNED' if same_direction else 'CONFLICT'
            elif own_rate is not None or allowed_rate is not None:
                intersection = 'ONE_SIDED'
            else:
                intersection = 'OFF'
            effect = _v11_effect(weighted, total_n)
            credibility = min(1.0, total_n / 15.0)
            rank = abs(float(weighted or 0.5) - 0.5) * credibility
            rows.append({
                'scenario_id': condition['scenario_id'],
                'title': condition['title'],
                'description': condition['description'],
                'group': condition['group'],
                'source_team': source_team,
                'target_market': _v11_market_label(market),
                'target_side': market.get('side'),
                'target_line': market.get('line'),
                'own': own,
                'opponent_allowed': allowed,
                'combined_n': total_n,
                'combined_rate': weighted,
                'intersection': intersection,
                'effect': effect,
                'sample_label': _v11_sample_label(total_n),
                'rank': rank,
            })
    # Do not count overlapping variants from one family as independent. Keep the
    # strongest result per source team and group.
    best: dict[tuple[str, str], dict[str, Any]] = {}
    for row in rows:
        key = (str(row['source_team']), str(row['group']))
        if key not in best or float(row['rank']) > float(best[key]['rank']):
            best[key] = row
    selected = sorted(best.values(), key=lambda r: (float(r['rank']), int(r['combined_n'])), reverse=True)
    support = [r for r in selected if r['effect'] in {'STRONG_SUPPORT', 'SUPPORT'}]
    conflict = [r for r in selected if r['effect'] in {'STRONG_CONFLICT', 'WEAKEN'}]
    used = support[:3] + conflict[:1]
    if used:
        weights = [max(0.01, float(r['rank'])) for r in used]
        scenario_probability = sum(float(r['combined_rate']) * w for r, w in zip(used, weights)) / sum(weights)
        scenario_n = max(int(r['combined_n']) for r in used)
    else:
        scenario_probability = 0.50
        scenario_n = 0
    return {
        'p_scenario_miner': scenario_probability,
        'scenario_n': scenario_n,
        'support_count': len(support),
        'conflict_count': len(conflict),
        'top_support': support[:3],
        'top_conflict': conflict[:2],
        'all_selected_patterns': selected,
    }


def _v11_history_values(canonical: dict[str, Any], spec: dict[str, Any]) -> list[float]:
    market = {**spec, 'side': 'OVER', 'line': 0.5}
    values: list[float] = []
    if spec.get('team'):
        team = spec['team']
        own_key = 'team_a' if team == canonical['home_team'] else 'team_b'
        opp_key = 'team_b' if own_key == 'team_a' else 'team_a'
        for game in canonical.get('history', {}).get(own_key) or []:
            value = _segment_value(game, market, team)
            if value is not None:
                values.extend([float(value)] * 5)
        for game in canonical.get('history', {}).get(opp_key) or []:
            value = _segment_value(game, market, team, opponent_allowed=True)
            if value is not None:
                values.extend([float(value)] * 3)
    else:
        for key in ('team_a', 'team_b'):
            for game in canonical.get('history', {}).get(key) or []:
                value = _segment_value(game, market)
                if value is not None:
                    values.append(float(value))
    return values


def _v11_model_center(canonical: dict[str, Any], spec: dict[str, Any]) -> Optional[float]:
    values = _v11_history_values(canonical, spec)
    if not values:
        return None
    return _v11_round_half(statistics.median(values))


def _v11_synthetic_market(spec: dict[str, Any], line: float, side: str, *, tag: str = 'MODEL') -> dict[str, Any]:
    stable = f"{spec.get('market_type')}|{spec.get('team')}|{spec.get('segment')}|{side}|{line}|{tag}"
    return {
        'market_id': 'V11-' + hashlib.sha256(stable.encode('utf-8')).hexdigest()[:16],
        'source_market_id': None,
        'market_type': spec['market_type'],
        'team': spec.get('team'),
        'segment': spec['segment'],
        'side': side,
        'line': float(line),
        'odds': ADVISOR_MODEL_ODDS,
        'bookmaker': 'MODEL',
        'source_bucket': 'advisor_model',
        'source_scope': spec['segment'],
        'raw_line_row': {'is_real_bookmaker_line': False, 'advisor_model': True},
        'parser_issues': [],
        'eligible_market': True,
        'is_model_line': True,
    }


def _v11_serious_blockers(evaluation: dict[str, Any], *, exceptional_edge: bool = False) -> list[str]:
    ignored = {
        'HISTORY_ZONE_BELOW_75',
        'LIVE_EDGE_BELOW_3',
        'LIVE_EDGE_BELOW_0_50',
        'Q3_EXCEPTIONAL_PROBABILITY_BELOW_80',
        'Q3_EXCEPTIONAL_STATS_REQUIRED',
        'PRODUCTION_ROUTER_BLOCK',
    }
    if exceptional_edge:
        ignored.add('SCENARIO_DIRECTION_CONFLICT')
    serious = []
    for blocker in evaluation.get('blockers') or []:
        code = str(blocker.get('rule_id') or '')
        if code not in ignored:
            serious.append(code)
    return serious


def _v11_enrich_evaluation(evaluation: dict[str, Any], canonical: dict[str, Any], *, is_model: bool = False, mine_scenarios: bool = True) -> dict[str, Any]:
    item = evaluation
    history = item.get('history') or {}
    live = item.get('live') or {}
    stat = item.get('stat_comparison') or {}
    zone = to_number(history.get('history_zone_rate'))
    p_final = float(item.get('p_final_system') or item.get('p_final') or 0.50)
    edge = to_number(live.get('line_edge'))
    p_live = float(live.get('p_live') or 0.50)
    p_hist = float(history.get('p_hist') or 0.50)
    p_scenario = float((item.get('scenario') or {}).get('p_scenario') or 0.50)
    exceptional = edge is not None and edge >= ADVISOR_EXCEPTIONAL_EDGE_MIN
    serious = _v11_serious_blockers(item, exceptional_edge=exceptional)
    if str(stat.get('stat_gate_status') or '') == 'AGAINST' and 'STAT_GATE_AGAINST' not in serious:
        serious.append('STAT_GATE_AGAINST')
    fake_against = bool(
        (item.get('side') == 'OVER' and stat.get('fake_over'))
        or (item.get('side') == 'UNDER' and stat.get('fake_under'))
    )
    odds = to_number(item.get('odds'))
    edge_ok = edge is not None and edge >= 0.0
    zone_ok = zone is not None and zone >= ADVISOR_HISTORY_ZONE_MIN
    scenario_miner = _v11_mine_scenarios(item, canonical) if mine_scenarios else {
        'p_scenario_miner': p_scenario, 'scenario_n': 0, 'support_count': 0,
        'conflict_count': 0, 'top_support': [], 'top_conflict': [],
        'all_selected_patterns': [], 'deferred': True,
    }

    if is_model:
        action = 'PASS'
        status = 'MODEL LINE / THEORETICAL TRIGGER'
    elif odds is None or odds < float(DEFAULT_CONFIG.get('odds_min', 1.44)):
        action, status = 'PASS', 'PASS — ODDS'
    elif serious or fake_against or not edge_ok:
        action, status = 'PASS', 'PASS — CONFLICT/BLOCKER'
    elif (zone_ok or exceptional) and p_final >= ADVISOR_PLAY_MIN:
        action, status = 'PLAY', 'PLAY — ADVISOR CLEAN'
    elif (zone_ok or exceptional) and p_final >= ADVISOR_RISK_MIN:
        action, status = 'RISK', 'RISK — ADVISOR'
    else:
        action, status = 'PASS', 'PASS — ADVISOR'

    directions = []
    directions.append('HISTORY_' + ('SUPPORT' if p_hist >= 0.60 else 'AGAINST' if p_hist <= 0.40 else 'NEUTRAL'))
    directions.append('SCENARIO_' + ('SUPPORT' if p_scenario >= 0.60 else 'AGAINST' if p_scenario <= 0.40 else 'NEUTRAL'))
    directions.append('LIVE_' + ('SUPPORT' if p_live >= 0.60 and edge_ok else 'AGAINST' if p_live <= 0.40 or (edge is not None and edge < 0) else 'NEUTRAL'))
    support_n = sum(code.endswith('SUPPORT') for code in directions)
    against_n = sum(code.endswith('AGAINST') for code in directions)
    if support_n == 3:
        alignment = 'TRIPLE ALIGNED'
    elif against_n and support_n:
        alignment = 'HARD CONFLICT' if against_n >= 1 and support_n >= 2 else 'SOFT CONFLICT'
    elif support_n >= 2:
        alignment = 'PARTIAL ALIGNED'
    else:
        alignment = 'NEUTRAL'

    item['advisor'] = {
        'version': ADVISOR_VERSION,
        'is_model_line': is_model,
        'history_zone_rate': zone,
        'history_zone_eligible': zone_ok,
        'exceptional_edge': exceptional,
        'exceptional_edge_min': ADVISOR_EXCEPTIONAL_EDGE_MIN,
        'telegram_line_eligible': zone_ok or exceptional,
        'p_hist': p_hist,
        'p_scenario_core': p_scenario,
        'p_scenario_miner': scenario_miner['p_scenario_miner'],
        'p_live': p_live,
        'p_final': p_final,
        'projection_used': live.get('projection_used'),
        'line_edge': edge,
        'alignment': alignment,
        'direction_components': directions,
        'fake_over': bool(stat.get('fake_over')),
        'fake_under': bool(stat.get('fake_under')),
        'stat_gate_status': stat.get('stat_gate_status'),
        'serious_blockers': serious,
        'scenario_miner': scenario_miner,
        'action': action,
        'status': status,
    }
    item['system_action'] = action
    item['system_status'] = status
    item['stake'] = '0%' if action == 'PASS' else ('10-15% live-limit' if action == 'RISK' else '15-20% live-limit')
    item['p_final_system'] = p_final
    return item


def _v11_light_model_evaluation(
    spec: dict[str, Any],
    line: float,
    side: str,
    projection: float,
    values: list[float],
    canonical: dict[str, Any],
    probe: Optional[dict[str, Any]],
) -> dict[str, Any]:
    wins = sum(1 for value in values if _v11_hit(value, line, side) is True)
    pushes = sum(1 for value in values if _v11_hit(value, line, side) is None)
    n = len(values)
    raw = wins / n if n else None
    p_hist = smoothed_probability(wins, n) if n else 0.50
    sigma = _stage_sigma(str(spec.get('market_type')), str(canonical.get('stage')), DEFAULT_CONFIG)
    if canonical.get('data_mode') == 'SCORE_TIME_HISTORY':
        sigma *= 1.20
    edge = projection - line if side == 'OVER' else line - projection
    p_live = normal_cdf(edge / max(0.1, sigma))
    # Preliminary model grid uses the exact-line history as a conservative
    # scenario prior. The full Scenario Miner is run only for selected lines.
    p_scenario = 0.50 + 0.50 * (p_hist - 0.50)
    market = _v11_synthetic_market(spec, line, side)
    weights = _v9_stage_weights(market, canonical, DEFAULT_CONFIG)
    p_raw = weights['hist'] * p_hist + weights['scenario'] * p_scenario + weights['live'] * p_live
    cap = 1.0
    if n < 15:
        cap = min(cap, 0.72)
    elif n < 20:
        cap = min(cap, 0.74)
    if canonical.get('data_mode') == 'SCORE_TIME_HISTORY':
        cap = min(cap, 0.79)
    stat = deepcopy((probe or {}).get('stat_comparison') or {
        'stat_gate_status': 'OFF', 'fake_over': False, 'fake_under': False,
    })
    if str(stat.get('stat_gate_status') or '') == 'AGAINST':
        cap = min(cap, 0.67)
    p_final = min(p_raw, cap)
    history = {
        'p_hist': p_hist,
        'history_zone_rate': raw,
        'history_zone_hits': wins,
        'history_zone_n': n,
        'history_zone_source': 'V11_MODEL_EXACT_HISTORY',
        'pooled': {'wins': wins, 'pushes': pushes, 'n': n, 'raw_pct': raw, 'p_smoothed': p_hist},
    }
    return {
        **market,
        'history': history,
        'scenario': {
            'p_scenario': p_scenario,
            'scenario_support': 'MODEL_GRID_PRIOR',
            'patterns_found': [], 'patterns_used': [], 'patterns_rejected': [],
        },
        'live': {
            'projection_used': projection,
            'line_edge': edge,
            'p_live': p_live,
            'projection_source': 'V11_MODEL_HISTORY_LIVE_CENTER',
        },
        'stat_comparison': stat,
        'q4_context': deepcopy((probe or {}).get('q4_context') or {'applicable': False, 'status': 'OFF'}),
        'weights': {'normalized': weights},
        'p_raw': p_raw,
        'p_final': p_final,
        'p_final_system': p_final,
        'router': _router(market, canonical),
        'caps': ([{'rule_id': 'V11_MODEL_CAP', 'cap': cap, 'reason': 'Model-line sample/data cap'}] if cap < 1.0 else []),
        'blockers': [],
        'parser_issues': [],
        'is_model_line': True,
    }


def _v11_evaluate_model_grid(
    calculator: SuperBasketCalculator,
    canonical: dict[str, Any],
    spec: dict[str, Any],
) -> list[dict[str, Any]]:
    values = _v11_history_values(canonical, spec)
    center = _v11_model_center(canonical, spec)
    if center is None or not values:
        return []
    # One full probe per market gives live/stat context. The line grid itself is
    # evaluated with a lightweight exact-history/Phi calculation.
    probe_market = _v11_synthetic_market(spec, center, 'OVER', tag='PROBE')
    probe = calculator.evaluate_market(probe_market, canonical)
    core_projection = to_number((probe.get('live') or {}).get('projection_used'))
    projection = core_projection
    if projection is None or projection <= 0 or projection < center * 0.55 or projection > center * 1.65:
        projection = center

    p10 = percentile(values, 0.10) or min(values)
    p90 = percentile(values, 0.90) or max(values)
    realistic_low = max(0.5, _v11_round_half(p10 - 2.0))
    realistic_high = _v11_round_half(p90 + 2.0)

    candidate_lines: set[tuple[str, float]] = set()
    candidate_lines.add(('OVER', _v11_round_half(projection)))
    candidate_lines.add(('UNDER', _v11_round_half(projection)))
    for offset in ADVISOR_MODEL_OFFSETS:
        candidate_lines.add(('OVER', _v11_round_half(projection - offset)))
        candidate_lines.add(('UNDER', _v11_round_half(projection + offset)))
    evaluations: list[dict[str, Any]] = []
    for side, line in sorted(candidate_lines, key=lambda row: (row[0], row[1])):
        if line < realistic_low or line > realistic_high:
            continue
        evaluated = _v11_light_model_evaluation(spec, line, side, float(projection), values, canonical, probe)
        evaluations.append(_v11_enrich_evaluation(evaluated, canonical, is_model=True, mine_scenarios=False))
    return evaluations

def _v112_number(value: Any, default: float) -> float:
    number = to_number(value)
    return default if number is None else float(number)


def _v112_edge(item: dict[str, Any], default: float = -999.0) -> float:
    return _v112_number((item.get('advisor') or {}).get('line_edge'), default)


def _v112_market_is_currently_supported(item: dict[str, Any]) -> bool:
    """Exclude settled/invalid markets from dispatch and primary selection."""
    terminal_issues = {
        'NO_LINE', 'SYNTHETIC_LINE', 'UNSUPPORTED_MARKET', 'UNKNOWN_QUARTER',
        'INVALID_QUARTER', 'PAST_QUARTER', 'FUTURE_QUARTER',
        'NO_CURRENT_QUARTER', 'NO_EXACT_CURRENT_QUARTER_TIME',
        'NO_CURRENT_QUARTER_SCORE',
    }
    if any(str(code) in terminal_issues for code in item.get('parser_issues') or []):
        return False
    router = item.get('router') or {}
    if str(router.get('status') or '').upper() == 'BLOCK' and bool(router.get('hard_block', True)):
        return False
    return item.get('line') is not None


def _v112_input_state_gate(source: dict[str, Any], canonical: dict[str, Any], checkpoint: Optional[int]) -> dict[str, Any]:
    """Prevent final-score/future-checkpoint leakage from stale parser files."""
    match = source.get('match') if isinstance(source.get('match'), dict) else {}
    raw_stage = str(match.get('stage') or source.get('stage') or source.get('status') or '').upper().strip()
    finished = raw_stage in {'FT', 'FINAL', 'FINISHED', 'ENDED', 'COMPLETED', 'AFTER_OT'}
    current = to_int(canonical.get('current_quarter'))
    expected = checkpoint + 1 if checkpoint in {1, 2, 3} else None
    stale = bool(expected is not None and current is not None and current > expected)
    allowed = not finished and not stale
    if finished:
        reason = 'MATCH_ALREADY_FINISHED'
    elif stale:
        reason = f'CHECKPOINT_STALE: expected Q{expected}, source already at Q{current}'
    else:
        reason = 'OK'
    return {
        'allowed': allowed,
        'finished': finished,
        'stale_checkpoint': stale,
        'reason': reason,
        'raw_stage': raw_stage or None,
        'checkpoint': checkpoint,
        'current_quarter': current,
        'expected_quarter': expected,
    }


def _v11_model_summary(model_evaluations: list[dict[str, Any]], canonical: dict[str, Any]) -> list[dict[str, Any]]:
    groups: dict[tuple[str, Optional[str], str], list[dict[str, Any]]] = {}
    for item in model_evaluations:
        groups.setdefault(_v11_market_key(item), []).append(item)
    output: list[dict[str, Any]] = []
    for key, rows in groups.items():
        projection = next((to_number((r.get('live') or {}).get('projection_used')) for r in rows if to_number((r.get('live') or {}).get('projection_used')) is not None), None)
        side_rows: dict[str, list[dict[str, Any]]] = {'OVER': [], 'UNDER': []}
        for row in rows:
            side_rows[str(row.get('side'))].append(row)
        recommendation: dict[str, Any] = {'market_key': list(key), 'projection_used': projection, 'market_label': _v11_market_label(rows[0])}
        for side in ('OVER', 'UNDER'):
            candidates = sorted(
                side_rows[side],
                key=lambda r: (
                    abs(_v112_edge(r, 999.0)),
                    -float((r.get('advisor') or {}).get('p_final') or 0.0),
                ),
            )
            risk_pool = [r for r in candidates if (r.get('advisor') or {}).get('p_final', 0) >= ADVISOR_RISK_MIN and (r.get('advisor') or {}).get('history_zone_eligible') and _v112_edge(r, -999.0) >= 0 and not (r.get('advisor') or {}).get('serious_blockers')]
            play_pool = [r for r in candidates if (r.get('advisor') or {}).get('p_final', 0) >= ADVISOR_PLAY_MIN and (r.get('advisor') or {}).get('history_zone_eligible') and _v112_edge(r, -999.0) >= 0 and not (r.get('advisor') or {}).get('serious_blockers')]
            # The nearest qualifying line is authoritative, not the easiest
            # absurdly distant line with the highest model probability.
            risk = min(risk_pool, key=lambda r: abs(_v112_edge(r, 999.0)), default=None)
            play = min(play_pool, key=lambda r: abs(_v112_edge(r, 999.0)), default=None)
            best = candidates[0] if candidates else None
            for selected_row in (best, risk, play):
                if selected_row and (selected_row.get('advisor') or {}).get('scenario_miner', {}).get('deferred'):
                    mined = _v11_mine_scenarios(selected_row, canonical)
                    selected_row['advisor']['scenario_miner'] = mined
                    selected_row['advisor']['p_scenario_miner'] = mined.get('p_scenario_miner')
            recommendation[side.lower()] = {
                'best_model': _v11_compact_line(best) if best else None,
                'risk_trigger': _v11_compact_line(risk) if risk else None,
                'play_trigger': _v11_compact_line(play) if play else None,
            }
        output.append(recommendation)
    def summary_rank(row: dict[str, Any]) -> tuple[float, float, float]:
        trigger, _ = _v11_best_model_trigger(row)
        if not trigger:
            return (0.0, -999.0, 0.0)
        is_play = 1.0 if float(trigger.get('p_final') or 0.0) >= ADVISOR_PLAY_MIN else 0.0
        return (is_play, -abs(_v112_number(trigger.get('line_edge'), 999.0)), float(trigger.get('p_final') or 0.0))
    output.sort(key=summary_rank, reverse=True)
    return output


def _v11_compact_line(item: Optional[dict[str, Any]]) -> Optional[dict[str, Any]]:
    if not item:
        return None
    adv = item.get('advisor') or {}
    hist = item.get('history') or {}
    return {
        'market_type': item.get('market_type'),
        'team': item.get('team'),
        'segment': item.get('segment'),
        'side': item.get('side'),
        'line': item.get('line'),
        'odds': item.get('odds'),
        'bookmaker': item.get('bookmaker'),
        'is_model_line': bool(adv.get('is_model_line')),
        'action': adv.get('action'),
        'status': adv.get('status'),
        'history_zone_rate': adv.get('history_zone_rate'),
        'history_zone_hits': hist.get('history_zone_hits'),
        'history_zone_n': hist.get('history_zone_n'),
        'p_hist': adv.get('p_hist'),
        'p_scenario_core': adv.get('p_scenario_core'),
        'p_scenario_miner': adv.get('p_scenario_miner'),
        'p_live': adv.get('p_live'),
        'p_raw': item.get('p_raw'),
        'p_final': adv.get('p_final'),
        'projection_used': adv.get('projection_used'),
        'line_edge': adv.get('line_edge'),
        'alignment': adv.get('alignment'),
        'fake_over': adv.get('fake_over'),
        'fake_under': adv.get('fake_under'),
        'stat_gate_status': adv.get('stat_gate_status'),
        'serious_blockers': adv.get('serious_blockers'),
        'scenario_miner': adv.get('scenario_miner'),
    }


def _v11_primary_sort(item: dict[str, Any]) -> tuple[float, ...]:
    adv = item.get('advisor') or {}
    action_rank = {'PLAY': 3.0, 'RISK': 2.0, 'PASS': 1.0}.get(str(adv.get('action')), 0.0)
    return (
        action_rank,
        float(adv.get('p_final') or 0.0),
        float(adv.get('history_zone_rate') or 0.0),
        _v112_number(adv.get('line_edge'), -999.0),
        float(item.get('odds') or 0.0),
    )


def _v11_select_advisor_lines(evaluations: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    qualifying = [
        item for item in evaluations
        if not (item.get('advisor') or {}).get('is_model_line')
        and _v112_market_is_currently_supported(item)
        and ((item.get('advisor') or {}).get('history_zone_eligible') or (item.get('advisor') or {}).get('exceptional_edge'))
    ]
    qualifying.sort(key=_v11_primary_sort, reverse=True)
    primary = qualifying[:ADVISOR_MAX_PRIMARY]
    return primary, qualifying


def _v11_theoretical_for_pass(
    calculator: SuperBasketCalculator,
    canonical: dict[str, Any],
    evaluation: Optional[dict[str, Any]],
) -> Optional[dict[str, Any]]:
    if not evaluation:
        return None
    projection = to_number((evaluation.get('live') or {}).get('projection_used'))
    if projection is None:
        return None
    spec = {k: evaluation.get(k) for k in ('market_type', 'team', 'segment')}
    side = str(evaluation.get('side') or '')
    rows = []
    for offset in ADVISOR_MODEL_OFFSETS:
        line = _v11_round_half(projection - offset if side == 'OVER' else projection + offset)
        if line <= 0:
            continue
        model = calculator.evaluate_market(_v11_synthetic_market(spec, line, side, tag='PASS_TRIGGER'), canonical)
        rows.append(_v11_enrich_evaluation(model, canonical, is_model=True))
    candidates = sorted(rows, key=lambda r: float((r.get('advisor') or {}).get('p_final') or 0.0), reverse=True)
    play = next((r for r in candidates if (r.get('advisor') or {}).get('history_zone_eligible') and (r.get('advisor') or {}).get('p_final', 0) >= ADVISOR_PLAY_MIN and _v112_edge(r, -999.0) >= 0 and not (r.get('advisor') or {}).get('serious_blockers')), None)
    risk = next((r for r in candidates if (r.get('advisor') or {}).get('history_zone_eligible') and (r.get('advisor') or {}).get('p_final', 0) >= ADVISOR_RISK_MIN and _v112_edge(r, -999.0) >= 0 and not (r.get('advisor') or {}).get('serious_blockers')), None)
    return {
        'market_label': _v11_market_label(evaluation),
        'side': side,
        'projection_used': projection,
        'risk_trigger': _v11_compact_line(risk),
        'play_trigger': _v11_compact_line(play),
    }


def _v11_pct(value: Any) -> str:
    number = to_number(value)
    return 'N/A' if number is None else f'{number:.1%}'


def _v11_num(value: Any, digits: int = 1) -> str:
    number = to_number(value)
    return 'N/A' if number is None else f'{number:.{digits}f}'


def _v11_scenario_lines(scenario: dict[str, Any], limit: int = 2) -> list[str]:
    rows = (scenario or {}).get('top_support') or []
    if not rows:
        rows = (scenario or {}).get('top_conflict') or []
    output: list[str] = []
    for row in rows[:limit]:
        rate = row.get('combined_rate')
        own = row.get('own') or {}
        allowed = row.get('opponent_allowed') or {}
        output.append(
            f"• {row.get('title')}: {_v11_pct(rate)} для цієї лінії "
            f"(N={row.get('combined_n')}, own {own.get('hits')}/{own.get('n')}, "
            f"суперник {allowed.get('hits')}/{allowed.get('n')}); {row.get('effect')}."
        )
        impacts = own
        impact_bits = []
        if impacts.get('team_won_match_rate') is not None:
            impact_bits.append(f"виграш матчу {_v11_pct(impacts.get('team_won_match_rate'))}")
        if impacts.get('won_next_quarter_rate') is not None:
            impact_bits.append(f"виграш наступної чверті {_v11_pct(impacts.get('won_next_quarter_rate'))}")
        if impacts.get('won_all_4_quarters_rate') is not None:
            impact_bits.append(f"4–0 по чвертях {_v11_pct(impacts.get('won_all_4_quarters_rate'))}")
        if impacts.get('under_21_all_4_quarters_rate') is not None:
            impact_bits.append(f"<21 у всіх чвертях {_v11_pct(impacts.get('under_21_all_4_quarters_rate'))}")
        if impact_bits:
            output.append('  Далі: ' + '; '.join(impact_bits) + '.')
    return output


def _v11_line_block(item: dict[str, Any], index: int) -> str:
    adv = item.get('advisor') or {}
    hist = item.get('history') or {}
    scenario = adv.get('scenario_miner') or {}
    line_type = 'MODEL' if adv.get('is_model_line') else 'REAL'
    side = str(item.get('side') or '')
    action = str(adv.get('action') or 'PASS')
    zone = adv.get('history_zone_rate')
    hits = hist.get('history_zone_hits')
    n = hist.get('history_zone_n')
    zone_fact = f'{hits}/{n}' if hits is not None and n else 'N/A'
    lines = [
        f'<b>#{index} {html.escape(action)} — {html.escape(_v11_market_label(item))}</b>',
        f'<b>Лінія:</b> {html.escape(side)} {float(item.get("line")):.1f} | {line_type}' + (f' | @{float(item.get("odds")):.2f}' if not adv.get('is_model_line') and item.get('odds') else ''),
        f'<b>Історична зона:</b> {_v11_pct(zone)} ({html.escape(zone_fact)})',
        f'<b>P_history:</b> {_v11_pct(adv.get("p_hist"))} | <b>P_scenario:</b> {_v11_pct(adv.get("p_scenario_core"))} | <b>P_live:</b> {_v11_pct(adv.get("p_live"))}',
        f'<b>P_final:</b> {_v11_pct(adv.get("p_final"))}',
        f'<b>LiveProjection:</b> {_v11_num(adv.get("projection_used"))} | <b>Edge:</b> {_v11_num(adv.get("line_edge"))}',
        f'<b>History/Scenario/Live:</b> {html.escape(str(adv.get("alignment")))}',
        f'<b>Стата:</b> {html.escape(str(adv.get("stat_gate_status") or "OFF"))} | FAKE OVER: {"YES" if adv.get("fake_over") else "NO"} | FAKE UNDER: {"YES" if adv.get("fake_under") else "NO"}',
    ]
    scenario_lines = _v11_scenario_lines(scenario)
    if scenario_lines:
        lines.append('<b>Сценарії та вплив:</b>')
        lines.extend(html.escape(text) for text in scenario_lines)
    if adv.get('serious_blockers'):
        lines.append('<b>Чому не брати:</b> ' + html.escape(', '.join(adv['serious_blockers'])))
    return '\n'.join(lines)


def _v11_best_model_trigger(summary: dict[str, Any]) -> tuple[Optional[dict[str, Any]], str]:
    plays = []
    risks = []
    for side_key, side_label in (('over', 'OVER'), ('under', 'UNDER')):
        group = summary.get(side_key) or {}
        if group.get('play_trigger'):
            plays.append((group['play_trigger'], side_label))
        if group.get('risk_trigger'):
            risks.append((group['risk_trigger'], side_label))
    pool = plays if plays else risks
    if not pool:
        return None, ''
    row, side_label = min(
        pool,
        key=lambda pair: (
            abs(_v112_number(pair[0].get('line_edge'), 999.0)),
            -float(pair[0].get('p_final') or 0.0),
        ),
    )
    return row, side_label


def _v11_model_trigger_text(summary: dict[str, Any], index: int) -> str:
    lines = [f'<b>MODEL #{index}: {html.escape(str(summary.get("market_label")))}</b>', f'<b>LiveProjection:</b> {_v11_num(summary.get("projection_used"))}']
    row, side_label = _v11_best_model_trigger(summary)
    if row:
        label = 'theoretical PLAY' if float(row.get('p_final') or 0.0) >= ADVISOR_PLAY_MIN else 'theoretical RISK'
        lines.append(html.escape(
            f'• {side_label} {float(row["line"]):.1f}: {label}, '
            f'P_final {_v11_pct(row.get("p_final"))}, зона {_v11_pct(row.get("history_zone_rate"))}, '
            f'edge {_v11_num(row.get("line_edge"))}'
        ))
    else:
        lines.append('Підтвердженого theoretical trigger у зоні 75–100% не знайдено.')
    return '\n'.join(lines)



def _v11_audit_block(items: list[dict[str, Any]], primary: list[dict[str, Any]]) -> str:
    primary_keys = {
        (row.get('market_type'), row.get('team'), row.get('segment'), row.get('side'), to_number(row.get('line')))
        for row in primary
    }
    remaining = [
        row for row in items
        if (row.get('market_type'), row.get('team'), row.get('segment'), row.get('side'), to_number(row.get('line'))) not in primary_keys
    ]
    if not remaining:
        return ''
    lines = ['<b>ДОДАТКОВИЙ АУДИТ УСІХ ЛІНІЙ 75–100% / EXCEPTIONAL EDGE</b>']
    for row in remaining:
        adv = row.get('advisor') or {}
        market = html.escape(_v11_market_label(row))
        side = html.escape(str(row.get('side') or 'N/A'))
        line = _v11_num(row.get('line'))
        action = html.escape(str(adv.get('action') or 'PASS'))
        lines.append(
            f'• <b>{action}</b> — {market}: {side} {line}; '
            f'зона {_v11_pct(adv.get("history_zone_rate"))}; '
            f'P_h {_v11_pct(adv.get("p_hist"))}; P_s {_v11_pct(adv.get("p_scenario_core"))}; '
            f'P_l {_v11_pct(adv.get("p_live"))}; P_f {_v11_pct(adv.get("p_final"))}; '
            f'Proj {_v11_num(adv.get("projection_used"))}; edge {_v11_num(adv.get("line_edge"))}; '
            f'{html.escape(str(adv.get("alignment") or "N/A"))}; '
            f'FO={"YES" if adv.get("fake_over") else "NO"}; FU={"YES" if adv.get("fake_under") else "NO"}'
        )
    return '\n'.join(lines)

def _v11_build_messages(advisor: dict[str, Any], calculation: dict[str, Any]) -> list[str]:
    snapshot = calculation['canonical_snapshot']
    score = snapshot.get('score') or {}
    quarters = snapshot.get('quarters') or []
    qtext = ' | '.join(
        f"Q{i + 1} {q.get('home')}:{q.get('away')}"
        for i, q in enumerate(quarters)
        if q.get('home') is not None and q.get('away') is not None
    )
    action = advisor['action']
    icon = '✅' if action == 'PLAY' else '⚠️' if action == 'RISK' else '❌'
    header = '\n'.join([
        f'<b>{icon} {html.escape(action)}</b>',
        f'<b>Матч:</b> {html.escape(str(snapshot.get("name")))}',
        f'<b>Стадія:</b> {html.escape(str(snapshot.get("stage")))} | <b>Рахунок:</b> {score.get("home")}:{score.get("away")}',
        f'<b>Чверті:</b> {html.escape(qtext or "N/A")}',
        f'<b>Чому матч у Telegram:</b> {html.escape(str(advisor.get("dispatch_reason")))}',
    ])
    primary_lines = advisor.get('primary_lines') or []
    blocks = [_v11_line_block(item, idx) for idx, item in enumerate(primary_lines, 1)]
    if not blocks:
        blocks = [_v11_model_trigger_text(row, idx) for idx, row in enumerate((advisor.get('model_summary') or [])[:ADVISOR_MAX_PRIMARY], 1)]
    audit_block = _v11_audit_block(advisor.get('all_qualifying_real_lines') or [], primary_lines)
    if audit_block:
        blocks.append(audit_block)
    theoretical = advisor.get('nearest_theoretical_play')
    if theoretical:
        trigger_lines = ['<b>НАЙБЛИЖЧИЙ ТЕОРЕТИЧНИЙ PLAY</b>', f'<b>Ринок:</b> {html.escape(str(theoretical.get("market_label")))}', f'<b>LiveProjection:</b> {_v11_num(theoretical.get("projection_used"))}']
        if theoretical.get('play_trigger'):
            row = theoretical['play_trigger']
            trigger_lines.append(f"PLAY при {html.escape(str(row.get('side')))} {float(row.get('line')):.1f}; P_final {_v11_pct(row.get('p_final'))}; зона {_v11_pct(row.get('history_zone_rate'))}")
        elif theoretical.get('risk_trigger'):
            row = theoretical['risk_trigger']
            trigger_lines.append(f"Поки лише RISK при {html.escape(str(row.get('side')))} {float(row.get('line')):.1f}; P_final {_v11_pct(row.get('p_final'))}")
        else:
            trigger_lines.append('Підтвердженої лінії 75–100% поки немає.')
        blocks.append('\n'.join(trigger_lines))
    if not blocks:
        blocks.append('Розрахунок виконано, але придатної реальної або модельної лінії не знайдено.')

    messages: list[str] = []
    current = header
    for block in blocks:
        candidate = current + '\n\n' + block
        if len(candidate) > 3900 and current != header:
            messages.append(current)
            current = '<b>Продовження розрахунку</b>\n' + block
        elif len(candidate) > 3900:
            # Hard split long block by lines.
            messages.append(current)
            current = '<b>Продовження розрахунку</b>'
            for line in block.splitlines():
                if len(current) + len(line) + 1 > 3900:
                    messages.append(current)
                    current = '<b>Продовження розрахунку</b>\n' + line
                else:
                    current += '\n' + line
        else:
            current = candidate
    if current:
        messages.append(current)
    return messages


def _v11_delivery_connect(db_path: str | Path) -> sqlite3.Connection:
    path = Path(db_path).expanduser()
    path.parent.mkdir(parents=True, exist_ok=True)
    connection = sqlite3.connect(path)
    connection.execute(
        '''CREATE TABLE IF NOT EXISTS advisor_deliveries (
               advisor_key TEXT PRIMARY KEY,
               match_id TEXT NOT NULL,
               input_hash TEXT NOT NULL,
               action TEXT NOT NULL,
               status TEXT NOT NULL,
               telegram_status TEXT NOT NULL,
               created_at TEXT NOT NULL
           )'''
    )
    connection.commit()
    return connection


def _v11_send_messages(
    messages: list[str],
    *,
    telegram_sender: Optional[Callable[[str], dict[str, Any]]] = None,
) -> dict[str, Any]:
    sender = telegram_sender or (lambda value: send_telegram_message(value))
    results: list[dict[str, Any]] = []
    for message in messages:
        raw = sender(message)
        results.append(raw if isinstance(raw, dict) else {'sent': False, 'status': 'INVALID_SENDER_RESPONSE'})
    sent_count = sum(bool(row.get('sent')) for row in results)
    all_sent = bool(results) and sent_count == len(results)
    any_sent = sent_count > 0
    if all_sent:
        status = 'SENT'
    elif any_sent:
        status = 'PARTIAL_SEND_FAILED'
    else:
        status = results[0].get('status') if results else 'NO_MESSAGES'
    return {
        'status': status,
        'sent': all_sent,
        'partially_sent': any_sent and not all_sent,
        'sent_count': sent_count,
        'message_id': next((row.get('message_id') for row in results if row.get('message_id')), None),
        'parts': results,
        'message_count': len(messages),
    }


def process_vps_match_file(
    match_path: str | Path,
    *,
    output_path: str | Path | None = None,
    zones_path: str | Path | None = None,
    db_path: str | Path = 'super_basket.sqlite3',
    mode: str = 'ACTION',
    require_gpt: bool = False,
    enable_gpt: bool = False,
    enable_telegram: bool = True,
    dry_run: bool = False,
    strict_schema: bool = False,
    checkpoint: Optional[int] = None,
    gpt_reviewer: Optional[Callable[[dict[str, Any], dict[str, Any]], dict[str, Any]]] = None,
    telegram_sender: Optional[Callable[[str], dict[str, Any]]] = None,
) -> dict[str, Any]:
    """Process one parser JSON as a full Telegram adviser.

    Telegram dispatch is intentionally narrow:
      1) at least one real line has an exact historical zone >=75%; or
      2) no supported real total/IT line exists; or
      3) a real line has a directional live edge >= configured 15 points.
    """
    del require_gpt, enable_gpt, gpt_reviewer  # deterministic advisor; GPT is not required
    mode = mode.upper()
    if mode not in {'ACTION', 'STRICT'}:
        raise ValueError('mode must be ACTION or STRICT')
    source_path = Path(match_path).expanduser().resolve()
    source = load_json(source_path)
    if checkpoint is None:
        checkpoint = _v11_checkpoint_from_filename(source_path)
    if checkpoint is not None:
        checkpoint = int(checkpoint)
        if checkpoint not in {1, 2, 3}:
            raise ValueError('checkpoint must be 1, 2 or 3')
        context = source.get('analysis_context') if isinstance(source.get('analysis_context'), dict) else {}
        source['analysis_context'] = {**context, 'trigger_checkpoint': checkpoint}
    zones, zones_metadata = resolve_team_relative_zones(source, zones_path=zones_path)
    calculator = SuperBasketCalculator(deepcopy(DEFAULT_CONFIG), zones, zones_metadata)
    core_result = calculator.calculate(source, dispatch_threshold=float(DEFAULT_CONFIG['dispatch_threshold']), strict_schema=strict_schema)
    calculation = core_result['super_basket_calculation']
    canonical = adapt_match(source, deepcopy(DEFAULT_CONFIG), strict_schema)
    canonical['data_gate']['team_relative_zones'] = deepcopy(zones_metadata)
    canonical['coursework_forecast'] = build_coursework_remaining_forecast(canonical)
    input_state_gate = _v112_input_state_gate(source, canonical, checkpoint)

    real_evaluations = [_v11_enrich_evaluation(item, canonical, is_model=False) for item in calculation.get('market_evaluations') or []]
    calculation['market_evaluations'] = real_evaluations
    primary, qualifying = _v11_select_advisor_lines(real_evaluations)

    real_keys = {
        _v11_market_key(item)
        for item in real_evaluations
        if _v112_market_is_currently_supported(item)
    }
    supported_real_count = len(real_keys)
    model_evaluations: list[dict[str, Any]] = []
    for spec in _v11_relevant_market_specs(canonical):
        if _v11_market_key(spec) not in real_keys:
            model_evaluations.extend(_v11_evaluate_model_grid(calculator, canonical, spec))
    model_summary = _v11_model_summary(model_evaluations, canonical)

    supported_real_evaluations = [item for item in real_evaluations if _v112_market_is_currently_supported(item)]
    zone_dispatch = any((item.get('advisor') or {}).get('history_zone_eligible') for item in supported_real_evaluations)
    edge_dispatch = any((item.get('advisor') or {}).get('exceptional_edge') for item in supported_real_evaluations)
    no_line_dispatch = supported_real_count == 0
    should_dispatch = bool(input_state_gate['allowed'] and (zone_dispatch or edge_dispatch or no_line_dispatch))
    reasons: list[str] = []
    if zone_dispatch:
        reasons.append('Є реальна лінія в історичній зоні 75–100%.')
    if edge_dispatch:
        reasons.append(f'Є винятковий live-edge ≥{ADVISOR_EXCEPTIONAL_EDGE_MIN:g} очок.')
    if no_line_dispatch:
        reasons.append('Підтримуваних реальних total/IT ліній немає — показано model triggers.')
    dispatch_reason = ' '.join(reasons) if reasons else 'Матч не відповідає Telegram-фільтру.'
    if not input_state_gate['allowed']:
        dispatch_reason = 'Telegram заблоковано: ' + str(input_state_gate['reason'])

    active = [item for item in primary if (item.get('advisor') or {}).get('action') in {'PLAY', 'RISK'}]
    active.sort(key=_v11_primary_sort, reverse=True)
    if active and input_state_gate['allowed']:
        top = active[0]
        action = str((top.get('advisor') or {}).get('action'))
    else:
        top = primary[0] if primary else None
        action = 'PASS'

    theoretical = None
    if action == 'PASS' and top and input_state_gate['allowed']:
        theoretical = _v11_theoretical_for_pass(calculator, canonical, top)
    if action == 'PASS' and input_state_gate['allowed'] and theoretical is None and model_summary:
        # The model summary itself is the theoretical recommendation when no
        # real market exists.
        first_summary = model_summary[0]
        best_row, _ = _v11_best_model_trigger(first_summary)
        is_play = bool(best_row and float(best_row.get('p_final') or 0.0) >= ADVISOR_PLAY_MIN)
        theoretical = {
            'market_label': first_summary.get('market_label'),
            'projection_used': first_summary.get('projection_used'),
            'side': (best_row or {}).get('side'),
            'play_trigger': best_row if is_play else None,
            'risk_trigger': best_row if best_row and not is_play else None,
        }

    advisor = {
        'version': ADVISOR_VERSION,
        'action': action,
        'status': (
            str((top.get('advisor') or {}).get('status'))
            if top is not None and action in {'PLAY', 'RISK'}
            else ('PASS — INPUT STATE GATE' if not input_state_gate['allowed'] else 'PASS — TELEGRAM ADVISOR')
        ),
        'dispatch': should_dispatch,
        'dispatch_reason': dispatch_reason,
        'filter': {
            'history_zone_min': ADVISOR_HISTORY_ZONE_MIN,
            'exceptional_edge_min': ADVISOR_EXCEPTIONAL_EDGE_MIN,
            'supported_real_market_count': supported_real_count,
            'zone_dispatch': zone_dispatch,
            'edge_dispatch': edge_dispatch,
            'no_line_dispatch': no_line_dispatch,
            'input_state_gate': input_state_gate,
        },
        'primary_lines': primary,
        'all_qualifying_real_lines': qualifying,
        'model_summary': model_summary,
        'nearest_theoretical_play': theoretical,
    }
    messages = _v11_build_messages(advisor, calculation) if should_dispatch else []
    advisor['telegram_messages'] = messages

    # Build a backwards-compatible decision block around the selected real line.
    if top:
        top['system_action'] = action
        top['system_status'] = advisor['status']
        decision = build_decision(top if action in {'PLAY', 'RISK'} else None, top, calculation, mode)
    else:
        decision = build_decision(None, None, calculation, mode)
    decision['action'] = action
    decision['deterministic_action'] = action
    decision['status'] = advisor['status']
    if not input_state_gate['allowed']:
        decision['reason_codes'] = [str(input_state_gate['reason'])]
        decision['market'] = None
        decision['signal_id'] = None
        decision['explanation_uk'] = 'Рекомендацію заблоковано через стан вхідного snapshot: ' + str(input_state_gate['reason'])
        decision['main_risk_uk'] = 'Не використовувати завершені або застарілі дані для live-рекомендації.'
        decision['trigger_uk'] = 'Очікувати новий актуальний snapshot поточного матчу.'
    elif top:
        top_adv = top.get('advisor') or {}
        top_policy = top_adv.get('audit_zone_policy') or {}
        decision['reason_codes'] = list(top_adv.get('serious_blockers') or [])
        if top_policy.get('applied'):
            decision['audit_policy_classification'] = top_policy.get('classification')
        projection = to_number(top_adv.get('projection_used'))
        line_value = to_number(top.get('line'))
        edge_value = to_number(top_adv.get('line_edge'))
        if projection is not None and line_value is not None:
            relation = 'на рівні лінії'
            if projection > line_value:
                relation = f'вище лінії на {projection - line_value:.1f}'
            elif projection < line_value:
                relation = f'нижче лінії на {line_value - projection:.1f}'
            decision['explanation_uk'] = (
                f"P_final {float(top_adv.get('p_final') or 0.0):.1%}: "
                f"LiveProjection {projection:.1f}, {relation}; "
                f"лінія {line_value:.1f}; історична зона "
                f"{float(top_adv.get('history_zone_rate') or 0.0):.1%}; "
                f"edge сторони {edge_value:.1f}."
            )
        if action == 'RISK':
            decision['main_risk_uk'] = (
                f"RISK через P_final {float(top_adv.get('p_final') or 0.0):.1%} "
                f"та/або неповне підтвердження. " + str(top_policy.get('reason_uk') or '')
            ).strip()
        elif action == 'PLAY':
            decision['main_risk_uk'] = 'Основний ризик — зміна live-лінії, темпу або статистичного профілю після snapshot.'
        else:
            decision['main_risk_uk'] = 'PASS: ' + (', '.join(decision['reason_codes']) if decision['reason_codes'] else str(top_policy.get('reason_uk') or dispatch_reason))
    else:
        decision['reason_codes'] = []
    decision['advisor_dispatch'] = should_dispatch
    decision['advisor_dispatch_reason'] = dispatch_reason
    decision['nearest_theoretical_play'] = theoretical
    decision['gpt_status'] = 'NOT_REQUIRED_V11_DETERMINISTIC'
    decision.pop('_evaluation', None)

    target = Path(output_path).expanduser().resolve() if output_path else source_path.with_name(source_path.stem + '_advisor_result.json')
    delivery = {'status': 'SKIPPED_FILTER', 'sent': False, 'message_id': None, 'message_count': len(messages)}
    duplicate = False
    advisor_key = hashlib.sha256((calculation['input_snapshot_hash'] + '|' + ADVISOR_VERSION).encode('utf-8')).hexdigest()
    connection = _v11_delivery_connect(db_path)
    try:
        row = connection.execute('SELECT telegram_status FROM advisor_deliveries WHERE advisor_key=?', (advisor_key,)).fetchone()
        duplicate = row is not None and row[0] == 'SENT'
        if not should_dispatch:
            delivery = {'status': 'SKIPPED_FILTER', 'sent': False, 'message_id': None, 'message_count': 0}
        elif duplicate:
            delivery = {'status': 'SKIPPED_DUPLICATE_ALREADY_SENT', 'sent': False, 'message_id': None, 'message_count': len(messages)}
        elif dry_run:
            delivery = {'status': 'DRY_RUN_NOT_SENT', 'sent': False, 'message_id': None, 'message_count': len(messages)}
        elif not enable_telegram:
            delivery = {'status': 'SKIPPED_TELEGRAM_DISABLED', 'sent': False, 'message_id': None, 'message_count': len(messages)}
        else:
            delivery = _v11_send_messages(messages, telegram_sender=telegram_sender)
        connection.execute(
            'INSERT OR REPLACE INTO advisor_deliveries(advisor_key,match_id,input_hash,action,status,telegram_status,created_at) VALUES(?,?,?,?,?,?,?)',
            (advisor_key, canonical['match_id'], calculation['input_snapshot_hash'], action, advisor['status'], delivery['status'], utc_now()),
        )
        connection.commit()
    finally:
        connection.close()

    # v11.0/v11.1 accidentally bypassed the LearningStore used by `report`,
    # `settle` and `settle-match`. Persist every real PLAY/RISK again and mark
    # every processed snapshot, while keeping Telegram delivery deduplication
    # in the dedicated advisor_deliveries table.
    learning_duplicate = False
    learning_store = LearningStore(db_path)
    try:
        decision['telegram_status'] = delivery['status']
        if action in {'PLAY', 'RISK'} and decision.get('signal_id') and decision.get('market'):
            existing_signal, learning_duplicate = learning_store.record_signal(decision, calculation)
            learning_store.update_delivery(
                decision['signal_id'],
                action,
                decision.get('gpt_status', 'NOT_REQUIRED_V11_DETERMINISTIC'),
                delivery['status'],
                delivery.get('message_id'),
            )
        learning_store.mark_processed(
            calculation['input_snapshot_hash'],
            str(source_path),
            str(target),
            'OK' if input_state_gate['allowed'] else str(input_state_gate['reason']),
        )
    finally:
        learning_store.close()

    line_recommendations = [_v11_compact_line(item) for item in real_evaluations]
    calculation['line_recommendations'] = line_recommendations
    calculation['active_line_recommendations'] = [row for row in line_recommendations if row and row.get('action') in {'PLAY', 'RISK'}]
    calculation['advisor_model_evaluations'] = [_v11_compact_line(item) for item in model_evaluations]
    calculation['advisor'] = advisor

    system = {
        'version': ADVISOR_VERSION,
        'processed_at': utc_now(),
        'input_hash': calculation['input_snapshot_hash'],
        'mode': mode,
        'status': 'OK',
        'data_gate': calculation.get('data_gate'),
        'format_gate': format_gate(calculation),
        'market_audit': calculation.get('market_audit'),
        'line_coverage': calculation.get('line_coverage'),
        'line_recommendations': line_recommendations,
        'active_line_recommendations': calculation.get('active_line_recommendations'),
        'decision': decision,
        'decision_text': f'{action} | {dispatch_reason}',
        'gpt_review': {'status': 'NOT_REQUIRED_V11_DETERMINISTIC', 'approved': True, 'action': action},
        'risk_post_filter': {'enabled': False, 'policy': 'V11_ADVISOR'},
        'telegram_delivery': {**delivery, 'duplicate_signal': duplicate},
        'learning_store': {'signal_recorded': action in {'PLAY', 'RISK'} and bool(decision.get('signal_id')), 'duplicate_signal': learning_duplicate},
        'advisor': advisor,
        'input_state_gate': input_state_gate,
        'files': {'source': str(source_path), 'result': str(target)},
    }
    core_result['super_basket_system'] = system
    save_json(target, core_result)
    append_verdict_log({
        'timestamp': system['processed_at'],
        'match_id': canonical['match_id'],
        'match_name': canonical['name'],
        'checkpoint': canonical['stage'],
        'trigger_checkpoint': canonical.get('trigger_checkpoint'),
        'verdict': action,
        'verdict_status': advisor['status'],
        'p_final': (top.get('advisor') or {}).get('p_final') if top else None,
        'market': _v11_compact_line(top) if top else None,
        'description': dispatch_reason,
        'reason_codes': (top.get('advisor') or {}).get('serious_blockers') if top else [],
        'input_hash': calculation['input_snapshot_hash'],
        'gpt_status': 'NOT_REQUIRED_V11_DETERMINISTIC',
        'telegram_status': delivery['status'],
        'advisor_filter': advisor['filter'],
        'nearest_theoretical_play': theoretical,
        'files': {'source': str(source_path), 'result': str(target)},
    })
    if ENABLE_EXCEL_AUDIT:
        append_excel_audit(core_result)
    return core_result



def _v11_checkpoint_from_filename(path: Path) -> Optional[int]:
    match = re.search(r'(?:^|_)q([123])(?:_|$)', path.stem, flags=re.IGNORECASE)
    return int(match.group(1)) if match else None


def watch_inbox(
    inbox: str | Path,
    outbox: str | Path,
    *,
    zones_path: str | Path | None,
    db_path: str | Path,
    mode: str,
    require_gpt: bool,
    enable_gpt: bool,
    enable_telegram: bool,
    poll_seconds: float,
) -> None:
    """Watch parser JSONs, including names such as *_q1_result.json.

    The legacy watcher skipped every *_result.json, which also skipped the
    parser's own Q1/HT/Q3 files. v11 only skips files generated by the advisor
    and infers checkpoint 1/2/3 from the parser filename when available.
    """
    inbox_path = Path(inbox).expanduser().resolve()
    outbox_path = Path(outbox).expanduser().resolve()
    inbox_path.mkdir(parents=True, exist_ok=True)
    outbox_path.mkdir(parents=True, exist_ok=True)
    signatures: dict[str, tuple[int, int]] = {}
    stable: dict[str, int] = {}
    processed: dict[str, tuple[int, int]] = {}
    print(f'WATCHING {inbox_path} -> {outbox_path}', flush=True)
    while True:
        for path in sorted(inbox_path.glob('*.json')):
            if path.name.endswith(('_advisor_result.json', '_calculated.json')):
                continue
            try:
                stat_result = path.stat()
                signature = (stat_result.st_size, stat_result.st_mtime_ns)
            except OSError:
                continue
            key = str(path)
            if signatures.get(key) == signature:
                stable[key] = stable.get(key, 0) + 1
            else:
                signatures[key] = signature
                stable[key] = 0
            if stable[key] < 1 or processed.get(key) == signature:
                continue
            output = outbox_path / f'{path.stem}_advisor_result.json'
            checkpoint = _v11_checkpoint_from_filename(path)
            try:
                result = process_vps_match_file(
                    path,
                    output_path=output,
                    zones_path=zones_path,
                    db_path=db_path,
                    mode=mode,
                    require_gpt=require_gpt,
                    enable_gpt=enable_gpt,
                    enable_telegram=enable_telegram,
                    checkpoint=checkpoint,
                )
                decision = result['super_basket_system']['decision']
                print(
                    f"{utc_now()} {path.name}: checkpoint={checkpoint or 'auto'} "
                    f"{decision['action']} {decision['status']}",
                    flush=True,
                )
                processed[key] = signature
            except (OSError, ValueError, KeyError, json.JSONDecodeError, sqlite3.Error) as exc:
                print(f'{utc_now()} ERROR {path.name}: {type(exc).__name__}: {exc}', file=sys.stderr, flush=True)
        time.sleep(max(0.5, poll_seconds))



# ===== v11.1 ASYMMETRIC STRONG-ZONE AUDIT POLICY =====
# The mathematical core remains unchanged. This layer only changes the final
# adviser decision filter and explanation according to the completed strong-zone
# audit. The audit found materially different behaviour for OVER and UNDER:
# - strong OVER zones tolerated a modest negative live edge much better;
# - strong UNDER zones failed frequently when live projection was above the line.
# P_history, P_scenario, projections, P_live, P_raw and P_final are not modified.

ADVISOR_VERSION = '11.2.0-BUGFIX-ASYMMETRIC-ZONE-ADVISOR'
ADVISOR_OVER_AUDIT_ZONE_MIN = float(os.getenv('SUPER_BASKET_OVER_AUDIT_ZONE_MIN', '0.80'))
ADVISOR_OVER_SMALL_DEFICIT_MAX = float(os.getenv('SUPER_BASKET_OVER_SMALL_DEFICIT_MAX', '8.0'))
ADVISOR_OVER_LARGE_DEFICIT_MIN = float(os.getenv('SUPER_BASKET_OVER_LARGE_DEFICIT_MIN', '8.0'))
ADVISOR_UNDER_AUDIT_ZONE_MIN = float(os.getenv('SUPER_BASKET_UNDER_AUDIT_ZONE_MIN', '0.80'))
ADVISOR_UNDER_HARD_CONFLICT_EDGE = float(os.getenv('SUPER_BASKET_UNDER_HARD_CONFLICT_EDGE', '8.0'))
ADVISOR_UNDER_ELITE_ZONE_MIN = float(os.getenv('SUPER_BASKET_UNDER_ELITE_ZONE_MIN', '0.90'))
ADVISOR_OVER_SCENARIO_SUPPORT_MIN = float(os.getenv('SUPER_BASKET_OVER_SCENARIO_SUPPORT_MIN', '0.60'))
ADVISOR_OVER_STRONG_SCENARIO_MIN = float(os.getenv('SUPER_BASKET_OVER_STRONG_SCENARIO_MIN', '0.75'))

_V111_BASE_ENRICH = _v11_enrich_evaluation
_V111_BASE_COMPACT = _v11_compact_line
_V111_BASE_LINE_BLOCK = _v11_line_block


def _v111_stat_confirmed(item: dict[str, Any], side: str) -> bool:
    stat = item.get('stat_comparison') or {}
    if side == 'OVER':
        return bool(stat.get('real_over')) or str(stat.get('over_gate_status') or '') == 'CONFIRMED' or str(stat.get('stat_gate_status') or '') == 'CONFIRMED'
    return bool(stat.get('real_under')) or str(stat.get('under_gate_status') or '') == 'CONFIRMED' or str(stat.get('stat_gate_status') or '') == 'CONFIRMED'


def _v111_scenario_support(adv: dict[str, Any], *, strong: bool = False) -> bool:
    core = float(adv.get('p_scenario_core') or 0.50)
    miner = float(adv.get('p_scenario_miner') or 0.50)
    scenario = adv.get('scenario_miner') or {}
    n = int(scenario.get('scenario_n') or 0)
    threshold = ADVISOR_OVER_STRONG_SCENARIO_MIN if strong else ADVISOR_OVER_SCENARIO_SUPPORT_MIN
    # The core scenario may be based on a valid sample even when the miner has
    # no matching pattern. The miner itself must have at least five cases.
    return core >= threshold or (miner >= threshold and n >= 5)


def _v111_remove_blockers(blockers: list[str], allowed: set[str]) -> list[str]:
    return [code for code in blockers if code not in allowed]


def _v111_audit_policy(
    item: dict[str, Any],
    adv: dict[str, Any],
    canonical: dict[str, Any],
) -> dict[str, Any]:
    side = str(item.get('side') or '')
    zone = to_number(adv.get('history_zone_rate'))
    edge = to_number(adv.get('line_edge'))
    stat = item.get('stat_comparison') or {}
    p_final = float(adv.get('p_final') or 0.50)
    p_live = float(adv.get('p_live') or 0.50)
    scenario_support = _v111_scenario_support(adv, strong=False)
    strong_scenario = _v111_scenario_support(adv, strong=True)
    stat_confirmed = _v111_stat_confirmed(item, side)
    fake_against = bool((side == 'OVER' and stat.get('fake_over')) or (side == 'UNDER' and stat.get('fake_under')))

    result = {
        'applied': False,
        'side': side,
        'zone': zone,
        'edge': edge,
        'classification': 'STANDARD',
        'permission': False,
        'hard_block': False,
        'scenario_support': scenario_support,
        'strong_scenario': strong_scenario,
        'stat_confirmed': stat_confirmed,
        'p_final_unchanged': p_final,
        'p_live_unchanged': p_live,
        'reason_uk': '',
        'audit_evidence_uk': '',
        'removed_blockers': [],
        'added_blockers': [],
    }
    if zone is None or edge is None:
        return result

    if side == 'OVER' and zone >= ADVISOR_OVER_AUDIT_ZONE_MIN:
        result['applied'] = True
        if edge >= 0:
            result.update({
                'classification': 'OVER_ZONE_LIVE_ALIGNED',
                'permission': True,
                'reason_uk': 'Сильна OVER-зона та live-проєкція на/вище лінії.',
                'audit_evidence_uk': 'В аудиті OVER 80%+ при проєкції на/вище лінії: 4/4 WIN.',
            })
        elif edge >= -ADVISOR_OVER_SMALL_DEFICIT_MAX:
            result.update({
                'classification': 'OVER_ZONE_SMALL_LIVE_DEFICIT',
                'permission': bool(scenario_support or stat_confirmed),
                'reason_uk': (
                    f'Live-проєкція нижче OVER-лінії лише на {abs(edge):.1f} очка; '
                    'невеликий мінус не є автоматичним блокером за умови підтримки сценарію або статистики.'
                ),
                'audit_evidence_uk': 'В аудиті сильні OVER-зони з дефіцитом live 0–8 очок дали 7/7 WIN.',
            })
        else:
            result.update({
                'classification': 'OVER_ZONE_LARGE_LIVE_DEFICIT',
                'permission': bool(stat_confirmed or strong_scenario),
                'reason_uk': (
                    f'Live-проєкція нижче OVER-лінії на {abs(edge):.1f} очка; '
                    'потрібне FULL_STAT або сильне сценарне підтвердження.'
                ),
                'audit_evidence_uk': 'В аудиті дефіцит OVER понад 8 очок дав 3/5 WIN, тому автоматичне підвищення заборонене.',
            })
        if fake_against:
            result['permission'] = False
            result['hard_block'] = True
            result['added_blockers'].append('AUDIT_FAKE_OVER_BLOCK')
            result['reason_uk'] += ' Виявлено FAKE OVER, тому дозвіл скасовано.'
        return result

    if side == 'UNDER' and zone >= ADVISOR_UNDER_AUDIT_ZONE_MIN:
        result['applied'] = True
        projection_above = edge < 0
        if projection_above and abs(edge) > ADVISOR_UNDER_HARD_CONFLICT_EDGE:
            result.update({
                'classification': 'UNDER_ZONE_HARD_LIVE_CONFLICT_GT8',
                'hard_block': True,
                'permission': False,
                'reason_uk': f'Live-проєкція вище UNDER-лінії на {abs(edge):.1f} очка — жорсткий конфлікт.',
                'audit_evidence_uk': 'В аудиті UNDER 80%+ при проєкції вище лінії більш ніж на 8 очок: 0/7 WIN.',
            })
            result['added_blockers'].append('AUDIT_UNDER_LIVE_CONFLICT_GT8')
        elif projection_above and zone >= ADVISOR_UNDER_ELITE_ZONE_MIN:
            result.update({
                'classification': 'UNDER_ELITE_ZONE_ANY_LIVE_CONFLICT',
                'hard_block': True,
                'permission': False,
                'reason_uk': f'UNDER-зона {zone:.1%}, але live-проєкція вище лінії на {abs(edge):.1f} очка.',
                'audit_evidence_uk': 'В аудиті UNDER 90–100% із live-проєкцією вище лінії: 0/4 WIN.',
            })
            result['added_blockers'].append('AUDIT_UNDER_90_LIVE_CONFLICT')
        elif edge >= 0:
            result.update({
                'classification': 'UNDER_ZONE_LIVE_ALIGNED',
                'permission': True,
                'reason_uk': 'Сильна UNDER-зона підтверджена live-проєкцією на/нижче лінії.',
                'audit_evidence_uk': 'В аудиті UNDER 80%+ при проєкції на/нижче лінії: 3/4 WIN; вибірка мала, тому потрібна обережність.',
            })
        else:
            result.update({
                'classification': 'UNDER_ZONE_SOFT_LIVE_CONFLICT',
                'permission': False,
                'reason_uk': f'Live-проєкція вище UNDER-лінії на {abs(edge):.1f} очка.',
                'audit_evidence_uk': 'Сильні UNDER-зони в аудиті були значно слабші за OVER; конфлікт live не ігнорується.',
            })
        if fake_against:
            result['hard_block'] = True
            result['permission'] = False
            result['added_blockers'].append('AUDIT_FAKE_UNDER_BLOCK')
        return result

    return result


def _v11_enrich_evaluation(
    evaluation: dict[str, Any],
    canonical: dict[str, Any],
    *,
    is_model: bool = False,
    mine_scenarios: bool = True,
) -> dict[str, Any]:
    """v11.1 final decision filter.

    Calls the unchanged v11/core calculation first, then applies only an
    asymmetric adviser policy to the final action. Numeric P values remain
    exactly as calculated by the core.
    """
    item = _V111_BASE_ENRICH(evaluation, canonical, is_model=is_model, mine_scenarios=mine_scenarios)
    adv = item.get('advisor') or {}
    policy = _v111_audit_policy(item, adv, canonical)
    side = str(item.get('side') or '')
    zone = to_number(adv.get('history_zone_rate'))
    edge = to_number(adv.get('line_edge'))
    p_final = float(adv.get('p_final') or 0.50)
    p_live = float(adv.get('p_live') or 0.50)
    p_scenario = max(float(adv.get('p_scenario_core') or 0.50), float(adv.get('p_scenario_miner') or 0.50))
    odds = to_number(item.get('odds'))
    stat = item.get('stat_comparison') or {}
    fake_against = bool((side == 'OVER' and stat.get('fake_over')) or (side == 'UNDER' and stat.get('fake_under')))
    exceptional = bool(adv.get('exceptional_edge'))
    zone_ok = zone is not None and zone >= ADVISOR_HISTORY_ZONE_MIN
    serious = list(adv.get('serious_blockers') or [])

    # The audit gives permission to remove only legacy no-stat/live-direction
    # blockers. All metadata, fake-profile, router, Q4 and stat-against blockers
    # remain in force.
    removable = {
        'LIVE_DIRECTION_OR_EDGE_FAILED',
        'NO_STAT_SUPPORT_TOO_LOW',
        'STRONG_HISTORY_LIVE_CONFLICT',
        'LIVE_EDGE_BELOW_3',
        'LIVE_EDGE_BELOW_0_50',
    }
    if side == 'OVER' and policy.get('permission'):
        before = set(serious)
        serious = _v111_remove_blockers(serious, removable)
        policy['removed_blockers'] = sorted(before - set(serious))
    for code in policy.get('added_blockers') or []:
        if code not in serious:
            serious.append(code)

    # Recalculate only the final action. P_final and every component remain
    # unchanged and are displayed alongside the audit decision.
    if is_model:
        action, status = 'PASS', 'MODEL LINE / THEORETICAL TRIGGER'
    elif odds is None or odds < float(DEFAULT_CONFIG.get('odds_min', 1.44)):
        action, status = 'PASS', 'PASS — ODDS'
    elif serious or fake_against or policy.get('hard_block'):
        action, status = 'PASS', 'PASS — AUDIT CONFLICT/BLOCKER'
    else:
        allowed_direction = bool(edge is not None and edge >= 0)
        if side == 'OVER' and policy.get('permission'):
            allowed_direction = True
        eligible = bool(zone_ok or exceptional)
        if eligible and allowed_direction and p_final >= ADVISOR_PLAY_MIN:
            action, status = 'PLAY', 'PLAY — ASYMMETRIC ZONE ADVISOR'
        elif eligible and allowed_direction and p_final >= ADVISOR_RISK_MIN:
            action, status = 'RISK', 'RISK — ASYMMETRIC ZONE ADVISOR'
        else:
            action, status = 'PASS', 'PASS — ASYMMETRIC ZONE ADVISOR'

    # Strong UNDER zones were much less reliable in the audit. Every settled
    # strong UNDER in that audit was NO_STAT and the combined hit rate was only
    # 29.4%. Therefore a non-confirmed-stat UNDER is capped at RISK even when
    # history, scenario and live all look strong. FULL_STAT confirmation is
    # required for a clean UNDER PLAY.
    stat_confirmed = _v111_stat_confirmed(item, side)
    if action == 'PLAY' and side == 'UNDER' and not stat_confirmed:
        action = 'RISK'
        status = 'RISK — UNDER AUDIT CAUTION'
        policy['under_play_downgraded'] = True
        policy['reason_uk'] = (policy.get('reason_uk') or '') + ' Без підтвердженої live-статистики сильний UNDER має максимум RISK, не clean PLAY.'

    # Alignment is asymmetric too: a tolerated small OVER deficit is neutral,
    # not a hard live conflict. The actual edge remains visible numerically.
    if side == 'OVER' and policy.get('permission') and edge is not None and edge < 0:
        if float(adv.get('p_hist') or 0.50) >= 0.60 and p_scenario >= 0.60:
            adv['alignment'] = 'HISTORY + SCENARIO ALIGNED / LIVE SMALL DEFICIT'
        else:
            adv['alignment'] = 'OVER AUDIT TOLERANCE / LIVE DEFICIT'

    adv['version'] = ADVISOR_VERSION
    adv['audit_zone_policy'] = policy
    adv['serious_blockers'] = serious
    adv['action'] = action
    adv['status'] = status
    adv['p_final'] = p_final
    adv['p_final_formula_changed'] = False
    item['advisor'] = adv
    item['system_action'] = action
    item['system_status'] = status
    item['stake'] = '0%' if action == 'PASS' else ('10-15% live-limit' if action == 'RISK' else '15-20% live-limit')
    item['p_final_system'] = p_final
    return item


def _v11_compact_line(item: Optional[dict[str, Any]]) -> Optional[dict[str, Any]]:
    row = _V111_BASE_COMPACT(item)
    if row is not None and item is not None:
        row['audit_zone_policy'] = deepcopy((item.get('advisor') or {}).get('audit_zone_policy') or {})
        row['p_final_formula_changed'] = False
    return row


def _v11_line_block(item: dict[str, Any], index: int) -> str:
    text = _V111_BASE_LINE_BLOCK(item, index)
    policy = (item.get('advisor') or {}).get('audit_zone_policy') or {}
    if not policy.get('applied'):
        return text
    lines = [text, '<b>Асиметричний аудит зон:</b>']
    lines.append(html.escape(str(policy.get('reason_uk') or 'Правило не активне.')))
    if policy.get('audit_evidence_uk'):
        lines.append(html.escape(str(policy.get('audit_evidence_uk'))))
    if policy.get('removed_blockers'):
        lines.append('<b>Не застосовані старі блокери:</b> ' + html.escape(', '.join(policy['removed_blockers'])))
    lines.append('<b>Важливо:</b> P_history, P_scenario, P_live, P_raw і P_final не змінювалися; змінений лише фінальний advisor-filter.')
    return '\n'.join(lines)


DEFAULT_CONFIG.setdefault('advisor_audit_policy', {})
DEFAULT_CONFIG['advisor_audit_policy'].update({
    'version': ADVISOR_VERSION,
    'formula_components_unchanged': ['P_history', 'P_scenario', 'projections', 'P_live', 'P_raw', 'P_final'],
    'over_zone_min': ADVISOR_OVER_AUDIT_ZONE_MIN,
    'over_small_live_deficit_tolerated': ADVISOR_OVER_SMALL_DEFICIT_MAX,
    'under_hard_conflict_projection_above': ADVISOR_UNDER_HARD_CONFLICT_EDGE,
    'under_elite_zone_min': ADVISOR_UNDER_ELITE_ZONE_MIN,
    'source_note': 'SUPER_BASKET detailed audit of strong historical zones; v11.2 bugfixes quarter routing, stale snapshots, zero-edge selection and delivery integrity.',
})

SYSTEM_VERSION = ADVISOR_VERSION
DEFAULT_CONFIG['engine_version'] = ADVISOR_VERSION


# ============================================================================
# v11.3 CLEAR TELEGRAM EXPLANATIONS — OUTPUT ONLY
# ============================================================================
# This release changes only the human-readable Telegram rendering. It does not
# modify P_history, P_scenario, live projections, P_live, P_raw, P_final,
# scenario mining, caps, blockers, routing or PLAY/RISK/PASS selection.

ADVISOR_VERSION = '11.3.0-CLEAR-TELEGRAM-EXPLANATIONS'


def _v113_side_uk(side: Any) -> str:
    return 'БІЛЬШЕ' if str(side or '').upper() == 'OVER' else 'МЕНШЕ'


def _v113_effect_uk(effect: Any) -> str:
    return {
        'STRONG_SUPPORT': 'сильно підсилює',
        'SUPPORT': 'підсилює',
        'NEUTRAL': 'не дає помітної переваги',
        'WEAKEN': 'послаблює',
        'STRONG_CONFLICT': 'сильно суперечить',
        'INSUFFICIENT': 'має недостатню вибірку',
    }.get(str(effect or '').upper(), str(effect or 'невідомий вплив'))


def _v113_sample_uk(label: Any) -> str:
    return {
        'STRONG_SAMPLE': 'велика вибірка',
        'RELIABLE': 'надійна вибірка',
        'NORMAL': 'нормальна вибірка',
        'SMALL_SAMPLE': 'мала вибірка',
        'INSUFFICIENT': 'недостатня вибірка',
    }.get(str(label or '').upper(), str(label or 'вибірка не визначена'))


def _v113_intersection_uk(value: Any) -> str:
    return {
        'ALIGNED': 'історія команди та профіль суперника підтверджують один напрямок',
        'CONFLICT': 'історія команди та профіль суперника суперечать одне одному',
        'ONE_SIDED': 'підтвердження є лише з одного боку',
        'OFF': 'достатнього перетину немає',
    }.get(str(value or '').upper(), str(value or 'перетин не визначений'))


def _v113_segment_phrase(item: dict[str, Any]) -> str:
    market_type = str(item.get('market_type') or '')
    segment = str(item.get('segment') or '')
    if market_type in {'MATCH_TOTAL', 'TEAM_IT_MATCH'} or segment == 'MATCH':
        return 'за весь матч'
    if market_type in {'H1_TOTAL', 'TEAM_IT_H1'} or segment == 'H1':
        return 'у 1-й половині'
    if market_type in {'H2_TOTAL', 'TEAM_IT_H2'} or segment == 'H2':
        return 'у 2-й половині'
    if segment.startswith('Q'):
        return f'у {segment}'
    return f'у сегменті {segment}' if segment else 'у цьому ринку'


def _v113_bet_sentence(item: dict[str, Any], *, with_odds: bool = True) -> str:
    side = _v113_side_uk(item.get('side'))
    line = _v11_num(item.get('line'))
    team = str(item.get('team') or '').strip()
    market_type = str(item.get('market_type') or '')
    period = _v113_segment_phrase(item)
    if market_type in {'TEAM_IT_MATCH', 'TEAM_IT_H1', 'TEAM_IT_H2', 'CURRENT_QUARTER_TEAM_IT'} or team:
        text = f'{team or "Команда"} набере {side} {line} очка {period}'
    else:
        text = f'команди разом наберуть {side} {line} очка {period}'
    odds = to_number(item.get('odds'))
    if with_odds and odds is not None and not bool((item.get('advisor') or {}).get('is_model_line')):
        text += f' за коефіцієнт {odds:.2f}'
    return text


def _v113_target_sentence_from_pattern(row: dict[str, Any], fallback_item: dict[str, Any]) -> str:
    # The pattern already stores the exact target market/side/line. Rebuild a
    # human sentence from the current evaluated item so team names are never lost.
    item = dict(fallback_item)
    if row.get('target_side'):
        item['side'] = row.get('target_side')
    if row.get('target_line') is not None:
        item['line'] = row.get('target_line')
    return _v113_bet_sentence(item, with_odds=False)


def _v113_meaningful_rate(label: str, value: Any) -> Optional[str]:
    number = to_number(value)
    if number is None:
        return None
    # Do not clutter Telegram with ordinary 40–60% observations. Show only
    # genuinely informative continuation patterns.
    if 0.30 < number < 0.70:
        return None
    qualifier = 'часто' if number >= 0.70 else 'рідко'
    return f'{label} {qualifier}: {_v11_pct(number)}'


def _v113_scenario_lines(scenario: dict[str, Any], item: dict[str, Any], limit: int = 2) -> list[str]:
    support = list((scenario or {}).get('top_support') or [])
    conflict = list((scenario or {}).get('top_conflict') or [])
    rows = support[:limit]
    if not rows and conflict:
        rows = conflict[:limit]
    elif conflict and len(rows) < limit:
        rows += conflict[: limit - len(rows)]
    output: list[str] = []
    for idx, row in enumerate(rows, 1):
        own = row.get('own') or {}
        allowed = row.get('opponent_allowed') or {}
        own_hits, own_n = int(own.get('hits') or 0), int(own.get('n') or 0)
        opp_hits, opp_n = int(allowed.get('hits') or 0), int(allowed.get('n') or 0)
        total_hits, total_n = own_hits + opp_hits, own_n + opp_n
        target = _v113_target_sentence_from_pattern(row, item)
        source_team = str(row.get('source_team') or 'Команда')
        description = str(row.get('description') or row.get('title') or 'Схожий сценарій')
        rate = row.get('combined_rate')
        effect = _v113_effect_uk(row.get('effect'))
        sample = _v113_sample_uk(row.get('sample_label'))
        intersection = _v113_intersection_uk(row.get('intersection'))

        output.append(f'<b>Сценарій {idx}:</b> {html.escape(description)}')
        if total_n:
            output.append(
                f'У схожих матчах варіант «{html.escape(target)}» пройшов '
                f'<b>{total_hits} із {total_n}</b> разів; скоригована сценарна оцінка — <b>{_v11_pct(rate)}</b>.'
            )
        else:
            output.append(f'Скоригована сценарна оцінка для цієї лінії — <b>{_v11_pct(rate)}</b>.')
        detail_bits = []
        if own_n:
            detail_bits.append(f'за історією {source_team}: {own_hits}/{own_n}')
        if opp_n:
            detail_bits.append(f'за дзеркальним профілем суперника: {opp_hits}/{opp_n}')
        if detail_bits:
            output.append(html.escape('Джерела: ' + '; '.join(detail_bits) + '.'))
        output.append(
            f'<b>Вплив:</b> сценарій {html.escape(effect)} цю рекомендацію; '
            f'{html.escape(intersection)}; {html.escape(sample)}.'
        )

        continuation = []
        # Match/next-quarter outcomes are useful at both strong and very weak
        # frequencies. Rare 4–0 or <21-in-all-four facts are omitted unless they
        # are genuinely common, otherwise the message becomes harder to read.
        for label, value in (
            (f'{source_team} вигравала матч', own.get('team_won_match_rate')),
            (f'{source_team} вигравала наступну чверть', own.get('won_next_quarter_rate')),
        ):
            text = _v113_meaningful_rate(label, value)
            if text:
                continuation.append(text)
        sweep = to_number(own.get('won_all_4_quarters_rate'))
        if sweep is not None and sweep >= 0.50:
            continuation.append(f'{source_team} завершувала матч 4–0 по чвертях часто: {_v11_pct(sweep)}')
        no21 = to_number(own.get('under_21_all_4_quarters_rate'))
        if no21 is not None and no21 >= 0.50:
            continuation.append(f'{source_team} не набирала 21 очко в жодній чверті часто: {_v11_pct(no21)}')
        if continuation:
            output.append('<b>Що часто відбувалося далі:</b> ' + html.escape('; '.join(continuation) + '.'))
    return output


def _v113_projection_explanation(item: dict[str, Any]) -> str:
    adv = item.get('advisor') or {}
    projection = to_number(adv.get('projection_used'))
    line = to_number(item.get('line'))
    edge = to_number(adv.get('line_edge'))
    if projection is None or line is None:
        return 'Live-проєкція або лінія відсутня.'
    side = str(item.get('side') or '').upper()
    if projection > line:
        relative = f'на {projection - line:.1f} очка вище лінії'
    elif projection < line:
        relative = f'на {line - projection:.1f} очка нижче лінії'
    else:
        relative = 'точно на рівні лінії'
    direction = 'підтримує ставку' if edge is not None and edge >= 0 else 'не підтримує ставку'
    return f'Модель очікує {projection:.1f} очка — це {relative}; для {side} така проєкція {direction}.'


def _v113_plain_conclusion(item: dict[str, Any]) -> list[str]:
    adv = item.get('advisor') or {}
    action = str(adv.get('action') or 'PASS')
    bet = _v113_bet_sentence(item)
    zone = _v11_pct(adv.get('history_zone_rate'))
    p_final = _v11_pct(adv.get('p_final'))
    scenario_miner = adv.get('scenario_miner') or {}
    scenario_rate = scenario_miner.get('p_scenario_miner')
    lines = ['<b>ПРОСТИЙ ВИСНОВОК</b>']
    if action == 'PLAY':
        lines.append(f'<b>БРАТИ:</b> {html.escape(bet)}.')
    elif action == 'RISK':
        lines.append(f'<b>МОЖНА РОЗГЛЯДАТИ ЯК RISK:</b> {html.escape(bet)}.')
    else:
        lines.append(f'<b>ЗАРАЗ НЕ БРАТИ:</b> {html.escape(bet)}.')
    lines.append(
        f'Історична зона — <b>{zone}</b>, підсумкова оцінка P_final — <b>{p_final}</b>. '
        + html.escape(_v113_projection_explanation(item))
    )
    if scenario_rate is not None:
        lines.append(
            f'Автоматичний пошук схожих сценаріїв оцінює підтримку цієї конкретної лінії у <b>{_v11_pct(scenario_rate)}</b>. '
            'Це пояснювальна сценарна оцінка; формулу P_final вона не підміняє.'
        )
    blockers = list(adv.get('serious_blockers') or [])
    if action == 'PASS' and blockers:
        lines.append('<b>Головна причина PASS:</b> ' + html.escape(', '.join(blockers)) + '.')
    elif action == 'PLAY':
        lines.append('Історія, сценарії та live не мають критичного конфлікту, тому це найкращий чистий варіант із поточного snapshot.')
    elif action == 'RISK':
        lines.append('Напрямок має підтримку, але не всі умови достатньо сильні для чистого PLAY.')
    return lines


def _v11_line_block(item: dict[str, Any], index: int) -> str:
    adv = item.get('advisor') or {}
    hist = item.get('history') or {}
    scenario = adv.get('scenario_miner') or {}
    line_type = 'MODEL' if adv.get('is_model_line') else 'REAL'
    action = str(adv.get('action') or 'PASS')
    zone = adv.get('history_zone_rate')
    hits = hist.get('history_zone_hits')
    n = hist.get('history_zone_n')
    zone_fact = f'{hits}/{n}' if hits is not None and n else 'N/A'
    recommendation = _v113_bet_sentence(item)
    lines = [
        f'<b>#{index} {html.escape(action)}</b>',
        f'<b>Конкретна рекомендація:</b> {html.escape(recommendation)}.',
        f'<b>Тип лінії:</b> {line_type}' + (f' | <b>Букмекер:</b> {html.escape(str(item.get("bookmaker") or "N/A"))}' if not adv.get('is_model_line') else ''),
        f'<b>Історична зона:</b> {_v11_pct(zone)} ({html.escape(zone_fact)})',
        f'<b>P_history:</b> {_v11_pct(adv.get("p_hist"))}',
        f'<b>P_scenario у формулі:</b> {_v11_pct(adv.get("p_scenario_core"))}',
        f'<b>P_live:</b> {_v11_pct(adv.get("p_live"))} | <b>P_final:</b> {_v11_pct(adv.get("p_final"))}',
        f'<b>LiveProjection:</b> {_v11_num(adv.get("projection_used"))} | <b>Edge для ставки:</b> {_v11_num(adv.get("line_edge"))}',
        f'<b>Пояснення проєкції:</b> {html.escape(_v113_projection_explanation(item))}',
        f'<b>Узгодження:</b> {html.escape(str(adv.get("alignment") or "N/A"))}',
        f'<b>Статистика:</b> {html.escape(str(adv.get("stat_gate_status") or "OFF"))} | FAKE OVER: {"YES" if adv.get("fake_over") else "NO"} | FAKE UNDER: {"YES" if adv.get("fake_under") else "NO"}',
    ]
    scenario_lines = _v113_scenario_lines(scenario, item)
    if scenario_lines:
        lines.append('<b>ЩО ОЗНАЧАЮТЬ СЦЕНАРІЇ ДЛЯ ЦІЄЇ СТАВКИ</b>')
        lines.extend(scenario_lines)
    else:
        lines.append('<b>Сценарії:</b> достатнього повторюваного сценарію для зрозумілого висновку не знайдено.')
    policy = adv.get('audit_zone_policy') or {}
    if policy.get('applied'):
        lines.append('<b>Асиметричний аудит OVER/UNDER:</b>')
        lines.append(html.escape(str(policy.get('reason_uk') or 'Правило не активне.')))
        if policy.get('audit_evidence_uk'):
            lines.append(html.escape(str(policy.get('audit_evidence_uk'))))
        lines.append('Математичні P_history, P_scenario, P_live, P_raw і P_final не змінювалися; це лише фінальний advisor-filter.')
    if adv.get('serious_blockers'):
        lines.append('<b>Що блокує ставку:</b> ' + html.escape(', '.join(adv['serious_blockers'])))
    lines.extend(_v113_plain_conclusion(item))
    return '\n'.join(lines)


def _v113_trigger_sentence(row: dict[str, Any], *, play: bool) -> str:
    item = {
        'market_type': row.get('market_type'),
        'team': row.get('team'),
        'segment': row.get('segment'),
        'side': row.get('side'),
        'line': row.get('line'),
        'odds': row.get('odds'),
        'advisor': {'is_model_line': True},
    }
    side = str(row.get('side') or '').upper()
    threshold = 'або нижче' if side == 'OVER' else 'або вище'
    kind = 'PLAY' if play else 'RISK'
    return (
        f'{kind} може з’явитися, якщо букмекер дасть лінію, за якої '
        f'{_v113_bet_sentence(item, with_odds=False)} ({threshold}). '
        f'Очікуваний P_final — {_v11_pct(row.get("p_final"))}, edge — {_v11_num(row.get("line_edge"))}.'
    )


def _v11_model_trigger_text(summary: dict[str, Any], index: int) -> str:
    lines = [
        f'<b>MODEL #{index}: {html.escape(str(summary.get("market_label")))}</b>',
        f'<b>Поточна live-проєкція:</b> {_v11_num(summary.get("projection_used"))}',
    ]
    row, _ = _v11_best_model_trigger(summary)
    if row:
        is_play = float(row.get('p_final') or 0.0) >= ADVISOR_PLAY_MIN
        lines.append(html.escape(_v113_trigger_sentence(row, play=is_play)))
    else:
        lines.append('Підтвердженого theoretical trigger в історичній зоні 75–100% не знайдено.')
    return '\n'.join(lines)


def _v11_audit_block(items: list[dict[str, Any]], primary: list[dict[str, Any]]) -> str:
    primary_keys = {
        (row.get('market_type'), row.get('team'), row.get('segment'), row.get('side'), to_number(row.get('line')))
        for row in primary
    }
    remaining = [
        row for row in items
        if (row.get('market_type'), row.get('team'), row.get('segment'), row.get('side'), to_number(row.get('line'))) not in primary_keys
    ]
    if not remaining:
        return ''
    lines = ['<b>ДОДАТКОВІ ЛІНІЇ 75–100% / EXCEPTIONAL EDGE</b>']
    for row in remaining:
        adv = row.get('advisor') or {}
        action = str(adv.get('action') or 'PASS')
        bet = _v113_bet_sentence(row)
        lines.append(
            f'• <b>{html.escape(action)}</b>: {html.escape(bet)}. '
            f'Зона {_v11_pct(adv.get("history_zone_rate"))}; '
            f'P_final {_v11_pct(adv.get("p_final"))}; '
            f'live-проєкція {_v11_num(adv.get("projection_used"))}; '
            f'edge {_v11_num(adv.get("line_edge"))}.'
        )
    return '\n'.join(lines)


def _v113_global_conclusion(advisor: dict[str, Any]) -> str:
    action = str(advisor.get('action') or 'PASS')
    primary = list(advisor.get('primary_lines') or [])
    selected = next((row for row in primary if str((row.get('advisor') or {}).get('action') or '') == action), None)
    selected = selected or (primary[0] if primary else None)
    lines = ['<b>ФІНАЛЬНА РЕКОМЕНДАЦІЯ ПРОСТИМИ СЛОВАМИ</b>']
    if selected:
        bet = _v113_bet_sentence(selected)
        if action == 'PLAY':
            lines.append(f'<b>Найкращий варіант зараз:</b> {html.escape(bet)}.')
            lines.append('Це PLAY: ставка має найкраще узгодження історії, сценаріїв і live серед перевірених ліній.')
        elif action == 'RISK':
            lines.append(f'<b>Найкращий ризиковий варіант:</b> {html.escape(bet)}.')
            lines.append('Це не чистий PLAY: підтримка є, але залишається ризик або неповне підтвердження.')
        else:
            lines.append(f'<b>Поточну лінію не брати:</b> {html.escape(bet)}.')
    else:
        lines.append('<b>Зараз немає реальної лінії, яку радник дозволяє брати.</b>')

    theoretical = advisor.get('nearest_theoretical_play') or {}
    trigger = theoretical.get('play_trigger') or theoretical.get('risk_trigger')
    if trigger:
        lines.append('<b>Що може стати хорошою ставкою далі:</b> ' + html.escape(
            _v113_trigger_sentence(trigger, play=bool(theoretical.get('play_trigger')))
        ))
    elif action == 'PASS':
        lines.append('Зрозумілого теоретичного тригера в історичній зоні 75–100% поки немає.')
    return '\n'.join(lines)


def _v11_build_messages(advisor: dict[str, Any], calculation: dict[str, Any]) -> list[str]:
    snapshot = calculation['canonical_snapshot']
    score = snapshot.get('score') or {}
    quarters = snapshot.get('quarters') or []
    qtext = ' | '.join(
        f"Q{i + 1} {q.get('home')}:{q.get('away')}"
        for i, q in enumerate(quarters)
        if q.get('home') is not None and q.get('away') is not None
    )
    action = advisor['action']
    icon = '✅' if action == 'PLAY' else '⚠️' if action == 'RISK' else '❌'
    header = '\n'.join([
        f'<b>{icon} {html.escape(action)}</b>',
        f'<b>Матч:</b> {html.escape(str(snapshot.get("name")))}',
        f'<b>Стадія:</b> {html.escape(str(snapshot.get("stage")))} | <b>Рахунок:</b> {score.get("home")}:{score.get("away")}',
        f'<b>Чверті:</b> {html.escape(qtext or "N/A")}',
        f'<b>Чому матч надіслано:</b> {html.escape(str(advisor.get("dispatch_reason")))}',
    ])
    primary_lines = advisor.get('primary_lines') or []
    blocks = [_v11_line_block(item, idx) for idx, item in enumerate(primary_lines, 1)]
    if not blocks:
        blocks = [_v11_model_trigger_text(row, idx) for idx, row in enumerate((advisor.get('model_summary') or [])[:ADVISOR_MAX_PRIMARY], 1)]
    audit_block = _v11_audit_block(advisor.get('all_qualifying_real_lines') or [], primary_lines)
    if audit_block:
        blocks.append(audit_block)
    theoretical = advisor.get('nearest_theoretical_play')
    if theoretical:
        trigger_lines = [
            '<b>НАЙБЛИЖЧА ТЕОРЕТИЧНА МОЖЛИВІСТЬ</b>',
            f'<b>Ринок:</b> {html.escape(str(theoretical.get("market_label")))}',
            f'<b>Поточна live-проєкція:</b> {_v11_num(theoretical.get("projection_used"))}',
        ]
        if theoretical.get('play_trigger'):
            trigger_lines.append(html.escape(_v113_trigger_sentence(theoretical['play_trigger'], play=True)))
        elif theoretical.get('risk_trigger'):
            trigger_lines.append(html.escape(_v113_trigger_sentence(theoretical['risk_trigger'], play=False)))
        else:
            trigger_lines.append('Підтвердженої лінії 75–100% поки немає.')
        blocks.append('\n'.join(trigger_lines))
    blocks.append(_v113_global_conclusion(advisor))

    messages: list[str] = []
    current = header
    for block in blocks:
        candidate = current + '\n\n' + block
        if len(candidate) > 3900 and current != header:
            messages.append(current)
            current = '<b>Продовження розрахунку</b>\n' + block
        elif len(candidate) > 3900:
            messages.append(current)
            current = '<b>Продовження розрахунку</b>'
            for line in block.splitlines():
                if len(current) + len(line) + 1 > 3900:
                    messages.append(current)
                    current = '<b>Продовження розрахунку</b>\n' + line
                else:
                    current += '\n' + line
        else:
            current = candidate
    if current:
        messages.append(current)
    return messages


DEFAULT_CONFIG.setdefault('telegram_explanation_policy', {})
DEFAULT_CONFIG['telegram_explanation_policy'].update({
    'version': ADVISOR_VERSION,
    'math_changed': False,
    'team_name_required_for_team_it': True,
    'plain_conclusion_at_end': True,
    'scenario_target_line_explained': True,
    'scenario_meaningful_continuation_only': True,
})
SYSTEM_VERSION = ADVISOR_VERSION
DEFAULT_CONFIG['engine_version'] = ADVISOR_VERSION


# =============================================================================
# v11.4 SCORE-FLOOR / ALREADY-CROSSED-LINE BUGFIX
# =============================================================================
# Fixes impossible recommendations such as:
# - a live projection below points already scored in the same scope;
# - an OVER trigger below a team's current score (already crossed);
# - model triggers for completed scopes or an in-play quarter without reliable time.
# The historical/scenario formulas are unchanged. P_live/P_final change only where
# the former live projection violated the mathematical score floor.

ADVISOR_VERSION = '11.4.0-SCORE-FLOOR-CROSSED-LINE-BUGFIX'
_V114_CALCULATE_LIVE_PROJECTION_BASE = calculate_live_projection


def _v114_scope_state(spec_or_market: dict[str, Any], canonical: dict[str, Any]) -> dict[str, Any]:
    """Return the score/time already consumed by the exact market scope."""
    market = {
        'market_type': spec_or_market.get('market_type'),
        'team': spec_or_market.get('team'),
        'segment': spec_or_market.get('segment') or 'MATCH',
        'side': spec_or_market.get('side') or 'OVER',
        'line': float(to_number(spec_or_market.get('line')) or 0.5),
    }
    try:
        clock = _segment_clock(market, canonical)
    except Exception:
        return {
            'valid': False,
            'current_points': None,
            'remaining_seconds': None,
            'elapsed_seconds': None,
            'reason': 'SCOPE_CLOCK_ERROR',
        }
    current = to_number(clock.get('current_points'))
    remaining = to_number(clock.get('remaining_seconds'))
    elapsed = to_number(clock.get('elapsed_seconds'))
    segment = str(market.get('segment') or '')
    current_quarter = to_int(canonical.get('current_quarter'))
    target_quarter = int(segment[1:]) if segment.startswith('Q') and segment[1:].isdigit() else None
    reliable = bool((canonical.get('data_gate') or {}).get('time_reliable', True))
    if remaining is not None and remaining <= 0:
        return {
            'valid': False,
            'current_points': current,
            'remaining_seconds': remaining,
            'elapsed_seconds': elapsed,
            'reason': 'SCOPE_ALREADY_COMPLETED',
        }
    if target_quarter is not None:
        if current_quarter is not None and target_quarter < current_quarter:
            return {
                'valid': False,
                'current_points': current,
                'remaining_seconds': remaining,
                'elapsed_seconds': elapsed,
                'reason': 'PAST_QUARTER',
            }
        if current_quarter == target_quarter and not reliable:
            return {
                'valid': False,
                'current_points': current,
                'remaining_seconds': remaining,
                'elapsed_seconds': elapsed,
                'reason': 'CURRENT_QUARTER_TIME_UNRELIABLE',
            }
    return {
        'valid': current is not None and remaining is not None,
        'current_points': current,
        'remaining_seconds': remaining,
        'elapsed_seconds': elapsed,
        'reason': 'OK',
    }


def _v114_recalculate_live_probability(result: dict[str, Any], market: dict[str, Any], projection: float) -> None:
    line = float(market['line'])
    side = str(market.get('side') or '').upper()
    sigma = float(to_number(result.get('sigma')) or 1.0)
    edge_over = projection - line
    edge_under = line - projection
    edge = edge_over if side == 'OVER' else edge_under
    z_score = edge / sigma if sigma > 0 else 0.0
    result['projection_used'] = projection
    result['Projection_used'] = projection
    result['projection_model_live'] = projection
    result['line_edge'] = edge
    result['line_edge_over'] = edge_over
    result['line_edge_under'] = edge_under
    result['projection_minus_line'] = projection - line
    result['z_score'] = z_score
    result['p_live'] = normal_cdf(z_score) if sigma > 0 else 0.50


def calculate_live_projection(
    market: dict[str, Any],
    canonical: dict[str, Any],
    history: dict[str, Any],
    scenario: dict[str, Any],
    config: dict[str, Any],
    stat: Optional[dict[str, Any]] = None,
) -> dict[str, Any]:
    result = _V114_CALCULATE_LIVE_PROJECTION_BASE(
        market, canonical, history, scenario, config, stat
    )
    current = to_number(result.get('current_points'))
    remaining = to_number(result.get('remaining_seconds'))
    projection = to_number(result.get('projection_used'))
    if current is None or projection is None:
        return result

    original = float(projection)
    corrected = original
    reason = None
    if remaining is not None and remaining <= 0:
        corrected = float(current)
        reason = 'COMPLETED_SCOPE_EQUALS_FINAL_SCORE'
    elif original < float(current):
        # A final projection can never be lower than points already on the board.
        # Prefer a valid score/time channel; otherwise use the strict score floor.
        score_time = to_number(result.get('projection_score_time'))
        simple = to_number(result.get('projection_simple'))
        valid_live = [
            float(value) for value in (score_time, simple)
            if value is not None and float(value) >= float(current)
        ]
        corrected = min(valid_live) if valid_live else float(current)
        reason = 'PROJECTION_BELOW_CURRENT_SCORE'

    if corrected != original:
        _v114_recalculate_live_probability(result, market, corrected)
        result['projection_floor_applied'] = True
        result['projection_before_score_floor'] = original
        result['projection_score_floor'] = float(current)
        result['projection_floor_reason'] = reason
        result['projection_formula'] = (
            str(result.get('projection_formula') or '')
            + ' Final projection is clamped to points already scored in this market scope.'
        ).strip()
    else:
        result['projection_floor_applied'] = False
    return result


def _v114_projection_from_probe(
    probe: dict[str, Any],
    current_points: float,
    remaining_seconds: float,
) -> tuple[Optional[float], dict[str, Any]]:
    live = probe.get('live') or {}
    raw_projection = to_number(live.get('projection_used'))
    score_time = to_number(live.get('projection_score_time'))
    simple = to_number(live.get('projection_simple'))
    candidates = []
    for source, value in (
        ('projection_used', raw_projection),
        ('projection_score_time', score_time),
        ('projection_simple', simple),
    ):
        if value is not None and float(value) >= current_points:
            candidates.append((source, float(value)))
    if remaining_seconds <= 0:
        return None, {'reason': 'SCOPE_ALREADY_COMPLETED'}
    if not candidates:
        return None, {
            'reason': 'NO_PROJECTION_ABOVE_CURRENT_SCORE',
            'raw_projection': raw_projection,
            'score_time': score_time,
            'simple': simple,
            'current_points': current_points,
        }
    # Keep the clean core projection when valid; otherwise use the valid score/time channel.
    source, projection = candidates[0]
    return projection, {
        'reason': 'OK',
        'selected_source': source,
        'raw_projection': raw_projection,
        'current_points': current_points,
    }


def _v114_model_line_is_open(line: float, current_points: float) -> bool:
    # Both a live OVER and live UNDER line must remain above the score already made.
    # Otherwise the OVER has already crossed and the UNDER has already lost/closed.
    return float(line) >= _v11_round_half(float(current_points) + 0.01)


def _v11_evaluate_model_grid(
    calculator: SuperBasketCalculator,
    canonical: dict[str, Any],
    spec: dict[str, Any],
) -> list[dict[str, Any]]:
    scope = _v114_scope_state(spec, canonical)
    if not scope.get('valid'):
        return []
    current_points = float(scope['current_points'])
    remaining_seconds = float(scope['remaining_seconds'])

    values = _v11_history_values(canonical, spec)
    center = _v11_model_center(canonical, spec)
    if center is None or not values:
        return []

    probe_line = max(float(center), _v11_round_half(current_points + 0.01))
    probe_market = _v11_synthetic_market(spec, probe_line, 'OVER', tag='PROBE_V114')
    probe = calculator.evaluate_market(probe_market, canonical)
    projection, projection_audit = _v114_projection_from_probe(
        probe, current_points, remaining_seconds
    )
    if projection is None:
        return []
    projection = max(float(projection), current_points)

    p10 = percentile(values, 0.10) or min(values)
    p90 = percentile(values, 0.90) or max(values)
    realistic_low = max(
        0.5,
        _v11_round_half(p10 - 2.0),
        _v11_round_half(current_points + 0.01),
    )
    realistic_high = _v11_round_half(p90 + 2.0)
    if realistic_high < realistic_low:
        # History is already below the live score; no unopened historical-zone line exists.
        return []

    candidate_lines: set[tuple[str, float]] = set()
    for side in ('OVER', 'UNDER'):
        center_line = _v11_round_half(max(projection, current_points + 0.01))
        candidate_lines.add((side, center_line))
    for offset in ADVISOR_MODEL_OFFSETS:
        candidate_lines.add(('OVER', _v11_round_half(projection - offset)))
        candidate_lines.add(('UNDER', _v11_round_half(projection + offset)))

    evaluations: list[dict[str, Any]] = []
    for side, line in sorted(candidate_lines, key=lambda row: (row[0], row[1])):
        if line < realistic_low or line > realistic_high:
            continue
        if not _v114_model_line_is_open(line, current_points):
            continue
        evaluated = _v11_light_model_evaluation(
            spec, line, side, projection, values, canonical, probe
        )
        evaluated['live']['current_points'] = current_points
        evaluated['live']['remaining_seconds'] = remaining_seconds
        evaluated['live']['projection_floor_applied'] = bool(
            projection_audit.get('raw_projection') is not None
            and float(projection_audit['raw_projection']) < current_points
        )
        evaluated['model_scope_guard'] = {
            'current_points': current_points,
            'remaining_seconds': remaining_seconds,
            'line_above_current_score': True,
            'projection_audit': projection_audit,
        }
        evaluations.append(
            _v11_enrich_evaluation(
                evaluated, canonical, is_model=True, mine_scenarios=False
            )
        )
    return evaluations


def _v11_compact_line(item: Optional[dict[str, Any]]) -> Optional[dict[str, Any]]:
    if not item:
        return None
    adv = item.get('advisor') or {}
    hist = item.get('history') or {}
    live = item.get('live') or {}
    return {
        'market_type': item.get('market_type'),
        'team': item.get('team'),
        'segment': item.get('segment'),
        'side': item.get('side'),
        'line': item.get('line'),
        'odds': item.get('odds'),
        'bookmaker': item.get('bookmaker'),
        'is_model_line': bool(adv.get('is_model_line')),
        'action': adv.get('action'),
        'status': adv.get('status'),
        'history_zone_rate': adv.get('history_zone_rate'),
        'history_zone_hits': hist.get('history_zone_hits'),
        'history_zone_n': hist.get('history_zone_n'),
        'p_hist': adv.get('p_hist'),
        'p_scenario_core': adv.get('p_scenario_core'),
        'p_scenario_miner': adv.get('p_scenario_miner'),
        'p_live': adv.get('p_live'),
        'p_raw': item.get('p_raw'),
        'p_final': adv.get('p_final'),
        'projection_used': adv.get('projection_used'),
        'line_edge': adv.get('line_edge'),
        'current_points': live.get('current_points'),
        'remaining_seconds': live.get('remaining_seconds'),
        'projection_floor_applied': live.get('projection_floor_applied'),
        'alignment': adv.get('alignment'),
        'fake_over': adv.get('fake_over'),
        'fake_under': adv.get('fake_under'),
        'stat_gate_status': adv.get('stat_gate_status'),
        'serious_blockers': adv.get('serious_blockers'),
        'scenario_miner': adv.get('scenario_miner'),
    }


def _v114_compact_trigger_is_valid(row: Optional[dict[str, Any]]) -> bool:
    if not row:
        return False
    line = to_number(row.get('line'))
    current = to_number(row.get('current_points'))
    projection = to_number(row.get('projection_used'))
    remaining = to_number(row.get('remaining_seconds'))
    if line is None or projection is None:
        return False
    if current is not None:
        if projection < current:
            return False
        if not _v114_model_line_is_open(line, current):
            return False
    if remaining is not None and remaining <= 0:
        return False
    return True


def _v11_model_summary(model_evaluations: list[dict[str, Any]], canonical: dict[str, Any]) -> list[dict[str, Any]]:
    groups: dict[tuple[str, Optional[str], str], list[dict[str, Any]]] = {}
    for item in model_evaluations:
        groups.setdefault(_v11_market_key(item), []).append(item)
    output: list[dict[str, Any]] = []
    for key, rows in groups.items():
        rows = [row for row in rows if _v114_compact_trigger_is_valid(_v11_compact_line(row))]
        if not rows:
            continue
        projection = next((
            to_number((r.get('live') or {}).get('projection_used'))
            for r in rows
            if to_number((r.get('live') or {}).get('projection_used')) is not None
        ), None)
        side_rows: dict[str, list[dict[str, Any]]] = {'OVER': [], 'UNDER': []}
        for row in rows:
            side_rows[str(row.get('side'))].append(row)
        recommendation: dict[str, Any] = {
            'market_key': list(key),
            'projection_used': projection,
            'market_label': _v11_market_label(rows[0]),
        }
        for side in ('OVER', 'UNDER'):
            candidates = sorted(
                side_rows[side],
                key=lambda r: (
                    abs(_v112_edge(r, 999.0)),
                    -float((r.get('advisor') or {}).get('p_final') or 0.0),
                ),
            )
            risk_pool = [
                r for r in candidates
                if (r.get('advisor') or {}).get('p_final', 0) >= ADVISOR_RISK_MIN
                and (r.get('advisor') or {}).get('history_zone_eligible')
                and _v112_edge(r, -999.0) >= 0
                and not (r.get('advisor') or {}).get('serious_blockers')
                and _v114_compact_trigger_is_valid(_v11_compact_line(r))
            ]
            play_pool = [
                r for r in candidates
                if (r.get('advisor') or {}).get('p_final', 0) >= ADVISOR_PLAY_MIN
                and (r.get('advisor') or {}).get('history_zone_eligible')
                and _v112_edge(r, -999.0) >= 0
                and not (r.get('advisor') or {}).get('serious_blockers')
                and _v114_compact_trigger_is_valid(_v11_compact_line(r))
            ]
            risk = min(risk_pool, key=lambda r: abs(_v112_edge(r, 999.0)), default=None)
            play = min(play_pool, key=lambda r: abs(_v112_edge(r, 999.0)), default=None)
            best = candidates[0] if candidates else None
            for selected_row in (best, risk, play):
                if selected_row and (selected_row.get('advisor') or {}).get('scenario_miner', {}).get('deferred'):
                    mined = _v11_mine_scenarios(selected_row, canonical)
                    selected_row['advisor']['scenario_miner'] = mined
                    selected_row['advisor']['p_scenario_miner'] = mined.get('p_scenario_miner')
            recommendation[side.lower()] = {
                'best_model': _v11_compact_line(best) if best else None,
                'risk_trigger': _v11_compact_line(risk) if risk else None,
                'play_trigger': _v11_compact_line(play) if play else None,
            }
        output.append(recommendation)

    def summary_rank(row: dict[str, Any]) -> tuple[float, float, float]:
        trigger, _ = _v11_best_model_trigger(row)
        if not _v114_compact_trigger_is_valid(trigger):
            return (0.0, -999.0, 0.0)
        is_play = 1.0 if float(trigger.get('p_final') or 0.0) >= ADVISOR_PLAY_MIN else 0.0
        return (
            is_play,
            -abs(_v112_number(trigger.get('line_edge'), 999.0)),
            float(trigger.get('p_final') or 0.0),
        )
    output = [row for row in output if _v114_compact_trigger_is_valid(_v11_best_model_trigger(row)[0])]
    output.sort(key=summary_rank, reverse=True)
    return output


def _v11_theoretical_for_pass(
    calculator: SuperBasketCalculator,
    canonical: dict[str, Any],
    evaluation: Optional[dict[str, Any]],
) -> Optional[dict[str, Any]]:
    if not evaluation:
        return None
    spec = {k: evaluation.get(k) for k in ('market_type', 'team', 'segment')}
    scope = _v114_scope_state(spec, canonical)
    if not scope.get('valid'):
        return None
    current_points = float(scope['current_points'])
    remaining_seconds = float(scope['remaining_seconds'])
    projection = to_number((evaluation.get('live') or {}).get('projection_used'))
    if projection is None or projection < current_points:
        projection = to_number((evaluation.get('live') or {}).get('projection_score_time'))
    if projection is None or projection < current_points:
        return None
    projection = float(projection)
    side = str(evaluation.get('side') or '')
    rows = []
    for offset in ADVISOR_MODEL_OFFSETS:
        line = _v11_round_half(projection - offset if side == 'OVER' else projection + offset)
        if line <= 0 or not _v114_model_line_is_open(line, current_points):
            continue
        model = calculator.evaluate_market(
            _v11_synthetic_market(spec, line, side, tag='PASS_TRIGGER_V114'),
            canonical,
        )
        model_live = model.get('live') or {}
        model_live['current_points'] = current_points
        model_live['remaining_seconds'] = remaining_seconds
        enriched = _v11_enrich_evaluation(model, canonical, is_model=True)
        if _v114_compact_trigger_is_valid(_v11_compact_line(enriched)):
            rows.append(enriched)
    candidates = sorted(
        rows,
        key=lambda r: float((r.get('advisor') or {}).get('p_final') or 0.0),
        reverse=True,
    )
    play = next((
        r for r in candidates
        if (r.get('advisor') or {}).get('history_zone_eligible')
        and (r.get('advisor') or {}).get('p_final', 0) >= ADVISOR_PLAY_MIN
        and _v112_edge(r, -999.0) >= 0
        and not (r.get('advisor') or {}).get('serious_blockers')
    ), None)
    risk = next((
        r for r in candidates
        if (r.get('advisor') or {}).get('history_zone_eligible')
        and (r.get('advisor') or {}).get('p_final', 0) >= ADVISOR_RISK_MIN
        and _v112_edge(r, -999.0) >= 0
        and not (r.get('advisor') or {}).get('serious_blockers')
    ), None)
    result = {
        'market_label': _v11_market_label(evaluation),
        'side': side,
        'projection_used': projection,
        'current_points': current_points,
        'remaining_seconds': remaining_seconds,
        'risk_trigger': _v11_compact_line(risk),
        'play_trigger': _v11_compact_line(play),
    }
    if not _v114_compact_trigger_is_valid(result.get('play_trigger')):
        result['play_trigger'] = None
    if not _v114_compact_trigger_is_valid(result.get('risk_trigger')):
        result['risk_trigger'] = None
    return result if result.get('play_trigger') or result.get('risk_trigger') else None


def _v11_best_model_trigger(summary: dict[str, Any]) -> tuple[Optional[dict[str, Any]], str]:
    plays = []
    risks = []
    for side_key, side_label in (('over', 'OVER'), ('under', 'UNDER')):
        group = summary.get(side_key) or {}
        if _v114_compact_trigger_is_valid(group.get('play_trigger')):
            plays.append((group['play_trigger'], side_label))
        if _v114_compact_trigger_is_valid(group.get('risk_trigger')):
            risks.append((group['risk_trigger'], side_label))
    pool = plays if plays else risks
    if not pool:
        return None, ''
    row, side_label = min(
        pool,
        key=lambda pair: (
            abs(_v112_number(pair[0].get('line_edge'), 999.0)),
            -float(pair[0].get('p_final') or 0.0),
        ),
    )
    return row, side_label


def _v113_trigger_sentence(row: dict[str, Any], *, play: bool) -> str:
    if not _v114_compact_trigger_is_valid(row):
        return 'Теоретичний тригер скасовано: лінія вже перетнута рахунком або проєкція некоректна.'
    item = {
        'market_type': row.get('market_type'),
        'team': row.get('team'),
        'segment': row.get('segment'),
        'side': row.get('side'),
        'line': row.get('line'),
        'odds': row.get('odds'),
        'advisor': {'is_model_line': True},
    }
    side = str(row.get('side') or '').upper()
    threshold = 'або нижче' if side == 'OVER' else 'або вище'
    kind = 'PLAY' if play else 'RISK'
    current = to_number(row.get('current_points'))
    current_text = (
        f' Зараз у цьому ринку вже набрано {current:.1f} очка.'
        if current is not None else ''
    )
    return (
        f'{kind} може з’явитися, якщо букмекер дасть лінію, за якої '
        f'{_v113_bet_sentence(item, with_odds=False)} ({threshold}).'
        f'{current_text} Очікуваний P_final — {_v11_pct(row.get("p_final"))}, '
        f'edge — {_v11_num(row.get("line_edge"))}.'
    )


_V114_BUILD_MESSAGES_BASE = _v11_build_messages

def _v11_build_messages(advisor: dict[str, Any], calculation: dict[str, Any]) -> list[str]:
    # Reuse the clear v11.3 message format, but add the exact clock and remove any
    # stale model trigger that fails the score-floor guard.
    advisor = deepcopy(advisor)
    valid_summaries = []
    for summary in advisor.get('model_summary') or []:
        trigger, _ = _v11_best_model_trigger(summary)
        if _v114_compact_trigger_is_valid(trigger):
            valid_summaries.append(summary)
    advisor['model_summary'] = valid_summaries
    theoretical = advisor.get('nearest_theoretical_play') or {}
    if theoretical:
        if not _v114_compact_trigger_is_valid(theoretical.get('play_trigger')):
            theoretical['play_trigger'] = None
        if not _v114_compact_trigger_is_valid(theoretical.get('risk_trigger')):
            theoretical['risk_trigger'] = None
        if not theoretical.get('play_trigger') and not theoretical.get('risk_trigger'):
            advisor['nearest_theoretical_play'] = None
    messages = _V114_BUILD_MESSAGES_BASE(advisor, calculation)
    snapshot = calculation.get('canonical_snapshot') or {}
    clock = snapshot.get('clock')
    if clock:
        updated = []
        for index, message in enumerate(messages):
            if index == 0 and '<b>Стадія:</b>' in message and '<b>Час:' not in message:
                message = message.replace(
                    f'<b>Стадія:</b> {html.escape(str(snapshot.get("stage")))} |',
                    f'<b>Стадія:</b> {html.escape(str(snapshot.get("stage")))} | <b>Час:</b> {html.escape(str(clock))} |',
                    1,
                )
            updated.append(message)
        messages = updated
    return messages


DEFAULT_CONFIG.setdefault('score_floor_policy', {})
DEFAULT_CONFIG['score_floor_policy'].update({
    'version': ADVISOR_VERSION,
    'projection_must_not_be_below_current_score': True,
    'model_line_must_be_above_current_score': True,
    'completed_scope_model_triggers_disabled': True,
    'current_quarter_requires_reliable_time': True,
})
SYSTEM_VERSION = ADVISOR_VERSION
DEFAULT_CONFIG['engine_version'] = ADVISOR_VERSION


# =============================================================================
# v11.5 CLEAR TELEGRAM METRICS
# =============================================================================
# Display-only update requested by the user. It does not change P_history,
# P_scenario, projections, P_live, P_raw, P_final or PLAY/RISK/PASS logic.
# Every real line and model/theoretical trigger now explicitly prints:
# - P_history;
# - P_scenario used in the formula;
# - live projection;
# - the line;
# - human-readable projection-vs-line difference;
# - side-specific edge.

ADVISOR_VERSION = '11.5.0-CLEAR-TELEGRAM-METRICS'


def _v115_projection_difference_text(projection_value: Any, line_value: Any) -> str:
    projection = to_number(projection_value)
    line = to_number(line_value)
    if projection is None or line is None:
        return 'Різницю між проєкцією та лінією неможливо визначити.'
    delta = float(projection) - float(line)
    if delta > 0.0001:
        return f'проєкція на {abs(delta):.1f} очка ВИЩЕ лінії'
    if delta < -0.0001:
        return f'проєкція на {abs(delta):.1f} очка НИЖЧЕ лінії'
    return 'проєкція точно дорівнює лінії'


def _v115_probability_lines(source: dict[str, Any]) -> list[str]:
    scenario_miner = to_number(source.get('p_scenario_miner'))
    lines = [
        f'<b>P_history:</b> {_v11_pct(source.get("p_hist"))}',
        f'<b>P_scenario:</b> {_v11_pct(source.get("p_scenario_core"))}',
    ]
    if scenario_miner is not None:
        lines.append(f'<b>Scenario Miner:</b> {_v11_pct(scenario_miner)}')
    if source.get('p_live') is not None:
        lines.append(f'<b>P_live:</b> {_v11_pct(source.get("p_live"))}')
    if source.get('p_final') is not None:
        lines.append(f'<b>P_final:</b> {_v11_pct(source.get("p_final"))}')
    return lines


def _v115_line_projection_lines(source: dict[str, Any]) -> list[str]:
    side = str(source.get('side') or '').upper() or 'N/A'
    line = to_number(source.get('line'))
    projection = to_number(source.get('projection_used'))
    edge = to_number(source.get('line_edge'))
    line_text = 'N/A' if line is None else f'{line:.1f}'
    projection_text = 'N/A' if projection is None else f'{projection:.1f}'
    return [
        f'<b>Лінія:</b> {html.escape(side)} {line_text}',
        f'<b>Live-проєкція:</b> {projection_text}',
        f'<b>Різниця:</b> {html.escape(_v115_projection_difference_text(projection, line))}',
        f'<b>Edge для {html.escape(side)}:</b> {_v11_num(edge)}',
    ]


def _v11_line_block(item: dict[str, Any], index: int) -> str:
    adv = item.get('advisor') or {}
    hist = item.get('history') or {}
    scenario = adv.get('scenario_miner') or {}
    line_type = 'MODEL' if adv.get('is_model_line') else 'REAL'
    action = str(adv.get('action') or 'PASS')
    zone = adv.get('history_zone_rate')
    hits = hist.get('history_zone_hits')
    n = hist.get('history_zone_n')
    zone_fact = f'{hits}/{n}' if hits is not None and n else 'N/A'
    recommendation = _v113_bet_sentence(item)
    source = {
        **adv,
        'line': item.get('line'),
        'side': item.get('side'),
        'p_scenario_miner': adv.get('p_scenario_miner'),
    }
    lines = [
        f'<b>#{index} {html.escape(action)}</b>',
        f'<b>Конкретна рекомендація:</b> {html.escape(recommendation)}.',
        f'<b>Тип лінії:</b> {line_type}' + (
            f' | <b>Букмекер:</b> {html.escape(str(item.get("bookmaker") or "N/A"))}'
            if not adv.get('is_model_line') else ''
        ),
        f'<b>Історична зона:</b> {_v11_pct(zone)} ({html.escape(zone_fact)})',
    ]
    lines.extend(_v115_probability_lines(source))
    lines.extend(_v115_line_projection_lines(source))
    lines.extend([
        f'<b>Узгодження:</b> {html.escape(str(adv.get("alignment") or "N/A"))}',
        f'<b>Статистика:</b> {html.escape(str(adv.get("stat_gate_status") or "OFF"))} | '
        f'FAKE OVER: {"YES" if adv.get("fake_over") else "NO"} | '
        f'FAKE UNDER: {"YES" if adv.get("fake_under") else "NO"}',
    ])
    scenario_lines = _v113_scenario_lines(scenario, item)
    if scenario_lines:
        lines.append('<b>ЩО ОЗНАЧАЮТЬ СЦЕНАРІЇ ДЛЯ ЦІЄЇ СТАВКИ</b>')
        lines.extend(scenario_lines)
    else:
        lines.append('<b>Сценарії:</b> достатнього повторюваного сценарію для зрозумілого висновку не знайдено.')
    policy = adv.get('audit_zone_policy') or {}
    if policy.get('applied'):
        lines.append('<b>Асиметричний аудит OVER/UNDER:</b>')
        lines.append(html.escape(str(policy.get('reason_uk') or 'Правило не активне.')))
        if policy.get('audit_evidence_uk'):
            lines.append(html.escape(str(policy.get('audit_evidence_uk'))))
        lines.append('Математичні P_history, P_scenario, P_live, P_raw і P_final не змінювалися; це лише фінальний advisor-filter.')
    if adv.get('serious_blockers'):
        lines.append('<b>Що блокує ставку:</b> ' + html.escape(', '.join(adv['serious_blockers'])))
    lines.extend(_v113_plain_conclusion(item))
    return '\n'.join(lines)


def _v113_trigger_sentence(row: dict[str, Any], *, play: bool) -> str:
    if not _v114_compact_trigger_is_valid(row):
        return 'Теоретичний тригер скасовано: лінія вже перетнута рахунком або проєкція некоректна.'
    item = {
        'market_type': row.get('market_type'),
        'team': row.get('team'),
        'segment': row.get('segment'),
        'side': row.get('side'),
        'line': row.get('line'),
        'odds': row.get('odds'),
        'advisor': {'is_model_line': True},
    }
    side = str(row.get('side') or '').upper()
    threshold = 'або нижче' if side == 'OVER' else 'або вище'
    kind = 'PLAY' if play else 'RISK'
    current = to_number(row.get('current_points'))
    current_text = (
        f' Зараз у цьому ринку вже набрано {current:.1f} очка.'
        if current is not None else ''
    )
    return (
        f'{kind} може з’явитися, якщо букмекер дасть лінію, за якої '
        f'{_v113_bet_sentence(item, with_odds=False)} ({threshold}).'
        f'{current_text} P_history — {_v11_pct(row.get("p_hist"))}; '
        f'P_scenario — {_v11_pct(row.get("p_scenario_core"))}; '
        f'live-проєкція — {_v11_num(row.get("projection_used"))}; '
        f'лінія — {side} {_v11_num(row.get("line"))}; '
        f'{_v115_projection_difference_text(row.get("projection_used"), row.get("line"))}; '
        f'edge для {side} — {_v11_num(row.get("line_edge"))}; '
        f'P_final — {_v11_pct(row.get("p_final"))}.'
    )


def _v11_model_trigger_text(summary: dict[str, Any], index: int) -> str:
    row, _ = _v11_best_model_trigger(summary)
    lines = [f'<b>MODEL #{index}: {html.escape(str(summary.get("market_label")))}</b>']
    if not row:
        lines.append('Підтвердженого theoretical trigger в історичній зоні 75–100% не знайдено.')
        return '\n'.join(lines)
    source = dict(row)
    if source.get('projection_used') is None:
        source['projection_used'] = summary.get('projection_used')
    is_play = float(source.get('p_final') or 0.0) >= ADVISOR_PLAY_MIN
    lines.append('<b>Статус:</b> ' + ('теоретичний PLAY' if is_play else 'теоретичний RISK'))
    lines.extend(_v115_probability_lines(source))
    lines.extend(_v115_line_projection_lines(source))
    current = to_number(source.get('current_points'))
    if current is not None:
        lines.append(f'<b>Уже набрано в цьому ринку:</b> {current:.1f}')
    lines.append(html.escape(_v113_trigger_sentence(source, play=is_play)))
    return '\n'.join(lines)


def _v11_audit_block(items: list[dict[str, Any]], primary: list[dict[str, Any]]) -> str:
    primary_keys = {
        (row.get('market_type'), row.get('team'), row.get('segment'), row.get('side'), to_number(row.get('line')))
        for row in primary
    }
    remaining = [
        row for row in items
        if (row.get('market_type'), row.get('team'), row.get('segment'), row.get('side'), to_number(row.get('line'))) not in primary_keys
    ]
    if not remaining:
        return ''
    lines = ['<b>ДОДАТКОВІ ЛІНІЇ 75–100% / EXCEPTIONAL EDGE</b>']
    for row in remaining:
        adv = row.get('advisor') or {}
        action = str(adv.get('action') or 'PASS')
        bet = _v113_bet_sentence(row)
        difference = _v115_projection_difference_text(adv.get('projection_used'), row.get('line'))
        lines.extend([
            f'• <b>{html.escape(action)}</b>: {html.escape(bet)}.',
            f'  P_history {_v11_pct(adv.get("p_hist"))} | '
            f'P_scenario {_v11_pct(adv.get("p_scenario_core"))} | '
            f'P_final {_v11_pct(adv.get("p_final"))}.',
            f'  Лінія {html.escape(str(row.get("side") or ""))} {_v11_num(row.get("line"))} | '
            f'live-проєкція {_v11_num(adv.get("projection_used"))} | '
            f'{html.escape(difference)} | edge {_v11_num(adv.get("line_edge"))}.',
        ])
    return '\n'.join(lines)


DEFAULT_CONFIG.setdefault('telegram_metrics_policy', {})
DEFAULT_CONFIG['telegram_metrics_policy'].update({
    'version': ADVISOR_VERSION,
    'math_changed': False,
    'p_history_required': True,
    'p_scenario_required': True,
    'projection_required': True,
    'line_required': True,
    'projection_line_difference_required': True,
    'auto_split_long_messages': True,
})
SYSTEM_VERSION = ADVISOR_VERSION
DEFAULT_CONFIG['engine_version'] = ADVISOR_VERSION

# =============================================================================
# v11.6 COMPACT TELEGRAM + MARKET SCOPE AUDIT
# =============================================================================
# - Compact Telegram output with the final recommendation at the very top.
# - Keeps only one full primary calculation and up to two compact alternatives.
# - Explicitly reports whether Projection_used came from FULL_STAT/PARTIAL_STAT
#   or score/time fallback.
# - Corrects parser rows where a first-half team total was incorrectly placed
#   into home_ind_total/away_ind_total with scope=Match. The correction is based
#   on consistency with the real H1 total and match total, not on the team name.

ADVISOR_VERSION = '11.6.0-COMPACT-TELEGRAM-MARKET-SCOPE-AUDIT'

_V116_PARSE_MARKETS_BASE = parse_markets

_V116_CALCULATE_LIVE_PROJECTION_BASE = calculate_live_projection


def _v116_ambiguous_quarter_total_keys(source: dict[str, Any]) -> dict[tuple[str, float], dict[str, Any]]:
    containers = (
        source.get('lines')
        or source.get('bookmaker_lines')
        or source.get('bookmaker_markets')
        or source.get('markets')
        or {}
    )
    rows = containers.get('quarter_total') if isinstance(containers, dict) else None
    if not isinstance(rows, list):
        return {}
    grouped: dict[str, list[float]] = {}
    for row in rows:
        if not isinstance(row, dict):
            continue
        line = to_number(row.get('line'))
        if line is None:
            continue
        grouped.setdefault(_scope_text(row) or 'UNKNOWN', []).append(float(line))
    ambiguous: dict[tuple[str, float], dict[str, Any]] = {}
    for scope, values in grouped.items():
        unique = sorted(set(values))
        if len(unique) < 2:
            continue
        gaps = [(unique[i + 1] - unique[i], i) for i in range(len(unique) - 1)]
        gap, index = max(gaps, default=(0.0, 0))
        low = unique[: index + 1]
        high = unique[index + 1 :]
        if not low or not high:
            continue
        # A ~20-point line and a ~40-point line under the same Q scope are
        # not alternative quarter totals. The low row is almost certainly a
        # team IT that lost its team identity in the upstream parser.
        if gap >= 8.0 and max(low) <= 0.72 * min(high):
            for line in low:
                ambiguous[('quarter_total', line)] = {
                    'reason': 'LOW_QUARTER_TOTAL_CLUSTER_LOOKS_LIKE_TEAM_IT',
                    'scope': scope,
                    'low_cluster': low,
                    'normal_total_cluster': high,
                    'action': 'BLOCK_NO_TEAM_ID',
                }
    return ambiguous


def _v116_quarter_stat_projection(
    market: dict[str, Any],
    canonical: dict[str, Any],
    result: dict[str, Any],
) -> dict[str, Any]:
    mode = str(result.get('data_mode') or canonical.get('data_mode') or '')
    if mode not in {'FULL_STAT', 'PARTIAL_STAT'}:
        return result
    if market.get('market_type') not in {'CURRENT_QUARTER_TOTAL', 'CURRENT_QUARTER_TEAM_IT'}:
        return result
    remaining_seconds = to_number(result.get('remaining_seconds'))
    current_points = to_number(result.get('current_points'))
    if remaining_seconds is None or remaining_seconds <= 0 or current_points is None:
        return result

    rho, k_stage, minutes_played = _v10_stage_trust(canonical)
    remaining_minutes = float(remaining_seconds) / 60.0
    pre_home = _v10_pre_stat_team('home', market, canonical)
    pre_away = _v10_pre_stat_team('away', market, canonical)
    live_home = _v10_live_stat_team('home', market, canonical, pre_home, rho, remaining_minutes)
    live_away = _v10_live_stat_team('away', market, canonical, pre_away, rho, remaining_minutes)
    if market.get('team'):
        selected = live_home if market.get('team') == canonical.get('home_team') else live_away
        stat_projection = to_number(selected.get('LiveRaw_Team'))
    else:
        home_projection = to_number(live_home.get('LiveRaw_Team'))
        away_projection = to_number(live_away.get('LiveRaw_Team'))
        stat_projection = (
            float(home_projection) + float(away_projection)
            if home_projection is not None and away_projection is not None
            else None
        )
    if stat_projection is None:
        return result

    score_time = to_number(result.get('projection_score_time'))
    elapsed_seconds = max(0.0, float(to_number(result.get('elapsed_seconds')) or 0.0))
    full_scope_seconds = elapsed_seconds + float(remaining_seconds)
    if score_time is not None and elapsed_seconds > 0 and full_scope_seconds > 0:
        score_weight = min(0.75, max(0.25, elapsed_seconds / full_scope_seconds))
        projection = score_weight * float(score_time) + (1.0 - score_weight) * float(stat_projection)
        formula_mode = f'{mode}_CURRENT_QUARTER_STAT_SCORE_BLEND'
    else:
        score_weight = 0.0
        projection = float(stat_projection)
        formula_mode = f'{mode}_CURRENT_QUARTER_STAT_PACE'
    projection = max(float(current_points), projection)
    _v114_recalculate_live_probability(result, market, projection)
    result['projection_stat_live_only'] = float(stat_projection)
    result['projection_stat_adjusted'] = float(stat_projection)
    result['projection_formula_mode'] = formula_mode
    result['projection_formula'] = (
        'Current-quarter score plus expected remaining points from cumulative '
        'live FGA/FTA/2PA/3PA/ORB/TO rates. Cumulative rates are applied only '
        'to remaining quarter time; they are not treated as quarter-only counts.'
    )
    result['stat_projection_details'] = {
        'quarter_stat_projection_enabled': True,
        'rho_stage': rho,
        'K_stage': k_stage,
        'minutes_played_game': minutes_played,
        'remaining_minutes_quarter': remaining_minutes,
        'score_time_projection': score_time,
        'score_time_weight': score_weight,
        'stat_projection': stat_projection,
        'live_home': live_home,
        'live_away': live_away,
    }
    components = result.setdefault('components', {})
    components['projection_stat_live_only'] = {
        'value': float(stat_projection),
        'included': True,
        'role': 'live_cumulative_rates_applied_to_quarter_remaining_time',
    }
    if 'projection_score_time' in components:
        components['projection_score_time']['included'] = bool(score_time is not None and score_weight > 0)
    return result


def calculate_live_projection(
    market: dict[str, Any],
    canonical: dict[str, Any],
    history: dict[str, Any],
    scenario: dict[str, Any],
    config: dict[str, Any],
    stat: Optional[dict[str, Any]] = None,
) -> dict[str, Any]:
    result = _V116_CALCULATE_LIVE_PROJECTION_BASE(market, canonical, history, scenario, config, stat)
    return _v116_quarter_stat_projection(market, canonical, result)


def _v116_float_lines(rows: Any, *, scope_prefix: Optional[str] = None) -> list[float]:
    values: list[float] = []
    if not isinstance(rows, list):
        return values
    for row in rows:
        if not isinstance(row, dict):
            continue
        scope = _scope_text(row)
        if scope_prefix and not scope.startswith(scope_prefix):
            continue
        value = to_number(row.get('line'))
        if value is not None:
            values.append(float(value))
    return values


def _v116_team_it_scope_overrides(source: dict[str, Any]) -> dict[tuple[str, float], dict[str, Any]]:
    containers = (
        source.get('lines')
        or source.get('bookmaker_lines')
        or source.get('bookmaker_markets')
        or source.get('markets')
        or {}
    )
    if not isinstance(containers, dict):
        return {}

    match_totals = _v116_float_lines(containers.get('match_total'))
    h1_totals = _v116_float_lines(containers.get('half_total'), scope_prefix='H1')
    if not h1_totals:
        h1_totals = [
            float(row.get('line'))
            for row in (containers.get('half_total') or [])
            if isinstance(row, dict)
            and to_number(row.get('line')) is not None
            and _scope_text(row) in {'1H', 'FIRSTHALF'}
        ]

    buckets: dict[str, list[float]] = {}
    for bucket in ('home_ind_total', 'away_ind_total'):
        values: list[float] = []
        for row in containers.get(bucket) or []:
            if not isinstance(row, dict):
                continue
            if _scope_text(row) not in {'', 'MATCH', 'FULLMATCH'}:
                continue
            line = to_number(row.get('line'))
            if line is not None:
                values.append(float(line))
        buckets[bucket] = values

    overrides: dict[tuple[str, float], dict[str, Any]] = {}
    h1_pair_error: dict[tuple[str, float], float] = {}
    match_pair_error: dict[tuple[str, float], float] = {}

    home_values = buckets.get('home_ind_total') or []
    away_values = buckets.get('away_ind_total') or []
    for home_line in home_values:
        for away_line in away_values:
            pair_sum = home_line + away_line
            for total in h1_totals:
                tolerance = max(3.0, 0.045 * total)
                error = abs(pair_sum - total)
                if error <= tolerance:
                    for key in (('home_ind_total', home_line), ('away_ind_total', away_line)):
                        h1_pair_error[key] = min(h1_pair_error.get(key, 9999.0), error)
            for total in match_totals:
                tolerance = max(5.0, 0.045 * total)
                error = abs(pair_sum - total)
                if error <= tolerance:
                    for key in (('home_ind_total', home_line), ('away_ind_total', away_line)):
                        match_pair_error[key] = min(match_pair_error.get(key, 9999.0), error)

    match_center = statistics.median(match_totals) if match_totals else None
    h1_center = statistics.median(h1_totals) if h1_totals else None

    for bucket, values in buckets.items():
        unique = sorted(set(values))
        low_cluster: set[float] = set()
        if len(unique) >= 2:
            gaps = [(unique[i + 1] - unique[i], i) for i in range(len(unique) - 1)]
            largest_gap, split_index = max(gaps, default=(0.0, 0))
            if largest_gap >= max(8.0, 0.18 * statistics.median(unique)):
                low_cluster = set(unique[: split_index + 1])

        for line in unique:
            key = (bucket, line)
            h1_error = h1_pair_error.get(key)
            match_error = match_pair_error.get(key)
            reason = None
            confidence = None

            # Strongest evidence: home+away team totals add up to a real H1 line.
            if h1_error is not None and (match_error is None or h1_error + 0.5 < match_error):
                reason = 'HOME_AWAY_TEAM_IT_SUM_MATCHES_H1_TOTAL'
                confidence = 'HIGH'
            # Bimodal bucket: a low cluster near H1 scale and a high cluster near match scale.
            elif line in low_cluster and h1_center is not None and match_center is not None:
                ratio_h1 = line / h1_center if h1_center else 0.0
                ratio_match = line / match_center if match_center else 0.0
                high_values = [value for value in unique if value not in low_cluster]
                high_match_ok = any(0.32 <= value / match_center <= 0.68 for value in high_values) if match_center else False
                if 0.20 <= ratio_h1 <= 0.80 and ratio_match < 0.40 and high_match_ok:
                    reason = 'BIMODAL_TEAM_IT_BUCKET_LOW_CLUSTER_IS_H1'
                    confidence = 'HIGH'
            # Conservative fallback only when the line is clearly half-scale.
            elif h1_center is not None and match_center is not None:
                ratio_h1 = line / h1_center if h1_center else 0.0
                ratio_match = line / match_center if match_center else 0.0
                if line in low_cluster and 0.20 <= ratio_h1 <= 0.80 and ratio_match < 0.36:
                    reason = 'LOW_TEAM_IT_LINE_CONSISTENT_WITH_H1_NOT_MATCH'
                    confidence = 'MEDIUM'

            if reason:
                overrides[key] = {
                    'market_type': 'TEAM_IT_H1',
                    'segment': 'H1',
                    'reason': reason,
                    'confidence': confidence,
                    'h1_total_center': h1_center,
                    'match_total_center': match_center,
                }
    return overrides


def parse_markets(
    source: dict[str, Any],
    canonical: dict[str, Any],
    config: dict[str, Any],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    evaluations, audit = _V116_PARSE_MARKETS_BASE(source, canonical, config)
    overrides = _v116_team_it_scope_overrides(source)
    ambiguous_quarter = _v116_ambiguous_quarter_total_keys(source)

    for item in evaluations:
        bucket = str(item.get('source_bucket') or '')
        line = to_number(item.get('line'))
        ambiguous = ambiguous_quarter.get((bucket, float(line))) if line is not None else None
        if ambiguous and item.get('market_type') == 'CURRENT_QUARTER_TOTAL':
            item['ambiguous_market_scope'] = deepcopy(ambiguous)
            issues = list(item.get('parser_issues') or [])
            if 'AMBIGUOUS_QUARTER_TOTAL_LOOKS_TEAM_IT_NO_TEAM_ID' not in issues:
                issues.append('AMBIGUOUS_QUARTER_TOTAL_LOOKS_TEAM_IT_NO_TEAM_ID')
            item['parser_issues'] = issues
            item['eligible_market'] = False
        override = overrides.get((bucket, float(line))) if line is not None else None
        if not override:
            continue
        if item.get('market_type') == 'TEAM_IT_MATCH' and item.get('segment') == 'MATCH':
            item['market_type'] = override['market_type']
            item['segment'] = override['segment']
            item['source_scope_original'] = item.get('source_scope')
            item['source_scope'] = 'H1_CORRECTED'
            item['market_scope_correction'] = deepcopy(override)
            issues = [reason for reason in (item.get('parser_issues') or []) if reason != 'AMBIGUOUS_TEAM_IT_SCOPE']
            item['parser_issues'] = issues
            item['eligible_market'] = not issues

    for row in audit:
        bucket = str(row.get('bucket') or '')
        line = to_number(row.get('line'))
        ambiguous = ambiguous_quarter.get((bucket, float(line))) if line is not None else None
        if ambiguous and row.get('market_type') == 'CURRENT_QUARTER_TOTAL':
            row['ambiguous_market_scope'] = deepcopy(ambiguous)
            issues = list(row.get('issues') or [])
            if 'AMBIGUOUS_QUARTER_TOTAL_LOOKS_TEAM_IT_NO_TEAM_ID' not in issues:
                issues.append('AMBIGUOUS_QUARTER_TOTAL_LOOKS_TEAM_IT_NO_TEAM_ID')
            row['issues'] = issues
        override = overrides.get((bucket, float(line))) if line is not None else None
        if not override:
            continue
        if row.get('market_type') == 'TEAM_IT_MATCH' and row.get('segment') == 'MATCH':
            row['market_type_original'] = row.get('market_type')
            row['segment_original'] = row.get('segment')
            row['market_type'] = override['market_type']
            row['segment'] = override['segment']
            row['scope_correction'] = deepcopy(override)
    return evaluations, audit


def _v116_history_fact(item: dict[str, Any]) -> str:
    hist = item.get('history') or {}
    market_type = str(item.get('market_type') or '')
    if market_type in {'TEAM_IT_MATCH', 'TEAM_IT_H1', 'TEAM_IT_H2', 'CURRENT_QUARTER_TEAM_IT'}:
        own = hist.get('own_scored') or {}
        allowed = hist.get('opponent_allowed') or {}
        own_n = to_int(own.get('n')) or 0
        allowed_n = to_int(allowed.get('n')) or 0
        own_text = f"команда {to_int(own.get('wins')) or 0}/{own_n}" if own_n else 'команда N/A'
        allowed_text = f"суперник пропускав {to_int(allowed.get('wins')) or 0}/{allowed_n}" if allowed_n else 'суперник N/A'
        return f'{own_text}; {allowed_text}'
    pooled = hist.get('pooled') or {}
    n = to_int(pooled.get('n')) or 0
    if n:
        return f"{to_int(pooled.get('wins')) or 0}/{n}"
    team_a = hist.get('team_a') or {}
    team_b = hist.get('team_b') or {}
    parts = []
    for label, row in (('A', team_a), ('B', team_b)):
        rn = to_int(row.get('n')) or 0
        if rn:
            parts.append(f'{label} {to_int(row.get("wins")) or 0}/{rn}')
    return '; '.join(parts) or 'N/A'


def _v116_top_scenario_text(item: dict[str, Any]) -> str:
    adv = item.get('advisor') or {}
    miner = adv.get('scenario_miner') or {}
    supports = miner.get('top_support') or []
    conflicts = miner.get('top_conflict') or []
    pattern = supports[0] if supports else conflicts[0] if conflicts else None
    if not pattern:
        return 'Сильного повторюваного сценарію не знайдено.'
    description = str(pattern.get('description') or pattern.get('title') or 'Схожий сценарій')
    n = to_int(pattern.get('combined_n')) or 0
    rate = to_number(pattern.get('combined_rate'))
    hits = None
    own = pattern.get('own') or {}
    allowed = pattern.get('opponent_allowed') or {}
    if n:
        hits = (to_int(own.get('hits')) or 0) + (to_int(allowed.get('hits')) or 0)
    effect_map = {
        'STRONG_SUPPORT': 'сильно підсилює',
        'SUPPORT': 'підсилює',
        'NEUTRAL': 'нейтральний',
        'WEAKEN': 'послаблює',
        'STRONG_CONFLICT': 'сильно суперечить',
    }
    effect = effect_map.get(str(pattern.get('effect') or ''), str(pattern.get('effect') or ''))
    sample = f'{hits}/{n}' if hits is not None and n else f'N={n}' if n else 'N/A'
    rate_text = _v11_pct(rate)
    return f'{description} Для цієї лінії: {sample}, {rate_text}; сценарій {effect} ставку.'


def _v116_projection_source_text(item: dict[str, Any]) -> str:
    live = item.get('live') or {}
    mode = str(item.get('data_mode') or live.get('data_mode') or 'DATA_OFF')
    formula = str(live.get('projection_formula_mode') or '')
    stat_value = to_number(live.get('projection_stat_live_only'))
    score_value = to_number(live.get('projection_score_time'))
    used = to_number(live.get('projection_used'))
    if mode == 'FULL_STAT' and stat_value is not None and ('CURRENT_QUARTER_STAT_' in formula):
        return 'FULL_STAT: проєкція чверті розрахована зі статистичного темпу FGA/FTA/ORB/TO на час, що залишився.'
    if mode == 'FULL_STAT' and stat_value is not None and 'CURRENT_QUARTER_SCORE_TIME' not in formula:
        return (
            f'FULL_STAT: проєкція {used:.1f} розрахована зі статистики FGA/FTA/ORB/TO; score/time {score_value:.1f}.'
            if score_value is not None and used is not None
            else 'FULL_STAT: проєкція побудована зі статистики FGA/FTA/ORB/TO.'
        )
    if mode == 'PARTIAL_STAT' and stat_value is not None and 'CURRENT_QUARTER_SCORE_TIME' not in formula:
        return 'PARTIAL_STAT: проєкція використовує доступні статистичні поля з підвищеною невизначеністю.'
    if mode == 'FULL_STAT' and 'CURRENT_QUARTER_SCORE_TIME' in formula:
        return 'FULL_STAT є, але для окремої чверті використано score/time: boxscore у файлі накопичувальний за матч, а не окремий за чверть.'
    if mode == 'SCORE_TIME_HISTORY':
        return 'NO_STAT: live-проєкція рахується за рахунком і часом; історія та сценарій входять окремо у P_final.'
    return f'Режим проєкції: {mode or "DATA_OFF"}.'


def _v116_compact_main_block(item: dict[str, Any]) -> str:
    adv = item.get('advisor') or {}
    action = str(adv.get('action') or 'PASS')
    bet = _v113_bet_sentence(item)
    hist_fact = _v116_history_fact(item)
    zone = adv.get('history_zone_rate')
    projection = adv.get('projection_used')
    line = item.get('line')
    side = str(item.get('side') or '').upper()
    scenario = _v116_top_scenario_text(item)
    source_text = _v116_projection_source_text(item)

    # The recommendation is already placed in the first line of Telegram.
    # Do not repeat PLAY/RISK/PASS here; keep only the calculation that explains it.
    lines = [
        f'<b>🎯 Основний розрахунок:</b> {html.escape(bet)}.',
        f'<b>Історія:</b> зона {_v11_pct(zone)} ({html.escape(hist_fact)}); '
        f'P_history {_v11_pct(adv.get("p_hist"))}.',
        f'<b>Ймовірності:</b> P_scenario {_v11_pct(adv.get("p_scenario_core"))} | '
        f'P_live {_v11_pct(adv.get("p_live"))} | P_final {_v11_pct(adv.get("p_final"))}.',
        f'<b>Лінія / live:</b> {html.escape(side)} {_v11_num(line)} → {_v11_num(projection)}; '
        f'{html.escape(_v115_projection_difference_text(projection, line))}; edge {_v11_num(adv.get("line_edge"))}.',
        f'<b>Сценарій:</b> {html.escape(scenario)}',
    ]

    mode = str(item.get('data_mode') or (item.get('live') or {}).get('data_mode') or 'DATA_OFF')
    stat_bits = [mode]
    if adv.get('fake_over'):
        stat_bits.append('FAKE OVER')
    if adv.get('fake_under'):
        stat_bits.append('FAKE UNDER')
    gate = str(adv.get('stat_gate_status') or 'OFF')
    if gate not in {'OFF', 'N/A', 'NA', 'N_A_NO_STATS'}:
        stat_bits.append(f'gate {gate}')
    lines.append(f'<b>Статистика:</b> {html.escape("; ".join(stat_bits))}. {html.escape(source_text)}')

    blockers = adv.get('serious_blockers') or []
    if action == 'PLAY':
        lines.append('<b>Підсумок:</b> брати як PLAY, якщо лінія й коефіцієнт ще актуальні.')
    elif action == 'RISK':
        reason = ', '.join(blockers[:2]) if blockers else 'підтвердження неповне'
        lines.append(f'<b>Підсумок:</b> RISK PLAY, не чистий PLAY. Ризик: {html.escape(reason)}.')
    else:
        reason = ', '.join(blockers[:2]) if blockers else 'чистого підтвердження немає'
        lines.append(f'<b>Підсумок:</b> цю лінію не брати. Причина: {html.escape(reason)}.')

    correction = item.get('market_scope_correction') or {}
    if correction:
        lines.append('<b>Уточнення ринку:</b> це IT 1-ї половини, а не IT матчу; scope виправлено за загальною H1-лінією.')
    return '\n'.join(lines)

def _v116_compact_alternative(item: dict[str, Any], index: int) -> str:
    adv = item.get('advisor') or {}
    action = str(adv.get('action') or 'PASS')
    return (
        f'<b>{index}) {html.escape(action)}</b> — {html.escape(_v113_bet_sentence(item))}.\n'
        f'P_final {_v11_pct(adv.get("p_final"))}; зона {_v11_pct(adv.get("history_zone_rate"))}; '
        f'проєкція {_v11_num(adv.get("projection_used"))}; edge {_v11_num(adv.get("line_edge"))}.'
    )


def _v116_compact_trigger(summary: dict[str, Any], index: int) -> Optional[str]:
    row, _ = _v11_best_model_trigger(summary)
    if not row or not _v114_compact_trigger_is_valid(row):
        return None
    is_play = float(row.get('p_final') or 0.0) >= ADVISOR_PLAY_MIN
    item = {
        'market_type': row.get('market_type'),
        'team': row.get('team'),
        'segment': row.get('segment'),
        'side': row.get('side'),
        'line': row.get('line'),
        'odds': row.get('odds'),
        'advisor': {'is_model_line': True},
    }
    return (
        f'<b>{index}) Теоретичний {"PLAY" if is_play else "RISK"}</b> — '
        f'{html.escape(_v113_bet_sentence(item, with_odds=False))}.\n'
        f'P_history {_v11_pct(row.get("p_hist"))}; P_scenario {_v11_pct(row.get("p_scenario_core"))}; '
        f'P_final {_v11_pct(row.get("p_final"))}; проєкція {_v11_num(row.get("projection_used"))}; '
        f'edge {_v11_num(row.get("line_edge"))}.'
    )


def _v11_build_messages(advisor: dict[str, Any], calculation: dict[str, Any]) -> list[str]:
    snapshot = calculation.get('canonical_snapshot') or {}
    score = snapshot.get('score') or {}
    quarters = snapshot.get('quarters') or []
    qtext = ' | '.join(
        f"Q{i + 1} {q.get('home')}:{q.get('away')}"
        for i, q in enumerate(quarters)
        if q.get('home') is not None and q.get('away') is not None
        and not (float(q.get('home') or 0) == 0.0 and float(q.get('away') or 0) == 0.0 and i + 1 > _v11_completed_quarters(snapshot))
    )
    action = str(advisor.get('action') or 'PASS')
    icon = '✅' if action == 'PLAY' else '⚠️' if action == 'RISK' else '❌'
    primary = list(advisor.get('primary_lines') or [])
    selected = next((row for row in primary if str((row.get('advisor') or {}).get('action') or '') == action), None)
    selected = selected or (primary[0] if primary else None)

    if action == 'PASS':
        top_line = '<b>❌ PASS — ЗАРАЗ РЕАЛЬНУ СТАВКУ НЕ БРАТИ</b>'
    elif selected:
        top_line = f'<b>{icon} {html.escape(action)}</b> — {html.escape(_v113_bet_sentence(selected))}'
    else:
        top_line = f'<b>{icon} {html.escape(action)}</b> — реального сигналу немає'
    header = '\n'.join([
        top_line,
        f'<b>Матч:</b> {html.escape(str(snapshot.get("name") or "N/A"))}',
        f'<b>Стадія:</b> {html.escape(str(snapshot.get("stage") or "N/A"))} | '
        f'{_v118_clock_context(snapshot)} | '
        f'<b>Рахунок:</b> {_v11_num(score.get("home"))}:{_v11_num(score.get("away"))}',
        f'<b>Чверті:</b> {html.escape(qtext or "N/A")}',
    ])

    blocks: list[str] = []
    if selected:
        blocks.append(_v116_compact_main_block(selected))
        # For PASS the theoretical trigger is already the second useful calculation.
        # Do not add another losing real-line block and overload Telegram.
        alternatives = [] if action == 'PASS' else [
            row for row in primary
            if row is not selected and str((row.get('advisor') or {}).get('action') or '') == action
        ][:1]
        if alternatives:
            blocks.append('<b>ЩЕ ОДИН ВАРІАНТ</b>\n' + '\n\n'.join(
                _v116_compact_alternative(row, idx) for idx, row in enumerate(alternatives, 2)
            ))
    else:
        trigger_blocks = []
        for summary in advisor.get('model_summary') or []:
            text = _v116_compact_trigger(summary, len(trigger_blocks) + 1)
            if text:
                trigger_blocks.append(text)
            if len(trigger_blocks) >= 2:
                break
        if trigger_blocks:
            blocks.append('<b>НАЙБЛИЖЧІ ТЕОРЕТИЧНІ УМОВИ</b>\n' + '\n\n'.join(trigger_blocks))
        else:
            blocks.append('Немає реальної лінії та немає коректного теоретичного тригера.')

    if action == 'PASS':
        theoretical = advisor.get('nearest_theoretical_play') or {}
        trigger = theoretical.get('play_trigger') or theoretical.get('risk_trigger')
        if trigger and _v114_compact_trigger_is_valid(trigger):
            blocks.append('<b>ЩО МОЖЕ СТАТИ СТАВКОЮ</b>\n' + html.escape(
                _v113_trigger_sentence(trigger, play=bool(theoretical.get('play_trigger')))
            ))

    text = header + '\n\n' + '\n\n'.join(blocks)
    # Keep the message compact. Split only at section boundaries when needed.
    if len(text) <= 3900:
        return [text]
    messages: list[str] = []
    current = header
    for block in blocks:
        candidate = current + '\n\n' + block
        if len(candidate) > 3900 and current != header:
            messages.append(current)
            current = '<b>Продовження</b>\n' + block
        else:
            current = candidate
    if current:
        messages.append(current)
    return messages


DEFAULT_CONFIG.setdefault('compact_telegram_policy', {})
DEFAULT_CONFIG['compact_telegram_policy'].update({
    'version': ADVISOR_VERSION,
    'final_recommendation_first': True,
    'one_full_primary_block': True,
    'max_compact_alternatives': 1,
    'full_audit_kept_in_json_not_telegram': True,
    'market_scope_consistency_guard': True,
    'stat_projection_source_disclosed': True,
    'ambiguous_low_quarter_total_blocked': True,
    'full_stat_quarter_projection_uses_cumulative_rate': True,
})
SYSTEM_VERSION = ADVISOR_VERSION
DEFAULT_CONFIG['engine_version'] = ADVISOR_VERSION


# =============================================================================
# v11.8 CLOCK SEMANTICS FIX + LABELED PROJECTION METRICS
# =============================================================================
# Display-only update. Mathematical calculations, market selection, Scenario
# Miner, projections, P_history/P_scenario/P_live/P_final and verdict logic are
# unchanged. Every recommendation block now explicitly identifies whose/which
# segment projection is shown and always prints history zone, line and edge.

ADVISOR_VERSION = '11.8.0-CLOCK-SEMANTICS-FIX'


def _v117_period_label(item: dict[str, Any]) -> str:
    market_type = str(item.get('market_type') or '')
    segment = str(item.get('segment') or '')
    if market_type in {'MATCH_TOTAL', 'TEAM_IT_MATCH'} or segment == 'MATCH':
        return 'за весь матч'
    if market_type in {'H1_TOTAL', 'TEAM_IT_H1'} or segment == 'H1':
        return 'у 1-й половині'
    if market_type in {'H2_TOTAL', 'TEAM_IT_H2'} or segment == 'H2':
        return 'у 2-й половині'
    if segment.startswith('Q'):
        return f'у {segment}'
    return f'у сегменті {segment}' if segment else 'у цьому ринку'


def _v117_projection_label(item: dict[str, Any]) -> str:
    market_type = str(item.get('market_type') or '')
    team = str(item.get('team') or '').strip()
    segment = str(item.get('segment') or '')
    if market_type in {'TEAM_IT_MATCH', 'TEAM_IT_H1', 'TEAM_IT_H2', 'CURRENT_QUARTER_TEAM_IT'} or team:
        return f'Live-проєкція команди {team or "N/A"} {_v117_period_label(item)}'
    if market_type == 'MATCH_TOTAL' or segment == 'MATCH':
        return 'Live-проєкція загального тоталу матчу'
    if market_type == 'H1_TOTAL' or segment == 'H1':
        return 'Live-проєкція загального тоталу 1-ї половини'
    if market_type == 'H2_TOTAL' or segment == 'H2':
        return 'Live-проєкція загального тоталу 2-ї половини'
    if market_type == 'CURRENT_QUARTER_TOTAL' or segment.startswith('Q'):
        return f'Live-проєкція загального тоталу {segment or "поточної чверті"}'
    return f'Live-проєкція ринку {_v11_market_label(item)}'


def _v117_history_zone_line(source: dict[str, Any]) -> str:
    zone = source.get('history_zone_rate')
    return f'<b>Історична зона:</b> {_v11_pct(zone)}'


def _v117_labeled_projection_line(source: dict[str, Any]) -> str:
    label = _v117_projection_label(source)
    return f'<b>{html.escape(label)}:</b> {_v11_num(source.get("projection_used"))}'


def _v116_compact_main_block(item: dict[str, Any]) -> str:
    adv = item.get('advisor') or {}
    action = str(adv.get('action') or 'PASS')
    bet = _v113_bet_sentence(item)
    hist_fact = _v116_history_fact(item)
    zone = adv.get('history_zone_rate')
    projection = adv.get('projection_used')
    line = item.get('line')
    side = str(item.get('side') or '').upper()
    scenario = _v116_top_scenario_text(item)
    source_text = _v116_projection_source_text(item)

    lines = [
        f'<b>🎯 Основний розрахунок:</b> {html.escape(bet)}.',
        f'<b>Історична зона:</b> {_v11_pct(zone)} ({html.escape(hist_fact)}); '
        f'P_history {_v11_pct(adv.get("p_hist"))}.',
        f'<b>Ймовірності:</b> P_scenario {_v11_pct(adv.get("p_scenario_core"))} | '
        f'P_live {_v11_pct(adv.get("p_live"))} | P_final {_v11_pct(adv.get("p_final"))}.',
        f'<b>Лінія:</b> {html.escape(side)} {_v11_num(line)}.',
        _v117_labeled_projection_line({**item, **adv, 'projection_used': projection}),
        f'<b>Різниця та edge:</b> {html.escape(_v115_projection_difference_text(projection, line))}; '
        f'edge для {html.escape(side)} {_v11_num(adv.get("line_edge"))}.',
        f'<b>Сценарій:</b> {html.escape(scenario)}',
    ]

    mode = str(item.get('data_mode') or (item.get('live') or {}).get('data_mode') or 'DATA_OFF')
    stat_bits = [mode]
    if adv.get('fake_over'):
        stat_bits.append('FAKE OVER')
    if adv.get('fake_under'):
        stat_bits.append('FAKE UNDER')
    gate = str(adv.get('stat_gate_status') or 'OFF')
    if gate not in {'OFF', 'N/A', 'NA', 'N_A_NO_STATS'}:
        stat_bits.append(f'gate {gate}')
    lines.append(f'<b>Статистика:</b> {html.escape("; ".join(stat_bits))}. {html.escape(source_text)}')

    blockers = adv.get('serious_blockers') or []
    if action == 'PLAY':
        lines.append('<b>Підсумок:</b> брати як PLAY, якщо лінія й коефіцієнт ще актуальні.')
    elif action == 'RISK':
        reason = ', '.join(blockers[:2]) if blockers else 'підтвердження неповне'
        lines.append(f'<b>Підсумок:</b> RISK PLAY, не чистий PLAY. Ризик: {html.escape(reason)}.')
    else:
        reason = ', '.join(blockers[:2]) if blockers else 'чистого підтвердження немає'
        lines.append(f'<b>Підсумок:</b> цю лінію не брати. Причина: {html.escape(reason)}.')

    correction = item.get('market_scope_correction') or {}
    if correction:
        lines.append('<b>Уточнення ринку:</b> це IT 1-ї половини, а не IT матчу; scope виправлено за загальною H1-лінією.')
    return '\n'.join(lines)


def _v116_compact_alternative(item: dict[str, Any], index: int) -> str:
    adv = item.get('advisor') or {}
    action = str(adv.get('action') or 'PASS')
    side = str(item.get('side') or '').upper()
    source = {**item, **adv, 'projection_used': adv.get('projection_used')}
    return '\n'.join([
        f'<b>{index}) {html.escape(action)}</b> — {html.escape(_v113_bet_sentence(item))}.',
        f'<b>Історична зона:</b> {_v11_pct(adv.get("history_zone_rate"))}; '
        f'P_history {_v11_pct(adv.get("p_hist"))}; P_scenario {_v11_pct(adv.get("p_scenario_core"))}; '
        f'P_final {_v11_pct(adv.get("p_final"))}.',
        f'<b>Лінія:</b> {html.escape(side)} {_v11_num(item.get("line"))}.',
        _v117_labeled_projection_line(source),
        f'<b>Edge для {html.escape(side)}:</b> {_v11_num(adv.get("line_edge"))} '
        f'({_v115_projection_difference_text(adv.get("projection_used"), item.get("line"))}).',
    ])


def _v116_compact_trigger(summary: dict[str, Any], index: int) -> Optional[str]:
    row, _ = _v11_best_model_trigger(summary)
    if not row or not _v114_compact_trigger_is_valid(row):
        return None
    is_play = float(row.get('p_final') or 0.0) >= ADVISOR_PLAY_MIN
    item = {
        'market_type': row.get('market_type'),
        'team': row.get('team'),
        'segment': row.get('segment'),
        'side': row.get('side'),
        'line': row.get('line'),
        'odds': row.get('odds'),
        'projection_used': row.get('projection_used'),
        'advisor': {'is_model_line': True},
    }
    side = str(row.get('side') or '').upper()
    return '\n'.join([
        f'<b>{index}) Теоретичний {"PLAY" if is_play else "RISK"}</b> — '
        f'{html.escape(_v113_bet_sentence(item, with_odds=False))}.',
        f'<b>Історична зона:</b> {_v11_pct(row.get("history_zone_rate"))}; '
        f'P_history {_v11_pct(row.get("p_hist"))}; P_scenario {_v11_pct(row.get("p_scenario_core"))}; '
        f'P_final {_v11_pct(row.get("p_final"))}.',
        f'<b>Лінія:</b> {html.escape(side)} {_v11_num(row.get("line"))}.',
        _v117_labeled_projection_line(item),
        f'<b>Edge для {html.escape(side)}:</b> {_v11_num(row.get("line_edge"))} '
        f'({_v115_projection_difference_text(row.get("projection_used"), row.get("line"))}).',
    ])


def _v113_trigger_sentence(row: dict[str, Any], *, play: bool) -> str:
    if not _v114_compact_trigger_is_valid(row):
        return 'Теоретичний тригер скасовано: лінія вже перетнута рахунком або проєкція некоректна.'
    item = {
        'market_type': row.get('market_type'),
        'team': row.get('team'),
        'segment': row.get('segment'),
        'side': row.get('side'),
        'line': row.get('line'),
        'odds': row.get('odds'),
        'projection_used': row.get('projection_used'),
        'advisor': {'is_model_line': True},
    }
    side = str(row.get('side') or '').upper()
    threshold = 'або нижче' if side == 'OVER' else 'або вище'
    kind = 'PLAY' if play else 'RISK'
    current = to_number(row.get('current_points'))
    current_text = f' Зараз у цьому ринку вже набрано {current:.1f} очка.' if current is not None else ''
    return (
        f'{kind} може з’явитися, якщо букмекер дасть лінію, за якої '
        f'{_v113_bet_sentence(item, with_odds=False)} ({threshold}).'
        f'{current_text} Історична зона — {_v11_pct(row.get("history_zone_rate"))}; '
        f'P_history — {_v11_pct(row.get("p_hist"))}; P_scenario — {_v11_pct(row.get("p_scenario_core"))}; '
        f'{_v117_projection_label(item)} — {_v11_num(row.get("projection_used"))}; '
        f'лінія — {side} {_v11_num(row.get("line"))}; '
        f'{_v115_projection_difference_text(row.get("projection_used"), row.get("line"))}; '
        f'edge для {side} — {_v11_num(row.get("line_edge"))}; P_final — {_v11_pct(row.get("p_final"))}.'
    )


DEFAULT_CONFIG.setdefault('telegram_recommendation_metrics_policy', {})
DEFAULT_CONFIG['telegram_recommendation_metrics_policy'].update({
    'version': ADVISOR_VERSION,
    'math_changed': False,
    'every_recommendation_has_history_zone': True,
    'every_recommendation_has_named_projection': True,
    'every_recommendation_has_line': True,
    'every_recommendation_has_edge': True,
    'team_it_projection_names_team': True,
    'quarter_projection_names_quarter': True,
})
SYSTEM_VERSION = ADVISOR_VERSION
DEFAULT_CONFIG['engine_version'] = ADVISOR_VERSION

# =============================================================================
# v11.8 CLOCK DISPLAY AND SEMANTICS
# =============================================================================
def _v118_mmss(seconds: Any) -> str:
    value = to_number(seconds)
    if value is None:
        return 'N/A'
    total = max(0, int(round(value)))
    return f'{total // 60:02d}:{total % 60:02d}'


def _v118_clock_context(snapshot: dict[str, Any]) -> str:
    """Human-readable clock: explicitly distinguish played vs remaining time."""
    stage = str(snapshot.get('stage') or '').upper()
    explicit = str(snapshot.get('explicit_stage') or '').upper()
    if stage == 'PRE_MATCH':
        return '<b>Час:</b> матч ще не почався'
    if stage == 'HT':
        return '<b>Час:</b> перерва після Q2'
    if any(token in explicit for token in ('FINISHED', 'FINAL', 'ENDED', 'ЗАВЕРШ', 'КІНЕЦЬ')):
        return '<b>Час:</b> матч завершено'

    quarter = to_int(snapshot.get('current_quarter'))
    q_seconds = to_number(snapshot.get('quarter_seconds'))
    remaining = to_number(snapshot.get('quarter_seconds_remaining'))
    if q_seconds is not None and remaining is not None:
        remaining = max(0.0, min(q_seconds, remaining))
        played = max(0.0, q_seconds - remaining)
        q_label = f'Q{quarter}' if quarter else 'поточна чверть'
        return (
            f'<b>{html.escape(q_label)} зіграно:</b> {_v118_mmss(played)} | '
            f'<b>залишилось:</b> {_v118_mmss(remaining)}'
        )

    clock = snapshot.get('clock')
    if clock:
        return f'<b>Залишилось у чверті:</b> {html.escape(str(clock))}'
    return '<b>Час:</b> N/A'


DEFAULT_CONFIG.setdefault('clock_semantics_policy', {})
DEFAULT_CONFIG['clock_semantics_policy'].update({
    'version': ADVISOR_VERSION,
    'parser_clock_is_remaining': True,
    'derive_played_from_remaining': True,
    'telegram_shows_played_and_remaining': True,
    'projection_math_changed_only_when_elapsed_was_missing': True,
})
SYSTEM_VERSION = ADVISOR_VERSION
DEFAULT_CONFIG['engine_version'] = ADVISOR_VERSION


# =============================================================================
# v12.1 ALWAYS-RANKED ADVISOR / REFERENCE LINES WHEN BK HAS NO TOTALS
# =============================================================================
# The mathematical core is unchanged. This layer changes only the final advisor:
# - every valid live snapshot receives the strongest 1-2 real bookmaker options;
# - PASS is never sent to Telegram;
# - weak-but-valid options are explicitly labelled MICRO (3-5% of bankroll);
# - better options are WORKING (10-15%); clean strong options are STRONG (25-33%);
# - truly invalid/stale/settled markets still produce no Telegram message.

ADVISOR_VERSION = '12.1.0-REFERENCE-LINES-NO-BK'
V12_MAX_PRIMARY = max(1, min(2, int(os.getenv('SUPER_BASKET_V12_MAX_PRIMARY', '2'))))
V12_SECOND_MAX_SCORE_GAP = float(os.getenv('SUPER_BASKET_V12_SECOND_MAX_SCORE_GAP', '0.14'))
V12_MICRO_SCORE_MAX = float(os.getenv('SUPER_BASKET_V12_MICRO_SCORE_MAX', '0.60'))
V12_STRONG_SCORE_MIN = float(os.getenv('SUPER_BASKET_V12_STRONG_SCORE_MIN', '0.78'))
V12_STRONG_P_FINAL_MIN = float(os.getenv('SUPER_BASKET_V12_STRONG_P_FINAL_MIN', '0.75'))

_V12_FATAL_PARSER_ISSUES = {
    'NO_LINE', 'NO_ODDS', 'SYNTHETIC_LINE', 'UNSUPPORTED_MARKET',
    'ODDS_BELOW_MINIMUM', 'ODDS_ABOVE_MAXIMUM',
    'UNKNOWN_QUARTER', 'INVALID_QUARTER', 'PAST_QUARTER', 'FUTURE_QUARTER',
    'NO_CURRENT_QUARTER', 'NO_EXACT_CURRENT_QUARTER_TIME',
    'NO_CURRENT_QUARTER_SCORE',
}

_V12_FATAL_BLOCKERS = {
    'SCHEMA_ERROR',
    'PRODUCTION_ROUTER_BLOCK',
    'LINE_BELOW_CURRENT_SCOPE_SCORE',
    'SOURCE_SCOPE_SEGMENT_MISMATCH',
    'Q4_MISSING_MANDATORY_CONTEXT',
    'TEAM_IT_REQUIRED_LIVE_UNREALISTIC',
}

_V12_MARKET_SEMANTIC_BASE = _market_semantic_issues


def _market_semantic_issues(market: dict[str, Any], canonical: dict[str, Any]) -> list[dict[str, Any]]:
    """Accept normal bookmaker full-match scopes such as MATCH(OT).

    v11.8 treated MATCH(OT) as a mismatch even though the normalized segment is
    correctly MATCH. Preserve every other semantic guard.
    """
    issues = _V12_MARKET_SEMANTIC_BASE(market, canonical)
    source_scope = str(market.get('source_scope') or '').upper().replace(' ', '')
    segment = str(market.get('segment') or '').upper()
    match_scope_ok = bool(
        segment == 'MATCH'
        and (
            source_scope.startswith('MATCH')
            or source_scope in {'FT', 'FULLMATCH', 'FULLMATCH(OT)', 'REGULATION+OT'}
        )
    )
    if match_scope_ok:
        issues = [row for row in issues if row.get('rule_id') != 'SOURCE_SCOPE_SEGMENT_MISMATCH']
    return issues


_V12_MAJOR_PENALTY_CODES = {
    'TEAM_IT_WEAKEST_BLOCK',
    'TEAM_IT_WEAKEST_BELOW_70',
    'FAKE_OVER', 'FAKE_UNDER',
    'Q4_UNDER_NO_DRY',
    'Q4_UNDER_MEDIUM_DRY_NO_STRONG_EDGE',
    'Q4_OVER_CONFIRMATION_FAILED',
    'Q4_UNDER_DANGER',
    'Q4_MATCH_UNDER_OT_TAIL',
    'Q4_LOW_FOUL_CONVERSION',
    'STAT_GATE_DIRECTLY_AGAINST',
    'STRONG_HISTORY_LIVE_CONFLICT',
    'AUDIT_UNDER_LIVE_CONFLICT_GT8',
    'AUDIT_UNDER_90_LIVE_CONFLICT',
    'AUDIT_FAKE_UNDER_BLOCK',
    'NO_SAME_FORMAT_HISTORY',
}


def _v12_rule_codes(rows: Any) -> list[str]:
    output: list[str] = []
    for row in rows or []:
        if isinstance(row, dict):
            code = str(row.get('rule_id') or '').strip()
        else:
            code = str(row or '').strip()
        if code:
            output.append(code)
    return list(dict.fromkeys(output))


def _v12_fatal_reasons(item: dict[str, Any]) -> list[str]:
    reasons = [str(code) for code in item.get('parser_issues') or [] if str(code) in _V12_FATAL_PARSER_ISSUES]
    router = item.get('router') or {}
    if str(router.get('status') or '').upper() == 'BLOCK' and bool(router.get('hard_block', True)):
        reasons.append('PRODUCTION_ROUTER_BLOCK')
    reasons.extend(code for code in _v12_rule_codes(item.get('blockers')) if code in _V12_FATAL_BLOCKERS)
    q4 = item.get('q4_context') or {}
    if q4.get('applicable') and q4.get('mandatory_missing'):
        reasons.append('Q4_MISSING_MANDATORY_CONTEXT')
    return list(dict.fromkeys(reasons))


def _v12_probability(item: dict[str, Any], key: str, default: float = 0.50) -> float:
    adv = item.get('advisor') or {}
    history = item.get('history') or {}
    scenario = item.get('scenario') or {}
    live = item.get('live') or {}
    mapping = {
        'p_final': adv.get('p_final', item.get('p_final_system', item.get('p_final'))),
        'p_live': adv.get('p_live', live.get('p_live')),
        'p_hist': adv.get('p_hist', history.get('p_hist')),
        'zone': adv.get('history_zone_rate', history.get('history_zone_rate')),
        'p_scenario': max(
            float(adv.get('p_scenario_core') or scenario.get('p_scenario') or 0.50),
            float(adv.get('p_scenario_miner') or 0.50),
        ),
    }
    value = to_number(mapping.get(key))
    return default if value is None else max(0.0, min(1.0, float(value)))


def _v12_edge_quality(edge: float) -> float:
    # 0.50 at the line, approaches 1.00 for a large positive edge and 0.00 for
    # a large negative edge. This is a ranking feature, not a new P_final.
    return max(0.0, min(1.0, 0.5 + 0.5 * math.tanh(edge / 7.0)))


def _v12_recommendation_score(item: dict[str, Any]) -> dict[str, Any]:
    p_final = _v12_probability(item, 'p_final')
    p_live = _v12_probability(item, 'p_live')
    p_hist = _v12_probability(item, 'p_hist')
    zone = _v12_probability(item, 'zone', p_hist)
    p_scenario = _v12_probability(item, 'p_scenario')
    adv = item.get('advisor') or {}
    live = item.get('live') or {}
    stat = item.get('stat_comparison') or {}
    edge = to_number(adv.get('line_edge'))
    if edge is None:
        edge = to_number(live.get('line_edge')) or 0.0
    edge = float(edge)

    stat_status = str(adv.get('stat_gate_status') or stat.get('stat_gate_status') or 'OFF').upper()
    stat_support = str(stat.get('stat_support') or item.get('data_mode') or 'OFF').upper()
    if stat_status == 'CONFIRMED':
        stat_quality = 1.0
    elif stat_status == 'LIMITED':
        stat_quality = 0.62
    elif stat_status in {'OFF', 'N/A', 'NA', 'N_A_NO_STATS'}:
        stat_quality = 0.50
    else:
        stat_quality = 0.25

    score = (
        0.42 * p_final
        + 0.22 * p_live
        + 0.14 * zone
        + 0.10 * p_scenario
        + 0.07 * stat_quality
        + 0.05 * _v12_edge_quality(edge)
    )

    blocker_codes = _v12_rule_codes(item.get('blockers'))
    cap_codes = _v12_rule_codes(item.get('caps'))
    all_codes = list(dict.fromkeys(blocker_codes + cap_codes + list((adv.get('serious_blockers') or []))))
    major = [code for code in all_codes if code in _V12_MAJOR_PENALTY_CODES]
    soft = [code for code in all_codes if code not in _V12_MAJOR_PENALTY_CODES and code not in _V12_FATAL_BLOCKERS]
    score -= min(0.18, 0.055 * len(major))
    score -= min(0.10, 0.018 * len(soft))
    if edge < 0:
        score -= min(0.12, abs(edge) / 40.0)
    if stat_support == 'OFF':
        score -= 0.04
    elif stat_support == 'LIMITED':
        score -= 0.02
    odds = to_number(item.get('odds'))
    if odds is not None:
        score += min(0.012, max(0.0, (float(odds) - 1.44) * 0.01))
    score = max(0.01, min(0.99, score))

    fake_against = bool(
        (str(item.get('side') or '').upper() == 'OVER' and stat.get('fake_over'))
        or (str(item.get('side') or '').upper() == 'UNDER' and stat.get('fake_under'))
    )
    clean = bool(
        edge >= 0
        and not major
        and not blocker_codes
        and not fake_against
        and stat_status != 'AGAINST'
    )
    return {
        'score': score,
        'p_final': p_final,
        'p_live': p_live,
        'p_hist': p_hist,
        'history_zone_rate': zone,
        'p_scenario': p_scenario,
        'line_edge': edge,
        'stat_status': stat_status,
        'stat_support': stat_support,
        'clean': clean,
        'major_penalties': major,
        'soft_penalties': soft,
    }


def _v12_tier(metrics: dict[str, Any]) -> dict[str, str]:
    score = float(metrics['score'])
    p_final = float(metrics['p_final'])
    clean = bool(metrics['clean'])
    stat_support = str(metrics.get('stat_support') or 'OFF')

    if score >= V12_STRONG_SCORE_MIN and p_final >= V12_STRONG_P_FINAL_MIN and clean and stat_support != 'OFF':
        if score >= 0.88 and p_final >= 0.85:
            stake = '30-33% від бюджету'
        elif score >= 0.83:
            stake = '27-30% від бюджету'
        else:
            stake = '25-27% від бюджету'
        return {'tier': 'STRONG', 'action': 'PLAY', 'status': 'STRONG — НАЙСИЛЬНІШИЙ ВХІД', 'stake': stake}

    if score >= V12_MICRO_SCORE_MAX or p_final >= 0.60:
        if score >= 0.72 or p_final >= 0.72:
            stake = '12-15% від бюджету'
        else:
            stake = '10-12% від бюджету'
        return {'tier': 'WORKING', 'action': 'RISK', 'status': 'WORKING — РОБОЧИЙ ВХІД', 'stake': stake}

    return {'tier': 'MICRO', 'action': 'RISK', 'status': 'MICRO — СЛАБКИЙ, АЛЕ НАЙКРАЩИЙ ВАРІАНТ', 'stake': '3-5% від бюджету'}


def _v12_market_group(item: dict[str, Any]) -> tuple[Any, ...]:
    return (item.get('market_type'), item.get('team'), item.get('segment'))


def _v12_rank_recommendations(evaluations: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    ranked: list[dict[str, Any]] = []
    rejected: list[dict[str, Any]] = []
    odds_min = float(DEFAULT_CONFIG.get('odds_min', 1.44))
    for original in evaluations:
        item = deepcopy(original)
        fatal = _v12_fatal_reasons(item)
        odds = to_number(item.get('odds'))
        if odds is None or odds < odds_min:
            fatal.append('ODDS_BELOW_MINIMUM')
        if not _v112_market_is_currently_supported(item):
            fatal.append('MARKET_NOT_CURRENTLY_SUPPORTED')
        fatal = list(dict.fromkeys(fatal))
        if fatal:
            rejected.append({'market': _v11_compact_line(item), 'fatal_reasons': fatal})
            continue
        metrics = _v12_recommendation_score(item)
        tier = _v12_tier(metrics)
        if bool(item.get('is_reference_line')):
            # A synthetic threshold can be WORKING/STRONG only when its evidence is
            # clean. Any direct stat/history conflict is always MICRO 3-5%.
            if (
                metrics.get('major_penalties')
                or str(metrics.get('stat_status') or '').upper() == 'AGAINST'
                or float(metrics.get('history_zone_rate') or 0.0) < 0.45
            ):
                tier = {
                    'tier': 'MICRO',
                    'action': 'RISK',
                    'status': 'MICRO — MODEL LINE З КОНФЛІКТОМ',
                    'stake': '3-5% від бюджету',
                }
            elif tier.get('tier') == 'STRONG' and (
                float(metrics.get('history_zone_rate') or 0.0) < 0.75
                or str(metrics.get('stat_status') or '').upper() != 'CONFIRMED'
            ):
                tier = {
                    'tier': 'WORKING',
                    'action': 'RISK',
                    'status': 'WORKING — MODEL LINE',
                    'stake': '12-15% від бюджету',
                }
        adv = item.setdefault('advisor', {})
        adv['original_action_before_v12'] = adv.get('action')
        adv['original_status_before_v12'] = adv.get('status')
        adv['action'] = tier['action']
        adv['status'] = tier['status']
        adv['recommendation_tier'] = tier['tier']
        adv['recommendation_score'] = metrics['score']
        adv['stake_budget'] = tier['stake']
        adv['always_ranked_policy'] = True
        adv['fatal_reasons'] = []
        adv['ranking_metrics'] = metrics
        item['system_action'] = tier['action']
        item['system_status'] = tier['status']
        item['stake'] = tier['stake']
        item['p_final_system'] = metrics['p_final']
        item['v12_rank_score'] = metrics['score']
        ranked.append(item)

    # First choose one best offer/side per logical market. This prevents two
    # neighbouring lines of the same total from occupying both Telegram slots.
    best_by_group: dict[tuple[Any, ...], dict[str, Any]] = {}
    for item in ranked:
        key = _v12_market_group(item)
        current = best_by_group.get(key)
        candidate_edge = float(((item.get('advisor') or {}).get('ranking_metrics') or {}).get('line_edge') or 0.0)
        current_edge = float((((current or {}).get('advisor') or {}).get('ranking_metrics') or {}).get('line_edge') or 0.0)
        candidate_key = (
            1.0 if candidate_edge >= 0 else 0.0,
            float(item.get('v12_rank_score') or 0.0),
            float((item.get('advisor') or {}).get('p_final') or 0.0),
            float(item.get('odds') or 0.0),
        )
        current_key = (
            1.0 if current_edge >= 0 else 0.0,
            float(current.get('v12_rank_score') or 0.0),
            float((current.get('advisor') or {}).get('p_final') or 0.0),
            float(current.get('odds') or 0.0),
        ) if current else (-1.0, -1.0, -1.0, -1.0)
        if current is None or candidate_key > current_key:
            best_by_group[key] = item

    groups = sorted(
        best_by_group.values(),
        key=lambda item: (
            1.0 if float((((item.get('advisor') or {}).get('ranking_metrics') or {}).get('line_edge')) or 0.0) >= 0 else 0.0,
            {'STRONG': 3, 'WORKING': 2, 'MICRO': 1}.get(str((item.get('advisor') or {}).get('recommendation_tier')), 0),
            float(item.get('v12_rank_score') or 0.0),
            float((item.get('advisor') or {}).get('p_final') or 0.0),
            float(item.get('odds') or 0.0),
        ),
        reverse=True,
    )
    if not groups:
        return [], rejected
    selected = [groups[0]]
    top_score = float(groups[0].get('v12_rank_score') or 0.0)
    for candidate in groups[1:]:
        if len(selected) >= V12_MAX_PRIMARY:
            break
        score = float(candidate.get('v12_rank_score') or 0.0)
        if score < top_score - V12_SECOND_MAX_SCORE_GAP:
            continue
        selected.append(candidate)
    return selected, rejected


def _v12_budget_recommendation(action: str, status: str, stake: str) -> dict[str, Any]:
    match = re.search(r'(\d+(?:\.\d+)?)\s*[-–]\s*(\d+(?:\.\d+)?)%', stake or '')
    pct_min = float(match.group(1)) if match else 0.0
    pct_max = float(match.group(2)) if match else 0.0
    bankroll = to_number(os.getenv('SUPER_BASKET_BANKROLL'))
    fallback_limit = to_number(os.getenv('SUPER_BASKET_LIVE_LIMIT'))
    base_amount = bankroll if bankroll is not None else fallback_limit
    base_type = 'BANKROLL' if bankroll is not None else 'LIVE_LIMIT_FALLBACK' if fallback_limit is not None else 'PERCENT_ONLY'
    currency = os.getenv('SUPER_BASKET_CURRENCY', 'USDT')
    return {
        'action': action,
        'status': status,
        'base_type': base_type,
        'base_amount': base_amount,
        'currency': currency,
        'percent_min': pct_min,
        'percent_max': pct_max,
        'amount_min': round(base_amount * pct_min / 100.0, 2) if base_amount is not None else None,
        'amount_max': round(base_amount * pct_max / 100.0, 2) if base_amount is not None else None,
        'text': stake,
        'educational_note': 'Розмір входу задається від окремого betting-bankroll і не гарантує результат.',
    }


def _v12_reason_text(item: dict[str, Any]) -> str:
    adv = item.get('advisor') or {}
    metrics = adv.get('ranking_metrics') or {}
    side = str(item.get('side') or '').upper()
    projection = to_number(adv.get('projection_used'))
    if projection is None:
        projection = to_number((item.get('live') or {}).get('projection_used'))
    line = to_number(item.get('line'))
    edge = to_number(metrics.get('line_edge')) or 0.0
    if projection is not None and line is not None:
        aligned = (side == 'OVER' and projection > line) or (side == 'UNDER' and projection < line)
        relation = 'вище' if projection > line else 'нижче' if projection < line else 'на рівні'
        direction_text = 'підтримує сторону' if aligned else 'суперечить стороні'
        return (
            f'Проєкція {projection:.1f} {relation} лінії {line:.1f} і {direction_text}; '
            f'edge {edge:+.1f}, P_final {_v11_pct(metrics.get("p_final"))}, '
            f'історична зона {_v11_pct(metrics.get("history_zone_rate"))}.'
        )
    return f'Найвищий комплексний рейтинг серед усіх актуальних ліній БК: {float(metrics.get("score") or 0.0):.1%}.'


def _v12_risk_text(item: dict[str, Any]) -> str:
    adv = item.get('advisor') or {}
    metrics = adv.get('ranking_metrics') or {}
    penalties = list(metrics.get('major_penalties') or []) + list(metrics.get('soft_penalties') or [])
    labels = {
        'FULL_STAT_CONFIRMATION_NOT_ON': 'немає повного підтвердження live-статою',
        'Q4_OVER_CONFIRMATION_FAILED': 'Q4 Over не пройшов повний foul/volume/efficiency gate',
        'Q4_UNDER_NO_DRY': 'для Q4 Under недостатньо dry-підтвердження',
        'Q4_UNDER_MEDIUM_DRY_NO_STRONG_EDGE': 'dry середній, а edge недостатньо сильний',
        'STAT_GATE_DIRECTLY_AGAINST': 'live-статистика суперечить цій стороні',
        'STRONG_HISTORY_LIVE_CONFLICT': 'історія та live-проєкція конфліктують',
        'LIVE_DIRECTION_OR_EDGE_FAILED': 'live-напрямок або edge слабкий',
        'HISTORY_CONFIRMATION_BELOW_DYNAMIC_MIN': 'історичне підтвердження нижче робочого рівня',
        'SCENARIO_CONFIRMATION_BELOW_DYNAMIC_MIN': 'схожі сценарії недостатньо підтримують сторону',
        'TEAM_IT_WEAKEST_BLOCK': 'один із Team IT gate: own scored / opponent allowed — слабкий',
        'FAKE_OVER': 'ризик fake over',
        'FAKE_UNDER': 'ризик fake under і відскоку результативності',
        'V11_MODEL_CAP': 'модельна лінія має data/sample cap',
        'STAT_GATE_AGAINST': 'live-статистика не підтверджує модельний напрямок',
    }
    if penalties:
        return '; '.join(labels.get(str(code), str(code).replace('_', ' ').lower()) for code in penalties[:3])
    if str(metrics.get('stat_support') or '').upper() == 'OFF':
        return 'немає повної live-статистики; оцінка більше спирається на рахунок, час, історію та сценарій'
    return 'зміна темпу, лінії або статистичного профілю після snapshot'


def _v12_build_messages(advisor: dict[str, Any], calculation: dict[str, Any]) -> list[str]:
    snapshot = calculation.get('canonical_snapshot') or {}
    score = snapshot.get('score') or {}
    rows = list(advisor.get('primary_lines') or [])
    if not rows:
        return []
    header = '\n'.join([
        '<b>🔥 SUPER BASKET — НАЙСИЛЬНІШІ ВАРІАНТИ</b>',
        f'<b>Матч:</b> {html.escape(str(snapshot.get("name") or "N/A"))}',
        f'<b>Стадія:</b> {html.escape(str(snapshot.get("stage") or "N/A"))} | '
        f'{_v118_clock_context(snapshot)} | '
        f'<b>Рахунок:</b> {_v11_num(score.get("home"))}:{_v11_num(score.get("away"))}',
    ])
    blocks: list[str] = []
    for index, item in enumerate(rows, 1):
        adv = item.get('advisor') or {}
        metrics = adv.get('ranking_metrics') or {}
        budget = _v12_budget_recommendation(adv.get('action', 'RISK'), adv.get('status', ''), adv.get('stake_budget', '3-5% від бюджету'))
        amount = ''
        if budget.get('amount_min') is not None:
            amount = f' ({budget["amount_min"]:.2f}–{budget["amount_max"]:.2f} {html.escape(str(budget["currency"]))})'
        reference = bool(item.get('is_reference_line') or adv.get('is_reference_line'))
        ref_details = item.get('reference_line_details') or adv.get('reference_line_details') or {}
        type_line = (
            '<b>Тип:</b> МОДЕЛЬНА REFERENCE LINE — БК не дав total/IT; це поріг входу, не котирування букмекера.'
            if reference else '<b>Тип:</b> реальна лінія БК.'
        )
        odds_line = (
            f'<b>Коефіцієнт:</b> N/A; використовувати лише фактичний odds >= {float(DEFAULT_CONFIG.get("odds_min", 1.44)):.2f}.'
            if reference else f'<b>Коефіцієнт:</b> {_v11_num(item.get("odds"))}'
        )
        condition_line = (
            f'<b>Умова входу:</b> {html.escape(str(ref_details.get("entry_condition") or "знайти фактичну лінію не гірше reference threshold"))}.'
            if reference else '<b>Умова входу:</b> лінія, odds, рахунок і час мають залишатися актуальними.'
        )
        budget_label = (
            '<b>Рекомендований бюджет при фактичній лінії не гірше порога:</b>'
            if reference else '<b>Рекомендований бюджет:</b>'
        )
        block = [
            f'<b>{index}) {html.escape(str(adv.get("recommendation_tier") or "MICRO"))}</b> — {html.escape(_v113_bet_sentence(item))}',
            type_line,
            odds_line,
            f'{budget_label} {html.escape(str(adv.get("stake_budget") or "3-5% від бюджету"))}{amount}',
            f'<b>Рейтинг радника:</b> {float(metrics.get("score") or 0.0):.1%} | '
            f'<b>P_final:</b> {_v11_pct(metrics.get("p_final"))} | <b>P_live:</b> {_v11_pct(metrics.get("p_live"))}',
            f'<b>Історична зона:</b> {_v11_pct(metrics.get("history_zone_rate"))} | '
            f'<b>P_scenario:</b> {_v11_pct(metrics.get("p_scenario"))}',
            f'<b>Проєкція:</b> {_v11_num((item.get("live") or {}).get("projection_used"))} | '
            f'<b>Лінія:</b> {_v11_num(item.get("line"))} | <b>Edge:</b> {_v11_num(metrics.get("line_edge"))}',
            f'<b>Чому:</b> {html.escape(_v12_reason_text(item))}',
            f'<b>Головний ризик:</b> {html.escape(_v12_risk_text(item))}',
            condition_line,
        ]
        if reference:
            block.insert(5, (
                f'<b>Історичний центр:</b> {_v11_num(ref_details.get("history_center"))} '
                f'(mean {_v11_num(ref_details.get("mean"))}, median {_v11_num(ref_details.get("median"))}, N={int(ref_details.get("n") or 0)})'
            ))
        blocks.append('\n'.join(block))
    return [header + '\n\n' + '\n\n'.join(blocks)]


# =============================================================================
# v12.1 NO-BOOKMAKER REFERENCE LINE ENGINE
# =============================================================================
# A reference line is NOT presented as a bookmaker quote. It is a deterministic
# threshold anchored to the teams' same-format history and adjusted toward the
# current live projection. Telegram tells the user to act only if an actual book
# offers an equal-or-better line at odds >= configured odds_min.

V121_REFERENCE_HISTORY_WEIGHT_FULL_STAT = float(os.getenv('SUPER_BASKET_V121_REFERENCE_HISTORY_WEIGHT_FULL_STAT', '0.60'))
V121_REFERENCE_HISTORY_WEIGHT_NO_STAT = float(os.getenv('SUPER_BASKET_V121_REFERENCE_HISTORY_WEIGHT_NO_STAT', '0.72'))
V121_REFERENCE_MIN_SAMPLE = max(1, int(os.getenv('SUPER_BASKET_V121_REFERENCE_MIN_SAMPLE', '3')))
V121_REFERENCE_INTERNAL_ODDS = float(DEFAULT_CONFIG.get('odds_min', 1.44))


def _v121_trimmed_mean(values: list[float], trim_ratio: float = 0.10) -> Optional[float]:
    rows = sorted(float(value) for value in values if value is not None and math.isfinite(float(value)))
    if not rows:
        return None
    if len(rows) < 10:
        return statistics.fmean(rows)
    cut = int(len(rows) * trim_ratio)
    core = rows[cut:len(rows) - cut] if len(rows) - 2 * cut >= 3 else rows
    return statistics.fmean(core)


def _v121_history_reference_center(values: list[float]) -> dict[str, Any]:
    rows = [float(value) for value in values if value is not None and math.isfinite(float(value))]
    if not rows:
        return {'available': False, 'n': 0}
    mean = statistics.fmean(rows)
    median = statistics.median(rows)
    trimmed = _v121_trimmed_mean(rows)
    # Mean is explicitly the main anchor requested by the user; median and
    # trimmed mean reduce one-off overtime/blowout distortion.
    robust = 0.50 * mean + 0.35 * median + 0.15 * float(trimmed if trimmed is not None else mean)
    return {
        'available': True,
        'n': len(rows),
        'mean': mean,
        'median': median,
        'trimmed_mean': trimmed,
        'history_center': _v11_round_half(robust),
        'p25': percentile(rows, 0.25),
        'p75': percentile(rows, 0.75),
        'minimum': min(rows),
        'maximum': max(rows),
    }


def _v121_reference_line(
    history_center: float,
    live_projection: float,
    market: dict[str, Any],
    canonical: dict[str, Any],
) -> dict[str, Any]:
    stat_support = str(canonical.get('stat_support') or 'OFF').upper()
    history_weight = (
        V121_REFERENCE_HISTORY_WEIGHT_NO_STAT
        if stat_support in {'OFF', 'LIMITED'}
        else V121_REFERENCE_HISTORY_WEIGHT_FULL_STAT
    )
    live_weight = 1.0 - history_weight
    blended = history_weight * float(history_center) + live_weight * float(live_projection)
    clock = _segment_clock(market, canonical)
    current_points = float(clock.get('current_points') or 0.0)
    remaining_seconds = float(clock.get('remaining_seconds') or 0.0)
    # A live threshold cannot sit below points already scored in that exact scope.
    floor = current_points + 0.5 if remaining_seconds > 0 else current_points
    line = max(floor, _v11_round_half(blended))
    return {
        'line': line,
        'history_weight': history_weight,
        'live_weight': live_weight,
        'raw_blended_center': blended,
        'current_scope_points': current_points,
        'remaining_scope_seconds': remaining_seconds,
        'floor_applied': line > _v11_round_half(blended),
    }



def _v121_actual_sample_info(canonical: dict[str, Any], spec: dict[str, Any]) -> dict[str, int]:
    team = spec.get('team')
    h2h_n = len(canonical.get('history', {}).get('h2h') or [])
    if team:
        own_key = 'team_a' if team == canonical.get('home_team') else 'team_b'
        opp_key = 'team_b' if own_key == 'team_a' else 'team_a'
        own_n = len(canonical.get('history', {}).get(own_key) or [])
        allowed_n = len(canonical.get('history', {}).get(opp_key) or [])
        return {
            'own_n': own_n,
            'opponent_allowed_n': allowed_n,
            'h2h_n': h2h_n,
            'pooled_n': own_n + allowed_n,
            'total_evidence_n': own_n + allowed_n + h2h_n,
        }
    team_a_n = len(canonical.get('history', {}).get('team_a') or [])
    team_b_n = len(canonical.get('history', {}).get('team_b') or [])
    return {
        'team_a_n': team_a_n,
        'team_b_n': team_b_n,
        'h2h_n': h2h_n,
        'pooled_n': team_a_n + team_b_n,
        'total_evidence_n': team_a_n + team_b_n + h2h_n,
    }



def _v121_h2h_reference_values(canonical: dict[str, Any], spec: dict[str, Any]) -> list[float]:
    market = {**spec, 'side': 'OVER', 'line': 0.5}
    team = spec.get('team')
    values: list[float] = []
    for game in canonical.get('history', {}).get('h2h') or []:
        value = _segment_value(game, market, team if team else None)
        if value is not None:
            values.append(float(value))
    return values


def _v121_reference_evaluations(
    calculator: SuperBasketCalculator,
    canonical: dict[str, Any],
) -> list[dict[str, Any]]:
    output: list[dict[str, Any]] = []
    for spec in _v11_relevant_market_specs(canonical):
        values = _v11_history_values(canonical, spec)
        center_info = _v121_history_reference_center(values)
        sample_info = _v121_actual_sample_info(canonical, spec)
        actual_n = int(sample_info.get('pooled_n') or 0)
        if not center_info.get('available') or actual_n < V121_REFERENCE_MIN_SAMPLE:
            continue
        center_info['weighted_observation_n'] = int(center_info.get('n') or 0)
        center_info['n'] = actual_n
        center_info['sample_info'] = sample_info
        base_history_center = float(center_info['history_center'])
        h2h_values = _v121_h2h_reference_values(canonical, spec)
        h2h_info = _v121_history_reference_center(h2h_values)
        if h2h_info.get('available'):
            h2h_weight = 0.15 if spec.get('team') else 0.10
            history_center = _v11_round_half(
                (1.0 - h2h_weight) * base_history_center
                + h2h_weight * float(h2h_info['history_center'])
            )
        else:
            h2h_weight = 0.0
            history_center = base_history_center
        center_info['base_history_center'] = base_history_center
        center_info['h2h_center'] = h2h_info.get('history_center')
        center_info['h2h_n'] = len(h2h_values)
        center_info['h2h_weight'] = h2h_weight
        center_info['history_center'] = history_center
        probe_market = _v11_synthetic_market(spec, history_center, 'OVER', tag='V121_REFERENCE_PROBE')
        probe = calculator.evaluate_market(probe_market, canonical)
        projection = to_number((probe.get('live') or {}).get('projection_used'))
        if projection is None or not math.isfinite(float(projection)) or float(projection) <= 0:
            projection = history_center
        reference = _v121_reference_line(history_center, float(projection), probe_market, canonical)
        line = float(reference['line'])
        for side in ('OVER', 'UNDER'):
            evaluated = _v11_light_model_evaluation(
                spec,
                line,
                side,
                float(projection),
                values,
                canonical,
                probe,
            )
            evaluated = _v11_enrich_evaluation(evaluated, canonical, is_model=True, mine_scenarios=False)
            evaluated['odds'] = V121_REFERENCE_INTERNAL_ODDS
            evaluated['bookmaker'] = 'REFERENCE_MODEL_NO_BK'
            evaluated['source_bucket'] = 'reference_model_no_bk'
            evaluated['source_scope'] = spec.get('segment')
            evaluated['is_model_line'] = True
            evaluated['is_reference_line'] = True
            evaluated['actual_bookmaker_line_present'] = False
            evaluated['display_odds'] = None
            evaluated['reference_line_details'] = {
                **center_info,
                **reference,
                'method': '50% arithmetic mean + 35% median + 15% trimmed mean; H2H modifier; then history/live blend',
                'projection_used': float(projection),
                'entry_condition': (
                    f'OVER: actual line <= {line:.1f} and odds >= {float(DEFAULT_CONFIG.get("odds_min", 1.44)):.2f}'
                    if side == 'OVER'
                    else f'UNDER: actual line >= {line:.1f} and odds >= {float(DEFAULT_CONFIG.get("odds_min", 1.44)):.2f}'
                ),
            }
            adv = evaluated.setdefault('advisor', {})
            adv['is_model_line'] = True
            adv['is_reference_line'] = True
            adv['reference_line_details'] = deepcopy(evaluated['reference_line_details'])
            adv['actual_bookmaker_line_present'] = False
            output.append(evaluated)
    return output


def _v121_has_supported_real_lines(evaluations: list[dict[str, Any]]) -> bool:
    return any(
        _v112_market_is_currently_supported(item)
        and not bool(item.get('is_model_line'))
        and not bool(item.get('is_reference_line'))
        for item in evaluations
    )



_V12_PROCESS_BASE = process_vps_match_file


def process_vps_match_file(
    match_path: str | Path,
    *,
    output_path: str | Path | None = None,
    zones_path: str | Path | None = None,
    db_path: str | Path = 'super_basket.sqlite3',
    mode: str = 'ACTION',
    require_gpt: bool = False,
    enable_gpt: bool = False,
    enable_telegram: bool = True,
    dry_run: bool = False,
    strict_schema: bool = False,
    checkpoint: Optional[int] = None,
    gpt_reviewer: Optional[Callable[[dict[str, Any], dict[str, Any]], dict[str, Any]]] = None,
    telegram_sender: Optional[Callable[[str], dict[str, Any]]] = None,
) -> dict[str, Any]:
    """v12.1: rank real lines; when none exist, build 1-2 history-anchored reference lines; never send PASS to Telegram."""
    del require_gpt, enable_gpt, gpt_reviewer
    mode = mode.upper()
    if mode not in {'ACTION', 'STRICT'}:
        raise ValueError('mode must be ACTION or STRICT')
    source_path = Path(match_path).expanduser().resolve()
    source = load_json(source_path)
    if checkpoint is None:
        checkpoint = _v11_checkpoint_from_filename(source_path)
    if checkpoint is not None:
        checkpoint = int(checkpoint)
        if checkpoint not in {1, 2, 3}:
            raise ValueError('checkpoint must be 1, 2 or 3')
        context = source.get('analysis_context') if isinstance(source.get('analysis_context'), dict) else {}
        source['analysis_context'] = {**context, 'trigger_checkpoint': checkpoint}

    zones, zones_metadata = resolve_team_relative_zones(source, zones_path=zones_path)
    calculator = SuperBasketCalculator(deepcopy(DEFAULT_CONFIG), zones, zones_metadata)
    core_result = calculator.calculate(
        source,
        dispatch_threshold=float(DEFAULT_CONFIG.get('dispatch_threshold', 0.65)),
        strict_schema=strict_schema,
    )
    calculation = core_result['super_basket_calculation']
    canonical = adapt_match(source, deepcopy(DEFAULT_CONFIG), strict_schema)
    canonical['data_gate']['team_relative_zones'] = deepcopy(zones_metadata)
    canonical['coursework_forecast'] = build_coursework_remaining_forecast(canonical)
    input_state_gate = _v112_input_state_gate(source, canonical, checkpoint)

    real_evaluations = [
        _v11_enrich_evaluation(item, canonical, is_model=False)
        for item in calculation.get('market_evaluations') or []
    ]
    calculation['market_evaluations'] = real_evaluations
    supported_real_present = _v121_has_supported_real_lines(real_evaluations)
    reference_evaluations: list[dict[str, Any]] = []
    recommendation_pool = real_evaluations
    if not supported_real_present:
        # No bookmaker totals/IT at all: build the full relevant reference-line map.
        reference_evaluations = _v121_reference_evaluations(calculator, canonical)
        recommendation_pool = reference_evaluations
    elif _v123_is_pre_q1_or_q1_start(canonical):
        # A bookmaker may publish match/H1 lines but omit Q1.  At pre-match/Q1 start
        # still build Q1 reference totals/IT so the advisor can produce a first-quarter
        # recommendation instead of silently falling back to an unrelated match market.
        all_reference = _v121_reference_evaluations(calculator, canonical)
        real_keys = {
            _v11_market_key(item)
            for item in real_evaluations
            if _v112_market_is_currently_supported(item)
        }
        reference_evaluations = [
            item for item in all_reference
            if str(item.get('segment') or '').upper() == 'Q1'
            and _v11_market_key(item) not in real_keys
        ]
        recommendation_pool = real_evaluations + reference_evaluations
    calculation['reference_line_evaluations'] = reference_evaluations
    selected, rejected = _v12_rank_recommendations(recommendation_pool)

    can_dispatch = bool(input_state_gate.get('allowed') and selected)
    messages = _v12_build_messages({'primary_lines': selected}, calculation) if can_dispatch else []
    top = selected[0] if selected else None
    if top:
        top_adv = top.get('advisor') or {}
        action = str(top_adv.get('action') or 'RISK')
        status = str(top_adv.get('status') or 'MICRO — НАЙКРАЩИЙ ВАРІАНТ')
        top['system_action'] = action
        top['system_status'] = status
        top['stake'] = str(top_adv.get('stake_budget') or '3-5% від бюджету')
        decision = build_decision(top, top, calculation, mode)
        decision['action'] = action
        decision['deterministic_action'] = action
        decision['status'] = status
        decision['stake'] = top['stake']
        decision['budget_recommendation'] = _v12_budget_recommendation(action, status, top['stake'])
        decision['explanation_uk'] = _v12_reason_text(top)
        decision['main_risk_uk'] = _v12_risk_text(top)
        if bool(top.get('is_reference_line') or top_adv.get('is_reference_line')):
            ref = top.get('reference_line_details') or top_adv.get('reference_line_details') or {}
            decision['trigger_uk'] = str(ref.get('entry_condition') or 'Знайти фактичну лінію не гірше модельного порога та odds >=1.44.')
            if isinstance(decision.get('market'), dict):
                decision['market']['display_odds'] = None
                decision['market']['internal_reference_odds'] = decision['market'].get('odds')
                decision['market']['bookmaker'] = 'REFERENCE_MODEL_NO_BK'
                decision['market']['is_reference_line'] = True
                decision['market']['reference_line_details'] = deepcopy(ref)
        else:
            decision['trigger_uk'] = 'Рішення приймає користувач; лінія, коефіцієнт, рахунок і час мають залишатися актуальними.'
        decision['reason_codes'] = list((top_adv.get('ranking_metrics') or {}).get('major_penalties') or []) + list((top_adv.get('ranking_metrics') or {}).get('soft_penalties') or [])
    else:
        action = 'NONE'
        status = 'NO VALID REAL OR REFERENCE LINE — TELEGRAM SILENT'
        decision = build_decision(None, None, calculation, mode)
        decision.update({
            'action': 'NONE',
            'deterministic_action': 'NONE',
            'status': status,
            'stake': '0%',
            'budget_recommendation': _v12_budget_recommendation('NONE', status, '0-0%'),
            'explanation_uk': 'Немає валідної реальної лінії БК і недостатньо історії для побудови модельної reference line.',
            'main_risk_uk': 'Snapshot завершений, stale, ринок уже закритий або має критичну помилку scope/time.',
            'trigger_uk': 'Очікувати актуальний snapshot або достатню same-format історію для reference line.',
            'reason_codes': [str(input_state_gate.get('reason'))] if not input_state_gate.get('allowed') else ['NO_VALID_REAL_OR_REFERENCE_LINE'],
            'signal_id': None,
            'market': None,
        })
    decision['gpt_status'] = 'NOT_REQUIRED_V12_DETERMINISTIC'
    decision['advisor_dispatch'] = can_dispatch
    top_is_reference = bool(
        top and (
            top.get('is_reference_line')
            or (top.get('advisor') or {}).get('is_reference_line')
        )
    )
    decision['advisor_dispatch_reason'] = (
        (
            'Надіслано Q1 history-anchored reference signal, бо БК не дав відповідну Q1-лінію.'
            if top_is_reference and supported_real_present
            else 'Ліній БК немає: надіслано 1-2 history-anchored reference signals.'
            if top_is_reference
            else 'Надіслано 1-2 найсильніші реальні лінії.'
        )
        if can_dispatch else 'Telegram мовчить: немає валідної реальної або reference line.'
    )
    decision['alternative_recommendations'] = [_v11_compact_line(item) for item in selected[1:]]
    decision.pop('_evaluation', None)

    advisor = {
        'version': ADVISOR_VERSION,
        'action': action,
        'status': status,
        'dispatch': can_dispatch,
        'dispatch_reason': decision['advisor_dispatch_reason'],
        'policy': {
            'pass_telegram_disabled': True,
            'always_rank_valid_real_lines': True,
            'reference_lines_when_no_bk': True,
            'max_primary': V12_MAX_PRIMARY,
            'tiers': {
                'MICRO': '3-5% від бюджету',
                'WORKING': '10-15% від бюджету',
                'STRONG': '25-33% від бюджету',
            },
            'core_probability_math_changed': False,
        },
        'primary_lines': selected,
        'reference_lines_generated': reference_evaluations,
        'real_lines_present': supported_real_present,
        'rejected_fatal_lines': rejected,
        'telegram_messages': messages,
        'input_state_gate': input_state_gate,
    }

    target = Path(output_path).expanduser().resolve() if output_path else source_path.with_name(source_path.stem + '_advisor_result.json')
    delivery = {'status': 'SKIPPED_NO_VALID_SIGNAL', 'sent': False, 'message_id': None, 'message_count': len(messages)}
    duplicate = False
    advisor_key = hashlib.sha256((calculation['input_snapshot_hash'] + '|' + ADVISOR_VERSION).encode('utf-8')).hexdigest()
    connection = _v11_delivery_connect(db_path)
    try:
        row = connection.execute('SELECT telegram_status FROM advisor_deliveries WHERE advisor_key=?', (advisor_key,)).fetchone()
        duplicate = row is not None and row[0] == 'SENT'
        if not can_dispatch:
            delivery = {'status': 'SKIPPED_NO_VALID_SIGNAL', 'sent': False, 'message_id': None, 'message_count': 0}
        elif duplicate:
            delivery = {'status': 'SKIPPED_DUPLICATE_ALREADY_SENT', 'sent': False, 'message_id': None, 'message_count': len(messages)}
        elif dry_run:
            delivery = {'status': 'DRY_RUN_NOT_SENT', 'sent': False, 'message_id': None, 'message_count': len(messages)}
        elif not enable_telegram:
            delivery = {'status': 'SKIPPED_TELEGRAM_DISABLED', 'sent': False, 'message_id': None, 'message_count': len(messages)}
        else:
            delivery = _v11_send_messages(messages, telegram_sender=telegram_sender)
        connection.execute(
            'INSERT OR REPLACE INTO advisor_deliveries(advisor_key,match_id,input_hash,action,status,telegram_status,created_at) VALUES(?,?,?,?,?,?,?)',
            (advisor_key, canonical['match_id'], calculation['input_snapshot_hash'], action, status, delivery['status'], utc_now()),
        )
        connection.commit()
    finally:
        connection.close()

    learning_duplicate = False
    learning_store = LearningStore(db_path)
    try:
        decision['telegram_status'] = delivery['status']
        if top and decision.get('signal_id') and decision.get('market'):
            _, learning_duplicate = learning_store.record_signal(decision, calculation)
            learning_store.update_delivery(
                decision['signal_id'],
                action,
                decision.get('gpt_status', 'NOT_REQUIRED_V12_DETERMINISTIC'),
                delivery['status'],
                delivery.get('message_id'),
            )
        learning_store.mark_processed(
            calculation['input_snapshot_hash'],
            str(source_path),
            str(target),
            'OK' if input_state_gate.get('allowed') else str(input_state_gate.get('reason')),
        )
    finally:
        learning_store.close()

    all_evaluations_for_output = real_evaluations + reference_evaluations
    line_recommendations = [_v11_compact_line(item) for item in all_evaluations_for_output]
    selected_compact = [_v11_compact_line(item) for item in selected]
    calculation['line_recommendations'] = line_recommendations
    calculation['active_line_recommendations'] = selected_compact
    calculation['advisor'] = advisor
    calculation['advisor_ranked_recommendations'] = selected_compact
    calculation['advisor_rejected_fatal_lines'] = rejected
    calculation['reference_line_mode'] = {
        'active': not supported_real_present,
        'generated_count': len(reference_evaluations),
        'method': 'history mean/median/trimmed mean + live adjustment',
        'not_a_bookmaker_quote': True,
    }

    system = {
        'version': ADVISOR_VERSION,
        'processed_at': utc_now(),
        'input_hash': calculation['input_snapshot_hash'],
        'mode': mode,
        'status': 'OK',
        'data_gate': calculation.get('data_gate'),
        'format_gate': format_gate(calculation),
        'market_audit': calculation.get('market_audit'),
        'line_coverage': calculation.get('line_coverage'),
        'line_recommendations': line_recommendations,
        'active_line_recommendations': selected_compact,
        'decision': decision,
        'decision_text': f'{action} | {status}',
        'gpt_review': {'status': 'NOT_REQUIRED_V12_DETERMINISTIC', 'approved': True, 'action': action},
        'risk_post_filter': {'enabled': False, 'policy': 'V12_ALWAYS_RANKED'},
        'telegram_delivery': {**delivery, 'duplicate_signal': duplicate},
        'learning_store': {'signal_recorded': bool(top and decision.get('signal_id')), 'duplicate_signal': learning_duplicate},
        'advisor': advisor,
        'input_state_gate': input_state_gate,
        'files': {'source': str(source_path), 'result': str(target)},
    }
    core_result['super_basket_system'] = system
    save_json(target, core_result)
    append_verdict_log({
        'timestamp': system['processed_at'],
        'match_id': canonical['match_id'],
        'match_name': canonical['name'],
        'checkpoint': canonical['stage'],
        'trigger_checkpoint': canonical.get('trigger_checkpoint'),
        'verdict': action,
        'verdict_status': status,
        'p_final': ((top.get('advisor') or {}).get('p_final') if top else None),
        'market': (_v11_compact_line(top) if top else None),
        'alternatives': [_v11_compact_line(item) for item in selected[1:]],
        'description': decision['advisor_dispatch_reason'],
        'reason_codes': decision.get('reason_codes') or [],
        'input_hash': calculation['input_snapshot_hash'],
        'gpt_status': 'NOT_REQUIRED_V12_DETERMINISTIC',
        'telegram_status': delivery['status'],
        'advisor_policy': advisor['policy'],
        'files': {'source': str(source_path), 'result': str(target)},
    })
    if ENABLE_EXCEL_AUDIT:
        append_excel_audit(core_result)
    return core_result


DEFAULT_CONFIG.setdefault('v12_advisor_policy', {})
DEFAULT_CONFIG['v12_advisor_policy'].update({
    'version': ADVISOR_VERSION,
    'pass_telegram_disabled': True,
    'always_rank_valid_real_lines': True,
    'reference_lines_when_no_bk': True,
    'max_primary': V12_MAX_PRIMARY,
    'budget_tiers': {
        'MICRO': '3-5%',
        'WORKING': '10-15%',
        'STRONG': '25-33%',
    },
    'core_probability_math_changed': False,
})
SYSTEM_VERSION = ADVISOR_VERSION
DEFAULT_CONFIG['engine_version'] = ADVISOR_VERSION


# ===== v12.2 clock/score consistency guard =====
# The advisor must never silently combine a stale score source with a newer clock source.
# This wrapper cross-checks match.score, raw_data.main_match hs/as_, quarter sums and all
# available clock representations.  A resolvable disagreement is repaired; an unresolved
# disagreement blocks Telegram while keeping the full audit in the output JSON.
_V122_ADAPT_BASE = adapt_match
_V122_INPUT_STATE_GATE_BASE = _v112_input_state_gate


def _v122_pair_score(mapping: Any, home_keys: tuple[str, ...], away_keys: tuple[str, ...]) -> Optional[tuple[float, float]]:
    if not isinstance(mapping, dict):
        return None
    home = to_number(first(mapping, home_keys))
    away = to_number(first(mapping, away_keys))
    if home is None or away is None:
        return None
    return float(home), float(away)


def _v122_score_equal(left: Optional[tuple[float, float]], right: Optional[tuple[float, float]], tolerance: float=0.01) -> bool:
    return bool(
        left is not None and right is not None
        and abs(left[0] - right[0]) <= tolerance
        and abs(left[1] - right[1]) <= tolerance
    )


def _v122_quarter_score(canonical: dict[str, Any]) -> Optional[tuple[float, float]]:
    rows = canonical.get('quarters') or []
    current = to_int(canonical.get('current_quarter')) or 4
    home = away = 0.0
    used = 0
    for index, row in enumerate(rows[:max(1, min(4, current))]):
        if not isinstance(row, dict):
            continue
        qh = to_number(row.get('home'))
        qa = to_number(row.get('away'))
        if qh is None or qa is None:
            continue
        home += float(qh)
        away += float(qa)
        used += 1
    return (home, away) if used else None


def _v122_clock_candidates(source: dict[str, Any], canonical: dict[str, Any]) -> list[dict[str, Any]]:
    match = source.get('match') if isinstance(source.get('match'), dict) else {}
    raw = ((source.get('raw_data') or {}).get('main_match') or {}) if isinstance(source.get('raw_data'), dict) else {}
    quarter_minutes = int(canonical.get('quarter_minutes') or 10)
    quarter_seconds = quarter_minutes * 60
    full_seconds = int(canonical.get('full_game_seconds') or quarter_seconds * 4)
    rows: list[dict[str, Any]] = []

    elapsed = to_number(first(match, ['match_minute_played', 'elapsed_minutes']))
    if elapsed is not None:
        rows.append({'source': 'match.match_minute_played', 'elapsed_seconds': int(round(float(elapsed) * 60)), 'priority': 2})

    period = to_int(first(match, ['period', 'quarter', 'current_quarter']))
    played = to_number(first(match, ['period_minute_played', 'quarter_minute_played']))
    left = to_number(first(match, ['period_minute_left', 'quarter_minute_left']))
    if period is not None and played is not None:
        safe_played = max(0.0, min(float(quarter_minutes), float(played)))
        rows.append({
            'source': 'match.period_minute_played',
            'elapsed_seconds': int(round(((period - 1) * quarter_minutes + safe_played) * 60)),
            'period': period,
            'period_played_seconds': int(round(safe_played * 60)),
            'priority': 4,
        })
    if period is not None and left is not None:
        safe_left = max(0.0, min(float(quarter_minutes), float(left)))
        played_from_left = float(quarter_minutes) - safe_left
        rows.append({
            'source': 'match.period_minute_left',
            'elapsed_seconds': int(round(((period - 1) * quarter_minutes + played_from_left) * 60)),
            'period': period,
            'period_played_seconds': int(round(played_from_left * 60)),
            'priority': 5,
        })

    raw_status = str(first(raw, ['st', 'status']) or '')
    parsed = _parse_status_clock(raw_status, quarter_seconds, full_seconds)
    if parsed is not None:
        raw_elapsed, raw_period, raw_played = parsed
        rows.append({
            'source': 'raw_data.main_match.st',
            'elapsed_seconds': int(raw_elapsed),
            'period': raw_period,
            'period_played_seconds': raw_played,
            'priority': 3,
            'raw_status': raw_status,
        })

    # Clamp all candidates to regulation time.
    for row in rows:
        row['elapsed_seconds'] = max(0, min(full_seconds, int(row['elapsed_seconds'])))
    return rows


def _v122_choose_clock(candidates: list[dict[str, Any]], canonical: dict[str, Any]) -> tuple[Optional[dict[str, Any]], bool, list[dict[str, Any]]]:
    if not candidates:
        return None, False, []
    tolerance = 75  # provider status strings are usually minute-granular
    clusters: list[list[dict[str, Any]]] = []
    for row in sorted(candidates, key=lambda item: int(item['elapsed_seconds'])):
        placed = False
        for cluster in clusters:
            center = statistics.median(int(item['elapsed_seconds']) for item in cluster)
            if abs(int(row['elapsed_seconds']) - center) <= tolerance:
                cluster.append(row)
                placed = True
                break
        if not placed:
            clusters.append([row])
    clusters.sort(
        key=lambda cluster: (
            len(cluster),
            max(int(item.get('priority') or 0) for item in cluster),
        ),
        reverse=True,
    )
    winning = clusters[0]
    best = max(winning, key=lambda item: int(item.get('priority') or 0))
    unresolved = len(clusters) > 1 and len(winning) == len(clusters[1]) == 1
    return best, unresolved, clusters


def adapt_match(source: dict[str, Any], config: dict[str, Any], strict: bool=False) -> dict[str, Any]:
    canonical = _V122_ADAPT_BASE(source, config, strict)
    gate = canonical.setdefault('data_gate', {})
    match = source.get('match') if isinstance(source.get('match'), dict) else {}
    raw = ((source.get('raw_data') or {}).get('main_match') or {}) if isinstance(source.get('raw_data'), dict) else {}

    match_score = _v122_pair_score(match.get('score') or {}, ('home', 'home_score'), ('away', 'away_score'))
    raw_score = _v122_pair_score(raw, ('hs', 'home_score', 'homeScore'), ('as_', 'away_score', 'awayScore'))
    quarter_score = _v122_quarter_score(canonical)
    selected_before = (
        float((canonical.get('score') or {}).get('home') or 0.0),
        float((canonical.get('score') or {}).get('away') or 0.0),
    )
    available = {name: value for name, value in {
        'match.score': match_score,
        'raw_data.main_match': raw_score,
        'quarter_sum': quarter_score,
    }.items() if value is not None}
    score_conflicts = []
    names = list(available)
    for index, name in enumerate(names):
        for other in names[index + 1:]:
            if not _v122_score_equal(available[name], available[other]):
                score_conflicts.append({'left_source': name, 'left': list(available[name]), 'right_source': other, 'right': list(available[other])})

    resolved_score = selected_before
    resolved_source = 'base_priority'
    unresolved_score = False
    if quarter_score is not None:
        matches = [name for name, value in available.items() if name != 'quarter_sum' and _v122_score_equal(value, quarter_score)]
        if matches:
            resolved_score = quarter_score
            resolved_source = 'quarter_sum_confirmed_by_' + matches[0]
        elif len(available) >= 2 and score_conflicts:
            unresolved_score = True
    elif match_score is not None and raw_score is not None and not _v122_score_equal(match_score, raw_score):
        unresolved_score = True

    if resolved_score != selected_before:
        canonical['score'] = {
            'home': float(resolved_score[0]),
            'away': float(resolved_score[1]),
            'total': float(resolved_score[0] + resolved_score[1]),
            'margin_home': float(resolved_score[0] - resolved_score[1]),
        }

    clock_candidates = _v122_clock_candidates(source, canonical)
    chosen_clock, unresolved_time, clock_clusters = _v122_choose_clock(clock_candidates, canonical)
    if chosen_clock is not None and not unresolved_time:
        full_seconds = int(canonical.get('full_game_seconds') or 2400)
        quarter_seconds = int(canonical.get('quarter_seconds') or 600)
        elapsed_seconds = int(chosen_clock['elapsed_seconds'])
        period = to_int(chosen_clock.get('period'))
        if period is None and elapsed_seconds < full_seconds:
            period = min(4, elapsed_seconds // quarter_seconds + 1)
        if period is not None:
            played_seconds = elapsed_seconds - (period - 1) * quarter_seconds
            played_seconds = max(0, min(quarter_seconds, played_seconds))
            left_seconds = max(0, quarter_seconds - played_seconds)
        else:
            left_seconds = 0
        explicit_stage = str(canonical.get('explicit_stage') or '')
        canonical['elapsed_game_seconds'] = elapsed_seconds
        canonical['remaining_game_seconds'] = max(0, full_seconds - elapsed_seconds)
        canonical['current_quarter'] = period
        canonical['quarter_seconds_remaining'] = left_seconds
        canonical['clock'] = f'{left_seconds // 60:02d}:{left_seconds % 60:02d}' if period is not None else None
        canonical['stage'] = _stage(elapsed_seconds, full_seconds, quarter_seconds, explicit_stage)

    gate['score_consistency'] = {
        'sources': {name: list(value) for name, value in available.items()},
        'selected_before_guard': list(selected_before),
        'selected_after_guard': [canonical['score']['home'], canonical['score']['away']],
        'selected_source': resolved_source,
        'conflicts': score_conflicts,
        'unresolved': unresolved_score,
    }
    gate['clock_consistency'] = {
        'candidates': clock_candidates,
        'selected': deepcopy(chosen_clock),
        'clusters': [[deepcopy(item) for item in cluster] for cluster in clock_clusters],
        'unresolved': unresolved_time,
        'final_elapsed_seconds': canonical.get('elapsed_game_seconds'),
        'final_clock': canonical.get('clock'),
        'final_quarter': canonical.get('current_quarter'),
    }
    gate['score_conflict'] = unresolved_score
    gate['time_conflict'] = unresolved_time
    if unresolved_score and 'INPUT_SCORE_CONFLICT' not in gate.setdefault('schema_errors', []):
        gate['schema_errors'].append('INPUT_SCORE_CONFLICT')
    if unresolved_time and 'INPUT_TIME_CONFLICT' not in gate.setdefault('schema_errors', []):
        gate['schema_errors'].append('INPUT_TIME_CONFLICT')
    return canonical


def _v112_input_state_gate(source: dict[str, Any], canonical: dict[str, Any], checkpoint: Optional[int]) -> dict[str, Any]:
    result = _V122_INPUT_STATE_GATE_BASE(source, canonical, checkpoint)
    gate = canonical.get('data_gate') or {}
    score_conflict = bool(gate.get('score_conflict'))
    time_conflict = bool(gate.get('time_conflict'))
    if score_conflict or time_conflict:
        result['allowed'] = False
        if score_conflict and time_conflict:
            result['reason'] = 'INPUT_SCORE_AND_TIME_CONFLICT'
        elif score_conflict:
            result['reason'] = 'INPUT_SCORE_CONFLICT'
        else:
            result['reason'] = 'INPUT_TIME_CONFLICT'
    result['score_consistency'] = deepcopy(gate.get('score_consistency'))
    result['clock_consistency'] = deepcopy(gate.get('clock_consistency'))
    return result


ADVISOR_VERSION = '12.2.0-CLOCK-SCORE-CONSISTENCY-GUARD'
SYSTEM_VERSION = ADVISOR_VERSION
DEFAULT_CONFIG['engine_version'] = ADVISOR_VERSION
DEFAULT_CONFIG.setdefault('v12_advisor_policy', {})['clock_score_consistency_guard'] = True


# =============================================================================
# v12.3 PRE-MATCH + Q1-START ADVISOR
# =============================================================================
# Adds two production states without changing the historical/scenario/projection
# formulas:
# 1) PRE_MATCH: real bookmaker totals/IT or history-anchored reference lines can be
#    ranked even when the score is 0:0 and there are no live statistics.
# 2) Q1_START: Q1 total and both Q1 team-IT markets are evaluated from 0:0 or the
#    first live possessions, with exact score/clock semantics preserved.

V123_Q1_FORCE_WINDOW_SECONDS = max(
    0,
    int(os.getenv('SUPER_BASKET_V123_Q1_FORCE_WINDOW_SECONDS', '120')),
)
V123_PREMATCH_REFERENCE_Q1_ENABLED = env_bool(
    'SUPER_BASKET_V123_PREMATCH_REFERENCE_Q1_ENABLED', True
)


def _v123_is_pre_q1_or_q1_start(canonical: dict[str, Any]) -> bool:
    stage = str(canonical.get('stage') or '').upper()
    current = to_int(canonical.get('current_quarter'))
    elapsed = int(canonical.get('elapsed_game_seconds') or 0)
    qsec = int(canonical.get('quarter_seconds') or 600)
    return bool(
        stage == 'PRE_MATCH'
        or (current == 1 and 0 <= elapsed < min(qsec, V123_Q1_FORCE_WINDOW_SECONDS + 1))
    )


_V123_ADAPT_BASE = adapt_match


def adapt_match(source: dict[str, Any], config: dict[str, Any], strict: bool=False) -> dict[str, Any]:
    canonical = _V123_ADAPT_BASE(source, config, strict)
    current = to_int(canonical.get('current_quarter'))
    elapsed = int(canonical.get('elapsed_game_seconds') or 0)
    qsec = int(canonical.get('quarter_seconds') or 600)
    score = canonical.get('score') or {}
    rows = canonical.setdefault('quarters', [])
    while len(rows) < 4:
        rows.append({'home': None, 'away': None, 'total': None})
    backfilled = False
    # At pre-match or during Q1, the cumulative game score equals the Q1 score.
    # Some parser snapshots leave q1h/q1a empty until the first provider refresh;
    # backfill only this exact state so Q1 projections never use a false 0 total.
    if current == 1 and elapsed < qsec:
        q1 = rows[0]
        home = to_number(score.get('home'))
        away = to_number(score.get('away'))
        if home is not None and away is not None and (
            q1.get('home') is None or q1.get('away') is None
        ):
            q1['home'] = float(home)
            q1['away'] = float(away)
            q1['total'] = float(home + away)
            backfilled = True
    canonical.setdefault('data_gate', {})['q1_score_backfilled_from_match_score'] = backfilled
    explicit_upper = str(canonical.get('explicit_stage') or '').upper()
    explicit_q1_live = current == 1 and any(token in explicit_upper for token in ('LIVE', 'Q1', 'ЧВЕРТ', 'QUARTER'))
    canonical['data_gate']['advisor_phase'] = (
        'Q1_START'
        if explicit_q1_live and elapsed <= V123_Q1_FORCE_WINDOW_SECONDS
        else 'PRE_MATCH'
        if str(canonical.get('stage') or '').upper() == 'PRE_MATCH' and elapsed == 0
        else 'Q1_START'
        if current == 1 and elapsed <= V123_Q1_FORCE_WINDOW_SECONDS
        else str(canonical.get('stage') or 'UNKNOWN')
    )
    return canonical


_V123_CURRENT_QUARTER_ISSUE_BASE = _current_quarter_issue


def _current_quarter_issue(
    market_type: str,
    segment: str,
    canonical: dict[str, Any],
) -> Optional[str]:
    target = int(segment[1:]) if segment.startswith('Q') and segment[1:].isdigit() else None
    if (
        market_type in {'CURRENT_QUARTER_TOTAL', 'CURRENT_QUARTER_TEAM_IT'}
        and target == 1
    ):
        stage = str(canonical.get('stage') or '').upper()
        current = to_int(canonical.get('current_quarter'))
        elapsed = int(canonical.get('elapsed_game_seconds') or 0)
        qsec = int(canonical.get('quarter_seconds') or 600)
        # Before tip-off, a Q1 line is a valid pre-match market and does not need
        # a live clock or a populated Q1 boxscore row.
        if stage == 'PRE_MATCH' and elapsed == 0:
            return None
        # During Q1, retain the normal exact-time requirement.  The v12.3 adapter
        # backfills Q1 score from the cumulative score when the provider omits q1h/q1a.
        if current == 1 and elapsed < qsec and canonical.get('clock') is not None:
            return None
    return _V123_CURRENT_QUARTER_ISSUE_BASE(market_type, segment, canonical)


_V123_RELEVANT_MARKET_SPECS_BASE = _v11_relevant_market_specs


def _v11_relevant_market_specs(canonical: dict[str, Any]) -> list[dict[str, Any]]:
    specs = list(_V123_RELEVANT_MARKET_SPECS_BASE(canonical))
    if V123_PREMATCH_REFERENCE_Q1_ENABLED and _v123_is_pre_q1_or_q1_start(canonical):
        q1_specs = [
            {'market_type': 'CURRENT_QUARTER_TOTAL', 'segment': 'Q1', 'team': None},
            {'market_type': 'CURRENT_QUARTER_TEAM_IT', 'segment': 'Q1', 'team': canonical.get('home_team')},
            {'market_type': 'CURRENT_QUARTER_TEAM_IT', 'segment': 'Q1', 'team': canonical.get('away_team')},
        ]
        specs = q1_specs + specs
    output: list[dict[str, Any]] = []
    seen: set[tuple[Any, ...]] = set()
    for spec in specs:
        key = (spec.get('market_type'), spec.get('team'), spec.get('segment'))
        if key not in seen:
            seen.add(key)
            output.append(spec)
    return output


_V123_RANK_RECOMMENDATIONS_BASE = _v12_rank_recommendations


def _v123_eval_elapsed(item: dict[str, Any]) -> Optional[int]:
    value = to_number((item.get('live') or {}).get('elapsed_game_seconds'))
    return None if value is None else int(round(float(value)))


def _v12_rank_recommendations(
    evaluations: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    selected, rejected = _V123_RANK_RECOMMENDATIONS_BASE(evaluations)
    q1_pool = [
        item for item in evaluations
        if str(item.get('segment') or '').upper() == 'Q1'
        and (_v123_eval_elapsed(item) is None or _v123_eval_elapsed(item) <= V123_Q1_FORCE_WINDOW_SECONDS)
    ]
    if not q1_pool:
        return selected, rejected
    q1_selected, q1_rejected = _V123_RANK_RECOMMENDATIONS_BASE(q1_pool)
    if not q1_selected:
        return selected, rejected + q1_rejected

    primary = q1_selected[0]
    combined = [primary]
    used_groups = {_v12_market_group(primary)}
    # Preserve a second genuinely different strongest option (Q1 or another scope),
    # but never replace the requested first-quarter primary at pre-match/Q1 start.
    candidates = q1_selected[1:] + selected
    for item in candidates:
        group = _v12_market_group(item)
        if group in used_groups:
            continue
        combined.append(item)
        used_groups.add(group)
        if len(combined) >= V12_MAX_PRIMARY:
            break
    return combined[:V12_MAX_PRIMARY], rejected + q1_rejected


_V123_CHECKPOINT_FILENAME_BASE = _v11_checkpoint_from_filename


def _v11_checkpoint_from_filename(path: Path) -> Optional[int]:
    """Do not misread Q1-start/live files as the after-Q1 checkpoint.

    Explicit result/end/after names remain checkpoints.  PRE_MATCH/Q1_START/Q1_LIVE
    files are analysed with checkpoint=None so Q1 totals and Q1 team-IT stay active.
    """
    stem = path.stem.lower()
    if re.search(r'(?:^|_)(?:pre_?match|not_?started|q1_?start|start_?q1|q1_?live|live_?q1)(?:_|$)', stem):
        return None
    explicit = re.search(r'(?:^|_)(?:after|end|finished)_?q([123])(?:_|$)', stem)
    if explicit:
        return int(explicit.group(1))
    result = re.search(r'(?:^|_)q([123])_?(?:result|checkpoint|done)(?:_|$)', stem)
    if result:
        return int(result.group(1))
    # A bare q1 token is ambiguous and is commonly used for a live/start snapshot.
    # Only explicit q1_result/after_q1/end_q1 names may mean checkpoint 1.
    if re.search(r'(?:^|_)q1(?:_|$)', stem):
        return None
    return _V123_CHECKPOINT_FILENAME_BASE(path)


ADVISOR_VERSION = '12.3.0-PREMATCH-Q1-START-ADVISOR'
SYSTEM_VERSION = ADVISOR_VERSION
DEFAULT_CONFIG['engine_version'] = ADVISOR_VERSION
DEFAULT_CONFIG.setdefault('v12_advisor_policy', {}).update({
    'pre_match_recommendations': True,
    'q1_start_recommendations': True,
    'q1_reference_line_when_missing': True,
    'q1_force_window_seconds': V123_Q1_FORCE_WINDOW_SECONDS,
    'pre_match_live_points_required': False,
    'pre_match_max_tier_without_live_confirmation': 'WORKING',
})



# =============================================================================
# v12.4 FIVE CHECKPOINTS + STAT/PROJECTION PRESERVATION FIX
# =============================================================================
# Public checkpoint semantics (do not confuse with the legacy completed-quarter
# trigger used internally by scenario calculations):
#   1 = PREMATCH
#   2 = EARLY_LIVE_Q1, approximately 2:00 played in Q1
#   3 = EARLY_LIVE_Q2, approximately 2:00 played in Q2
#   4 = HT
#   5 = Q4_CONFIRMATION, approximately 4:00 played in Q4
# Exact provider score/clock remains authoritative. The checkpoint supplies a
# fallback clock only when the parser omitted it and never discards otherwise
# valid lines or parser live projections merely because of a naming/time mismatch.

ADVISOR_VERSION = '12.4.0-FIVE-CHECKPOINTS-STAT-PROJECTION-FIX'
SYSTEM_VERSION = ADVISOR_VERSION
DEFAULT_CONFIG['engine_version'] = ADVISOR_VERSION

V124_CHECKPOINT_NAMES = {
    1: 'PREMATCH',
    2: 'EARLY_LIVE_Q1',
    3: 'EARLY_LIVE_Q2',
    4: 'HT',
    5: 'Q4_CONFIRMATION',
}
V124_CHECKPOINT_TOLERANCE_SECONDS = max(
    60, int(os.getenv('SUPER_BASKET_V124_CHECKPOINT_TOLERANCE_SECONDS', '180'))
)
V124_PARSER_PROJECTION_STALE_SECONDS = max(
    60, int(os.getenv('SUPER_BASKET_V124_PARSER_PROJECTION_STALE_SECONDS', '240'))
)

# Runtime bridge used by the legacy CLI/watch loop. Processing is sequential in
# the shipped watcher; the bridge is cleared in a finally block after every file.
_V124_RUNTIME_PATH: Optional[Path] = None
_V124_RUNTIME_CHECKPOINT: Optional[int] = None
_V124_BASE_PROCESS_ACTIVE = False
_V124_CLI_FORCED_CHECKPOINT: Optional[int] = None
_V124_LAST_CHECKPOINT: Optional[int] = None


def _v124_checkpoint_expected(cp: int, quarter_seconds: int) -> dict[str, Any]:
    q = int(quarter_seconds or 600)
    table = {
        1: {'period': None, 'period_played_seconds': 0, 'elapsed_seconds': 0},
        2: {'period': 1, 'period_played_seconds': 120, 'elapsed_seconds': 120},
        3: {'period': 2, 'period_played_seconds': 120, 'elapsed_seconds': q + 120},
        4: {'period': None, 'period_played_seconds': None, 'elapsed_seconds': 2 * q},
        5: {'period': 4, 'period_played_seconds': 240, 'elapsed_seconds': 3 * q + 240},
    }
    return {'checkpoint': cp, 'name': V124_CHECKPOINT_NAMES[cp], **table[cp]}


def _v124_period_played(match: dict[str, Any], quarter_minutes: float=10.0) -> tuple[Optional[int], Optional[float]]:
    period = to_int(first(match, ['period', 'quarter', 'current_quarter']))
    played = to_number(first(match, ['period_minute_played', 'quarter_minute_played']))
    if played is None:
        left = to_number(first(match, ['period_minute_left', 'quarter_minute_left']))
        if left is not None:
            played = max(0.0, min(float(quarter_minutes), float(quarter_minutes) - float(left)))
    return period, played


def _v124_checkpoint_from_source(source: dict[str, Any], path: Optional[Path]=None) -> Optional[int]:
    context = source.get('analysis_context') if isinstance(source.get('analysis_context'), dict) else {}
    explicit = to_int(first(context, ['advisor_checkpoint', 'checkpoint_5stage', 'five_stage_checkpoint']))
    if explicit in V124_CHECKPOINT_NAMES:
        return explicit

    match = source.get('match') if isinstance(source.get('match'), dict) else {}
    stage = str(first(match, ['stage', 'status']) or source.get('stage') or source.get('status') or '').upper()
    quarter_minutes = to_number(first(match, ['quarter_minutes', 'period_minutes'])) or 10.0
    period, played = _v124_period_played(match, quarter_minutes)
    elapsed = to_number(first(match, ['match_minute_played', 'elapsed_minutes']))
    score = match.get('score') if isinstance(match.get('score'), dict) else {}
    score_total = to_number(score.get('total'))
    if score_total is None:
        home = to_number(score.get('home')); away = to_number(score.get('away'))
        score_total = home + away if home is not None and away is not None else None

    if any(token in stage for token in ('PRE_MATCH', 'PREMATCH', 'NOT_STARTED', 'SCHEDULED')):
        return 1
    if stage in {'HT', 'HALFTIME', 'HALF_TIME'} or 'ПЕРЕРВА' in stage or 'AFTER_Q2' in stage:
        return 4
    if period == 1 and (played is None or played <= 5.0):
        return 2
    if period == 2 and (played is None or played <= 5.0):
        return 3
    if period == 4:
        return 5
    if elapsed is not None:
        if float(elapsed) <= 0.01 and (score_total in (None, 0.0)):
            return 1
        if 0.0 < float(elapsed) <= 5.0:
            return 2
        if 10.0 <= float(elapsed) <= 15.0:
            return 3
        if abs(float(elapsed) - 20.0) <= 1.0:
            return 4
        if 30.0 <= float(elapsed) <= 40.0:
            return 5

    stem = (path.stem.lower() if path else '')
    if re.search(r'(?:^|_)(?:pre_?match|prematch|not_?started)(?:_|$)', stem):
        return 1
    if re.search(r'(?:^|_)(?:early_?live_?q1|q1_?2min|q1_?start|start_?q1)(?:_|$)', stem):
        return 2
    if re.search(r'(?:^|_)(?:early_?live_?q2|q2_?2min|q1_?result|after_?q1)(?:_|$)', stem):
        return 3
    if re.search(r'(?:^|_)(?:ht|half_?time|q2_?result|after_?q2)(?:_|$)', stem):
        return 4
    if re.search(r'(?:^|_)(?:q4_?confirmation|q4_?4min|early_?q4)(?:_|$)', stem):
        return 5
    return None


def _v124_prepare_source(source: dict[str, Any], cp: Optional[int]) -> dict[str, Any]:
    output = deepcopy(source)
    if cp not in V124_CHECKPOINT_NAMES:
        return output
    match = output.get('match') if isinstance(output.get('match'), dict) else {}
    output['match'] = match
    rules = output.get('rules') if isinstance(output.get('rules'), dict) else {}
    qmin = to_int(first(rules, ['quarter_minutes', 'period_minutes'])) or to_int(first(match, ['quarter_minutes', 'period_minutes'])) or 10
    qsec = int(qmin) * 60
    expected = _v124_checkpoint_expected(cp, qsec)
    context = output.get('analysis_context') if isinstance(output.get('analysis_context'), dict) else {}
    legacy = {1: None, 2: None, 3: 1, 4: 2, 5: 3}[cp]
    context.update({
        'advisor_checkpoint': cp,
        'advisor_checkpoint_name': V124_CHECKPOINT_NAMES[cp],
        'checkpoint_5stage': cp,
        'checkpoint_expected_elapsed_seconds': expected['elapsed_seconds'],
        'checkpoint_expected_period': expected['period'],
        'checkpoint_expected_period_played_seconds': expected['period_played_seconds'],
        'checkpoint_tolerance_seconds': V124_CHECKPOINT_TOLERANCE_SECONDS,
        'legacy_trigger_checkpoint': legacy,
    })
    if legacy is None:
        context.pop('trigger_checkpoint', None)
    else:
        context['trigger_checkpoint'] = legacy
    output['analysis_context'] = context

    # Exact parser time wins. Only synthesize missing fields.
    period, played = _v124_period_played(match, qmin)
    elapsed = to_number(first(match, ['match_minute_played', 'elapsed_minutes']))
    synthesized: list[str] = []
    if cp == 1:
        if not str(match.get('stage') or '').strip():
            match['stage'] = 'PRE_MATCH'; synthesized.append('stage')
        if elapsed is None:
            match['match_minute_played'] = 0.0; synthesized.append('match_minute_played')
        if not isinstance(match.get('score'), dict):
            match['score'] = {'home': 0, 'away': 0, 'total': 0, 'margin_home': 0}; synthesized.append('score')
    elif cp in {2, 3, 5}:
        target_period = int(expected['period'])
        target_played = float(expected['period_played_seconds']) / 60.0
        if period is None:
            match['period'] = target_period; synthesized.append('period')
            period = target_period
        if played is None:
            match['period_minute_played'] = target_played
            match['period_minute_left'] = max(0.0, float(qmin) - target_played)
            synthesized.extend(['period_minute_played', 'period_minute_left'])
        if elapsed is None:
            match['match_minute_played'] = float(expected['elapsed_seconds']) / 60.0
            match['match_minute_left'] = max(0.0, 4.0 * qmin - float(match['match_minute_played']))
            synthesized.extend(['match_minute_played', 'match_minute_left'])
        if not str(match.get('stage') or '').strip():
            match['stage'] = f'Q{target_period}_live'; synthesized.append('stage')
    elif cp == 4:
        if elapsed is None:
            match['match_minute_played'] = 2.0 * qmin
            match['match_minute_left'] = 2.0 * qmin
            synthesized.extend(['match_minute_played', 'match_minute_left'])
        if not str(match.get('stage') or '').strip():
            match['stage'] = 'HT'; synthesized.append('stage')
    context['checkpoint_synthesized_fields'] = synthesized
    context['checkpoint_time_synthesized'] = bool(synthesized)
    return output


# Source injection bridge.
_V124_LOAD_JSON_BASE = load_json


def load_json(path: str | Path) -> dict[str, Any]:
    data = _V124_LOAD_JSON_BASE(path)
    try:
        resolved = Path(path).expanduser().resolve()
    except Exception:
        resolved = None
    if _V124_RUNTIME_PATH is not None and resolved == _V124_RUNTIME_PATH:
        return _v124_prepare_source(data, _V124_RUNTIME_CHECKPOINT)
    return data


# New filename inference for the watcher. While the legacy v12.3 process body is
# executing, return None so it cannot reinterpret 1..5 as the old 1..3 system.
_V124_FILENAME_BASE = _v11_checkpoint_from_filename


def _v11_checkpoint_from_filename(path: Path) -> Optional[int]:
    if _V124_BASE_PROCESS_ACTIVE:
        return None
    try:
        source = _V124_LOAD_JSON_BASE(path)
    except Exception:
        source = {}
    checkpoint = _v124_checkpoint_from_source(source, Path(path))
    if checkpoint in V124_CHECKPOINT_NAMES:
        return checkpoint
    return None


def _v124_stat_integrity(canonical: dict[str, Any]) -> dict[str, Any]:
    live = canonical.get('live_stats') if isinstance(canonical.get('live_stats'), dict) else {}
    score = canonical.get('score') if isinstance(canonical.get('score'), dict) else {}
    sides: dict[str, Any] = {}
    total_present = 0
    total_critical = 0
    total_warnings = 0
    for side in ('home', 'away'):
        row = live.get(side) if isinstance(live.get(side), dict) else {}
        actual = to_number(score.get(side))
        present = [name for name in ('FGA','FGM','FTA','FTM','ORB','TO','Poss','eFG') if to_number(row.get(name)) is not None]
        total_present += len(present)
        critical: list[str] = []
        warnings: list[str] = []
        values = {key: to_number(row.get(key)) for key in ('FGA','FGM','2PA','2PM','3PA','3PM','FTA','FTM','ORB','TO','Poss','eFG')}
        for made, attempts in (('FGM','FGA'), ('2PM','2PA'), ('3PM','3PA'), ('FTM','FTA')):
            if values[made] is not None and values[attempts] is not None and float(values[made]) > float(values[attempts]) + 0.01:
                critical.append(f'{made}_GT_{attempts}')
        if values['FGA'] is not None and values['2PA'] is not None and values['3PA'] is not None:
            if abs(float(values['FGA']) - float(values['2PA']) - float(values['3PA'])) > 1.01:
                critical.append('FGA_NE_2PA_PLUS_3PA')
        if values['FGM'] is not None and values['2PM'] is not None and values['3PM'] is not None:
            if abs(float(values['FGM']) - float(values['2PM']) - float(values['3PM'])) > 1.01:
                warnings.append('FGM_NE_2PM_PLUS_3PM')
        reconstructed = None
        if values['2PM'] is not None and values['3PM'] is not None and values['FTM'] is not None:
            reconstructed = 2.0 * float(values['2PM']) + 3.0 * float(values['3PM']) + float(values['FTM'])
            if actual is not None and abs(reconstructed - float(actual)) > max(4.0, 0.20 * max(1.0, float(actual))):
                warnings.append('BOX_SCORE_POINTS_DIFFER_FROM_LIVE_SCORE')
        if values['eFG'] is not None and not (0.0 <= float(values['eFG']) <= 1.50):
            critical.append('EFG_OUT_OF_RANGE')
        if values['Poss'] is not None and float(values['Poss']) <= 0:
            critical.append('POSS_NON_POSITIVE')
        total_critical += len(critical)
        total_warnings += len(warnings)
        sides[side] = {
            'present_metrics': present,
            'actual_points': actual,
            'reconstructed_points': reconstructed,
            'critical_errors': critical,
            'warnings': warnings,
        }
    data_present = total_present >= 6
    if not data_present:
        status = 'OFF'
    elif total_critical == 0 and total_warnings <= 1:
        status = 'ON'
    else:
        status = 'LIMITED_INCONSISTENT'
    return {
        'data_present': data_present,
        'status': status,
        'trusted_directional_gate': status == 'ON',
        'critical_error_count': total_critical,
        'warning_count': total_warnings,
        'sides': sides,
    }


# Make data-mode classification quality-aware: present but incoherent statistics
# are PARTIAL/LIMITED, not OFF and not falsely FULL_STAT.
_V124_STAT_CHANNELS_BASE = _v9_stat_channels


def _v9_stat_channels(canonical: dict[str, Any]) -> dict[str, Any]:
    result = _V124_STAT_CHANNELS_BASE(canonical)
    integrity = _v124_stat_integrity(canonical)
    result['integrity'] = integrity
    if result.get('data_mode') == 'FULL_STAT' and not integrity.get('trusted_directional_gate'):
        result['data_mode'] = 'PARTIAL_STAT'
        result['stat_support'] = 'LIMITED'
        result['quality_downgrade_reason'] = 'LIVE_STATS_PRESENT_BUT_INTERNALLY_INCONSISTENT'
    return result


_V124_ADAPT_BASE = adapt_match


def adapt_match(source: dict[str, Any], config: dict[str, Any], strict: bool=False) -> dict[str, Any]:
    canonical = _V124_ADAPT_BASE(source, config, strict)
    context = source.get('analysis_context') if isinstance(source.get('analysis_context'), dict) else {}
    cp = to_int(first(context, ['advisor_checkpoint', 'checkpoint_5stage']))
    if cp not in V124_CHECKPOINT_NAMES:
        cp = _v124_checkpoint_from_source(source)
    if cp in V124_CHECKPOINT_NAMES:
        expected = _v124_checkpoint_expected(cp, int(canonical.get('quarter_seconds') or 600))
        actual = int(canonical.get('elapsed_game_seconds') or 0)
        delta = actual - int(expected['elapsed_seconds'])
        canonical['advisor_checkpoint'] = cp
        canonical['advisor_checkpoint_name'] = V124_CHECKPOINT_NAMES[cp]
        canonical['legacy_trigger_checkpoint'] = to_int(context.get('legacy_trigger_checkpoint'))
        checkpoint_gate = {
            **expected,
            'actual_elapsed_seconds': actual,
            'delta_seconds': delta,
            'within_tolerance': abs(delta) <= V124_CHECKPOINT_TOLERANCE_SECONDS,
            'exact_provider_time_authoritative': True,
            'time_synthesized': bool(context.get('checkpoint_time_synthesized')),
            'synthesized_fields': list(context.get('checkpoint_synthesized_fields') or []),
            'mismatch_is_warning_not_line_blocker': True,
        }
        gate = canonical.setdefault('data_gate', {})
        gate['advisor_checkpoint'] = cp
        gate['advisor_checkpoint_name'] = V124_CHECKPOINT_NAMES[cp]
        gate['checkpoint_semantics'] = checkpoint_gate
        if not checkpoint_gate['within_tolerance']:
            gate.setdefault('warnings', []).append('CHECKPOINT_TIME_DIFFERS_FROM_PROVIDER_TIME_SOURCE_TIME_USED')
    integrity = _v124_stat_integrity(canonical)
    classification = _v9_stat_channels(canonical)
    canonical['data_mode'] = classification['data_mode']
    canonical['stat_support'] = classification['stat_support']
    canonical['stat_channels'] = classification['groups']
    gate = canonical.setdefault('data_gate', {})
    gate['live_stats_present'] = integrity['data_present']
    gate['live_stat_integrity'] = integrity
    gate['data_mode'] = classification['data_mode']
    gate['stat_support'] = classification['stat_support']
    gate['stats_found'] = integrity['data_present']
    return canonical


def _v124_parser_stat_profile(canonical: dict[str, Any], side: str) -> dict[str, Any]:
    blocks = canonical.get('parser_blocks') if isinstance(canonical.get('parser_blocks'), dict) else {}
    alignment = blocks.get('stat_alignment') if isinstance(blocks.get('stat_alignment'), dict) else {}
    profile = alignment.get('live_stat_support') if isinstance(alignment.get('live_stat_support'), dict) else {}
    over = 0
    under = 0
    details: list[dict[str, Any]] = []
    for key in ('fga_poss', 'efg_2p_3p', 'fta_fouls', 'orb_to_extraposs', 'pace'):
        value = str(profile.get(key) or '').upper()
        direction = 'NEUTRAL'
        if value in {'HIGH', 'VERY_HIGH', 'REAL_HIGH'}:
            direction = 'OVER'; over += 1
        elif value in {'LOW', 'VERY_LOW', 'DRY'}:
            direction = 'UNDER'; under += 1
        details.append({'channel': key, 'value': value or None, 'direction': direction})
    candidate = under if side == 'UNDER' else over
    opposite = over if side == 'UNDER' else under
    direction = 'OVER' if over >= 3 and over > under else 'UNDER' if under >= 3 and under > over else 'MIXED'
    return {
        'available': bool(profile),
        'source_verdict': profile.get('verdict'),
        'over_confirmations': over,
        'under_confirmations': under,
        'candidate_confirmations': candidate,
        'opposite_confirmations': opposite,
        'direction': direction,
        'channels': details,
    }


_V124_STAT_GATE_BASE = calculate_stat_gate


def calculate_stat_gate(
    market: dict[str, Any],
    canonical: dict[str, Any],
    zones_data: Optional[dict[str, Any]],
    *,
    project_counts_to_scope_end: bool=True,
) -> dict[str, Any]:
    result = _V124_STAT_GATE_BASE(
        market, canonical, zones_data,
        project_counts_to_scope_end=project_counts_to_scope_end,
    )
    integrity = _v124_stat_integrity(canonical)
    profile = _v124_parser_stat_profile(canonical, str(market.get('side') or '').upper())
    result['stat_data_present'] = integrity['data_present']
    result['stat_integrity'] = integrity
    result['parser_live_stat_profile'] = profile
    result['stat_gate_status_raw_before_integrity_guard'] = result.get('stat_gate_status')
    if integrity['data_present'] and not integrity['trusted_directional_gate']:
        result['stat_support'] = 'LIMITED'
        result['data_mode'] = 'PARTIAL_STAT'
        # Do not convert internally inconsistent box-score counters into a hard
        # AGAINST/CONFIRMED verdict. They remain visible and contribute as LIMITED.
        if str(result.get('stat_gate_status') or '').upper() in {'AGAINST', 'CONFIRMED', 'CONFLICT'}:
            result['stat_gate_status'] = 'LIMITED'
        result['partial_independent_confirmations'] = max(
            int(result.get('partial_independent_confirmations') or 0),
            int(profile.get('candidate_confirmations') or 0),
        )
        result['partial_opposite_confirmations'] = max(
            int(result.get('partial_opposite_confirmations') or 0),
            int(profile.get('opposite_confirmations') or 0),
        )
        result['integrity_guard_applied'] = True
    else:
        result['integrity_guard_applied'] = False
    return result


def _v124_parser_projection_values(market: dict[str, Any], canonical: dict[str, Any], current_points: float) -> dict[str, Any]:
    blocks = canonical.get('parser_blocks') if isinstance(canonical.get('parser_blocks'), dict) else {}
    projections = blocks.get('projections') if isinstance(blocks.get('projections'), dict) else {}
    conditioned = blocks.get('stat_conditioned_line_profiles') if isinstance(blocks.get('stat_conditioned_line_profiles'), dict) else {}
    live_meta = conditioned.get('live_calibrated') if isinstance(conditioned.get('live_calibrated'), dict) else {}
    parser_minute = to_number(live_meta.get('min_played'))
    current_minute = float(canonical.get('elapsed_game_seconds') or 0) / 60.0
    stale_seconds = abs(float(parser_minute) - current_minute) * 60.0 if parser_minute is not None else None
    fresh = stale_seconds is None or stale_seconds <= V124_PARSER_PROJECTION_STALE_SECONDS
    side_key = 'home' if market.get('team') == canonical.get('home_team') else 'away' if market.get('team') else None
    market_type = str(market.get('market_type') or '')
    segment = str(market.get('segment') or 'MATCH').upper()
    values: dict[str, Any] = {}

    def put(name: str, value: Any, source: str) -> None:
        number = to_number(value)
        valid = number is not None and float(number) + 0.01 >= float(current_points) and fresh
        values[name] = {
            'value': float(number) if number is not None else None,
            'valid': valid,
            'source': source,
            'fresh': fresh,
            'stale_seconds': stale_seconds,
            'reason': None if valid else 'STALE' if not fresh else 'BELOW_CURRENT_SCOPE_POINTS' if number is not None else 'MISSING',
        }

    if segment == 'MATCH':
        live = projections.get('live_calibrated') if isinstance(projections.get('live_calibrated'), dict) else {}
        seg = projections.get('segment_projection') if isinstance(projections.get('segment_projection'), dict) else {}
        pre = projections.get('pre_match_stat') if isinstance(projections.get('pre_match_stat'), dict) else {}
        key = f'{side_key}_final' if side_key else 'total'
        put('parser_live_calibrated', live.get(key), f'projections.live_calibrated.{key}')
        put('parser_segment_projection', seg.get(key), f'projections.segment_projection.{key}')
        put('parser_pre_match_stat', pre.get(key), f'projections.pre_match_stat.{key}')
    else:
        quarter_keys = {
            'Q1': ['q1'], 'Q2': ['q2'], 'Q3': ['q3'], 'Q4': ['q4'],
            'H1': ['q1', 'q2'], 'H2': ['q3', 'q4'],
        }.get(segment, [])
        components = []
        for qkey in quarter_keys:
            row = projections.get(qkey) if isinstance(projections.get(qkey), dict) else {}
            key = ('team_a_center' if side_key == 'home' else 'team_b_center' if side_key == 'away' else 'total_center')
            number = to_number(row.get(key))
            if number is None:
                components = []
                break
            components.append(float(number))
        put('parser_segment_centers', sum(components) if components else None, f'projections.{"+".join(quarter_keys)}')
    values['projection_priority'] = deepcopy(projections.get('projection_priority') or {})
    return values


_V124_LIVE_PROJECTION_BASE = calculate_live_projection


def calculate_live_projection(
    market: dict[str, Any],
    canonical: dict[str, Any],
    history: dict[str, Any],
    scenario: dict[str, Any],
    config: dict[str, Any],
    stat: Optional[dict[str, Any]]=None,
) -> dict[str, Any]:
    result = _V124_LIVE_PROJECTION_BASE(market, canonical, history, scenario, config, stat)
    current_points = float(to_number(result.get('current_points')) or 0.0)
    parser = _v124_parser_projection_values(market, canonical, current_points)
    valid = [(name, float(row['value'])) for name, row in parser.items() if isinstance(row, dict) and row.get('valid')]
    core = to_number(result.get('projection_used'))
    result['parser_projection_preservation'] = parser
    if not valid:
        result['parser_projection_used'] = False
        return result

    priority = parser.get('projection_priority') if isinstance(parser.get('projection_priority'), dict) else {}
    main = str(priority.get('main_for_p_live') or '').lower()
    weighted: list[tuple[str, float, float]] = []
    if core is not None:
        weighted.append(('core_live_engine', float(core), 0.40))
    for name, value in valid:
        if name == 'parser_segment_projection':
            weight = 0.35 if 'segment' in main else 0.28
        elif name == 'parser_live_calibrated':
            weight = 0.30 if 'live_calibrated' in main else 0.24
        elif name == 'parser_segment_centers':
            weight = 0.32
        else:
            weight = 0.12 if canonical.get('stage') == 'PRE_MATCH' else 0.08
        weighted.append((name, value, weight))
    total_weight = sum(weight for _, _, weight in weighted)
    if total_weight <= 0:
        return result
    projection = sum(value * weight for _, value, weight in weighted) / total_weight
    projection = max(current_points, projection)
    _v114_recalculate_live_probability(result, market, projection)
    result['projection_parser_blended'] = projection
    result['parser_projection_used'] = True
    result['parser_projection_components_used'] = [
        {'name': name, 'value': value, 'weight': weight / total_weight}
        for name, value, weight in weighted
    ]
    result['projection_formula_mode'] = str(result.get('projection_formula_mode') or '') + '+PARSER_PROJECTION_PRESERVED'
    result['projection_formula'] = (
        str(result.get('projection_formula') or '')
        + ' Parser live_calibrated/segment/quarter projections are retained when fresh and scope-consistent.'
    ).strip()
    components = result.setdefault('components', {})
    for name, row in parser.items():
        if not isinstance(row, dict) or 'value' not in row:
            continue
        components[f'projection_{name}'] = {
            'value': row.get('value'),
            'included': bool(row.get('valid')),
            'role': 'fresh_scope_consistent_parser_projection',
            'exclusion_reason': row.get('reason'),
        }
    return result


# Checkpoint-aware quarter validation. Missing provider time can be recovered from
# the checkpoint fallback; an exact provider time always remains authoritative.
_V124_CURRENT_QUARTER_ISSUE_BASE = _current_quarter_issue


def _current_quarter_issue(market_type: str, segment: str, canonical: dict[str, Any]) -> Optional[str]:
    issue = _V124_CURRENT_QUARTER_ISSUE_BASE(market_type, segment, canonical)
    if market_type not in {'CURRENT_QUARTER_TOTAL', 'CURRENT_QUARTER_TEAM_IT'}:
        return issue
    cp = to_int(canonical.get('advisor_checkpoint'))
    target = to_int(segment[1:]) if segment.startswith('Q') else None
    expected_target = {2: 1, 3: 2, 5: 4}.get(cp)
    if expected_target is not None and target == expected_target:
        if canonical.get('clock') is not None and int(canonical.get('elapsed_game_seconds') or 0) < int(canonical.get('full_game_seconds') or 2400):
            return None
    return issue


_V124_SUPPORTED_BASE = _v112_market_is_currently_supported


def _v112_market_is_currently_supported(item: dict[str, Any]) -> bool:
    if _V124_SUPPORTED_BASE(item):
        return True
    issues = set(str(code) for code in item.get('parser_issues') or [])
    recoverable = {'NO_EXACT_CURRENT_QUARTER_TIME', 'NO_CURRENT_QUARTER_SCORE', 'NO_CURRENT_QUARTER'}
    nonrecoverable = issues - recoverable
    live = item.get('live') if isinstance(item.get('live'), dict) else {}
    router = item.get('router') if isinstance(item.get('router'), dict) else {}
    return bool(
        issues and not nonrecoverable
        and item.get('line') is not None
        and (live.get('clock') is not None or to_number(live.get('elapsed_game_seconds')) is not None)
        and not (str(router.get('status') or '').upper() == 'BLOCK' and bool(router.get('hard_block', True)))
    )


# Add checkpoint metadata to calculator output before Telegram formatting.
_V124_CALCULATE_BASE = SuperBasketCalculator.calculate


def _v124_calculate(self: SuperBasketCalculator, source: dict[str, Any], dispatch_threshold: Optional[float]=None, strict_schema: bool=False) -> dict[str, Any]:
    output = _V124_CALCULATE_BASE(self, source, dispatch_threshold, strict_schema)
    calculation = output.get('super_basket_calculation') or {}
    snapshot = calculation.get('canonical_snapshot') or {}
    context = source.get('analysis_context') if isinstance(source.get('analysis_context'), dict) else {}
    cp = to_int(first(context, ['advisor_checkpoint', 'checkpoint_5stage']))
    if cp in V124_CHECKPOINT_NAMES:
        snapshot['advisor_checkpoint'] = cp
        snapshot['advisor_checkpoint_name'] = V124_CHECKPOINT_NAMES[cp]
        snapshot['legacy_trigger_checkpoint'] = snapshot.get('trigger_checkpoint')
        calculation.setdefault('data_gate', {})['advisor_checkpoint'] = cp
        calculation['data_gate']['advisor_checkpoint_name'] = V124_CHECKPOINT_NAMES[cp]
    return output


SuperBasketCalculator.calculate = _v124_calculate


# Accurate Telegram wording: distinguish "stats absent" from "stats present but
# partial/inconsistent" and from "stats support the opposite side".
_V124_RISK_TEXT_BASE = _v12_risk_text


def _v124_stat_text(item: dict[str, Any]) -> Optional[str]:
    stat = item.get('stat_comparison') if isinstance(item.get('stat_comparison'), dict) else {}
    if not stat:
        return None
    integrity = stat.get('stat_integrity') if isinstance(stat.get('stat_integrity'), dict) else {}
    support = str(stat.get('stat_support') or '').upper()
    status = str(stat.get('stat_gate_status') or '').upper()
    over = int(stat.get('over_gate_score') or 0)
    under = int(stat.get('under_gate_score') or 0)
    side = str(item.get('side') or '').upper()
    if integrity.get('data_present') and not integrity.get('trusted_directional_gate'):
        profile = stat.get('parser_live_stat_profile') if isinstance(stat.get('parser_live_stat_profile'), dict) else {}
        return (
            'live-стата є й врахована, але має внутрішні неузгодженості, тому використана як LIMITED; '
            f'parser-профіль: {profile.get("direction") or "MIXED"}'
        )
    if support in {'OFF', 'N/A_NO_STATS', 'N/A'}:
        return 'live-статистика відсутня'
    if status == 'AGAINST':
        return f'live-статистика врахована, але підтримує протилежний бік: OVER {over}/5, UNDER {under}/5'
    if status == 'CONFIRMED':
        return f'live-статистика підтверджує {side}: OVER {over}/5, UNDER {under}/5'
    if status in {'LIMITED', 'NEUTRAL', 'CONFLICT'}:
        return f'live-статистика врахована частково/змішано: OVER {over}/5, UNDER {under}/5'
    return None


def _v12_risk_text(item: dict[str, Any]) -> str:
    base = _V124_RISK_TEXT_BASE(item)
    stat_text = _v124_stat_text(item)
    if stat_text is None:
        return base
    # Remove the old ambiguous phrases and prepend the precise status.
    parts = [part.strip() for part in str(base).split(';') if part.strip()]
    ambiguous = {
        'немає повного підтвердження live-статою',
        'live-статистика не підтверджує модельний напрямок',
        'live-статистика суперечить цій стороні',
    }
    parts = [part for part in parts if part not in ambiguous]
    return '; '.join([stat_text] + parts[:2])


_V124_REASON_TEXT_BASE = _v12_reason_text


def _v12_reason_text(item: dict[str, Any]) -> str:
    base = _V124_REASON_TEXT_BASE(item)
    stat_text = _v124_stat_text(item)
    return base if not stat_text else f'{base} Стата: {stat_text}.'


_V124_BUILD_MESSAGES_BASE = _v12_build_messages


def _v12_build_messages(advisor: dict[str, Any], calculation: dict[str, Any]) -> list[str]:
    messages = _V124_BUILD_MESSAGES_BASE(advisor, calculation)
    snapshot = calculation.get('canonical_snapshot') or {}
    cp = to_int(snapshot.get('advisor_checkpoint'))
    if cp not in V124_CHECKPOINT_NAMES:
        return messages
    checkpoint_line = f'<b>Чекпоінт:</b> {cp}/5 {html.escape(V124_CHECKPOINT_NAMES[cp])}'
    output = []
    for message in messages:
        token = '<b>Стадія:</b>'
        output.append(message.replace(token, checkpoint_line + '\n' + token, 1) if token in message else checkpoint_line + '\n' + message)
    return output


# Tolerant state gate: checkpoint-vs-provider time differences are audited but do
# not discard lines/projections. Finished games and unresolved score/time source
# conflicts remain hard stops.
_V124_INPUT_GATE_BASE = _v112_input_state_gate


def _v112_input_state_gate(source: dict[str, Any], canonical: dict[str, Any], checkpoint: Optional[int]) -> dict[str, Any]:
    result = _V124_INPUT_GATE_BASE(source, canonical, None)
    cp = to_int(canonical.get('advisor_checkpoint'))
    if cp in V124_CHECKPOINT_NAMES:
        expected = _v124_checkpoint_expected(cp, int(canonical.get('quarter_seconds') or 600))
        actual = int(canonical.get('elapsed_game_seconds') or 0)
        delta = actual - int(expected['elapsed_seconds'])
        result.update({
            'advisor_checkpoint': cp,
            'advisor_checkpoint_name': V124_CHECKPOINT_NAMES[cp],
            'expected_elapsed_seconds': expected['elapsed_seconds'],
            'actual_elapsed_seconds': actual,
            'checkpoint_delta_seconds': delta,
            'checkpoint_within_tolerance': abs(delta) <= V124_CHECKPOINT_TOLERANCE_SECONDS,
            'checkpoint_mismatch_policy': 'WARNING_USE_PROVIDER_TIME',
        })
        if not result.get('finished') and not bool((canonical.get('data_gate') or {}).get('score_conflict')) and not bool((canonical.get('data_gate') or {}).get('time_conflict')):
            result['allowed'] = True
            result['reason'] = 'OK_PROVIDER_TIME_AUTHORITATIVE' if abs(delta) > V124_CHECKPOINT_TOLERANCE_SECONDS else 'OK'
    return result


# Public 1..5 process wrapper. Internally, v12.3 receives no legacy CLI checkpoint;
# the prepared source carries the correct legacy completed-quarter trigger only
# where scenario math needs it (Q2 early/HT/Q4 confirmation).
_V124_PROCESS_BASE = process_vps_match_file


def process_vps_match_file(
    match_path: str | Path,
    *,
    output_path: str | Path | None=None,
    zones_path: str | Path | None=None,
    db_path: str | Path='super_basket.sqlite3',
    mode: str='ACTION',
    require_gpt: bool=False,
    enable_gpt: bool=False,
    enable_telegram: bool=True,
    dry_run: bool=False,
    strict_schema: bool=False,
    checkpoint: Optional[int]=None,
    gpt_reviewer: Optional[Callable[[dict[str, Any], dict[str, Any]], dict[str, Any]]]=None,
    telegram_sender: Optional[Callable[[str], dict[str, Any]]]=None,
) -> dict[str, Any]:
    global _V124_RUNTIME_PATH, _V124_RUNTIME_CHECKPOINT, _V124_BASE_PROCESS_ACTIVE, _V124_LAST_CHECKPOINT
    source_path = Path(match_path).expanduser().resolve()
    raw_source = _V124_LOAD_JSON_BASE(source_path)
    cp = checkpoint if checkpoint is not None else _V124_CLI_FORCED_CHECKPOINT
    if cp is None:
        cp = _v124_checkpoint_from_source(raw_source, source_path)
    if cp is not None:
        cp = int(cp)
        if cp not in V124_CHECKPOINT_NAMES:
            raise ValueError('checkpoint must be 1..5: 1=prematch, 2=early_live_q1, 3=early_live_q2, 4=ht, 5=q4_confirmation')
    _V124_RUNTIME_PATH = source_path
    _V124_RUNTIME_CHECKPOINT = cp
    _V124_BASE_PROCESS_ACTIVE = True
    _V124_LAST_CHECKPOINT = cp
    try:
        result = _V124_PROCESS_BASE(
            source_path,
            output_path=output_path,
            zones_path=zones_path,
            db_path=db_path,
            mode=mode,
            require_gpt=require_gpt,
            enable_gpt=enable_gpt,
            enable_telegram=enable_telegram,
            dry_run=dry_run,
            strict_schema=strict_schema,
            checkpoint=None,
            gpt_reviewer=gpt_reviewer,
            telegram_sender=telegram_sender,
        )
    finally:
        _V124_BASE_PROCESS_ACTIVE = False
        _V124_RUNTIME_PATH = None
        _V124_RUNTIME_CHECKPOINT = None

    calculation = result.get('super_basket_calculation') or {}
    snapshot = calculation.get('canonical_snapshot') or {}
    if cp in V124_CHECKPOINT_NAMES:
        snapshot['legacy_trigger_checkpoint'] = snapshot.get('trigger_checkpoint')
        snapshot['advisor_checkpoint'] = cp
        snapshot['advisor_checkpoint_name'] = V124_CHECKPOINT_NAMES[cp]
        calculation.setdefault('data_gate', {})['advisor_checkpoint'] = cp
        calculation['data_gate']['advisor_checkpoint_name'] = V124_CHECKPOINT_NAMES[cp]
        system = result.get('super_basket_system') or {}
        system['version'] = ADVISOR_VERSION
        system['advisor_checkpoint'] = cp
        system['advisor_checkpoint_name'] = V124_CHECKPOINT_NAMES[cp]
        if isinstance(system.get('advisor'), dict):
            system['advisor']['checkpoint'] = cp
            system['advisor']['checkpoint_name'] = V124_CHECKPOINT_NAMES[cp]
        if isinstance(system.get('decision'), dict):
            system['decision']['advisor_checkpoint'] = cp
            system['decision']['advisor_checkpoint_name'] = V124_CHECKPOINT_NAMES[cp]
    target = Path(output_path).expanduser().resolve() if output_path else source_path.with_name(source_path.stem + '_advisor_result.json')
    save_json(target, result)
    return result


# CLI compatibility: accept --checkpoint 1..5 even though the embedded legacy
# argparse parser still documents 1..3. The option is stripped before delegating;
# the public wrapper above receives it via the runtime bridge.
_V124_CLI_BASE = _single_file_cli


def _single_file_cli(argv: list[str] | None=None) -> int:
    global _V124_CLI_FORCED_CHECKPOINT
    args = list(sys.argv[1:] if argv is None else argv)
    forced = None
    if '--checkpoint' in args:
        index = args.index('--checkpoint')
        if index + 1 >= len(args):
            raise SystemExit('--checkpoint requires 1..5')
        try:
            forced = int(args[index + 1])
        except ValueError as error:
            raise SystemExit('--checkpoint requires an integer 1..5') from error
        if forced not in V124_CHECKPOINT_NAMES:
            raise SystemExit('--checkpoint must be 1..5')
        del args[index:index + 2]
    _V124_CLI_FORCED_CHECKPOINT = forced
    try:
        return _V124_CLI_BASE(args)
    finally:
        _V124_CLI_FORCED_CHECKPOINT = None


DEFAULT_CONFIG.setdefault('v12_advisor_policy', {}).update({
    'version': ADVISOR_VERSION,
    'five_checkpoints': deepcopy(V124_CHECKPOINT_NAMES),
    'provider_time_authoritative': True,
    'checkpoint_time_mismatch_is_warning': True,
    'preserve_fresh_parser_live_projection': True,
    'stat_presence_distinct_from_stat_direction_confirmation': True,
    'inconsistent_live_stats_downgrade_to_partial': True,
})



# ===== v12.4.1 directional stat-profile and prematch cleanup =====
# Prematch never treats leftover raw box-score fields as live data. When the
# parser's aggregate live-stat profile is fresh, it resolves the direction of an
# otherwise internally inconsistent raw box score instead of calling stats absent.

_V1241_STAT_INTEGRITY_BASE = _v124_stat_integrity


def _v124_stat_integrity(canonical: dict[str, Any]) -> dict[str, Any]:
    if str(canonical.get('stage') or '').upper() == 'PRE_MATCH' or to_int(canonical.get('advisor_checkpoint')) == 1:
        return {
            'data_present': False,
            'status': 'PREMATCH_NOT_APPLICABLE',
            'trusted_directional_gate': False,
            'critical_error_count': 0,
            'warning_count': 0,
            'sides': {},
            'note': 'Live box score is not applicable before tip-off; pre-match stat projection remains available separately.',
        }
    return _V1241_STAT_INTEGRITY_BASE(canonical)


_V1241_STAT_CHANNELS_BASE = _v9_stat_channels


def _v9_stat_channels(canonical: dict[str, Any]) -> dict[str, Any]:
    result = _V1241_STAT_CHANNELS_BASE(canonical)
    if str(canonical.get('stage') or '').upper() == 'PRE_MATCH' or to_int(canonical.get('advisor_checkpoint')) == 1:
        result['data_mode'] = 'SCORE_TIME_HISTORY'
        result['stat_support'] = 'N/A_NO_STATS'
        result['quality_downgrade_reason'] = 'PREMATCH_LIVE_STATS_NOT_APPLICABLE'
    return result


_V1241_PARSER_STAT_PROFILE_BASE = _v124_parser_stat_profile


def _v124_parser_stat_profile(canonical: dict[str, Any], side: str) -> dict[str, Any]:
    profile = _V1241_PARSER_STAT_PROFILE_BASE(canonical, side)
    if str(canonical.get('stage') or '').upper() == 'PRE_MATCH' or to_int(canonical.get('advisor_checkpoint')) == 1:
        return {**profile, 'available': False, 'fresh': False, 'reason': 'PREMATCH_NOT_APPLICABLE'}
    blocks = canonical.get('parser_blocks') if isinstance(canonical.get('parser_blocks'), dict) else {}
    conditioned = blocks.get('stat_conditioned_line_profiles') if isinstance(blocks.get('stat_conditioned_line_profiles'), dict) else {}
    live_meta = conditioned.get('live_calibrated') if isinstance(conditioned.get('live_calibrated'), dict) else {}
    parser_minute = to_number(live_meta.get('min_played'))
    current_minute = float(canonical.get('elapsed_game_seconds') or 0) / 60.0
    stale_seconds = abs(float(parser_minute) - current_minute) * 60.0 if parser_minute is not None else None
    fresh = stale_seconds is None or stale_seconds <= V124_PARSER_PROJECTION_STALE_SECONDS
    profile['fresh'] = fresh
    profile['stale_seconds'] = stale_seconds
    if not fresh:
        profile['available'] = False
        profile['reason'] = 'PARSER_STAT_PROFILE_STALE'
    return profile


_V1241_STAT_GATE_BASE = calculate_stat_gate


def calculate_stat_gate(
    market: dict[str, Any],
    canonical: dict[str, Any],
    zones_data: Optional[dict[str, Any]],
    *,
    project_counts_to_scope_end: bool=True,
) -> dict[str, Any]:
    result = _V1241_STAT_GATE_BASE(
        market, canonical, zones_data,
        project_counts_to_scope_end=project_counts_to_scope_end,
    )
    integrity = result.get('stat_integrity') if isinstance(result.get('stat_integrity'), dict) else _v124_stat_integrity(canonical)
    profile = _v124_parser_stat_profile(canonical, str(market.get('side') or '').upper())
    result['parser_live_stat_profile'] = profile
    if integrity.get('data_present') and not integrity.get('trusted_directional_gate') and profile.get('available'):
        candidate = int(profile.get('candidate_confirmations') or 0)
        opposite = int(profile.get('opposite_confirmations') or 0)
        result['partial_independent_confirmations'] = candidate
        result['partial_opposite_confirmations'] = opposite
        if candidate >= 3 and candidate > opposite:
            result['stat_gate_status'] = 'CONFIRMED'
            result['limited_confirmation_source'] = 'FRESH_PARSER_LIVE_STAT_PROFILE'
        elif opposite >= 3 and opposite > candidate:
            result['stat_gate_status'] = 'AGAINST'
            result['limited_confirmation_source'] = 'FRESH_PARSER_LIVE_STAT_PROFILE_OPPOSITE'
        else:
            result['stat_gate_status'] = 'LIMITED'
            result['limited_confirmation_source'] = 'MIXED_PARSER_LIVE_STAT_PROFILE'
    return result


# Parser projection freshness is source-specific: live_calibrated and live segment
# projections require a synchronized snapshot; pre-match projections and static
# quarter centers remain usable as structural/model components.
def _v124_parser_projection_values(market: dict[str, Any], canonical: dict[str, Any], current_points: float) -> dict[str, Any]:
    blocks = canonical.get('parser_blocks') if isinstance(canonical.get('parser_blocks'), dict) else {}
    projections = blocks.get('projections') if isinstance(blocks.get('projections'), dict) else {}
    conditioned = blocks.get('stat_conditioned_line_profiles') if isinstance(blocks.get('stat_conditioned_line_profiles'), dict) else {}
    live_meta = conditioned.get('live_calibrated') if isinstance(conditioned.get('live_calibrated'), dict) else {}
    parser_minute = to_number(live_meta.get('min_played'))
    current_minute = float(canonical.get('elapsed_game_seconds') or 0) / 60.0
    stale_seconds = abs(float(parser_minute) - current_minute) * 60.0 if parser_minute is not None else None
    live_fresh = stale_seconds is None or stale_seconds <= V124_PARSER_PROJECTION_STALE_SECONDS
    side_key = 'home' if market.get('team') == canonical.get('home_team') else 'away' if market.get('team') else None
    segment = str(market.get('segment') or 'MATCH').upper()
    values: dict[str, Any] = {}

    def put(name: str, value: Any, source: str, *, requires_live_fresh: bool) -> None:
        number = to_number(value)
        fresh = live_fresh if requires_live_fresh else True
        valid = number is not None and float(number) + 0.01 >= float(current_points) and fresh
        values[name] = {
            'value': float(number) if number is not None else None,
            'valid': valid,
            'source': source,
            'fresh': fresh,
            'stale_seconds': stale_seconds if requires_live_fresh else None,
            'reason': None if valid else 'STALE' if not fresh else 'BELOW_CURRENT_SCOPE_POINTS' if number is not None else 'MISSING',
        }

    if segment == 'MATCH':
        live = projections.get('live_calibrated') if isinstance(projections.get('live_calibrated'), dict) else {}
        seg = projections.get('segment_projection') if isinstance(projections.get('segment_projection'), dict) else {}
        pre = projections.get('pre_match_stat') if isinstance(projections.get('pre_match_stat'), dict) else {}
        key = f'{side_key}_final' if side_key else 'total'
        put('parser_live_calibrated', live.get(key), f'projections.live_calibrated.{key}', requires_live_fresh=True)
        put('parser_segment_projection', seg.get(key), f'projections.segment_projection.{key}', requires_live_fresh=True)
        put('parser_pre_match_stat', pre.get(key), f'projections.pre_match_stat.{key}', requires_live_fresh=False)
    else:
        quarter_keys = {'Q1':['q1'],'Q2':['q2'],'Q3':['q3'],'Q4':['q4'],'H1':['q1','q2'],'H2':['q3','q4']}.get(segment, [])
        components = []
        for qkey in quarter_keys:
            row = projections.get(qkey) if isinstance(projections.get(qkey), dict) else {}
            key = 'team_a_center' if side_key == 'home' else 'team_b_center' if side_key == 'away' else 'total_center'
            number = to_number(row.get(key))
            if number is None:
                components = []
                break
            components.append(float(number))
        put('parser_segment_centers', sum(components) if components else None, f'projections.{"+".join(quarter_keys)}', requires_live_fresh=False)
    values['projection_priority'] = deepcopy(projections.get('projection_priority') or {})
    return values


ADVISOR_VERSION = '12.4.1-FIVE-CHECKPOINTS-STAT-DIRECTION-PROJECTION-FIX'
SYSTEM_VERSION = ADVISOR_VERSION
DEFAULT_CONFIG['engine_version'] = ADVISOR_VERSION
DEFAULT_CONFIG.setdefault('v12_advisor_policy', {}).update({
    'version': ADVISOR_VERSION,
    'prematch_live_boxscore_ignored': True,
    'fresh_parser_stat_profile_resolves_limited_direction': True,
})


# =============================================================================
# v13.1 STRICT VALUE / SAFETY / AUDIT LAYER
# =============================================================================
# Authoritative hierarchy:
#   parser/canonical validation -> exact history -> scenario -> independent live
#   projection -> P_live -> normalized stage weights -> caps/blockers -> P_final
#   -> strict directional gates -> market fair probability / EV -> bankroll stake.
#
# This patch deliberately does NOT claim that P_final is an empirically calibrated
# probability. It exposes P_final as an internal model score and creates a
# conservative probability estimate for price comparison. True calibration must
# be learned from a much larger settled out-of-sample database.

ADVISOR_VERSION = '13.1.0-STRICT-AUDITED-VALUE'
SYSTEM_VERSION = ADVISOR_VERSION
DEFAULT_CONFIG['engine_version'] = ADVISOR_VERSION

V131_MAX_ACTIVE_PER_SNAPSHOT = 1
V131_MAX_DISPLAY = 2
V131_MAX_STAKE_PCT = 1.50
V131_KELLY_FRACTION = 0.25
V131_MIN_EV = 0.02
V131_MIN_PRICE_EDGE = 0.015

_V131_BASE_PROCESS = process_vps_match_file


def _v131_market_key(item: dict[str, Any]) -> tuple[Any, ...]:
    return (
        item.get('source_market_id') or item.get('math_market_key') or item.get('market_id'),
        item.get('market_type'), item.get('segment'), item.get('team'),
        to_number(item.get('line')), item.get('bookmaker'),
    )


def _v131_general_total(item: dict[str, Any]) -> bool:
    return str(item.get('market_type') or '') in {
        'MATCH_TOTAL', 'H1_TOTAL', 'H2_TOTAL', 'CURRENT_QUARTER_TOTAL'
    }


def _v131_market_priority(item: dict[str, Any], canonical: dict[str, Any]) -> float:
    market_type = str(item.get('market_type') or '')
    segment = str(item.get('segment') or '').upper()
    stage = str(canonical.get('stage') or '').upper()
    checkpoint = to_int(canonical.get('advisor_checkpoint') or canonical.get('trigger_checkpoint'))
    if checkpoint == 1 or stage == 'PRE_MATCH':
        table = {'MATCH_TOTAL': 2.5, 'H1_TOTAL': 1.4, 'CURRENT_QUARTER_TOTAL': 1.2,
                 'TEAM_IT_MATCH': 0.9, 'TEAM_IT_H1': 0.75, 'CURRENT_QUARTER_TEAM_IT': 0.6}
    elif checkpoint == 2:
        table = {'CURRENT_QUARTER_TOTAL': 2.3 if segment == 'Q1' else 0.8,
                 'H1_TOTAL': 1.7, 'MATCH_TOTAL': 1.6, 'TEAM_IT_H1': 0.9,
                 'TEAM_IT_MATCH': 0.8, 'CURRENT_QUARTER_TEAM_IT': 0.7}
    elif checkpoint == 3:
        table = {'H1_TOTAL': 2.3, 'MATCH_TOTAL': 1.8,
                 'CURRENT_QUARTER_TOTAL': 1.3 if segment == 'Q2' else 0.8,
                 'TEAM_IT_H1': 0.9, 'TEAM_IT_MATCH': 0.8}
    elif checkpoint == 4 or stage == 'HT':
        table = {'MATCH_TOTAL': 2.5, 'H2_TOTAL': 2.1,
                 'CURRENT_QUARTER_TOTAL': 1.5 if segment == 'Q3' else 0.8,
                 'TEAM_IT_MATCH': 0.9, 'TEAM_IT_H2': 0.85}
    elif checkpoint == 5 or stage in {'AFTER_3Q', 'Q4_CONFIRMATION'}:
        table = {'MATCH_TOTAL': 2.6,
                 'CURRENT_QUARTER_TOTAL': 2.0 if segment == 'Q4' else 0.7,
                 'TEAM_IT_MATCH': 0.9, 'CURRENT_QUARTER_TEAM_IT': 0.8,
                 'H2_TOTAL': 0.6}
    else:
        table = {'MATCH_TOTAL': 1.3, 'H1_TOTAL': 1.1, 'H2_TOTAL': 1.1,
                 'CURRENT_QUARTER_TOTAL': 1.0, 'TEAM_IT_MATCH': 0.8,
                 'TEAM_IT_H1': 0.7, 'TEAM_IT_H2': 0.7,
                 'CURRENT_QUARTER_TEAM_IT': 0.65}
    return float(table.get(market_type, 0.4))


def _v131_edge_minimum(item: dict[str, Any]) -> float:
    return {
        'MATCH_TOTAL': 3.0,
        'H1_TOTAL': 2.5,
        'H2_TOTAL': 2.5,
        'CURRENT_QUARTER_TOTAL': 1.5,
        'TEAM_IT_MATCH': 2.5,
        'TEAM_IT_H1': 2.0,
        'TEAM_IT_H2': 2.0,
        'CURRENT_QUARTER_TEAM_IT': 1.5,
    }.get(str(item.get('market_type') or ''), 2.0)


def _v131_paired_prices(evaluations: list[dict[str, Any]]) -> dict[tuple[Any, ...], dict[str, float]]:
    grouped: dict[tuple[Any, ...], dict[str, float]] = {}
    for item in evaluations:
        if bool(item.get('is_reference_line')):
            continue
        odds = to_number(item.get('odds'))
        side = str(item.get('side') or '').upper()
        if odds is None or odds <= 1.0 or side not in {'OVER', 'UNDER'}:
            continue
        grouped.setdefault(_v131_market_key(item), {})[side] = float(odds)
    result: dict[tuple[Any, ...], dict[str, float]] = {}
    for key, row in grouped.items():
        if 'OVER' not in row or 'UNDER' not in row:
            continue
        q_over = 1.0 / row['OVER']
        q_under = 1.0 / row['UNDER']
        total = q_over + q_under
        if total <= 0:
            continue
        result[key] = {
            'OVER': q_over / total,
            'UNDER': q_under / total,
            'overround': total - 1.0,
        }
    return result


def _v131_history_zone(item: dict[str, Any], canonical: dict[str, Any]) -> dict[str, Any]:
    line = to_number(item.get('line'))
    side = str(item.get('side') or '').upper()
    if line is None or side not in {'OVER', 'UNDER'}:
        return {'available': False, 'reason': 'NO_LINE_OR_SIDE'}
    spec = {'market_type': item.get('market_type'), 'segment': item.get('segment'), 'team': item.get('team')}
    values = [float(v) for v in _v11_history_values(canonical, spec) if v is not None]
    if not values:
        return {'available': False, 'reason': 'NO_SAME_FORMAT_HISTORY', 'n': 0}

    def row_at(candidate: float) -> dict[str, Any]:
        outcomes = [settle(value, candidate, side) for value in values]
        wins = outcomes.count('win')
        losses = outcomes.count('loss')
        pushes = outcomes.count('push')
        n = len(values)
        return {
            'line': float(candidate), 'side': side, 'wins': wins, 'losses': losses,
            'pushes': pushes, 'n': n, 'raw_probability': wins / n if n else None,
            'smoothed_probability': smoothed_probability(wins, n, 1.0, 1.0) if n else None,
        }

    exact = row_at(float(line))
    candidates = []
    start = math.floor(min(values)) - 0.5
    stop = math.ceil(max(values)) + 0.5
    cursor = start
    while cursor <= stop + 1e-9:
        if (side == 'OVER' and cursor <= line + 1e-9) or (side == 'UNDER' and cursor >= line - 1e-9):
            candidates.append(row_at(round(cursor, 1)))
        cursor += 1.0
    strong = [row for row in candidates if float(row.get('smoothed_probability') or 0.0) >= 0.75]
    nearest = min(strong, key=lambda row: abs(float(row['line']) - float(line))) if strong else None
    return {
        'available': True,
        'exact': exact,
        'nearest_75_zone': nearest,
        'n': len(values),
        'method': 'same-format history with Beta(1,1) smoothing',
    }


def _v131_conservative_probability(item: dict[str, Any], zone: dict[str, Any]) -> dict[str, Any]:
    p_final = to_number((item.get('advisor') or {}).get('p_final'))
    if p_final is None:
        p_final = to_number(item.get('p_final_system') or item.get('p_final')) or 0.50
    p_final = max(0.0, min(1.0, float(p_final)))
    mode = str(item.get('data_mode') or (item.get('live') or {}).get('data_mode') or 'NO_STAT').upper()
    base = {'FULL_STAT': 0.62, 'PARTIAL_STAT': 0.50, 'SCORE_TIME_HISTORY': 0.40,
            'NO_STAT': 0.40, 'DATA_OFF': 0.25}.get(mode, 0.35)
    n = int(zone.get('n') or 0)
    sample_factor = max(0.45, min(1.0, math.sqrt(max(1, n) / 70.0)))
    reliability = max(0.15, min(0.65, base * sample_factor))
    estimate = 0.50 + reliability * (p_final - 0.50)
    stat_status = str((item.get('stat_comparison') or {}).get('stat_gate_status') or
                      (item.get('advisor') or {}).get('stat_gate_status') or 'OFF').upper()
    if stat_status == 'NEUTRAL':
        estimate -= 0.005
    elif stat_status in {'OFF', 'N/A', 'NA', 'N_A_NO_STATS'}:
        estimate -= 0.015
    estimate = max(0.02, min(0.98, estimate))
    return {
        'probability': estimate, 'p_final_internal_score': p_final,
        'reliability': reliability, 'sample_factor': sample_factor,
        'method': 'conservative shrinkage; NOT empirical calibration',
    }


def _v131_stat_status(item: dict[str, Any]) -> str:
    return str((item.get('stat_comparison') or {}).get('stat_gate_status') or
               (item.get('advisor') or {}).get('stat_gate_status') or 'OFF').upper()


def _v131_fake_against(item: dict[str, Any]) -> bool:
    stat = item.get('stat_comparison') or {}
    side = str(item.get('side') or '').upper()
    return bool((side == 'OVER' and stat.get('fake_over')) or
                (side == 'UNDER' and stat.get('fake_under')))


def _v131_projection_sanity(item: dict[str, Any], canonical: dict[str, Any]) -> tuple[bool, list[str]]:
    reasons: list[str] = []
    projection = to_number((item.get('live') or {}).get('projection_used'))
    line = to_number(item.get('line'))
    if projection is None or not math.isfinite(float(projection)):
        reasons.append('PROJECTION_MISSING_OR_NONFINITE')
        return False, reasons
    current = to_number((item.get('live') or {}).get('current_points'))
    if current is not None and float(projection) + 1e-9 < float(current):
        reasons.append('PROJECTION_BELOW_CURRENT_POINTS')
    sigma = to_number((item.get('live') or {}).get('sigma'))
    if sigma is not None and (float(sigma) <= 0 or not math.isfinite(float(sigma))):
        reasons.append('SIGMA_INVALID')
    if line is not None and abs(float(projection) - float(line)) > 80:
        reasons.append('PROJECTION_LINE_DISTANCE_IMPLAUSIBLE')
    return not reasons, reasons


def _v131_team_it_gate(item: dict[str, Any]) -> tuple[bool, list[str]]:
    if str(item.get('market_type') or '') not in {
        'TEAM_IT_MATCH', 'TEAM_IT_H1', 'TEAM_IT_H2', 'CURRENT_QUARTER_TEAM_IT'
    }:
        return True, []
    history = item.get('history') or {}
    own = to_number((history.get('own_scored') or {}).get('p_smoothed'))
    allowed = to_number((history.get('opponent_allowed') or {}).get('p_smoothed'))
    reasons: list[str] = []
    if own is None:
        reasons.append('TEAM_IT_OWN_SCORED_MISSING')
    if allowed is None:
        reasons.append('TEAM_IT_OPPONENT_ALLOWED_MISSING')
    if own is not None and own < 0.70:
        reasons.append('TEAM_IT_OWN_SCORED_BELOW_70')
    if allowed is not None and allowed < 0.70:
        reasons.append('TEAM_IT_OPPONENT_ALLOWED_BELOW_70')
    return not reasons, reasons


def _v131_q4_gate(item: dict[str, Any], canonical: dict[str, Any]) -> tuple[bool, list[str]]:
    q4 = item.get('q4_context') or {}
    if not q4.get('applicable'):
        return True, []
    reasons: list[str] = []
    side = str(item.get('side') or '').upper()
    if q4.get('mandatory_missing'):
        reasons.append('Q4_MANDATORY_CONTEXT_MISSING')
    foul = to_number(q4.get('foul_tail_score')) or 0.0
    kill = to_number(q4.get('kill_chase_score')) or 0.0
    dry = to_number(q4.get('dry_score')) or 0.0
    projection = to_number((item.get('live') or {}).get('projection_used'))
    line = to_number(item.get('line'))
    if side == 'UNDER':
        if foul >= 0.70: reasons.append('Q4_UNDER_FOUL_TAIL_HIGH')
        if kill >= 0.65: reasons.append('Q4_UNDER_KILL_CHASE_HIGH')
        if bool(q4.get('bonus_path')): reasons.append('Q4_UNDER_BONUS_PATH')
        if bool(q4.get('three_pa_chase')): reasons.append('Q4_UNDER_3PA_CHASE')
        if bool(q4.get('leader_ft_path')): reasons.append('Q4_UNDER_LEADER_FT_PATH')
        if projection is not None and line is not None and projection >= line:
            reasons.append('Q4_UNDER_PROJECTION_NOT_BELOW_LINE')
        if dry < 0.50: reasons.append('Q4_UNDER_DRY_NOT_SUPPORTIVE')
    elif side == 'OVER':
        if projection is not None and line is not None and projection < line:
            reasons.append('Q4_OVER_PROJECTION_BELOW_LINE')
        if max(foul, kill, to_number(q4.get('volume_score')) or 0.0) < 0.45:
            reasons.append('Q4_OVER_CONTEXT_NOT_SUPPORTIVE')
    return not reasons, reasons


def _v131_hard_gate(item: dict[str, Any], canonical: dict[str, Any], metrics: dict[str, Any]) -> list[str]:
    reasons: list[str] = []
    if bool(item.get('is_reference_line')):
        reasons.append('REFERENCE_LINE_NO_REAL_ODDS')
    router = item.get('router') or {}
    if str(router.get('status') or '').upper() == 'BLOCK' and bool(router.get('hard_block', True)):
        reasons.append('PRODUCTION_ROUTER_BLOCK')
    reasons.extend(str(row.get('rule_id')) for row in item.get('blockers') or []
                   if str(row.get('rule_id') or '') in _V12_FATAL_BLOCKERS)
    stat_status = _v131_stat_status(item)
    if stat_status in {'AGAINST', 'CONFLICT'}:
        reasons.append('STAT_DIRECTION_' + stat_status)
    if _v131_fake_against(item):
        reasons.append('FAKE_PROFILE_AGAINST_CANDIDATE')
    sane, sanity = _v131_projection_sanity(item, canonical)
    if not sane:
        reasons.extend(sanity)
    team_ok, team_reasons = _v131_team_it_gate(item)
    if not team_ok:
        reasons.extend(team_reasons)
    q4_ok, q4_reasons = _v131_q4_gate(item, canonical)
    if not q4_ok:
        reasons.extend(q4_reasons)
    edge = float(metrics.get('line_edge') or -999.0)
    if edge < _v131_edge_minimum(item):
        reasons.append('LIVE_EDGE_BELOW_STRICT_MARKET_MIN')
    zone_prob = to_number(metrics.get('history_exact_probability'))
    history_min = 0.75 if edge < 3.0 else 0.60
    if zone_prob is None or zone_prob < history_min:
        reasons.append('EXACT_HISTORY_BELOW_DYNAMIC_MIN')
    p_final = float(metrics.get('p_final_internal_score') or 0.50)
    if p_final < 0.65:
        reasons.append('P_FINAL_INTERNAL_BELOW_65')
    ev = to_number(metrics.get('ev'))
    if ev is None or ev < V131_MIN_EV:
        reasons.append('EV_BELOW_MINIMUM')
    price_edge = to_number(metrics.get('price_edge'))
    if price_edge is not None and price_edge < V131_MIN_PRICE_EDGE:
        reasons.append('PRICE_EDGE_BELOW_MINIMUM')
    return list(dict.fromkeys(reasons))


def _v131_quarter_kelly(probability: float, odds: float) -> float:
    if odds <= 1.0:
        return 0.0
    b = odds - 1.0
    full = max(0.0, (b * probability - (1.0 - probability)) / b)
    return min(V131_MAX_STAKE_PCT, full * V131_KELLY_FRACTION * 100.0)


def _v131_assess(item: dict[str, Any], canonical: dict[str, Any], price_map: dict[tuple[Any, ...], dict[str, float]]) -> dict[str, Any]:
    zone = _v131_history_zone(item, canonical)
    conservative = _v131_conservative_probability(item, zone)
    p = float(conservative['probability'])
    odds = to_number(item.get('odds'))
    side = str(item.get('side') or '').upper()
    pair = price_map.get(_v131_market_key(item), {})
    fair = to_number(pair.get(side))
    break_even = 1.0 / float(odds) if odds is not None and odds > 1.0 and not bool(item.get('is_reference_line')) else None
    ev = p * float(odds) - 1.0 if break_even is not None else None
    price_edge = p - float(fair) if fair is not None else (p - break_even if break_even is not None else None)
    base = _v12_recommendation_score(item)
    exact_prob = to_number((zone.get('exact') or {}).get('smoothed_probability'))
    metrics = {
        **base,
        **conservative,
        'p_final_internal_score': conservative['p_final_internal_score'],
        'history_zone': zone,
        'history_exact_probability': exact_prob,
        'market_fair_probability': fair,
        'break_even_probability': break_even,
        'ev': ev,
        'price_edge': price_edge,
        'market_priority': _v131_market_priority(item, canonical),
    }
    reasons = _v131_hard_gate(item, canonical, metrics)
    active = not reasons
    if active:
        stake = _v131_quarter_kelly(p, float(odds)) if odds is not None else 0.0
        if stake < 0.10:
            tier, status = 'LEAN', 'LEAN — МАЛА ПЕРЕВАГА'
        elif stake < 0.50:
            tier, status = 'RISK', 'RISK — ОБЕРЕЖНИЙ ВХІД'
        elif stake < 1.00:
            tier, status = 'PLAY', 'PLAY — РОБОЧИЙ VALUE'
        else:
            tier, status = 'STRONG', 'STRONG — НАЙСИЛЬНІШИЙ VALUE'
        stake_min = max(0.10, stake * 0.80)
        stake_text = f'{stake_min:.2f}-{stake:.2f}% від банку'
        action = 'PLAY' if tier in {'PLAY', 'STRONG'} else tier
    else:
        tier, status, action, stake_text = 'FORECAST', 'FORECAST ONLY — СТАВКА 0%', 'FORECAST', '0.00-0.00% від банку'
    rank_score = (
        0.34 * p + 0.16 * float(metrics.get('p_live') or 0.50)
        + 0.12 * float(exact_prob or 0.50)
        + 0.08 * float(metrics.get('p_scenario') or 0.50)
        + 0.08 * max(0.0, min(1.0, 0.5 + float(metrics.get('line_edge') or 0.0) / 20.0))
        + 0.12 * min(1.0, max(0.0, float(ev or 0.0) * 5.0))
        + 0.10 * min(1.0, _v131_market_priority(item, canonical) / 2.6)
    )
    return {
        'item': item, 'metrics': metrics, 'active': active, 'hard_gate_reasons': reasons,
        'tier': tier, 'status': status, 'action': action, 'stake': stake_text,
        'rank_score': rank_score,
    }


def _v131_3x3(canonical: dict[str, Any]) -> bool:
    text = ' '.join(str(canonical.get(key) or '') for key in ('name','tournament','home_team','away_team')).lower()
    return '3x3' in text or '3х3' in text


def _v131_select(pool: list[dict[str, Any]], canonical: dict[str, Any]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    price_map = _v131_paired_prices(pool)
    assessed: list[dict[str, Any]] = []
    rejected: list[dict[str, Any]] = []
    for item in pool:
        if _v131_3x3(canonical) and str(item.get('market_type') or '') != 'MATCH_TOTAL':
            rejected.append({'market': _v11_compact_line(item), 'fatal_reasons': ['3X3_MATCH_TOTAL_ONLY']})
            continue
        fatal = _v12_fatal_reasons(item)
        if not _v112_market_is_currently_supported(item):
            fatal.append('MARKET_NOT_CURRENTLY_SUPPORTED')
        if fatal:
            rejected.append({'market': _v11_compact_line(item), 'fatal_reasons': list(dict.fromkeys(fatal))})
            continue
        assessed.append(_v131_assess(deepcopy(item), canonical, price_map))
    if not assessed:
        return [], rejected

    # One side/offer per logical market.
    best_by_group: dict[tuple[Any, ...], dict[str, Any]] = {}
    for row in assessed:
        item = row['item']
        key = (item.get('market_type'), item.get('segment'), item.get('team'))
        current = best_by_group.get(key)
        candidate_key = (1 if row['active'] else 0, row['rank_score'], float(row['metrics'].get('line_edge') or -999))
        current_key = ((1 if current['active'] else 0), current['rank_score'], float(current['metrics'].get('line_edge') or -999)) if current else (-1,-1,-999)
        if current is None or candidate_key > current_key:
            best_by_group[key] = row
    rows = list(best_by_group.values())
    rows.sort(key=lambda row: (1 if row['active'] else 0, row['rank_score']), reverse=True)

    # Exactly one monetary signal per snapshot. The second line is forecast-only.
    active_rows = [row for row in rows if row['active']]
    selected_rows: list[dict[str, Any]] = []
    if active_rows:
        best_active = max(active_rows, key=lambda row: row['rank_score'])
        selected_rows.append(best_active)
        for row in rows:
            if row is best_active:
                continue
            row = deepcopy(row)
            row['active'] = False
            row['tier'] = 'FORECAST'
            row['status'] = 'FORECAST ONLY — ДРУГИЙ ВАРІАНТ, СТАВКА 0%'
            row['action'] = 'FORECAST'
            row['stake'] = '0.00-0.00% від банку'
            row['hard_gate_reasons'] = list(dict.fromkeys(row['hard_gate_reasons'] + ['MAX_ONE_ACTIVE_BET_PER_SNAPSHOT']))
            selected_rows.append(row)
            break
    else:
        selected_rows = rows[:V131_MAX_DISPLAY]

    output: list[dict[str, Any]] = []
    for row in selected_rows[:V131_MAX_DISPLAY]:
        item = deepcopy(row['item'])
        adv = item.setdefault('advisor', {})
        adv.update({
            'version': ADVISOR_VERSION,
            'action': row['action'], 'status': row['status'],
            'recommendation_tier': row['tier'], 'stake_budget': row['stake'],
            'strict_active': row['active'], 'hard_gate_reasons_v13_1': row['hard_gate_reasons'],
            'ranking_metrics_v13_1': row['metrics'], 'recommendation_score_v13_1': row['rank_score'],
        })
        item['system_action'] = row['action']
        item['system_status'] = row['status']
        item['stake'] = row['stake']
        item['v13_1_rank_score'] = row['rank_score']
        output.append(item)
    return output, rejected


def _v131_pct(value: Any) -> str:
    number = to_number(value)
    return 'N/A' if number is None else f'{float(number)*100:.1f}%'


def _v131_messages(selected: list[dict[str, Any]], calculation: dict[str, Any]) -> list[str]:
    if not selected:
        return []
    snapshot = calculation.get('canonical_snapshot') or {}
    score = snapshot.get('score') or {}
    lines = [
        '<b>🏀 SUPER BASKET v13.1 — STRICT AUDITED</b>',
        f'<b>Матч:</b> {html.escape(str(snapshot.get("name") or "N/A"))}',
        f'<b>Стадія:</b> {html.escape(str(snapshot.get("stage") or "N/A"))} | '
        f'<b>Рахунок:</b> {_v11_num(score.get("home"))}:{_v11_num(score.get("away"))}',
        '<i>P_final нижче показано як внутрішній model score, а не як доведену реальну ймовірність.</i>',
    ]
    for idx, item in enumerate(selected, 1):
        adv = item.get('advisor') or {}
        m = adv.get('ranking_metrics_v13_1') or {}
        zone = m.get('history_zone') or {}
        exact = zone.get('exact') or {}
        near = zone.get('nearest_75_zone') or {}
        reference = bool(item.get('is_reference_line'))
        odds_text = 'N/A' if reference else _v11_num(item.get('odds'), 2)
        fair = _v131_pct(m.get('market_fair_probability'))
        ev = to_number(m.get('ev'))
        price_edge = to_number(m.get('price_edge'))
        reason_text = ', '.join(str(v) for v in adv.get('hard_gate_reasons_v13_1') or []) or 'усі strict gates пройдено'
        history_text = (
            f'{int(exact.get("wins") or 0)}/{int(exact.get("n") or 0)}; '
            f'Bayes {_v131_pct(exact.get("smoothed_probability"))}'
            if zone.get('available') else 'N/A'
        )
        if near:
            history_text += f'; найближча 75%+ зона {item.get("side")} {float(near.get("line")):.1f} ({_v131_pct(near.get("smoothed_probability"))})'
        lines.extend([
            '',
            f'<b>{idx}) {html.escape(str(adv.get("status") or "FORECAST"))}</b>',
            f'{html.escape(_v113_bet_sentence(item, with_odds=not reference))}',
            f'<b>Проєкція / лінія / edge:</b> {_v11_num((item.get("live") or {}).get("projection_used"))} / '
            f'{_v11_num(item.get("line"))} / {_v11_num(m.get("line_edge"))}',
            f'<b>Історія:</b> {html.escape(history_text)}',
            f'<b>Model score:</b> {_v131_pct(m.get("p_final_internal_score"))} | '
            f'<b>Conservative P:</b> {_v131_pct(m.get("probability"))}',
            f'<b>Odds:</b> {odds_text} | <b>fair:</b> {fair} | '
            f'<b>EV:</b> {"N/A" if ev is None else f"{ev:+.1%}"} | '
            f'<b>price edge:</b> {"N/A" if price_edge is None else f"{price_edge:+.1%}"}',
            f'<b>Stat gate:</b> {html.escape(_v131_stat_status(item))}',
            f'<b>Бюджет:</b> {html.escape(str(adv.get("stake_budget") or "0%"))}',
            f'<b>Strict audit:</b> {html.escape(reason_text)}',
        ])
    text = '\n'.join(lines)
    if len(text) <= 3900:
        return [text]
    return [text[i:i+3900] for i in range(0, len(text), 3900)]


def process_vps_match_file(
    match_path: str | Path,
    *,
    output_path: str | Path | None=None,
    zones_path: str | Path | None=None,
    db_path: str | Path='super_basket.sqlite3',
    mode: str='ACTION',
    require_gpt: bool=False,
    enable_gpt: bool=False,
    enable_telegram: bool=True,
    dry_run: bool=False,
    strict_schema: bool=False,
    checkpoint: Optional[int]=None,
    gpt_reviewer: Optional[Callable[[dict[str, Any], dict[str, Any]], dict[str, Any]]]=None,
    telegram_sender: Optional[Callable[[str], dict[str, Any]]]=None,
) -> dict[str, Any]:
    # Prevent the legacy advisor from sending before strict post-processing.
    result = _V131_BASE_PROCESS(
        match_path, output_path=output_path, zones_path=zones_path, db_path=db_path,
        mode=mode, require_gpt=require_gpt, enable_gpt=enable_gpt,
        enable_telegram=False, dry_run=dry_run, strict_schema=strict_schema,
        checkpoint=checkpoint, gpt_reviewer=gpt_reviewer, telegram_sender=telegram_sender,
    )
    calculation = result.get('super_basket_calculation') or {}
    canonical_snapshot = calculation.get('canonical_snapshot') or {}
    # The compact canonical_snapshot intentionally omits raw last35 games. Rebuild
    # the full canonical object for exact historical zones, then preserve the
    # checkpoint-resolved state/time from the core calculation.
    source_for_history = load_json(Path(match_path).expanduser().resolve())
    canonical = adapt_match(source_for_history, deepcopy(DEFAULT_CONFIG), strict_schema)
    for key in (
        'stage', 'trigger_checkpoint', 'advisor_checkpoint', 'advisor_checkpoint_name',
        'current_quarter', 'elapsed_game_seconds', 'remaining_game_seconds',
        'quarter_seconds_remaining', 'clock', 'score', 'quarters', 'snapshot_datetime'
    ):
        if key in canonical_snapshot:
            canonical[key] = deepcopy(canonical_snapshot[key])
    system = result.get('super_basket_system') or {}
    old_advisor = system.get('advisor') if isinstance(system.get('advisor'), dict) else {}
    pool = list(calculation.get('market_evaluations') or [])
    references = list(calculation.get('reference_line_evaluations') or [])
    real_supported = bool(old_advisor.get('real_lines_present'))
    if not real_supported:
        pool = references
    elif references:
        pool = pool + references
    selected, rejected = _v131_select(pool, canonical)
    messages = _v131_messages(selected, calculation)
    input_gate = old_advisor.get('input_state_gate') or system.get('input_state_gate') or {}
    dispatch = bool(selected and input_gate.get('allowed', True))
    top = selected[0] if selected else None
    if top:
        adv = top.get('advisor') or {}
        action = str(adv.get('action') or 'FORECAST')
        status = str(adv.get('status') or 'FORECAST ONLY')
        stake = str(adv.get('stake_budget') or '0.00-0.00% від банку')
        metrics = adv.get('ranking_metrics_v13_1') or {}
        decision = system.get('decision') if isinstance(system.get('decision'), dict) else {}
        decision.update({
            'action': action, 'deterministic_action': action, 'status': status,
            'stake': stake, 'budget_recommendation': _v12_budget_recommendation(action, status, stake),
            'market': _v11_compact_line(top),
            'probabilities': {
                'p_final_internal_score': metrics.get('p_final_internal_score'),
                'conservative_probability_not_calibrated': metrics.get('probability'),
                'p_live': metrics.get('p_live'), 'p_hist': metrics.get('p_hist'),
                'p_scenario': metrics.get('p_scenario'),
            },
            'market_value': {
                'market_fair_probability': metrics.get('market_fair_probability'),
                'break_even_probability': metrics.get('break_even_probability'),
                'expected_value': metrics.get('ev'), 'price_edge': metrics.get('price_edge'),
            },
            'strict_active': bool(adv.get('strict_active')),
            'strict_gate_reasons': list(adv.get('hard_gate_reasons_v13_1') or []),
            'trigger_uk': 'Лінія, коефіцієнт, рахунок і час мають залишатися актуальними.',
        })
        system['decision'] = decision
        system['decision_text'] = f'{action} | {status}'
    else:
        system['decision'] = {
            'action': 'NONE', 'status': 'NO VALID FORECAST', 'stake': '0%',
            'strict_active': False, 'strict_gate_reasons': ['NO_VALID_MARKET'],
        }
    advisor = {
        'version': ADVISOR_VERSION,
        'action': (top.get('advisor') or {}).get('action') if top else 'NONE',
        'status': (top.get('advisor') or {}).get('status') if top else 'NO VALID FORECAST',
        'dispatch': dispatch,
        'dispatch_reason': 'Strict audited forecast; at most one monetary signal per snapshot.',
        'policy': {
            'p_final_is_internal_score_not_calibrated_probability': True,
            'stat_against_or_conflict_money_block': True,
            'fake_profile_money_block': True,
            'team_it_weakest_gate_70': True,
            'q4_context_hard_gate': True,
            'positive_ev_required': True,
            'one_active_bet_per_snapshot': True,
            'general_totals_priority': True,
            'reference_lines_zero_stake': True,
            'max_stake_percent_bankroll': V131_MAX_STAKE_PCT,
            '3x3_match_total_only': True,
        },
        'primary_lines': selected,
        'reference_lines_generated': references,
        'real_lines_present': real_supported,
        'rejected_fatal_lines': rejected,
        'telegram_messages': messages,
        'input_state_gate': input_gate,
    }
    system['advisor'] = advisor
    system['version'] = ADVISOR_VERSION
    calculation['engine_version'] = ADVISOR_VERSION
    calculation['formula_registry_v13_1'] = {
        'authoritative_chain': 'validation -> exact history -> scenario -> independent live projection -> Phi P_live -> normalized weights -> caps/blockers -> P_final internal score -> strict gates -> no-vig/EV -> stake',
        'p_final_status': 'internal model score; not empirical probability',
        'strict_stat_policy': 'AGAINST or CONFLICT => forecast only, stake 0',
        'team_it_policy': 'own scored and opponent allowed each >=70%',
        'money_policy': 'positive EV and price edge required; quarter Kelly capped at 1.5% bankroll',
    }
    delivery = {'status': 'SKIPPED_NO_MESSAGES', 'sent': False, 'message_count': len(messages)}
    if dispatch and messages and not dry_run and enable_telegram:
        delivery = _v11_send_messages(messages, telegram_sender=telegram_sender)
    elif dispatch and messages and dry_run:
        delivery = {'status': 'DRY_RUN_NOT_SENT', 'sent': False, 'message_count': len(messages)}
    elif dispatch and messages and not enable_telegram:
        delivery = {'status': 'SKIPPED_TELEGRAM_DISABLED', 'sent': False, 'message_count': len(messages)}
    system['telegram_delivery'] = delivery
    target = Path(output_path).expanduser().resolve() if output_path else Path(match_path).expanduser().resolve().with_name(Path(match_path).stem + '_advisor_result.json')
    save_json(target, result)
    return result


DEFAULT_CONFIG.setdefault('v13_1_policy', {}).update({
    'version': ADVISOR_VERSION,
    'strict_stat_conflict_block': True,
    'one_active_per_snapshot': True,
    'quarter_kelly_cap_percent': V131_MAX_STAKE_PCT,
    'p_final_not_labeled_calibrated_probability': True,
})


# =============================================================================
# v13.2 MARKET-SCOPE / REAL-SAMPLE / PREMATCH-PROJECTION AUDIT
# =============================================================================
# Fixes three production defects found on Al Hekmeh Beirut vs Al Riyadi:
# 1) half-time team totals incorrectly labelled by the parser as full-match IT;
# 2) weighted copies (5x own + 3x allowed) displayed as 196 independent games;
# 3) pre-match projections of incompatible scales blended with the bookmaker line.

ADVISOR_VERSION = '13.2.0-SCOPE-HISTORY-PROJECTION-AUDITED'
SYSTEM_VERSION = ADVISOR_VERSION
DEFAULT_CONFIG['engine_version'] = ADVISOR_VERSION


def _v132_line_containers(source: dict[str, Any]) -> dict[str, Any]:
    value = (
        source.get('lines')
        or source.get('bookmaker_lines')
        or source.get('bookmaker_markets')
        or source.get('markets')
        or {}
    )
    return value if isinstance(value, dict) else {}


def _v132_projection_blocks(source: dict[str, Any], canonical: dict[str, Any]) -> dict[str, Any]:
    direct = source.get('projections')
    if isinstance(direct, dict):
        return direct
    blocks = canonical.get('parser_blocks') if isinstance(canonical.get('parser_blocks'), dict) else {}
    value = blocks.get('projections')
    return value if isinstance(value, dict) else {}


def _v132_team_projection(projections: dict[str, Any], side_key: str, segment: str) -> Optional[float]:
    segment = str(segment or '').upper()
    key = 'home_final' if side_key == 'home' else 'away_final'
    if segment == 'MATCH':
        row = projections.get('segment_projection') if isinstance(projections.get('segment_projection'), dict) else {}
        return to_number(row.get(key))
    quarter_key = 'team_a_center' if side_key == 'home' else 'team_b_center'
    quarter_names = {
        'H1': ('q1', 'q2'), 'H2': ('q3', 'q4'),
        'Q1': ('q1',), 'Q2': ('q2',), 'Q3': ('q3',), 'Q4': ('q4',),
    }.get(segment, ())
    values: list[float] = []
    for name in quarter_names:
        row = projections.get(name) if isinstance(projections.get(name), dict) else {}
        number = to_number(row.get(quarter_key))
        if number is None:
            return None
        values.append(float(number))
    return sum(values) if values else None


def _v132_match_total_center(source: dict[str, Any]) -> Optional[float]:
    rows = _v132_line_containers(source).get('match_total') or []
    values = [float(v) for row in rows if isinstance(row, dict) and (v := to_number(row.get('line'))) is not None]
    return statistics.median(values) if values else None


def _v132_individual_lines(source: dict[str, Any]) -> dict[str, list[float]]:
    containers = _v132_line_containers(source)
    output: dict[str, list[float]] = {'home': [], 'away': []}
    for key, side in (('home_ind_total', 'home'), ('away_ind_total', 'away')):
        for row in containers.get(key) or []:
            if not isinstance(row, dict):
                continue
            value = to_number(row.get('line'))
            if value is not None:
                output[side].append(float(value))
    return output


_V132_SCOPE_OVERRIDES_BASE = _v116_team_it_scope_overrides


def _v116_team_it_scope_overrides(source: dict[str, Any]) -> dict[tuple[str, float], dict[str, Any]]:
    overrides = dict(_V132_SCOPE_OVERRIDES_BASE(source))
    containers = _v132_line_containers(source)
    projections = source.get('projections') if isinstance(source.get('projections'), dict) else {}
    match_total = _v132_match_total_center(source)
    lines = _v132_individual_lines(source)
    if not lines['home'] or not lines['away']:
        return overrides

    for home_line in lines['home']:
        for away_line in lines['away']:
            pair_sum = home_line + away_line
            pair_half_evidence = bool(
                match_total is not None
                and 0.40 <= pair_sum / float(match_total) <= 0.60
                and abs(pair_sum - 0.50 * float(match_total)) <= max(6.0, 0.08 * float(match_total))
            )
            details: dict[str, Any] = {
                'pair_sum': pair_sum,
                'match_total_center': match_total,
                'pair_half_evidence': pair_half_evidence,
            }
            all_h1_closer = True
            for side, line in (('home', home_line), ('away', away_line)):
                full_projection = _v132_team_projection(projections, side, 'MATCH')
                h1_projection = _v132_team_projection(projections, side, 'H1')
                if full_projection is None or h1_projection is None:
                    all_h1_closer = False
                    continue
                full_error = abs(line - full_projection) / max(1.0, abs(full_projection))
                h1_error = abs(line - h1_projection) / max(1.0, abs(h1_projection))
                details[f'{side}_full_projection'] = full_projection
                details[f'{side}_h1_projection'] = h1_projection
                details[f'{side}_full_error'] = full_error
                details[f'{side}_h1_error'] = h1_error
                if not (h1_error <= 0.20 and h1_error + 0.12 < full_error):
                    all_h1_closer = False
            if pair_half_evidence and all_h1_closer:
                for bucket, line in (('home_ind_total', home_line), ('away_ind_total', away_line)):
                    overrides[(bucket, float(line))] = {
                        'market_type': 'TEAM_IT_H1',
                        'segment': 'H1',
                        'reason': 'PAIR_SUM_AND_PROJECTION_SCALE_IDENTIFY_H1_TEAM_TOTAL',
                        'confidence': 'HIGH',
                        **deepcopy(details),
                    }
    return overrides


_V132_PARSE_BASE = parse_markets


def parse_markets(
    source: dict[str, Any],
    canonical: dict[str, Any],
    config: dict[str, Any],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    evaluations, audit = _V132_PARSE_BASE(source, canonical, config)
    projections = _v132_projection_blocks(source, canonical)

    def inspect(row: dict[str, Any], *, audit_row: bool=False) -> None:
        market_type = str(row.get('market_type') or '')
        if market_type != 'TEAM_IT_MATCH':
            return
        team = row.get('team')
        side_key = 'home' if team == canonical.get('home_team') else 'away' if team == canonical.get('away_team') else None
        line = to_number(row.get('line'))
        if side_key is None or line is None:
            return
        full_projection = _v132_team_projection(projections, side_key, 'MATCH')
        h1_projection = _v132_team_projection(projections, side_key, 'H1')
        if full_projection is None:
            return
        ratio = float(line) / max(1.0, float(full_projection))
        if 0.65 <= ratio <= 1.35:
            return
        issues_key = 'issues' if audit_row else 'parser_issues'
        issues = list(row.get(issues_key) or [])
        if h1_projection is not None:
            h1_error = abs(float(line) - float(h1_projection)) / max(1.0, abs(float(h1_projection)))
            full_error = abs(float(line) - float(full_projection)) / max(1.0, abs(float(full_projection)))
            if h1_error <= 0.20 and h1_error + 0.12 < full_error:
                row['market_type_original'] = row.get('market_type')
                row['segment_original'] = row.get('segment')
                row['market_type'] = 'TEAM_IT_H1'
                row['segment'] = 'H1'
                row['source_scope_original'] = row.get('source_scope')
                row['source_scope'] = 'H1_CORRECTED_BY_SCALE'
                row['market_scope_correction'] = {
                    'reason': 'TEAM_IT_MATCH_LINE_IMPLAUSIBLE_AND_H1_PROJECTION_MATCHES',
                    'confidence': 'HIGH',
                    'line': float(line),
                    'full_match_projection': float(full_projection),
                    'h1_projection': float(h1_projection),
                    'line_to_full_projection_ratio': ratio,
                }
                row[issues_key] = [x for x in issues if x != 'AMBIGUOUS_TEAM_IT_SCOPE']
                if not audit_row:
                    row['eligible_market'] = not row[issues_key]
                return
        code = 'IMPLAUSIBLE_TEAM_IT_MATCH_SCOPE'
        if code not in issues:
            issues.append(code)
        row[issues_key] = issues
        row['market_scope_sanity'] = {
            'line': float(line),
            'full_match_projection': float(full_projection),
            'h1_projection': float(h1_projection) if h1_projection is not None else None,
            'line_to_full_projection_ratio': ratio,
            'action': 'BLOCK_AMBIGUOUS_SCOPE',
        }
        if not audit_row:
            row['eligible_market'] = False

    for item in evaluations:
        inspect(item, audit_row=False)
    for row in audit:
        inspect(row, audit_row=True)
    return evaluations, audit


_V132_PARSER_VALUES_BASE = _v124_parser_projection_values


def _v124_parser_projection_values(market: dict[str, Any], canonical: dict[str, Any], current_points: float) -> dict[str, Any]:
    values = _V132_PARSER_VALUES_BASE(market, canonical, current_points)
    stage = str(canonical.get('stage') or '').upper()
    segment = str(market.get('segment') or '').upper()
    if segment != 'MATCH':
        return values
    trusted = values.get('parser_segment_projection') if isinstance(values.get('parser_segment_projection'), dict) else None
    trusted_value = to_number((trusted or {}).get('value'))
    for name in ('parser_live_calibrated', 'parser_pre_match_stat'):
        row = values.get(name)
        if not isinstance(row, dict):
            continue
        number = to_number(row.get('value'))
        if stage == 'PRE_MATCH' and name == 'parser_live_calibrated':
            row['valid'] = False
            row['reason'] = 'PREMATCH_LIVE_CALIBRATED_NOT_APPLICABLE'
            continue
        if trusted_value is not None and number is not None:
            ratio = float(number) / max(1.0, abs(float(trusted_value)))
            if ratio < 0.65 or ratio > 1.35:
                row['valid'] = False
                row['reason'] = 'SCOPE_SCALE_MISMATCH_VS_SEGMENT_PROJECTION'
                row['scale_ratio'] = ratio
    return values


_V132_LIVE_BASE = calculate_live_projection


def calculate_live_projection(
    market: dict[str, Any],
    canonical: dict[str, Any],
    history: dict[str, Any],
    scenario: dict[str, Any],
    config: dict[str, Any],
    stat: Optional[dict[str, Any]]=None,
) -> dict[str, Any]:
    result = _V132_LIVE_BASE(market, canonical, history, scenario, config, stat)
    if str(canonical.get('stage') or '').upper() != 'PRE_MATCH':
        return result
    parser = result.get('parser_projection_preservation') if isinstance(result.get('parser_projection_preservation'), dict) else {}
    segment = str(market.get('segment') or '').upper()
    preferred = ('parser_segment_projection', 'parser_pre_match_stat') if segment == 'MATCH' else ('parser_segment_centers',)
    chosen_name = None
    chosen_value = None
    for name in preferred:
        row = parser.get(name)
        if isinstance(row, dict) and row.get('valid') and to_number(row.get('value')) is not None:
            chosen_name = name
            chosen_value = float(row['value'])
            break
    if chosen_value is None:
        return result
    current_points = float(to_number(result.get('current_points')) or 0.0)
    projection = max(current_points, chosen_value)
    _v114_recalculate_live_probability(result, market, projection)
    result['projection_parser_blended'] = projection
    result['parser_projection_used'] = True
    result['parser_projection_components_used'] = [{'name': chosen_name, 'value': projection, 'weight': 1.0}]
    result['projection_formula_mode'] = 'PREMATCH_SCOPE_PRIMARY_PARSER_PROJECTION'
    result['projection_formula'] = (
        'Pre-match uses one scope-compatible projection only: full-match segment_projection for MATCH, '
        'or summed quarter centers for H1/H2/Q markets. The bookmaker line and incompatible-scale parser '
        'fields are not blended into the projection.'
    )
    components = result.setdefault('components', {})
    for key, row in parser.items():
        if isinstance(row, dict) and 'value' in row:
            components[f'projection_{key}'] = {
                'value': row.get('value'),
                'included': key == chosen_name,
                'role': 'prematch_scope_primary' if key == chosen_name else 'excluded_prematch_component',
                'exclusion_reason': None if key == chosen_name else row.get('reason') or 'NOT_PRIMARY_FOR_SCOPE',
            }
    return result


def _v132_game_team_segment(game: dict[str, Any], team: str, segment: str) -> Optional[float]:
    side = 'home' if str(game.get('home_team') or '') == str(team) else 'away' if str(game.get('away_team') or '') == str(team) else None
    if side is None:
        return None
    segment = str(segment or 'MATCH').upper()
    if segment == 'MATCH':
        return to_number(game.get('home_score') if side == 'home' else game.get('away_score'))
    quarters = game.get('quarters') or []
    indices = {'H1': (0,1), 'H2': (2,3), 'Q1': (0,), 'Q2': (1,), 'Q3': (2,), 'Q4': (3,)}.get(segment, ())
    values: list[float] = []
    for idx in indices:
        if idx >= len(quarters) or not isinstance(quarters[idx], dict):
            return None
        number = to_number(quarters[idx].get(side))
        if number is None:
            return None
        values.append(float(number))
    return sum(values) if values else None


def _v132_perspective_segment(game: dict[str, Any], segment: str, *, allowed: bool=False) -> Optional[float]:
    segment = str(segment or 'MATCH').upper()
    if segment == 'MATCH':
        return to_number(game.get('opponent_score') if allowed else game.get('team_score'))
    values = game.get('opponent_quarters') if allowed else game.get('team_quarters')
    values = values or []
    indices = {'H1': (0,1), 'H2': (2,3), 'Q1': (0,), 'Q2': (1,), 'Q3': (2,), 'Q4': (3,)}.get(segment, ())
    selected: list[float] = []
    for idx in indices:
        if idx >= len(values):
            return None
        number = to_number(values[idx])
        if number is None:
            return None
        selected.append(float(number))
    return sum(selected) if selected else None


def _v132_probability_row(values: list[float], line: float, side: str) -> dict[str, Any]:
    clean = [float(v) for v in values if v is not None and math.isfinite(float(v))]
    outcomes = [settle(v, float(line), side) for v in clean]
    wins = outcomes.count('win')
    losses = outcomes.count('loss')
    pushes = outcomes.count('push')
    n = len(clean)
    return {
        'wins': wins, 'losses': losses, 'pushes': pushes, 'n': n,
        'raw_probability': wins / n if n else None,
        'smoothed_probability': smoothed_probability(wins, n, 1.0, 1.0) if n else None,
    }


_V132_HISTORY_ZONE_BASE = _v131_history_zone


def _v131_history_zone(item: dict[str, Any], canonical: dict[str, Any]) -> dict[str, Any]:
    market_type = str(item.get('market_type') or '')
    if market_type not in {'TEAM_IT_MATCH', 'TEAM_IT_H1', 'TEAM_IT_H2', 'CURRENT_QUARTER_TEAM_IT'}:
        return _V132_HISTORY_ZONE_BASE(item, canonical)
    line = to_number(item.get('line'))
    side = str(item.get('side') or '').upper()
    team = str(item.get('team') or '')
    segment = str(item.get('segment') or 'MATCH').upper()
    if line is None or side not in {'OVER','UNDER'} or not team:
        return {'available': False, 'reason': 'NO_LINE_SIDE_OR_TEAM'}
    own_key = 'team_a' if team == canonical.get('home_team') else 'team_b'
    opp_key = 'team_b' if own_key == 'team_a' else 'team_a'
    own_values = [
        value for game in canonical.get('history', {}).get(own_key) or []
        if (value := _v132_perspective_segment(game, segment, allowed=False)) is not None
    ]
    allowed_values = [
        value for game in canonical.get('history', {}).get(opp_key) or []
        if (value := _v132_perspective_segment(game, segment, allowed=True)) is not None
    ]
    h2h_values = [
        value for game in canonical.get('history', {}).get('h2h') or []
        if (value := _v132_game_team_segment(game, team, segment)) is not None
    ]
    if not own_values or not allowed_values:
        return {'available': False, 'reason': 'TEAM_IT_OWN_OR_ALLOWED_HISTORY_MISSING'}

    def composite(candidate: float) -> dict[str, Any]:
        own = _v132_probability_row(own_values, candidate, side)
        allowed = _v132_probability_row(allowed_values, candidate, side)
        h2h = _v132_probability_row(h2h_values, candidate, side) if h2h_values else {'n':0,'smoothed_probability':None}
        components = [('own_scored', own, 0.50), ('opponent_allowed', allowed, 0.35)]
        if h2h.get('n'):
            components.append(('h2h_it', h2h, 0.15))
        total_weight = sum(weight for _, row, weight in components if row.get('smoothed_probability') is not None)
        weighted = (
            sum(float(row['smoothed_probability']) * weight for _, row, weight in components if row.get('smoothed_probability') is not None) / total_weight
            if total_weight > 0 else None
        )
        weakest = min(float(own.get('smoothed_probability') or 0.0), float(allowed.get('smoothed_probability') or 0.0))
        return {
            'line': float(candidate), 'side': side,
            'own_scored': own, 'opponent_allowed': allowed, 'h2h_it': h2h,
            'smoothed_probability': weighted,
            'weighted_probability': weighted,
            'weakest_gate': weakest,
            'n_effective': min(70, int(own.get('n') or 0) + int(allowed.get('n') or 0)),
            'method': '0.50 own scored + 0.35 opponent allowed + 0.15 H2H; real samples, no duplicated weights',
        }

    exact = composite(float(line))
    all_values = own_values + allowed_values + h2h_values
    start = math.floor(min(all_values)) - 0.5
    stop = math.ceil(max(all_values)) + 0.5
    candidates: list[dict[str, Any]] = []
    cursor = start
    while cursor <= stop + 1e-9:
        if (side == 'OVER' and cursor <= float(line) + 1e-9) or (side == 'UNDER' and cursor >= float(line) - 1e-9):
            row = composite(round(cursor,1))
            if float(row.get('weighted_probability') or 0.0) >= 0.75 and float(row.get('weakest_gate') or 0.0) >= 0.70:
                candidates.append(row)
        cursor += 1.0
    nearest = min(candidates, key=lambda row: abs(float(row['line']) - float(line))) if candidates else None
    return {
        'available': True,
        'kind': 'TEAM_IT_COMPONENTS',
        'exact': exact,
        'nearest_75_zone': nearest,
        'n': exact.get('n_effective'),
        'method': exact.get('method'),
    }


# Over/Under prices must be paired by logical market, not side-specific market_id.
def _v131_market_key(item: dict[str, Any]) -> tuple[Any, ...]:
    common_source = item.get('source_market_id')
    if common_source not in (None, ''):
        return ('SOURCE', common_source, item.get('market_type'), item.get('segment'), item.get('team'), to_number(item.get('line')), item.get('bookmaker'))
    return (
        'LOGICAL', item.get('source_bucket'), item.get('source_scope'),
        item.get('market_type'), item.get('segment'), item.get('team'),
        to_number(item.get('line')), item.get('bookmaker'),
    )


_V132_HARD_GATE_BASE = _v131_hard_gate


def _v131_hard_gate(item: dict[str, Any], canonical: dict[str, Any], metrics: dict[str, Any]) -> list[str]:
    reasons = list(_V132_HARD_GATE_BASE(item, canonical, metrics))
    issues = {str(v) for v in item.get('parser_issues') or []}
    if 'IMPLAUSIBLE_TEAM_IT_MATCH_SCOPE' in issues:
        reasons.append('IMPLAUSIBLE_TEAM_IT_MATCH_SCOPE')
    return list(dict.fromkeys(reasons))


def _v132_history_text(item: dict[str, Any], zone: dict[str, Any]) -> str:
    if not zone.get('available'):
        return 'N/A'
    exact = zone.get('exact') or {}
    near = zone.get('nearest_75_zone') or {}
    if zone.get('kind') == 'TEAM_IT_COMPONENTS':
        own = exact.get('own_scored') or {}
        allowed = exact.get('opponent_allowed') or {}
        h2h = exact.get('h2h_it') or {}
        parts = [
            f'власні очки {int(own.get("wins") or 0)}/{int(own.get("n") or 0)} (Bayes {_v131_pct(own.get("smoothed_probability"))})',
            f'суперник пропускав {int(allowed.get("wins") or 0)}/{int(allowed.get("n") or 0)} (Bayes {_v131_pct(allowed.get("smoothed_probability"))})',
        ]
        if int(h2h.get('n') or 0):
            parts.append(f'H2H {int(h2h.get("wins") or 0)}/{int(h2h.get("n") or 0)} (Bayes {_v131_pct(h2h.get("smoothed_probability"))})')
        parts.append(f'P_hist IT {_v131_pct(exact.get("weighted_probability"))}; weakest {_v131_pct(exact.get("weakest_gate"))}')
        text = '; '.join(parts)
    else:
        text = f'{int(exact.get("wins") or 0)}/{int(exact.get("n") or 0)}; Bayes {_v131_pct(exact.get("smoothed_probability"))}'
    if near:
        probability = near.get('weighted_probability') if zone.get('kind') == 'TEAM_IT_COMPONENTS' else near.get('smoothed_probability')
        text += f'; найближча 75%+ зона {item.get("side")} {float(near.get("line")):.1f} ({_v131_pct(probability)})'
    return text


_V132_REASON_LABELS = {
    'MAX_ONE_ACTIVE_BET_PER_SNAPSHOT': 'другий варіант: активний фінансовий сигнал у цьому snapshot уже вибрано',
    'LIVE_EDGE_BELOW_STRICT_MARKET_MIN': 'edge нижчий за мінімум цього ринку',
    'EV_BELOW_MINIMUM': 'EV нижче мінімуму',
    'PRICE_EDGE_BELOW_MINIMUM': 'цінова перевага нижче мінімуму',
    'TEAM_IT_OPPONENT_ALLOWED_BELOW_70': 'суперник пропускав цю лінію менш ніж у 70% вибірки',
    'TEAM_IT_OWN_SCORED_BELOW_70': 'команда пробивала цю лінію менш ніж у 70% вибірки',
    'P_FINAL_INTERNAL_BELOW_65': 'внутрішній model score нижче 65%',
    'IMPLAUSIBLE_TEAM_IT_MATCH_SCOPE': 'лінія не відповідає масштабу IT за весь матч',
}


def _v131_messages(selected: list[dict[str, Any]], calculation: dict[str, Any]) -> list[str]:
    if not selected:
        return []
    snapshot = calculation.get('canonical_snapshot') or {}
    score = snapshot.get('score') or {}
    lines = [
        '<b>🏀 SUPER BASKET v13.2 — SCOPE & HISTORY AUDITED</b>',
        f'<b>Матч:</b> {html.escape(str(snapshot.get("name") or "N/A"))}',
        f'<b>Стадія:</b> {html.escape(str(snapshot.get("stage") or "N/A"))} | '
        f'<b>Рахунок:</b> {_v11_num(score.get("home"))}:{_v11_num(score.get("away"))}',
        '<i>P_final — внутрішній model score. Кількість матчів ніколи не збільшується ваговими копіями.</i>',
    ]
    for idx, item in enumerate(selected, 1):
        adv = item.get('advisor') or {}
        m = adv.get('ranking_metrics_v13_1') or {}
        zone = m.get('history_zone') or {}
        reference = bool(item.get('is_reference_line'))
        odds_text = 'N/A' if reference else _v11_num(item.get('odds'), 2)
        fair = _v131_pct(m.get('market_fair_probability'))
        ev = to_number(m.get('ev'))
        price_edge = to_number(m.get('price_edge'))
        reasons = list(adv.get('hard_gate_reasons_v13_1') or [])
        reason_text = '; '.join(_V132_REASON_LABELS.get(str(v), str(v)) for v in reasons) or 'усі strict gates пройдено'
        lines.extend([
            '',
            f'<b>{idx}) {html.escape(str(adv.get("status") or "FORECAST"))}</b>',
            f'{html.escape(_v113_bet_sentence(item, with_odds=not reference))}',
        ])
        correction = item.get('market_scope_correction')
        if isinstance(correction, dict):
            lines.append(
                '<b>Корекція scope:</b> парсер позначив IT як весь матч, але лінія та проєкції '
                f'однозначно відповідають H1; розраховано як IT першої половини. '
                f'Full projection {_v11_num(correction.get("away_full_projection") or correction.get("home_full_projection") or correction.get("full_match_projection"))}, '
                f'H1 projection {_v11_num(correction.get("away_h1_projection") or correction.get("home_h1_projection") or correction.get("h1_projection"))}.'
            )
        lines.extend([
            f'<b>Проєкція / лінія / edge:</b> {_v11_num((item.get("live") or {}).get("projection_used"))} / '
            f'{_v11_num(item.get("line"))} / {_v11_num(m.get("line_edge"))}',
            f'<b>Історія:</b> {html.escape(_v132_history_text(item, zone))}',
            f'<b>Model score:</b> {_v131_pct(m.get("p_final_internal_score"))} | '
            f'<b>Conservative P:</b> {_v131_pct(m.get("probability"))}',
            f'<b>Odds:</b> {odds_text} | <b>fair:</b> {fair} | '
            f'<b>EV:</b> {"N/A" if ev is None else f"{ev:+.1%}"} | '
            f'<b>price edge:</b> {"N/A" if price_edge is None else f"{price_edge:+.1%}"}',
            f'<b>Stat gate:</b> {html.escape(_v131_stat_status(item))}',
            f'<b>Бюджет:</b> {html.escape(str(adv.get("stake_budget") or "0%"))}',
            f'<b>Strict audit:</b> {html.escape(reason_text)}',
        ])
    text = '\n'.join(lines)
    return [text] if len(text) <= 3900 else [text[i:i+3900] for i in range(0, len(text), 3900)]


_V132_PROCESS_BASE = process_vps_match_file


def process_vps_match_file(
    match_path: str | Path,
    *,
    output_path: str | Path | None=None,
    zones_path: str | Path | None=None,
    db_path: str | Path='super_basket.sqlite3',
    mode: str='ACTION',
    require_gpt: bool=False,
    enable_gpt: bool=False,
    enable_telegram: bool=True,
    dry_run: bool=False,
    strict_schema: bool=False,
    checkpoint: Optional[int]=None,
    gpt_reviewer: Optional[Callable[[dict[str, Any], dict[str, Any]], dict[str, Any]]]=None,
    telegram_sender: Optional[Callable[[str], dict[str, Any]]]=None,
) -> dict[str, Any]:
    result = _V132_PROCESS_BASE(
        match_path, output_path=output_path, zones_path=zones_path, db_path=db_path,
        mode=mode, require_gpt=require_gpt, enable_gpt=enable_gpt,
        enable_telegram=enable_telegram, dry_run=dry_run, strict_schema=strict_schema,
        checkpoint=checkpoint, gpt_reviewer=gpt_reviewer, telegram_sender=telegram_sender,
    )
    system = result.get('super_basket_system') or {}
    advisor = system.get('advisor') if isinstance(system.get('advisor'), dict) else {}
    calculation = result.get('super_basket_calculation') or {}
    system['version'] = ADVISOR_VERSION
    if advisor:
        advisor['version'] = ADVISOR_VERSION
        advisor.setdefault('policy', {}).update({
            'market_scope_scale_guard': True,
            'team_it_real_sample_display': True,
            'prematch_scope_primary_projection': True,
            'logical_over_under_price_pairing': True,
        })
    calculation['engine_version'] = ADVISOR_VERSION
    calculation['v13_2_audit'] = {
        'team_it_weighted_copies_not_counted_as_games': True,
        'half_scale_team_it_scope_correction': True,
        'prematch_incompatible_projection_scales_blocked': True,
        'bookmaker_line_not_used_as_prematch_projection': True,
    }
    target = Path(output_path).expanduser().resolve() if output_path else Path(match_path).expanduser().resolve().with_name(Path(match_path).stem + '_advisor_result.json')
    save_json(target, result)
    return result


DEFAULT_CONFIG.setdefault('v13_2_policy', {}).update({
    'version': ADVISOR_VERSION,
    'market_scope_scale_guard': True,
    'team_it_real_sample_display': True,
    'prematch_scope_primary_projection': True,
    'logical_price_pairing': True,
})


# ===== v13.2.1 user-facing decision consistency cleanup =====
ADVISOR_VERSION = '13.2.1-SCOPE-HISTORY-PROJECTION-AUDITED'
SYSTEM_VERSION = ADVISOR_VERSION
DEFAULT_CONFIG['engine_version'] = ADVISOR_VERSION

_V1321_PROCESS_BASE = process_vps_match_file


def _v132_reason_texts(codes: list[Any]) -> list[str]:
    extra = {
        'EXACT_HISTORY_BELOW_DYNAMIC_MIN': 'точна історична зона нижча за обов’язковий поріг',
        'REFERENCE_LINE_NO_REAL_ODDS': 'немає реальної лінії та коефіцієнта БК',
        'STAT_DIRECTION_AGAINST': 'live-статистика підтримує протилежний напрямок',
        'STAT_DIRECTION_CONFLICT': 'live-статистика конфліктує з напрямком',
        'FAKE_PROFILE_AGAINST_CANDIDATE': 'виявлено fake-профіль проти кандидата',
        'PRODUCTION_ROUTER_BLOCK': 'цей ринок не дозволений на поточному checkpoint',
    }
    return [_V132_REASON_LABELS.get(str(code), extra.get(str(code), str(code))) for code in codes]


def process_vps_match_file(
    match_path: str | Path,
    *,
    output_path: str | Path | None=None,
    zones_path: str | Path | None=None,
    db_path: str | Path='super_basket.sqlite3',
    mode: str='ACTION',
    require_gpt: bool=False,
    enable_gpt: bool=False,
    enable_telegram: bool=True,
    dry_run: bool=False,
    strict_schema: bool=False,
    checkpoint: Optional[int]=None,
    gpt_reviewer: Optional[Callable[[dict[str, Any], dict[str, Any]], dict[str, Any]]]=None,
    telegram_sender: Optional[Callable[[str], dict[str, Any]]]=None,
) -> dict[str, Any]:
    result = _V1321_PROCESS_BASE(
        match_path, output_path=output_path, zones_path=zones_path, db_path=db_path,
        mode=mode, require_gpt=require_gpt, enable_gpt=enable_gpt,
        enable_telegram=enable_telegram, dry_run=dry_run, strict_schema=strict_schema,
        checkpoint=checkpoint, gpt_reviewer=gpt_reviewer, telegram_sender=telegram_sender,
    )
    system = result.get('super_basket_system') or {}
    advisor = system.get('advisor') if isinstance(system.get('advisor'), dict) else {}
    selected = list(advisor.get('primary_lines') or [])
    decision = system.get('decision') if isinstance(system.get('decision'), dict) else {}
    if selected:
        top = selected[0]
        adv = top.get('advisor') or {}
        metrics = adv.get('ranking_metrics_v13_1') or {}
        projection = to_number((top.get('live') or {}).get('projection_used'))
        line = to_number(top.get('line'))
        edge = to_number(metrics.get('line_edge'))
        reasons = list(adv.get('hard_gate_reasons_v13_1') or [])
        readable = _v132_reason_texts(reasons)
        direction = str(top.get('side') or '')
        decision['explanation_uk'] = (
            f'Проєкція {_v11_num(projection)} проти лінії {_v11_num(line)}; '
            f'напрямок {direction}, edge {_v11_num(edge)}. '
            + ('Strict gates пройдено.' if not readable else 'Фінансовий вхід заблоковано: ' + '; '.join(readable) + '.')
        )
        decision['main_risk_uk'] = 'немає окремого ризику понад strict audit' if not readable else '; '.join(readable)
        decision['reason_codes'] = reasons
        decision['alternative_recommendations'] = [
            {
                **(_v11_compact_line(item) or {}),
                'strict_active': bool((item.get('advisor') or {}).get('strict_active')),
                'strict_gate_reasons': list((item.get('advisor') or {}).get('hard_gate_reasons_v13_1') or []),
                'history_zone': deepcopy(((item.get('advisor') or {}).get('ranking_metrics_v13_1') or {}).get('history_zone') or {}),
            }
            for item in selected[1:]
        ]
        # Remove legacy weighted pseudo-count fields from the final decision payload.
        market = decision.get('market') if isinstance(decision.get('market'), dict) else {}
        market.pop('history_zone_hits', None)
        market.pop('history_zone_n', None)
    system['version'] = ADVISOR_VERSION
    if advisor:
        advisor['version'] = ADVISOR_VERSION
        advisor.setdefault('policy', {})['legacy_weighted_pseudocounts_removed_from_decision'] = True
    calculation = result.get('super_basket_calculation') or {}
    calculation['engine_version'] = ADVISOR_VERSION
    calculation.setdefault('v13_2_audit', {})['legacy_decision_fields_rebuilt_from_strict_selection'] = True
    target = Path(output_path).expanduser().resolve() if output_path else Path(match_path).expanduser().resolve().with_name(Path(match_path).stem + '_advisor_result.json')
    save_json(target, result)
    return result

DEFAULT_CONFIG.setdefault('v13_2_policy', {})['version'] = ADVISOR_VERSION


# Human-readable audit labels completed after regression test.
_V132_REASON_LABELS.update({
    'EXACT_HISTORY_BELOW_DYNAMIC_MIN': 'точна історична зона нижча за обов’язковий поріг',
    'REFERENCE_LINE_NO_REAL_ODDS': 'немає реальної лінії та коефіцієнта БК',
    'STAT_DIRECTION_AGAINST': 'live-статистика підтримує протилежний напрямок',
    'STAT_DIRECTION_CONFLICT': 'live-статистика конфліктує з напрямком',
    'FAKE_PROFILE_AGAINST_CANDIDATE': 'виявлено fake-профіль проти кандидата',
    'PRODUCTION_ROUTER_BLOCK': 'цей ринок не дозволений на поточному checkpoint',
})


if __name__ == '__main__':
    raise SystemExit(_single_file_cli())
